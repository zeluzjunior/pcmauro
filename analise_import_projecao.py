"""
Análise do processo de importação ProjecaoGasto.
Compara: colunas do Excel | campos do modelo | campos mapeados no upload.

Uso: python analise_import_projecao.py "caminho/para/9. PCM - Previsão de Gastos 2025 e 2026.xlsx"
"""
import openpyxl
import sys
import re


# Colunas que o upload_projecao_gastos_from_file procura (utils.py)
UPLOAD_COLUMN_NAMES = {
    'ID': ['ID', 'id', 'Id'],
    'SETOR': ['SETOR ', 'SETOR', 'setor'],
    'SOLICITANTE': ['SOLICITANTE', 'solicitante'],
    'FORNECEDOR NOME FANTASIA': ['FORNECEDOR\nNOME FANTASIA', 'FORNECEDOR NOME FANTASIA', 'FORNECEDOR'],
    'FORNECEDOR CNPJ': ['FORNECEDOR\nCNPJ', 'FORNECEDOR CNPJ'],
    'DESCRIÇÃO DO SERVIÇO': ['DESCRIÇÃO DO SERVIÇO', 'DESCRI��O DO SERVI�O', 'DESCRICAO DO SERVICO'],
    'VALOR TOTAL': ['VALOR TOTAL', 'valor_total'],
    'PREVISÃO P/ EXECUÇÃO': ['PREVISÃO \nP/ EXECUÇÃO', 'PREVISÃO P/ EXECUÇÃO', 'PREVIS�O \nP/ EXECU��O'],
    'USO CONTÁBIL': ['USO \nCONTÁBIL', 'USO CONTÁBIL', 'USO \nCONT�BIL'],
    'NÚMERO DA NOTA FISCAL': ['NÚMERO DA \nNOTA FISCAL', 'NÚMERO DA NOTA FISCAL', 'N�MERO DA \nNOTA FISCAL'],
    'TIPO DE SOLICITAÇÃO': ['TIPO DE \nSOLICITAÇÃO', 'TIPO DE SOLICITAÇÃO', 'TIPO DE \nSOLICITACAO', 'TIPO DE SOLICITACAO'],
    'ORDEM DE SERVIÇO': ['ORDEM \nDE SERVIÇO', 'ORDEM DE SERVIÇO', 'ORDEM \nDE SERVI�O'],
    'DATA DE ABERTURA DA REQUISIÇÃO': ['DATA DE ABERTURA \nDA REQUISIÇÃO', 'DATA DE ABERTURA DA REQUISIÇÃO', 'DATA DE ABERTURA \nDA REQUISI��O'],
    'NÚMERO DA REQUISIÇÃO DE COMPRA': ['NÚMERO DA REQUISIÇÃO \nDE COMPRA', 'NÚMERO DA REQUISIÇÃO DE COMPRA', 'N�EMRO DA REQUISI��O \nDE COMPRA'],
    'NÚMERO DO PEDIDO DE COMPRA': ['NÚMERO DO \nPEDIDO DE COMPRA', 'NÚMERO DO PEDIDO DE COMPRA', 'N�MERO DO \nPEDIDO DE COMPRA'],
    'SERVIÇO CONCLUÍDO': ['SERVIÇO CONCLUÍDO', 'SERVI�O CONCLU�DO'],
    'NF DE SERVIÇO RECEBIDA': ['NF DE SERVIÇO\n RECEBIDA', 'NF DE SERVIÇO RECEBIDA', 'NF DE SERVI�O\n RECEBIDA'],
    'NF ENVIADA PARA LANÇAMENTO': ['NF ENVIADA\n PARA LANÇAMENTO ', 'NF ENVIADA PARA LANÇAMENTO', 'NF ENVIADA\n PARA LANAMENTO '],
    'OBSERVAÇÕES': ['OBSERVAÇÕES', 'OBSERVAES', 'OBSERVACOES', 'OBSERVAÇÕES '],
}

# Campos do modelo ProjecaoGasto (models.py) - principais
MODEL_FIELDS = {
    'id_excel', 'setor', 'solicitante', 'descricao', 'tipo_solicitacao', 'valor_total',
    'data_abertura_requisicao', 'previsao_execucao', 'mes_referencia', 'ano_referencia',
    'fornecedor_nome_fantasia', 'fornecedor_cnpj', 'uso_contabil', 'numero_nf',
    'numero_ordem_servico', 'numero_requisicao_compra', 'numero_pedido_compra',
    'servico_concluido', 'nf_servico_recebida', 'nf_enviada_lancamento', 'observacoes',
    'tipo', 'centro_atividade', 'nome_centro_atividade', 'valor_planejado', 'valor_realizado',
    'valor_projetado', 'data_requisicao', 'data_planejada', 'data_realizada', 'fornecedor',
    'numero_requisicao', 'status', 'dados_adicionais'
}


def normalize_header(header):
    if not header:
        return None
    normalized = re.sub(r'\s+', ' ', str(header).replace('\n', ' ').strip())
    return normalized.upper()


def analyze_excel(file_path):
    print("=" * 90)
    print("ANÁLISE DO PROCESSO DE IMPORTAÇÃO - Projeção de Gastos")
    print("=" * 90)
    print(f"\nArquivo: {file_path}\n")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Planilha GASTOS (usada pelo import)
        sheet_name = 'GASTOS'
        if sheet_name not in wb.sheetnames:
            print(f"[ERRO] Planilha '{sheet_name}' não encontrada!")
            print(f"Planilhas disponíveis: {wb.sheetnames}")
            print("\nDica: O import usa apenas a planilha 'GASTOS'. Verifique se o arquivo PCM tem essa aba.")
            return
        print(f"[OK] Planilha '{sheet_name}' encontrada")
        if len(wb.sheetnames) > 1:
            print(f"    Outras planilhas no arquivo: {[s for s in wb.sheetnames if s != sheet_name]}")

        sheet = wb[sheet_name]
        headers_raw = [cell.value for cell in sheet[1]]
        headers = [h for h in headers_raw if h]

        print(f"\n--- COLUNAS NO EXCEL ({len(headers)} colunas) ---")
        excel_cols_normalized = {}
        for i, h in enumerate(headers):
            norm = normalize_header(h)
            excel_cols_normalized[norm] = (h, i + 1)
            print(f"  {i+1:2}. {repr(h)}")

        # Verificar quais colunas do upload estão no Excel
        print("\n--- MAPEAMENTO: UPLOAD → EXCEL ---")
        upload_found = {}
        upload_not_found = []
        for logical_name, variations in UPLOAD_COLUMN_NAMES.items():
            found = False
            matched = None
            for var in variations:
                norm = normalize_header(var)
                if norm in excel_cols_normalized:
                    found = True
                    matched = excel_cols_normalized[norm][0]
                    break
            if found:
                upload_found[logical_name] = matched
            else:
                upload_not_found.append(logical_name)

        for name, excel_col in sorted(upload_found.items()):
            print(f"  [OK] {name:40} → Excel: {repr(excel_col)}")
        for name in upload_not_found:
            print(f"  [X]  {name:40} → NÃO ENCONTRADO NO EXCEL")

        # Colunas no Excel que NÃO estão no mapeamento de upload
        all_upload_norms = set()
        for variations in UPLOAD_COLUMN_NAMES.values():
            for v in variations:
                n = normalize_header(v)
                if n:
                    all_upload_norms.add(n)

        excel_only = [h for n, (h, _) in excel_cols_normalized.items() if n not in all_upload_norms]
        print("\n--- COLUNAS NO EXCEL QUE NÃO SÃO IMPORTADAS ---")
        if excel_only:
            for h in excel_only:
                print(f"  - {repr(h)}")
            print(f"\n  Total: {len(excel_only)} colunas não mapeadas → podem ser armazenadas em 'dados_adicionais'")
        else:
            print("  (Nenhuma - todas as colunas do Excel estão mapeadas)")

        # Campos do modelo que não recebem dados do Excel (além de id_excel, created_at, updated_at)
        fields_from_excel = {
            'setor', 'solicitante', 'descricao', 'tipo_solicitacao', 'valor_total',
            'data_abertura_requisicao', 'previsao_execucao', 'mes_referencia', 'ano_referencia',
            'fornecedor_nome_fantasia', 'fornecedor_cnpj', 'uso_contabil', 'numero_nf',
            'numero_ordem_servico', 'numero_requisicao_compra', 'numero_pedido_compra',
            'servico_concluido', 'nf_servico_recebida', 'nf_enviada_lancamento', 'observacoes',
            'centro_atividade', 'fornecedor', 'data_requisicao', 'numero_requisicao', 'tipo'
        }
        model_not_fed = MODEL_FIELDS - fields_from_excel - {'id_excel', 'dados_adicionais', 'created_at', 'updated_at'}
        # Remover campos derivados (mes_referencia, ano_referencia vêm de previsao_execucao)
        model_not_fed -= {'mes_referencia', 'ano_referencia'}

        print("\n--- CAMPOS DO MODELO SEM FONTE DIRETA NO EXCEL ---")
        for f in sorted(model_not_fed):
            print(f"  - {f}")

        # Amostra de dados (primeira linha com dados)
        print("\n--- AMOSTRA: Primeira linha de dados ---")
        for row_num in range(2, min(4, sheet.max_row + 1)):
            row_vals = [cell.value for cell in sheet[row_num]]
            if any(row_vals):
                print(f"\nLinha {row_num}:")
                for h, v in zip(headers_raw, row_vals):
                    if h and v is not None and str(v).strip():
                        print(f"  {repr(h)}: {v}")
                break

        print("\n" + "=" * 90)
        print("RECOMENDAÇÕES:")
        print("1. Se há colunas no Excel não importadas → adicione mapeamento em utils.upload_projecao_gastos_from_file")
        print("2. Se o Excel usa nomes diferentes → inclua variações em find_column_value")
        print("3. Campos extras podem ir em dados_adicionais (JSON) sem alterar o modelo")
        print("=" * 90)

    except Exception as e:
        print(f"[ERRO] {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python analise_import_projecao.py \"caminho/para/9. PCM - Previsão de Gastos 2025 e 2026.xlsx\"")
        print("\nExemplo:")
        print('  python analise_import_projecao.py "C:\\Users\\zeluzjunior\\Downloads\\9. PCM - Previsão de Gastos 2025 e 2026.xlsx"')
        sys.exit(1)
    analyze_excel(sys.argv[1])

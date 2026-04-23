"""
Utility functions for file uploads and data processing
"""
import csv
import io
import unicodedata
import warnings
from typing import List, Dict, Tuple
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction

# Suprimir aviso do openpyxl sobre Data Validation (não afeta a leitura)
warnings.filterwarnings('ignore', message='.*Data Validation.*')

import openpyxl


def read_excel_file(file, sheet_name=None, sheet_index=None, header_row=1):
    """
    Lê um arquivo Excel (.xlsx, .xls, .xlsm) e retorna os dados.
    
    Args:
        file: Arquivo Excel (Django UploadedFile ou path)
        sheet_name: Nome da planilha a ser lida (None para usar sheet_index ou planilha ativa)
        sheet_index: Índice 0-based da planilha (usa quando sheet_name é None; evita problemas de encoding no nome)
        header_row: Linha do cabeçalho (1-based). Muitos PCMs têm título na linha 1 e cabeçalho na linha 2.
    
    Returns:
        Lista de dicionários com os dados
    """
    try:
        import re
        import warnings
        from datetime import datetime, date
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            if hasattr(file, 'read'):
                file.seek(0)
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            else:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        
        if sheet_index is not None and 0 <= sheet_index < len(wb.worksheets):
            ws = wb.worksheets[sheet_index]
        elif sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Ler cabeçalhos da linha especificada
        header_row_cells = list(ws[header_row])
        headers = []
        for cell in header_row_cells:
            header_value = cell.value if cell.value is not None and str(cell.value).strip() else f'col_{len(headers)}'
            if isinstance(header_value, str):
                header_value = header_value.strip().replace('\xa0', ' ').replace('\u00a0', ' ')
                header_value = re.sub(r'\s+', ' ', header_value).strip()
            headers.append(str(header_value) if header_value else f'col_{len(headers)}')
        
        # Ler dados (a partir da linha seguinte ao cabeçalho)
        data = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for idx, cell_value in enumerate(row):
                    header = headers[idx] if idx < len(headers) else f'col_{idx}'
                    if isinstance(cell_value, (datetime, date)):
                        row_dict[header] = cell_value
                    elif isinstance(cell_value, str):
                        cell_value = cell_value.strip().replace('\xa0', ' ').replace('\u00a0', ' ')
                        cell_value = re.sub(r'\s+', ' ', cell_value).strip()
                        row_dict[header] = cell_value
                    else:
                        row_dict[header] = cell_value
                data.append(row_dict)
        
        return data
    
    except Exception as e:
        raise ValidationError(f"Erro ao ler arquivo Excel: {str(e)}")


def _strip_cell_csv(s: str) -> str:
    """Remove BOM, NBSP, zero-width e espaços — evita P 'vazio' com caracteres invisíveis."""
    if s is None:
        return ''
    t = str(s).replace('\ufeff', '').replace('\u00a0', ' ').replace('\u200b', '').replace('\u200c', '')
    return t.strip()


def _detect_plano_preventiva_p_q_r_indices(header_cells: List[str]) -> Tuple[int, int, int]:
    """
    Localiza colunas Funcionário (P), Nome Funcionário (Q) e coluna à direita (R) pelo cabeçalho.
    Fallback: Excel clássico P=15, Q=16, R=17 (índices 0-based).

    Importante: não usar o subtexto 'funcion' solto — em Python, 'funcion' in 'funcionamento'
    e em 'funcionalidade' é True, o que deslocava P para a coluna errada e quebrava o ajuste.
    """
    import unicodedata

    def norm(s: str) -> str:
        s = (s or '').strip().lower()
        s = unicodedata.normalize('NFKD', s)
        return ''.join(c for c in s if not unicodedata.combining(c))

    idx_p = None
    idx_q = None
    for i, h in enumerate(header_cells):
        hn = norm(str(h))
        # Q: "Nome Funcionário", "Nome do Funcionário", etc. (rejeita "Nome Funcionamento" se existir)
        if hn.startswith('nome') and 'funcionario' in hn and 'funcionamento' not in hn:
            idx_q = i
        # P: título exatamente "Funcionário" (normalizado → funcionario), não "Funcionamento"
        elif hn == 'funcionario':
            idx_p = i

    if idx_p is None:
        idx_p = 15
    if idx_q is None:
        idx_q = 16
    idx_r = idx_q + 1
    return idx_p, idx_q, idx_r


def _is_bc_only_row(stripped: List[str], idx_b: int = 1, idx_c: int = 2) -> bool:
    """
    Retorna True se a linha tem dados APENAS nas colunas B e C:
    B com números, C com nome (texto). Todas as outras colunas vazias.
    """
    if len(stripped) <= max(idx_b, idx_c):
        return False
    b_val = (stripped[idx_b] if idx_b < len(stripped) else '').strip()
    c_val = (stripped[idx_c] if idx_c < len(stripped) else '').strip()
    if not b_val or not c_val:
        return False
    b_clean = b_val.replace(',', '.').replace(' ', '').replace('\u00a0', '')
    try:
        float(b_clean)
        b_is_num = True
    except ValueError:
        b_is_num = False
    if not b_is_num:
        return False
    c_has_letter = any(c.isalpha() for c in c_val)
    if not c_has_letter:
        return False
    for i, v in enumerate(stripped):
        if i != idx_b and i != idx_c and (v and str(v).strip()):
            return False
    return True


def clean_plano_preventiva_delimitado_csv(text: str, delimiter: str = ';') -> Tuple[str, Dict[str, int]]:
    """
    Remove linhas inválidas do export PLANO_PREVENTIVA_DELIMITADO (CSV com ;).

    Etapa 1:
    - Remove apenas linhas totalmente vazias (só separadores / em branco).
    - (Não remove mais linhas “só B+C” para evitar perda de dados em exportações variadas.)

    Etapa 2:
    - Ajuste P/Q/R: se a coluna **R** (à direita de Nome Funcionário) tiver dado **e P estiver vazio**,
      copia Q→P e R→Q e esvazia R. Se P já tiver conteúdo, a linha não é alterada.
    - Índices P/Q/R vêm do **primeiro registro (cabeçalho)** quando reconhecível; senão usa 15/16/17.

    Retorna (texto CSV gerado, estatísticas).
    """
    import csv as _csv

    reader = _csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_out: List[List[str]] = []
    removed_blank = 0
    total_in = 0

    removed_bc_only = 0
    for row in reader:
        total_in += 1
        cells = [str(c) if c is not None else '' for c in row]
        stripped = [x.strip() for x in cells]

        if not any(stripped):
            removed_blank += 1
            continue

        if _is_bc_only_row(stripped):
            removed_bc_only += 1
            continue

        rows_out.append(cells)

    if not rows_out:
        buf = io.StringIO()
        return buf.getvalue(), {
            'total_in': total_in,
            'removed_blank': removed_blank,
            'removed_bc_only': removed_bc_only,
            'kept': 0,
            'adjusted_pq_rows': 0,
            'skipped_pq_nonempty_p': 0,
            'idx_p': 15,
            'idx_q': 16,
            'idx_r': 17,
        }

    idx_p, idx_q, idx_r = _detect_plano_preventiva_p_q_r_indices(rows_out[0])
    min_len = max(idx_p, idx_q, idx_r) + 1

    # Etapa 2: R tem dado e P vazio → Q→P, R→Q, R vazio (não sobrescrever P)
    adjusted_pq_rows = 0
    skipped_pq_nonempty_p = 0
    for row in rows_out:
        while len(row) < min_len:
            row.append('')
        r_has = _strip_cell_csv(row[idx_r]) if idx_r < len(row) else ''
        p_has = _strip_cell_csv(row[idx_p]) if idx_p < len(row) else ''
        if r_has:
            if not p_has:
                row[idx_p] = row[idx_q]
                row[idx_q] = row[idx_r]
                row[idx_r] = ''
                adjusted_pq_rows += 1
            else:
                skipped_pq_nonempty_p += 1

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter, lineterminator='\r\n', quoting=_csv.QUOTE_MINIMAL)
    for row in rows_out:
        writer.writerow(row)

    stats = {
        'total_in': total_in,
        'removed_blank': removed_blank,
        'removed_bc_only': removed_bc_only,
        'kept': len(rows_out),
        'adjusted_pq_rows': adjusted_pq_rows,
        'skipped_pq_nonempty_p': skipped_pq_nonempty_p,
        'idx_p': idx_p,
        'idx_q': idx_q,
        'idx_r': idx_r,
    }
    return buf.getvalue(), stats


def clean_roteiro_preventiva_csv(text: str, delimiter: str = ';') -> Tuple[str, Dict[str, int]]:
    """
    Remove linhas inválidas do export ROTEIRO_PREVENTIVA_DADOS (CSV com ;).

    Etapa 1: Remove linhas totalmente vazias.
    Etapa 2: Remove linhas que têm dados APENAS nas colunas A, B ou C (índices 0, 1, 2).
             Mantém apenas linhas com pelo menos um dado em coluna D ou superior (índice >= 3).

    Retorna (texto CSV gerado, estatísticas).
    """
    import csv as _csv

    reader = _csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_out: List[List[str]] = []
    removed_blank = 0
    removed_abc_only = 0
    total_in = 0

    for row in reader:
        total_in += 1
        cells = [str(c) if c is not None else '' for c in row]
        stripped = [x.strip() for x in cells]

        if not any(stripped):
            removed_blank += 1
            continue

        # Linha tem dados apenas em A, B ou C? (colunas D onwards vazias)
        has_data_from_d_onwards = any(
            s for i, s in enumerate(stripped) if i >= 3 and s
        )
        if not has_data_from_d_onwards:
            removed_abc_only += 1
            continue

        rows_out.append(cells)

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter, lineterminator='\r\n', quoting=_csv.QUOTE_MINIMAL)
    for row in rows_out:
        writer.writerow(row)

    stats = {
        'total_in': total_in,
        'removed_blank': removed_blank,
        'removed_abc_only': removed_abc_only,
        'kept': len(rows_out),
    }
    return buf.getvalue(), stats


def _parse_csv_dict_rows_from_text(content: str, delimiter: str) -> List[Dict]:
    """Parse texto CSV já decodificado em lista de dicts (mesma regra de read_csv_file)."""
    csv_reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    data = []
    for row in csv_reader:
        row_dict = {}
        for key, value in row.items():
            normalized_key = key.strip() if key else f'col_{len(row_dict)}'
            if value is not None:
                normalized_value = value.strip() if str(value).strip() else ''
            else:
                normalized_value = None
            row_dict[normalized_key] = normalized_value
        if any(v for v in row_dict.values() if v):
            data.append(row_dict)
    return data


def read_csv_file(file, encoding='utf-8', delimiter=','):
    """
    LÃª um arquivo CSV e retorna os dados
    
    Args:
        file: Arquivo CSV (Django UploadedFile ou path)
        encoding: Encoding do arquivo (padrÃ£o: utf-8)
        delimiter: Delimitador do CSV (padrÃ£o: vÃ­rgula)
    
    Returns:
        Lista de dicionÃ¡rios com os dados
    """
    try:
        if hasattr(file, 'read'):
            file.seek(0)
            content = file.read().decode(encoding)
        else:
            with open(file, 'r', encoding=encoding) as f:
                content = f.read()
        return _parse_csv_dict_rows_from_text(content, delimiter)
    except Exception as e:
        raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")


def upload_ordens_corretivas_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de ordens de serviÃ§o corretivas a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import OrdemServicoCorretiva, OrdemServicoCorretivaFicha
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
            # O arquivo usa delimitador ponto e vÃ­rgula (;) conforme instruÃ§Ãµes na pÃ¡gina
            data = None
            encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
            
            for encoding in encodings_to_try:
                try:
                    file.seek(0)  # Resetar arquivo para o inÃ­cio
                    data = read_csv_file(file, encoding=encoding, delimiter=';')
                    break  # Se conseguir ler, sair do loop
                except (UnicodeDecodeError, ValidationError) as e:
                    if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                        raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                    continue  # Tentar prÃ³ximo encoding
                except Exception as e:
                    # Outros erros (nÃ£o relacionados a encoding)
                    raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
            
            if data is None:
                raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_FUNCIOMANU;NOME_FUNCIOMANU;FUNCIOMANU_ID;CD_SETORMANUT;DESCR_SETORMANUT;...
                    
                    # Unidade
                    cd_unid = _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid'))
                    nome_unid = _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255)
                    
                    # FuncionÃ¡rio
                    cd_funciomanu = _safe_str(row_data.get('CD_FUNCIOMANU') or row_data.get('cd_funciomanu') or row_data.get('Cd_Funciomanu'), max_length=100)
                    nome_funciomanu = _safe_str(row_data.get('NOME_FUNCIOMANU') or row_data.get('nome_funciomanu') or row_data.get('Nome_Funciomanu'), max_length=255)
                    funciomanu_id = _safe_int(row_data.get('FUNCIOMANU_ID') or row_data.get('funciomanu_id') or row_data.get('Funciomanu_Id'))
                    
                    # Setor
                    cd_setormanut = _safe_str(row_data.get('CD_SETORMANUT') or row_data.get('cd_setormanut') or row_data.get('Cd_Setormanut'), max_length=50)
                    descr_setormanut = _safe_str(row_data.get('DESCR_SETORMANUT') or row_data.get('descr_setormanut') or row_data.get('Descr_Setormanut'), max_length=255)
                    
                    # Tipo Centro de Atividade
                    cd_tpcentativ = _safe_int(row_data.get('CD_TPCENTATIV') or row_data.get('cd_tpcentativ') or row_data.get('Cd_Tpcentativ'))
                    descr_abrev_tpcentativ = _safe_str(row_data.get('DESCR_ABREV_TPCENTATIV') or row_data.get('descr_abrev_tpcentativ') or row_data.get('Descr_Abrev_Tpcentativ'), max_length=255)
                    
                    # Ordem de ServiÃ§o
                    dt_abertura = _safe_str(row_data.get('DT_ABERTURA') or row_data.get('dt_abertura') or row_data.get('Dt_Abertura'), max_length=50)
                    cd_ordemserv = _safe_int(row_data.get('CD_ORDEMSERV') or row_data.get('cd_ordemserv') or row_data.get('Cd_Ordemserv'))
                    ordemserv_id = _safe_int(row_data.get('ORDEMSERV_ID') or row_data.get('ordemserv_id') or row_data.get('Ordemserv_Id'))
                    
                    # MÃ¡quina
                    cd_maquina = _safe_int(row_data.get('CD_MAQUINA') or row_data.get('cd_maquina') or row_data.get('Cd_Maquina'))
                    descr_maquina = _safe_str(row_data.get('DESCR_MAQUINA') or row_data.get('descr_maquina') or row_data.get('Descr_Maquina'), max_length=500)
                    
                    # Validar que temos pelo menos cÃ³digo da ordem de serviÃ§o
                    if not cd_ordemserv:
                        errors.append(f"Linha {row_num}: CÃ³digo da ordem de serviÃ§o (CD_ORDEMSERV) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    # Remover campos que nÃ£o existem no modelo
                    ordem_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'cd_setormanut': cd_setormanut,
                        'descr_setormanut': descr_setormanut,
                        'cd_tpcentativ': cd_tpcentativ,
                        'descr_abrev_tpcentativ': descr_abrev_tpcentativ,
                        'cd_ordemserv': cd_ordemserv,
                        'cd_maquina': cd_maquina,
                        'descr_maquina': descr_maquina,
                    }
                    
                    # Mapear campos de funcionÃ¡rio e data se existirem no CSV
                    # FuncionÃ¡rio solicitante (se disponÃ­vel no CSV)
                    if cd_funciomanu or nome_funciomanu:
                        ordem_data['cd_func_solic_os'] = cd_funciomanu
                        ordem_data['nm_func_solic_os'] = nome_funciomanu
                    
                    # Data de abertura (mapear para dt_aberordser se disponÃ­vel)
                    if dt_abertura:
                        ordem_data['dt_aberordser'] = dt_abertura
                    
                    # Adicionar outros campos do CSV se existirem
                    # Data Entrada
                    dt_entrada = _safe_str(row_data.get('DT_ENTRADA') or row_data.get('dt_entrada') or row_data.get('Dt_Entrada'), max_length=50)
                    if dt_entrada:
                        ordem_data['dt_entrada'] = dt_entrada
                    
                    # FuncionÃ¡rio Executor
                    cd_func_exec = _safe_str(row_data.get('CD_FUNC_EXEC') or row_data.get('cd_func_exec') or row_data.get('NM_FUNC_EXEC') or row_data.get('nm_func_exec'), max_length=100)
                    nm_func_exec = _safe_str(row_data.get('NM_FUNC_EXEC') or row_data.get('nm_func_exec') or row_data.get('NOME_FUNC_EXEC') or row_data.get('nome_func_exec'), max_length=255)
                    if cd_func_exec:
                        ordem_data['cd_func_exec'] = cd_func_exec
                    if nm_func_exec:
                        ordem_data['nm_func_exec'] = nm_func_exec
                    
                    # FuncionÃ¡rio Solicitante (se nÃ£o foi mapeado acima)
                    cd_func_solic = _safe_str(row_data.get('CD_FUNC_SOLIC_OS') or row_data.get('cd_func_solic_os') or row_data.get('NM_FUNC_SOLIC_OS') or row_data.get('nm_func_solic_os'), max_length=100)
                    nm_func_solic = _safe_str(row_data.get('NM_FUNC_SOLIC_OS') or row_data.get('nm_func_solic_os') or row_data.get('NOME_FUNC_SOLIC_OS') or row_data.get('nome_func_solic_os'), max_length=255)
                    if cd_func_solic and not ordem_data.get('cd_func_solic_os'):
                        ordem_data['cd_func_solic_os'] = cd_func_solic
                    if nm_func_solic and not ordem_data.get('nm_func_solic_os'):
                        ordem_data['nm_func_solic_os'] = nm_func_solic
                    
                    # Data Encerramento
                    dt_encordmanu = _safe_str(row_data.get('DT_ENCORDMANU') or row_data.get('dt_encordmanu') or row_data.get('Dt_Encordmanu'), max_length=50)
                    if dt_encordmanu:
                        ordem_data['dt_encordmanu'] = dt_encordmanu
                    
                    # DescriÃ§Ã£o Queixa
                    descr_queixa = _safe_str(row_data.get('DESCR_QUEIXA') or row_data.get('descr_queixa') or row_data.get('Descr_Queixa'))
                    if descr_queixa:
                        ordem_data['descr_queixa'] = descr_queixa
                    
                    # ExecuÃ§Ã£o Tarefas
                    exec_tarefas = _safe_str(row_data.get('EXEC_TAREFAS') or row_data.get('exec_tarefas') or row_data.get('Exec_Tarefas'))
                    if exec_tarefas:
                        ordem_data['exec_tarefas'] = exec_tarefas
                    
                    # Unidade ExecuÃ§Ã£o
                    cd_unid_exec = _safe_int(row_data.get('CD_UNID_EXEC') or row_data.get('cd_unid_exec') or row_data.get('Cd_Unid_Exec'))
                    nome_unid_exec = _safe_str(row_data.get('NOME_UNID_EXEC') or row_data.get('nome_unid_exec') or row_data.get('Nome_Unid_Exec'), max_length=255)
                    if cd_unid_exec:
                        ordem_data['cd_unid_exec'] = cd_unid_exec
                    if nome_unid_exec:
                        ordem_data['nome_unid_exec'] = nome_unid_exec
                    
                    # Data Abertura SolicitaÃ§Ã£o
                    dt_abertura_solicita = _safe_str(row_data.get('DT_ABERTURA_SOLICITA') or row_data.get('dt_abertura_solicita') or row_data.get('Dt_Abertura_Solicita'), max_length=50)
                    if dt_abertura_solicita:
                        ordem_data['dt_abertura_solicita'] = dt_abertura_solicita
                    
                    # ObservaÃ§Ãµes da Ordem de ServiÃ§o
                    descr_obsordserv = _safe_str(row_data.get('DESCR_OBSORDSERV') or row_data.get('descr_obsordserv') or row_data.get('Descr_Obsordserv'))
                    if descr_obsordserv:
                        ordem_data['descr_obsordserv'] = descr_obsordserv
                    
                    # Data Abertura Ordem ServiÃ§o
                    dt_aberordser = _safe_str(row_data.get('DT_ABERORDSER') or row_data.get('dt_aberordser') or row_data.get('Dt_Aberordser'), max_length=50)
                    if dt_aberordser:
                        ordem_data['dt_aberordser'] = dt_aberordser
                    
                    # Datas de Parada de ManutenÃ§Ã£o
                    dt_iniparmanu = _safe_str(row_data.get('DT_INIPARMANU') or row_data.get('dt_iniparmanu') or row_data.get('Dt_Iniparmanu'), max_length=50)
                    dt_fimparmanu = _safe_str(row_data.get('DT_FIMPARMANU') or row_data.get('dt_fimparmanu') or row_data.get('Dt_Fimparmanu'), max_length=50)
                    if dt_iniparmanu:
                        ordem_data['dt_iniparmanu'] = dt_iniparmanu
                    if dt_fimparmanu:
                        ordem_data['dt_fimparmanu'] = dt_fimparmanu
                    
                    # Data Prevista ExecuÃ§Ã£o
                    dt_prev_exec = _safe_str(row_data.get('DT_PREV_EXEC') or row_data.get('dt_prev_exec') or row_data.get('Dt_Prev_Exec'), max_length=50)
                    if dt_prev_exec:
                        ordem_data['dt_prev_exec'] = dt_prev_exec
                    
                    # Tipo de Ordem de ServiÃ§o
                    cd_tpordservtv = _safe_int(row_data.get('CD_TPORDSERTV') or row_data.get('cd_tpordservtv') or row_data.get('Cd_Tpordservtv'))
                    descr_tpordservtv = _safe_str(row_data.get('DESCR_TPORDSERTV') or row_data.get('descr_tpordservtv') or row_data.get('Descr_Tpordservtv'), max_length=255)
                    if cd_tpordservtv:
                        ordem_data['cd_tpordservtv'] = cd_tpordservtv
                    if descr_tpordservtv:
                        ordem_data['descr_tpordservtv'] = descr_tpordservtv
                    
                    # SituaÃ§Ã£o Ordem ServiÃ§o
                    descr_sitordsetv = _safe_str(row_data.get('DESCR_SITORDSETV') or row_data.get('descr_sitordsetv') or row_data.get('Descr_Sitordsetv'), max_length=255)
                    if descr_sitordsetv:
                        ordem_data['descr_sitordsetv'] = descr_sitordsetv
                    
                    # RecomendaÃ§Ãµes
                    descr_recomenos = _safe_str(row_data.get('DESCR_RECOMENOS') or row_data.get('descr_recomenos') or row_data.get('Descr_Recomenos'))
                    if descr_recomenos:
                        ordem_data['descr_recomenos'] = descr_recomenos
                    
                    # SequÃªncia Plano ManutenÃ§Ã£o
                    descr_seqplamanu = _safe_str(row_data.get('DESCR_SEQPLAMANU') or row_data.get('descr_seqplamanu') or row_data.get('Descr_Seqplamanu'), max_length=255)
                    if descr_seqplamanu:
                        ordem_data['descr_seqplamanu'] = descr_seqplamanu
                    
                    # Tipo de ManutenÃ§Ã£o
                    cd_tpmanuttv = _safe_int(row_data.get('CD_TPMANUTTV') or row_data.get('cd_tpmanuttv') or row_data.get('Cd_Tpmanuttv'))
                    descr_tpmanuttv = _safe_str(row_data.get('DESCR_TPMANUTTV') or row_data.get('descr_tpmanuttv') or row_data.get('Descr_Tpmanuttv'), max_length=255)
                    if cd_tpmanuttv:
                        ordem_data['cd_tpmanuttv'] = cd_tpmanuttv
                    if descr_tpmanuttv:
                        ordem_data['descr_tpmanuttv'] = descr_tpmanuttv
                    
                    # ClassificaÃ§Ã£o Origem OS
                    cd_clasorigos = _safe_int(row_data.get('CD_CLASORIGOS') or row_data.get('cd_clasorigos') or row_data.get('Cd_Clasorigos'))
                    descr_clasorigos = _safe_str(row_data.get('DESCR_CLASORIGOS') or row_data.get('descr_clasorigos') or row_data.get('Descr_Clasorigos'), max_length=255)
                    if cd_clasorigos:
                        ordem_data['cd_clasorigos'] = cd_clasorigos
                    if descr_clasorigos:
                        ordem_data['descr_clasorigos'] = descr_clasorigos
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        ordem_obj, created = OrdemServicoCorretiva.objects.update_or_create(
                            cd_ordemserv=cd_ordemserv,
                            defaults=ordem_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        ordem_obj, created = OrdemServicoCorretiva.objects.get_or_create(
                            cd_ordemserv=cd_ordemserv,
                            defaults=ordem_data
                        )
                        if created:
                            created_count += 1
                    
                    # Criar Ficha de ManutenÃ§Ã£o se os campos estiverem presentes no CSV
                    cd_func_exec_os = _safe_str(row_data.get('CD_FUNC_EXEC_OS') or row_data.get('cd_func_exec_os') or row_data.get('Cd_Func_Exec_Os'), max_length=100)
                    nm_func_exec_os = _safe_str(row_data.get('NM_FUNC_EXEC_OS') or row_data.get('nm_func_exec_os') or row_data.get('Nm_Func_Exec_Os'), max_length=255)
                    dt_ficapomanu = _safe_str(row_data.get('DT_FICAPOMANU') or row_data.get('dt_ficapomanu') or row_data.get('Dt_Ficapomanu'), max_length=50)
                    dt_inic_iteficmanu = _safe_str(row_data.get('DT_INIC_ITEFICMANU') or row_data.get('dt_inic_iteficmanu') or row_data.get('Dt_Inic_Iteficmanu'), max_length=50)
                    dt_fim_iteficmanu = _safe_str(row_data.get('DT_FIM_ITEFICMANU') or row_data.get('dt_fim_iteficmanu') or row_data.get('Dt_Fim_Iteficmanu'), max_length=50)
                    
                    # Criar ficha apenas se houver pelo menos um campo de ficha preenchido
                    if cd_func_exec_os or nm_func_exec_os or dt_ficapomanu or dt_inic_iteficmanu or dt_fim_iteficmanu:
                        ficha_data = {
                            'ordem_servico': ordem_obj,
                        }
                        if cd_func_exec_os:
                            ficha_data['cd_func_exec_os'] = cd_func_exec_os
                        if nm_func_exec_os:
                            ficha_data['nm_func_exec_os'] = nm_func_exec_os
                        if dt_ficapomanu:
                            ficha_data['dt_ficapomanu'] = dt_ficapomanu
                        if dt_inic_iteficmanu:
                            ficha_data['dt_inic_iteficmanu'] = dt_inic_iteficmanu
                        if dt_fim_iteficmanu:
                            ficha_data['dt_fim_iteficmanu'] = dt_fim_iteficmanu
                        
                        # Criar ficha (permitir mÃºltiplas fichas para a mesma ordem)
                        try:
                            OrdemServicoCorretivaFicha.objects.create(**ficha_data)
                        except Exception as e:
                            errors.append(f"Linha {row_num}: Erro ao criar ficha de manutenÃ§Ã£o - {str(e)}")
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        return 0, 0, errors


def upload_ordens_preventivas_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de ordens de serviço preventivas a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import OrdemServicoPreventiva, OrdemServicoPreventivaFicha
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensão
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes encodings - começar com latin-1 (mais comum para arquivos brasileiros)
            # O arquivo usa delimitador ponto e vírgula (;) conforme instruções na página
            data = None
            encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
            
            for encoding in encodings_to_try:
                try:
                    file.seek(0)  # Resetar arquivo para o início
                    data = read_csv_file(file, encoding=encoding, delimiter=';')
                    break  # Se conseguir ler, sair do loop
                except (UnicodeDecodeError, ValidationError) as e:
                    if encoding == encodings_to_try[-1]:  # Se for o último encoding
                        raise ValidationError(f"Erro ao ler arquivo CSV: Não foi possível decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                    continue  # Tentar próximo encoding
                except Exception as e:
                    # Outros erros (não relacionados a encoding)
                    raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
            
            if data is None:
                raise ValidationError("Erro ao ler arquivo CSV: Não foi possível processar o arquivo.")
        else:
            raise ValidationError("Formato de arquivo não suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados válidos")
        
        # Processar dados em transação
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # Começar em 2 (linha 1 é cabeçalho)
                try:
                    # Verificar se a linha está vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # Unidade
                    cd_unid = _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid'))
                    nome_unid = _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255)
                    cd_unid_exec = _safe_int(row_data.get('CD_UNID_EXEC') or row_data.get('cd_unid_exec') or row_data.get('Cd_Unid_Exec'))
                    nome_unid_exec = _safe_str(row_data.get('NOME_UNID_EXEC') or row_data.get('nome_unid_exec') or row_data.get('Nome_Unid_Exec'), max_length=255)
                    
                    # Setor de Manutenção
                    cd_setormanut = _safe_str(row_data.get('CD_SETORMANUT') or row_data.get('cd_setormanut') or row_data.get('Cd_Setormanut'), max_length=50)
                    descr_setormanut = _safe_str(row_data.get('DESCR_SETORMANUT') or row_data.get('descr_setormanut') or row_data.get('Descr_Setormanut'), max_length=255)
                    
                    # Centro de Atividade
                    cd_tpcentativ = _safe_int(row_data.get('CD_TPCENTATIV') or row_data.get('cd_tpcentativ') or row_data.get('Cd_Tpcentativ'))
                    descr_abrev_tpcentativ = _safe_str(row_data.get('DESCR_ABREV_TPCENTATIV') or row_data.get('descr_abrev_tpcentativ') or row_data.get('Descr_Abrev_Tpcentativ'), max_length=255)
                    
                    # Máquina
                    cd_maquina = _safe_int(row_data.get('CD_MAQUINA') or row_data.get('cd_maquina') or row_data.get('Cd_Maquina'))
                    descr_maquina = _safe_str(row_data.get('DESCR_MAQUINA') or row_data.get('descr_maquina') or row_data.get('Descr_Maquina'), max_length=500)
                    
                    # Ordem de Serviço
                    cd_ordemserv = _safe_int(row_data.get('CD_ORDEMSERV') or row_data.get('cd_ordemserv') or row_data.get('Cd_Ordemserv'))
                    
                    # Validar que temos pelo menos código da ordem de serviço
                    if not cd_ordemserv:
                        errors.append(f"Linha {row_num}: Código da ordem de serviço (CD_ORDEMSERV) é obrigatório")
                        continue
                    
                    # Preparar dados para criação/atualização
                    ordem_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'cd_unid_exec': cd_unid_exec,
                        'nome_unid_exec': nome_unid_exec,
                        'cd_setormanut': cd_setormanut,
                        'descr_setormanut': descr_setormanut,
                        'cd_tpcentativ': cd_tpcentativ,
                        'descr_abrev_tpcentativ': descr_abrev_tpcentativ,
                        'cd_ordemserv': cd_ordemserv,
                        'cd_maquina': cd_maquina,
                        'descr_maquina': descr_maquina,
                    }
                    
                    # Datas
                    dt_entrada = _safe_str(row_data.get('DT_ENTRADA') or row_data.get('dt_entrada') or row_data.get('Dt_Entrada'), max_length=50)
                    dt_abertura_solicita = _safe_str(row_data.get('DT_ABERTURA_SOLICITA') or row_data.get('dt_abertura_solicita') or row_data.get('Dt_Abertura_Solicita'), max_length=50)
                    dt_encordmanu = _safe_str(row_data.get('DT_ENCORDMANU') or row_data.get('dt_encordmanu') or row_data.get('Dt_Encordmanu'), max_length=50)
                    dt_aberordser = _safe_str(row_data.get('DT_ABERORDSER') or row_data.get('dt_aberordser') or row_data.get('Dt_Aberordser'), max_length=50)
                    dt_iniparmanu = _safe_str(row_data.get('DT_INIPARMANU') or row_data.get('dt_iniparmanu') or row_data.get('Dt_Iniparmanu'), max_length=50)
                    dt_fimparmanu = _safe_str(row_data.get('DT_FIMPARMANU') or row_data.get('dt_fimparmanu') or row_data.get('Dt_Fimparmanu'), max_length=50)
                    dt_prev_exec = _safe_str(row_data.get('DT_PREV_EXEC') or row_data.get('dt_prev_exec') or row_data.get('Dt_Prev_Exec'), max_length=50)
                    
                    if dt_entrada:
                        ordem_data['dt_entrada'] = dt_entrada
                    if dt_abertura_solicita:
                        ordem_data['dt_abertura_solicita'] = dt_abertura_solicita
                    if dt_encordmanu:
                        ordem_data['dt_encordmanu'] = dt_encordmanu
                    if dt_aberordser:
                        ordem_data['dt_aberordser'] = dt_aberordser
                    if dt_iniparmanu:
                        ordem_data['dt_iniparmanu'] = dt_iniparmanu
                    if dt_fimparmanu:
                        ordem_data['dt_fimparmanu'] = dt_fimparmanu
                    if dt_prev_exec:
                        ordem_data['dt_prev_exec'] = dt_prev_exec
                    
                    # Funcionários
                    cd_func_solic_os = _safe_str(row_data.get('CD_FUNC_SOLIC_OS') or row_data.get('cd_func_solic_os') or row_data.get('Cd_Func_Solic_Os'), max_length=100)
                    nm_func_solic_os = _safe_str(row_data.get('NM_FUNC_SOLIC_OS') or row_data.get('nm_func_solic_os') or row_data.get('Nm_Func_Solic_Os'), max_length=255)
                    cd_func_exec = _safe_str(row_data.get('CD_FUNC_EXEC') or row_data.get('cd_func_exec') or row_data.get('Cd_Func_Exec'), max_length=100)
                    nm_func_exec = _safe_str(row_data.get('NM_FUNC_EXEC') or row_data.get('nm_func_exec') or row_data.get('Nm_Func_Exec'), max_length=255)
                    
                    if cd_func_solic_os:
                        ordem_data['cd_func_solic_os'] = cd_func_solic_os
                    if nm_func_solic_os:
                        ordem_data['nm_func_solic_os'] = nm_func_solic_os
                    if cd_func_exec:
                        ordem_data['cd_func_exec'] = cd_func_exec
                    if nm_func_exec:
                        ordem_data['nm_func_exec'] = nm_func_exec
                    
                    # Descrições
                    descr_queixa = _safe_str(row_data.get('DESCR_QUEIXA') or row_data.get('descr_queixa') or row_data.get('Descr_Queixa'))
                    exec_tarefas = _safe_str(row_data.get('EXEC_TAREFAS') or row_data.get('exec_tarefas') or row_data.get('Exec_Tarefas'))
                    descr_obsordserv = _safe_str(row_data.get('DESCR_OBSORDSERV') or row_data.get('descr_obsordserv') or row_data.get('Descr_Obsordserv'))
                    
                    if descr_queixa:
                        ordem_data['descr_queixa'] = descr_queixa
                    if exec_tarefas:
                        ordem_data['exec_tarefas'] = exec_tarefas
                    if descr_obsordserv:
                        ordem_data['descr_obsordserv'] = descr_obsordserv
                    
                    # Tipo de Ordem de Serviço
                    cd_tpordservtv = _safe_int(row_data.get('CD_TPORDSERTV') or row_data.get('cd_tpordservtv') or row_data.get('Cd_Tpordservtv'))
                    descr_tpordservtv = _safe_str(row_data.get('DESCR_TPORDSERTV') or row_data.get('descr_tpordservtv') or row_data.get('Descr_Tpordservtv'), max_length=255)
                    descr_sitordsetv = _safe_str(row_data.get('DESCR_SITORDSETV') or row_data.get('descr_sitordsetv') or row_data.get('Descr_Sitordsetv'), max_length=255)
                    
                    if cd_tpordservtv:
                        ordem_data['cd_tpordservtv'] = cd_tpordservtv
                    if descr_tpordservtv:
                        ordem_data['descr_tpordservtv'] = descr_tpordservtv
                    if descr_sitordsetv:
                        ordem_data['descr_sitordsetv'] = descr_sitordsetv
                    
                    # Recomendações e Sequência
                    descr_recomenos = _safe_str(row_data.get('DESCR_RECOMENOS') or row_data.get('descr_recomenos') or row_data.get('Descr_Recomenos'))
                    descr_seqplamanu = _safe_str(row_data.get('DESCR_SEQPLAMANU') or row_data.get('descr_seqplamanu') or row_data.get('Descr_Seqplamanu'), max_length=255)
                    
                    if descr_recomenos:
                        ordem_data['descr_recomenos'] = descr_recomenos
                    if descr_seqplamanu:
                        ordem_data['descr_seqplamanu'] = descr_seqplamanu
                    
                    # Tipo de Manutenção
                    cd_tpmanuttv = _safe_int(row_data.get('CD_TPMANUTTV') or row_data.get('cd_tpmanuttv') or row_data.get('Cd_Tpmanuttv'))
                    descr_tpmanuttv = _safe_str(row_data.get('DESCR_TPMANUTTV') or row_data.get('descr_tpmanuttv') or row_data.get('Descr_Tpmanuttv'), max_length=255)
                    
                    if cd_tpmanuttv:
                        ordem_data['cd_tpmanuttv'] = cd_tpmanuttv
                    if descr_tpmanuttv:
                        ordem_data['descr_tpmanuttv'] = descr_tpmanuttv
                    
                    # Classificação Origem OS
                    cd_clasorigos = _safe_int(row_data.get('CD_CLASORIGOS') or row_data.get('cd_clasorigos') or row_data.get('Cd_Clasorigos'))
                    descr_clasorigos = _safe_str(row_data.get('DESCR_CLASORIGOS') or row_data.get('descr_clasorigos') or row_data.get('Descr_Clasorigos'), max_length=255)
                    
                    if cd_clasorigos:
                        ordem_data['cd_clasorigos'] = cd_clasorigos
                    if descr_clasorigos:
                        ordem_data['descr_clasorigos'] = descr_clasorigos
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        ordem_obj, created = OrdemServicoPreventiva.objects.update_or_create(
                            cd_ordemserv=cd_ordemserv,
                            defaults=ordem_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        ordem_obj, created = OrdemServicoPreventiva.objects.get_or_create(
                            cd_ordemserv=cd_ordemserv,
                            defaults=ordem_data
                        )
                        if created:
                            created_count += 1
                    
                    # Criar Ficha de Manutenção se os campos estiverem presentes no CSV
                    cd_func_exec_os = _safe_str(row_data.get('CD_FUNC_EXEC_OS') or row_data.get('cd_func_exec_os') or row_data.get('Cd_Func_Exec_Os'), max_length=100)
                    nm_func_exec_os = _safe_str(row_data.get('NM_FUNC_EXEC_OS') or row_data.get('nm_func_exec_os') or row_data.get('Nm_Func_Exec_Os'), max_length=255)
                    dt_ficapomanu = _safe_str(row_data.get('DT_FICAPOMANU') or row_data.get('dt_ficapomanu') or row_data.get('Dt_Ficapomanu'), max_length=50)
                    dt_inic_iteficmanu = _safe_str(row_data.get('DT_INIC_ITEFICMANU') or row_data.get('dt_inic_iteficmanu') or row_data.get('Dt_Inic_Iteficmanu'), max_length=50)
                    dt_fim_iteficmanu = _safe_str(row_data.get('DT_FIM_ITEFICMANU') or row_data.get('dt_fim_iteficmanu') or row_data.get('Dt_Fim_Iteficmanu'), max_length=50)
                    
                    # Criar ficha apenas se houver pelo menos um campo de ficha preenchido
                    if cd_func_exec_os or nm_func_exec_os or dt_ficapomanu or dt_inic_iteficmanu or dt_fim_iteficmanu:
                        ficha_data = {
                            'ordem_servico': ordem_obj,
                        }
                        if cd_func_exec_os:
                            ficha_data['cd_func_exec_os'] = cd_func_exec_os
                        if nm_func_exec_os:
                            ficha_data['nm_func_exec_os'] = nm_func_exec_os
                        if dt_ficapomanu:
                            ficha_data['dt_ficapomanu'] = dt_ficapomanu
                        if dt_inic_iteficmanu:
                            ficha_data['dt_inic_iteficmanu'] = dt_inic_iteficmanu
                        if dt_fim_iteficmanu:
                            ficha_data['dt_fim_iteficmanu'] = dt_fim_iteficmanu
                        
                        # Criar ficha (permitir múltiplas fichas para a mesma ordem)
                        try:
                            OrdemServicoPreventivaFicha.objects.create(**ficha_data)
                        except Exception as e:
                            errors.append(f"Linha {row_num}: Erro ao criar ficha de manutenção - {str(e)}")
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        return 0, 0, errors


def upload_ordens_lubrificacao_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de ordens de serviço de lubrificação a partir de arquivo CSV ou Excel.
    Aceita lubrificacao_aberta.csv e lubrificacao_fechada.csv (mesma estrutura).
    Ordem -> OrdemServicoLubrificacao; Ficha/Apontamento -> OrdemServicoLubrificacaoFicha.
    """
    from app.models import OrdemServicoLubrificacao, OrdemServicoLubrificacaoFicha

    errors = []
    created_count = 0
    updated_count = 0
    file_name = file.name.lower()

    try:
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            data = None
            for enc in ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']:
                try:
                    file.seek(0)
                    data = read_csv_file(file, encoding=enc, delimiter=';')
                    break
                except (UnicodeDecodeError, ValidationError):
                    if enc == 'cp1252':
                        raise ValidationError("Erro ao ler CSV: não foi possível decodificar.")
                    continue
            if data is None:
                raise ValidationError("Erro ao ler arquivo CSV.")
        else:
            raise ValidationError("Formato não suportado. Use .xlsx, .xls, .xlsm ou .csv")

        if not data:
            raise ValidationError("Arquivo vazio ou sem dados válidos")

        def _g(d, *keys):
            for k in keys:
                v = d.get(k) or d.get(k.lower()) or d.get(k.replace('_', ' '))
                if v is not None and str(v).strip():
                    return v
            return None

        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):
                try:
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue

                    cd_ordemserv = _safe_int(_g(row_data, 'CD_ORDEMSERV', 'cd_ordemserv'))
                    if not cd_ordemserv:
                        errors.append(f"Linha {row_num}: CD_ORDEMSERV é obrigatório")
                        continue

                    ordem_data = {
                        'cd_unid': _safe_int(_g(row_data, 'CD_UNID')),
                        'nome_unid': _safe_str(_g(row_data, 'NOME_UNID'), 255),
                        'cd_unid_exec': _safe_int(_g(row_data, 'CD_UNID_EXEC')),
                        'nome_unid_exec': _safe_str(_g(row_data, 'NOME_UNID_EXEC'), 255),
                        'cd_setormanut': _safe_str(_g(row_data, 'CD_SETORMANUT'), 50),
                        'descr_setormanut': _safe_str(_g(row_data, 'DESCR_SETORMANUT'), 255),
                        'cd_tpcentativ': _safe_int(_g(row_data, 'CD_TPCENTATIV')),
                        'descr_abrev_tpcentativ': _safe_str(_g(row_data, 'DESCR_ABREV_TPCENTATIV'), 255),
                        'cd_maquina': _safe_int(_g(row_data, 'CD_MAQUINA')),
                        'descr_maquina': _safe_str(_g(row_data, 'DESCR_MAQUINA'), 500),
                        'cd_ordemserv': cd_ordemserv,
                    }
                    for fld, key in [
                        ('dt_entrada', 'DT_ENTRADA'), ('dt_abertura_solicita', 'DT_ABERTURA_SOLICITA'),
                        ('cd_func_solic_os', 'CD_FUNC_SOLIC_OS'), ('nm_func_solic_os', 'NM_FUNC_SOLIC_OS'),
                        ('descr_queixa', 'DESCR_QUEIXA'), ('exec_tarefas', 'EXEC_TAREFAS'),
                        ('cd_func_exec', 'CD_FUNC_EXEC'), ('nm_func_exec', 'NM_FUNC_EXEC'),
                        ('descr_obsordserv', 'DESCR_OBSORDSERV'), ('dt_encordmanu', 'DT_ENCORDMANU'),
                        ('dt_aberordser', 'DT_ABERORDSER'), ('dt_iniparmanu', 'DT_INIPARMANU'),
                        ('dt_fimparmanu', 'DT_FIMPARMANU'), ('dt_prev_exec', 'DT_PREV_EXEC'),
                        ('cd_tpordservtv', 'CD_TPORDSERTV'), ('descr_tpordservtv', 'DESCR_TPORDSERTV'),
                        ('descr_sitordsetv', 'DESCR_SITORDSETV'), ('descr_recomenos', 'DESCR_RECOMENOS'),
                        ('descr_seqplamanu', 'DESCR_SEQPLAMANU'), ('cd_tpmanuttv', 'CD_TPMANUTTV'),
                        ('descr_tpmanuttv', 'DESCR_TPMANUTTV'), ('cd_clasorigos', 'CD_CLASORIGOS'),
                        ('descr_clasorigos', 'DESCR_CLASORIGOS'),
                    ]:
                        v = _g(row_data, key)
                        if v is not None:
                            if fld.startswith('cd_') and fld != 'cd_ordemserv':
                                ordem_data[fld] = _safe_int(v)
                            elif fld.startswith('dt_'):
                                ordem_data[fld] = _safe_str(v, 50)
                            else:
                                ordem_data[fld] = _safe_str(v, 500 if 'descr_queixa' in fld or 'exec' in fld else 255)

                    if update_existing:
                        ordem_obj, created = OrdemServicoLubrificacao.objects.update_or_create(
                            cd_ordemserv=cd_ordemserv, defaults=ordem_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        ordem_obj, created = OrdemServicoLubrificacao.objects.get_or_create(
                            cd_ordemserv=cd_ordemserv, defaults=ordem_data
                        )
                        if created:
                            created_count += 1

                    cd_func_exec_os = _safe_str(_g(row_data, 'CD_FUNC_EXEC_OS'), 100)
                    nm_func_exec_os = _safe_str(_g(row_data, 'NM_FUNC_EXEC_OS'), 255)
                    dt_ficapomanu = _safe_str(_g(row_data, 'DT_FICAPOMANU'), 50)
                    dt_inic_iteficmanu = _safe_str(_g(row_data, 'DT_INIC_ITEFICMANU'), 50)
                    dt_fim_iteficmanu = _safe_str(_g(row_data, 'DT_FIM_ITEFICMANU'), 50)

                    if cd_func_exec_os or nm_func_exec_os or dt_ficapomanu or dt_inic_iteficmanu or dt_fim_iteficmanu:
                        ficha_data = {'ordem_servico': ordem_obj}
                        if cd_func_exec_os:
                            ficha_data['cd_func_exec_os'] = cd_func_exec_os
                        if nm_func_exec_os:
                            ficha_data['nm_func_exec_os'] = nm_func_exec_os
                        if dt_ficapomanu:
                            ficha_data['dt_ficapomanu'] = dt_ficapomanu
                        if dt_inic_iteficmanu:
                            ficha_data['dt_inic_iteficmanu'] = dt_inic_iteficmanu
                        if dt_fim_iteficmanu:
                            ficha_data['dt_fim_iteficmanu'] = dt_fim_iteficmanu
                        try:
                            OrdemServicoLubrificacaoFicha.objects.create(**ficha_data)
                        except Exception as e:
                            errors.append(f"Linha {row_num}: Erro ao criar ficha - {str(e)}")

                except Exception as e:
                    errors.append(f"Linha {row_num}: {str(e)}")

        return created_count, updated_count, errors

    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        errors.append(f"Erro geral: {str(e)}")
        return 0, 0, errors


def upload_ordens_geral_from_file(file, update_existing=False) -> Dict[str, object]:
    """
    Importa um único arquivo com ordens mistas e direciona para os importadores corretos.

    Regras de roteamento por tipo:
    - PREVENTIVA -> upload_ordens_preventivas_from_file
    - CORRETIVA/OUTROS -> upload_ordens_corretivas_from_file
    - LUBRIFICACAO/LUBRIFICAÇÃO -> upload_ordens_lubrificacao_from_file
    """
    file_name = file.name.lower()
    errors: List[str] = []

    if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
        data = read_excel_file(file)
    elif file_name.endswith('.csv'):
        data = None
        for enc in ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']:
            try:
                file.seek(0)
                data = read_csv_file(file, encoding=enc, delimiter=';')
                break
            except (UnicodeDecodeError, ValidationError):
                if enc == 'cp1252':
                    raise ValidationError("Erro ao ler CSV: não foi possível decodificar.")
                continue
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV.")
    else:
        raise ValidationError("Formato não suportado. Use .xlsx, .xls, .xlsm ou .csv")

    if not data:
        raise ValidationError("Arquivo vazio ou sem dados válidos")

    def _normalize_text(value) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        if not text:
            return ''
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(c for c in text if not unicodedata.combining(c))
        return text.upper()

    def _resolve_target(row_data: Dict) -> str:
        descr = row_data.get('DESCR_TPORDSERTV') or row_data.get('descr_tpordservtv') or row_data.get('Descr_Tpordservtv')
        code = row_data.get('CD_TPORDSERTV') or row_data.get('cd_tpordservtv') or row_data.get('Cd_Tpordservtv')
        descr_norm = _normalize_text(descr)
        code_norm = _normalize_text(code)

        if 'PREVENTIVA' in descr_norm or code_norm == '2':
            return 'preventiva'
        if 'LUBRIFICAC' in descr_norm:
            return 'lubrificacao'
        if 'CORRETIVA' in descr_norm or 'OUTROS' in descr_norm or code_norm in ('1', '3'):
            return 'corretiva_outros'
        return 'unknown'

    rows_by_target = {
        'corretiva_outros': [],
        'preventiva': [],
        'lubrificacao': [],
    }
    ignored_count = 0

    for idx, row_data in enumerate(data, start=2):
        if not any(str(v).strip() if v else '' for v in row_data.values()):
            continue
        target = _resolve_target(row_data)
        if target in rows_by_target:
            rows_by_target[target].append(row_data)
        else:
            ignored_count += 1
            cd_ordemserv = row_data.get('CD_ORDEMSERV') or row_data.get('cd_ordemserv') or 'sem_cd_ordemserv'
            errors.append(f"Linha {idx}: tipo de ordem não reconhecido para CD_ORDEMSERV {cd_ordemserv}.")

    def _rows_to_uploaded_csv(rows: List[Dict], upload_name: str):
        if not rows:
            return None
        headers: List[str] = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers, delimiter=';', extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow({k: '' if v is None else v for k, v in row.items()})
        content = buffer.getvalue().encode('latin-1', errors='ignore')
        return SimpleUploadedFile(upload_name, content, content_type='text/csv')

    counts = {
        'corretiva_outros': {'created': 0, 'updated': 0},
        'preventiva': {'created': 0, 'updated': 0},
        'lubrificacao': {'created': 0, 'updated': 0},
    }

    file_corretiva = _rows_to_uploaded_csv(rows_by_target['corretiva_outros'], 'ordens_corretivas_outros_auto.csv')
    if file_corretiva:
        created, updated, err = upload_ordens_corretivas_from_file(file_corretiva, update_existing=update_existing)
        counts['corretiva_outros']['created'] += created
        counts['corretiva_outros']['updated'] += updated
        errors.extend(err)

    file_preventiva = _rows_to_uploaded_csv(rows_by_target['preventiva'], 'ordens_preventivas_auto.csv')
    if file_preventiva:
        created, updated, err = upload_ordens_preventivas_from_file(file_preventiva, update_existing=update_existing)
        counts['preventiva']['created'] += created
        counts['preventiva']['updated'] += updated
        errors.extend(err)

    file_lubrificacao = _rows_to_uploaded_csv(rows_by_target['lubrificacao'], 'ordens_lubrificacao_auto.csv')
    if file_lubrificacao:
        created, updated, err = upload_ordens_lubrificacao_from_file(file_lubrificacao, update_existing=update_existing)
        counts['lubrificacao']['created'] += created
        counts['lubrificacao']['updated'] += updated
        errors.extend(err)

    created_total = sum(item['created'] for item in counts.values())
    updated_total = sum(item['updated'] for item in counts.values())

    return {
        'created_total': created_total,
        'updated_total': updated_total,
        'ignored_count': ignored_count,
        'counts': counts,
        'errors': errors,
    }


def upload_maquinas_from_file(file, update_existing=False, update_fields=None) -> Tuple[int, int, List[str], List]:
    """
    Faz upload de mÃ¡quinas a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
        update_fields: Lista de campos a serem atualizados. Se None, atualiza todos os campos.
                      Se update_existing=False, este parÃ¢metro Ã© ignorado.
    
    Returns:
        Tupla (created_count, updated_count, errors, missing_machines)
        missing_machines: Lista de máquinas que estão no DB mas não no arquivo (apenas ativas)
    """
    from app.models import Maquina
    
    errors = []
    created_count = 0
    updated_count = 0
    cd_maquinas_in_file = set()  # Para rastrear quais máquinas estão no arquivo
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes encodings e delimitadores (; usado em MAQUINA_TODOS.csv e outros CSVs brasileiros)
            def _read_maq_csv(enc='utf-8', delim=','):
                file.seek(0)
                return read_csv_file(file, encoding=enc, delimiter=delim)

            data = None
            for encoding in ('utf-8', 'latin-1'):
                for delimiter in (';', ','):
                    try:
                        data = _read_maq_csv(enc=encoding, delim=delimiter)
                        if data and len(data) > 0 and any('CD_MAQUINA' in str(k).upper() for k in data[0].keys()):
                            break
                        data = None
                    except (UnicodeDecodeError, ValidationError):
                        data = None
                        continue
                if data:
                    break
            if not data:
                raise ValidationError("Erro ao ler CSV. Verifique encoding (UTF-8/Latin-1) e delimitador (; ou ,).")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CS_TT_MAQUINA;DESCR_MAQUINA;CD_MAQUINA;...
                    
                    # Unidade
                    cd_unid = _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid'))
                    nome_unid = _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255)
                    
                    # MÃ¡quina
                    cs_tt_maquina = _safe_int(row_data.get('CS_TT_MAQUINA') or row_data.get('cs_tt_maquina') or row_data.get('Cs_Tt_Maquina'))
                    descr_maquina = _safe_str(row_data.get('DESCR_MAQUINA') or row_data.get('descr_maquina') or row_data.get('Descr_Maquina'), max_length=500)
                    cd_maquina = _safe_int(row_data.get('CD_MAQUINA') or row_data.get('cd_maquina') or row_data.get('Cd_Maquina'))
                    
                    # Setor ManutenÃ§Ã£o
                    cd_setormanut = _safe_str(row_data.get('CD_SETORMANUT') or row_data.get('cd_setormanut') or row_data.get('Cd_Setormanut'), max_length=50)
                    descr_setormanut = _safe_str(row_data.get('DESCR_SETORMANUT') or row_data.get('descr_setormanut') or row_data.get('Descr_Setormanut'), max_length=255)
                    
                    # Prioridade
                    cd_priomaqutv = _safe_int(row_data.get('CD_PRIOMAQUTV') or row_data.get('cd_priomaqutv') or row_data.get('Cd_Priomaqutv'))
                    
                    # PatrimÃ´nio
                    nro_patrimonio = _safe_str(row_data.get('NRO_PATRIMONIO') or row_data.get('nro_patrimonio') or row_data.get('Nro_Patrimonio'), max_length=100)
                    
                    # Modelo e Grupo
                    cd_modelo = _safe_int(row_data.get('CD_MODELO') or row_data.get('cd_modelo') or row_data.get('Cd_Modelo'))
                    cd_grupo = _safe_int(row_data.get('CD_GRUPO') or row_data.get('cd_grupo') or row_data.get('Cd_Grupo'))
                    
                    # Tipo Centro de Atividade
                    cd_tpcentativ = _safe_int(row_data.get('CD_TPCENTATIV') or row_data.get('cd_tpcentativ') or row_data.get('Cd_Tpcentativ'))
                    
                    # GerÃªncia
                    descr_gerenc = _safe_str(row_data.get('DESCR_GERENC') or row_data.get('descr_gerenc') or row_data.get('Descr_Gerenc'), max_length=255)
                    
                    # Validar que temos pelo menos cÃ³digo da mÃ¡quina
                    if not cd_maquina:
                        errors.append(f"Linha {row_num}: CÃ³digo da mÃ¡quina (CD_MAQUINA) Ã© obrigatÃ³rio")
                        continue
                    
                    # Adicionar cd_maquina ao conjunto de máquinas no arquivo
                    cd_maquinas_in_file.add(cd_maquina)
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    maquina_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'cs_tt_maquina': cs_tt_maquina,
                        'descr_maquina': descr_maquina,
                        'cd_setormanut': cd_setormanut,
                        'descr_setormanut': descr_setormanut,
                        'cd_priomaqutv': cd_priomaqutv,
                        'nro_patrimonio': nro_patrimonio,
                        'cd_modelo': cd_modelo,
                        'cd_grupo': cd_grupo,
                        'cd_tpcentativ': cd_tpcentativ,
                        'descr_gerenc': descr_gerenc,
                        'ativo': True,  # Sempre marcar como ativo quando aparece no arquivo (reativação automática)
                    }
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        # Se update_fields foi especificado, filtrar apenas os campos selecionados
                        if update_fields:
                            # Buscar mÃ¡quina existente primeiro
                            try:
                                maquina_obj = Maquina.objects.get(cd_maquina=cd_maquina)
                                # Atualizar apenas campos selecionados
                                for field in update_fields:
                                    if field in maquina_data:
                                        setattr(maquina_obj, field, maquina_data[field])
                                # Sempre reativar se estava inativa (independente de update_fields)
                                if not maquina_obj.ativo:
                                    maquina_obj.ativo = True
                                maquina_obj.save()
                                updated_count += 1
                            except Maquina.DoesNotExist:
                                # Se nÃ£o existe, criar novo registro com todos os campos
                                maquina_obj = Maquina.objects.create(
                                    cd_maquina=cd_maquina,
                                    **maquina_data
                                )
                                created_count += 1
                        else:
                            # Comportamento padrÃ£o: atualizar todos os campos
                            maquina_obj, created = Maquina.objects.update_or_create(
                            cd_maquina=cd_maquina,
                            defaults=maquina_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        maquina_obj, created = Maquina.objects.get_or_create(
                            cd_maquina=cd_maquina,
                            defaults=maquina_data
                        )
                        if created:
                            created_count += 1
                        else:
                            # Se já existe, reativar se estava inativa
                            if not maquina_obj.ativo:
                                maquina_obj.ativo = True
                                maquina_obj.save()
                                updated_count += 1
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        # Identificar máquinas que estão no DB mas não no arquivo (apenas as ativas)
        missing_machines = []
        if cd_maquinas_in_file:
            # Buscar máquinas ativas que não estão no arquivo
            maquinas_ativas_nao_no_arquivo = Maquina.objects.filter(
                ativo=True
            ).exclude(
                cd_maquina__in=cd_maquinas_in_file
            ).values('id', 'cd_maquina', 'descr_maquina', 'cd_setormanut', 'descr_setormanut')
            
            missing_machines = list(maquinas_ativas_nao_no_arquivo)
        
        return created_count, updated_count, errors, missing_machines
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors, []
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        import traceback
        traceback.print_exc()
        return 0, 0, errors, []


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_itens_estoque_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de itens de estoque a partir de um arquivo Excel ou CSV
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import ItemEstoque
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes encodings
            try:
                data = read_csv_file(file, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    file.seek(0)  # Resetar arquivo
                    data = read_csv_file(file, encoding='latin-1')
                except Exception as e:
                    raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_ITEM;DESCR_ITEM;...
                    
                    # Unidade
                    cd_unid = _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid'))
                    nome_unid = _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255)
                    
                    # Item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    descr_item = _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500)
                    
                    # Unidade de Medida
                    unidade_medida = _safe_str(row_data.get('UNIDADE_MEDIDA') or row_data.get('unidade_medida') or row_data.get('Unidade_Medida'), max_length=50)
                    
                    # Quantidade
                    qtde = _safe_decimal(row_data.get('QTDE') or row_data.get('qtde') or row_data.get('Qtde'))
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    item_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'descr_item': descr_item,
                        'unidade_medida': unidade_medida,
                        'qtde': qtde,
                    }
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        item_obj, created = ItemEstoque.objects.update_or_create(
                            codigo_item=cd_item,
                            defaults=item_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        item_obj, created = ItemEstoque.objects.get_or_create(
                            codigo_item=cd_item,
                            defaults=item_data
                        )
                        if created:
                            created_count += 1
                        
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_cas_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de Centros de Atividade (CA) a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import CentroAtividade
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes delimitadores e encodings
            # Primeiro tentar com ponto e vÃ­rgula (;) e encoding latin-1 (comum em arquivos brasileiros)
            data = None
            encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
            delimiters_to_try = [';', ',']
            
            last_error = None
            for delimiter in delimiters_to_try:
                for encoding in encodings_to_try:
                    try:
                        file.seek(0)  # Resetar arquivo
                        data = read_csv_file(file, encoding=encoding, delimiter=delimiter)
                        # Se conseguiu ler e tem dados, usar este formato
                        if data and len(data) > 0:
                            # Verificar se pelo menos uma coluna importante foi encontrada
                            first_row = data[0] if data else {}
                            if 'CA' in first_row or 'ca' in first_row or any('ca' in k.lower() for k in first_row.keys()):
                                break
                    except UnicodeDecodeError as e:
                        last_error = f"Erro de encoding com {encoding}: {str(e)}"
                        continue
                    except Exception as e:
                        last_error = f"Erro ao ler CSV com delimiter '{delimiter}' e encoding '{encoding}': {str(e)}"
                        continue
                if data and len(data) > 0:
                    break
                if data and len(data) > 0:
                    # Verificar novamente antes de sair do loop externo
                    first_row = data[0] if data else {}
                    if 'CA' in first_row or 'ca' in first_row or any('ca' in k.lower() for k in first_row.keys()):
                        break
            
            if not data or len(data) == 0:
                error_msg = "Erro ao ler arquivo CSV. Verifique o formato e encoding do arquivo."
                if last_error:
                    error_msg += f" Último erro: {last_error}"
                raise ValidationError(error_msg)
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                # Usar savepoint para cada linha, permitindo que erros em uma linha não afetem as outras
                savepoint_id = transaction.savepoint()
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        transaction.savepoint_commit(savepoint_id)
                        continue
                    
                    # Função auxiliar para encontrar coluna com múltiplas tentativas
                    def find_column_value(row_data, possible_names, partial_keywords=None):
                        """Tenta encontrar valor da coluna usando múltiplas estratégias"""
                        # Primeiro: tentar match exato (case-insensitive, ignorando espaços e encoding)
                        for key, value in row_data.items():
                            key_normalized = str(key).strip().upper().replace(' ', '').replace('_', '')
                            for name in possible_names:
                                name_normalized = str(name).strip().upper().replace(' ', '').replace('_', '')
                                if key_normalized == name_normalized:
                                    return value
                        
                        # Segundo: tentar match parcial
                        if partial_keywords:
                            result = _find_column_by_partial_match(row_data, partial_keywords)
                            if result is not None:
                                return result
                        
                        # Terceiro: tentar match direto com variações
                        for name in possible_names:
                            if name in row_data:
                                return row_data[name]
                        
                        # Quarto: tentar match case-insensitive direto
                        for key, value in row_data.items():
                            if str(key).strip().upper() == str(possible_names[0]).strip().upper():
                                return value
                        
                        return None
                    
                    # Extrair valores das colunas com múltiplas tentativas
                    ca_value = find_column_value(
                        row_data, 
                        ['CA', 'ca', 'Ca', 'CENTRO', 'centro', 'ATIVIDADE', 'atividade'],
                        ['ca', 'centro', 'atividade']
                    )
                    
                    sigla_value = find_column_value(
                        row_data,
                        ['SIGLA', 'sigla', 'Sigla'],
                        ['sigla']
                    )
                    
                    descricao_value = find_column_value(
                        row_data,
                        ['DESCRIÇÃO', 'DESCRI��O', 'DESCRIO', 'Descrição', 'Descri��o', 'Descrio', 
                         'descrição', 'descri��o', 'descrio', 'DESCRIÃ‡ÃƒO', 'DescriÃ§Ã£o'],
                        ['descricao', 'descrio']
                    )
                    
                    indice_value = find_column_value(
                        row_data,
                        ['INDICE', 'INDICE', 'ÍNDICE', 'ÃNDICE', 'ÃNDICE', 'indice', 'índice', 
                         'Ã­ndice', 'Indice', 'Índice'],
                        ['indice', 'ndice']
                    )
                    
                    encarregado_value = find_column_value(
                        row_data,
                        ['ENCARREGADO RESPONSÁVEL', 'ENCARREGADO RESPONS�VEL', 'ENCARREGADO RESPONSVEL',
                         'Encarregado Responsável', 'Encarregado Respons�vel', 'Encarregado Responsavel',
                         'encarregado responsável', 'encarregado respons�vel', 'encarregado responsavel',
                         'ENCARREGADO RESPONSÃVEL', 'Encarregado ResponsÃ¡vel'],
                        ['encarregado', 'responsavel', 'responsvel']
                    )
                    
                    local_value = find_column_value(
                        row_data,
                        ['LOCAL', 'local', 'Local'],
                        ['local']
                    )
                    
                    imagem_value = find_column_value(
                        row_data,
                        ['IMAGEM', 'imagem', 'Imagem', 'IMAGE', 'image', 'Image'],
                        ['imagem', 'image']
                    )
                    
                    # Validar que temos pelo menos o cÃ³digo CA
                    if not ca_value:
                        errors.append(f"Linha {row_num}: Campo 'CA' Ã© obrigatÃ³rio")
                        transaction.savepoint_commit(savepoint_id)
                        continue
                    
                    # Converter CA para inteiro
                    try:
                        ca_int = int(float(str(ca_value)))
                    except (ValueError, TypeError):
                        errors.append(f"Linha {row_num}: Valor de CA invÃ¡lido: {ca_value}")
                        transaction.savepoint_commit(savepoint_id)
                        continue
                    
                    # Converter Ã­ndice para inteiro se existir
                    indice_int = None
                    if indice_value:
                        try:
                            indice_int = int(float(str(indice_value)))
                        except (ValueError, TypeError):
                            pass  # Ãndice Ã© opcional
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o do CA
                    local_str = _safe_str(local_value, max_length=255) if local_value else None
                    # Preservar imagem mesmo se for string vazia (será None se realmente vazio)
                    imagem_str = None
                    if imagem_value is not None:
                        if str(imagem_value).strip():  # Se tem conteúdo após remover espaços
                            imagem_str = _safe_str(imagem_value, max_length=500)
                        # Se for string vazia, manter como None (não salvar strings vazias)
                    
                    # Normalizar caminho da imagem: remover 'static/' do início se presente
                    # pois Django's {% static %} tag já adiciona isso
                    if imagem_str and imagem_str.startswith('static/'):
                        imagem_str = imagem_str.replace('static/', '', 1)
                    
                    ca_data = {
                        'sigla': _safe_str(sigla_value, max_length=50),
                        'descricao': _safe_str(descricao_value, max_length=500),
                        'indice': indice_int,
                        'encarregado_responsavel': _safe_str(encarregado_value, max_length=255),
                        'local': local_str,
                        'imagem': imagem_str,
                    }
                    
                    # Criar ou atualizar CA
                    # Como CA não é mais único, precisamos verificar se já existe um registro com os mesmos dados
                    # ou criar um novo registro
                    if update_existing:
                        # Tentar encontrar um registro existente com o mesmo CA e dados similares
                        existing_ca = CentroAtividade.objects.filter(
                            ca=ca_int,
                            sigla=ca_data.get('sigla'),
                            descricao=ca_data.get('descricao'),
                            local=ca_data.get('local')
                        ).first()
                        
                        if existing_ca:
                            # Atualizar registro existente
                            for key, value in ca_data.items():
                                setattr(existing_ca, key, value)
                            existing_ca.save()
                            updated_count += 1
                        else:
                            # Criar novo registro
                            CentroAtividade.objects.create(ca=ca_int, **ca_data)
                            created_count += 1
                    else:
                        # Verificar se já existe um registro idêntico
                        existing_ca = CentroAtividade.objects.filter(
                            ca=ca_int,
                            sigla=ca_data.get('sigla'),
                            descricao=ca_data.get('descricao'),
                            local=ca_data.get('local')
                        ).first()
                        
                        if not existing_ca:
                            # Criar novo registro apenas se não existir um idêntico
                            CentroAtividade.objects.create(ca=ca_int, **ca_data)
                            created_count += 1
                    
                    # Confirmar savepoint se tudo correu bem
                    transaction.savepoint_commit(savepoint_id)
                    
                except Exception as e:
                    # Reverter savepoint em caso de erro
                    transaction.savepoint_rollback(savepoint_id)
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
                    # Continuar processando outras linhas mesmo se esta falhar
                    continue
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def _find_column_by_partial_match(row_data, keywords):
    """
    Tenta encontrar uma coluna em row_data que contenha qualquer uma das palavras-chave.
    Ãštil para lidar com problemas de encoding ou pequenas variaÃ§Ãµes.
    """
    for key, value in row_data.items():
        normalized_key = str(key).strip().lower().replace(' ', '_')
        for keyword in keywords:
            if keyword in normalized_key:
                return value
    return None


def upload_manutentores_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de manutentores a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import Manutentor
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar diferentes encodings
            try:
                data = read_csv_file(file, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    file.seek(0)  # Resetar arquivo
                    data = read_csv_file(file, encoding='latin-1')
                except Exception as e:
                    raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    matricula = _safe_str(row_data.get('MATRICULA') or row_data.get('matricula') or row_data.get('Matricula') or row_data.get('CADASTRO') or row_data.get('cadastro') or row_data.get('Cadastro'), max_length=1000)
                    nome = _safe_str(row_data.get('NOME') or row_data.get('nome') or row_data.get('Nome'), max_length=1000)
                    cargo = _safe_str(row_data.get('CARGO') or row_data.get('cargo') or row_data.get('Cargo'), max_length=1000)
                    horario_inicio_str = row_data.get('HORARIO_INICIO') or row_data.get('horario_inicio') or row_data.get('Horario_Inicio')
                    horario_fim_str = row_data.get('HORARIO_FIM') or row_data.get('horario_fim') or row_data.get('Horario_Fim')
                    tempo_trabalho = _safe_str(row_data.get('TEMPO_TRABALHO') or row_data.get('tempo_trabalho') or row_data.get('Tempo_Trabalho'), max_length=250)
                    turno = _safe_str(row_data.get('TURNO') or row_data.get('turno') or row_data.get('Turno'), max_length=25)
                    local_trab = _safe_str(row_data.get('LOCAL_TRAB') or row_data.get('local_trab') or row_data.get('Local_Trab') or row_data.get('LOCAL_TRABALHO') or row_data.get('local_trabalho') or row_data.get('Local_Trabalho'), max_length=40)
                    
                    # Validar que temos pelo menos a matrÃ­cula
                    if not matricula:
                        errors.append(f"Linha {row_num}: Campo 'MATRICULA' Ã© obrigatÃ³rio")
                        continue
                    
                    # Converter horÃ¡rios
                    horario_inicio_time = None
                    horario_fim_time = None
                    if horario_inicio_str:
                        try:
                            horario_inicio_time = datetime.strptime(str(horario_inicio_str).strip(), '%H:%M:%S').time()
                        except ValueError:
                            try:
                                horario_inicio_time = datetime.strptime(str(horario_inicio_str).strip(), '%H:%M').time()
                            except ValueError:
                                pass  # HorÃ¡rio Ã© opcional
                    
                    if horario_fim_str:
                        try:
                            horario_fim_time = datetime.strptime(str(horario_fim_str).strip(), '%H:%M:%S').time()
                        except ValueError:
                            try:
                                horario_fim_time = datetime.strptime(str(horario_fim_str).strip(), '%H:%M').time()
                            except ValueError:
                                pass  # HorÃ¡rio Ã© opcional
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    manutentor_data = {
                        'Nome': nome,
                        'Cargo': cargo,
                        'horario_inicio': horario_inicio_time,
                        'horario_fim': horario_fim_time,
                        'tempo_trabalho': tempo_trabalho,
                        'turno': turno,
                        'local_trab': local_trab,
                    }
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        manutentor_obj, created = Manutentor.objects.update_or_create(
                            Matricula=matricula,
                            defaults=manutentor_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        manutentor_obj, created = Manutentor.objects.get_or_create(
                            Matricula=matricula,
                            defaults=manutentor_data
                        )
                        if created:
                            created_count += 1
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def _safe_int(value, default=None):
    """Converte valor para inteiro de forma segura"""
    if value is None or value == '':
        return default
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return default


def _safe_str(value, max_length=None, default=''):
    """Converte valor para string de forma segura"""
    if value is None:
        return default
    str_value = str(value).strip()
    if max_length and len(str_value) > max_length:
        str_value = str_value[:max_length]
    return str_value


def _format_dt_execucao_plano(val):
    """Normaliza data de execução para DD/MM/AAAA (Excel pode enviar datetime/date)."""
    from datetime import datetime, date
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y')
    return _safe_str(val, max_length=50) if val else None


def _get_row_value(row_data, *keys, default=''):
    """Obtém valor do row_data tentando múltiplas chaves possíveis (suporta encoding variants)."""
    for key in keys:
        if key is None:
            continue
        if key in row_data:
            val = row_data.get(key)
            if val is not None and str(val).strip() != '':
                return val
    return default


def _get_row_value_by_normalized_key(row_data, *target_names, default=''):
    """
    Obtém valor do row_data buscando chave por nome normalizado (remove acentos, lowercase).
    Útil para CSVs com encoding variado (ex: PLANO_PREVENTIVA_DELIMITADO_DADOS.csv).
    target_names: ex. ('unidade', 'nome_unidade') - usa o primeiro match exato.
    """
    import unicodedata
    for k, v in row_data.items():
        if v is None or str(v).strip() == '':
            continue
        k_norm = unicodedata.normalize('NFKD', str(k)).encode('ascii', 'ignore').decode().lower()
        k_norm = k_norm.replace(' ', '_').replace('-', '_')
        for target in target_names:
            t = str(target).lower().replace(' ', '_')
            if not t:
                continue
            if k_norm == t:
                return v
            if len(k_norm) >= len(t) and (k_norm.startswith(t + '_') or k_norm.startswith(t + ' ') or k_norm == t):
                return v
    return default


def _classify_value_type(value):
    """
    Classifica um valor (string ou outro) em: 'integer', 'float', 'date', 'text', 'empty'.
    Usado para análise de padrões em colunas importadas.
    """
    import re
    from datetime import datetime
    if value is None:
        return 'empty'
    s = str(value).strip()
    if not s:
        return 'empty'
    # Integer: apenas dígitos, opcionalmente com sinal
    if re.match(r'^-?\d+$', s):
        return 'integer'
    # Float: dígitos com ponto ou vírgula decimal
    if re.match(r'^-?\d+[.,]\d+$', s) or re.match(r'^-?\d{1,3}(\.\d{3})*,\d+$', s):
        return 'float'
    # Date: formatos comuns dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd, etc.
    date_patterns = [
        r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        r'^\d{1,2}-\d{1,2}-\d{2,4}$',
        r'^\d{4}-\d{2}-\d{2}$',
        r'^\d{1,2}\.\d{1,2}\.\d{2,4}$',
        r'^\d{2}/\d{2}/\d{4}\s+\d{1,2}:\d{2}',
        r'^\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}',
    ]
    for pat in date_patterns:
        if re.match(pat, s):
            return 'date'
    # Tentar parse explícito de data
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%Y'):
        try:
            datetime.strptime(s[:10], fmt)
            return 'date'
        except (ValueError, TypeError):
            continue
    return 'text'


def analyze_maquinas_column_patterns(maquinas_queryset, columns_config):
    """
    Analisa padrões de dados nas colunas de máquinas importadas.
    Detecta inconsistências: ex. maioria inteiros mas alguns textos, ou datas em coluna numérica.

    columns_config: dict { field_name: {'expected': 'integer'|'text'|'date', 'label': 'Nome exibido'} }
    Retorna lista de dicts com: field, label, expected, total, type_counts, has_problem, problem_desc, sample_problematic
    """
    from collections import Counter
    fields = list(columns_config.keys())
    if not fields:
        return []
    # Uma única query para todos os campos
    rows = list(maquinas_queryset.values(*fields))
    results = []
    for field_name, config in columns_config.items():
        expected = config.get('expected', 'text')
        label = config.get('label', field_name)
        type_counts = Counter()
        problematic_values = []
        for row in rows:
            val = row.get(field_name)
            if val is not None and not isinstance(val, str):
                val = str(val)
            classified = _classify_value_type(val)
            type_counts[classified] += 1
            if expected == 'integer' and classified not in ('integer', 'empty'):
                problematic_values.append((str(val)[:80] if val is not None else 'NULL', classified))
            elif expected == 'date' and classified not in ('date', 'empty'):
                problematic_values.append((str(val)[:80] if val is not None else 'NULL', classified))
            elif expected == 'text' and classified == 'date':
                problematic_values.append((str(val)[:80] if val is not None else 'NULL', classified))
        total = len(rows)
        non_empty = total - type_counts.get('empty', 0)
        has_problem = False
        problem_desc = ''
        if total > 0 and non_empty > 0:
            sorted_types = [t for t, _ in type_counts.most_common() if t != 'empty']
            majority_type = sorted_types[0] if sorted_types else 'empty'
            majority_count = type_counts.get(majority_type, 0)
            minority_total = non_empty - majority_count
            if minority_total > 0 and (minority_total / non_empty) > 0.05:
                has_problem = True
                minority_types = [t for t, c in type_counts.items() if t != majority_type and t != 'empty' and c > 0]
                problem_desc = f"Maioria é '{majority_type}' ({majority_count}), mas {minority_total} valor(es) com tipo diferente: {', '.join(minority_types)}"
            if expected == 'integer' and (type_counts.get('date', 0) > 0 or type_counts.get('text', 0) > 0):
                has_problem = True
                if not problem_desc:
                    problem_desc = f"Coluna esperada como numérica, mas contém datas ({type_counts.get('date', 0)}) ou texto ({type_counts.get('text', 0)})"
            elif expected == 'date' and type_counts.get('integer', 0) > 0 and type_counts.get('date', 0) == 0:
                has_problem = True
                problem_desc = "Coluna esperada como data, mas a maioria dos valores são numéricos (possível coluna trocada)"
        results.append({
            'field': field_name,
            'label': label,
            'expected': expected,
            'total': total,
            'type_counts': dict(type_counts),
            'has_problem': has_problem,
            'problem_desc': problem_desc,
            'sample_problematic': problematic_values[:10],
        })
    return results


def _safe_decimal(value, default=None):
    """Converte valor para Decimal de forma segura
    
    Suporta formatos:
    - US: 1234.56 (ponto como separador decimal)
    - Europeu: 1234,56 ou 1.234,56 (vÃ­rgula como separador decimal, ponto como separador de milhares)
    """
    from decimal import Decimal, InvalidOperation
    if value is None or value == '':
        return default
    
    try:
        value_str = str(value).strip()
        
        # Se vazio apÃ³s strip, retornar default
        if not value_str:
            return default
        
        # Remover prefixo R$ (pode estar antes ou depois do sinal negativo)
        # Exemplos: "R$ 40,01", "-R$ 40,01", "R$ -40,01"
        import re
        value_str = re.sub(r'R\$\s*', '', value_str, flags=re.IGNORECASE)
        value_str = value_str.strip()
        
        # Se vazio apÃ³s remover R$, retornar default
        if not value_str:
            return default
        
        # Verificar se hÃ¡ vÃ­rgula (formato europeu)
        if ',' in value_str:
            # Verificar se hÃ¡ ponto antes da vÃ­rgula (separador de milhares)
            comma_pos = value_str.rfind(',')
            period_before_comma = value_str.rfind('.', 0, comma_pos)
            
            if period_before_comma != -1:
                # Formato europeu com separador de milhares: 1.234,56
                # Remover pontos (separadores de milhares) e substituir vÃ­rgula por ponto
                value_str = value_str.replace('.', '').replace(',', '.')
            else:
                # Formato europeu simples: 1234,56
                # Substituir vÃ­rgula por ponto
                value_str = value_str.replace(',', '.')
        
        return Decimal(value_str)
    except (InvalidOperation, ValueError, TypeError):
        return default


def _safe_uso_contabil(value, default=''):
    """Extrai apenas a parte inteira do uso contábil
    
    Converte valores como "242,00", "242.00", "242" para "242"
    Remove qualquer parte decimal e retorna apenas o número inteiro como string
    """
    if value is None or value == '':
        return default
    
    # Converter para string e remover espaços
    str_value = str(value).strip()
    if not str_value:
        return default
    
    import re
    
    # Remover tudo exceto dígitos, vírgula e ponto
    cleaned = re.sub(r'[^\d,.]', '', str_value)
    
    if not cleaned:
        return default
    
    # Extrair número inteiro
    # Se tiver vírgula, assumir formato brasileiro (242,00)
    if ',' in cleaned:
        # Formato brasileiro: 242,00 ou 1.242,00
        parts = cleaned.split(',')
        integer_part = parts[0].replace('.', '')  # Remover separador de milhares se houver
    # Se tiver ponto, verificar se é decimal ou separador de milhares
    elif '.' in cleaned:
        parts = cleaned.split('.')
        # Se a parte após o ponto tem 1-2 dígitos, é provavelmente decimal (242.00)
        # Caso contrário, é separador de milhares (1.242)
        if len(parts) == 2 and len(parts[1]) <= 2:
            integer_part = parts[0]  # Formato US: 242.00
        else:
            integer_part = cleaned.replace('.', '')  # Separador de milhares: 1.242
    else:
        # Apenas números inteiros
        integer_part = cleaned
    
    # Garantir que temos apenas dígitos
    integer_part = re.sub(r'\D', '', integer_part)
    
    return integer_part if integer_part else default


def _safe_date(value, default=None):
    """Converte valor para date de forma segura
    
    Suporta:
    - Objetos datetime/date do Python
    - Strings em vários formatos brasileiros e internacionais
    - Objetos datetime do Excel (openpyxl)
    """
    from datetime import datetime, date
    
    if value is None or value == '':
        return default
    
    # Excel usa "-", "---", "N/A" etc. para indicar ausência de data
    if isinstance(value, str):
        v = value.strip()
        if v in ('-', '--', '---', 'N/A', 'n/a', 'N/D', 'n/d', 'NA', 'ND', ''):
            return default
    
    # Se já for um objeto date, retornar diretamente
    if isinstance(value, date):
        return value
    
    # Se for datetime, extrair apenas a data
    if isinstance(value, datetime):
        return value.date()
    
    # Excel serial number (ex: 45678.0, 45500.0 ou "45678" para datas)
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta
            num = int(float(value))
            if 1 <= num <= 2958465:  # range Excel (1900-9999)
                return (datetime(1899, 12, 30) + timedelta(days=num)).date()
        except (ValueError, OverflowError):
            pass
    # String que parece número serial do Excel (ex: "45678", "45500.0")
    if isinstance(value, str):
        stripped = value.strip()
        cleaned = stripped.replace('.', '').replace(',', '')
        if cleaned.isdigit() or (cleaned.replace('-', '').isdigit() and '-' in stripped):
            try:
                from datetime import timedelta
                num = int(float(stripped.replace(',', '.')))
                if 1 <= num <= 2958465:
                    return (datetime(1899, 12, 30) + timedelta(days=num)).date()
            except (ValueError, OverflowError):
                pass
    
    # Tentar converter string para date
    value_str = str(value).strip()
    if not value_str:
        return default
    
    # Remover hora se existir
    if ' ' in value_str:
        date_part = value_str.split(' ')[0]
    else:
        date_part = value_str
    
    # Tentar diferentes formatos de data
    date_formats = [
        '%d/%m/%Y',      # 26/09/2025
        '%d-%m-%Y',      # 26-09-2025
        '%d.%m.%Y',      # 26.09.2025
        '%Y-%m-%d',      # 2025-09-26
        '%Y/%m/%d',      # 2025/09/26
        '%d/%m/%y',      # 26/09/25
        '%d-%m-%y',      # 26-09-25
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(date_part, fmt).date()
        except (ValueError, TypeError):
            continue
    
    # Se nenhum formato funcionou, tentar parse manual para formato brasileiro comum
    if '/' in date_part:
        parts = date_part.split('/')
        if len(parts) == 3:
            try:
                day, month, year = parts
                # Se ano tem 2 dígitos, assumir 2000+
                if len(year) == 2:
                    year = '20' + year
                return date(int(year), int(month), int(day))
            except (ValueError, TypeError):
                pass
    
        return default


def _fix_funcionario_columns(row_data):
    """
    Corrige deslocamento de colunas para 'Funcionário' e 'Nome Funcionário'
    Se 'Funcionário' estiver vazio, assume que os dados corretos estão na próxima coluna
    Suporta encoding variants (FuncionÃ¡rio, Funcionário, etc.)
    """
    import unicodedata
    funcionario_key = None
    nome_funcionario_key = None

    def _norm(s):
        return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower()

    # Encontrar as chaves corretas
    for key in row_data.keys():
        key_lower = str(key).lower().strip()
        key_norm = _norm(key)
        if 'funcionario' in key_norm or 'funcionÃ¡rio' in key_lower or 'funcionário' in key_lower:
            if 'nome' in key_norm or 'nome' in key_lower:
                nome_funcionario_key = key
            else:
                funcionario_key = key
    
    if not funcionario_key or not nome_funcionario_key:
        return row_data
    
    funcionario_value = row_data.get(funcionario_key, '').strip() if row_data.get(funcionario_key) else ''
    nome_funcionario_value = row_data.get(nome_funcionario_key, '').strip() if row_data.get(nome_funcionario_key) else ''
    
    # Se "FuncionÃ¡rio" estÃ¡ vazio, procurar dados nas prÃ³ximas colunas
    if not funcionario_value:
        # Encontrar todas as chaves apÃ³s "FuncionÃ¡rio"
        keys_list = list(row_data.keys())
        funcionario_idx = keys_list.index(funcionario_key) if funcionario_key in keys_list else -1
        
        if funcionario_idx >= 0:
            # Procurar primeira coluna nÃ£o vazia apÃ³s "FuncionÃ¡rio"
            for i in range(funcionario_idx + 1, len(keys_list)):
                next_key = keys_list[i]
                next_value = row_data.get(next_key, '').strip() if row_data.get(next_key) else ''
                
                if next_value:
                    # Se o valor contÃ©m apenas nÃºmeros, Ã© provavelmente o cÃ³digo do funcionÃ¡rio
                    if next_value.isdigit():
                        row_data[funcionario_key] = next_value
                        # Procurar prÃ³xima coluna nÃ£o vazia para o nome
                        for j in range(i + 1, len(keys_list)):
                            next_name_key = keys_list[j]
                            next_name_value = row_data.get(next_name_key, '').strip() if row_data.get(next_name_key) else ''
                            if next_name_value and not next_name_value.isdigit():
                                row_data[nome_funcionario_key] = next_name_value
                                # Limpar colunas deslocadas
                                row_data[next_key] = ''
                                if j < len(keys_list):
                                    row_data[next_name_key] = ''
                                break
                        break
    
    # Se "Nome FuncionÃ¡rio" contÃ©m apenas nÃºmeros, assumir que o nome correto estÃ¡ na prÃ³xima coluna
    if nome_funcionario_value and nome_funcionario_value.isdigit():
        # Mover nÃºmero para "FuncionÃ¡rio" se estiver vazio
        if not funcionario_value:
            row_data[funcionario_key] = nome_funcionario_value
        
        # Encontrar prÃ³xima coluna nÃ£o vazia para o nome
        keys_list = list(row_data.keys())
        nome_funcionario_idx = keys_list.index(nome_funcionario_key) if nome_funcionario_key in keys_list else -1
        
        if nome_funcionario_idx >= 0:
            for i in range(nome_funcionario_idx + 1, len(keys_list)):
                next_key = keys_list[i]
                next_value = row_data.get(next_key, '').strip() if row_data.get(next_key) else ''
                if next_value and not next_value.isdigit():
                    row_data[nome_funcionario_key] = next_value
                    # Limpar coluna deslocada
                    row_data[next_key] = ''
                    break
    
    return row_data


def _score_plano_preventiva_csv_rows(data: List[Dict]) -> int:
    """Conta linhas com Plano e Máquina reconhecíveis (usado para escolher encoding do CSV)."""
    if not data:
        return 0
    score = 0
    for row_data in data[: min(400, len(data))]:
        try:
            row_data = _fix_funcionario_columns(dict(row_data))

            def _v(*keys):
                return _get_row_value(row_data, *keys)

            def _vn(*nk):
                return _get_row_value_by_normalized_key(row_data, *nk)

            npi = _safe_int(_v('NUMERO_PLANO', 'numero_plano', 'Plano') or _vn('plano'))
            mqi = _safe_int(_v('CD_MAQUINA', 'cd_maquina', 'Máquina', 'Maquina') or _vn('maquina'))
            if npi is not None and mqi is not None:
                score += 1
        except Exception:
            continue
    return score


def read_csv_file_auto_encoding(file, delimiter=';') -> List[Dict]:
    """
    Decodifica CSV em bytes e escolhe o encoding que maximiza linhas com Plano+Máquina válidos.
    Tenta: UTF-8 com BOM, UTF-8, cp1252 (Excel Windows), latin-1.
    """
    if hasattr(file, 'read'):
        file.seek(0)
        raw = file.read()
    else:
        with open(file, 'rb') as fb:
            raw = fb.read()
    encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'iso-8859-1']
    best_data = None
    best_score = -1
    for enc in encodings:
        try:
            text = raw.decode(enc)
        except UnicodeDecodeError:
            continue
        try:
            data = _parse_csv_dict_rows_from_text(text, delimiter)
        except Exception:
            continue
        score = _score_plano_preventiva_csv_rows(data)
        if score > best_score:
            best_score = score
            best_data = data
    if best_data is None:
        raise ValidationError(
            'Não foi possível decodificar o CSV. Use UTF-8 ou Windows-1252 e delimitador ; (ponto e vírgula).'
        )
    if best_score == 0:
        for enc in encodings:
            try:
                text = raw.decode(enc)
            except UnicodeDecodeError:
                continue
            try:
                alt = _parse_csv_dict_rows_from_text(text, ',')
            except Exception:
                continue
            if _score_plano_preventiva_csv_rows(alt) > 0:
                return alt
    return best_data


def upload_plano_preventiva_from_file(file, update_existing=False) -> Tuple[int, int, int, List[str]]:
    """
    Faz upload de planos de manutenÃ§Ã£o preventiva a partir de um arquivo CSV ou Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, skipped_count, errors)
    """
    from app.models import PlanoPreventiva, Maquina
    
    errors = []
    created_count = 0
    updated_count = 0
    skipped_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            file.seek(0)
            data = read_csv_file_auto_encoding(file, delimiter=';')
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Cache de mÃ¡quinas para melhorar performance
        maquinas_cache = {}
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Corrigir deslocamento de colunas para Funcionário
                    row_data = _fix_funcionario_columns(row_data)
                    
                    # Mapear colunas do CSV para campos do modelo
                    # Suporta: (1) CD_UNID, NUMERO_PLANO, etc. e (2) PLANO_PREVENTIVA_DELIMITADO_DADOS.csv
                    # (Unidade, Nome Unidade, Setor, Máquina, Plano, Sequência Manutenção, Data Execução, etc.)
                    
                    def _v(*keys):
                        return _get_row_value(row_data, *keys)
                    
                    def _v_norm(*norm_keys):
                        return _get_row_value_by_normalized_key(row_data, *norm_keys)
                    
                    # Unidade
                    cd_unid = _safe_int(_v('CD_UNID', 'cd_unid', 'Unidade') or _v_norm('unidade'))
                    nome_unid = _safe_str(_v('NOME_UNID', 'nome_unid', 'Nome Unidade') or _v_norm('nome_unidade'), max_length=255)
                    
                    # Setor e Atividade (PLANO_PREVENTIVA_DELIMITADO_DADOS.csv)
                    cd_setor = _safe_str(_v('CD_SETOR', 'cd_setor', 'Setor') or _v_norm('setor'), max_length=50)
                    descr_setor = _safe_str(_v('DESCR_SETOR', 'descr_setor', 'Descrição Setor', 'Descricao Setor') or _v_norm('descricao_setor', 'descr_setor'), max_length=255)
                    cd_atividade = _safe_int(_v('CD_ATIVIDADE', 'cd_atividade', 'Atividade') or _v_norm('atividade'))
                    
                    # Plano
                    numero_plano = _safe_int(_v('NUMERO_PLANO', 'numero_plano', 'Plano') or _v_norm('plano'))
                    descr_plano = _safe_str(_v('DESCR_PLANO', 'descr_plano', 'Descrição Plano', 'Descricao Plano') or _v_norm('descricao_plano', 'descr_plano'), max_length=255)
                    
                    # Máquina
                    cd_maquina = _safe_int(_v('CD_MAQUINA', 'cd_maquina', 'Máquina', 'Maquina') or _v_norm('maquina'))
                    descr_maquina = _safe_str(_v('DESCR_MAQUINA', 'descr_maquina', 'Descrição Máquina', 'Descricao Maquina') or _v_norm('descricao_maquina', 'descr_maquina'), max_length=500)
                    nro_patrimonio = _safe_str(_v('NRO_PATRIMONIO', 'nro_patrimonio', 'Nº Patrimônio', 'N Patrimonio', 'Patrimonio') or _v_norm('patrimonio', 'nro_patrimonio'), max_length=100)
                    
                    # Sequências
                    sequencia_tarefa = _safe_int(_v('SEQUENCIA_TAREFA', 'sequencia_tarefa', 'Sequência Tarefa', 'Sequencia Tarefa') or _v_norm('sequencia_tarefa'))
                    sequencia_manutencao = _safe_int(_v('SEQUENCIA_MANUTENCAO', 'sequencia_manutencao', 'Sequência Manutenção', 'Sequencia Manutencao') or _v_norm('sequencia_manutencao'))
                    
                    # Tarefa
                    descr_tarefa = _safe_str(_v('DESCR_TAREFA', 'descr_tarefa', 'Descrição Tarefa', 'Descricao Tarefa') or _v_norm('descricao_tarefa', 'descr_tarefa'))
                    
                    # Quantidade Período (PLANO_PREVENTIVA_DELIMITADO_DADOS.csv)
                    quantidade_periodo = _safe_int(_v('QUANTIDADE_PERIODO', 'quantidade_periodo', 'Quantidade Período', 'Quantidade Periodo') or _v_norm('quantidade_periodo'))
                    
                    # Funcionário
                    cd_funcionario = _safe_str(_v('FUNCIONARIO', 'FUNCIONÁRIO', 'Funcionário', 'Funcionario', 'CD_FUNCIONARIO', 'cd_funcionario') or _v_norm('funcionario'), max_length=100)
                    nome_funcionario = _safe_str(_v('NOME_FUNCIONARIO', 'NOME_FUNCIONÁRIO', 'Nome Funcionário', 'Nome Funcionario') or _v_norm('nome_funcionario'), max_length=255)
                    
                    # Data Execução - PlanoPreventiva usa dt_execucao (CharField DD/MM/YYYY); Excel pode vir como datetime
                    data_execucao_raw = _v('DATA_EXECUCAO', 'data_execucao', 'Data Execução', 'Data Execucao') or _v_norm('data_execucao')
                    dt_execucao = _format_dt_execucao_plano(data_execucao_raw)
                    
                    # Validar campos obrigatórios (Plano e Máquina para identificar unicamente)
                    # Usar 'is None' para aceitar 0 como valor válido (ex: cd_maquina=0, numero_plano=0)
                    if numero_plano is None:
                        # Pular linhas malformadas (ex: apenas código e nome de funcionário em colunas erradas)
                        non_empty = [str(v).strip() for v in row_data.values() if v and str(v).strip()]
                        if len(non_empty) <= 2 and all(
                            (s.isdigit() and len(s) >= 4) or (not s.isdigit() and len(s) > 3)
                            for s in non_empty
                        ):
                            continue  # Linha malformada ignorada
                        errors.append(f"Linha {row_num}: Campo 'Plano' ou 'NUMERO_PLANO' é obrigatório")
                        continue
                    
                    if cd_maquina is None:
                        errors.append(f"Linha {row_num}: Campo 'Máquina' ou 'CD_MAQUINA' é obrigatório")
                        continue
                    
                    # Tentar encontrar máquina relacionada (opcional - importa mesmo se não existir)
                    maquina = None
                    if cd_maquina:
                        if cd_maquina in maquinas_cache:
                            maquina = maquinas_cache[cd_maquina]
                        else:
                            try:
                                maquina = Maquina.objects.get(cd_maquina=cd_maquina)
                                maquinas_cache[cd_maquina] = maquina
                            except Maquina.DoesNotExist:
                                maquina = None  # Importa sem vínculo com Maquina
                    
                    # Preparar dados para criação/atualização (PlanoPreventiva model fields)
                    plano_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'cd_setor': cd_setor,
                        'descr_setor': descr_setor,
                        'cd_atividade': cd_atividade,
                        'descr_plano': descr_plano,
                        'maquina': maquina,
                        'cd_maquina': cd_maquina,
                        'descr_maquina': descr_maquina,
                        'nro_patrimonio': nro_patrimonio,
                        'sequencia_tarefa': sequencia_tarefa,
                        'sequencia_manutencao': sequencia_manutencao,
                        'descr_tarefa': descr_tarefa,
                        'quantidade_periodo': quantidade_periodo,
                        'cd_funcionario': cd_funcionario,
                        'nome_funcionario': nome_funcionario,
                        'dt_execucao': dt_execucao,
                    }
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        plano_obj, created = PlanoPreventiva.objects.update_or_create(
                            numero_plano=numero_plano,
                            cd_maquina=cd_maquina,
                            sequencia_tarefa=sequencia_tarefa,
                            sequencia_manutencao=sequencia_manutencao,
                            defaults=plano_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        plano_obj, created = PlanoPreventiva.objects.get_or_create(
                            numero_plano=numero_plano,
                            cd_maquina=cd_maquina,
                            sequencia_tarefa=sequencia_tarefa,
                            sequencia_manutencao=sequencia_manutencao,
                            defaults=plano_data
                        )
                        if created:
                            created_count += 1
                        else:
                            skipped_count += 1
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, skipped_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")  # Debug
        import traceback
        traceback.print_exc()
        return 0, 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_roteiro_preventiva_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de roteiro de manutenÃ§Ã£o preventiva a partir de um arquivo CSV
    O arquivo deve ter colunas no formato: CD_UNID;NOME_UNID;CD_FUNCIOMANU;...
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RoteiroPreventiva, Maquina
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo baseado na extensÃ£o
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
        elif file_name.endswith('.csv'):
            # Tentar encodings: UTF-8 primeiro (evita mojibake em arquivos com Ç, Á, Õ etc).
            # Latin-1 quase nunca falha, então se usado primeiro corrompe textos UTF-8.
            data = None
            for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
                for delimiter in (';', ','):
                    try:
                        file.seek(0)
                        data = read_csv_file(file, encoding=encoding, delimiter=delimiter)
                        if data and len(data) > 0:
                            break
                    except (UnicodeDecodeError, Exception):
                        continue
                if data and len(data) > 0:
                    break
            if not data or len(data) == 0:
                raise ValidationError("Arquivo CSV vazio ou não foi possível decodificar com UTF-8, CP1252 ou Latin-1.")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls, .xlsm ou .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Cache de mÃ¡quinas para melhorar performance
        maquinas_cache = {}
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_FUNCIOMANU;NOME_FUNCIOMANU;FUNCIOMANU_ID;CD_SETORMANUT;DESCR_SETORMANUT;...
                    
                    # Unidade
                    cd_unid = _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid'))
                    nome_unid = _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255)
                    
                    # FuncionÃ¡rio
                    cd_funciomanu = _safe_str(row_data.get('CD_FUNCIOMANU') or row_data.get('cd_funciomanu') or row_data.get('Cd_Funciomanu'), max_length=100)
                    nome_funciomanu = _safe_str(row_data.get('NOME_FUNCIOMANU') or row_data.get('nome_funciomanu') or row_data.get('Nome_Funciomanu'), max_length=255)
                    funciomanu_id = _safe_int(row_data.get('FUNCIOMANU_ID') or row_data.get('funciomanu_id') or row_data.get('Funciomanu_Id'))
                    
                    # Setor
                    cd_setormanut = _safe_str(row_data.get('CD_SETORMANUT') or row_data.get('cd_setormanut') or row_data.get('Cd_Setormanut'), max_length=50)
                    descr_setormanut = _safe_str(row_data.get('DESCR_SETORMANUT') or row_data.get('descr_setormanut') or row_data.get('Descr_Setormanut'), max_length=255)
                    
                    # Tipo Centro de Atividade
                    cd_tpcentativ = _safe_int(row_data.get('CD_TPCENTATIV') or row_data.get('cd_tpcentativ') or row_data.get('Cd_Tpcentativ'))
                    descr_abrev_tpcentativ = _safe_str(row_data.get('DESCR_ABREV_TPCENTATIV') or row_data.get('descr_abrev_tpcentativ') or row_data.get('Descr_Abrev_Tpcentativ'), max_length=255)
                    
                    # Ordem de ServiÃ§o
                    dt_abertura = _safe_str(row_data.get('DT_ABERTURA') or row_data.get('dt_abertura') or row_data.get('Dt_Abertura'), max_length=50)
                    cd_ordemserv = _safe_int(row_data.get('CD_ORDEMSERV') or row_data.get('cd_ordemserv') or row_data.get('Cd_Ordemserv'))
                    ordemserv_id = _safe_int(row_data.get('ORDEMSERV_ID') or row_data.get('ordemserv_id') or row_data.get('Ordemserv_Id'))
                    
                    # MÃ¡quina
                    cd_maquina = _safe_int(row_data.get('CD_MAQUINA') or row_data.get('cd_maquina') or row_data.get('Cd_Maquina'))
                    descr_maquina = _safe_str(row_data.get('DESCR_MAQUINA') or row_data.get('descr_maquina') or row_data.get('Descr_Maquina'), max_length=500)
                    
                    # Plano de ManutenÃ§Ã£o
                    cd_planmanut = _safe_int(row_data.get('CD_PLANMANUT') or row_data.get('cd_planmanut') or row_data.get('Cd_Planmanut'))
                    descr_planmanut = _safe_str(row_data.get('DESCR_PLANMANUT') or row_data.get('descr_planmanut') or row_data.get('Descr_Planmanut'), max_length=255)
                    descr_recomenos = _safe_str(row_data.get('DESCR_RECOMENOS') or row_data.get('descr_recomenos') or row_data.get('Descr_Recomenos'))
                    cf_dt_final_execucao = _safe_str(row_data.get('CF_DT_FINAL_EXECUCAO') or row_data.get('cf_dt_final_execucao') or row_data.get('Cf_Dt_Final_Execucao'), max_length=50)
                    cs_qtde_periodo_max = _safe_int(row_data.get('CS_QTDE_PERIODO_MAX') or row_data.get('cs_qtde_periodo_max') or row_data.get('Cs_Qtde_Periodo_Max'))
                    cs_tot_temp = _safe_str(row_data.get('CS_TOT_TEMP') or row_data.get('cs_tot_temp') or row_data.get('Cs_Tot_Temp'), max_length=50)
                    cf_tot_temp = _safe_str(row_data.get('CF_TOT_TEMP') or row_data.get('cf_tot_temp') or row_data.get('Cf_Tot_Temp'), max_length=50)
                    
                    # SequÃªncia Plano ManutenÃ§Ã£o
                    seq_seqplamanu = _safe_int(row_data.get('SEQ_SEQPLAMANU') or row_data.get('seq_seqplamanu') or row_data.get('Seq_Seqplamanu'))
                    
                    # Tarefa ManutenÃ§Ã£o
                    cd_tarefamanu = _safe_int(row_data.get('CD_TAREFAMANU') or row_data.get('cd_tarefamanu') or row_data.get('Cd_Tarefamanu'))
                    descr_tarefamanu = _safe_str(row_data.get('DESCR_TAREFAMANU') or row_data.get('descr_tarefamanu') or row_data.get('Descr_Tarefamanu'))
                    descr_periodo = _safe_str(row_data.get('DESCR_PERIODO') or row_data.get('descr_periodo') or row_data.get('Descr_Periodo'), max_length=255)
                    
                    # ExecuÃ§Ã£o
                    dt_primexec = _safe_str(row_data.get('DT_PRIMEXEC') or row_data.get('dt_primexec') or row_data.get('Dt_Primexec'), max_length=50)
                    tempo_prev = _safe_str(row_data.get('TEMPO_PREV') or row_data.get('tempo_prev') or row_data.get('Tempo_Prev'), max_length=50)
                    qtde_periodo = _safe_int(row_data.get('QTDE_PERIODO') or row_data.get('qtde_periodo') or row_data.get('Qtde_Periodo'))
                    descr_seqplamanu = _safe_str(row_data.get('DESCR_SEQPLAMANU') or row_data.get('descr_seqplamanu') or row_data.get('Descr_Seqplamanu'), max_length=255)
                    cf_temp_prev = _safe_str(row_data.get('CF_TEMP_PREV') or row_data.get('cf_temp_prev') or row_data.get('Cf_Temp_Prev'), max_length=50)
                    
                    # Item do Plano
                    itemplanma_id = _safe_int(row_data.get('ITEMPLANMA_ID') or row_data.get('itemplanma_id') or row_data.get('Itemplanma_Id'))
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    descr_item = _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500)
                    item_id = _safe_int(row_data.get('ITEM_ID') or row_data.get('item_id') or row_data.get('Item_Id'))
                    qtde = _safe_int(row_data.get('QTDE') or row_data.get('qtde') or row_data.get('Qtde'))
                    qtde_saldo = _safe_int(row_data.get('QTDE_SALDO') or row_data.get('qtde_saldo') or row_data.get('Qtde_Saldo'))
                    qtde_reserva = _safe_int(row_data.get('QTDE_RESERVA') or row_data.get('qtde_reserva') or row_data.get('Qtde_Reserva'))
                    
                    # Validar que temos pelo menos cÃ³digo da mÃ¡quina ou descriÃ§Ã£o
                    if not cd_maquina and not descr_maquina:
                        errors.append(f"Linha {row_num}: CÃ³digo da mÃ¡quina ou descriÃ§Ã£o Ã© obrigatÃ³rio")
                        continue
                    
                    # Tentar encontrar mÃ¡quina relacionada
                    maquina = None
                    if cd_maquina:
                        # Verificar cache primeiro
                        if cd_maquina in maquinas_cache:
                            maquina = maquinas_cache[cd_maquina]
                        else:
                            try:
                                maquina = Maquina.objects.get(cd_maquina=cd_maquina)
                                maquinas_cache[cd_maquina] = maquina
                            except Maquina.DoesNotExist:
                                # Criar mÃ¡quina bÃ¡sica se nÃ£o existir
                                maquina = Maquina.objects.create(
                                    cd_maquina=cd_maquina,
                                    descr_maquina=descr_maquina or f'MÃ¡quina {cd_maquina}',
                                    cd_unid=cd_unid,
                                    nome_unid=nome_unid,
                                    cd_setormanut=cd_setormanut,
                                    descr_setormanut=descr_setormanut,
                                    cd_tpcentativ=cd_tpcentativ,
                                )
                                maquinas_cache[cd_maquina] = maquina
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    roteiro_data = {
                        'cd_unid': cd_unid,
                        'nome_unid': nome_unid,
                        'cd_funciomanu': cd_funciomanu,
                        'nome_funciomanu': nome_funciomanu,
                        'funciomanu_id': funciomanu_id,
                        'cd_setormanut': cd_setormanut,
                        'descr_setormanut': descr_setormanut,
                        'cd_tpcentativ': cd_tpcentativ,
                        'descr_abrev_tpcentativ': descr_abrev_tpcentativ,
                        'dt_abertura': dt_abertura,
                        'cd_ordemserv': cd_ordemserv,
                        'ordemserv_id': ordemserv_id,
                        'maquina': maquina,
                        'cd_maquina': cd_maquina,
                        'descr_maquina': descr_maquina,
                        'cd_planmanut': cd_planmanut,
                        'descr_planmanut': descr_planmanut,
                        'descr_recomenos': descr_recomenos,
                        'cf_dt_final_execucao': cf_dt_final_execucao,
                        'cs_qtde_periodo_max': cs_qtde_periodo_max,
                        'cs_tot_temp': cs_tot_temp,
                        'cf_tot_temp': cf_tot_temp,
                        'seq_seqplamanu': seq_seqplamanu,
                        'cd_tarefamanu': cd_tarefamanu,
                        'descr_tarefamanu': descr_tarefamanu,
                        'descr_periodo': descr_periodo,
                        'dt_primexec': dt_primexec,
                        'tempo_prev': tempo_prev,
                        'qtde_periodo': qtde_periodo,
                        'descr_seqplamanu': descr_seqplamanu,
                        'cf_temp_prev': cf_temp_prev,
                        'itemplanma_id': itemplanma_id,
                        'cd_item': cd_item,
                        'descr_item': descr_item,
                        'item_id': item_id,
                        'qtde': qtde,
                        'qtde_saldo': qtde_saldo,
                        'qtde_reserva': qtde_reserva,
                    }
                    
                    # Criar ou atualizar registro
                    if update_existing:
                        roteiro_obj, created = RoteiroPreventiva.objects.update_or_create(
                            cd_ordemserv=cd_ordemserv,
                            cd_planmanut=cd_planmanut,
                            seq_seqplamanu=seq_seqplamanu,
                            cd_tarefamanu=cd_tarefamanu,
                            defaults=roteiro_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        roteiro_obj, created = RoteiroPreventiva.objects.get_or_create(
                            cd_ordemserv=cd_ordemserv,
                            cd_planmanut=cd_planmanut,
                            seq_seqplamanu=seq_seqplamanu,
                            cd_tarefamanu=cd_tarefamanu,
                            defaults=roteiro_data
                        )
                        if created:
                            created_count += 1
                    
                except Exception as e:
                    errors.append(f"Linha {row_num}: {str(e)}")
                    continue
        
        return created_count, updated_count, errors
    
    except Exception as e:
        errors.append(f"Erro geral: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(row_data.get('CD_ITEM') or row_data.get('cd_item') or row_data.get('Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(row_data.get('CD_UNID') or row_data.get('cd_unid') or row_data.get('Cd_Unid')),
                        'nome_unid': _safe_str(row_data.get('NOME_UNID') or row_data.get('nome_unid') or row_data.get('Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(row_data.get('CD_USO_CTB') or row_data.get('cd_uso_ctb') or row_data.get('Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(row_data.get('DESCR_USO_CTB') or row_data.get('descr_uso_ctb') or row_data.get('Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(row_data.get('CD_DEPO') or row_data.get('cd_depo') or row_data.get('Cd_Depo')),
                        'descr_depo': _safe_str(row_data.get('DESCR_DEPO') or row_data.get('descr_depo') or row_data.get('Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(row_data.get('CD_LOCAL_FISIC') or row_data.get('cd_local_fisic') or row_data.get('Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(row_data.get('DESCR_LOCAL_FISIC') or row_data.get('descr_local_fisic') or row_data.get('Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(row_data.get('CD_EMBALAGEM') or row_data.get('cd_embalagem') or row_data.get('Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(row_data.get('DESCR_ITEM') or row_data.get('descr_item') or row_data.get('Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(row_data.get('CD_OPERACAO') or row_data.get('cd_operacao') or row_data.get('Cd_Operacao')),
                        'descr_operacao': _safe_str(row_data.get('DESCR_OPERACAO') or row_data.get('descr_operacao') or row_data.get('Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(row_data.get('CD_UNID_MEDIDA') or row_data.get('cd_unid_medida') or row_data.get('Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(row_data.get('QTDE_MOVTO_ESTOQ') or row_data.get('qtde_movto_estoq') or row_data.get('Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ') or row_data.get('vlr_movto_estoq') or row_data.get('Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(row_data.get('VLR_MOVTO_ESTOQ_REAV') or row_data.get('vlr_movto_estoq_reav') or row_data.get('Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(row_data.get('CD_UNID_BAIXA') or row_data.get('cd_unid_baixa') or row_data.get('Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(row_data.get('CD_CENTRO_ATIV') or row_data.get('cd_centro_ativ') or row_data.get('Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(row_data.get('CD_USU_CRIOU') or row_data.get('cd_usu_criou') or row_data.get('Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(row_data.get('CD_USU_ATEND') or row_data.get('cd_usu_atend') or row_data.get('Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(row_data.get('OBS RM') or row_data.get('obs_rm') or row_data.get('Obs_Rm')),
                        'obs_item': _safe_str(row_data.get('OBS ITEM') or row_data.get('obs_item') or row_data.get('Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_52_semanas_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de semanas (52 semanas) a partir de um arquivo Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import Semana52
    from datetime import datetime
    import re
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Mapeamento de meses em portuguÃªs para inglÃªs
    meses_pt_para_en = {
        'janeiro': 'January', 'fevereiro': 'February', 'marÃ§o': 'March', 'marco': 'March',
        'abril': 'April', 'maio': 'May', 'junho': 'June',
        'julho': 'July', 'agosto': 'August', 'setembro': 'September',
        'outubro': 'October', 'novembro': 'November', 'dezembro': 'December'
    }
    
    def limpar_texto(texto):
        """Remove caracteres especiais e normaliza espaÃ§os"""
        if not texto:
            return None
        texto = str(texto).strip()
        # Remover \xa0 (non-breaking space) e outros caracteres especiais
        texto = texto.replace('\xa0', ' ').replace('\u00a0', ' ')
        # Normalizar espaÃ§os mÃºltiplos
        texto = re.sub(r'\s+', ' ', texto)
        return texto.strip()
    
    def converter_data_pt_para_en(data_str):
        """Converte data em portuguÃªs para formato que datetime.strptime entende"""
        if not data_str:
            return None
        
        data_str = limpar_texto(data_str)
        if not data_str:
            return None
        
        # Tentar substituir meses em portuguÃªs por inglÃªs
        data_str_lower = data_str.lower()
        for mes_pt, mes_en in meses_pt_para_en.items():
            if mes_pt in data_str_lower:
                # Substituir mantendo o case original
                pattern = re.compile(re.escape(mes_pt), re.IGNORECASE)
                data_str = pattern.sub(mes_en, data_str)
                break
        
        return data_str
    
    def parse_date_value(value):
        """
        Tenta fazer parse de uma data em vÃ¡rios formatos diferentes.
        Suporta: strings de data, objetos datetime/date, nÃºmeros Excel, etc.
        """
        if not value:
            return None
        
        # Se jÃ¡ for um objeto date ou datetime
        from datetime import date, datetime
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        
        # Se for um nÃºmero (possÃ­vel data serial do Excel)
        if isinstance(value, (int, float)):
            try:
                # Excel usa 1 de janeiro de 1900 como base (mas tem bug, entÃ£o 1900-01-01 = 1)
                # Na verdade, Excel conta dias desde 1899-12-30
                from datetime import datetime, timedelta
                excel_epoch = datetime(1899, 12, 30)
                date_obj = excel_epoch + timedelta(days=int(value))
                return date_obj.date()
            except (ValueError, OverflowError):
                pass
        
        # Converter para string e tentar diferentes formatos
        value_str = str(value).strip()
        if not value_str:
            return None
        
        # Limpar e converter portuguÃªs para inglÃªs
        value_str = converter_data_pt_para_en(value_str)
        if not value_str:
            return None
        
        # Lista de formatos para tentar
        date_formats = [
            '%d %B %Y',           # "22 December 2025"
            '%d/%m/%Y',           # "22/12/2025"
            '%Y-%m-%d',           # "2025-12-22"
            '%d-%m-%Y',           # "22-12-2025"
            '%d.%m.%Y',           # "22.12.2025"
            '%d/%m/%y',           # "22/12/25"
            '%d-%m-%y',           # "22-12-25"
            '%d %b %Y',           # "22 Dec 2025" (abreviaÃ§Ã£o)
            '%d %B %y',           # "22 December 25"
            '%Y/%m/%d',           # "2025/12/22"
            '%m/%d/%Y',           # "12/22/2025" (formato US)
            '%d de %B de %Y',     # "22 de December de 2025" (com "de")
            '%d de %b de %Y',     # "22 de Dec de 2025"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue
        
        # Tentar parse com dateutil se disponÃ­vel (mais flexÃ­vel)
        try:
            from dateutil import parser
            return parser.parse(value_str, dayfirst=True).date()
        except (ImportError, ValueError, TypeError):
            pass
        
        return None
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo Excel
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            data = read_excel_file(file)
            print(f"DEBUG: Arquivo lido com sucesso. Total de linhas: {len(data)}")
            if data:
                print(f"DEBUG: Primeira linha de dados: {data[0]}")
        else:
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .xlsx, .xls ou .xlsm")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_idx, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 porque linha 1 Ã© cabeÃ§alho
                try:
                    # Verificar se a linha estÃ¡ vazia
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Normalizar nomes de colunas (case-insensitive)
                    semana_value = None
                    inicio_value = None
                    fim_value = None
                    
                    for key, value in row_data.items():
                        key_lower = str(key).lower().strip()
                        if 'semana' in key_lower:
                            semana_value = value
                        elif 'inicio' in key_lower or 'inÃ­cio' in key_lower:
                            inicio_value = value
                        elif 'fim' in key_lower:
                            fim_value = value
                    
                    # Debug: mostrar o que foi encontrado
                    if row_idx == 2:  # Apenas para primeira linha de dados
                        print(f"DEBUG Linha {row_idx}: semana_value={semana_value}, inicio_value={inicio_value}, fim_value={fim_value}")
                        print(f"DEBUG Linha {row_idx}: row_data keys={list(row_data.keys())}")
                    
                    if not semana_value:
                        errors.append(f"Linha {row_idx}: Campo 'semana' estÃ¡ vazio. Colunas encontradas: {', '.join(str(k) for k in row_data.keys())}")
                        continue
                    
                    # Limpar valores de semana
                    semana_value = limpar_texto(semana_value)
                    if not semana_value:
                        errors.append(f"Linha {row_idx}: Campo 'semana' estÃ¡ vazio apÃ³s limpeza")
                        continue
                    
                    # Converter datas usando funÃ§Ã£o melhorada
                    inicio_date = None
                    fim_date = None
                    
                    if inicio_value:
                        try:
                            inicio_date = parse_date_value(inicio_value)
                            if not inicio_date:
                                errors.append(f"Linha {row_idx}: Data de inÃ­cio invÃ¡lida ou formato nÃ£o reconhecido: {inicio_value}")
                        except Exception as e:
                            errors.append(f"Linha {row_idx}: Erro ao processar data de inÃ­cio '{inicio_value}': {str(e)}")
                    
                    if fim_value:
                        try:
                            fim_date = parse_date_value(fim_value)
                            if not fim_date:
                                errors.append(f"Linha {row_idx}: Data de fim invÃ¡lida ou formato nÃ£o reconhecido: {fim_value}")
                        except Exception as e:
                            errors.append(f"Linha {row_idx}: Erro ao processar data de fim '{fim_value}': {str(e)}")
                    
                    # Criar ou atualizar registro
                    # Usar semana + inicio como chave Ãºnica composta para permitir mÃºltiplas semanas com mesmo nome em anos diferentes
                    if update_existing:
                        if inicio_date:
                            # Se temos data de inÃ­cio, usar semana + inicio como chave Ãºnica
                            semana_obj, created = Semana52.objects.update_or_create(
                                semana=semana_value,
                                inicio=inicio_date,
                                defaults={
                                    'fim': fim_date
                                }
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            # Se nÃ£o temos data de inÃ­cio, verificar se existe registro com mesmo nome e sem data
                            try:
                                semana_obj = Semana52.objects.get(semana=semana_value, inicio__isnull=True)
                                # Atualizar registro existente
                                semana_obj.fim = fim_date
                                semana_obj.save()
                                updated_count += 1
                            except Semana52.DoesNotExist:
                                # Criar novo registro
                                semana_obj = Semana52.objects.create(
                                    semana=semana_value,
                                    inicio=inicio_date,
                                    fim=fim_date
                                )
                                created_count += 1
                            except Semana52.MultipleObjectsReturned:
                                # MÃºltiplos registros encontrados, atualizar o primeiro
                                semana_obj = Semana52.objects.filter(semana=semana_value, inicio__isnull=True).first()
                                if semana_obj:
                                    semana_obj.fim = fim_date
                                    semana_obj.save()
                                    updated_count += 1
                                else:
                                    # Criar novo registro se nenhum foi encontrado
                                    semana_obj = Semana52.objects.create(
                                        semana=semana_value,
                                        inicio=inicio_date,
                                        fim=fim_date
                                    )
                                    created_count += 1
                    else:
                        if inicio_date:
                            # Se temos data de inÃ­cio, usar semana + inicio como chave Ãºnica
                            semana_obj, created = Semana52.objects.get_or_create(
                                semana=semana_value,
                                inicio=inicio_date,
                                defaults={
                                    'fim': fim_date
                                }
                            )
                        else:
                            # Se nÃ£o temos data de inÃ­cio, verificar se jÃ¡ existe semana com mesmo nome e sem data
                            semana_obj, created = Semana52.objects.get_or_create(
                                semana=semana_value,
                                inicio__isnull=True,
                                defaults={
                                    'inicio': inicio_date,
                                    'fim': fim_date
                                }
                            )
                        if created:
                            created_count += 1
                        else:
                            # Registro jÃ¡ existe, ignorar
                            pass
                            
                except Exception as e:
                    error_msg = f"Linha {row_idx}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_idx}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_requisicoes_almoxarifado_from_file(file, data_requisicao, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de requisiÃ§Ãµes de almoxarifado a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        data_requisicao: Data da requisiÃ§Ã£o (datetime.date) - serÃ¡ associada a todos os itens
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Validar data_requisicao
    if not data_requisicao:
        errors.append("Data da requisiÃ§Ã£o Ã© obrigatÃ³ria")
        return 0, 0, errors
    
    # Converter data_requisicao para date se necessÃ¡rio
    if isinstance(data_requisicao, str):
        try:
            data_requisicao = datetime.strptime(data_requisicao, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Formato de data invÃ¡lido: {data_requisicao}. Use YYYY-MM-DD")
            return 0, 0, errors
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vÃ­rgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo nÃ£o suportado. Use .csv")
        
        # Tentar diferentes encodings - comeÃ§ar com latin-1 (mais comum para arquivos brasileiros)
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)  # Resetar arquivo para o inÃ­cio
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break  # Se conseguir ler, sair do loop
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:  # Se for o Ãºltimo encoding
                    raise ValidationError(f"Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel decodificar o arquivo com nenhum encoding testado (latin-1, iso-8859-1, utf-8, cp1252). Erro original: {str(e)}")
                continue  # Tentar prÃ³ximo encoding
            except Exception as e:
                # Outros erros (nÃ£o relacionados a encoding)
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: NÃ£o foi possÃ­vel processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados vÃ¡lidos")
        
        # Processar dados em transaÃ§Ã£o
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # ComeÃ§ar em 2 (linha 1 Ã© cabeÃ§alho)
                try:
                    # Verificar se a linha estÃ¡ vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # O CSV tem: CD_UNID;NOME_UNID;CD_USO_CTB;DESCR_USO_CTB;...
                    
                    # Função auxiliar para buscar valor com múltiplas variações de nome (incluindo espaços)
                    def _get_value(row_data, *keys):
                        """Busca valor no row_data tentando múltiplas chaves (case-insensitive, com/sem espaços)"""
                        for key in keys:
                            # Tentar exato
                            if key in row_data:
                                val = row_data[key]
                                if val is not None and str(val).strip():
                                    return val
                            # Tentar case-insensitive e com espaços normalizados
                            for row_key, row_val in row_data.items():
                                if row_key and row_key.strip().upper() == key.strip().upper():
                                    if row_val is not None and str(row_val).strip():
                                        return row_val
                        return None
                    
                    # Validar que temos pelo menos cÃ³digo do item
                    cd_item = _safe_int(_get_value(row_data, 'CD_ITEM', 'cd_item', 'Cd_Item'))
                    if not cd_item:
                        errors.append(f"Linha {row_num}: CÃ³digo do item (CD_ITEM) Ã© obrigatÃ³rio")
                        continue
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    requisicao_data = {
                        'data_requisicao': data_requisicao,
                        'cd_unid': _safe_int(_get_value(row_data, 'CD_UNID', 'cd_unid', 'Cd_Unid')),
                        'nome_unid': _safe_str(_get_value(row_data, 'NOME_UNID', 'nome_unid', 'Nome_Unid'), max_length=255),
                        'cd_uso_ctb': _safe_int(_get_value(row_data, 'CD_USO_CTB', 'cd_uso_ctb', 'Cd_Uso_Ctb')),
                        'descr_uso_ctb': _safe_str(_get_value(row_data, 'DESCR_USO_CTB', 'descr_uso_ctb', 'Descr_Uso_Ctb'), max_length=255),
                        'cd_depo': _safe_int(_get_value(row_data, 'CD_DEPO', 'cd_depo', 'Cd_Depo')),
                        'descr_depo': _safe_str(_get_value(row_data, 'DESCR_DEPO', 'descr_depo', 'Descr_Depo'), max_length=255),
                        'cd_local_fisic': _safe_int(_get_value(row_data, 'CD_LOCAL_FISIC', 'cd_local_fisic', 'Cd_Local_Fisic')),
                        'descr_local_fisic': _safe_str(_get_value(row_data, 'DESCR_LOCAL_FISIC', 'descr_local_fisic', 'Descr_Local_Fisic'), max_length=255),
                        'cd_item': cd_item,
                        'cd_embalagem': _safe_str(_get_value(row_data, 'CD_EMBALAGEM', 'cd_embalagem', 'Cd_Embalagem'), max_length=50),
                        'descr_item': _safe_str(_get_value(row_data, 'DESCR_ITEM', 'descr_item', 'Descr_Item'), max_length=500),
                        'cd_operacao': _safe_int(_get_value(row_data, 'CD_OPERACAO', 'cd_operacao', 'Cd_Operacao')),
                        'descr_operacao': _safe_str(_get_value(row_data, 'DESCR_OPERACAO', 'descr_operacao', 'Descr_Operacao'), max_length=255),
                        'cd_unid_medida': _safe_str(_get_value(row_data, 'CD_UNID_MEDIDA', 'cd_unid_medida', 'Cd_Unid_Medida'), max_length=50),
                        'qtde_movto_estoq': _safe_decimal(_get_value(row_data, 'QTDE_MOVTO_ESTOQ', 'qtde_movto_estoq', 'Qtde_Movto_Estoq')),
                        'vlr_movto_estoq': _safe_decimal(_get_value(row_data, 'VLR_MOVTO_ESTOQ', 'vlr_movto_estoq', 'Vlr_Movto_Estoq')),
                        'vlr_movto_estoq_reav': _safe_decimal(_get_value(row_data, 'VLR_MOVTO_ESTOQ_REAV', 'vlr_movto_estoq_reav', 'Vlr_Movto_Estoq_Reav')),
                        'cd_unid_baixa': _safe_int(_get_value(row_data, 'CD_UNID_BAIXA', 'cd_unid_baixa', 'Cd_Unid_Baixa')),
                        'cd_centro_ativ': _safe_int(_get_value(row_data, 'CD_CENTRO_ATIV', 'cd_centro_ativ', 'Cd_Centro_Ativ')),
                        'cd_usu_criou': _safe_str(_get_value(row_data, 'CD_USU_CRIOU', 'cd_usu_criou', 'Cd_Usu_Criou'), max_length=255),
                        'cd_usu_atend': _safe_str(_get_value(row_data, 'CD_USU_ATEND', 'cd_usu_atend', 'Cd_Usu_Atend'), max_length=255),
                        'obs_rm': _safe_str(_get_value(row_data, 'OBS RM', 'obs_rm', 'Obs_Rm')),
                        'obs_item': _safe_str(_get_value(row_data, 'OBS ITEM', 'obs_item', 'Obs_Item')),
                    }
                    
                    # Criar ou atualizar registro
                    # Usar data_requisicao + cd_item como chave Ãºnica (um item pode ser requisitado mÃºltiplas vezes na mesma data)
                    # Mas vamos usar apenas cd_item + data_requisicao para identificar duplicados
                    if update_existing:
                        requisicao, created = RequisicaoAlmoxarifado.objects.update_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Apenas criar novos registros
                        requisicao, created = RequisicaoAlmoxarifado.objects.get_or_create(
                            data_requisicao=data_requisicao,
                            cd_item=cd_item,
                            defaults=requisicao_data
                        )
                        if created:
                            created_count += 1
                        # Se nÃ£o foi criado, ignorar (jÃ¡ existe)
                            
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
        
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_msg = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_msg)
        print(f"Erro geral: {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors
def upload_notas_fiscais_from_file(file, update_existing=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de notas fiscais a partir de um arquivo CSV
    
    Args:
        file: Arquivo Django UploadedFile (CSV)
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import NotaFiscal
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vírgula)
        # Precisamos ler manualmente para tratar colunas duplicadas (ex: duas colunas "Situação")
        if file_name.endswith('.csv'):
            # utf-8-sig primeiro: arquivos gerados no sistema (download ESTF0198 etc.) vêm com BOM.
            # Se usar latin-1 antes, o BOM vira "ï»¿" no nome da 1ª coluna e o mapeamento quebra (Emitente vazio).
            data = None
            encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            
            for encoding in encodings_to_try:
                try:
                    file.seek(0)  # Resetar arquivo para o início
                    # Ler conteúdo do arquivo
                    if hasattr(file, 'read'):
                        content = file.read().decode(encoding)
                    else:
                        with open(file, 'r', encoding=encoding) as f:
                            content = f.read()
                    content = content.lstrip('\ufeff')
                    
                    # Ler CSV com csv.reader para suportar campos entre aspas (com ; dentro)
                    import csv
                    import io
                    import unicodedata
                    reader = csv.reader(io.StringIO(content), delimiter=';')
                    rows = list(reader)
                    if not rows:
                        raise ValidationError("Arquivo CSV vazio")
                    
                    headers_raw = [h.strip() for h in rows[0]]
                    header_count = {}
                    headers = []
                    for header in headers_raw:
                        header_lower = header.lower()
                        header_normalized = unicodedata.normalize('NFKD', header_lower).encode('ASCII', 'ignore').decode('ASCII')
                        if header_normalized in header_count:
                            header_count[header_normalized] += 1
                            headers.append(f"{header}_{header_count[header_normalized]}")
                        else:
                            header_count[header_normalized] = 1
                            headers.append(header)
                    
                    data = []
                    for line_num, values in enumerate(rows[1:], start=2):
                        if not values or not any(str(v).strip() for v in values):
                            continue
                        row_dict = {}
                        for idx, value in enumerate(values):
                            if idx < len(headers):
                                v = value.strip() if value else ''
                                if v:
                                    row_dict[headers[idx]] = v
                        if row_dict:
                            data.append(row_dict)
                    
                    break  # Se conseguir ler, sair do loop
                except (UnicodeDecodeError, ValidationError) as e:
                    if encoding == encodings_to_try[-1]:  # Se for o último encoding
                        raise ValidationError(
                            f"Erro ao ler arquivo CSV: Não foi possível decodificar o arquivo com nenhum encoding testado "
                            f"({'/'.join(encodings_to_try)}). Erro original: {str(e)}"
                        )
                    continue  # Tentar próximo encoding
                except Exception as e:
                    # Outros erros (não relacionados a encoding)
                    raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
            
            if data is None:
                raise ValidationError("Erro ao ler arquivo CSV: Não foi possível processar o arquivo.")
        else:
            raise ValidationError("Formato de arquivo não suportado. Use .csv")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados válidos")
        
        # Processar dados em transação
        with transaction.atomic():
            for row_num, row_data in enumerate(data, start=2):  # Começar em 2 (linha 1 é cabeçalho)
                try:
                    # Verificar se a linha está vazia ou tem apenas valores vazios
                    # Também ignorar linhas que começam com "Unidade" (linhas de resumo)
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # Verificar se é linha de resumo (começa com "Unidade")
                    primeiro_valor = str(list(row_data.values())[0] if row_data.values() else '').strip()
                    if primeiro_valor.upper() == 'UNIDADE':
                        continue
                    
                    # Mapear colunas do CSV para campos do modelo
                    # Emitente
                    emitente = _safe_str(row_data.get('Emitente') or row_data.get('emitente') or row_data.get('EMITENTE'), max_length=100)
                    nome_fantasia_emitente = _safe_str(row_data.get('Nome fantasia emitente') or row_data.get('nome fantasia emitente') or row_data.get('NOME FANTASIA EMITENTE'), max_length=500)
                    
                    # Dados da Nota Fiscal
                    nota = _safe_str(row_data.get('Nota') or row_data.get('nota') or row_data.get('NOTA'), max_length=50)
                    serie = _safe_str(row_data.get('Série') or row_data.get('série') or row_data.get('SÉRIE') or row_data.get('Serie') or row_data.get('SERIE'), max_length=50)
                    modelo = _safe_str(row_data.get('Modelo') or row_data.get('modelo') or row_data.get('MODELO'), max_length=50)
                    
                    # Total Nota (formato brasileiro: 2.982,99)
                    total_nota_str = row_data.get('Total Nota') or row_data.get('total nota') or row_data.get('TOTAL NOTA') or row_data.get('Total nota')
                    total_nota = _safe_decimal(total_nota_str)
                    
                    # Validar que temos pelo menos emitente e nota
                    if not emitente or not nota:
                        errors.append(f"Linha {row_num}: Emitente e Nota são obrigatórios")
                        continue
                    
                    # Datas
                    data_emissao = _safe_str(row_data.get('Data emissão') or row_data.get('data emissão') or row_data.get('DATA EMISSÃO') or row_data.get('Data emissao') or row_data.get('DATA EMISSAO'), max_length=50)
                    data_vencimento = _safe_str(row_data.get('Data vencimento') or row_data.get('data vencimento') or row_data.get('DATA VENCIMENTO'), max_length=50)
                    data_inclusao = _safe_str(row_data.get('Data inclusão') or row_data.get('data inclusão') or row_data.get('DATA INCLUSÃO') or row_data.get('Data inclusao') or row_data.get('DATA INCLUSAO'), max_length=50)
                    data_autorizacao = _safe_str(row_data.get('Data autorização') or row_data.get('data autorização') or row_data.get('DATA AUTORIZAÇÃO') or row_data.get('Data autorizacao') or row_data.get('DATA AUTORIZACAO'), max_length=50)
                    data_ult_sit_fechada = _safe_str(row_data.get('Data ult sit fechada') or row_data.get('data ult sit fechada') or row_data.get('DATA ULT SIT FECHADA') or row_data.get('Data ult sit fechada'), max_length=50)
                    
                    # Controle
                    ctrle = _safe_str(row_data.get('Ctrle') or row_data.get('ctrle') or row_data.get('CTRLE'), max_length=50)
                    
                    # Unidade
                    unidade = _safe_str(row_data.get('Unidade') or row_data.get('unidade') or row_data.get('UNIDADE'), max_length=50)
                    nome_unidade = _safe_str(row_data.get('Nome unidade') or row_data.get('nome unidade') or row_data.get('NOME UNIDADE'), max_length=255)
                    unidade_autorizacao = _safe_str(row_data.get('Unidade autorização') or row_data.get('unidade autorização') or row_data.get('UNIDADE AUTORIZAÇÃO') or row_data.get('Unidade autorizacao') or row_data.get('UNIDADE AUTORIZACAO'), max_length=50)
                    nome_unidade_autorizacao = _safe_str(row_data.get('Nome unidade autorização') or row_data.get('nome unidade autorização') or row_data.get('NOME UNIDADE AUTORIZAÇÃO') or row_data.get('Nome unidade autorizacao') or row_data.get('NOME UNIDADE AUTORIZACAO'), max_length=255)
                    
                    # Centro de Atividade
                    centro_atividade = _safe_str(row_data.get('Centro atividade') or row_data.get('centro atividade') or row_data.get('CENTRO ATIVIDADE'), max_length=50)
                    nome_centro_atividade = _safe_str(row_data.get('Nome centro atividade') or row_data.get('nome centro atividade') or row_data.get('NOME CENTRO ATIVIDADE'), max_length=255)
                    
                    # Situação - primeira coluna (ex: "AUTORIZADA")
                    situacao = _safe_str(
                        row_data.get('Situação') or 
                        row_data.get('situação') or 
                        row_data.get('SITUAÇÃO') or 
                        row_data.get('Situacao') or 
                        row_data.get('SITUACAO') or
                        row_data.get('Situação_1') or  # Primeira ocorrência quando há duplicatas
                        row_data.get('situação_1') or
                        row_data.get('SITUAÇÃO_1'),
                        max_length=100
                    )
                    
                    # Situação detalhada - segunda coluna (ex: "Solicitação autorizada por: ...")
                    # Nota: duplicatas são nomeadas como Situação_1, Situação_2, etc.
                    situacao_detalhada = _safe_str(
                        row_data.get('Situação_2') or row_data.get('Situação_1') or
                        row_data.get('situação_2') or row_data.get('situação_1') or
                        row_data.get('SITUAÇÃO_2') or row_data.get('SITUAÇÃO_1') or
                        row_data.get('Situacao_2') or row_data.get('Situacao_1') or
                        row_data.get('SITUACAO_2') or row_data.get('SITUACAO_1')
                    )
                    
                    # Usuário e Autorização (suporta encoding latin-1 e utf-8)
                    nome_usuario = _safe_str(
                        row_data.get('Nome usuário') or row_data.get('Nome usuÃ¡rio') or
                        row_data.get('nome usuário') or row_data.get('Nome usuario') or
                        row_data.get('NOME USUARIO'), max_length=255
                    )
                    autorizador = _safe_str(row_data.get('Autorizador') or row_data.get('autorizador') or row_data.get('AUTORIZADOR'), max_length=255)
                    
                    # Observações (suporta encoding latin-1 e utf-8)
                    observacoes = _safe_str(
                        row_data.get('Observações') or row_data.get('ObservaÃ§Ãµes') or
                        row_data.get('observações') or row_data.get('Observacoes') or
                        row_data.get('OBSERVACOES')
                    )
                    observacoes_csc = _safe_str(
                        row_data.get('Observações CSC') or row_data.get('ObservaÃ§Ãµes CSC') or
                        row_data.get('observações csc') or row_data.get('Observacoes CSC') or
                        row_data.get('OBSERVACOES CSC')
                    )
                    observacoes_autorizacao = _safe_str(
                        row_data.get('Observações autorização') or row_data.get('ObservaÃ§Ãµes autorizaÃ§Ã£o') or
                        row_data.get('observações autorização') or row_data.get('Observacoes autorizacao') or
                        row_data.get('OBSERVACOES AUTORIZACAO')
                    )
                    
                    # LanÃ§amento
                    lancamento_tesf0028 = _safe_str(row_data.get('LANCAMENTO TESF0028') or row_data.get('lancamento tesf0028') or row_data.get('Lancamento TESF0028'), max_length=255)
                    
                    # Uso Contábil - extrair apenas parte inteira (242,00 -> 242)
                    uso_contabil_raw = (
                        row_data.get('Uso contábil') or row_data.get('Uso contabil') or
                        row_data.get('uso contábil') or row_data.get('uso contabil') or
                        row_data.get('USO CONTÁBIL') or row_data.get('USO CONTABIL')
                    )
                    uso_contabil = _safe_uso_contabil(uso_contabil_raw, default='')
                    
                    # Preparar dados para criaÃ§Ã£o/atualizaÃ§Ã£o
                    nota_data = {
                        'emitente': emitente,
                        'nome_fantasia_emitente': nome_fantasia_emitente,
                        'nota': nota,
                        'serie': serie,
                        'modelo': modelo,
                        'total_nota': total_nota,
                        'uso_contabil': uso_contabil,
                        'data_emissao': data_emissao,
                        'data_vencimento': data_vencimento,
                        'data_inclusao': data_inclusao,
                        'data_autorizacao': data_autorizacao,
                        'data_ult_sit_fechada': data_ult_sit_fechada,
                        'ctrle': ctrle,
                        'unidade': unidade,
                        'nome_unidade': nome_unidade,
                        'unidade_autorizacao': unidade_autorizacao,
                        'nome_unidade_autorizacao': nome_unidade_autorizacao,
                        'centro_atividade': centro_atividade,
                        'nome_centro_atividade': nome_centro_atividade,
                        'situacao': situacao,
                        'situacao_detalhada': situacao_detalhada,
                        'nome_usuario': nome_usuario,
                        'autorizador': autorizador,
                        'observacoes': observacoes,
                        'observacoes_csc': observacoes_csc,
                        'observacoes_autorizacao': observacoes_autorizacao,
                        'lancamento_tesf0028': lancamento_tesf0028,
                    }
                    
                    # Normalizar serie/modelo vazios para '' (consistência com unique_together)
                    serie_norm = (serie or '').strip()
                    modelo_norm = (modelo or '').strip()
                    
                    # Lookup considera NULL e '' como equivalentes (registros antigos podem ter NULL)
                    from django.db.models import Q
                    lookup = Q(emitente=emitente, nota=nota)
                    if serie_norm:
                        lookup &= Q(serie=serie_norm)
                    else:
                        lookup &= (Q(serie='') | Q(serie__isnull=True))
                    if modelo_norm:
                        lookup &= Q(modelo=modelo_norm)
                    else:
                        lookup &= (Q(modelo='') | Q(modelo__isnull=True))
                    
                    if update_existing:
                        nota_obj = None
                        # Priorizar registro com serie/modelo vazios (legado) para evitar duplicatas
                        # Ex: NF 20100 METALURGICA BALENA - BD tem (emitente, nota, '', '') e (emitente, nota, 'S', '99')
                        if serie_norm or modelo_norm:
                            lookup_fallback = Q(emitente=emitente, nota=nota) & (
                                Q(serie='') | Q(serie__isnull=True)
                            ) & (Q(modelo='') | Q(modelo__isnull=True))
                            candidates = list(NotaFiscal.objects.filter(lookup_fallback))
                            if len(candidates) == 1:
                                nota_obj = candidates[0]
                                # Remover duplicata com serie/modelo do CSV para evitar unique_together ao salvar
                                dup_lookup = Q(emitente=emitente, nota=nota)
                                if serie_norm:
                                    dup_lookup &= Q(serie=serie_norm)
                                else:
                                    dup_lookup &= (Q(serie='') | Q(serie__isnull=True))
                                if modelo_norm:
                                    dup_lookup &= Q(modelo=modelo_norm)
                                else:
                                    dup_lookup &= (Q(modelo='') | Q(modelo__isnull=True))
                                NotaFiscal.objects.filter(dup_lookup).exclude(id=nota_obj.id).delete()
                        if nota_obj is None:
                            try:
                                nota_obj = NotaFiscal.objects.get(lookup)
                            except NotaFiscal.DoesNotExist:
                                pass
                            except NotaFiscal.MultipleObjectsReturned:
                                nota_obj = NotaFiscal.objects.filter(lookup).first()
                        if nota_obj:
                            for key, val in nota_data.items():
                                setattr(nota_obj, key, val)
                            nota_obj.save()
                            updated_count += 1
                        else:
                            nota_obj = NotaFiscal.objects.create(**nota_data)
                            created_count += 1
                    else:
                        if NotaFiscal.objects.filter(lookup).exists():
                            pass  # Já existe - ignorar
                        else:
                            NotaFiscal.objects.create(**nota_data)
                            created_count += 1
                    
                except Exception as e:
                    error_msg = f"Linha {row_num}: Erro ao processar registro - {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
        
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_projecao_gastos_from_file(file, update_existing=False, debug=False) -> Tuple[int, int, List[str]]:
    """
    Faz upload de projeções de gastos a partir de um arquivo Excel

    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
        debug: Se True, adiciona informações de estrutura do arquivo em errors.

    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import ProjecaoGasto
    from datetime import datetime
    import re
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    print(f"[ProjecaoGastos] Import: {file.name}")
    
    try:
        # Ler arquivo baseado na extensão
        if file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            # Construir lista de planilhas a tentar: GASTOS exato, depois com "GASTO", depois todas as demais
            sheets_to_try = ['GASTOS', 'Gastos', 'gastos']
            num_sheets = 0
            try:
                if hasattr(file, 'seek'):
                    file.seek(0)
                wb_temp = openpyxl.load_workbook(file, read_only=True, data_only=True)
                num_sheets = len(wb_temp.sheetnames)
                gasto_sheets = [s for s in wb_temp.sheetnames if 'GASTO' in str(s).upper() and s not in sheets_to_try]
                other_sheets = [s for s in wb_temp.sheetnames if s not in sheets_to_try and s not in gasto_sheets]
                sheets_to_try = [s for s in sheets_to_try if s in wb_temp.sheetnames] + gasto_sheets + other_sheets
                sheets_to_try.append(None)  # planilha ativa (fallback)
                try:
                    wb_temp.close()
                except Exception:
                    pass
            except Exception:
                sheets_to_try = ['GASTOS', 'Gastos', 'gastos', None]

            data = None
            used_header_row = 1
            sheet_used = None
            # Normalizar nome de coluna para detecção (permite acentos/encoding diferentes entre arquivos)
            def _header_has_required_keys(keys):
                """Exige coluna exatamente 'ID' e coluna SETOR (ou CENTRO)."""
                if not keys:
                    return False
                import unicodedata
                has_id = False
                has_setor = False
                for k in keys:
                    if not k:
                        continue
                    n = str(k).replace('\n', ' ').replace('\r', ' ').strip().upper()
                    n = unicodedata.normalize('NFD', n)
                    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
                    n = re.sub(r'\s+', ' ', n).strip()
                    if n == 'ID':
                        has_id = True
                    if n and ('SETOR' in n or 'CENTRO' in n):
                        has_setor = True
                    if has_id and has_setor:
                        return True
                return False
            def _accept_data(data_try, keys_try):
                return _header_has_required_keys(keys_try)
            # Primeiro: tentar por índice de planilha (evita problemas de encoding no nome, ex: "1. Manutenção Turno A")
            for idx in range(max(num_sheets, 1)):
                for header_row_try in [2, 1, 3, 4, 5, 6, 7]:
                    try:
                        if hasattr(file, 'seek'):
                            file.seek(0)
                        data_try = read_excel_file(file, sheet_index=idx, header_row=header_row_try)
                        if data_try and len(data_try) > 0:
                            keys_try = list(data_try[0].keys())
                            if _accept_data(data_try, keys_try):
                                data = data_try
                                used_header_row = header_row_try
                                first_row_keys = keys_try
                                sheet_used = f'(planilha {idx + 1})'
                                break
                    except Exception:
                        continue
                if data:
                    break
            # Depois: tentar por nome da planilha (para compatibilidade)
            if not data:
                for sheet_try in sheets_to_try:
                    for header_row_try in [2, 1, 3, 4, 5, 6, 7]:
                        try:
                            if hasattr(file, 'seek'):
                                file.seek(0)
                            data_try = read_excel_file(file, sheet_name=sheet_try, header_row=header_row_try)
                            if data_try and len(data_try) > 0:
                                keys_try = list(data_try[0].keys())
                                if _accept_data(data_try, keys_try):
                                    data = data_try
                                    used_header_row = header_row_try
                                    first_row_keys = keys_try
                                    sheet_used = sheet_try if sheet_try else '(planilha ativa)'
                                    break
                        except Exception:
                            continue
                    if data:
                        break
            if debug and sheet_used:
                errors.append(f"[DEBUG] Planilha usada: {sheet_used}, linha de cabeçalho: {used_header_row}, {len(data)} linha(s) de dados.")
            print(f"[ProjecaoGastos] Sheet: {sheet_used}, header_row: {used_header_row}, rows: {len(data) if data else 0}")
            if not data:
                raise ValidationError(
                    "Nenhuma planilha com dados encontrada. A planilha deve ter uma coluna chamada exatamente 'ID' e uma coluna 'SETOR' (ou 'CENTRO')."
                )
            
            first_row_keys = list(data[0].keys())
            
            # Inferir setor do nome do arquivo quando a coluna SETOR estiver vazia (ex: "1. Manutenção Turno A - Previsão....xlsx" -> "1. Manutenção Turno A")
            setor_from_filename = None
            base_name = (file.name or '').strip()
            for ext in ('.xlsx', '.xls', '.xlsm'):
                if base_name.lower().endswith(ext):
                    base_name = base_name[:-len(ext)].strip()
                    break
            if base_name and ' - ' in base_name:
                setor_from_filename = base_name.split(' - ')[0].strip()
                if len(setor_from_filename) > 100:
                    setor_from_filename = setor_from_filename[:100]
            elif base_name and re.match(r'^\d+\.\s*.+', base_name):
                setor_from_filename = base_name[:100].strip()
        
        # Helper para normalizar texto (remove acentos, colapsa espaços) para matching flexível
        def _normalize_key(s):
            if not s:
                return ''
            s = str(s).replace('\n', ' ').replace('\r', ' ')
            s = re.sub(r'\s+', ' ', s).strip().upper()
            # Remover acentos para matching mais flexível
            import unicodedata
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s

        # Função auxiliar para encontrar valor de coluna (case-insensitive e normaliza quebras de linha)
        def find_column_value(row_data, possible_names, fallback_keywords=None):
            """Encontra valor de coluna considerando variações de nome. fallback_keywords: lista de palavras que devem estar no nome da coluna."""
            for name in possible_names:
                # Tentar nome exato
                if name in row_data:
                    return row_data[name]
                # Tentar normalizado (sem quebras de linha, espaços normalizados)
                normalized_name = re.sub(r'\s+', ' ', name.replace('\n', ' ').replace('\r', ' ').strip())
                for key in row_data.keys():
                    normalized_key = re.sub(r'\s+', ' ', key.replace('\n', ' ').replace('\r', ' ').strip())
                    if normalized_key.upper() == normalized_name.upper():
                        return row_data[key]
                # Tentar com acentos removidos
                for key in row_data.keys():
                    if _normalize_key(key) == _normalize_key(name):
                        return row_data[key]
            # Fallback: buscar por palavras-chave (ex: coluna que contenha "SERVI" e "CONCLU")
            if fallback_keywords:
                for key in row_data.keys():
                    key_norm = _normalize_key(key)
                    if all(kw in key_norm for kw in fallback_keywords):
                        return row_data[key]
            return None
        
        # Função para extrair mês e ano de previsão
        def extract_mes_ano(previsao_str):
            """Extrai mês e ano de uma string como 'OUTUBRO / 2025'"""
            if not previsao_str:
                return None, None
            previsao_str = str(previsao_str).strip().upper()
            # Mapear meses em português
            meses = {
                'JANEIRO': '01', 'FEVEREIRO': '02', 'MARÇO': '03', 'MARCO': '03',
                'ABRIL': '04', 'MAIO': '05', 'JUNHO': '06',
                'JULHO': '07', 'AGOSTO': '08', 'SETEMBRO': '09',
                'OUTUBRO': '10', 'NOVEMBRO': '11', 'DEZEMBRO': '12'
            }
            # Procurar padrão "MÊS / ANO" ou "MÊS/ANO"
            match = re.search(r'([A-ZÇ]+)\s*/\s*(\d{4})', previsao_str)
            if match:
                mes_nome = match.group(1)
                ano = int(match.group(2))
                mes = meses.get(mes_nome)
                return mes, ano
            return None, None
        
        # Processar dados em transação
        excel_servico_imported_count = 0
        with transaction.atomic():
            for row_num, row in enumerate(data, start=used_header_row + 1):
                row_data = row
                try:
                    # Verificar se a linha está vazia ou tem apenas valores vazios
                    if not any(str(v).strip() if v else '' for v in row_data.values()):
                        continue
                    
                    # ID: apenas da coluna cujo cabeçalho é exatamente "ID" (case-insensitive, ignora espaços)
                    id_excel_raw = find_column_value(row_data, ['ID', 'id', 'Id'])
                    if id_excel_raw is None or (isinstance(id_excel_raw, str) and not str(id_excel_raw).strip()):
                        errors.append(f"Linha {row_num}: coluna ID não encontrada ou vazia. Linha ignorada.")
                        continue
                    try:
                        s = str(id_excel_raw).strip()
                        id_excel = int(float(s))
                    except (ValueError, TypeError):
                        errors.append(f"Linha {row_num}: ID inválido '{id_excel_raw}'. Linha ignorada.")
                        continue
                    
                    # Processar todas as linhas com ID válido (não pular por falta de outros campos)
                    # Se a linha só tem ID e todos os outros campos estão vazios, pular
                    # Verificar se pelo menos um campo importante tem valor
                    # (verificamos apenas alguns campos principais para eficiência)
                    has_meaningful_data = False
                    
                    # Lista de grupos de campos para verificar (cada grupo tem variações possíveis)
                    field_groups = [
                        ['SETOR ', 'SETOR', 'setor', 'CENTRO', 'CENTRO DE CUSTO'],
                        ['SOLICITANTE', 'solicitante'],
                        ['FORNECEDOR\nNOME FANTASIA', 'FORNECEDOR NOME FANTASIA', 'FORNECEDOR'],
                        ['FORNECEDOR\nCNPJ', 'FORNECEDOR CNPJ'],
                        ['DESCRIÇÃO DO SERVIÇO', 'DESCRI��O DO SERVI�O', 'DESCRICAO DO SERVICO'],
                        ['VALOR TOTAL', 'valor_total'],
                        ['PREVISÃO \nP/ EXECUÇÃO', 'PREVISÃO P/ EXECUÇÃO', 'PREVIS�O \nP/ EXECU��O'],
                        ['USO \nCONTÁBIL', 'USO CONTÁBIL', 'USO \nCONT�BIL'],
                        ['NÚMERO DA \nNOTA FISCAL', 'NÚMERO DA NOTA FISCAL', 'N�MERO DA \nNOTA FISCAL'],
                        ['TIPO DE \nSOLICITAÇÃO', 'TIPO DE SOLICITAÇÃO', 'TIPO DE \nSOLICITACAO', 'TIPO DE SOLICITACAO'],
                        ['ORDEM \nDE SERVIÇO', 'ORDEM DE SERVIÇO', 'ORDEM \nDE SERVI�O'],
                        ['DATA DE ABERTURA \nDA REQUISIÇÃO', 'DATA DE ABERTURA DA REQUISIÇÃO', 'DATA DE ABERTURA \nDA REQUISI��O'],
                        ['NÚMERO DA REQUISIÇÃO \nDE COMPRA', 'NÚMERO DA REQUISIÇÃO DE COMPRA', 'N�EMRO DA REQUISI��O \nDE COMPRA'],
                        ['NÚMERO DO \nPEDIDO DE COMPRA', 'NÚMERO DO PEDIDO DE COMPRA', 'N�MERO DO \nPEDIDO DE COMPRA'],
                        ['SERVIÇO CONCLUÍDO', 'SERVIÇO\nCONCLUÍDO', 'SERVICO CONCLUIDO', 'SERVI�O CONCLU�DO'],
                        ['NF DE SERVIÇO RECEBIDA', 'NF DE SERVIÇO\n RECEBIDA', 'NF DE SERVICO RECEBIDA', 'NF DE SERVI�O\n RECEBIDA'],
                        ['NF ENVIADA PARA LANÇAMENTO', 'NF ENVIADA\n PARA LANÇAMENTO ', 'NF ENVIADA PARA LANCAMENTO', 'NF ENVIADA\n PARA LANAMENTO '],
                        ['OBSERVAÇÕES', 'OBSERVAES', 'OBSERVACOES', 'OBSERVAÇÕES ']
                    ]
                    
                    # Verificar se pelo menos um campo além do ID tem valor
                    for field_group in field_groups:
                        value = find_column_value(row_data, field_group)
                        if value and str(value).strip():
                            has_meaningful_data = True
                            break
                    # Se a linha só tem ID e nenhum outro dado, não importar (evita linhas vazias em planilhas com muitas linhas)
                    if not has_meaningful_data:
                        continue
                    
                    # Mapear colunas do Excel para campos do modelo
                    setor = _safe_str(find_column_value(row_data, ['SETOR ', 'SETOR', 'setor', 'CENTRO', 'CENTRO DE CUSTO']), max_length=100)
                    solicitante = _safe_str(find_column_value(row_data, ['SOLICITANTE', 'solicitante']), max_length=100)
                    fornecedor_nome = _safe_str(find_column_value(row_data, ['FORNECEDOR\nNOME FANTASIA', 'FORNECEDOR NOME FANTASIA', 'FORNECEDOR']), max_length=255)
                    fornecedor_cnpj = _safe_str(find_column_value(row_data, ['FORNECEDOR\nCNPJ', 'FORNECEDOR CNPJ']), max_length=20)
                    descricao = _safe_str(find_column_value(row_data, ['DESCRIÇÃO DO SERVIÇO', 'DESCRI��O DO SERVI�O', 'DESCRICAO DO SERVICO']), max_length=500)
                    valor_total = _safe_decimal(find_column_value(row_data, ['VALOR TOTAL', 'valor_total']))
                    previsao_execucao = _safe_str(find_column_value(row_data, ['PREVISÃO \nP/ EXECUÇÃO', 'PREVISÃO P/ EXECUÇÃO', 'PREVIS�O \nP/ EXECU��O']), max_length=50)
                    uso_contabil = _safe_str(find_column_value(row_data, ['USO \nCONTÁBIL', 'USO CONTÁBIL', 'USO \nCONT�BIL']), max_length=100)
                    numero_nf = _safe_str(find_column_value(row_data, ['NÚMERO DA \nNOTA FISCAL', 'NÚMERO DA NOTA FISCAL', 'N�MERO DA \nNOTA FISCAL']), max_length=100)
                    tipo_solicitacao = _safe_str(find_column_value(row_data, ['TIPO DE \nSOLICITAÇÃO', 'TIPO DE SOLICITAÇÃO', 'TIPO DE \nSOLICITACAO', 'TIPO DE SOLICITACAO']), max_length=100)
                    numero_ordem_servico = _safe_str(find_column_value(row_data, ['ORDEM \nDE SERVIÇO', 'ORDEM DE SERVIÇO', 'ORDEM \nDE SERVI�O']), max_length=100)
                    data_abertura = _safe_date(find_column_value(row_data, ['DATA DE ABERTURA \nDA REQUISIÇÃO', 'DATA DE ABERTURA DA REQUISIÇÃO', 'DATA DE ABERTURA \nDA REQUISI��O']))
                    numero_requisicao_compra = _safe_str(find_column_value(row_data, ['NÚMERO DA REQUISIÇÃO \nDE COMPRA', 'NÚMERO DA REQUISIÇÃO DE COMPRA', 'N�EMRO DA REQUISI��O \nDE COMPRA']), max_length=100)
                    numero_pedido_compra = _safe_str(find_column_value(row_data, ['NÚMERO DO \nPEDIDO DE COMPRA', 'NÚMERO DO PEDIDO DE COMPRA', 'N�MERO DO \nPEDIDO DE COMPRA']), max_length=100)
                    # Importar como está (sem análise/tratamento): valores brutos do Excel
                    def _as_is(val, max_len=255):
                        if val is None: return None
                        s = str(val).strip()
                        return s[:max_len] if s else None
                    def _get_by_name_or_pos(row, names, fallback_kw, col_idx):
                        """Tenta por nome, depois por posição (PCM: P=15, Q=16, R=17)."""
                        v = find_column_value(row, names, fallback_keywords=fallback_kw)
                        if v is not None and str(v).strip():
                            return v
                        keys = list(row.keys())
                        if col_idx < len(keys):
                            return row.get(keys[col_idx])
                        return None
                    servico_concluido = _as_is(_get_by_name_or_pos(row_data, ['SERVIÇO CONCLUÍDO', 'SERVIÇO\nCONCLUÍDO', 'SERVICO CONCLUIDO', 'SERVI�O CONCLU�DO'], ['SERVICO', 'CONCLU'], 15))
                    if servico_concluido:
                        excel_servico_imported_count += 1
                    nf_servico_recebida = _as_is(_get_by_name_or_pos(row_data, ['NF DE SERVIÇO RECEBIDA', 'NF DE SERVIÇO\n RECEBIDA', 'NF DE SERVICO RECEBIDA', 'NF DE SERVI�O\n RECEBIDA'], ['RECEBIDA', 'SERVI'], 16))
                    nf_enviada_lancamento = _as_is(_get_by_name_or_pos(row_data, ['NF ENVIADA PARA LANÇAMENTO', 'NF ENVIADA\n PARA LANÇAMENTO ', 'NF ENVIADA PARA LANCAMENTO', 'NF ENVIADA\n PARA LANAMENTO '], ['ENVIADA', 'LANCAMENTO'], 17))
                    observacoes = _safe_str(find_column_value(row_data, ['OBSERVAÇÕES', 'OBSERVAES', 'OBSERVACOES', 'OBSERVAÇÕES ']), max_length=None)
                    
                    # Extrair mês e ano da previsão
                    mes_referencia, ano_referencia = extract_mes_ano(previsao_execucao)
                    
                    # Colunas mapeadas para campos do modelo (nomes normalizados UPPER para comparação)
                    mapped_column_names = {
                        'ID', 'SETOR', 'SETOR ', 'SOLICITANTE',
                        'FORNECEDOR NOME FANTASIA', 'FORNECEDOR CNPJ', 'FORNECEDOR',
                        'DESCRIÇÃO DO SERVIÇO', 'DESCRICAO DO SERVICO', 'VALOR TOTAL',
                        'PREVISÃO P/ EXECUÇÃO', 'PREVISO P/ EXECUO', 'USO CONTÁBIL', 'USO CONTBIL',
                        'NÚMERO DA NOTA FISCAL', 'NMERO DA NOTA FISCAL',
                        'TIPO DE SOLICITAÇÃO', 'TIPO DE SOLICITACAO', 'ORDEM DE SERVIÇO',
                        'DATA DE ABERTURA DA REQUISIÇÃO', 'DATA DE ABERTURA DA REQUISIO',
                        'NÚMERO DA REQUISIÇÃO DE COMPRA', 'NMERO DA REQUISIO DE COMPRA',
                        'NÚMERO DO PEDIDO DE COMPRA', 'NMERO DO PEDIDO DE COMPRA',
                        'SERVIÇO CONCLUÍDO', 'SERVICO CONCLUIDO', 'SERVIO CONCLUDO',
                        'NF DE SERVIÇO RECEBIDA', 'NF DE SERVIO RECEBIDA',
                        'NF ENVIADA PARA LANÇAMENTO', 'NF ENVIADA PARA LANAMENTO',
                        'OBSERVAÇÕES', 'OBSERVACOES'
                    }
                    
                    # Coletar colunas do Excel não mapeadas em dados_adicionais
                    dados_adicionais = {}
                    for excel_key, val in row_data.items():
                        if val is None or (isinstance(val, str) and not str(val).strip()):
                            continue
                        norm_key = re.sub(r'\s+', ' ', str(excel_key).replace('\n', ' ').strip()).upper()
                        # Verificar se esta coluna já está mapeada
                        if norm_key in mapped_column_names:
                            continue
                        # Serializar valor para JSON
                        if hasattr(val, 'isoformat'):
                            dados_adicionais[excel_key] = val.isoformat()
                        elif isinstance(val, (int, float)):
                            dados_adicionais[excel_key] = int(val) if isinstance(val, float) and val == int(val) else val
                        else:
                            dados_adicionais[excel_key] = _safe_str(val, max_length=500)
                    
                    # SETOR: apenas da coluna SETOR do Excel; se vazio, usar "SEM SETOR" (não inferir do nome do arquivo)
                    setor_value = re.sub(r'\s+', ' ', str(setor).strip()) if setor and str(setor).strip() else 'SEM SETOR'
                    
                    # Preparar dados para criação/atualização
                    projecao_data = {
                        'setor': setor_value,
                        'solicitante': solicitante,
                        'descricao': descricao,
                        'tipo_solicitacao': tipo_solicitacao,
                        'valor_total': valor_total,
                        'data_abertura_requisicao': data_abertura,
                        'previsao_execucao': previsao_execucao,
                        'mes_referencia': mes_referencia,
                        'ano_referencia': ano_referencia,
                        'fornecedor_nome_fantasia': fornecedor_nome,
                        'fornecedor_cnpj': fornecedor_cnpj,
                        'uso_contabil': uso_contabil,
                        'numero_nf': numero_nf,
                        'numero_ordem_servico': numero_ordem_servico,
                        'numero_requisicao_compra': numero_requisicao_compra,
                        'numero_pedido_compra': numero_pedido_compra,
                        'servico_concluido': servico_concluido,
                        'nf_servico_recebida': nf_servico_recebida,
                        'nf_enviada_lancamento': nf_enviada_lancamento,
                        'observacoes': observacoes,  # Campo OBSERVAÇÕES do Excel
                        'dados_adicionais': dados_adicionais if dados_adicionais else None,
                    }
                    
                    # Atualização/criação: apenas por id_excel + setor (exato); sem fallbacks
                    if update_existing:
                        projecao_obj, created = ProjecaoGasto.objects.update_or_create(
                            id_excel=id_excel,
                            setor=setor_value,
                            defaults=projecao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        projecao_obj, created = ProjecaoGasto.objects.get_or_create(
                            id_excel=id_excel,
                            setor=setor_value,
                            defaults=projecao_data
                        )
                        if created:
                            created_count += 1
                        else:
                            if servico_concluido is not None:
                                projecao_obj.servico_concluido = servico_concluido
                                projecao_obj.save(update_fields=['servico_concluido'])
                                updated_count += 1
                    
                        
                except Exception as e:
                    error_msg = f"Linha {row_num}: {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro ao processar linha {row_num}: {error_msg}")
                    continue
        
        # Transação commitada ao sair do with
        if debug:
            total_in_db = ProjecaoGasto.objects.count()
            errors.append(f"[DEBUG] Após import: {total_in_db} registro(s) na tabela Projeção de Gastos.")
        if excel_servico_imported_count > 0:
            errors.append(f"Info: {excel_servico_imported_count} linha(s) com SERVIÇO CONCLUÍDO importadas.")
        elif data and len(data) > 1:
            errors.append("Aviso: Nenhum valor de SERVIÇO CONCLUÍDO foi encontrado. Verifique se a coluna existe e tem dados.")
        if created_count == 0 and updated_count == 0 and data and len(data) > 1 and not any('ID não encontrado' in e for e in errors):
            total_rows = len(data)
            errors.append(f"[INFO] Nenhum registro criado ou atualizado. Verifique: 1) Opção 'Atualizar registros existentes' está marcada? 2) ID e SETOR no Excel correspondem aos do banco? 3) {total_rows} linha(s) de dados foram lidas.")
        print(f"[ProjecaoGastos] Result: created={created_count}, updated={updated_count}, errors={len(errors)}")
        if errors:
            for e in errors[:5]:
                print(f"[ProjecaoGastos]   {e[:120]}")
        return created_count, updated_count, errors
    
    except ValidationError as e:
        errors.append(str(e))
        print(f"[ProjecaoGastos] ValidationError: {e}")
        return 0, 0, errors
    except Exception as e:
        error_detail = f"Erro geral ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"[ProjecaoGastos] Erro geral: {error_detail}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors


def upload_controle_rc_e_nf_from_file(file, update_existing=False) -> Tuple[int, int, List[str], List[dict]]:
    """
    Faz upload de controle RC e NF a partir de um arquivo Excel
    
    Args:
        file: Arquivo Django UploadedFile
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
    
    Returns:
        Tupla (created_count, updated_count, errors, duplicates)
    """
    from app.models import ControleRCeNF
    from datetime import datetime, date
    from decimal import Decimal, InvalidOperation
    
    errors = []
    duplicates = []
    created_count = 0
    updated_count = 0
    
    try:
        # Ler arquivo Excel
        if hasattr(file, 'read'):
            file.seek(0)
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        
        # Selecionar a planilha "CONTROLE RC E NF"
        try:
            ws = wb['CONTROLE RC E NF']
        except KeyError:
            raise ValidationError("Planilha 'CONTROLE RC E NF' não encontrada no arquivo")
        
        # Ler cabeçalhos da linha 3 (linha 1 está vazia, linha 2 tem título)
        headers = []
        for cell in ws[3]:
            header_value = cell.value if cell.value else None
            if isinstance(header_value, str):
                header_value = header_value.strip().replace('\xa0', ' ').replace('\u00a0', ' ')
                import re
                header_value = re.sub(r'\s+', ' ', header_value).strip()
            headers.append(header_value)
        
        # Função auxiliar para converter valores
        def _safe_str(value, max_length=None):
            if value is None:
                return None
            if isinstance(value, (int, float)):
                value = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
            value = str(value).strip()
            if max_length and len(value) > max_length:
                value = value[:max_length]
            return value if value else None
        
        def _safe_decimal(value):
            if value is None or value == '':
                return None
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            if isinstance(value, str):
                # Remover formatação brasileira (pontos e vírgulas)
                value = value.replace('.', '').replace(',', '.')
                try:
                    return Decimal(value)
                except (InvalidOperation, ValueError):
                    return None
            return None
        
        def _safe_datetime(value):
            if value is None or value == '':
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, date) and not isinstance(value, datetime):
                return datetime.combine(value, datetime.min.time())
            if isinstance(value, str):
                # Tentar diferentes formatos
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        return datetime.strptime(value.strip(), fmt)
                    except ValueError:
                        continue
            return None
        
        # Processar dados em transação
        with transaction.atomic():
            # Ler dados a partir da linha 4
            for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=False), start=4):
                try:
                    # Verificar se a linha está vazia
                    if not any(cell.value for cell in row if cell.value):
                        continue
                    
                    # Criar dicionário com os dados da linha
                    row_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_data[headers[idx]] = cell.value
                    
                    # Obter ID do Excel (identificador único)
                    id_excel = _safe_str(row_data.get('ID'))
                    
                    # Verificar se tem dados mínimos (ID é obrigatório, mas manter compatibilidade com RC/NF Saída)
                    rc = _safe_str(row_data.get('RC'))
                    nf_saida = _safe_str(row_data.get('NF SAÍDA'))
                    
                    # Se não tem ID nem RC nem NF Saída, pular
                    if not id_excel and not rc and not nf_saida:
                        continue
                    
                    # Buscar registro existente (usar ID como identificador único principal)
                    existing = None
                    duplicate_reason = None
                    campos_duplicados = {}
                    found_by_id_excel = False  # Flag para indicar se encontrou por id_excel
                    
                    # Prioridade 1: Usar ID do Excel (chave primária - sempre atualizar se encontrado)
                    # Se id_excel existe, sempre atualizar (é a chave primária, reutilizável após deleção)
                    if id_excel:
                        existing = ControleRCeNF.objects.filter(id_excel=id_excel).first()
                        if existing:
                            found_by_id_excel = True
                            duplicate_reason = "ID já existe no banco de dados"
                            campos_duplicados = {'ID': id_excel}
                            # Verificar se outros campos também coincidem
                            if rc and existing.rc == rc:
                                campos_duplicados['RC'] = rc
                            if nf_saida and existing.nf_saida == nf_saida:
                                campos_duplicados['NF Saída'] = nf_saida
                            pedido_excel = _safe_str(row_data.get('PEDIDO'))
                            if pedido_excel and existing.pedido == pedido_excel:
                                campos_duplicados['Pedido'] = pedido_excel
                    
                    # Fallback: Se não encontrou por ID, tentar por RC (compatibilidade com dados antigos)
                    # IMPORTANTE: Só usar fallback se id_excel NÃO foi fornecido no Excel
                    # Se id_excel foi fornecido, não usar fallback (id_excel é a chave primária)
                    if not existing and not id_excel and rc:
                        existing = ControleRCeNF.objects.filter(rc=rc).first()
                        if existing:
                            duplicate_reason = "RC já existe no banco de dados (sem ID no Excel)"
                            campos_duplicados = {'RC': rc}
                            # Verificar se outros campos também coincidem
                            if nf_saida and existing.nf_saida == nf_saida:
                                campos_duplicados['NF Saída'] = nf_saida
                            pedido_excel = _safe_str(row_data.get('PEDIDO'))
                            if pedido_excel and existing.pedido == pedido_excel:
                                campos_duplicados['Pedido'] = pedido_excel
                    
                    # Fallback: Se não encontrou por ID nem RC, tentar por NF Saída (compatibilidade com dados antigos)
                    # IMPORTANTE: Só usar fallback se id_excel NÃO foi fornecido no Excel
                    if not existing and not id_excel and nf_saida:
                        existing = ControleRCeNF.objects.filter(nf_saida=nf_saida).first()
                        if existing:
                            duplicate_reason = "NF Saída já existe no banco de dados (sem ID no Excel)"
                            campos_duplicados = {'NF Saída': nf_saida}
                    
                    # Se encontrou por id_excel, sempre atualizar (id_excel é chave primária, reutilizável)
                    # Se encontrou por fallback (RC/NF), respeitar a flag update_existing
                    if existing and not found_by_id_excel and not update_existing:
                        # Registrar como duplicata apenas para fallback matching
                        duplicates.append({
                            'linha_excel': row_num,
                            'id_excel': id_excel or 'N/A',
                            'rc': rc or 'N/A',
                            'nf_saida': nf_saida or 'N/A',
                            'empresa': _safe_str(row_data.get('EMPRESA')) or 'N/A',
                            'solicitante': _safe_str(row_data.get('SOLICITANTE')) or 'N/A',
                            'pedido': _safe_str(row_data.get('PEDIDO')) or 'N/A',
                            'motivo': duplicate_reason,
                            'campos_duplicados': campos_duplicados,
                            'registro_existente_id': existing.id_excel,  # id_excel is now the primary key
                            'registro_existente_id_excel': existing.id_excel or 'N/A',
                            'registro_existente_rc': existing.rc or 'N/A',
                            'registro_existente_nf': existing.nf_saida or 'N/A',
                            'registro_existente_empresa': existing.empresa or 'N/A',
                            'registro_existente_pedido': existing.pedido or 'N/A',
                        })
                        continue
                    
                    # Preparar dados
                    # Se não tem ID do Excel, gerar um baseado na linha (id_excel é obrigatório como chave primária)
                    if not id_excel:
                        id_excel = f"IMPORT_{row_num}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    
                    data_dict = {
                        'id_excel': id_excel,  # ID é obrigatório (chave primária)
                        'solicitante': _safe_str(row_data.get('SOLICITANTE')),
                        'empresa': _safe_str(row_data.get('EMPRESA')),
                        'nf_saida': nf_saida,
                        'descricao_servico': _safe_str(row_data.get('DESCRIÇÃO DO SERVIÇO'), max_length=1000),
                        'ca_rateio': _safe_str(row_data.get('C.A/RATEIO')),
                        'uso': _safe_str(row_data.get('USO')),
                        'quem_abriu_rc': _safe_str(row_data.get('QUEM ABRIU \nA RC') or row_data.get('QUEM ABRIU A RC')),
                        'orcamento': _safe_str(row_data.get('ORÇAMENTO'), max_length=500),
                        'os': _safe_str(row_data.get('O.S')),
                        'classificacao': _safe_str(row_data.get('CLASSIF.')),
                        'justificativa_classificacao': _safe_str(row_data.get('JUSTIFICATIVA CLASSIFICAÇÃO'), max_length=1000),
                        'spaf0009_acesso_portaria': _safe_str(row_data.get('SPAF0009 - Acesso portária p/ classif. 5 e 8'), max_length=255),
                        'rc': rc,
                        'data_rc': _safe_datetime(row_data.get('DATA RC')),
                        'pedido': _safe_str(row_data.get('PEDIDO')),
                        'valor_total_pedido': _safe_decimal(row_data.get('VALOR TOTAL DO PEDIDO')),
                        'previsao_para_uso': _safe_datetime(row_data.get('PREVISÃO PARA USO')),
                        'nf_servico': _safe_str(row_data.get('NF SERVIÇO')),
                        'nf_retorno_e_data_lancamento': _safe_str(row_data.get('NF RETORNO E DATA LANÇAMENTO'), max_length=255),
                        'cnpj_aurora': _safe_str(row_data.get('CNPJ AURORA')),
                        'simples_nacional': _safe_str(row_data.get('SIMPLES NACIONAL')),
                        'valor_nf': _safe_decimal(row_data.get('VALOR NF')),
                        'emissao': _safe_datetime(row_data.get('EMISSÃO')),
                        'inclusao_198': _safe_datetime(row_data.get('INCLUSÃO 198')),
                        'status': _safe_str(row_data.get('STATUS'), max_length=255),
                        'obs': _safe_str(row_data.get('OBS'), max_length=1000),
                        'saldo_residual_pedido': _safe_decimal(row_data.get('SALDO RESIDUAL PEDIDO ') or row_data.get('SALDO RESIDUAL PEDIDO')),
                    }
                    
                    if existing:
                        # Atualizar registro existente
                        for key, value in data_dict.items():
                            setattr(existing, key, value)
                        existing.save()
                        updated_count += 1
                    else:
                        # Criar novo registro
                        # Se não tem ID do Excel, gerar um baseado na linha (para compatibilidade)
                        if not data_dict.get('id_excel'):
                            data_dict['id_excel'] = f"IMPORT_{row_num}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        try:
                            ControleRCeNF.objects.create(**data_dict)
                            created_count += 1
                        except Exception as e:
                            # Se falhar por ID duplicado, tentar atualizar
                            if 'id_excel' in str(e).lower() or 'unique' in str(e).lower():
                                existing_by_id = ControleRCeNF.objects.filter(id_excel=data_dict['id_excel']).first()
                                if existing_by_id:
                                    for key, value in data_dict.items():
                                        setattr(existing_by_id, key, value)
                                    existing_by_id.save()
                                    updated_count += 1
                                else:
                                    raise
                            else:
                                raise
                        
                except Exception as e:
                    error_msg = f"Linha {row_num}: {str(e)}"
                    errors.append(error_msg)
                    print(f"Erro na linha {row_num}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
        
        return created_count, updated_count, errors, duplicates
        
    except Exception as e:
        error_detail = f"Erro ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        print(f"Erro geral: {error_detail}")
        import traceback
        traceback.print_exc()
        return 0, 0, errors, []


def upload_paradas_maquina_from_file(file, update_existing=False, excluir_antigos=False, substituir_meses=None) -> Tuple[int, int, List[str]]:
    """
    Faz upload de paradas de máquinas a partir de um arquivo CSV.

    Args:
        file: Arquivo Django UploadedFile (CSV)
        update_existing: Se True, atualiza registros existentes. Se False, ignora duplicados.
        excluir_antigos: Se True, remove todos os registros de ParadaMaquina antes de importar;
            somente os dados do arquivo serão mantidos (evita duplicados ao reimportar arquivo completo).
        substituir_meses: Tupla (ano, [lista_de_meses]) - quando informado, remove APENAS os registros
            cujo campo data pertence ao ano e meses especificados, depois importa. Ex.: (2026, [2]) para
            substituir apenas Fevereiro/2026, deixando Janeiro e demais meses intactos.

    Returns:
        Tupla (created_count, updated_count, errors)
    """
    from app.models import ParadaMaquina
    from datetime import datetime, date, time
    import re
    
    errors = []
    created_count = 0
    updated_count = 0
    
    # Determinar tipo de arquivo
    file_name = file.name.lower()
    
    try:
        # Ler arquivo CSV (usar delimitador ponto e vírgula)
        if not file_name.endswith('.csv'):
            raise ValidationError("Formato de arquivo não suportado. Use .csv")
        
        # Tentar diferentes encodings
        data = None
        encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                file.seek(0)
                data = read_csv_file(file, encoding=encoding, delimiter=';')
                break
            except (UnicodeDecodeError, ValidationError) as e:
                if encoding == encodings_to_try[-1]:
                    raise ValidationError(f"Erro ao ler arquivo CSV: Não foi possível decodificar o arquivo com nenhum encoding testado. Erro original: {str(e)}")
                continue
            except Exception as e:
                raise ValidationError(f"Erro ao ler arquivo CSV: {str(e)}")
        
        if data is None:
            raise ValidationError("Erro ao ler arquivo CSV: Não foi possível processar o arquivo.")
        
        if not data:
            raise ValidationError("Arquivo vazio ou sem dados válidos")
        
        # Helper function para obter valor do row_data
        def _get_value(row_data, *keys):
            """Tenta obter um valor de row_data usando várias chaves, primeiro exato, depois case-insensitive e com strip."""
            for key in keys:
                if key in row_data:
                    return row_data[key]
                normalized_key = key.strip().lower()
                for existing_key, value in row_data.items():
                    if existing_key and existing_key.strip().lower() == normalized_key:
                        return value
            return None
        
        # Helper function para parse de data (DD/MM/YYYY)
        def _parse_date(date_str):
            """Parse data no formato DD/MM/YYYY"""
            if not date_str:
                return None
            date_str = str(date_str).strip()
            if not date_str:
                return None
            formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d']
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None
        
        # Helper function para parse de hora (HH:MM)
        def _parse_time(time_str):
            """Parse hora no formato HH:MM"""
            if not time_str:
                return None
            time_str = str(time_str).strip()
            if not time_str:
                return None
            formats = ['%H:%M', '%H:%M:%S']
            for fmt in formats:
                try:
                    return datetime.strptime(time_str, fmt).time()
                except ValueError:
                    continue
            return None
        
        # Processar dados em transação
        with transaction.atomic():
            if excluir_antigos:
                ParadaMaquina.objects.all().delete()
            elif substituir_meses:
                ano_subst, meses_subst = substituir_meses
                if ano_subst and meses_subst:
                    from django.db.models import Q
                    mes_conditions = Q()
                    for m in meses_subst:
                        if 1 <= m <= 12:
                            mes_conditions |= Q(data__month=m)
                    ParadaMaquina.objects.filter(data__year=ano_subst).filter(mes_conditions).delete()

            for row_num, row_data in enumerate(data, start=2):
                try:
                    # Verificar se a linha está vazia
                    if not row_data or not any(str(v).strip() for v in row_data.values() if v):
                        continue
                    
                    # Parse de data
                    data_parada = _parse_date(_get_value(row_data, 'DATA', 'Data', 'data'))
                    if not data_parada:
                        errors.append(f"Linha {row_num}: Data inválida ou não fornecida")
                        continue
                    
                    # Parse de horários
                    horario_inicial = _parse_time(_get_value(row_data, 'HORARIO INICIAL', 'Horario Inicial', 'horario_inicial'))
                    horario_final = _parse_time(_get_value(row_data, 'HORARIO FINAL', 'Horario Final', 'horario_final'))
                    
                    # Parse de códigos e valores
                    cod_recurso = _safe_int(_get_value(row_data, 'CÓD. RECURSO', 'Cód. Recurso', 'cod_recurso'))
                    
                    # Verificar se já existe (usar data + cod_recurso + horario_inicial como chave única)
                    # Quando excluir_antigos=True, não há registros antigos; sempre criar.
                    existing = None
                    if not excluir_antigos and cod_recurso:
                        existing = ParadaMaquina.objects.filter(
                            data=data_parada,
                            cod_recurso=cod_recurso,
                            horario_inicial=horario_inicial
                        ).first()

                    if existing and not update_existing:
                        continue

                    # Preparar dados
                    parada_data = {
                        'unid': _safe_int(_get_value(row_data, 'UNID', 'Unid', 'unid')),
                        'nome_unidade': _safe_str(_get_value(row_data, 'NOME_UNIDADE', 'Nome Unidade', 'nome_unidade'), max_length=255),
                        'linha_producao': _safe_int(_get_value(row_data, 'LINHA_PRODUCAO', 'Linha Produção', 'linha_producao')),
                        'descr_linha_producao': _safe_str(_get_value(row_data, 'DESCR_LINHA_PRODUCAO', 'Descr Linha Produção', 'descr_linha_producao'), max_length=255),
                        'turno': _safe_str(_get_value(row_data, 'TURNO', 'Turno', 'turno'), max_length=10),
                        'data': data_parada,
                        'cod_item': _safe_str(_get_value(row_data, 'CÓD. ITEM', 'Cód. Item', 'cod_item'), max_length=100),
                        'descr_item': _safe_str(_get_value(row_data, 'DESCR. ITEM', 'Descr. Item', 'descr_item'), max_length=500),
                        'cod_grupo_recurso': _safe_int(_get_value(row_data, 'CÓD. GRUPO DE RECURSO', 'Cód. Grupo de Recurso', 'cod_grupo_recurso')),
                        'grupo_recurso': _safe_str(_get_value(row_data, 'GRUPO RECURSO', 'Grupo Recurso', 'grupo_recurso'), max_length=255),
                        'cod_parada': _safe_str(_get_value(row_data, 'CÓD. PARADA', 'Cód. Parada', 'cod_parada'), max_length=50),
                        'descr_parada': _safe_str(_get_value(row_data, 'DESCR. PARADA', 'Descr. Parada', 'descr_parada'), max_length=255),
                        'nro': _safe_str(_get_value(row_data, 'NRO', 'Nro', 'nro'), max_length=50),
                        'cod_recurso': cod_recurso,
                        'descr_recurso': _safe_str(_get_value(row_data, 'DESCR. RECURSO', 'Descr. Recurso', 'descr_recurso'), max_length=255),
                        'horario_inicial': horario_inicial,
                        'horario_final': horario_final,
                        'dif_hora': _safe_decimal(_get_value(row_data, 'DIF. HORA', 'Dif. Hora', 'dif_hora')),
                        'capacidade': _safe_decimal(_get_value(row_data, 'CAPACIDADE', 'Capacidade', 'capacidade')),
                        'motivo': _safe_str(_get_value(row_data, 'MOTIVO', 'Motivo', 'motivo')),
                        'acao': _safe_str(_get_value(row_data, 'AÇÃO', 'Ação', 'acao')),
                    }
                    
                    # Criar ou atualizar
                    if existing:
                        for key, value in parada_data.items():
                            setattr(existing, key, value)
                        existing.save()
                        updated_count += 1
                    else:
                        ParadaMaquina.objects.create(**parada_data)
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f"Linha {row_num}: Erro ao processar - {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue
        
        return created_count, updated_count, errors
        
    except Exception as e:
        error_detail = f"Erro ao processar arquivo: {str(e)}"
        errors.append(error_detail)
        return 0, 0, errors


def normalize_nota_numero(numero_str):
    """
    Normaliza o número da nota fiscal para comparação.
    Remove prefixos como "NF", "Nº", espaços, e sufixos como " - 1", " - 2", etc.
    
    Exemplos:
    - "NF 05777 - 1" -> "05777"
    - "05777" -> "05777"
    - "NF 12345" -> "12345"
    - "Nº 99999 - 2" -> "99999"
    """
    if not numero_str:
        return None
    
    import re
    # Converter para string e remover espaços extras
    numero_str = str(numero_str).strip().upper()
    
    # Remover prefixos comuns (NF, Nº, N, NOTA, etc.)
    numero_str = re.sub(r'^(NF|Nº|N|NOTA|NOTA\s*FISCAL)\s*', '', numero_str, flags=re.IGNORECASE)
    
    # Remover sufixos como " - 1", " - 2", " - A", etc.
    numero_str = re.sub(r'\s*-\s*\d+$', '', numero_str)
    numero_str = re.sub(r'\s*-\s*[A-Z]$', '', numero_str)
    
    # Remover espaços e caracteres especiais, manter apenas números
    numero_str = re.sub(r'[^\d]', '', numero_str)
    
    return numero_str if numero_str else None


def normalize_fornecedor_nome(nome_str):
    """
    Normaliza o nome do fornecedor para comparação.
    Remove espaços extras, converte para maiúsculas, remove caracteres especiais.
    """
    if not nome_str:
        return None
    
    import re
    # Converter para string, remover espaços extras e converter para maiúsculas
    nome_str = str(nome_str).strip().upper()
    
    # Remover espaços múltiplos
    nome_str = re.sub(r'\s+', ' ', nome_str)
    
    # Remover caracteres especiais comuns (mantém letras, números e espaços)
    nome_str = re.sub(r'[^\w\s]', '', nome_str)
    
    return nome_str if nome_str else None


def match_projecao_nota_fiscal(projecao, nota_fiscal, tolerance_percent=0.01):
    """
    Verifica se uma ProjecaoGasto corresponde a uma NotaFiscal.
    
    Args:
        projecao: Instância de ProjecaoGasto
        nota_fiscal: Instância de NotaFiscal
        tolerance_percent: Tolerância percentual para valores (padrão 1%)
    
    Returns:
        Tuple (bool, dict): (True se match, dict com detalhes do match)
    """
    from decimal import Decimal
    
    match_details = {
        'numero_match': False,
        'fornecedor_match': False,
        'valor_match': False,
        'score': 0.0
    }
    
    # 1. Comparar número da nota
    proj_numero = normalize_nota_numero(projecao.numero_nf) if projecao.numero_nf else None
    nota_numero = normalize_nota_numero(nota_fiscal.nota) if nota_fiscal.nota else None
    
    if proj_numero and nota_numero:
        match_details['numero_match'] = (proj_numero == nota_numero)
    elif not proj_numero and not nota_numero:
        # Se ambos são None, não podemos comparar, mas não é um match negativo
        pass
    else:
        match_details['numero_match'] = False
    
    # 2. Comparar fornecedor
    proj_fornecedor = normalize_fornecedor_nome(projecao.fornecedor_nome_fantasia) if projecao.fornecedor_nome_fantasia else None
    nota_fornecedor = normalize_fornecedor_nome(nota_fiscal.nome_fantasia_emitente) if nota_fiscal.nome_fantasia_emitente else None
    
    if proj_fornecedor and nota_fornecedor:
        # Comparação exata ou parcial (se um contém o outro)
        match_details['fornecedor_match'] = (
            proj_fornecedor == nota_fornecedor or
            proj_fornecedor in nota_fornecedor or
            nota_fornecedor in proj_fornecedor
        )
    elif not proj_fornecedor and not nota_fornecedor:
        pass
    else:
        match_details['fornecedor_match'] = False
    
    # 3. Comparar valor (com tolerância)
    proj_valor = projecao.valor_total if projecao.valor_total else None
    nota_valor = nota_fiscal.total_nota if nota_fiscal.total_nota else None
    
    if proj_valor and nota_valor:
        # Converter para Decimal se necessário
        if not isinstance(proj_valor, Decimal):
            proj_valor = Decimal(str(proj_valor))
        if not isinstance(nota_valor, Decimal):
            nota_valor = Decimal(str(nota_valor))
        
        # Calcular diferença percentual
        if proj_valor > 0:
            diff_percent = abs(proj_valor - nota_valor) / proj_valor
            match_details['valor_match'] = (diff_percent <= tolerance_percent)
        else:
            match_details['valor_match'] = (proj_valor == nota_valor)
    elif not proj_valor and not nota_valor:
        pass
    else:
        match_details['valor_match'] = False
    
    # Calcular score (percentual de match)
    matches = sum([
        match_details['numero_match'],
        match_details['fornecedor_match'],
        match_details['valor_match']
    ])
    match_details['score'] = (matches / 3.0) * 100 if matches > 0 else 0.0
    
    # Considerar match se pelo menos 2 dos 3 critérios correspondem
    # OU se número e valor correspondem (mais importante)
    is_match = (
        matches >= 2 or
        (match_details['numero_match'] and match_details['valor_match'])
    )
    
    return is_match, match_details


def find_matching_notas_fiscais(projecao):
    """
    Encontra todas as Notas Fiscais que correspondem a uma ProjecaoGasto.
    
    Args:
        projecao: Instância de ProjecaoGasto
    
    Returns:
        List[Tuple]: Lista de tuplas (nota_fiscal, match_details)
    """
    from app.models import NotaFiscal
    
    if not projecao:
        return []
    
    # Buscar todas as notas fiscais
    todas_notas = NotaFiscal.objects.all()
    
    matches = []
    for nota in todas_notas:
        is_match, match_details = match_projecao_nota_fiscal(projecao, nota)
        if is_match:
            matches.append((nota, match_details))
    
    # Ordenar por score (melhor match primeiro)
    matches.sort(key=lambda x: x[1]['score'], reverse=True)
    
    return matches

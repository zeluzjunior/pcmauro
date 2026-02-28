from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q
import os


def handle_form_errors(form, request):
    """Helper function to handle form validation errors with improved messages"""
    missing_required = []
    for field, errors in form.errors.items():
        field_label = form.fields[field].label
        for error in errors:
            error_str = str(error).lower()
            if 'required' in error_str or 'obrigatório' in error_str or 'este campo é obrigatório' in error_str:
                if field_label not in missing_required:
                    missing_required.append(field_label)
                messages.warning(request, f'<strong>{field_label}</strong>: Este campo é obrigatório e deve ser preenchido.')
            else:
                messages.error(request, f'<strong>{field_label}</strong>: {error}')
    
    if missing_required:
        messages.warning(request, f'<strong>Atenção:</strong> {len(missing_required)} campo(s) obrigatório(s) não preenchido(s). Por favor, preencha todos os campos marcados com <span class="text-danger">*</span>.')
    elif form.errors:
        messages.error(request, 'Por favor, corrija os erros no formulário antes de continuar.')


def home(request):
    """Home page view - Data filtered by current month"""
    from app.models import OrdemServicoCorretiva, RequisicaoAlmoxarifado, Maquina, Manutentor
    from datetime import datetime, timedelta, date
    from django.db.models import Sum, Count, Q
    from decimal import Decimal
    from calendar import monthrange
    
    hoje = date.today()
    ultimo_dia_mes = monthrange(hoje.year, hoje.month)[1]
    data_inicio_mes = date(hoje.year, hoje.month, 1)
    data_fim_mes = date(hoje.year, hoje.month, ultimo_dia_mes)
    mes_ano_grafico = f"{hoje.year}-{str(hoje.month).zfill(2)}"
    meses_nomes = ('', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro')
    mes_ano_label = f"{meses_nomes[hoje.month]} {hoje.year}"
    
    # ========== KPIs BASEADOS NO MÊS ATUAL ==========
    
    # 1. Manutenções Corretivas no mês atual
    manutencoes_corretivas = 0
    if data_inicio_mes and data_fim_mes:
        try:
            ordens_mes = OrdemServicoCorretiva.objects.exclude(
                dt_entrada__isnull=True
            ).exclude(dt_entrada='')
            for ordem in ordens_mes:
                try:
                    dt_str = ordem.dt_entrada.strip()
                    if ' ' in dt_str:
                        date_part = dt_str.split(' ')[0]
                    else:
                        date_part = dt_str
                    if '/' in date_part:
                        parts = date_part.split('/')
                        if len(parts) == 3:
                            day, month, year = parts
                            ordem_date = date(int(year), int(month), int(day))
                            if data_inicio_mes <= ordem_date <= data_fim_mes:
                                manutencoes_corretivas += 1
                except:
                    continue
        except Exception as e:
            print(f"Erro ao contar manutenções corretivas: {e}")
    
    # 2. Manutenções Preventivas no mês atual
    try:
        from app.models import MeuPlanoPreventiva
        manutencoes_preventivas = MeuPlanoPreventiva.objects.filter(
            data_planejada__gte=data_inicio_mes,
            data_planejada__lte=data_fim_mes
        ).count()
    except:
        manutencoes_preventivas = 0
    
    # 3. Requisições de Almoxarifado no mês atual
    requisicoes_mes = RequisicaoAlmoxarifado.objects.filter(
        data_requisicao__gte=data_inicio_mes,
        data_requisicao__lte=data_fim_mes
    )
    total_requisicoes_semana = requisicoes_mes.count()
    valor_total_semana = requisicoes_mes.aggregate(
        total=Sum('vlr_movto_estoq')
    )['total'] or Decimal('0')
    
    # 4. Máquinas e Manutentores (total geral)
    total_maquinas = Maquina.objects.count()
    total_manutentores = Manutentor.objects.count()
    
    # 5. Calendário - eventos no mês atual (sem ordens de serviço corretiva)
    eventos = []
    try:
        from app.models import MeuPlanoPreventiva
        preventivas_mes = MeuPlanoPreventiva.objects.filter(
            data_planejada__gte=data_inicio_mes,
            data_planejada__lte=data_fim_mes
        )
        for preventiva in preventivas_mes:
            eventos.append({
                'title': f'Preventiva - {preventiva.cd_maquina.cd_maquina if preventiva.cd_maquina else "N/A"}',
                'start': preventiva.data_planejada.strftime('%Y-%m-%d'),
                'color': '#28a745',
                'url': f'/planejamento/meu-plano/?search={preventiva.cd_maquina.cd_maquina if preventiva.cd_maquina else ""}'
            })
    except:
        pass
    try:
        from app.models import ManutencaoTerceiro
        for manutencao in ManutencaoTerceiro.objects.filter(
            data__date__gte=data_inicio_mes,
            data__date__lte=data_fim_mes
        ):
            if manutencao.data:
                eventos.append({
                    'title': f'Manutenção Terceiro - {manutencao.titulo[:30]}',
                    'start': manutencao.data.strftime('%Y-%m-%d'),
                    'color': '#ff9800',
                    'url': '/manutencao-terceiro/consultar/'
                })
    except Exception as e:
        print(f"Erro ao adicionar eventos de Manutenção Terceiro: {e}")
    try:
        from app.models import Visitas
        for visita in Visitas.objects.filter(
            data__date__gte=data_inicio_mes,
            data__date__lte=data_fim_mes
        ):
            if visita.data:
                eventos.append({
                    'title': f'Visita - {visita.titulo[:30]}',
                    'start': visita.data.strftime('%Y-%m-%d'),
                    'color': '#9c27b0',
                    'url': '/visitas/consultar/'
                })
    except Exception as e:
        print(f"Erro ao adicionar eventos de Visitas: {e}")
    
    # ========== GRÁFICO: ORDENS FECHADAS NO MÊS ATUAL ==========
    ordens_fechadas_labels = []
    ordens_fechadas_data = []
    
    def parse_date_from_string(date_str):
        """
        Tenta fazer parse de uma data em vários formatos diferentes.
        Retorna um objeto date ou None se não conseguir fazer parse.
        """
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir (formato: "dd/mm/yyyy hh:mm" ou "dd/mm/yyyy hh:mm:ss")
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y',      # 26/09/2025
            '%d-%m-%Y',      # 26-09-2025
            '%d.%m.%Y',      # 26.09.2025
            '%Y-%m-%d',      # 2025-09-26
            '%Y/%m/%d',      # 2025/09/26
            '%d/%m/%y',      # 26/09/25
            '%d-%m-%y',      # 26-09-25
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt).date()
            except (ValueError, TypeError):
                continue
        
        # Se nenhum formato funcionou, tentar parse manual para formato brasileiro comum
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    # Se ano tem 2 dígitos, assumir 2000+
                    if len(year) == 2:
                        year = '20' + year
                    return date(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    if data_inicio_mes and data_fim_mes:
        from collections import defaultdict
        ordens_por_dia = defaultdict(int)
        ordens_fechadas = OrdemServicoCorretiva.objects.exclude(
            dt_encordmanu__isnull=True
        ).exclude(dt_encordmanu='')
        for ordem in ordens_fechadas:
            try:
                ordem_date = parse_date_from_string(ordem.dt_encordmanu)
                if ordem_date and data_inicio_mes <= ordem_date <= data_fim_mes:
                    data_key = ordem_date.strftime('%Y-%m-%d')
                    ordens_por_dia[data_key] += 1
            except Exception:
                continue
        current_date = data_inicio_mes
        while current_date <= data_fim_mes:
            data_key = current_date.strftime('%Y-%m-%d')
            ordens_fechadas_labels.append(current_date.strftime('%d/%m'))
            ordens_fechadas_data.append(ordens_por_dia.get(data_key, 0))
            current_date += timedelta(days=1)
    
    import json
    ordens_fechadas_labels_json = json.dumps(ordens_fechadas_labels)
    ordens_fechadas_data_json = json.dumps(ordens_fechadas_data)
    
    # ========== ORÇAMENTO, GASTOS E REQUISIÇÕES (ano atual, todos os meses) ==========
    from app.models import DadosOrcamento, ProjecaoGasto, NotaFiscal, RequisicaoAlmoxarifado
    from django.db.models import Sum, Q
    ano_orcamento = hoje.year
    total_orcamento_disponivel = DadosOrcamento.objects.filter(
        ano=ano_orcamento
    ).aggregate(total=Sum('valor_orcamento'))['total'] or Decimal('0')
    q_ano_proj = (
        Q(ano_referencia=ano_orcamento) |
        (Q(ano_referencia__isnull=True) & Q(created_at__year=ano_orcamento))
    )
    projecoes_ano = ProjecaoGasto.objects.filter(q_ano_proj)
    valor_servico_concluido_nao = projecoes_ano.filter(
        servico_concluido__iexact='NÃO'
    ).aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    requisicoes_ano = RequisicaoAlmoxarifado.objects.filter(
        data_requisicao__year=ano_orcamento
    )
    mensagem_nf_requisicao = "REQUISIÇÃO GERADA PELO PROCESSO DE ENTRADA DE NOTA FISCAL"
    requisicoes_sem_nf = requisicoes_ano.exclude(obs_rm__icontains=mensagem_nf_requisicao)
    valor_total_requisicoes_sem_nf = Decimal('0.00')
    for req in requisicoes_sem_nf.filter(cd_depo=1):
        if req.vlr_movto_estoq:
            valor_total_requisicoes_sem_nf += abs(req.vlr_movto_estoq)
    todas_notas_lancadas = NotaFiscal.objects.filter(
        Q(situacao__icontains='LANÇADA') | Q(situacao__icontains='LANCADA')
    ).filter(uso_contabil='242')
    def _parse_date_emissao(date_str):
        if not date_str:
            return None
        date_str = str(date_str).strip()
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d'):
            try:
                return datetime.strptime(date_str, fmt)
            except (ValueError, TypeError):
                continue
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                try:
                    d, m, y = parts
                    if len(y) == 2:
                        y = '20' + y
                    return datetime(int(y), int(m), int(d))
                except (ValueError, TypeError):
                    pass
        return None
    valor_total_notas_lancadas = Decimal('0')
    for nota in todas_notas_lancadas:
        de = _parse_date_emissao(nota.data_emissao)
        if de and de.year == ano_orcamento:
            valor_total_notas_lancadas += (nota.total_nota or Decimal('0'))
    saldo_parcial = total_orcamento_disponivel - valor_total_requisicoes_sem_nf - valor_servico_concluido_nao - valor_total_notas_lancadas
    
    context = {
        'page_title': 'Home',
        'active_page': 'home',
        'eventos': eventos,
        'data_inicio_mes': data_inicio_mes,
        'data_fim_mes': data_fim_mes,
        'mes_ano_label': mes_ano_label,
        'mes_ano_grafico': mes_ano_grafico,
        # KPIs
        'manutencoes_corretivas': manutencoes_corretivas,
        'manutencoes_preventivas': manutencoes_preventivas,
        'total_requisicoes_semana': total_requisicoes_semana,
        'valor_total_semana': abs(valor_total_semana),  # Usar valor absoluto
        'total_maquinas': total_maquinas,
        'total_manutentores': total_manutentores,
        # Dados do gráfico de ordens fechadas
        'ordens_fechadas_labels': ordens_fechadas_labels_json,
        'ordens_fechadas_data': ordens_fechadas_data_json,
        # Orçamento, Gastos e Requisições (ano atual)
        'total_orcamento_disponivel': total_orcamento_disponivel,
        'valor_total_requisicoes_sem_nf': valor_total_requisicoes_sem_nf,
        'valor_servico_concluido_nao': valor_servico_concluido_nao,
        'valor_total_notas_lancadas': valor_total_notas_lancadas,
        'saldo_parcial': saldo_parcial,
    }
    return render(request, 'home.html', context)


def centros_de_atividade(request):
    """Centros de Atividade listing page view - filtered by FRIGORÍFICO, INDÚSTRIA, UTILIDADES, EXTERNA, and APOIO"""
    from app.models import CentroAtividade
    from django.db.models import Q
    
    # Buscar Centros de Atividade filtrados por local
    centros_frigorifico = CentroAtividade.objects.filter(
        Q(local__iexact='FRIGORÍFICO') | Q(local__icontains='FRIGOR')
    ).distinct().order_by('ca')
    
    centros_industria = CentroAtividade.objects.filter(
        Q(local__iexact='INDÚSTRIA') | Q(local__icontains='IND')
    ).distinct().order_by('ca')
    
    centros_utilidades = CentroAtividade.objects.filter(
        Q(local__iexact='UTILIDADES') | Q(local__icontains='UTILIDADE')
    ).distinct().order_by('ca')
    
    centros_externa = CentroAtividade.objects.filter(
        Q(local__iexact='EXTERNA') | Q(local__icontains='EXTERN')
    ).distinct().order_by('ca')
    
    centros_apoio = CentroAtividade.objects.filter(
        Q(local__iexact='APOIO') | Q(local__icontains='APOIO')
    ).distinct().order_by('ca')
    
    total_frigorifico = centros_frigorifico.count()
    total_industria = centros_industria.count()
    total_utilidades = centros_utilidades.count()
    total_externa = centros_externa.count()
    total_apoio = centros_apoio.count()
    total_geral = total_frigorifico + total_industria + total_utilidades + total_externa + total_apoio
    
    context = {
        'page_title': 'Centros de Atividade',
        'active_page': 'centros_de_atividade',
        'centros_frigorifico': centros_frigorifico,
        'centros_industria': centros_industria,
        'centros_utilidades': centros_utilidades,
        'centros_externa': centros_externa,
        'centros_apoio': centros_apoio,
        'total_frigorifico': total_frigorifico,
        'total_industria': total_industria,
        'total_utilidades': total_utilidades,
        'total_externa': total_externa,
        'total_apoio': total_apoio,
        'total_geral': total_geral,
    }
    return render(request, 'centro_de_atividades/analise_centro_de_atividade.html', context)


def about(request):
    """About page view"""
    context = {
        'page_title': 'Sobre',
        'active_page': 'about'
    }
    return render(request, 'about.html', context)


def em_desenvolvimento(request):
    """Página em desenvolvimento"""
    context = {
        'page_title': 'Página em Desenvolvimento',
        'active_page': 'em_desenvolvimento'
    }
    return render(request, 'em_desenvolvimento.html', context)


def dados_producao_diaria(request):
    """Página de Dados de Produção Diária: accordion por mês com Suínos Abatidos e Produção Indústria por dia.
    Relaciona cada dia com o nome do dia da semana (PT-BR) e com a Semana52."""
    from app.models import ProducaoDiaria, Semana52
    from calendar import monthrange, weekday
    from datetime import date, timedelta
    from decimal import Decimal, InvalidOperation

    # Nomes dos dias da semana em português (segunda=0 ... domingo=6)
    DIAS_SEMANA_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
    DIAS_SEMANA_PT_SHORT = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']

    meses = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]
    # Ano: em POST usar form hidden; fallback para GET (ex.: ?ano=2026 na action URL)
    if request.method == 'POST':
        ano_atual = request.POST.get('ano') or request.GET.get('ano')
    else:
        ano_atual = request.GET.get('ano')
    try:
        ano_atual = int(ano_atual) if ano_atual else None
    except (ValueError, TypeError):
        ano_atual = None
    if not ano_atual:
        ano_atual = date.today().year

    def _normalize_br_number(s):
        """Converte número no padrão BR (1.234,56) para formato Python (1234.56)."""
        s = str(s).strip()
        if not s:
            return ''
        # Padrão BR: ponto = milhar, vírgula = decimal
        s = s.replace('.', '').replace(',', '.')
        return s

    def _parse_int(val):
        if val is None or str(val).strip() == '':
            return None
        s = _normalize_br_number(val)
        if not s:
            return None
        try:
            return int(Decimal(s))
        except (ValueError, TypeError, InvalidOperation):
            return None

    def _parse_decimal(val):
        if val is None or str(val).strip() == '':
            return None
        s = _normalize_br_number(val)
        if not s:
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError, TypeError):
            return None

    if request.method == 'POST':
        print(f"[producao_diaria] POST recebido ano={ano_atual} campos={len(request.POST)}")
        for mes_num, _ in meses:
            _, ultimo_dia = monthrange(ano_atual, mes_num)
            for dia in range(1, ultimo_dia + 1):
                suinos = _parse_int(request.POST.get(f'mes_{mes_num}_dia_{dia}_suinos_abatidos'))
                producao = _parse_decimal(request.POST.get(f'mes_{mes_num}_dia_{dia}_producao_industria'))
                ProducaoDiaria.objects.update_or_create(
                    ano=ano_atual,
                    mes=mes_num,
                    dia=dia,
                    defaults={
                        'suinos_abatidos': suinos,
                        'producao_industria': producao,
                    }
                )
        messages.success(request, 'Dados de produção diária salvos com sucesso.')
        from django.urls import reverse
        return redirect(reverse('dados_producao_diaria') + f'?ano={ano_atual}')

    # GET: carregar dados do ano e enriquecer com dia da semana (PT-BR) e Semana52
    registros = {
        (r.mes, r.dia): r
        for r in ProducaoDiaria.objects.filter(ano=ano_atual)
    }
    # Cache de Semana52 por data para evitar N queries
    _semana52_por_data = {}

    def _semana52_para_data(d):
        if d not in _semana52_por_data:
            s = Semana52.objects.filter(inicio__lte=d, fim__gte=d).first()
            _semana52_por_data[d] = (s.semana if s else None)
        return _semana52_por_data[d]

    config_por_mes = []
    for mes_num, mes_nome in meses:
        _, ultimo_dia = monthrange(ano_atual, mes_num)
        dias_dados = []
        for dia in range(1, ultimo_dia + 1):
            d = date(ano_atual, mes_num, dia)
            wd = weekday(ano_atual, mes_num, dia)  # 0=Segunda, 6=Domingo
            dia_nome = DIAS_SEMANA_PT_SHORT[wd]
            dia_nome_long = DIAS_SEMANA_PT[wd]
            semana52_label = _semana52_para_data(d)
            obj = registros.get((mes_num, dia))
            dias_dados.append((dia, obj, dia_nome, dia_nome_long, semana52_label, wd))

        # Layout por semanas civis (Seg-Dom). Cada linha = uma semana de calendário (segunda a domingo).
        empty_slot = (0, None, '', '', None, -1)
        semanas_dict = {}  # monday_date -> list of (dia, obj, ...) for that week
        for dia, obj, dia_nome, dia_nome_long, semana52_label, wd in dias_dados:
            d = date(ano_atual, mes_num, dia)
            monday_of_week = d - timedelta(days=wd)
            if monday_of_week not in semanas_dict:
                semanas_dict[monday_of_week] = [None] * 7
            semanas_dict[monday_of_week][wd] = (dia, obj, dia_nome, dia_nome_long, semana52_label, wd)

        semanas_list = []
        for semana_num, (monday_date, ordered_raw) in enumerate(sorted(semanas_dict.items()), 1):
            ordered = [item if item is not None else empty_slot for item in ordered_raw]
            dias_presentes = [item[0] for item in ordered if item[0]]
            start_dia = min(dias_presentes) if dias_presentes else 0
            end_dia = max(dias_presentes) if dias_presentes else 0
            week_label = f"Semana {semana_num} (dias {start_dia}-{end_dia})" if dias_presentes else f"Semana {semana_num}"
            semanas_list.append((week_label, ordered))

        config_por_mes.append((mes_num, mes_nome, ultimo_dia, dias_dados, semanas_list))

    mes_atual = date.today().month
    context = {
        'page_title': 'Dados de Produção Diária',
        'active_page': 'dados_producao_diaria',
        'ano_atual': ano_atual,
        'mes_atual': mes_atual,
        'meses': meses,
        'config_por_mes': config_por_mes,
    }
    return render(request, 'producao_diaria/dados_producao_diaria.html', context)


def consultar_producao_diaria(request):
    """Página Consultar Produção Diária"""
    context = {
        'page_title': 'Consultar Produção Diária',
        'active_page': 'consultar_producao_diaria',
    }
    return render(request, 'producao_diaria/consultar_producao_diaria.html', context)


def analise_calibracao(request):
    """Análise de Calibrações - Página em desenvolvimento"""
    context = {
        'page_title': 'Análise de Calibrações',
        'active_page': 'analise_calibracao'
    }
    return render(request, 'calibracao/analise_calibracao.html', context)


def consultar_calibracoes(request):
    """Consultar Ordens de Calibração - Página em desenvolvimento"""
    context = {
        'page_title': 'Consultar Ordens de Calibração',
        'active_page': 'consultar_calibracoes'
    }
    return render(request, 'calibracao/consultar_calibracoes.html', context)


def analise_requisicoes(request):
    """Análise de requisições de almoxarifado"""
    from app.models import RequisicaoAlmoxarifado
    from decimal import Decimal
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Q, Avg, Max, Case, When, F, Value
    from django.db.models.fields import DecimalField
    from collections import defaultdict
    import json
    from calendar import monthrange
    
    # Obter anos e meses disponíveis no banco de dados
    anos_disponiveis = RequisicaoAlmoxarifado.objects.values_list('data_requisicao__year', flat=True).distinct().order_by('-data_requisicao__year')
    meses_disponiveis = {}
    for ano in anos_disponiveis:
        meses = RequisicaoAlmoxarifado.objects.filter(data_requisicao__year=ano).values_list('data_requisicao__month', flat=True).distinct().order_by('data_requisicao__month')
        meses_disponiveis[ano] = list(meses)
    
    # Obter filtros de ano e meses (múltiplos) - similar ao analise_corretiva_outros
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e mês atual (carregar com filtro aplicado ao mês atual)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar; se não informado, usar mês atual
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    else:
        # Sem mês na URL: aplicar filtro ao mês atual
        meses_filtro_int = [hoje.month]
    
    # Construir queryset base com filtros
    queryset_base = RequisicaoAlmoxarifado.objects.all()
    
    # Aplicar filtro de ano
    if ano_filtro:
        queryset_base = queryset_base.filter(data_requisicao__year=ano_filtro)
        
        # Aplicar filtro de meses (múltiplos)
        if meses_filtro_int:
            # Filtrar por múltiplos meses usando Q objects
            from django.db.models import Q
            mes_conditions = Q()
            for mes in meses_filtro_int:
                mes_conditions |= Q(data_requisicao__month=mes)
            queryset_base = queryset_base.filter(mes_conditions)
    
    # Separar requisições geradas por NF e requisições normais
    mensagem_nf = "REQUISIÇÃO GERADA PELO PROCESSO DE ENTRADA DE NOTA FISCAL"
    queryset_nf = queryset_base.filter(obs_rm__icontains=mensagem_nf)
    queryset_normal = queryset_base.exclude(obs_rm__icontains=mensagem_nf)
    
    # Debug: verificar quantos registros após filtro
    # total_antes = RequisicaoAlmoxarifado.objects.count()
    # total_depois = queryset_base.count()
    # print(f"DEBUG - Total antes: {total_antes}, Total depois: {total_depois}")
    
    # Estatísticas gerais (usando queryset filtrado)
    hoje = datetime.now().date()
    total_requisicoes = queryset_base.count()
    total_requisicoes_nf = queryset_nf.count()
    total_requisicoes_normal = queryset_normal.count()
    
    # Últimos 30 dias (apenas se não houver filtros)
    if not ano_filtro:
        data_30_dias_atras = hoje - timedelta(days=30)
        requisicoes_recentes = queryset_base.filter(
            data_requisicao__gte=data_30_dias_atras
        ).count()
        requisicoes_recentes_nf = queryset_nf.filter(
            data_requisicao__gte=data_30_dias_atras
        ).count()
        requisicoes_recentes_normal = queryset_normal.filter(
            data_requisicao__gte=data_30_dias_atras
        ).count()
    else:
        # Se há filtros, mostrar total filtrado
        requisicoes_recentes = total_requisicoes
        requisicoes_recentes_nf = total_requisicoes_nf
        requisicoes_recentes_normal = total_requisicoes_normal
    
    # Mês atual (apenas se não houver filtros)
    if not ano_filtro:
        primeiro_dia_mes = hoje.replace(day=1)
        requisicoes_mes_atual = queryset_base.filter(
            data_requisicao__gte=primeiro_dia_mes
        ).count()
        requisicoes_mes_atual_nf = queryset_nf.filter(
            data_requisicao__gte=primeiro_dia_mes
        ).count()
        requisicoes_mes_atual_normal = queryset_normal.filter(
            data_requisicao__gte=primeiro_dia_mes
        ).count()
    else:
        # Se há filtros, mostrar total filtrado
        requisicoes_mes_atual = total_requisicoes
        requisicoes_mes_atual_nf = total_requisicoes_nf
        requisicoes_mes_atual_normal = total_requisicoes_normal
    
    # Itens únicos
    itens_unicos = queryset_base.values('cd_item').distinct().count()
    itens_unicos_nf = queryset_nf.values('cd_item').distinct().count()
    itens_unicos_normal = queryset_normal.values('cd_item').distinct().count()
    
    # Centros de atividade únicos
    centros_unicos = queryset_base.exclude(
        cd_centro_ativ__isnull=True
    ).values('cd_centro_ativ').distinct().count()
    centros_unicos_nf = queryset_nf.exclude(
        cd_centro_ativ__isnull=True
    ).values('cd_centro_ativ').distinct().count()
    centros_unicos_normal = queryset_normal.exclude(
        cd_centro_ativ__isnull=True
    ).values('cd_centro_ativ').distinct().count()
    
    # Calcular valor total (vlr_movto_estoq já é o valor total da linha, não precisa multiplicar por quantidade)
    # IMPORTANTE: Apenas considerar cd_depo == 1 para custos (itens novos que geram gasto)
    # cd_depo == 3 são itens reutilizados que não geram custo
    valor_total = Decimal('0.00')
    valor_total_reused = Decimal('0.00')  # Valor dos itens reutilizados (cd_depo == 3)
    quantidade_total = Decimal('0.00')
    
    valor_total_nf = Decimal('0.00')
    valor_total_reused_nf = Decimal('0.00')
    quantidade_total_nf = Decimal('0.00')
    
    valor_total_normal = Decimal('0.00')
    valor_total_reused_normal = Decimal('0.00')
    quantidade_total_normal = Decimal('0.00')
    
    for req in queryset_base:
        if req.vlr_movto_estoq:
            # vlr_movto_estoq já representa o valor total da transação (pode ser negativo para saídas)
            if req.cd_depo == 1:
                # Apenas cd_depo == 1 gera custo
                valor_total += abs(req.vlr_movto_estoq)
                if mensagem_nf in (req.obs_rm or ''):
                    valor_total_nf += abs(req.vlr_movto_estoq)
                else:
                    valor_total_normal += abs(req.vlr_movto_estoq)
            elif req.cd_depo == 3:
                # cd_depo == 3 são itens reutilizados (não geram custo, mas vamos rastrear)
                valor_total_reused += abs(req.vlr_movto_estoq)
                if mensagem_nf in (req.obs_rm or ''):
                    valor_total_reused_nf += abs(req.vlr_movto_estoq)
                else:
                    valor_total_reused_normal += abs(req.vlr_movto_estoq)
        if req.qtde_movto_estoq:
            quantidade_total += abs(req.qtde_movto_estoq)
            if mensagem_nf in (req.obs_rm or ''):
                quantidade_total_nf += abs(req.qtde_movto_estoq)
            else:
                quantidade_total_normal += abs(req.qtde_movto_estoq)
    
    # Valor médio por requisição
    valor_medio = valor_total / total_requisicoes if total_requisicoes > 0 else Decimal('0.00')
    valor_medio_nf = valor_total_nf / total_requisicoes_nf if total_requisicoes_nf > 0 else Decimal('0.00')
    valor_medio_normal = valor_total_normal / total_requisicoes_normal if total_requisicoes_normal > 0 else Decimal('0.00')
    
    # Evolução temporal (últimos 12 meses ou período filtrado)
    meses_labels = []
    meses_data = []
    meses_valor = []
    meses_data_nf = []
    meses_valor_nf = []
    meses_data_normal = []
    meses_valor_normal = []
    
    # Determinar período para evolução temporal
    if ano_filtro:
        try:
            periodo_inicio = datetime(ano_filtro, 1, 1).date()
            if meses_filtro_int:
                # Se há meses selecionados, usar apenas o primeiro e último mês selecionado
                primeiro_mes = min(meses_filtro_int)
                ultimo_mes = max(meses_filtro_int)
                periodo_inicio = datetime(ano_filtro, primeiro_mes, 1).date()
                ultimo_dia = monthrange(ano_filtro, ultimo_mes)[1]
                periodo_fim = datetime(ano_filtro, ultimo_mes, ultimo_dia).date()
                if periodo_fim > hoje:
                    periodo_fim = hoje
            else:
                # Todos os meses do ano
                periodo_fim = datetime(ano_filtro, 12, 31).date()
                if periodo_fim > hoje:
                    periodo_fim = hoje
        except (ValueError, TypeError):
            periodo_inicio = (hoje - timedelta(days=365)).replace(day=1)
            periodo_fim = hoje
    else:
        periodo_inicio = (hoje - timedelta(days=365)).replace(day=1)
        periodo_fim = hoje
    
    # Gerar meses do período
    data_atual = periodo_inicio.replace(day=1)
    while data_atual <= periodo_fim:
        # Calcular último dia do mês
        ultimo_dia_mes = monthrange(data_atual.year, data_atual.month)[1]
        fim_mes_calc = datetime(data_atual.year, data_atual.month, ultimo_dia_mes).date()
        fim_mes = fim_mes_calc if fim_mes_calc <= periodo_fim else periodo_fim
        
        count = queryset_base.filter(
            data_requisicao__gte=data_atual,
            data_requisicao__lte=fim_mes
        ).count()
        
        count_nf = queryset_nf.filter(
            data_requisicao__gte=data_atual,
            data_requisicao__lte=fim_mes
        ).count()
        
        count_normal = queryset_normal.filter(
            data_requisicao__gte=data_atual,
            data_requisicao__lte=fim_mes
        ).count()
        
        valor_mes = Decimal('0.00')
        valor_mes_nf = Decimal('0.00')
        valor_mes_normal = Decimal('0.00')
        
        for req in queryset_base.filter(
            data_requisicao__gte=data_atual,
            data_requisicao__lte=fim_mes
        ):
            if req.vlr_movto_estoq and req.cd_depo == 1:
                # Apenas considerar cd_depo == 1 para custos
                valor_mes += abs(req.vlr_movto_estoq)
                if mensagem_nf in (req.obs_rm or ''):
                    valor_mes_nf += abs(req.vlr_movto_estoq)
                else:
                    valor_mes_normal += abs(req.vlr_movto_estoq)
        
        meses_labels.append(data_atual.strftime('%b/%Y'))
        meses_data.append(count)
        meses_valor.append(float(valor_mes))
        meses_data_nf.append(count_nf)
        meses_valor_nf.append(float(valor_mes_nf))
        meses_data_normal.append(count_normal)
        meses_valor_normal.append(float(valor_mes_normal))
        
        # Próximo mês
        if data_atual.month == 12:
            data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=1)
        else:
            data_atual = data_atual.replace(month=data_atual.month + 1, day=1)
    
    # Quantidade "requisitada" = saídas do estoque: no sistema de origem valor negativo = saída.
    # Interpretamos como quantidade retirada: somamos (-qtde_movto_estoq) quando qtde_movto_estoq < 0.
    qtd_saida_expr = Sum(
        Case(
            When(qtde_movto_estoq__lt=0, then=-F('qtde_movto_estoq')),
            default=Value(0),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        )
    )
    
    # Top 10 itens mais requisitados (por quantidade) - Geral (agrupar só por cd_item para não fragmentar)
    top_itens_qtd = queryset_base.exclude(
        qtde_movto_estoq__isnull=True
    ).values('cd_item').annotate(
        total_qtd=qtd_saida_expr,
        descr_item=Max('descr_item')
    ).order_by('-total_qtd')[:10]
    
    top_itens_labels = []
    top_itens_data = []
    for item in top_itens_qtd:
        descr = (item.get('descr_item') or '').strip() or f"Item {item['cd_item']}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_labels.append(f"{item['cd_item']} - {descr}")
        top_itens_data.append(float(item['total_qtd'] or 0))
    
    # Top 10 itens mais requisitados (por quantidade) - NF (agrupar só por cd_item)
    top_itens_qtd_nf = queryset_nf.exclude(
        qtde_movto_estoq__isnull=True
    ).values('cd_item').annotate(
        total_qtd=qtd_saida_expr,
        descr_item=Max('descr_item')
    ).order_by('-total_qtd')[:10]
    
    top_itens_labels_nf = []
    top_itens_data_nf = []
    for item in top_itens_qtd_nf:
        descr = (item.get('descr_item') or '').strip() or f"Item {item['cd_item']}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_labels_nf.append(f"{item['cd_item']} - {descr}")
        top_itens_data_nf.append(float(item['total_qtd'] or 0))
    
    # Top 10 itens mais requisitados (por quantidade) - Normal (agrupar só por cd_item)
    top_itens_qtd_normal = queryset_normal.exclude(
        qtde_movto_estoq__isnull=True
    ).values('cd_item').annotate(
        total_qtd=qtd_saida_expr,
        descr_item=Max('descr_item')
    ).order_by('-total_qtd')[:10]
    
    top_itens_labels_normal = []
    top_itens_data_normal = []
    for item in top_itens_qtd_normal:
        descr = (item.get('descr_item') or '').strip() or f"Item {item['cd_item']}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_labels_normal.append(f"{item['cd_item']} - {descr}")
        top_itens_data_normal.append(float(item['total_qtd'] or 0))
    
    # Top 10 itens por valor - Geral
    top_itens_valor = []
    itens_valor_dict = defaultdict(lambda: Decimal('0.00'))
    itens_valor_dict_nf = defaultdict(lambda: Decimal('0.00'))
    itens_valor_dict_normal = defaultdict(lambda: Decimal('0.00'))
    
    for req in queryset_base.exclude(vlr_movto_estoq__isnull=True):
        if req.vlr_movto_estoq and req.cd_depo == 1:
            # Apenas considerar cd_depo == 1 para custos
            # vlr_movto_estoq já representa o valor total da transação
            itens_valor_dict[req.cd_item] += abs(req.vlr_movto_estoq)
            if mensagem_nf in (req.obs_rm or ''):
                itens_valor_dict_nf[req.cd_item] += abs(req.vlr_movto_estoq)
            else:
                itens_valor_dict_normal[req.cd_item] += abs(req.vlr_movto_estoq)
    
    # Ordenar e pegar top 10 - Geral
    sorted_itens = sorted(itens_valor_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_itens_valor_labels = []
    top_itens_valor_data = []
    for cd_item, valor in sorted_itens:
        req = queryset_base.filter(cd_item=cd_item).first()
        descr = req.descr_item if req and req.descr_item else f"Item {cd_item}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_valor_labels.append(f"{cd_item} - {descr}")
        top_itens_valor_data.append(float(valor))
    
    # Ordenar e pegar top 10 - NF
    sorted_itens_nf = sorted(itens_valor_dict_nf.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_itens_valor_labels_nf = []
    top_itens_valor_data_nf = []
    for cd_item, valor in sorted_itens_nf:
        req = queryset_nf.filter(cd_item=cd_item).first()
        descr = req.descr_item if req and req.descr_item else f"Item {cd_item}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_valor_labels_nf.append(f"{cd_item} - {descr}")
        top_itens_valor_data_nf.append(float(valor))
    
    # Ordenar e pegar top 10 - Normal
    sorted_itens_normal = sorted(itens_valor_dict_normal.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_itens_valor_labels_normal = []
    top_itens_valor_data_normal = []
    for cd_item, valor in sorted_itens_normal:
        req = queryset_normal.filter(cd_item=cd_item).first()
        descr = req.descr_item if req and req.descr_item else f"Item {cd_item}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_itens_valor_labels_normal.append(f"{cd_item} - {descr}")
        top_itens_valor_data_normal.append(float(valor))
    
    # Distribuição por centro de atividade (top 10)
    centros_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0.00')})
    
    for req in queryset_base.exclude(cd_centro_ativ__isnull=True):
        centros_dict[req.cd_centro_ativ]['count'] += 1
        if req.vlr_movto_estoq and req.cd_depo == 1:
            # Apenas considerar cd_depo == 1 para custos
            # vlr_movto_estoq já representa o valor total da transação
            centros_dict[req.cd_centro_ativ]['valor'] += abs(req.vlr_movto_estoq)
    
    sorted_centros = sorted(centros_dict.items(), key=lambda x: x[1]['valor'], reverse=True)[:10]
    
    centros_labels = []
    centros_data_count = []
    centros_data_valor = []
    for centro_id, dados in sorted_centros:
        centros_labels.append(str(centro_id))
        centros_data_count.append(dados['count'])
        centros_data_valor.append(float(dados['valor']))
    
    # Distribuição por operação (top 10)
    operacoes_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0.00')})
    
    for req in queryset_base.exclude(descr_operacao__isnull=True).exclude(descr_operacao=''):
        operacoes_dict[req.descr_operacao]['count'] += 1
        if req.vlr_movto_estoq and req.cd_depo == 1:
            # Apenas considerar cd_depo == 1 para custos
            # vlr_movto_estoq já representa o valor total da transação
            operacoes_dict[req.descr_operacao]['valor'] += abs(req.vlr_movto_estoq)
    
    sorted_operacoes = sorted(operacoes_dict.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
    
    operacoes_labels = []
    operacoes_data = []
    for operacao, dados in sorted_operacoes:
        if len(operacao) > 30:
            operacao = operacao[:27] + "..."
        operacoes_labels.append(operacao)
        operacoes_data.append(dados['count'])
    
    # Requisições recentes (últimas 20)
    requisicoes_recentes_list = queryset_base.order_by('-data_requisicao', '-created_at')[:20]
    
    # Requisições associadas a NF — lista paginada para a tabela na análise
    requisicoes_nf_ordered = queryset_nf.order_by('-data_requisicao', 'cd_item')
    paginator_nf = Paginator(requisicoes_nf_ordered, 50)
    page_nf = request.GET.get('page_nf', 1)
    try:
        requisicoes_nf_page = paginator_nf.page(int(page_nf))
    except (ValueError, TypeError):
        requisicoes_nf_page = paginator_nf.page(1)
    
    # Top 10 usuários que criaram requisições
    usuarios_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0.00')})
    
    for req in queryset_base.exclude(cd_usu_criou__isnull=True).exclude(cd_usu_criou=''):
        usuarios_dict[req.cd_usu_criou]['count'] += 1
        if req.vlr_movto_estoq and req.cd_depo == 1:
            # Apenas considerar cd_depo == 1 para custos
            usuarios_dict[req.cd_usu_criou]['valor'] += abs(req.vlr_movto_estoq)
    
    sorted_usuarios = sorted(usuarios_dict.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
    
    usuarios_labels = []
    usuarios_data_count = []
    usuarios_data_valor = []
    for usuario, dados in sorted_usuarios:
        usuarios_labels.append(str(usuario) if usuario else 'Não informado')
        usuarios_data_count.append(dados['count'])
        usuarios_data_valor.append(float(dados['valor']))
    
    # Dados diários para o mês selecionado (para o gráfico de evolução diária)
    if ano_filtro and meses_filtro_int:
        try:
            # Usar o primeiro mês selecionado para o gráfico diário
            mes = meses_filtro_int[0]
            primeiro_dia_mes_atual = datetime(ano_filtro, mes, 1).date()
            ultimo_dia_mes_atual = datetime(ano_filtro, mes, monthrange(ano_filtro, mes)[1]).date()
            if ultimo_dia_mes_atual > hoje:
                ultimo_dia_mes_atual = hoje
        except (ValueError, TypeError):
            primeiro_dia_mes_atual = hoje.replace(day=1)
            if hoje.month == 12:
                ultimo_dia_mes_atual = hoje.replace(year=hoje.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                ultimo_dia_mes_atual = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)
            if ultimo_dia_mes_atual > hoje:
                ultimo_dia_mes_atual = hoje
    else:
        primeiro_dia_mes_atual = hoje.replace(day=1)
        if hoje.month == 12:
            ultimo_dia_mes_atual = hoje.replace(year=hoje.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            ultimo_dia_mes_atual = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)
        if ultimo_dia_mes_atual > hoje:
            ultimo_dia_mes_atual = hoje
    
    dias_labels = []
    dias_data = []
    dias_valor = []
    dias_valor_reused = []  # Valores dos itens reutilizados (cd_depo == 3)
    dias_data_nf = []
    dias_valor_nf = []
    dias_data_normal = []
    dias_valor_normal = []
    
    for dia in range(1, ultimo_dia_mes_atual.day + 1):
        data_dia = primeiro_dia_mes_atual.replace(day=dia)
        count = queryset_base.filter(data_requisicao=data_dia).count()
        count_nf = queryset_nf.filter(data_requisicao=data_dia).count()
        count_normal = queryset_normal.filter(data_requisicao=data_dia).count()
        
        valor_dia = Decimal('0.00')
        valor_dia_reused = Decimal('0.00')
        valor_dia_nf = Decimal('0.00')
        valor_dia_normal = Decimal('0.00')
        
        for req in queryset_base.filter(data_requisicao=data_dia):
            if req.vlr_movto_estoq:
                # vlr_movto_estoq já representa o valor total da transação
                if req.cd_depo == 1:
                    # Apenas cd_depo == 1 gera custo
                    valor_dia += abs(req.vlr_movto_estoq)
                    if mensagem_nf in (req.obs_rm or ''):
                        valor_dia_nf += abs(req.vlr_movto_estoq)
                    else:
                        valor_dia_normal += abs(req.vlr_movto_estoq)
                elif req.cd_depo == 3:
                    # cd_depo == 3 são itens reutilizados (não geram custo, mas vamos rastrear)
                    valor_dia_reused += abs(req.vlr_movto_estoq)
        
        dias_labels.append(data_dia.strftime('%d/%m'))
        dias_data.append(count)
        dias_valor.append(float(valor_dia))
        dias_valor_reused.append(float(valor_dia_reused))
        dias_data_nf.append(count_nf)
        dias_valor_nf.append(float(valor_dia_nf))
        dias_data_normal.append(count_normal)
        dias_valor_normal.append(float(valor_dia_normal))
    
    # Determinar mês selecionado para o gráfico diário
    if ano_filtro and meses_filtro_int:
        mes_selecionado_grafico = f"{ano_filtro}-{str(meses_filtro_int[0]).zfill(2)}"
    else:
        mes_selecionado_grafico = hoje.strftime('%Y-%m')
    
    # Nomes dos meses em português
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Obter lista de anos disponíveis
    anos_disponiveis_list = list(anos_disponiveis)
    if not anos_disponiveis_list:
        anos_disponiveis_list = [hoje.year]
    
    context = {
        'page_title': 'Análise de Requisições',
        'active_page': 'analise_requisicoes',
        'total_requisicoes': total_requisicoes,
        'requisicoes_recentes': requisicoes_recentes,
        'requisicoes_mes_atual': requisicoes_mes_atual,
        'itens_unicos': itens_unicos,
        'centros_unicos': centros_unicos,
        'valor_total': valor_total,
        'quantidade_total': quantidade_total,
        'valor_medio': valor_medio,
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'meses_valor': json.dumps(meses_valor),
        'meses_data_nf': json.dumps(meses_data_nf),
        'meses_valor_nf': json.dumps(meses_valor_nf),
        'meses_data_normal': json.dumps(meses_data_normal),
        'meses_valor_normal': json.dumps(meses_valor_normal),
        'dias_labels': json.dumps(dias_labels),
        'dias_data': json.dumps(dias_data),
        'dias_valor': json.dumps(dias_valor),
        'dias_valor_reused': json.dumps(dias_valor_reused),
        'dias_data_nf': json.dumps(dias_data_nf),
        'dias_valor_nf': json.dumps(dias_valor_nf),
        'dias_data_normal': json.dumps(dias_data_normal),
        'dias_valor_normal': json.dumps(dias_valor_normal),
        'valor_total_reused': valor_total_reused,
        'mes_selecionado': mes_selecionado_grafico,
        'top_itens_labels': json.dumps(top_itens_labels),
        'top_itens_data': json.dumps(top_itens_data),
        'top_itens_valor_labels': json.dumps(top_itens_valor_labels),
        'top_itens_valor_data': json.dumps(top_itens_valor_data),
        # Estatísticas separadas NF vs Normal
        'total_requisicoes_nf': total_requisicoes_nf,
        'total_requisicoes_normal': total_requisicoes_normal,
        'requisicoes_recentes_nf': requisicoes_recentes_nf,
        'requisicoes_recentes_normal': requisicoes_recentes_normal,
        'requisicoes_mes_atual_nf': requisicoes_mes_atual_nf,
        'requisicoes_mes_atual_normal': requisicoes_mes_atual_normal,
        'itens_unicos_nf': itens_unicos_nf,
        'itens_unicos_normal': itens_unicos_normal,
        'centros_unicos_nf': centros_unicos_nf,
        'centros_unicos_normal': centros_unicos_normal,
        'valor_total_nf': valor_total_nf,
        'valor_total_normal': valor_total_normal,
        'valor_total_reused_nf': valor_total_reused_nf,
        'valor_total_reused_normal': valor_total_reused_normal,
        'quantidade_total_nf': quantidade_total_nf,
        'quantidade_total_normal': quantidade_total_normal,
        'valor_medio_nf': valor_medio_nf,
        'valor_medio_normal': valor_medio_normal,
        'top_itens_labels_nf': json.dumps(top_itens_labels_nf),
        'top_itens_data_nf': json.dumps(top_itens_data_nf),
        'top_itens_labels_normal': json.dumps(top_itens_labels_normal),
        'top_itens_data_normal': json.dumps(top_itens_data_normal),
        'top_itens_valor_labels_nf': json.dumps(top_itens_valor_labels_nf),
        'top_itens_valor_data_nf': json.dumps(top_itens_valor_data_nf),
        'top_itens_valor_labels_normal': json.dumps(top_itens_valor_labels_normal),
        'top_itens_valor_data_normal': json.dumps(top_itens_valor_data_normal),
        'centros_labels': json.dumps(centros_labels),
        'centros_data_count': json.dumps(centros_data_count),
        'centros_data_valor': json.dumps(centros_data_valor),
        'operacoes_labels': json.dumps(operacoes_labels),
        'operacoes_data': json.dumps(operacoes_data),
        'usuarios_labels': json.dumps(usuarios_labels),
        'usuarios_data_count': json.dumps(usuarios_data_count),
        'usuarios_data_valor': json.dumps(usuarios_data_valor),
        'requisicoes_recentes_list': requisicoes_recentes_list,
        'requisicoes_nf_page': requisicoes_nf_page,
        # Filtros
        'anos_disponiveis': anos_disponiveis_list,
        'meses_nomes': meses_nomes,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
    }
    return render(request, 'almoxarifado/analise_requisicoes.html', context)


def api_meses_por_ano(request):
    """API endpoint para obter meses disponíveis para um ano específico"""
    from django.http import JsonResponse
    from app.models import RequisicaoAlmoxarifado
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    ano = request.GET.get('ano')
    if not ano:
        return JsonResponse({'error': 'Parâmetro ano é obrigatório'}, status=400)
    
    try:
        ano_int = int(ano)
        meses = RequisicaoAlmoxarifado.objects.filter(
            data_requisicao__year=ano_int
        ).values_list('data_requisicao__month', flat=True).distinct().order_by('data_requisicao__month')
        
        meses_nomes = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
            5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
            9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }
        
        meses_list = []
        for mes_num in meses:
            meses_list.append({
                'value': mes_num,
                'label': meses_nomes[mes_num]
            })
        
        return JsonResponse({'meses': meses_list})
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': f'Ano inválido: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar dados: {str(e)}'}, status=500)


def api_dados_diarios_requisicoes(request):
    """API endpoint para obter dados diários de requisições, manutenções terceiro e visitas para um mês específico"""
    from django.http import JsonResponse
    from calendar import monthrange
    from app.models import RequisicaoAlmoxarifado, ManutencaoTerceiro, Visitas
    from decimal import Decimal
    from datetime import datetime
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    mes_ano = request.GET.get('mes_ano')  # Formato: YYYY-MM
    if not mes_ano:
        return JsonResponse({'error': 'Parâmetro mes_ano é obrigatório'}, status=400)
    
    try:
        year, month = map(int, mes_ano.split('-'))
        primeiro_dia = datetime(year, month, 1).date()
        
        # Calcular último dia do mês
        ultimo_dia_num = monthrange(year, month)[1]
        ultimo_dia = datetime(year, month, ultimo_dia_num).date()
        
        # Hoje para limitar se for o mês atual
        hoje = datetime.now().date()
        if primeiro_dia.year == hoje.year and primeiro_dia.month == hoje.month:
            ultimo_dia = hoje
        
        dias_labels = []
        dias_data = []  # Requisições
        dias_valor = []  # Valor das requisições (apenas cd_depo == 1)
        dias_valor_reused = []  # Valor dos itens reutilizados (cd_depo == 3)
        dias_data_nf = []  # Requisições NF
        dias_valor_nf = []  # Valor das requisições NF
        dias_data_normal = []  # Requisições Normais
        dias_valor_normal = []  # Valor das requisições Normais
        dias_manutencao_terceiro = []  # Manutenções Terceiro
        dias_visitas = []  # Visitas
        
        mensagem_nf = "REQUISIÇÃO GERADA PELO PROCESSO DE ENTRADA DE NOTA FISCAL"
        
        for dia in range(1, ultimo_dia.day + 1):
            data_dia = primeiro_dia.replace(day=dia)
            
            # Requisições de Almoxarifado
            queryset_dia = RequisicaoAlmoxarifado.objects.filter(data_requisicao=data_dia)
            queryset_dia_nf = queryset_dia.filter(obs_rm__icontains=mensagem_nf)
            queryset_dia_normal = queryset_dia.exclude(obs_rm__icontains=mensagem_nf)
            
            count = queryset_dia.count()
            count_nf = queryset_dia_nf.count()
            count_normal = queryset_dia_normal.count()
            
            valor_dia = Decimal('0.00')
            valor_dia_reused = Decimal('0.00')
            valor_dia_nf = Decimal('0.00')
            valor_dia_normal = Decimal('0.00')
            
            for req in queryset_dia:
                if req.vlr_movto_estoq:
                    if req.cd_depo == 1:
                        # Apenas cd_depo == 1 gera custo
                        valor_dia += abs(req.vlr_movto_estoq)
                        if mensagem_nf in (req.obs_rm or ''):
                            valor_dia_nf += abs(req.vlr_movto_estoq)
                        else:
                            valor_dia_normal += abs(req.vlr_movto_estoq)
                    elif req.cd_depo == 3:
                        # cd_depo == 3 são itens reutilizados (não geram custo, mas vamos rastrear)
                        valor_dia_reused += abs(req.vlr_movto_estoq)
            
            # Manutenções Terceiro (filtrar por data, que é DateTimeField)
            manutencao_count = ManutencaoTerceiro.objects.filter(
                data__date=data_dia
            ).count()
            
            # Visitas (filtrar por data, que é DateTimeField)
            visitas_count = Visitas.objects.filter(
                data__date=data_dia
            ).count()
            
            dias_labels.append(data_dia.strftime('%d/%m'))
            dias_data.append(count)
            dias_valor.append(float(valor_dia))
            dias_valor_reused.append(float(valor_dia_reused))
            dias_data_nf.append(count_nf)
            dias_valor_nf.append(float(valor_dia_nf))
            dias_data_normal.append(count_normal)
            dias_valor_normal.append(float(valor_dia_normal))
            dias_manutencao_terceiro.append(manutencao_count)
            dias_visitas.append(visitas_count)
        
        return JsonResponse({
            'labels': dias_labels,
            'data': dias_data,
            'valor': dias_valor,
            'valor_reused': dias_valor_reused,
            'data_nf': dias_data_nf,
            'valor_nf': dias_valor_nf,
            'data_normal': dias_data_normal,
            'valor_normal': dias_valor_normal,
            'manutencao_terceiro': dias_manutencao_terceiro,
            'visitas': dias_visitas
        })
        
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': f'Formato de data inválido: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar dados: {str(e)}'}, status=500)


def testes(request):
    """Página de testes - Hierarquia de Máquinas Primárias e Secundárias"""
    from app.models import Maquina, MaquinaPrimariaSecundaria
    import json
    
    # Buscar todas as máquinas primárias (descr_gerenc = "MÁQUINAS PRINCIPAL")
    maquinas_primarias = Maquina.objects.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).order_by('cd_maquina')
    
    # Buscar todos os relacionamentos
    relacionamentos = MaquinaPrimariaSecundaria.objects.select_related(
        'maquina_primaria', 'maquina_secundaria'
    ).order_by('maquina_primaria__cd_maquina', 'maquina_secundaria__cd_maquina')
    
    # Construir lista de nós no formato básico do OrgChartJS
    # Formato: { id: X, pid: Y, name: "..." }
    nodes = []
    
    # Adicionar máquinas primárias como nós raiz (sem pid)
    for maq_prim in maquinas_primarias:
        nodes.append({
            'id': maq_prim.id,
            'name': f"{maq_prim.cd_maquina} - {maq_prim.descr_maquina or 'Sem descrição'}"
        })
    
    # Adicionar máquinas secundárias como nós filhos (com pid)
    for rel in relacionamentos:
        maq_sec = rel.maquina_secundaria
        nodes.append({
            'id': maq_sec.id,
            'pid': rel.maquina_primaria.id,
            'name': f"{maq_sec.cd_maquina} - {maq_sec.descr_maquina or 'Sem descrição'}"
        })
    
    # Debug: imprimir informações
    print(f"Total de máquinas primárias: {maquinas_primarias.count()}")
    print(f"Total de relacionamentos: {relacionamentos.count()}")
    print(f"Total de nós criados: {len(nodes)}")
    
    # Serializar JSON
    try:
        dados_json_str = json.dumps(nodes, ensure_ascii=False, default=str)
    except Exception as e:
        print(f"Erro ao serializar JSON: {e}")
        dados_json_str = json.dumps([{
            'id': 0,
            'name': 'Erro ao processar dados'
        }], ensure_ascii=False)
    
    context = {
        'page_title': 'Testes - Hierarquia de Máquinas',
        'active_page': 'testes',
        'dados_json': dados_json_str,
        'total_primarias': maquinas_primarias.count(),
        'total_relacionamentos': relacionamentos.count()
    }
    return render(request, 'testes/testes.html', context)


def analise_plano_preventiva(request):
    """Análise de Plano Preventiva"""
    context = {
        'page_title': 'Análise de Plano Preventiva',
        'active_page': 'analise_plano_preventiva'
    }
    return render(request, 'analise/analise_plano_preventiva.html', context)


def analise_roteiro_plano_preventiva(request):
    """Análise de Roteiro e Plano de Preventiva - Encontrar relações baseadas em campos específicos"""
    from app.models import PlanoPreventiva, RoteiroPreventiva, MeuPlanoPreventiva, Maquina
    from django.core.paginator import Paginator
    from django.db import transaction
    from django.contrib import messages
    
    # Verificar se é uma ação de confirmação e salvamento
    if request.method == 'POST':
        # Debug: imprimir dados recebidos
        print(f"=== DEBUG CONFIRMAR RELAÇÃO ===")
        print(f"POST data: {request.POST}")
        print(f"POST keys: {list(request.POST.keys())}")
        print(f"confirmar_relacao in POST: {'confirmar_relacao' in request.POST}")
        
        if 'confirmar_todos' in request.POST:
            print(f"=== DEBUG CONFIRMAR TODOS ===")
            print(f"POST data: {request.POST}")
            print(f"POST keys: {list(request.POST.keys())}")
            # Bulk confirmation - confirm all pending relationships
            relacionamentos_confirmados = 0
            relacionamentos_erro = 0
            
            # Get all planos and roteiros
            planos = PlanoPreventiva.objects.all()
            roteiros = RoteiroPreventiva.objects.all()
            
            # Helper function to check if fields match (same as campos_correspondem)
            def campos_correspondem(plano, roteiro):
                if not plano.cd_maquina or not roteiro.cd_maquina:
                    return False
                if plano.cd_maquina != roteiro.cd_maquina:
                    return False
                
                descr_plano = (plano.descr_maquina or '').strip().upper()
                descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
                if descr_plano and descr_roteiro:
                    if descr_plano != descr_roteiro:
                        return False
                elif descr_plano or descr_roteiro:
                    return False
                
                if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
                    return False
                if plano.sequencia_tarefa != roteiro.cd_tarefamanu:
                    return False
                
                descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
                descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
                if descr_tarefa_plano and descr_tarefa_roteiro:
                    if descr_tarefa_plano != descr_tarefa_roteiro:
                        return False
                elif descr_tarefa_plano or descr_tarefa_roteiro:
                    return False
                
                if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
                    return False
                if plano.sequencia_manutencao != roteiro.seq_seqplamanu:
                    return False
                
                return True
            
            # Process all relationships
            with transaction.atomic():
                for plano in planos:
                    for roteiro in roteiros:
                        if campos_correspondem(plano, roteiro):
                            # Check if already saved
                            ja_existe = MeuPlanoPreventiva.objects.filter(
                                cd_maquina=plano.cd_maquina,
                                numero_plano=plano.numero_plano,
                                sequencia_manutencao=plano.sequencia_manutencao,
                                sequencia_tarefa=plano.sequencia_tarefa
                            ).exists()
                            
                            if not ja_existe:
                                try:
                                    meu_plano, created = MeuPlanoPreventiva.objects.get_or_create(
                                        cd_maquina=plano.cd_maquina,
                                        numero_plano=plano.numero_plano,
                                        sequencia_manutencao=plano.sequencia_manutencao,
                                        sequencia_tarefa=plano.sequencia_tarefa,
                                        defaults={
                                            'cd_unid': plano.cd_unid,
                                            'nome_unid': plano.nome_unid,
                                            'cd_setor': plano.cd_setor,
                                            'descr_setor': plano.descr_setor,
                                            'cd_atividade': plano.cd_atividade,
                                            'descr_maquina': plano.descr_maquina,
                                            'nro_patrimonio': plano.nro_patrimonio,
                                            'descr_plano': plano.descr_plano,
                                            'dt_execucao': plano.dt_execucao,
                                            'quantidade_periodo': plano.quantidade_periodo,
                                            'descr_tarefa': plano.descr_tarefa,
                                            'cd_funcionario': plano.cd_funcionario,
                                            'nome_funcionario': plano.nome_funcionario,
                                            'descr_seqplamanu': roteiro.descr_seqplamanu,
                                            'desc_detalhada_do_roteiro_preventiva': roteiro.descr_seqplamanu,
                                            'roteiro_preventiva': roteiro,
                                            'maquina': plano.maquina,
                                        }
                                    )
                                    
                                    if not created:
                                        meu_plano.desc_detalhada_do_roteiro_preventiva = roteiro.descr_seqplamanu
                                        meu_plano.descr_seqplamanu = roteiro.descr_seqplamanu
                                        meu_plano.roteiro_preventiva = roteiro
                                        meu_plano.cd_unid = plano.cd_unid
                                        meu_plano.nome_unid = plano.nome_unid
                                        meu_plano.cd_setor = plano.cd_setor
                                        meu_plano.descr_setor = plano.descr_setor
                                        meu_plano.cd_atividade = plano.cd_atividade
                                        meu_plano.descr_maquina = plano.descr_maquina
                                        meu_plano.nro_patrimonio = plano.nro_patrimonio
                                        meu_plano.descr_plano = plano.descr_plano
                                        meu_plano.dt_execucao = plano.dt_execucao
                                        meu_plano.quantidade_periodo = plano.quantidade_periodo
                                        meu_plano.descr_tarefa = plano.descr_tarefa
                                        meu_plano.cd_funcionario = plano.cd_funcionario
                                        meu_plano.nome_funcionario = plano.nome_funcionario
                                        meu_plano.maquina = plano.maquina
                                        meu_plano.save()
                                    
                                    relacionamentos_confirmados += 1
                                except Exception as e:
                                    relacionamentos_erro += 1
                                    print(f"Erro ao confirmar relação Plano {plano.id} - Roteiro {roteiro.id}: {str(e)}")
            
            if relacionamentos_confirmados > 0:
                messages.success(request, f'{relacionamentos_confirmados} relação(ões) confirmada(s) e salva(s) com sucesso!')
            if relacionamentos_erro > 0:
                messages.warning(request, f'{relacionamentos_erro} relação(ões) apresentaram erro ao salvar.')
            if relacionamentos_confirmados == 0 and relacionamentos_erro == 0:
                messages.info(request, 'Nenhuma relação pendente para confirmar.')
            
            return redirect('analise_roteiro_plano_preventiva')
        
        elif 'confirmar_relacao' in request.POST:
            plano_id = request.POST.get('plano_id')
            roteiro_id = request.POST.get('roteiro_id')
            
            print(f"plano_id: {plano_id}")
            print(f"roteiro_id: {roteiro_id}")
            
            if not plano_id or not roteiro_id:
                messages.error(request, 'Plano ID ou Roteiro ID não fornecido.')
                return redirect('analise_roteiro_plano_preventiva')
            
            try:
                plano = PlanoPreventiva.objects.get(id=plano_id)
                roteiro = RoteiroPreventiva.objects.get(id=roteiro_id)
                
                # Usar transaction.atomic para garantir integridade dos dados
                with transaction.atomic():
                    # Verificar se já existe um MeuPlanoPreventiva para esta combinação específica
                    # Usar uma combinação mais específica para evitar duplicatas
                    meu_plano, created = MeuPlanoPreventiva.objects.get_or_create(
                        cd_maquina=plano.cd_maquina,
                        numero_plano=plano.numero_plano,
                        sequencia_manutencao=plano.sequencia_manutencao,
                        sequencia_tarefa=plano.sequencia_tarefa,
                        defaults={
                            'cd_unid': plano.cd_unid,
                            'nome_unid': plano.nome_unid,
                            'cd_setor': plano.cd_setor,
                            'descr_setor': plano.descr_setor,
                            'cd_atividade': plano.cd_atividade,
                            'descr_maquina': plano.descr_maquina,
                            'nro_patrimonio': plano.nro_patrimonio,
                            'descr_plano': plano.descr_plano,
                            'dt_execucao': plano.dt_execucao,
                            'quantidade_periodo': plano.quantidade_periodo,
                            'descr_tarefa': plano.descr_tarefa,
                            'cd_funcionario': plano.cd_funcionario,
                            'nome_funcionario': plano.nome_funcionario,
                            'descr_seqplamanu': roteiro.descr_seqplamanu,
                            'desc_detalhada_do_roteiro_preventiva': roteiro.descr_seqplamanu,
                            'roteiro_preventiva': roteiro,
                            'maquina': plano.maquina,
                        }
                    )
                    
                    # Se já existia, atualizar com os dados do roteiro
                    if not created:
                        meu_plano.desc_detalhada_do_roteiro_preventiva = roteiro.descr_seqplamanu
                        meu_plano.descr_seqplamanu = roteiro.descr_seqplamanu
                        meu_plano.roteiro_preventiva = roteiro
                        # Atualizar outros campos que possam ter mudado
                        meu_plano.cd_unid = plano.cd_unid
                        meu_plano.nome_unid = plano.nome_unid
                        meu_plano.cd_setor = plano.cd_setor
                        meu_plano.descr_setor = plano.descr_setor
                        meu_plano.cd_atividade = plano.cd_atividade
                        meu_plano.descr_maquina = plano.descr_maquina
                        meu_plano.nro_patrimonio = plano.nro_patrimonio
                        meu_plano.descr_plano = plano.descr_plano
                        meu_plano.dt_execucao = plano.dt_execucao
                        meu_plano.quantidade_periodo = plano.quantidade_periodo
                        meu_plano.descr_tarefa = plano.descr_tarefa
                        meu_plano.cd_funcionario = plano.cd_funcionario
                        meu_plano.nome_funcionario = plano.nome_funcionario
                        meu_plano.maquina = plano.maquina
                        meu_plano.save()
                
                messages.success(request, f'Relação confirmada e salva com sucesso! Plano {plano.id} vinculado ao Roteiro {roteiro.id} em MeuPlanoPreventiva.')
                # Redirecionar para evitar reenvio do formulário
                return redirect('analise_roteiro_plano_preventiva')
            except PlanoPreventiva.DoesNotExist:
                messages.error(request, 'Plano não encontrado.')
                return redirect('analise_roteiro_plano_preventiva')
            except RoteiroPreventiva.DoesNotExist:
                messages.error(request, 'Roteiro não encontrado.')
                return redirect('analise_roteiro_plano_preventiva')
            except Exception as e:
                messages.error(request, f'Erro ao salvar relação: {str(e)}')
                import traceback
                print(f"Erro ao salvar relação: {traceback.format_exc()}")
                return redirect('analise_roteiro_plano_preventiva')
    
    # Buscar todos os registros
    planos = PlanoPreventiva.objects.all()
    roteiros = RoteiroPreventiva.objects.all()
    
    # Estatísticas gerais
    total_planos = planos.count()
    total_roteiros = roteiros.count()
    
    # Encontrar relacionamentos baseados em correspondência exata dos campos
    relacionamentos = []
    planos_sem_relacao = []
    roteiros_sem_relacao = []
    
    # Processar planos e encontrar relacionamentos
    planos_processados = set()
    roteiros_processados = set()
    
    def campos_correspondem(plano, roteiro):
        """Verifica se os campos principais correspondem exatamente"""
        # Comparar cd_maquina (ambos devem ter valor e serem iguais)
        if not plano.cd_maquina or not roteiro.cd_maquina:
            return False
        if plano.cd_maquina != roteiro.cd_maquina:
            return False
        
        # Comparar descr_maquina (ignorar case e espaços, mas ambos devem ter valor)
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if descr_plano and descr_roteiro:
            if descr_plano != descr_roteiro:
                return False
        elif descr_plano or descr_roteiro:
            # Se apenas um tem valor, não corresponde
            return False
        
        # Comparar sequencia_tarefa (Plano) com cd_tarefamanu (Roteiro)
        if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
            return False
        if plano.sequencia_tarefa != roteiro.cd_tarefamanu:
            return False
        
        # Comparar descr_tarefa (Plano) com descr_tarefamanu (Roteiro) - ignorar case e espaços
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if descr_tarefa_plano and descr_tarefa_roteiro:
            if descr_tarefa_plano != descr_tarefa_roteiro:
                return False
        elif descr_tarefa_plano or descr_tarefa_roteiro:
            # Se apenas um tem valor, não corresponde
            return False
        
        # Comparar sequencia_manutencao (Plano) com seq_seqplamanu (Roteiro)
        if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
            return False
        if plano.sequencia_manutencao != roteiro.seq_seqplamanu:
            return False
        
        return True
    
    # Função para calcular score parcial de match
    def calcular_score_parcial(plano, roteiro):
        """Calcula um score de correspondência parcial (0-100)"""
        score = 0
        total = 0
        
        # cd_maquina (peso 20)
        if plano.cd_maquina and roteiro.cd_maquina:
            total += 20
            if plano.cd_maquina == roteiro.cd_maquina:
                score += 20
        
        # descr_maquina (peso 20)
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if descr_plano and descr_roteiro:
            total += 20
            if descr_plano == descr_roteiro:
                score += 20
        
        # sequencia_tarefa vs cd_tarefamanu (peso 20)
        if plano.sequencia_tarefa and roteiro.cd_tarefamanu:
            total += 20
            if plano.sequencia_tarefa == roteiro.cd_tarefamanu:
                score += 20
        
        # descr_tarefa vs descr_tarefamanu (peso 20)
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if descr_tarefa_plano and descr_tarefa_roteiro:
            total += 20
            if descr_tarefa_plano == descr_tarefa_roteiro:
                score += 20
        
        # sequencia_manutencao vs seq_seqplamanu (peso 20)
        if plano.sequencia_manutencao and roteiro.seq_seqplamanu:
            total += 20
            if plano.sequencia_manutencao == roteiro.seq_seqplamanu:
                score += 20
        
        if total == 0:
            return 0
        return (score / total * 100)
    
    for plano in planos:
        melhor_match = None
        melhor_score = 0
        
        # Buscar roteiros que correspondem exatamente
        for roteiro in roteiros:
            if campos_correspondem(plano, roteiro):
                melhor_match = roteiro
                melhor_score = 100
                break
        
        if melhor_match:
            # Verificar se já foi salvo em MeuPlanoPreventiva
            # Usar a mesma combinação de campos usada no get_or_create
            ja_salvo = MeuPlanoPreventiva.objects.filter(
                cd_maquina=plano.cd_maquina,
                numero_plano=plano.numero_plano,
                sequencia_manutencao=plano.sequencia_manutencao,
                sequencia_tarefa=plano.sequencia_tarefa
            ).exists()
            
            relacionamentos.append({
                'plano': plano,
                'roteiro': melhor_match,
                'descr_seqplamanu': melhor_match.descr_seqplamanu,
                'ja_salvo': ja_salvo,
            })
            planos_processados.add(plano.id)
            roteiros_processados.add(melhor_match.id)
        else:
            # Encontrar melhor match parcial para exibição (sempre encontrar o melhor, mesmo que score < 40%)
            # Isso permite mostrar análise de erros mesmo quando não há match parcial bom
            melhor_match_parcial = None
            melhor_score_parcial = 0
            for roteiro in roteiros:
                if roteiro.id not in roteiros_processados:
                    score = calcular_score_parcial(plano, roteiro)
                    if score > melhor_score_parcial:  # Sempre encontrar o melhor, mesmo que baixo
                        melhor_score_parcial = score
                        melhor_match_parcial = roteiro
            
            planos_sem_relacao.append({
                'plano': plano,
                'melhor_match_parcial': melhor_match_parcial,
                'score_parcial': melhor_score_parcial,
            })
    
    # Encontrar roteiros sem plano correspondente
    for roteiro in roteiros:
        if roteiro.id not in roteiros_processados:
            # Encontrar melhor match parcial para exibição (sempre encontrar o melhor, mesmo que score < 40%)
            # Isso permite mostrar análise de erros mesmo quando não há match parcial bom
            melhor_match_parcial = None
            melhor_score_parcial = 0
            for plano in planos:
                if plano.id not in planos_processados:
                    score = calcular_score_parcial(plano, roteiro)
                    if score > melhor_score_parcial:  # Sempre encontrar o melhor, mesmo que baixo
                        melhor_score_parcial = score
                        melhor_match_parcial = plano
            
            roteiros_sem_relacao.append({
                'roteiro': roteiro,
                'melhor_match_parcial': melhor_match_parcial,
                'score_parcial': melhor_score_parcial,
            })
    
    # Filtros - Obter valores ANTES de aplicar
    filter_maquina = request.GET.get('filter_maquina', '').strip()
    filter_descr_seqplamanu = request.GET.get('filter_descr_seqplamanu', '').strip()
    filter_tipo = request.GET.get('filter_tipo', 'all')  # all, matched, planos_sem, roteiros_sem, salvos
    filter_status = request.GET.get('filter_status', 'all')  # all, pendentes, salvos
    
    # Aplicar filtros
    relacionamentos_filtrados = relacionamentos.copy()
    planos_sem_relacao_filtrados = planos_sem_relacao.copy()
    roteiros_sem_relacao_filtrados = roteiros_sem_relacao.copy()
    
    if filter_maquina:
        try:
            maquina_num = int(float(filter_maquina))
            relacionamentos_filtrados = [r for r in relacionamentos_filtrados if (r['plano'].cd_maquina and r['plano'].cd_maquina == maquina_num) or (r['roteiro'].cd_maquina and r['roteiro'].cd_maquina == maquina_num)]
            planos_sem_relacao_filtrados = [p for p in planos_sem_relacao_filtrados if p['plano'].cd_maquina and p['plano'].cd_maquina == maquina_num]
            roteiros_sem_relacao_filtrados = [r for r in roteiros_sem_relacao_filtrados if r['roteiro'].cd_maquina and r['roteiro'].cd_maquina == maquina_num]
        except (ValueError, TypeError):
            filter_maquina_str = str(filter_maquina).lower()
            relacionamentos_filtrados = [r for r in relacionamentos_filtrados if 
                (r['plano'].cd_maquina and filter_maquina_str in str(r['plano'].cd_maquina).lower()) or 
                (r['roteiro'].cd_maquina and filter_maquina_str in str(r['roteiro'].cd_maquina).lower()) or
                (r['plano'].descr_maquina and filter_maquina_str in str(r['plano'].descr_maquina).lower()) or
                (r['roteiro'].descr_maquina and filter_maquina_str in str(r['roteiro'].descr_maquina).lower())]
            planos_sem_relacao_filtrados = [p for p in planos_sem_relacao_filtrados if 
                (p['plano'].cd_maquina and filter_maquina_str in str(p['plano'].cd_maquina).lower()) or
                (p['plano'].descr_maquina and filter_maquina_str in str(p['plano'].descr_maquina).lower())]
            roteiros_sem_relacao_filtrados = [r for r in roteiros_sem_relacao_filtrados if 
                (r['roteiro'].cd_maquina and filter_maquina_str in str(r['roteiro'].cd_maquina).lower()) or
                (r['roteiro'].descr_maquina and filter_maquina_str in str(r['roteiro'].descr_maquina).lower())]
    
    if filter_descr_seqplamanu:
        filter_descr_str = filter_descr_seqplamanu.lower()
        relacionamentos_filtrados = [r for r in relacionamentos_filtrados if r.get('descr_seqplamanu') and filter_descr_str in r['descr_seqplamanu'].lower()]
    
    if filter_status == 'pendentes':
        relacionamentos_filtrados = [r for r in relacionamentos_filtrados if not r.get('ja_salvo', False)]
    elif filter_status == 'salvos':
        relacionamentos_filtrados = [r for r in relacionamentos_filtrados if r.get('ja_salvo', False)]
    
    # Estatísticas de relacionamentos APÓS filtros
    total_relacionamentos = len(relacionamentos_filtrados)
    total_planos_sem_relacao = len(planos_sem_relacao_filtrados)
    total_roteiros_sem_relacao = len(roteiros_sem_relacao_filtrados)
    total_salvos = sum(1 for rel in relacionamentos_filtrados if rel.get('ja_salvo', False))
    total_pendentes = total_relacionamentos - total_salvos
    
    # Paginação para relacionamentos - usar listas filtradas
    if filter_tipo == 'matched':
        items_to_paginate = relacionamentos_filtrados
    elif filter_tipo == 'planos_sem':
        items_to_paginate = planos_sem_relacao_filtrados
    elif filter_tipo == 'roteiros_sem':
        items_to_paginate = roteiros_sem_relacao_filtrados
    else:
        items_to_paginate = relacionamentos_filtrados
    
    paginator = Paginator(items_to_paginate, 50)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    # Preparar dados para o contexto baseado no tipo de filtro - usar listas filtradas
    if filter_tipo == 'matched':
        relacionamentos_display = list(page_obj)
        planos_sem_display = planos_sem_relacao_filtrados[:50]
        roteiros_sem_display = roteiros_sem_relacao_filtrados[:50]
    elif filter_tipo == 'planos_sem':
        relacionamentos_display = relacionamentos_filtrados[:50]
        planos_sem_display = list(page_obj)
        roteiros_sem_display = roteiros_sem_relacao_filtrados[:50]
    elif filter_tipo == 'roteiros_sem':
        relacionamentos_display = relacionamentos_filtrados[:50]
        planos_sem_display = planos_sem_relacao_filtrados[:50]
        roteiros_sem_display = list(page_obj)
    else:  # all
        relacionamentos_display = relacionamentos_filtrados[:100]
        planos_sem_display = planos_sem_relacao_filtrados[:100]
        roteiros_sem_display = roteiros_sem_relacao_filtrados[:100]
    
    context = {
        'page_title': 'Análise de Roteiro e Plano de Preventiva',
        'active_page': 'analise_roteiro_plano_preventiva',
        'relacionamentos': relacionamentos_display,
        'planos_sem_relacao': planos_sem_display,
        'roteiros_sem_relacao': roteiros_sem_display,
        'total_planos': total_planos,
        'total_roteiros': total_roteiros,
        'total_relacionamentos': total_relacionamentos,
        'total_planos_sem_relacao': total_planos_sem_relacao,
        'total_roteiros_sem_relacao': total_roteiros_sem_relacao,
        'total_salvos': total_salvos,
        'total_pendentes': total_pendentes,
        'filter_maquina': filter_maquina,
        'filter_descr_seqplamanu': filter_descr_seqplamanu,
        'filter_tipo': filter_tipo,
        'filter_status': filter_status,
        'page_obj': page_obj,
    }
    return render(request, 'planejamento/analise_roteiro_plano_preventiva.html', context)


def ajustar_maquinas_outros(request):
    """Página para análise e ajuste de máquinas com descr_gerenc = OUTROS"""
    from .models import Maquina
    from django.db.models import Count, Q
    from django.core.paginator import Paginator
    
    # Queryset base: máquinas OUTROS (inclui NULL/vazio)
    base_queryset = Maquina.objects.filter(ativo=True).filter(
        Q(descr_gerenc__iexact='OUTROS') |
        Q(descr_gerenc__isnull=True) |
        Q(descr_gerenc='')
    ).order_by('cd_maquina')
    
    # Filtros da tabela (query params)
    filter_codigo = request.GET.get('filter_codigo', '').strip()
    filter_descricao = request.GET.get('filter_descricao', '').strip()
    filter_gerenc = request.GET.get('filter_gerenc', '').strip()
    filter_setor = request.GET.get('filter_setor', '').strip()
    filter_patrimonio = request.GET.get('filter_patrimonio', '').strip()
    
    maquinas_queryset = base_queryset
    if filter_codigo:
        try:
            codigo_int = int(filter_codigo)
            maquinas_queryset = maquinas_queryset.filter(cd_maquina=codigo_int)
        except ValueError:
            from django.db.models.functions import Cast
            from django.db.models import CharField
            maquinas_queryset = maquinas_queryset.annotate(
                cd_str=Cast('cd_maquina', CharField())
            ).filter(cd_str__icontains=filter_codigo)
    if filter_descricao:
        maquinas_queryset = maquinas_queryset.filter(descr_maquina__icontains=filter_descricao)
    if filter_gerenc:
        if filter_gerenc.upper() == 'NULL':
            maquinas_queryset = maquinas_queryset.filter(
                Q(descr_gerenc__isnull=True) | Q(descr_gerenc='')
            )
        else:
            maquinas_queryset = maquinas_queryset.filter(descr_gerenc__icontains=filter_gerenc)
    if filter_setor:
        maquinas_queryset = maquinas_queryset.filter(cd_setormanut=filter_setor)
    if filter_patrimonio:
        maquinas_queryset = maquinas_queryset.filter(nro_patrimonio__icontains=filter_patrimonio)
    
    # Paginação
    paginator = Paginator(maquinas_queryset, 50)
    page_number = request.GET.get('page', 1)
    maquinas_outros_paginated = paginator.get_page(page_number)
    
    # Valores únicos para dropdowns
    filter_gerenc_options = list(
        base_queryset.exclude(descr_gerenc__isnull=True)
        .exclude(descr_gerenc='')
        .values_list('descr_gerenc', flat=True)
        .distinct()
        .order_by('descr_gerenc')
    )
    filter_setor_options = list(
        base_queryset.exclude(cd_setormanut__isnull=True)
        .exclude(cd_setormanut='')
        .values_list('cd_setormanut', flat=True)
        .distinct()
        .order_by('cd_setormanut')
    )
    
    # Buscar todos os valores únicos de descr_gerenc (para resumo/cards)
    valores_gerenc = Maquina.objects.filter(
        ativo=True
    ).exclude(
        descr_gerenc__isnull=True
    ).exclude(
        descr_gerenc=''
    ).values('descr_gerenc').annotate(
        total=Count('id')
    ).order_by('descr_gerenc')
    
    maquinas_por_gerenc = {}
    maquinas_outros = []
    
    for item in valores_gerenc:
        gerenc = item['descr_gerenc']
        total = item['total']
        if gerenc and gerenc.upper().strip() == 'OUTROS':
            maquinas_outros.append({
                'descr_gerenc': gerenc,
                'total': total,
                'maquinas': Maquina.objects.filter(
                    ativo=True,
                    descr_gerenc__iexact='OUTROS'
                ).order_by('cd_maquina')
            })
        else:
            maquinas_por_gerenc[gerenc] = {
                'descr_gerenc': gerenc,
                'total': total,
                'maquinas': Maquina.objects.filter(
                    ativo=True,
                    descr_gerenc=gerenc
                ).order_by('cd_maquina')[:10]
            }
    
    maquinas_null = Maquina.objects.filter(
        ativo=True
    ).filter(
        Q(descr_gerenc__isnull=True) | Q(descr_gerenc='')
    ).count()
    
    if maquinas_null > 0:
        maquinas_outros.append({
            'descr_gerenc': None,
            'total': maquinas_null,
            'maquinas': Maquina.objects.filter(
                ativo=True
            ).filter(
                Q(descr_gerenc__isnull=True) | Q(descr_gerenc='')
            ).order_by('cd_maquina')
        })
    
    total_maquinas = Maquina.objects.filter(ativo=True).count()
    total_outros = sum(item['total'] for item in maquinas_outros)
    total_com_gerenc_valida = total_maquinas - total_outros
    gerenc_choices = sorted(set(maquinas_por_gerenc.keys()))
    
    # Build query string for pagination (preserve filters)
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    
    context = {
        'page_title': 'Ajustar Máquinas - OUTROS',
        'active_page': 'ajustar_maquinas_outros',
        'maquinas_outros': maquinas_outros,
        'maquinas_outros_paginated': maquinas_outros_paginated,
        'maquinas_por_gerenc': maquinas_por_gerenc,
        'total_maquinas': total_maquinas,
        'total_outros': total_outros,
        'total_com_gerenc_valida': total_com_gerenc_valida,
        'percentual_outros': round((total_outros / total_maquinas * 100) if total_maquinas > 0 else 0, 1),
        'gerenc_choices': gerenc_choices,
        'filter_codigo': filter_codigo,
        'filter_descricao': filter_descricao,
        'filter_gerenc': filter_gerenc,
        'filter_setor': filter_setor,
        'filter_patrimonio': filter_patrimonio,
        'filter_gerenc_options': filter_gerenc_options,
        'filter_setor_options': filter_setor_options,
        'query_params': query_params,
    }
    return render(request, 'maquinas/ajustar_maquinas_outros.html', context)


def maquina_primaria_secundaria(request):
    """Agrupar Máquinas Primárias e Secundárias"""
    from .models import Maquina, MaquinaPrimariaSecundaria
    from django.contrib import messages
    
    # Filtro por setor (descr_setormanut) - GET
    filtro_setor = (request.GET.get('setor') or request.GET.get('descr_setormanut') or '').strip() or None
    setores_disponiveis = list(
        Maquina.objects.exclude(descr_setormanut__isnull=True)
        .exclude(descr_setormanut='')
        .values_list('descr_setormanut', flat=True)
        .distinct()
        .order_by('descr_setormanut')
    )
    
    # Buscar máquinas primárias (descr_gerenc = "MÁQUINAS PRINCIPAL") - apenas ativas; opcionalmente filtrar por setor
    maquinas_primarias = Maquina.objects.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL',
        ativo=True
    )
    if filtro_setor:
        maquinas_primarias = maquinas_primarias.filter(descr_setormanut=filtro_setor)
    maquinas_primarias = maquinas_primarias.order_by('cd_maquina')
    
    # Buscar máquinas secundárias que ainda não estão relacionadas - apenas ativas; opcionalmente filtrar por setor
    maquinas_secundarias_relacionadas = MaquinaPrimariaSecundaria.objects.values_list('maquina_secundaria_id', flat=True)
    maquinas_secundarias = Maquina.objects.filter(
        ativo=True
    ).exclude(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).exclude(
        id__in=maquinas_secundarias_relacionadas
    )
    if filtro_setor:
        maquinas_secundarias = maquinas_secundarias.filter(descr_setormanut=filtro_setor)
    maquinas_secundarias = maquinas_secundarias.order_by('cd_maquina')
    
    # Buscar relacionamentos existentes; opcionalmente filtrar por setor da máquina primária
    relacionamentos = MaquinaPrimariaSecundaria.objects.select_related('maquina_primaria', 'maquina_secundaria').order_by('-created_at')
    if filtro_setor:
        relacionamentos = relacionamentos.filter(maquina_primaria__descr_setormanut=filtro_setor)
    
    # Processar POST para criar relacionamentos
    if request.method == 'POST':
        # Debug: imprimir dados recebidos
        import traceback
        print(f"=== DEBUG MAQUINA PRIMARIA SECUNDARIA ===")
        print(f"POST data: {request.POST}")
        print(f"POST keys: {list(request.POST.keys())}")
        print(f"criar_relacionamento: {'criar_relacionamento' in request.POST}")
        print(f"remover_relacionamento: {'remover_relacionamento' in request.POST}")
        
        if 'criar_relacionamento' in request.POST:
            maquina_primaria_id = request.POST.get('maquina_primaria')
            maquinas_secundarias_ids = request.POST.getlist('maquinas_secundarias')
            observacoes = request.POST.get('observacoes', '').strip()
            
            print(f"maquina_primaria_id: {maquina_primaria_id}")
            print(f"maquinas_secundarias_ids: {maquinas_secundarias_ids}")
            print(f"observacoes: {observacoes}")
            
            if not maquina_primaria_id:
                messages.error(request, 'Por favor, selecione uma máquina primária.')
            elif not maquinas_secundarias_ids or (len(maquinas_secundarias_ids) == 1 and not maquinas_secundarias_ids[0]):
                messages.error(request, 'Por favor, selecione pelo menos uma máquina secundária.')
            else:
                try:
                    maquina_primaria = Maquina.objects.get(id=maquina_primaria_id)
                    if maquina_primaria.descr_gerenc and maquina_primaria.descr_gerenc.upper() != 'MÁQUINAS PRINCIPAL':
                        messages.error(request, 'A máquina selecionada não é uma máquina primária.')
                    else:
                        relacionamentos_criados = 0
                        relacionamentos_duplicados = 0
                        
                        for secundaria_id in maquinas_secundarias_ids:
                            try:
                                maquina_secundaria = Maquina.objects.get(id=secundaria_id)
                                
                                # Verificar se já existe o relacionamento
                                if MaquinaPrimariaSecundaria.objects.filter(
                                    maquina_primaria=maquina_primaria,
                                    maquina_secundaria=maquina_secundaria
                                ).exists():
                                    relacionamentos_duplicados += 1
                                else:
                                    MaquinaPrimariaSecundaria.objects.create(
                                        maquina_primaria=maquina_primaria,
                                        maquina_secundaria=maquina_secundaria,
                                        observacoes=observacoes if observacoes else None
                                    )
                                    relacionamentos_criados += 1
                            except Maquina.DoesNotExist:
                                continue
                        
                        if relacionamentos_criados > 0:
                            messages.success(request, f'{relacionamentos_criados} relacionamento(s) criado(s) com sucesso.')
                        if relacionamentos_duplicados > 0:
                            messages.warning(request, f'{relacionamentos_duplicados} relacionamento(s) já existia(m) e foi(ram) ignorado(s).')
                except Maquina.DoesNotExist:
                    messages.error(request, 'Máquina primária não encontrada.')
                except Exception as e:
                    messages.error(request, f'Erro ao criar relacionamento: {str(e)}')
        
        elif 'remover_relacionamento' in request.POST:
            relacionamento_id = request.POST.get('relacionamento_id')
            if relacionamento_id:
                try:
                    relacionamento = MaquinaPrimariaSecundaria.objects.get(id=relacionamento_id)
                    relacionamento.delete()
                    messages.success(request, 'Relacionamento removido com sucesso.')
                except MaquinaPrimariaSecundaria.DoesNotExist:
                    messages.error(request, 'Relacionamento não encontrado.')
                except Exception as e:
                    messages.error(request, f'Erro ao remover relacionamento: {str(e)}')
        
        from django.urls import reverse
        from urllib.parse import quote
        url = reverse('maquina_primaria_secundaria')
        # Manter filtro setor após POST (vem do GET na página ou do hidden no form)
        setor_redirect = request.POST.get('redirect_setor') or filtro_setor
        if setor_redirect:
            url += '?setor=' + quote(setor_redirect)
        return redirect(url)
    
    context = {
        'page_title': 'Agrupar Máquinas Primárias e Secundárias',
        'active_page': 'maquina_primaria_secundaria',
        'maquinas_primarias': maquinas_primarias,
        'maquinas_secundarias': maquinas_secundarias,
        'relacionamentos': relacionamentos,
        'setores_disponiveis': setores_disponiveis,
        'filtro_setor': filtro_setor,
    }
    return render(request, 'maquinas/maquina_primaria_secundaria.html', context)


def contact(request):
    """Contact page view"""
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        newsletter = request.POST.get('newsletter', False)
        
        # Basic validation
        if not all([name, email, subject, message]):
            messages.error(request, 'Por favor, preencha todos os campos obrigatórios.')
        else:
            # Here you would typically save to database or send email
            # For now, we'll just show a success message
            messages.success(request, f'Obrigado {name}! Sua mensagem foi enviada com sucesso. Entraremos em contato em breve.')
            
            # Optional: Send email notification
            try:
                send_mail(
                    f'Contato via site - {subject}',
                    f'Nome: {name}\nEmail: {email}\nTelefone: {phone}\nAssunto: {subject}\nMensagem: {message}',
                    settings.DEFAULT_FROM_EMAIL,
                    [settings.DEFAULT_FROM_EMAIL],
                    fail_silently=False,
                )
            except Exception as e:
                # Log the error but don't show it to the user
                print(f"Email sending failed: {e}")
            
            return redirect('contact')
    
    context = {
        'page_title': 'Contato',
        'active_page': 'contact'
    }
    return render(request, 'contact.html', context)


def services(request):
    """Services page view"""
    context = {
        'page_title': 'Serviços',
        'active_page': 'services'
    }
    return render(request, 'services.html', context)


def importar_maquinas(request):
    """Importar Máquinas page view"""
    if request.method == 'POST':
        # Verificar primeiro se é confirmação de inativação
        if 'confirm_inactivate' in request.POST:
            from app.models import Maquina
            maquinas_to_inactivate = request.POST.getlist('inactivate_machines')
            print(f"DEBUG - confirm_inactivate recebido. Máquinas selecionadas: {maquinas_to_inactivate}")
            
            if maquinas_to_inactivate:
                # Converter IDs para inteiros
                try:
                    maquinas_ids = [int(id) for id in maquinas_to_inactivate]
                    print(f"DEBUG - IDs convertidos: {maquinas_ids}")
                    
                    # Verificar quantas máquinas existem antes da atualização
                    maquinas_antes = Maquina.objects.filter(id__in=maquinas_ids, ativo=True).count()
                    print(f"DEBUG - Máquinas ativas antes da atualização: {maquinas_antes}")
                    
                    inactivated_count = Maquina.objects.filter(
                        id__in=maquinas_ids
                    ).update(ativo=False)
                    
                    print(f"DEBUG - Máquinas atualizadas: {inactivated_count}")
                    
                    # Verificar se realmente foram atualizadas
                    maquinas_depois = Maquina.objects.filter(id__in=maquinas_ids, ativo=False).count()
                    print(f"DEBUG - Máquinas inativas depois da atualização: {maquinas_depois}")
                    
                    if inactivated_count > 0:
                        messages.success(request, f'{inactivated_count} máquina(s) marcada(s) como inativa(s).')
                    else:
                        messages.warning(request, 'Nenhuma máquina foi atualizada. Verifique se as máquinas selecionadas existem.')
                    
                    # Limpar sessão
                    if 'missing_machines' in request.session:
                        del request.session['missing_machines']
                    if 'import_success' in request.session:
                        del request.session['import_success']
                    if 'created_count' in request.session:
                        del request.session['created_count']
                    if 'updated_count' in request.session:
                        del request.session['updated_count']
                except (ValueError, TypeError) as e:
                    print(f"DEBUG - Erro ao processar IDs: {str(e)}")
                    messages.error(request, f'Erro ao processar IDs das máquinas: {str(e)}')
            else:
                messages.warning(request, 'Nenhuma máquina foi selecionada para inativação.')
            
            # Redirecionar para evitar reenvio do formulário
            return redirect('importar_maquinas')
        
        # Se não é confirmação de inativação, verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Máquinas',
                'active_page': 'importar_maquinas'
            }
            return render(request, 'importar/importar_maquinas.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Máquinas',
                'active_page': 'importar_maquinas'
            }
            return render(request, 'importar/importar_maquinas.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        update_fields = []
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
            # Se update_existing estiver marcado, pegar lista de campos para atualizar
            if update_existing:
                update_fields = request.POST.getlist('update_fields')
                # Se nenhum campo foi selecionado, atualizar todos (comportamento padrão)
                if not update_fields:
                    update_fields = [
                        'cd_unid', 'nome_unid', 'cs_tt_maquina', 'descr_maquina',
                        'cd_setormanut', 'descr_setormanut', 'cd_priomaqutv',
                        'nro_patrimonio', 'cd_modelo', 'cd_grupo', 'cd_tpcentativ',
                        'descr_gerenc'
                    ]
        
        try:
            from app.utils import upload_maquinas_from_file
            
            # Fazer upload dos dados
            # Se only_new_records estiver marcado, update_existing será False (ignora duplicados)
            created_count, updated_count, errors, missing_machines = upload_maquinas_from_file(
                file, 
                update_existing=update_existing,
                update_fields=update_fields if update_existing else None
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(
                        request, 
                        f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.'
                    )
            
            if created_count > 0 or updated_count > 0:
                success_msg = f'Importação concluída com sucesso! '
                if created_count > 0:
                    success_msg += f'{created_count} registro(s) criado(s). '
                if updated_count > 0:
                    success_msg += f'{updated_count} registro(s) atualizado(s).'
                messages.success(request, success_msg)
            elif not errors:
                messages.info(request, 'Nenhum registro foi importado.')
            
            # Se há máquinas não encontradas no arquivo, armazenar na sessão para mostrar na página
            if missing_machines:
                # Converter para lista de dicionários simples para garantir serialização correta
                missing_machines_list = [
                    {
                        'id': m.get('id'),
                        'cd_maquina': m.get('cd_maquina'),
                        'descr_maquina': m.get('descr_maquina'),
                        'cd_setormanut': m.get('cd_setormanut'),
                        'descr_setormanut': m.get('descr_setormanut'),
                    }
                    for m in missing_machines
                ]
                request.session['missing_machines'] = missing_machines_list
                request.session['import_success'] = True
                request.session['created_count'] = created_count
                request.session['updated_count'] = updated_count
                print(f"DEBUG - Armazenando {len(missing_machines_list)} máquinas na sessão")
            
        except Exception as e:
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
    
    # Buscar máquinas não encontradas da sessão
    missing_machines = request.session.get('missing_machines', [])
    import_success = request.session.get('import_success', False)
    created_count = request.session.get('created_count', 0)
    updated_count = request.session.get('updated_count', 0)
    
    # Debug: verificar dados da sessão
    if missing_machines:
        print(f"DEBUG - Recuperando {len(missing_machines)} máquinas da sessão")
        if missing_machines:
            print(f"DEBUG - Primeira máquina da sessão: {missing_machines[0]}")
    
    context = {
        'page_title': 'Importar Máquinas',
        'active_page': 'importar_maquinas',
        'missing_machines': missing_machines,
        'import_success': import_success,
        'created_count': created_count,
        'updated_count': updated_count,
    }
    return render(request, 'importar/importar_maquinas.html', context)


def exportar_maquinas(request):
    """Exportar Máquinas - página e download Excel/CSV com dados da tabela Maquina"""
    from django.http import HttpResponse
    from app.models import Maquina
    import csv
    import io
    from datetime import datetime

    formato = (request.GET.get('formato') or '').strip().lower()
    if formato in ('csv', 'xlsx'):
        # Exportar todas as máquinas (tabela Maquina); filtro ativo opcional via GET
        incluir_inativas = request.GET.get('incluir_inativas') == '1'
        if incluir_inativas:
            maquinas = Maquina.objects.all().order_by('cd_maquina')
        else:
            maquinas = Maquina.objects.filter(ativo=True).order_by('cd_maquina')
        maquinas = maquinas.select_related('centro_atividade')
        filename_base = f'maquinas_export_{datetime.now().strftime("%Y%m%d_%H%M")}'

        # Cabeçalhos: ordem preferida (descr_setormanut, cd_setormanut, descr_maquina, cd_maquina, descr_gerenc primeiro), demais a seguir
        headers = [
            'descr_setormanut', 'cd_setormanut', 'descr_maquina', 'cd_maquina', 'descr_gerenc',
            'cd_unid', 'nome_unid', 'cs_tt_maquina', 'cd_priomaqutv', 'nro_patrimonio',
            'cd_modelo', 'cd_grupo', 'cd_tpcentativ',
            'centro_atividade_id', 'centro_atividade_ca', 'ativo', 'created_at', 'updated_at',
            'foto', 'placa_identificacao', 'codigo_aurora', 'codigo_fabricante',
            'arquivo_pdf', 'diagrama_eletrico', 'pecas_reposicao',
        ]

        def row_from_maquina(m):
            return [
                m.descr_setormanut or '',
                m.cd_setormanut or '',
                m.descr_maquina or '',
                m.cd_maquina,
                m.descr_gerenc or '',
                m.cd_unid,
                m.nome_unid or '',
                m.cs_tt_maquina,
                m.cd_priomaqutv,
                m.nro_patrimonio or '',
                m.cd_modelo,
                m.cd_grupo,
                m.cd_tpcentativ,
                m.centro_atividade_id,
                m.centro_atividade.ca if m.centro_atividade else '',
                m.ativo,
                m.created_at.strftime('%Y-%m-%d %H:%M:%S') if m.created_at else '',
                m.updated_at.strftime('%Y-%m-%d %H:%M:%S') if m.updated_at else '',
                m.foto.name if m.foto else '',
                m.placa_identificacao.name if m.placa_identificacao else '',
                m.codigo_aurora.name if m.codigo_aurora else '',
                m.codigo_fabricante.name if m.codigo_fabricante else '',
                m.arquivo_pdf.name if m.arquivo_pdf else '',
                m.diagrama_eletrico.name if m.diagrama_eletrico else '',
                m.pecas_reposicao.name if m.pecas_reposicao else '',
            ]

        if formato == 'csv':
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)
            for m in maquinas:
                writer.writerow(row_from_maquina(m))
            response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            return response

        elif formato == 'xlsx':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Máquinas'
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, m in enumerate(maquinas, 2):
                for col_idx, val in enumerate(row_from_maquina(m), 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return response

    context = {
        'page_title': 'Exportar Máquinas',
        'active_page': 'exportar_maquinas',
    }
    return render(request, 'exportar/exportar_maquinas.html', context)


def importar_manutentores(request):
    """Importar Manutentores page view"""
    if request.method == 'POST':
        print(f"DEBUG - POST recebido! Files: {list(request.FILES.keys())}")
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Manutentores',
                'active_page': 'importar_manutentores'
            }
            return render(request, 'importar/importar_manutentor.html', context)
        
        file = request.FILES['file']
        print(f"DEBUG - Arquivo recebido: {file.name}, Tamanho: {file.size}")
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Manutentores',
                'active_page': 'importar_manutentores'
            }
            return render(request, 'importar/importar_manutentor.html', context)
        
        # Verificar se deve atualizar registros existentes
        update_existing = request.POST.get('update_existing', 'off') == 'on'
        print(f"DEBUG - Update existing: {update_existing}")
        
        try:
            from app.utils import upload_manutentores_from_file
            
            # Fazer upload dos dados
            print("DEBUG - Iniciando upload...")
            created_count, updated_count, errors = upload_manutentores_from_file(
                file,
                update_existing=update_existing
            )
            
            print(f"DEBUG - Upload concluído: {created_count} criados, {updated_count} atualizados, {len(errors)} erros")
            
            # Exibir mensagens
            if created_count > 0:
                messages.success(request, f'{created_count} manutentor(es) criado(s) com sucesso!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} manutentor(es) atualizado(s) com sucesso!')
            if errors:
                for error in errors[:10]:  # Limitar a 10 erros para não sobrecarregar
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, 'Nenhum registro foi importado. Verifique se o arquivo contém dados válidos.')
        
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"DEBUG - Erro durante upload: {error_detail}")
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
        
        # Sempre redirecionar para consultar_manutentores após importação
        return redirect('consultar_manutentores')
    
    context = {
        'page_title': 'Importar Manutentores',
        'active_page': 'importar_manutentores'
    }
    return render(request, 'importar/importar_manutentor.html', context)


def importar_ordens_corretivas_e_outros(request):
    """Importar Ordens Corretivas e Outros page view"""
    if request.method == 'POST':
        print(f"DEBUG - POST recebido! Files: {list(request.FILES.keys())}")
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Ordens Corretivas e Outros',
                'active_page': 'importar_ordens_corretivas_e_outros'
            }
            return render(request, 'importar/importar_ordens_corretiva_outros.html', context)
        
        file = request.FILES['file']
        print(f"DEBUG - Arquivo recebido: {file.name}, Tamanho: {file.size}")
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Ordens Corretivas e Outros',
                'active_page': 'importar_ordens_corretivas_e_outros'
            }
            return render(request, 'importar/importar_ordens_corretiva_outros.html', context)
        
        # Verificar se deve atualizar registros existentes
        update_existing = request.POST.get('update_existing', 'off') == 'on'
        print(f"DEBUG - Update existing: {update_existing}")
        
        try:
            from app.utils import upload_ordens_corretivas_from_file
            
            # Fazer upload dos dados
            print("DEBUG - Iniciando upload...")
            created_count, updated_count, errors = upload_ordens_corretivas_from_file(
                file, 
                update_existing=update_existing
            )
            print(f"DEBUG - Upload concluído: criados={created_count}, atualizados={updated_count}, erros={len(errors)}")
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(
                        request, 
                        f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.'
                    )
            
            # Sempre redirecionar após tentativa de importação, independente do resultado
            if created_count > 0 or updated_count > 0:
                success_msg = f'Importação concluída com sucesso! '
                if created_count > 0:
                    success_msg += f'{created_count} registro(s) criado(s). '
                if updated_count > 0:
                    success_msg += f'{updated_count} registro(s) atualizado(s).'
                messages.success(request, success_msg)
            elif not errors:
                messages.info(request, 'Nenhum registro foi importado.')
            else:
                # Se houver apenas erros, ainda redireciona mas mostra os erros
                messages.warning(request, 'Importação concluída com erros. Verifique as mensagens acima.')
            
            # Redirecionar para a página de consulta após importação
            return redirect('consultar_corretivas_outros')
            
        except Exception as e:
            import traceback
            print(f"DEBUG - Erro: {str(e)}")
            print(f"DEBUG - Traceback: {traceback.format_exc()}")
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
            # Redirecionar mesmo em caso de erro para não ficar na página de importação
            return redirect('consultar_corretivas_outros')
    
    context = {
        'page_title': 'Importar Ordens Corretivas e Outros',
        'active_page': 'importar_ordens_corretivas_e_outros'
    }
    return render(request, 'importar/importar_ordens_corretiva_outros.html', context)


def importar_ordens_preventivas(request):
    """Importar Ordens Preventivas page view"""
    if request.method == 'POST':
        print(f"DEBUG - POST recebido! Files: {list(request.FILES.keys())}")
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Ordens Preventivas',
                'active_page': 'importar_ordens_preventivas'
            }
            return render(request, 'importar/importar_ordens_preventivas.html', context)
        
        file = request.FILES['file']
        print(f"DEBUG - Arquivo recebido: {file.name}, Tamanho: {file.size}")
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Ordens Preventivas',
                'active_page': 'importar_ordens_preventivas'
            }
            return render(request, 'importar/importar_ordens_preventivas.html', context)
        
        # Verificar se deve atualizar registros existentes
        update_existing = request.POST.get('update_existing', 'off') == 'on'

        print(f"DEBUG - Update existing: {update_existing}")
        
        try:
            from app.utils import upload_ordens_preventivas_from_file
            
            # Fazer upload dos dados
            print("DEBUG - Iniciando upload...")
            created_count, updated_count, errors = upload_ordens_preventivas_from_file(
                file, 
                update_existing=update_existing
            )
            print(f"DEBUG - Upload concluído: criados={created_count}, atualizados={updated_count}, erros={len(errors)}")
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(
                        request, 
                        f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.'
                    )
            
            # Sempre redirecionar após tentativa de importação, independente do resultado
            if created_count > 0 or updated_count > 0:
                success_msg = f'Importação concluída com sucesso! '
                if created_count > 0:
                    success_msg += f'{created_count} registro(s) criado(s). '
                if updated_count > 0:
                    success_msg += f'{updated_count} registro(s) atualizado(s).'
                messages.success(request, success_msg)
            elif not errors:
                messages.info(request, 'Nenhum registro foi importado.')
            else:
                # Se houver apenas erros, ainda redireciona mas mostra os erros
                messages.warning(request, 'Importação concluída com erros. Verifique as mensagens acima.')
            
            # Redirecionar de volta para a página de importação
            return redirect('importar_ordens_preventivas')
            
        except Exception as e:
            error_msg = f'Erro ao processar arquivo: {str(e)}'
            messages.error(request, error_msg)
            print(f"Erro na importação: {error_msg}")
            import traceback
            traceback.print_exc()
    
    context = {
        'page_title': 'Importar Ordens Preventivas',
        'active_page': 'importar_ordens_preventivas'
    }
    return render(request, 'importar/importar_ordens_preventivas.html', context)


def importar_plano_preventiva(request):
    """Importar Plano Preventiva page view"""
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Plano Preventiva',
                'active_page': 'importar_plano_preventiva'
            }
            return render(request, 'importar/importar_plano_preventiva.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Plano Preventiva',
                'active_page': 'importar_plano_preventiva'
            }
            return render(request, 'importar/importar_plano_preventiva.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_plano_preventiva_from_file
            
            # Fazer upload dos dados
            # Se only_new_records estiver marcado, update_existing será False (ignora duplicados)
            created_count, updated_count, errors = upload_plano_preventiva_from_file(
                file, 
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erros.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} registro(s) de plano preventiva criado(s) com sucesso!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} registro(s) de plano preventiva atualizado(s)!')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, 'Nenhum registro novo foi importado. Todos os registros já existem no banco de dados.')
            
        except Exception as e:
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
    
    context = {
        'page_title': 'Importar Plano Preventiva',
        'active_page': 'importar_plano_preventiva'
    }
    return render(request, 'importar/importar_plano_preventiva.html', context)


def importar_roteiro_preventiva(request):
    """Importar Roteiro Preventiva page view"""
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Roteiro Preventiva',
                'active_page': 'importar_roteiro_preventiva'
            }
            return render(request, 'importar/importar_roteiro_preventiva.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Roteiro Preventiva',
                'active_page': 'importar_roteiro_preventiva'
            }
            return render(request, 'importar/importar_roteiro_preventiva.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_roteiro_preventiva_from_file
            
            # Fazer upload dos dados
            # Se only_new_records estiver marcado, update_existing será False (ignora duplicados)
            created_count, updated_count, errors = upload_roteiro_preventiva_from_file(
                file, 
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erros.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} registro(s) de roteiro preventiva criado(s) com sucesso!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} registro(s) de roteiro preventiva atualizado(s)!')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, 'Nenhum registro novo foi importado. Todos os registros já existem no banco de dados.')
            
        except Exception as e:
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
    
    context = {
        'page_title': 'Importar Roteiro Preventiva',
        'active_page': 'importar_roteiro_preventiva'
    }
    return render(request, 'importar/importar_roteiro_preventiva.html', context)


def importar_52_semanas(request):
    """Importar 52 Semanas page view"""
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar 52 Semanas',
                'active_page': 'importar_52_semanas'
            }
            return render(request, 'importar/importar_52_semanas.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar 52 Semanas',
                'active_page': 'importar_52_semanas'
            }
            return render(request, 'importar/importar_52_semanas.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_52_semanas_from_file
            import traceback
            
            # Fazer upload dos dados
            # Se only_new_records estiver marcado, update_existing será False (ignora duplicados)
            created_count, updated_count, errors = upload_52_semanas_from_file(
                file, 
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erros.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} semana(s) criada(s) com sucesso!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} semana(s) atualizada(s)!')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, 'Nenhum registro novo foi importado. Todas as semanas já existem no banco de dados.')
            
        except Exception as e:
            error_msg = f'Erro ao importar arquivo: {str(e)}'
            messages.error(request, error_msg)
            # Log detalhado do erro para debug
            import traceback
            print(f"Erro ao importar 52 semanas: {error_msg}")
            traceback.print_exc()
    
    context = {
        'page_title': 'Importar 52 Semanas',
        'active_page': 'importar_52_semanas'
    }
    return render(request, 'importar/importar_52_semanas.html', context)


def importar_notas_fiscais(request):
    """Importar Notas Fiscais page view"""
    from app.utils import upload_notas_fiscais_from_file
    from django.contrib import messages
    
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Notas Fiscais',
                'active_page': 'importar_notas_fiscais'
            }
            return render(request, 'importar/importar_notas_fiscais.html', context)
        
        file = request.FILES['file']
        update_existing = request.POST.get('update_existing') == 'on'
        only_new_records = request.POST.get('only_new_records') == 'on'
        
        # Se "Atualizar registros existentes" estiver marcado, respeitar (prioridade)
        # Caso contrário, se "only_new_records" estiver marcado, não atualizar
        if only_new_records and not update_existing:
            update_existing = False
        
        # Processar arquivo
        try:
            created_count, updated_count, errors = upload_notas_fiscais_from_file(file, update_existing=update_existing)
            
            # Mensagens de sucesso
            if created_count > 0:
                messages.success(request, f'{created_count} nota(s) fiscal(is) importada(s) com sucesso!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} nota(s) fiscal(is) atualizada(s).')
            if created_count == 0 and updated_count == 0:
                messages.warning(request, 'Nenhuma nota fiscal foi importada. Verifique se há novos registros no arquivo.')
            
            # Mensagens de erro
            if errors:
                for error in errors[:10]:  # Limitar a 10 erros para não sobrecarregar
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f'... e mais {len(errors) - 10} erro(s). Verifique o console para mais detalhes.')
        
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    context = {
        'page_title': 'Importar Notas Fiscais',
        'active_page': 'importar_notas_fiscais'
    }
    return render(request, 'importar/importar_notas_fiscais.html', context)


def importar_paradas_maquina(request):
    """Importar Paradas de Máquina page view"""
    from app.utils import upload_paradas_maquina_from_file
    from app.models import ParadaMaquina
    from django.contrib import messages
    from datetime import date
    
    # Contexto comum: anos disponíveis, meses, ano atual
    anos_paradas = list(ParadaMaquina.objects.exclude(data__isnull=True).values_list('data__year', flat=True).distinct().order_by('-data__year'))
    ano_atual = date.today().year
    if ano_atual not in anos_paradas:
        anos_paradas = [ano_atual] + anos_paradas
    anos_disponiveis_paradas = sorted(set(anos_paradas), reverse=True) if anos_paradas else [ano_atual]
    
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    base_context = {
        'page_title': 'Importar Paradas de Máquina',
        'active_page': 'importar_paradas_maquina',
        'anos_disponiveis_paradas': anos_disponiveis_paradas,
        'meses_nomes': meses_nomes,
        'ano_atual': ano_atual,
    }
    
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            return render(request, 'importar/importar_paradas_de_maquina.html', base_context)
        
        file = request.FILES['file']
        
        # Ano + meses obrigatórios: sempre substituir os meses selecionados
        try:
            ano_subst = int(request.POST.get('ano_substituir', ano_atual))
            meses_raw = request.POST.getlist('mes_substituir')
            meses_subst = [int(m) for m in meses_raw if m and 1 <= int(m) <= 12]
            meses_subst = sorted(set(meses_subst))
            if not ano_subst or not meses_subst:
                messages.error(request, 'Selecione pelo menos um mês que corresponde aos dados do arquivo (campo Data).')
                return render(request, 'importar/importar_paradas_de_maquina.html', base_context)
            substituir_meses = (ano_subst, meses_subst)
        except (ValueError, TypeError):
            messages.error(request, 'Ano ou mês(es) inválido(s).')
            return render(request, 'importar/importar_paradas_de_maquina.html', base_context)

        try:
            created_count, updated_count, errors = upload_paradas_maquina_from_file(
                file, update_existing=False, excluir_antigos=False,
                substituir_meses=substituir_meses
            )

            ano_s, meses_s = substituir_meses
            nomes = [meses_nomes.get(m, str(m)) for m in meses_s]
            messages.success(request, f'Dados de {", ".join(nomes)}/{ano_s} substituídos. {created_count} parada(s) importada(s) com sucesso.')
            if updated_count > 0:
                messages.info(request, f'{updated_count} parada(s) atualizada(s).')
            
            # Mensagens de erro
            if errors:
                for error in errors[:10]:  # Limitar a 10 erros para não sobrecarregar
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f'... e mais {len(errors) - 10} erro(s). Verifique o console para mais detalhes.')
        
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    return render(request, 'importar/importar_paradas_de_maquina.html', base_context)


def relatorio_nf_estf0198(request):
    """Relatório NF ESTF0198 page view - Ajuste de arquivo CSV"""
    from django.http import HttpResponse
    import csv
    import io
    import re
    
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo CSV para processar.')
            context = {
                'page_title': 'Relatório NF ESTF0198',
                'active_page': 'relatorio_nf_estf0198'
            }
            return render(request, 'analise_relatorios/relatorio_nf_estf0198.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        file_name = file.name.lower()
        if not file_name.endswith('.csv'):
            messages.error(
                request, 
                f'<strong>Formato inválido:</strong> O arquivo "{file.name}" não é um arquivo CSV. Por favor, selecione um arquivo com extensão .csv'
            )
            context = {
                'page_title': 'Relatório NF ESTF0198',
                'active_page': 'relatorio_nf_estf0198'
            }
            return render(request, 'analise_relatorios/relatorio_nf_estf0198.html', context)
        
        try:
            # Tentar diferentes encodings
            encodings_to_try = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
            content = None
            encoding_used = None
            
            for encoding in encodings_to_try:
                try:
                    file.seek(0)
                    content = file.read().decode(encoding)
                    encoding_used = encoding
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                raise ValueError("Não foi possível decodificar o arquivo com nenhum encoding testado")
            
            # Ler CSV manualmente primeiro para detectar problemas de formatação
            # O csv.reader pode não lidar bem com aspas não balanceadas
            lines = content.strip().split('\n')
            if not lines:
                raise ValueError("Arquivo CSV vazio")
            
            # Processar linha por linha manualmente para corrigir problemas
            all_rows = []
            for line in lines:
                if not line.strip():
                    continue
                
                # Tentar usar csv.reader primeiro
                csv_line = io.StringIO(line)
                try:
                    reader = csv.reader(csv_line, delimiter=';')
                    row = next(reader)
                    all_rows.append(row)
                except:
                    # Se falhar, fazer split manual e depois corrigir
                    row = line.split(';')
                    all_rows.append(row)
            
            if not all_rows:
                raise ValueError("Arquivo CSV vazio")
            
            # Ler cabeçalho
            headers = [h.strip() for h in all_rows[0]]
            
            # Encontrar índice da coluna "Observações" (pode ter variações de encoding)
            observacoes_index = None
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                # Normalizar para comparar (remover acentos)
                import unicodedata
                header_normalized = unicodedata.normalize('NFKD', header_lower).encode('ASCII', 'ignore').decode('ASCII')
                if 'observa' in header_normalized:
                    observacoes_index = idx
                    break
            
            if observacoes_index is None:
                # Tentar encontrar por posição (geralmente é a coluna 19 ou 20)
                if len(headers) > 19:
                    observacoes_index = 19  # Coluna T em Excel (0-indexed seria 19)
            
            # Processar linhas de dados
            fixed_rows = [headers]  # Começar com cabeçalho
            rows_processed = 0
            rows_fixed = 0
            
            for row_num, row in enumerate(all_rows[1:], start=2):
                if not row or not any(cell.strip() for cell in row):
                    continue
                
                # Verificar se é linha de resumo (começa com "Unidade")
                if row and len(row) > 0 and str(row[0]).strip().upper() == 'UNIDADE':
                    # Garantir que linha de resumo tenha número correto de colunas
                    while len(row) < len(headers):
                        row.append('')
                    fixed_rows.append(row)
                    continue
                
                # Garantir que temos valores suficientes
                while len(row) < len(headers):
                    row.append('')
                
                # CORREÇÃO PRINCIPAL: Detectar e corrigir células com semicolons que deveriam ser separadores
                # O problema: células como "texto;;data;;;" onde os ;; deveriam separar colunas
                # Exemplo: "ASSISTENCIA T�CNICA DO TANQUE DE CO2. - 112582 - 14441;;09/01/2026;;;;"
                # Deveria ser: "ASSISTENCIA..." na coluna atual, "" na próxima, "09/01/2026" na seguinte, etc.
                
                # Verificar cada célula a partir da coluna Observações
                start_check_index = observacoes_index if observacoes_index is not None else 19
                
                # Processar célula por célula, começando da coluna Observações
                col_idx = start_check_index
                while col_idx < len(row) and col_idx < len(headers):
                    cell_value = str(row[col_idx]).strip()
                    
                    # Detectar padrão de semicolons duplos/triplos que indicam campos concatenados
                    if ';;' in cell_value:
                        # Remover aspas não balanceadas primeiro
                        if cell_value.startswith('"') and not cell_value.endswith('"'):
                            if cell_value.count('"') == 1:
                                cell_value = cell_value.lstrip('"')
                        
                        # Usar regex para dividir por 2 ou mais semicolons consecutivos
                        import re
                        # Encontrar todos os padrões de semicolons múltiplos e suas posições
                        # Dividir a string preservando informações sobre quantos semicolons havia
                        parts = []
                        semicolon_counts = []
                        
                        # Encontrar todos os grupos de semicolons múltiplos
                        matches = list(re.finditer(r';{2,}', cell_value))
                        
                        if matches:
                            # Primeira parte (antes do primeiro grupo de semicolons)
                            first_part = cell_value[:matches[0].start()].strip()
                            if first_part:
                                parts.append(first_part)
                            
                            # Processar cada grupo de semicolons e a parte seguinte
                            for i, match in enumerate(matches):
                                semicolon_count = len(match.group())  # Quantos semicolons (;; = 2, ;;; = 3, etc.)
                                semicolon_counts.append(semicolon_count)
                                
                                # Parte após este grupo de semicolons
                                start_pos = match.end()
                                if i < len(matches) - 1:
                                    end_pos = matches[i+1].start()
                                else:
                                    end_pos = len(cell_value)
                                
                                next_part = cell_value[start_pos:end_pos].strip()
                                if next_part:
                                    parts.append(next_part)
                            
                            # Se encontramos partes separadas, distribuir nas colunas
                            if len(parts) > 1:
                                # Primeira parte vai na coluna atual
                                row[col_idx] = parts[0]
                                
                                # Distribuir partes restantes
                                current_col = col_idx
                                for part_idx, part in enumerate(parts[1:], start=1):
                                    # Quantos semicolons havia antes desta parte?
                                    if part_idx - 1 < len(semicolon_counts):
                                        semicolon_count = semicolon_counts[part_idx - 1]
                                        # ;; = 2 semicolons = 1 coluna vazia antes da próxima
                                        # ;;; = 3 semicolons = 2 colunas vazias antes da próxima
                                        # Então avançamos (semicolon_count - 1) colunas
                                        current_col += (semicolon_count - 1)
                                    
                                    # Avançar para próxima coluna
                                    current_col += 1
                                    
                                    # Garantir que não ultrapassamos os limites
                                    if current_col < len(headers):
                                        # Expandir row se necessário
                                        while len(row) <= current_col:
                                            row.append('')
                                        # Preencher a coluna com a parte atual
                                        row[current_col] = part
                                
                                rows_fixed += 1
                                # Avançar para próxima coluna após processar todas as partes
                                col_idx = current_col + 1
                            else:
                                # Não havia partes separadas, apenas limpar
                                row[col_idx] = cell_value.strip()
                                col_idx += 1
                        else:
                            # Não encontrou padrão, apenas limpar
                            row[col_idx] = cell_value.strip()
                            col_idx += 1
                    else:
                        # Limpar célula normalmente
                        row[col_idx] = cell_value.strip()
                        col_idx += 1
                
                # Garantir que temos o número correto de colunas
                while len(row) < len(headers):
                    row.append('')
                # Truncar se tiver mais valores que headers
                if len(row) > len(headers):
                    row = row[:len(headers)]
                
                # Limpar todos os campos (remover espaços extras)
                row = [str(cell).strip() for cell in row]
                
                fixed_rows.append(row)
                rows_processed += 1
            
            # Criar novo CSV em memória
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            
            for row in fixed_rows:
                # Garantir que todas as linhas tenham o mesmo número de colunas
                while len(row) < len(headers):
                    row.append('')
                # Truncar se tiver mais colunas
                if len(row) > len(headers):
                    row = row[:len(headers)]
                writer.writerow(row)
            
            # Armazenar arquivo processado na sessão para download posterior
            output.seek(0)
            processed_content = output.getvalue()
            
            # Armazenar na sessão (codificar em base64 para garantir serialização)
            import base64
            processed_content_encoded = base64.b64encode(processed_content.encode(encoding_used)).decode('utf-8')
            
            request.session['relatorio_nf_processed_file'] = processed_content_encoded
            request.session['relatorio_nf_encoding'] = encoding_used
            request.session['relatorio_nf_filename'] = f"ajustado_{file.name}"
            request.session['relatorio_nf_rows_processed'] = rows_processed
            request.session['relatorio_nf_rows_fixed'] = rows_fixed
            
            messages.success(
                request, 
                f'<i class="fas fa-check-circle me-2"></i><strong>Arquivo processado com sucesso!</strong> '
                f'{rows_processed} linha(s) processada(s), {rows_fixed} linha(s) corrigida(s). '
                f'Clique no botão abaixo para baixar o arquivo ajustado.'
            )
            
        except Exception as e:
            messages.error(
                request, 
                f'<i class="fas fa-exclamation-triangle me-2"></i><strong>Erro ao processar arquivo:</strong> {str(e)}'
            )
            import traceback
            print(f"Erro ao processar relatório NF ESTF0198: {traceback.format_exc()}")
    
    # Verificar se há arquivo processado na sessão
    has_processed_file = 'relatorio_nf_processed_file' in request.session
    rows_processed = request.session.get('relatorio_nf_rows_processed', 0)
    rows_fixed = request.session.get('relatorio_nf_rows_fixed', 0)
    
    context = {
        'page_title': 'Relatório NF ESTF0198',
        'active_page': 'relatorio_nf_estf0198',
        'has_processed_file': has_processed_file,
        'rows_processed': rows_processed,
        'rows_fixed': rows_fixed,
    }
    return render(request, 'analise_relatorios/relatorio_nf_estf0198.html', context)


def download_relatorio_nf_estf0198(request):
    """Download do arquivo CSV processado"""
    from django.http import HttpResponse
    import base64
    
    # Verificar se há arquivo processado na sessão
    if 'relatorio_nf_processed_file' not in request.session:
        messages.error(request, 'Nenhum arquivo processado encontrado. Por favor, processe um arquivo primeiro.')
        return redirect('relatorio_nf_estf0198')
    
    try:
        # Recuperar dados da sessão
        processed_content_encoded = request.session.get('relatorio_nf_processed_file')
        encoding_used = request.session.get('relatorio_nf_encoding', 'latin-1')
        filename = request.session.get('relatorio_nf_filename', 'ajustado_relatorio.csv')
        
        # Decodificar conteúdo
        processed_content = base64.b64decode(processed_content_encoded).decode(encoding_used)
        
        # Preparar resposta para download
        response = HttpResponse(processed_content.encode(encoding_used), content_type=f'text/csv; charset={encoding_used}')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Limpar sessão após download (opcional - você pode manter se quiser permitir múltiplos downloads)
        # del request.session['relatorio_nf_processed_file']
        # del request.session['relatorio_nf_encoding']
        # del request.session['relatorio_nf_filename']
        
        return response
        
    except Exception as e:
        messages.error(
            request, 
            f'<i class="fas fa-exclamation-triangle me-2"></i><strong>Erro ao gerar download:</strong> {str(e)}'
        )
        import traceback
        print(f"Erro ao fazer download do relatório NF ESTF0198: {traceback.format_exc()}")
        return redirect('relatorio_nf_estf0198')


def importar_requisicoes_almoxarifado(request):
    """Importar Requisições Almoxarifado page view"""
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Requisições Almoxarifado',
                'active_page': 'importar_requisicoes_almoxarifado'
            }
            return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)
        
        file = request.FILES['file']
        use_file_name_date = request.POST.get('use_file_name_date') == 'on'
        data_requisicao_str = request.POST.get('data_requisicao')
        
        # Se usar data do nome do arquivo, extrair do nome
        if use_file_name_date:
            import re
            file_name = file.name
            # Remover extensão
            name_without_ext = file_name.replace('.csv', '').replace('.CSV', '')
            # Tentar padrão DD.MM.YYYY
            date_pattern = r'(\d{2})\.(\d{2})\.(\d{4})'
            match = re.search(date_pattern, name_without_ext)
            
            if match:
                day = match.group(1)
                month = match.group(2)
                year = match.group(3)
                data_requisicao_str = f"{year}-{month}-{day}"
            else:
                messages.error(request, f'Não foi possível extrair a data do nome do arquivo "{file_name}". O formato esperado é DD.MM.YYYY (ex: 01.11.2025.csv).')
                context = {
                    'page_title': 'Importar Requisições Almoxarifado',
                    'active_page': 'importar_requisicoes_almoxarifado'
                }
                return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)
        
        # Verificar se há data da requisição
        if not data_requisicao_str:
            messages.error(request, 'Por favor, informe a data da requisição ou marque a opção para usar a data do nome do arquivo.')
            context = {
                'page_title': 'Importar Requisições Almoxarifado',
                'active_page': 'importar_requisicoes_almoxarifado'
            }
            return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)
        
        # Validar extensão do arquivo
        allowed_extensions = ['.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Requisições Almoxarifado',
                'active_page': 'importar_requisicoes_almoxarifado'
            }
            return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_requisicoes_almoxarifado_from_file
            from datetime import datetime
            
            # Converter data_requisicao_str para date
            try:
                data_requisicao = datetime.strptime(data_requisicao_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, f'Formato de data inválido: {data_requisicao_str}. Use YYYY-MM-DD')
                context = {
                    'page_title': 'Importar Requisições Almoxarifado',
                    'active_page': 'importar_requisicoes_almoxarifado'
                }
                return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)
            
            # Fazer upload dos dados
            created_count, updated_count, errors = upload_requisicoes_almoxarifado_from_file(
                file, 
                data_requisicao=data_requisicao,
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erros.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} requisição(ões) criada(s) com sucesso para a data {data_requisicao.strftime("%d/%m/%Y")}!')
            if updated_count > 0:
                messages.info(request, f'{updated_count} requisição(ões) atualizada(s) para a data {data_requisicao.strftime("%d/%m/%Y")}!')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, f'Nenhum registro novo foi importado para a data {data_requisicao.strftime("%d/%m/%Y")}. Todas as requisições já existem no banco de dados.')
            
        except Exception as e:
            error_msg = f'Erro ao importar arquivo: {str(e)}'
            messages.error(request, error_msg)
            # Log detalhado do erro para debug
            import traceback
            print(f"Erro ao importar requisições almoxarifado: {error_msg}")
            traceback.print_exc()
    
    context = {
        'page_title': 'Importar Requisições Almoxarifado',
        'active_page': 'importar_requisicoes_almoxarifado'
    }
    return render(request, 'importar/importar_requisicoes_almoxaridado.html', context)


def importar_projecao_gastos(request):
    """Importar Projeção de Gastos page view"""
    from django.contrib import messages
    from app.utils import upload_projecao_gastos_from_file
    
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Projeção de Gastos',
                'active_page': 'importar_projecao_gastos'
            }
            return render(request, 'importar/importar_projecao_gastos.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Projeção de Gastos',
                'active_page': 'importar_projecao_gastos'
            }
            return render(request, 'importar/importar_projecao_gastos.html', context)
        
        # Verificar opção de atualização
        update_existing = request.POST.get('update_existing', 'off') == 'on'
        debug_import = request.POST.get('debug_import', 'off') == 'on'
        
        try:
            # Fazer upload dos dados
            created_count, updated_count, errors = upload_projecao_gastos_from_file(
                file, 
                update_existing=update_existing,
                debug=debug_import
            )
            
            # Preparar mensagens (debug mostra mais linhas)
            if errors:
                limit = 30 if debug_import else 10
                for error in errors[:limit]:
                    (messages.info if error.startswith('[DEBUG]') else messages.warning)(request, error)
                if len(errors) > limit:
                    messages.warning(request, f'... e mais {len(errors) - limit} mensagens.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} projeção(ões) de gasto(s) criada(s) com sucesso! Os dados foram gravados no banco. Recarregue a página de consulta e verifique os filtros (ano, mês, setor) se não aparecerem.')
            if updated_count > 0:
                messages.info(request, f'{updated_count} projeção(ões) de gasto(s) atualizada(s)! Os dados foram gravados no banco. Recarregue a página de consulta e verifique os filtros (ano, mês, setor) se não aparecerem.')
            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, 'Nenhum registro criado ou atualizado. Verifique se a opção "Atualizar registros existentes" está marcada e se o arquivo tem a estrutura esperada (colunas ID, SETOR, etc.).')
            
            
        except Exception as e:
            error_msg = f'Erro ao importar arquivo: {str(e)}'
            messages.error(request, error_msg)
            # Log detalhado do erro para debug
            import traceback
            print(f"Erro ao importar projeção de gastos: {error_msg}")
            traceback.print_exc()
    
    context = {
        'page_title': 'Importar Projeção de Gastos',
        'active_page': 'importar_projecao_gastos'
    }
    return render(request, 'importar/importar_projecao_gastos.html', context)


def importar_controle_nf_e_rc(request):
    """Importar Controle RC e NF page view"""
    from django.contrib import messages
    from app.utils import upload_controle_rc_e_nf_from_file
    
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Controle RC e NF',
                'active_page': 'importar_controle_nf_e_rc'
            }
            return render(request, 'importar/importar_controle_nf_e_rc.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Controle RC e NF',
                'active_page': 'importar_controle_nf_e_rc'
            }
            return render(request, 'importar/importar_controle_nf_e_rc.html', context)
        
        # Verificar opção de atualização
        update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            # Fazer upload dos dados
            created_count, updated_count, errors, duplicates = upload_controle_rc_e_nf_from_file(
                file, 
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... e mais {len(errors) - 10} erros.')
            
            if created_count > 0:
                messages.success(request, f'{created_count} registro(s) criado(s) com sucesso!')
            
            if updated_count > 0:
                messages.success(request, f'{updated_count} registro(s) atualizado(s) com sucesso!')
            
            if duplicates:
                messages.info(request, f'{len(duplicates)} registro(s) duplicado(s) encontrado(s) e não foram importados. Veja detalhes abaixo.')
            
            if created_count == 0 and updated_count == 0 and not errors and not duplicates:
                messages.info(request, 'Nenhum registro foi importado. Verifique se o arquivo contém dados válidos.')
                
        except Exception as e:
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
            duplicates = []
    
    # Se não houve POST ou houve erro, inicializar duplicates vazio
    if request.method != 'POST' or 'duplicates' not in locals():
        duplicates = []
    
    context = {
        'page_title': 'Importar Controle RC e NF',
        'active_page': 'importar_controle_nf_e_rc',
        'duplicates': duplicates,
    }
    
    return render(request, 'importar/importar_controle_nf_e_rc.html', context)


def importar_estoque(request):
    """Importar Estoque page view"""
    if request.method == 'POST':
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Estoque',
                'active_page': 'importar_estoque'
            }
            return render(request, 'importar/estoque.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Estoque',
                'active_page': 'importar_estoque'
            }
            return render(request, 'importar/estoque.html', context)
        
        # Verificar se deve atualizar registros existentes
        update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_itens_estoque_from_file
            
            # Fazer upload dos dados
            created_count, updated_count, errors = upload_itens_estoque_from_file(
                file,
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(
                        request,
                        f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.'
                    )
            
            if created_count > 0 or updated_count > 0:
                success_msg = f'Importação concluída com sucesso! '
                if created_count > 0:
                    success_msg += f'{created_count} registro(s) criado(s). '
                if updated_count > 0:
                    success_msg += f'{updated_count} registro(s) atualizado(s).'
                messages.success(request, success_msg)
            elif not errors:
                messages.info(request, 'Nenhum registro foi importado.')
            else:
                messages.warning(request, 'Importação concluída com erros. Verifique as mensagens acima.')
            
        except Exception as e:
            import traceback
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
            print(f"DEBUG - Erro: {str(e)}")
            print(f"DEBUG - Traceback: {traceback.format_exc()}")
    
    context = {
        'page_title': 'Importar Estoque',
        'active_page': 'importar_estoque'
    }
    return render(request, 'importar/estoque.html', context)


def consultar_estoque(request):
    """Consultar/listar itens de estoque cadastrados com filtros avançados"""
    from app.models import ItemEstoque
    from decimal import Decimal
    
    # Buscar todos os itens de estoque
    itens_list = ItemEstoque.objects.all()
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    print(f"DEBUG consultar_estoque - search_query: '{search_query}'")
    print(f"DEBUG consultar_estoque - request.GET: {dict(request.GET)}")
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(codigo_item=search_num)
            print(f"DEBUG consultar_estoque - Added numeric search for codigo_item={search_num}")
        except (ValueError, TypeError):
            print(f"DEBUG consultar_estoque - Could not convert '{search_query}' to number")
            pass
        
        # Para campos de texto, usar icontains
        text_conditions = (
            Q(descricao_item__icontains=search_query) |
            Q(unidade_medida__icontains=search_query) |
            Q(descricao_dest_uso__icontains=search_query) |
            Q(classificacao_tempo_sem_consumo__icontains=search_query)
        )
        search_conditions |= text_conditions
        print(f"DEBUG consultar_estoque - Added text search conditions")
        
        itens_list = itens_list.filter(search_conditions)
        print(f"DEBUG consultar_estoque - Filtered count: {itens_list.count()}")
    
    # Filtros específicos
    # Filtro por Unidade de Medida
    filtro_unidade_medida = request.GET.get('filtro_unidade_medida', '')
    if filtro_unidade_medida:
        itens_list = itens_list.filter(unidade_medida__icontains=filtro_unidade_medida)
    
    # Filtro por Destino de Uso
    filtro_destino_uso = request.GET.get('filtro_destino_uso', '')
    if filtro_destino_uso:
        itens_list = itens_list.filter(descricao_dest_uso__icontains=filtro_destino_uso)
    
    # Filtro por Controla Estoque Mínimo
    filtro_controla_estoque = request.GET.get('filtro_controla_estoque', '')
    if filtro_controla_estoque:
        itens_list = itens_list.filter(controla_estoque_minimo__icontains=filtro_controla_estoque)
    
    # Filtro por Classificação Tempo Sem Consumo
    filtro_classificacao = request.GET.get('filtro_classificacao', '')
    if filtro_classificacao:
        itens_list = itens_list.filter(classificacao_tempo_sem_consumo__icontains=filtro_classificacao)
    
    # Filtro por Estante
    filtro_estante = request.GET.get('filtro_estante', '')
    if filtro_estante:
        try:
            estante_num = int(filtro_estante)
            itens_list = itens_list.filter(estante=estante_num)
        except ValueError:
            pass
    
    # Filtro por Prateleira
    filtro_prateleira = request.GET.get('filtro_prateleira', '')
    if filtro_prateleira:
        try:
            prateleira_num = int(filtro_prateleira)
            itens_list = itens_list.filter(prateleira=prateleira_num)
        except ValueError:
            pass
    
    # Filtro por Quantidade (mínima e máxima)
    quantidade_min = request.GET.get('quantidade_min', '')
    quantidade_max = request.GET.get('quantidade_max', '')
    if quantidade_min:
        try:
            qtd_min = Decimal(quantidade_min)
            itens_list = itens_list.filter(quantidade__gte=qtd_min)
        except (ValueError, TypeError):
            pass
    if quantidade_max:
        try:
            qtd_max = Decimal(quantidade_max)
            itens_list = itens_list.filter(quantidade__lte=qtd_max)
        except (ValueError, TypeError):
            pass
    
    # Ordenar por código do item
    itens_list = itens_list.order_by('codigo_item')
    
    # Paginação
    paginator = Paginator(itens_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    itens = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = ItemEstoque.objects.count()
    unidades_count = ItemEstoque.objects.exclude(unidade_medida__isnull=True).exclude(unidade_medida='').values('unidade_medida').distinct().count()
    destinos_count = ItemEstoque.objects.exclude(descricao_dest_uso__isnull=True).exclude(descricao_dest_uso='').values('descricao_dest_uso').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros
    unidades_medida_unicas = ItemEstoque.objects.exclude(
        unidade_medida__isnull=True
    ).exclude(
        unidade_medida=''
    ).values_list('unidade_medida', flat=True).distinct().order_by('unidade_medida')
    
    destinos_uso_unicos = ItemEstoque.objects.exclude(
        descricao_dest_uso__isnull=True
    ).exclude(
        descricao_dest_uso=''
    ).values_list('descricao_dest_uso', flat=True).distinct().order_by('descricao_dest_uso')
    
    classificacoes_unicas = ItemEstoque.objects.exclude(
        classificacao_tempo_sem_consumo__isnull=True
    ).exclude(
        classificacao_tempo_sem_consumo=''
    ).values_list('classificacao_tempo_sem_consumo', flat=True).distinct().order_by('classificacao_tempo_sem_consumo')
    
    context = {
        'page_title': 'Consultar Estoque',
        'active_page': 'consultar_estoque',
        'itens': itens,
        'total_count': total_count,
        'unidades_count': unidades_count,
        'destinos_count': destinos_count,
        # Valores dos filtros ativos
        'filtro_unidade_medida': filtro_unidade_medida,
        'filtro_destino_uso': filtro_destino_uso,
        'filtro_controla_estoque': filtro_controla_estoque,
        'filtro_classificacao': filtro_classificacao,
        'filtro_estante': filtro_estante,
        'filtro_prateleira': filtro_prateleira,
        'quantidade_min': quantidade_min,
        'quantidade_max': quantidade_max,
        # Valores únicos para dropdowns
        'unidades_medida_unicas': unidades_medida_unicas,
        'destinos_uso_unicos': destinos_uso_unicos,
        'classificacoes_unicas': classificacoes_unicas,
    }
    return render(request, 'consultar/consultar_estoque.html', context)


def importar_locais_e_cas(request):
    """Importar Locais e CAs page view"""
    print(f"DEBUG - Método da requisição: {request.method}")
    if request.method == 'POST':
        print("DEBUG - POST recebido!")
        # Verificar se há arquivo enviado
        if 'file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo para importar.')
            context = {
                'page_title': 'Importar Locais e CAs',
                'active_page': 'importar_locais_e_cas'
            }
            return render(request, 'importar/importar_centro_de_atividade.html', context)
        
        file = request.FILES['file']
        
        # Validar extensão do arquivo
        allowed_extensions = ['.xlsx', '.xls', '.xlsm', '.csv']
        file_extension = '.' + file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'Formato de arquivo não suportado. Use: {", ".join(allowed_extensions)}'
            )
            context = {
                'page_title': 'Importar Locais e CAs',
                'active_page': 'importar_locais_e_cas'
            }
            return render(request, 'importar/importar_centro_de_atividade.html', context)
        
        # Verificar se deve apenas adicionar novos registros (ignorar duplicados)
        only_new_records = request.POST.get('only_new_records', 'off') == 'on'
        
        # Verificar se deve atualizar registros existentes
        # Se only_new_records estiver marcado, update_existing será ignorado
        update_existing = False
        if not only_new_records:
            update_existing = request.POST.get('update_existing', 'off') == 'on'
        
        try:
            from app.utils import upload_cas_from_file
            
            # Fazer upload dos dados
            # Se only_new_records estiver marcado, update_existing será False (ignora duplicados)
            created_count, updated_count, errors = upload_cas_from_file(
                file, 
                update_existing=update_existing
            )
            
            # Preparar mensagens
            if errors:
                for error in errors[:10]:  # Mostrar apenas os primeiros 10 erros
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(
                        request, 
                        f'... e mais {len(errors) - 10} erro(s). Verifique o arquivo.'
                    )
            
            if created_count > 0 or updated_count > 0:
                success_msg = f'Importação concluída com sucesso! '
                if created_count > 0:
                    success_msg += f'{created_count} registro(s) criado(s). '
                if updated_count > 0:
                    success_msg += f'{updated_count} registro(s) atualizado(s).'
                messages.success(request, success_msg)
            elif not errors:
                messages.info(request, 'Nenhum registro foi importado.')
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messages.error(request, f'Erro ao importar arquivo: {str(e)}')
            print(f"DEBUG - Erro ao importar: {error_detail}")
    
    context = {
        'page_title': 'Importar Locais e CAs',
        'active_page': 'importar_locais_e_cas'
    }
    return render(request, 'importar/importar_centro_de_atividade.html', context)


def cadastrar_local_e_cas(request):
    """Cadastrar novo Centro de Atividade (CA) com local"""
    from app.forms import CentroAtividadeForm
    import uuid
    
    if request.method == 'POST':
        form = CentroAtividadeForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                centro_atividade = form.save(commit=False)
                
                # Processar upload de imagem se fornecido
                if 'imagem_upload' in request.FILES:
                    imagem_file = request.FILES['imagem_upload']
                    
                    # Gerar nome único para o arquivo
                    file_extension = os.path.splitext(imagem_file.name)[1]
                    unique_filename = f"ca_{centro_atividade.ca}_{uuid.uuid4().hex[:8]}{file_extension}"
                    
                    # Caminho completo para salvar
                    fotos_home_path = os.path.join(settings.STATICFILES_DIRS[0], 'fotos_home')
                    
                    # Criar diretório se não existir
                    os.makedirs(fotos_home_path, exist_ok=True)
                    
                    # Caminho completo do arquivo
                    file_path = os.path.join(fotos_home_path, unique_filename)
                    
                    # Salvar arquivo
                    with open(file_path, 'wb+') as destination:
                        for chunk in imagem_file.chunks():
                            destination.write(chunk)
                    
                    # Atualizar campo imagem com caminho relativo (sem 'static/')
                    imagem_path = f"fotos_home/{unique_filename}"
                    if imagem_path.startswith('static/'):
                        imagem_path = imagem_path.replace('static/', '', 1)
                    centro_atividade.imagem = imagem_path
                    messages.success(request, f'Foto do Centro de Atividade {centro_atividade.ca} cadastrada com sucesso!')
                else:
                    # Se não houve upload, normalizar o caminho do campo 'imagem' se fornecido
                    if centro_atividade.imagem and centro_atividade.imagem.startswith('static/'):
                        centro_atividade.imagem = centro_atividade.imagem.replace('static/', '', 1)
                
                centro_atividade.save()
                messages.success(request, f'Centro de Atividade {centro_atividade.ca} cadastrado com sucesso!')
                return redirect('consultar_locais_e_cas')
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                print(f"Erro ao cadastrar Centro de Atividade: {error_detail}")
                messages.error(request, f'Erro ao cadastrar Centro de Atividade: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = CentroAtividadeForm()
    
    context = {
        'page_title': 'Cadastrar Centro de Atividade',
        'active_page': 'cadastrar_centro_de_atividade',
        'form': form,
    }
    return render(request, 'cadastrar/cadastrar_centro_de_atividade.html', context)


def consultar_locais_e_cas(request):
    """Consultar/listar Centros de Atividade (CA) cadastrados"""
    from app.models import CentroAtividade
    
    # Buscar todos os centros de atividade
    cas_list = CentroAtividade.objects.all()
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(ca=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(sigla__icontains=search_query) |
            Q(descricao__icontains=search_query) |
            Q(encarregado_responsavel__icontains=search_query) |
            Q(local__icontains=search_query) |
            Q(observacoes__icontains=search_query)
        )
        
        cas_list = cas_list.filter(search_conditions).distinct()
    
    # Filtros por coluna individual
    filter_ca = request.GET.get('filter_ca', '').strip()
    if filter_ca:
        try:
            ca_num = int(float(filter_ca))
            cas_list = cas_list.filter(ca=ca_num)
        except (ValueError, TypeError):
            cas_list = cas_list.filter(ca__icontains=filter_ca)
    
    filter_sigla = request.GET.get('filter_sigla', '').strip()
    if filter_sigla:
        cas_list = cas_list.filter(sigla__icontains=filter_sigla)
    
    filter_descricao = request.GET.get('filter_descricao', '').strip()
    if filter_descricao:
        cas_list = cas_list.filter(descricao__icontains=filter_descricao)
    
    filter_indice = request.GET.get('filter_indice', '').strip()
    if filter_indice:
        try:
            indice_num = int(float(filter_indice))
            cas_list = cas_list.filter(indice=indice_num)
        except (ValueError, TypeError):
            cas_list = cas_list.filter(indice__icontains=filter_indice)
    
    filter_encarregado = request.GET.get('filter_encarregado', '').strip()
    if filter_encarregado:
        cas_list = cas_list.filter(encarregado_responsavel__icontains=filter_encarregado)
    
    filter_local = request.GET.get('filter_local', '').strip()
    if filter_local:
        cas_list = cas_list.filter(local__icontains=filter_local)
    
    # Ordenar por código CA
    cas_list = cas_list.order_by('ca').distinct()
    
    # Paginação
    paginator = Paginator(cas_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    cas = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = CentroAtividade.objects.count()
    siglas_count = CentroAtividade.objects.exclude(sigla__isnull=True).exclude(sigla='').values('sigla').distinct().count()
    encarregados_count = CentroAtividade.objects.exclude(encarregado_responsavel__isnull=True).exclude(encarregado_responsavel='').values('encarregado_responsavel').distinct().count()
    
    context = {
        'page_title': 'Consultar Centros de Atividades',
        'active_page': 'consultar_locais_e_cas',
        'cas': cas,
        'total_count': total_count,
        'siglas_count': siglas_count,
        'encarregados_count': encarregados_count,
        # Preservar filtros no contexto
        'filter_ca': filter_ca,
        'filter_sigla': filter_sigla,
        'filter_descricao': filter_descricao,
        'filter_indice': filter_indice,
        'filter_encarregado': filter_encarregado,
        'filter_local': filter_local,
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_centros_de_atividades.html', context)


def visualizar_centro_de_atividade(request, ca_id):
    """Visualizar detalhes de um Centro de Atividade específico"""
    from app.models import CentroAtividade, Maquina, MaquinaPrimariaSecundaria
    import json
    
    try:
        centro_atividade = CentroAtividade.objects.get(id=ca_id)
    except CentroAtividade.DoesNotExist:
        messages.error(request, 'Centro de Atividade não encontrado.')
        return redirect('consultar_locais_e_cas')
    
    # Buscar máquinas relacionadas a este Centro de Atividade
    # Máquinas podem estar relacionadas através de:
    # 1. centro_atividade (campo direto na tabela Maquina)
    # 2. cd_tpcentativ (campo direto na tabela Maquina que corresponde ao número do CA)
    from django.db.models import Q
    
    maquinas_do_ca = Maquina.objects.filter(
        Q(centro_atividade=centro_atividade) |
        Q(cd_tpcentativ=centro_atividade.ca)
    ).distinct()
    
    print(f"DEBUG: Centro de Atividade CA={centro_atividade.ca}, ID={centro_atividade.id}")
    print(f"DEBUG: Total de máquinas encontradas no CA (via centro_atividade): {Maquina.objects.filter(centro_atividade=centro_atividade).count()}")
    print(f"DEBUG: Total de máquinas encontradas no CA (via cd_tpcentativ={centro_atividade.ca}): {Maquina.objects.filter(cd_tpcentativ=centro_atividade.ca).count()}")
    print(f"DEBUG: Total de máquinas encontradas no CA (total combinado): {maquinas_do_ca.count()}")
    
    # Buscar máquinas primárias (descr_gerenc = "MÁQUINAS PRINCIPAL") relacionadas a este CA
    maquinas_primarias = maquinas_do_ca.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).order_by('cd_maquina')
    
    print(f"DEBUG: Máquinas primárias encontradas: {maquinas_primarias.count()}")
    if maquinas_primarias.exists():
        for mp in maquinas_primarias[:5]:  # Mostrar apenas as 5 primeiras para debug
            print(f"  - Máquina Primária: {mp.cd_maquina} - {mp.descr_maquina}")
    
    # Buscar relacionamentos entre máquinas primárias e secundárias para estas máquinas
    relacionamentos = MaquinaPrimariaSecundaria.objects.filter(
        maquina_primaria__in=maquinas_primarias
    ).select_related(
        'maquina_primaria', 'maquina_secundaria'
    ).order_by('maquina_primaria__cd_maquina', 'maquina_secundaria__cd_maquina')
    
    print(f"DEBUG: Relacionamentos encontrados: {relacionamentos.count()}")
    
    # Construir lista de nós no formato OrgChartJS
    nodes = []
    
    # Função auxiliar para construir URL da imagem
    def get_image_url(maquina):
        if maquina.foto:
            return request.build_absolute_uri(maquina.foto.url)
        return None
    
    # Adicionar máquinas primárias como nós raiz (sem pid)
    for maq_prim in maquinas_primarias:
        node_data = {
            'id': maq_prim.id,
            'field_0': maq_prim.descr_maquina or 'Sem descrição',
            'field_1': str(maq_prim.cd_maquina)
        }
        # Adicionar imagem se existir usando img_0
        foto_url = get_image_url(maq_prim)
        if foto_url:
            node_data['img_0'] = foto_url
        nodes.append(node_data)
    
    # Adicionar máquinas secundárias como nós filhos (com pid)
    for rel in relacionamentos:
        maq_sec = rel.maquina_secundaria
        node_data = {
            'id': maq_sec.id,
            'pid': rel.maquina_primaria.id,
            'field_0': maq_sec.descr_maquina or 'Sem descrição',
            'field_1': str(maq_sec.cd_maquina)
        }
        # Adicionar imagem se existir usando img_0
        foto_url = get_image_url(maq_sec)
        if foto_url:
            node_data['img_0'] = foto_url
        nodes.append(node_data)
    
    print(f"DEBUG: Total de nós criados: {len(nodes)}")
    if nodes:
        print(f"DEBUG: Primeiro nó: {nodes[0]}")
    
    # Serializar JSON
    try:
        dados_json_str = json.dumps(nodes, ensure_ascii=False, default=str)
    except Exception as e:
        print(f"Erro ao serializar JSON: {e}")
        dados_json_str = json.dumps([{
            'id': 0,
            'name': 'Erro ao processar dados'
        }], ensure_ascii=False)
    
    context = {
        'page_title': f'Visualizar CA {centro_atividade.ca}',
        'active_page': 'consultar_locais_e_cas',
        'ca': centro_atividade,
        'dados_json': dados_json_str,
        'total_primarias': maquinas_primarias.count(),
        'total_relacionamentos': relacionamentos.count(),
        'total_maquinas': maquinas_do_ca.count(),
        'has_maquinas': maquinas_do_ca.exists(),
        'has_primarias': maquinas_primarias.exists()
    }
    return render(request, 'visualizar/visualizar_centro_de_atividade.html', context)


def visualizar_local(request, ca_id):
    """Visualizar detalhes de um Centro de Atividade específico (anteriormente visualizar_local)"""
    from app.models import CentroAtividade, Maquina
    
    try:
        centro_atividade = CentroAtividade.objects.get(id=ca_id)
    except CentroAtividade.DoesNotExist:
        messages.error(request, 'Centro de Atividade não encontrado.')
        return redirect('consultar_locais_e_cas')
    
    # Buscar todas as máquinas relacionadas a este centro de atividade com classificação "MÁQUINAS PRINCIPAL"
    maquinas = Maquina.objects.filter(
        centro_atividade=centro_atividade,
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).order_by('cd_maquina')
    
    context = {
        'page_title': f'Visualizar CA {centro_atividade.ca} - {centro_atividade.local or "Sem local"}',
        'active_page': 'consultar_locais_e_cas',
        'centro_atividade': centro_atividade,
        'maquinas': maquinas,
    }
    return render(request, 'visualizar/visualizar_local.html', context)


def consultar_locais(request):
    """Consultar/listar Centros de Atividade com locais cadastrados"""
    from app.models import CentroAtividade
    from django.db.models import Q
    
    # Buscar todos os centros de atividade que têm local
    locais_list = CentroAtividade.objects.filter(local__isnull=False).exclude(local='')
    
    # Filtro de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Buscar em local, observações, CA e sigla do centro de atividade
        search_conditions = (
            Q(local__icontains=search_query) |
            Q(observacoes__icontains=search_query) |
            Q(ca__icontains=search_query) |
            Q(sigla__icontains=search_query) |
            Q(descricao__icontains=search_query)
        )
        locais_list = locais_list.filter(search_conditions)
    
    # Ordenar por local
    locais_list = locais_list.order_by('local')
    
    # Paginação
    paginator = Paginator(locais_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    locais = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = CentroAtividade.objects.filter(local__isnull=False).exclude(local='').count()
    centros_count = total_count
    
    context = {
        'page_title': 'Consultar Locais',
        'active_page': 'consultar_locais',
        'locais': locais,
        'total_count': total_count,
        'centros_count': centros_count,
    }
    return render(request, 'consultar/consultar_locais.html', context)


def editar_ca_e_locais(request, ca_id):
    """Editar Centro de Atividade (CA) existente com local"""
    from app.forms import CentroAtividadeForm
    from app.models import CentroAtividade
    import uuid
    
    try:
        centro_atividade = CentroAtividade.objects.get(id=ca_id)
    except CentroAtividade.DoesNotExist:
        messages.error(request, 'Centro de Atividade não encontrado.')
        return redirect('consultar_locais_e_cas')
    
    if request.method == 'POST':
        form = CentroAtividadeForm(request.POST, request.FILES, instance=centro_atividade)
        
        if form.is_valid():
            try:
                centro_atividade = form.save(commit=False)
                
                # Processar upload de imagem se fornecido
                # Se um arquivo foi enviado, ele tem prioridade sobre o campo de texto 'imagem'
                if 'imagem_upload' in request.FILES:
                    imagem_file = request.FILES['imagem_upload']
                    
                    # Gerar nome único para o arquivo
                    file_extension = os.path.splitext(imagem_file.name)[1]
                    unique_filename = f"ca_{centro_atividade.ca}_{uuid.uuid4().hex[:8]}{file_extension}"
                    
                    # Caminho completo para salvar
                    fotos_home_path = os.path.join(settings.STATICFILES_DIRS[0], 'fotos_home')
                    
                    # Criar diretório se não existir
                    os.makedirs(fotos_home_path, exist_ok=True)
                    
                    # Caminho completo do arquivo
                    file_path = os.path.join(fotos_home_path, unique_filename)
                    
                    # Salvar arquivo
                    with open(file_path, 'wb+') as destination:
                        for chunk in imagem_file.chunks():
                            destination.write(chunk)
                    
                    # Atualizar campo imagem com caminho relativo (sem 'static/')
                    # Normalizar: remover 'static/' do início se presente
                    imagem_path = f"fotos_home/{unique_filename}"
                    if imagem_path.startswith('static/'):
                        imagem_path = imagem_path.replace('static/', '', 1)
                    centro_atividade.imagem = imagem_path
                    messages.success(request, f'Foto do Centro de Atividade {centro_atividade.ca} atualizada com sucesso!')
                else:
                    # Se não houve upload, normalizar o caminho do campo 'imagem' se fornecido
                    if centro_atividade.imagem and centro_atividade.imagem.startswith('static/'):
                        centro_atividade.imagem = centro_atividade.imagem.replace('static/', '', 1)
                
                # Salvar todos os campos do form
                centro_atividade.save()
                messages.success(request, f'Centro de Atividade {centro_atividade.ca} atualizado com sucesso!')
                return redirect('consultar_locais_e_cas')
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                print(f"Erro ao atualizar Centro de Atividade: {error_detail}")
                messages.error(request, f'Erro ao atualizar Centro de Atividade: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = CentroAtividadeForm(instance=centro_atividade)
    
    context = {
        'page_title': f'Editar CA {centro_atividade.ca}',
        'active_page': 'consultar_locais_e_cas',
        'form': form,
        'centro_atividade': centro_atividade,
    }
    return render(request, 'editar/editar_centro_de_atividades.html', context)


def cadastrar_maquina(request):
    """Cadastrar nova máquina"""
    from app.forms import MaquinaForm
    
    if request.method == 'POST':
        form = MaquinaForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                maquina = form.save()
                messages.success(request, f'Máquina {maquina.cd_maquina} cadastrada com sucesso!')
                return redirect('consultar_maquinas')
            except Exception as e:
                messages.error(request, f'Erro ao cadastrar máquina: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = MaquinaForm()
    
    context = {
        'page_title': 'Cadastrar Máquina',
        'active_page': 'cadastrar_maquina',
        'form': form
    }
    return render(request, 'cadastrar/cadastrar_maquina.html', context)


def analise_maquinas(request):
    """Página de análise de máquinas com gráficos e estatísticas"""
    from app.models import Maquina, OrdemServicoCorretiva
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Estatísticas básicas
    total_count = Maquina.objects.count()
    setores_count = Maquina.objects.exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    unidades_count = Maquina.objects.exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
    
    # Máquinas por setor (cd_setormanut) - TODOS os setores
    maquinas_por_setor = Maquina.objects.exclude(
        cd_setormanut__isnull=True
    ).exclude(
        cd_setormanut=''
    ).values('cd_setormanut').annotate(
        total=Count('id')
    ).order_by('-total')  # Removido [:10] para mostrar todos
    
    setores_labels = [str(item['cd_setormanut']) for item in maquinas_por_setor]
    setores_data = [item['total'] for item in maquinas_por_setor]
    
    # Máquinas por unidade (top 10)
    maquinas_por_unidade = Maquina.objects.exclude(
        nome_unid__isnull=True
    ).exclude(
        nome_unid=''
    ).values('nome_unid').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    unidades_labels = [item['nome_unid'][:30] for item in maquinas_por_unidade]
    unidades_data = [item['total'] for item in maquinas_por_unidade]
    
    # Máquinas por mês (últimos 12 meses)
    maquinas_por_mes = defaultdict(int)
    maquinas = Maquina.objects.all().order_by('created_at')
    for maquina in maquinas:
        if maquina.created_at:
            mes_ano = maquina.created_at.strftime('%Y-%m')
            maquinas_por_mes[mes_ano] += 1
    
    # Ordenar por data e pegar últimos 12 meses
    meses_ordenados = sorted(maquinas_por_mes.keys())[-12:]
    meses_labels = [datetime.strptime(m, '%Y-%m').strftime('%b/%Y') for m in meses_ordenados]
    meses_data = [maquinas_por_mes[m] for m in meses_ordenados]
    
    # Distribuição por prioridade
    maquinas_por_prioridade = Maquina.objects.exclude(
        cd_priomaqutv__isnull=True
    ).values('cd_priomaqutv').annotate(
        total=Count('id')
    ).order_by('-cd_priomaqutv')[:10]
    
    prioridades_labels = [f"Prioridade {item['cd_priomaqutv']}" for item in maquinas_por_prioridade]
    prioridades_data = [item['total'] for item in maquinas_por_prioridade]
    
    # Distribuição por grupo
    maquinas_por_grupo = Maquina.objects.exclude(
        cd_grupo__isnull=True
    ).values('cd_grupo').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    grupos_labels = [f"Grupo {item['cd_grupo']}" for item in maquinas_por_grupo]
    grupos_data = [item['total'] for item in maquinas_por_grupo]
    
    # Distribuição por gerência
    maquinas_por_gerencia = Maquina.objects.exclude(
        descr_gerenc__isnull=True
    ).exclude(
        descr_gerenc=''
    ).values('descr_gerenc').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    gerencias_labels = [item['descr_gerenc'][:30] for item in maquinas_por_gerencia]
    gerencias_data = [item['total'] for item in maquinas_por_gerencia]
    
    # Máquinas com fotos
    maquinas_com_foto = Maquina.objects.exclude(foto__isnull=True).exclude(foto='').count()
    percentual_foto = (maquinas_com_foto / total_count * 100) if total_count > 0 else 0
    
    # Máquinas com placa
    maquinas_com_placa = Maquina.objects.exclude(placa_identificacao__isnull=True).exclude(placa_identificacao='').count()
    percentual_placa = (maquinas_com_placa / total_count * 100) if total_count > 0 else 0
    
    # Máquinas recentes (últimas 30 dias)
    data_30_dias_atras = datetime.now() - timedelta(days=30)
    maquinas_recentes = Maquina.objects.filter(created_at__gte=data_30_dias_atras).count()
    
    # Máquinas do mês atual
    mes_atual = datetime.now().replace(day=1)
    maquinas_mes_atual = Maquina.objects.filter(created_at__gte=mes_atual).count()
    
    # Top 10 máquinas com mais ordens de serviço
    top_maquinas_os = OrdemServicoCorretiva.objects.exclude(
        cd_maquina__isnull=True
    ).values('cd_maquina').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    # Buscar descrições das máquinas
    top_maquinas_os_list = []
    for item in top_maquinas_os:
        try:
            maquina = Maquina.objects.get(cd_maquina=item['cd_maquina'])
            top_maquinas_os_list.append({
                'cd_maquina': item['cd_maquina'],
                'descr_maquina': maquina.descr_maquina or 'Sem descrição',
                'total': item['total']
            })
        except Maquina.DoesNotExist:
            top_maquinas_os_list.append({
                'cd_maquina': item['cd_maquina'],
                'descr_maquina': 'Máquina não encontrada',
                'total': item['total']
            })
    
    maquinas_os_labels = [f"{item['cd_maquina']} - {item['descr_maquina'][:40]}" for item in top_maquinas_os_list]
    maquinas_os_data = [item['total'] for item in top_maquinas_os_list]
    
    # Máquinas com patrimônio
    maquinas_com_patrimonio = Maquina.objects.exclude(nro_patrimonio__isnull=True).exclude(nro_patrimonio='').count()
    percentual_patrimonio = (maquinas_com_patrimonio / total_count * 100) if total_count > 0 else 0
    
    # Estatísticas adicionais para novos cards
    # Máquinas ativas e inativas
    maquinas_ativas = Maquina.objects.filter(ativo=True).count()
    maquinas_inativas = Maquina.objects.filter(ativo=False).count()
    percentual_ativas = (maquinas_ativas / total_count * 100) if total_count > 0 else 0
    percentual_inativas = (maquinas_inativas / total_count * 100) if total_count > 0 else 0
    
    # Máquinas com código Aurora
    maquinas_com_codigo_aurora = Maquina.objects.exclude(codigo_aurora__isnull=True).exclude(codigo_aurora='').count()
    percentual_codigo_aurora = (maquinas_com_codigo_aurora / total_count * 100) if total_count > 0 else 0
    
    # Máquinas com código Fabricante
    maquinas_com_codigo_fabricante = Maquina.objects.exclude(codigo_fabricante__isnull=True).exclude(codigo_fabricante='').count()
    percentual_codigo_fabricante = (maquinas_com_codigo_fabricante / total_count * 100) if total_count > 0 else 0
    
    # Máquinas com documentos PDF
    maquinas_com_pdf = Maquina.objects.exclude(arquivo_pdf__isnull=True).exclude(arquivo_pdf='').count()
    percentual_pdf = (maquinas_com_pdf / total_count * 100) if total_count > 0 else 0
    
    # Máquinas principais (MÁQUINAS PRINCIPAL)
    maquinas_principais_count = Maquina.objects.filter(descr_gerenc__iexact='MÁQUINAS PRINCIPAL').count()
    percentual_principais = (maquinas_principais_count / total_count * 100) if total_count > 0 else 0
    
    # Máquinas com centro de atividade associado
    maquinas_com_centro_atividade = Maquina.objects.exclude(centro_atividade__isnull=True).count()
    percentual_com_centro = (maquinas_com_centro_atividade / total_count * 100) if total_count > 0 else 0
    
    # Total de grupos únicos
    grupos_unicos_count = Maquina.objects.exclude(cd_grupo__isnull=True).values('cd_grupo').distinct().count()
    
    # Total de prioridades únicas
    prioridades_unicas_count = Maquina.objects.exclude(cd_priomaqutv__isnull=True).values('cd_priomaqutv').distinct().count()
    
    # Máquinas "MÁQUINAS PRINCIPAL" agrupadas por cd_setormanut
    maquinas_principais = Maquina.objects.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).exclude(
        cd_setormanut__isnull=True
    ).exclude(
        cd_setormanut=''
    ).order_by('cd_setormanut', 'cd_maquina')
    
    # Agrupar máquinas por setor
    maquinas_por_setor_principal = {}
    for maquina in maquinas_principais:
        setor = str(maquina.cd_setormanut)
        if setor not in maquinas_por_setor_principal:
            maquinas_por_setor_principal[setor] = {
                'cd_setormanut': setor,
                'descr_setormanut': maquina.descr_setormanut or 'Sem descrição',
                'maquinas': []
            }
        maquinas_por_setor_principal[setor]['maquinas'].append({
            'id': maquina.id,
            'cd_maquina': maquina.cd_maquina,
            'descr_maquina': maquina.descr_maquina or 'Sem descrição',
            'nome_unid': maquina.nome_unid or '-',
            'nro_patrimonio': maquina.nro_patrimonio or '-'
        })
    
    # Converter para lista ordenada por cd_setormanut
    maquinas_principais_por_setor = sorted(
        maquinas_por_setor_principal.values(),
        key=lambda x: x['cd_setormanut']
    )
    
    # Buscar Centros de Atividade filtrados por local (INDÚSTRIA ou FRIGORÍFICO)
    from app.models import CentroAtividade
    centros_industria = list(CentroAtividade.objects.filter(
        local__iexact='INDÚSTRIA'
    ).distinct().order_by('ca'))
    
    centros_frigorifico = list(CentroAtividade.objects.filter(
        local__iexact='FRIGORÍFICO'
    ).distinct().order_by('ca'))
    
    # Preparar dados para o organograma (OrgChartJS)
    from app.models import MaquinaPrimariaSecundaria
    maquinas_primarias_org = Maquina.objects.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).order_by('cd_maquina')
    
    relacionamentos_org = MaquinaPrimariaSecundaria.objects.select_related(
        'maquina_primaria', 'maquina_secundaria'
    ).order_by('maquina_primaria__cd_maquina', 'maquina_secundaria__cd_maquina')
    
    # Construir lista de nós no formato do OrgChartJS
    nodes_org = []
    
    # Adicionar máquinas primárias como nós raiz (sem pid)
    for maq_prim in maquinas_primarias_org:
        nodes_org.append({
            'id': maq_prim.id,
            'name': f"{maq_prim.cd_maquina} - {maq_prim.descr_maquina or 'Sem descrição'}"
        })
    
    # Adicionar máquinas secundárias como nós filhos (com pid)
    for rel in relacionamentos_org:
        maq_sec = rel.maquina_secundaria
        nodes_org.append({
            'id': maq_sec.id,
            'pid': rel.maquina_primaria.id,
            'name': f"{maq_sec.cd_maquina} - {maq_sec.descr_maquina or 'Sem descrição'}"
        })
    
    # Serializar JSON para o organograma
    try:
        dados_json_org = json.dumps(nodes_org, ensure_ascii=False, default=str)
    except Exception as e:
        print(f"Erro ao serializar JSON do organograma: {e}")
        dados_json_org = json.dumps([{
            'id': 0,
            'name': 'Erro ao processar dados'
        }], ensure_ascii=False)
    
    context = {
        'page_title': 'Análise de Máquinas',
        'active_page': 'analise_maquinas',
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_recentes': maquinas_recentes,
        'maquinas_mes_atual': maquinas_mes_atual,
        'maquinas_com_foto': maquinas_com_foto,
        'percentual_foto': round(percentual_foto, 1),
        'maquinas_com_placa': maquinas_com_placa,
        'percentual_placa': round(percentual_placa, 1),
        'maquinas_com_patrimonio': maquinas_com_patrimonio,
        'percentual_patrimonio': round(percentual_patrimonio, 1),
        # Estatísticas adicionais
        'maquinas_ativas': maquinas_ativas,
        'maquinas_inativas': maquinas_inativas,
        'percentual_ativas': round(percentual_ativas, 1),
        'percentual_inativas': round(percentual_inativas, 1),
        'maquinas_com_codigo_aurora': maquinas_com_codigo_aurora,
        'percentual_codigo_aurora': round(percentual_codigo_aurora, 1),
        'maquinas_com_codigo_fabricante': maquinas_com_codigo_fabricante,
        'percentual_codigo_fabricante': round(percentual_codigo_fabricante, 1),
        'maquinas_com_pdf': maquinas_com_pdf,
        'percentual_pdf': round(percentual_pdf, 1),
        'maquinas_principais_count': maquinas_principais_count,
        'percentual_principais': round(percentual_principais, 1),
        'maquinas_com_centro_atividade': maquinas_com_centro_atividade,
        'percentual_com_centro': round(percentual_com_centro, 1),
        'grupos_unicos_count': grupos_unicos_count,
        'prioridades_unicas_count': prioridades_unicas_count,
        # Dados para gráficos (JSON)
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'prioridades_labels': json.dumps(prioridades_labels),
        'prioridades_data': json.dumps(prioridades_data),
        'grupos_labels': json.dumps(grupos_labels),
        'grupos_data': json.dumps(grupos_data),
        'gerencias_labels': json.dumps(gerencias_labels),
        'gerencias_data': json.dumps(gerencias_data),
        'maquinas_os_labels': json.dumps(maquinas_os_labels),
        'maquinas_os_data': json.dumps(maquinas_os_data),
        # Dados para tabelas
        'top_maquinas_os': top_maquinas_os_list,
        'maquinas_por_setor': maquinas_por_setor,
        'maquinas_por_unidade': maquinas_por_unidade,
        'maquinas_principais_por_setor': maquinas_principais_por_setor,
        'centros_industria': centros_industria,
        'centros_frigorifico': centros_frigorifico,
        # Dados para organograma
        'dados_json_org': dados_json_org,
        'total_primarias_org': maquinas_primarias_org.count(),
        'total_relacionamentos_org': relacionamentos_org.count(),
    }
    return render(request, 'maquinas/analise_geral_maquinas.html', context)


def analise_maquinas_importadas(request):
    """Página de análise de máquinas importadas com gráficos e estatísticas"""
    from app.models import Maquina
    from django.db.models import Count
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Estatísticas básicas - apenas máquinas importadas (com created_at) e ativas
    total_importadas = Maquina.objects.filter(ativo=True).exclude(created_at__isnull=True).count()
    total_geral = Maquina.objects.filter(ativo=True).count()
    percentual_importadas = (total_importadas / total_geral * 100) if total_geral > 0 else 0
    
    # Máquinas importadas recentes (últimas 30 dias) - apenas ativas
    data_30_dias_atras = datetime.now() - timedelta(days=30)
    importadas_recentes = Maquina.objects.filter(
        ativo=True,
        created_at__gte=data_30_dias_atras
    ).exclude(created_at__isnull=True).count()
    
    # Máquinas importadas do mês atual - apenas ativas
    mes_atual = datetime.now().replace(day=1)
    importadas_mes_atual = Maquina.objects.filter(
        ativo=True,
        created_at__gte=mes_atual
    ).exclude(created_at__isnull=True).count()
    
    # Máquinas importadas por setor - apenas ativas
    maquinas_importadas_por_setor = Maquina.objects.filter(
        ativo=True
    ).exclude(
        created_at__isnull=True
    ).exclude(
        cd_setormanut__isnull=True
    ).exclude(
        cd_setormanut=''
    ).values('cd_setormanut').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    setores_labels = [str(item['cd_setormanut']) for item in maquinas_importadas_por_setor]
    setores_data = [item['total'] for item in maquinas_importadas_por_setor]
    
    # Máquinas importadas por unidade (top 10) - apenas ativas
    maquinas_importadas_por_unidade = Maquina.objects.filter(
        ativo=True
    ).exclude(
        created_at__isnull=True
    ).exclude(
        nome_unid__isnull=True
    ).exclude(
        nome_unid=''
    ).values('nome_unid').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    unidades_labels = [item['nome_unid'][:30] for item in maquinas_importadas_por_unidade]
    unidades_data = [item['total'] for item in maquinas_importadas_por_unidade]
    
    # Máquinas importadas por mês (últimos 12 meses) - apenas ativas
    maquinas_importadas_por_mes = defaultdict(int)
    maquinas_importadas = Maquina.objects.filter(
        ativo=True
    ).exclude(created_at__isnull=True).order_by('created_at')
    for maquina in maquinas_importadas:
        if maquina.created_at:
            mes_ano = maquina.created_at.strftime('%Y-%m')
            maquinas_importadas_por_mes[mes_ano] += 1
    
    # Ordenar por data e pegar últimos 12 meses
    meses_ordenados = sorted(maquinas_importadas_por_mes.keys())[-12:]
    meses_labels = [datetime.strptime(m, '%Y-%m').strftime('%b/%Y') for m in meses_ordenados]
    meses_data = [maquinas_importadas_por_mes[m] for m in meses_ordenados]
    
    # Máquinas importadas recentes para exibir na tabela - apenas ativas
    maquinas_importadas_recentes = Maquina.objects.filter(
        ativo=True
    ).exclude(
        created_at__isnull=True
    ).order_by('-created_at')[:50]
    
    # Agrupar máquinas importadas por descr_gerenc - apenas ativas
    from collections import defaultdict
    maquinas_por_gerencia = defaultdict(list)
    maquinas_importadas_gerencia = Maquina.objects.filter(
        ativo=True
    ).exclude(
        created_at__isnull=True
    ).exclude(
        descr_gerenc__isnull=True
    ).exclude(
        descr_gerenc=''
    ).order_by('descr_gerenc', 'cd_maquina')
    
    for maquina in maquinas_importadas_gerencia:
        gerencia = maquina.descr_gerenc or 'Sem Classificação'
        maquinas_por_gerencia[gerencia].append(maquina)
    
    # Converter para lista ordenada por nome da gerência
    maquinas_por_gerencia_list = sorted(
        maquinas_por_gerencia.items(),
        key=lambda x: x[0]
    )
    
    context = {
        'page_title': 'Análise de Máquinas Importadas',
        'active_page': 'analise_maquinas_importadas',
        'total_importadas': total_importadas,
        'importadas_recentes': importadas_recentes,
        'importadas_mes_atual': importadas_mes_atual,
        'percentual_importadas': round(percentual_importadas, 1),
        'maquinas_importadas_recentes': maquinas_importadas_recentes,
        'maquinas_por_gerencia': maquinas_por_gerencia_list,
        # Dados para gráficos (JSON)
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
    }
    return render(request, 'maquinas/analise_maquinas_importadas.html', context)


def consultar_maquinas(request):
    """Consultar/listar máquinas cadastradas"""
    from app.models import Maquina
    
    # Verificar se deve incluir máquinas inativas
    include_inativas = request.GET.get('include_inativas', 'off') == 'on'
    
    # Buscar máquinas (apenas ativas por padrão)
    if include_inativas:
        maquinas_list = Maquina.objects.all()
    else:
        maquinas_list = Maquina.objects.filter(ativo=True)
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_maquina=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(cd_setormanut__icontains=search_query) |
            Q(descr_setormanut__icontains=search_query) |
            Q(nome_unid__icontains=search_query) |
            Q(nro_patrimonio__icontains=search_query)
        )
        
        maquinas_list = maquinas_list.filter(search_conditions)
    
    # Filtros por coluna individual
    filter_codigo = request.GET.get('filter_codigo', '').strip()
    if filter_codigo:
        try:
            codigo_num = int(float(filter_codigo))
            maquinas_list = maquinas_list.filter(cd_maquina=codigo_num)
        except (ValueError, TypeError):
            maquinas_list = maquinas_list.filter(cd_maquina__icontains=filter_codigo)
    
    filter_descricao = request.GET.get('filter_descricao', '').strip()
    if filter_descricao:
        maquinas_list = maquinas_list.filter(descr_maquina__icontains=filter_descricao)
    
    filter_setor = request.GET.get('filter_setor', '').strip()
    if filter_setor:
        # Se for um valor exato (da lista de opções), usar busca exata
        # Caso contrário, usar busca parcial
        maquinas_list = maquinas_list.filter(descr_setormanut=filter_setor)
    
    filter_unidade = request.GET.get('filter_unidade', '').strip()
    if filter_unidade:
        maquinas_list = maquinas_list.filter(nome_unid__icontains=filter_unidade)
    
    filter_patrimonio = request.GET.get('filter_patrimonio', '').strip()
    if filter_patrimonio:
        maquinas_list = maquinas_list.filter(nro_patrimonio__icontains=filter_patrimonio)
    
    filter_prioridade = request.GET.get('filter_prioridade', '').strip()
    if filter_prioridade:
        try:
            prioridade_num = int(float(filter_prioridade))
            maquinas_list = maquinas_list.filter(cd_priomaqutv=prioridade_num)
        except (ValueError, TypeError):
            maquinas_list = maquinas_list.filter(cd_priomaqutv__icontains=filter_prioridade)
    
    filter_gerenc = request.GET.get('filter_gerenc', '').strip()
    if filter_gerenc:
        maquinas_list = maquinas_list.filter(descr_gerenc=filter_gerenc)
    
    # Ordenar por código da máquina
    maquinas_list = maquinas_list.order_by('cd_maquina')
    
    # Paginação
    paginator = Paginator(maquinas_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    maquinas = paginator.get_page(page_number)
    
    # Estatísticas (usar mesmo filtro de ativo/inativo)
    if include_inativas:
        total_count = Maquina.objects.count()
        setores_count = Maquina.objects.exclude(descr_setormanut__isnull=True).exclude(descr_setormanut='').values('descr_setormanut').distinct().count()
        unidades_count = Maquina.objects.exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
        maquinas_principais_count = Maquina.objects.filter(descr_gerenc__iexact='MÁQUINAS PRINCIPAL').count()
        gerenc_choices = Maquina.objects.exclude(descr_gerenc__isnull=True).exclude(descr_gerenc='').values_list('descr_gerenc', flat=True).distinct().order_by('descr_gerenc')
    else:
        total_count = Maquina.objects.filter(ativo=True).count()
        setores_count = Maquina.objects.filter(ativo=True).exclude(descr_setormanut__isnull=True).exclude(descr_setormanut='').values('descr_setormanut').distinct().count()
        unidades_count = Maquina.objects.filter(ativo=True).exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
        maquinas_principais_count = Maquina.objects.filter(ativo=True, descr_gerenc__iexact='MÁQUINAS PRINCIPAL').count()
        gerenc_choices = Maquina.objects.filter(ativo=True).exclude(descr_gerenc__isnull=True).exclude(descr_gerenc='').values_list('descr_gerenc', flat=True).distinct().order_by('descr_gerenc')
    
    # Obter valores distintos de descr_setormanut para o select
    if include_inativas:
        setor_choices = Maquina.objects.exclude(descr_setormanut__isnull=True).exclude(descr_setormanut='').values_list('descr_setormanut', flat=True).distinct().order_by('descr_setormanut')
    else:
        setor_choices = Maquina.objects.filter(ativo=True).exclude(descr_setormanut__isnull=True).exclude(descr_setormanut='').values_list('descr_setormanut', flat=True).distinct().order_by('descr_setormanut')
    
    context = {
        'page_title': 'Consultar Máquinas',
        'active_page': 'consultar_maquinas',
        'maquinas': maquinas,
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_principais_count': maquinas_principais_count,
        'gerenc_choices': gerenc_choices,
        'setor_choices': setor_choices,
        'include_inativas': include_inativas,
        # Preservar filtros no contexto
        'filter_codigo': filter_codigo,
        'filter_descricao': filter_descricao,
        'filter_setor': filter_setor,
        'filter_unidade': filter_unidade,
        'filter_patrimonio': filter_patrimonio,
        'filter_prioridade': filter_prioridade,
        'filter_gerenc': filter_gerenc,
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_maquinas.html', context)


def consultar_manutencoes_preventivas(request):
    """Consultar/listar ordens de serviço preventivas cadastradas com filtros avançados"""
    from app.models import OrdemServicoPreventiva
    from datetime import datetime
    
    # Buscar todas as ordens preventivas
    ordens_list = OrdemServicoPreventiva.objects.all()
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_ordemserv=search_num)
            search_conditions |= Q(cd_maquina=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(cd_setormanut__icontains=search_query) |
            Q(descr_setormanut__icontains=search_query) |
            Q(nm_func_solic_os__icontains=search_query) |
            Q(nm_func_exec__icontains=search_query) |
            Q(descr_queixa__icontains=search_query) |
            Q(exec_tarefas__icontains=search_query)
        )
        
        ordens_list = ordens_list.filter(search_conditions)
    
    # Filtros específicos
    # Filtro por Setor de Manutenção
    filtro_setor = request.GET.get('filtro_setor', '')
    if filtro_setor:
        ordens_list = ordens_list.filter(descr_setormanut__icontains=filtro_setor)
    
    # Filtro por Unidade
    filtro_unidade = request.GET.get('filtro_unidade', '')
    if filtro_unidade:
        ordens_list = ordens_list.filter(nome_unid__icontains=filtro_unidade)
    
    # Filtro por Tipo de Ordem de Serviço
    filtro_tipo_os = request.GET.get('filtro_tipo_os', '')
    if filtro_tipo_os:
        ordens_list = ordens_list.filter(descr_tpordservtv__icontains=filtro_tipo_os)
    
    # Filtro por Situação da Ordem
    filtro_situacao = request.GET.get('filtro_situacao', '')
    if filtro_situacao:
        ordens_list = ordens_list.filter(descr_sitordsetv__icontains=filtro_situacao)
    
    # Filtro por Funcionário Solicitante
    filtro_solicitante = request.GET.get('filtro_solicitante', '')
    if filtro_solicitante:
        ordens_list = ordens_list.filter(nm_func_solic_os__icontains=filtro_solicitante)
    
    # Filtro por Funcionário Executor
    filtro_executor = request.GET.get('filtro_executor', '')
    if filtro_executor:
        ordens_list = ordens_list.filter(nm_func_exec__icontains=filtro_executor)
    
    # Filtro por Código da Máquina
    filtro_maquina = request.GET.get('filtro_maquina', '')
    if filtro_maquina:
        ordens_list = ordens_list.filter(cd_maquina__icontains=filtro_maquina)
    
    # Filtro por Data de Entrada (período)
    data_entrada_inicio = request.GET.get('data_entrada_inicio', '')
    data_entrada_fim = request.GET.get('data_entrada_fim', '')
    if data_entrada_inicio:
        try:
            data_inicio = datetime.strptime(data_entrada_inicio, '%Y-%m-%d')
            ordens_list = ordens_list.filter(created_at__gte=data_inicio)
        except ValueError:
            pass
    if data_entrada_fim:
        try:
            data_fim = datetime.strptime(data_entrada_fim, '%Y-%m-%d')
            # Adicionar 1 dia para incluir o dia final
            from datetime import timedelta
            data_fim = data_fim + timedelta(days=1)
            ordens_list = ordens_list.filter(created_at__lte=data_fim)
        except ValueError:
            pass
    
    # Filtro por Status da Ordem (Abertas/Fechadas)
    filtro_ordens_abertas = request.GET.get('filtro_ordens_abertas', '')
    filtro_ordens_fechadas = request.GET.get('filtro_ordens_fechadas', '')
    
    # Converter para boolean (se existe e não é vazio, é True)
    filtro_ordens_abertas = filtro_ordens_abertas == '1'
    filtro_ordens_fechadas = filtro_ordens_fechadas == '1'
    
    # Aplicar filtros baseado nos checkboxes marcados
    if filtro_ordens_abertas and filtro_ordens_fechadas:
        # Ambos marcados: mostrar todas (não aplicar filtro)
        pass
    elif filtro_ordens_abertas and not filtro_ordens_fechadas:
        # Apenas "Ordens Abertas" marcado: dt_encordmanu está vazio ou nulo
        ordens_list = ordens_list.filter(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    elif filtro_ordens_fechadas and not filtro_ordens_abertas:
        # Apenas "Ordens Fechadas" marcado: dt_encordmanu tem valor (não é nulo nem vazio)
        ordens_list = ordens_list.exclude(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    # Se nenhum está marcado, mostra todas (não aplicar filtro)
    
    # Ordenar por código da ordem de serviço (mais recente primeiro)
    ordens_list = ordens_list.order_by('-cd_ordemserv')
    
    # Paginação
    paginator = Paginator(ordens_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    ordens = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = OrdemServicoPreventiva.objects.count()
    setores_count = OrdemServicoPreventiva.objects.exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    unidades_count = OrdemServicoPreventiva.objects.exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros
    setores_unicos = OrdemServicoPreventiva.objects.exclude(
        descr_setormanut__isnull=True
    ).exclude(
        descr_setormanut=''
    ).values_list('descr_setormanut', flat=True).distinct().order_by('descr_setormanut')
    
    unidades_unicas = OrdemServicoPreventiva.objects.exclude(
        nome_unid__isnull=True
    ).exclude(
        nome_unid=''
    ).values_list('nome_unid', flat=True).distinct().order_by('nome_unid')
    
    tipos_os_unicos = OrdemServicoPreventiva.objects.exclude(
        descr_tpordservtv__isnull=True
    ).exclude(
        descr_tpordservtv=''
    ).values_list('descr_tpordservtv', flat=True).distinct().order_by('descr_tpordservtv')
    
    situacoes_unicas = OrdemServicoPreventiva.objects.exclude(
        descr_sitordsetv__isnull=True
    ).exclude(
        descr_sitordsetv=''
    ).values_list('descr_sitordsetv', flat=True).distinct().order_by('descr_sitordsetv')
    
    context = {
        'page_title': 'Consultar Ordens Preventivas',
        'active_page': 'consultar_manutencoes_preventivas',
        'ordens': ordens,
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        # Valores para dropdowns
        'setores_unicos': setores_unicos,
        'unidades_unicas': unidades_unicas,
        'tipos_os_unicos': tipos_os_unicos,
        'situacoes_unicas': situacoes_unicas,
        # Valores dos filtros ativos
        'filtro_setor': filtro_setor,
        'filtro_unidade': filtro_unidade,
        'filtro_tipo_os': filtro_tipo_os,
        'filtro_situacao': filtro_situacao,
        'filtro_solicitante': filtro_solicitante,
        'filtro_executor': filtro_executor,
        'filtro_maquina': filtro_maquina,
        'data_entrada_inicio': data_entrada_inicio,
        'data_entrada_fim': data_entrada_fim,
        'filtro_ordens_abertas': '1' if filtro_ordens_abertas else '',
        'filtro_ordens_fechadas': '1' if filtro_ordens_fechadas else '',
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_manutencoes_preventivas.html', context)


def consultar_meu_plano(request):
    """Consultar/listar Meus Planos Preventiva (MeuPlanoPreventiva)"""
    from app.models import MeuPlanoPreventiva
    
    # Buscar todos os meus planos preventiva
    planos_list = MeuPlanoPreventiva.objects.all().select_related('maquina', 'roteiro_preventiva')
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_maquina=search_num) | Q(numero_plano=search_num) | Q(sequencia_manutencao=search_num) | Q(sequencia_tarefa=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(descr_tarefa__icontains=search_query) |
            Q(nome_funcionario__icontains=search_query) |
            Q(cd_funcionario__icontains=search_query) |
            Q(cd_setor__icontains=search_query) |
            Q(descr_setor__icontains=search_query) |
            Q(nome_unid__icontains=search_query) |
            Q(descr_plano__icontains=search_query) |
            Q(desc_detalhada_do_roteiro_preventiva__icontains=search_query) |
            Q(descr_seqplamanu__icontains=search_query)
        )
        
        planos_list = planos_list.filter(search_conditions)
    
    # Filtros por coluna individual
    filter_maquina = request.GET.get('filter_maquina', '').strip()
    if filter_maquina:
        try:
            maquina_num = int(float(filter_maquina))
            planos_list = planos_list.filter(cd_maquina=maquina_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(
                Q(cd_maquina__icontains=filter_maquina) |
                Q(descr_maquina__icontains=filter_maquina)
            )
    
    filter_plano = request.GET.get('filter_plano', '').strip()
    if filter_plano:
        try:
            plano_num = int(float(filter_plano))
            planos_list = planos_list.filter(numero_plano=plano_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(
                Q(numero_plano__icontains=filter_plano) |
                Q(descr_plano__icontains=filter_plano)
            )
    
    filter_seq_manutencao = request.GET.get('filter_seq_manutencao', '').strip()
    if filter_seq_manutencao:
        try:
            seq_num = int(float(filter_seq_manutencao))
            planos_list = planos_list.filter(sequencia_manutencao=seq_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(sequencia_manutencao__icontains=filter_seq_manutencao)
    
    filter_data = request.GET.get('filter_data', '').strip()
    if filter_data:
        planos_list = planos_list.filter(dt_execucao__icontains=filter_data)
    
    filter_periodo = request.GET.get('filter_periodo', '').strip()
    if filter_periodo:
        try:
            periodo_num = int(float(filter_periodo))
            planos_list = planos_list.filter(quantidade_periodo=periodo_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(quantidade_periodo__icontains=filter_periodo)
    
    filter_seq_tarefa = request.GET.get('filter_seq_tarefa', '').strip()
    if filter_seq_tarefa:
        try:
            seq_tarefa_num = int(float(filter_seq_tarefa))
            planos_list = planos_list.filter(sequencia_tarefa=seq_tarefa_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(sequencia_tarefa__icontains=filter_seq_tarefa)
    
    filter_tarefa = request.GET.get('filter_tarefa', '').strip()
    if filter_tarefa:
        planos_list = planos_list.filter(descr_tarefa__icontains=filter_tarefa)
    
    filter_desc_detalhada = request.GET.get('filter_desc_detalhada', '').strip()
    if filter_desc_detalhada:
        planos_list = planos_list.filter(desc_detalhada_do_roteiro_preventiva__icontains=filter_desc_detalhada)
    
    filter_descr_seqplamanu = request.GET.get('filter_descr_seqplamanu', '').strip()
    if filter_descr_seqplamanu:
        planos_list = planos_list.filter(descr_seqplamanu__icontains=filter_descr_seqplamanu)
    
    filter_funcionario = request.GET.get('filter_funcionario', '').strip()
    if filter_funcionario:
        planos_list = planos_list.filter(
            Q(nome_funcionario__icontains=filter_funcionario) |
            Q(cd_funcionario__icontains=filter_funcionario)
        )
    
    filter_setor = request.GET.get('filter_setor', '').strip()
    if filter_setor:
        planos_list = planos_list.filter(
            Q(cd_setor__icontains=filter_setor) |
            Q(descr_setor__icontains=filter_setor)
        )
    
    filter_unidade = request.GET.get('filter_unidade', '').strip()
    if filter_unidade:
        try:
            unidade_num = int(float(filter_unidade))
            planos_list = planos_list.filter(cd_unid=unidade_num)
        except (ValueError, TypeError):
            planos_list = planos_list.filter(
                Q(nome_unid__icontains=filter_unidade)
            )
    
    # Ordenar por máquina, plano, sequência manutenção e sequência tarefa
    planos_list = planos_list.order_by('cd_maquina', 'numero_plano', 'sequencia_manutencao', 'sequencia_tarefa')
    
    # Paginação
    paginator = Paginator(planos_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    planos = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = MeuPlanoPreventiva.objects.count()
    maquinas_count = MeuPlanoPreventiva.objects.exclude(cd_maquina__isnull=True).values('cd_maquina').distinct().count()
    setores_count = MeuPlanoPreventiva.objects.exclude(cd_setor__isnull=True).exclude(cd_setor='').values('cd_setor').distinct().count()
    planos_com_roteiro = MeuPlanoPreventiva.objects.exclude(roteiro_preventiva__isnull=True).count()
    
    context = {
        'page_title': 'Consultar Meus Planos Preventiva',
        'active_page': 'consultar_meu_plano',
        'planos': planos,
        'total_count': total_count,
        'maquinas_count': maquinas_count,
        'setores_count': setores_count,
        'planos_com_roteiro': planos_com_roteiro,
        # Preservar filtros no contexto
        'search_query': search_query,
        'filter_maquina': filter_maquina,
        'filter_plano': filter_plano,
        'filter_seq_manutencao': filter_seq_manutencao,
        'filter_data': filter_data,
        'filter_periodo': filter_periodo,
        'filter_seq_tarefa': filter_seq_tarefa,
        'filter_tarefa': filter_tarefa,
        'filter_desc_detalhada': filter_desc_detalhada,
        'filter_descr_seqplamanu': filter_descr_seqplamanu,
        'filter_funcionario': filter_funcionario,
        'filter_setor': filter_setor,
    }
    return render(request, 'consultar/consultar_meu_plano.html', context)


def consultar_requisicoes_almoxarifado(request):
    """Consultar/listar requisições de almoxarifado com filtros avançados"""
    from app.models import RequisicaoAlmoxarifado
    from decimal import Decimal
    from datetime import datetime
    
    # Função auxiliar para parse de datas
    def parse_date(date_str):
        """Tenta fazer parse de data em vários formatos"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        if not date_str:
            return None
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses (None)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Se não há meses selecionados, usar todos os meses
    meses_para_mostrar = meses_filtro_int if meses_filtro_int else list(range(1, 13))
    
    # Busca geral
    search_query = request.GET.get('search', '').strip()
    todas_requisicoes = RequisicaoAlmoxarifado.objects.all()
    
    # Obter anos disponíveis (baseado em data_requisicao)
    anos_disponiveis = []
    for requisicao in todas_requisicoes:
        if requisicao.data_requisicao:
            anos_disponiveis.append(requisicao.data_requisicao.year)
    anos_disponiveis = sorted(list(set(anos_disponiveis)), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    if hoje.year not in anos_disponiveis:
        anos_disponiveis.insert(0, hoje.year)
    
    # Filtrar por ano e mês (data_requisicao)
    requisicoes_filtradas_por_data = []
    for requisicao in todas_requisicoes:
        if requisicao.data_requisicao and requisicao.data_requisicao.year == ano_filtro:
            if not meses_filtro_int or requisicao.data_requisicao.month in meses_filtro_int:
                requisicoes_filtradas_por_data.append(requisicao.id)
    
    if requisicoes_filtradas_por_data:
        requisicoes_list = todas_requisicoes.filter(id__in=requisicoes_filtradas_por_data)
    else:
        # Se não há requisições no ano/mês, retornar queryset vazio
        requisicoes_list = RequisicaoAlmoxarifado.objects.none()
    
    # Aplicar busca geral
    if search_query:
        requisicoes_list = requisicoes_list.filter(
            Q(cd_item__icontains=search_query) |
            Q(descr_item__icontains=search_query) |
            Q(cd_centro_ativ__icontains=search_query) |
            Q(cd_usu_criou__icontains=search_query) |
            Q(cd_usu_atend__icontains=search_query) |
            Q(descr_operacao__icontains=search_query) |
            Q(descr_local_fisic__icontains=search_query)
        )
    
    # Filtros específicos (manter compatibilidade com filtros antigos)
    filter_data = request.GET.get('filter_data', '').strip()
    if filter_data:
        try:
            data_obj = datetime.strptime(filter_data, '%Y-%m-%d').date()
            requisicoes_list = requisicoes_list.filter(data_requisicao=data_obj)
        except ValueError:
            requisicoes_list = requisicoes_list.filter(data_requisicao__icontains=filter_data)
    
    filter_item = request.GET.get('filter_item', '').strip()
    if filter_item:
        try:
            item_num = int(float(filter_item))
            requisicoes_list = requisicoes_list.filter(cd_item=item_num)
        except (ValueError, TypeError):
            requisicoes_list = requisicoes_list.filter(cd_item__icontains=filter_item)
    
    filter_descricao = request.GET.get('filter_descricao', '').strip()
    if filter_descricao:
        requisicoes_list = requisicoes_list.filter(descr_item__icontains=filter_descricao)
    
    filter_quantidade = request.GET.get('filter_quantidade', '').strip()
    if filter_quantidade:
        try:
            qtd_num = Decimal(filter_quantidade)
            requisicoes_list = requisicoes_list.filter(qtde_movto_estoq=qtd_num)
        except (ValueError, TypeError):
            requisicoes_list = requisicoes_list.filter(qtde_movto_estoq__icontains=filter_quantidade)
    
    filter_valor = request.GET.get('filter_valor', '').strip()
    if filter_valor:
        try:
            valor_num = Decimal(filter_valor)
            requisicoes_list = requisicoes_list.filter(vlr_movto_estoq=valor_num)
        except (ValueError, TypeError):
            requisicoes_list = requisicoes_list.filter(vlr_movto_estoq__icontains=filter_valor)
    
    filter_centro = request.GET.get('filter_centro', '').strip()
    if filter_centro:
        try:
            centro_num = int(float(filter_centro))
            requisicoes_list = requisicoes_list.filter(cd_centro_ativ=centro_num)
        except (ValueError, TypeError):
            requisicoes_list = requisicoes_list.filter(cd_centro_ativ__icontains=filter_centro)
    
    filter_usuario_criou = request.GET.get('filter_usuario_criou', '').strip()
    if filter_usuario_criou:
        requisicoes_list = requisicoes_list.filter(cd_usu_criou__icontains=filter_usuario_criou)
    
    filter_usuario_atend = request.GET.get('filter_usuario_atend', '').strip()
    if filter_usuario_atend:
        requisicoes_list = requisicoes_list.filter(cd_usu_atend__icontains=filter_usuario_atend)
    
    # Opções para filtro de Operação (distintos no dataset já filtrado por ano/mês e outros)
    operacoes_disponiveis = list(
        requisicoes_list.exclude(descr_operacao__isnull=True).exclude(descr_operacao='')
        .values_list('descr_operacao', flat=True).distinct().order_by('descr_operacao')
    )
    has_operacao_vazio = requisicoes_list.filter(Q(descr_operacao__isnull=True) | Q(descr_operacao='')).exists()
    filter_operacao = request.GET.get('filter_operacao', '').strip()
    if filter_operacao:
        if filter_operacao == '__VAZIO__':
            requisicoes_list = requisicoes_list.filter(Q(descr_operacao__isnull=True) | Q(descr_operacao=''))
        else:
            requisicoes_list = requisicoes_list.filter(descr_operacao=filter_operacao)
    
    filter_local = request.GET.get('filter_local', '').strip()
    if filter_local:
        requisicoes_list = requisicoes_list.filter(descr_local_fisic__icontains=filter_local)
    
    # Ordenação: Valor Total (asc/desc) ou padrão (data + item)
    sort_valor = request.GET.get('sort_valor', '').strip().lower()
    sort_valor = sort_valor if sort_valor in ('asc', 'desc') else ''
    if sort_valor == 'asc':
        from django.db.models import F
        requisicoes_list = requisicoes_list.order_by(F('vlr_movto_estoq').asc(nulls_last=True), '-data_requisicao', 'cd_item')
    elif sort_valor == 'desc':
        from django.db.models import F
        requisicoes_list = requisicoes_list.order_by(F('vlr_movto_estoq').desc(nulls_last=True), '-data_requisicao', 'cd_item')
    else:
        requisicoes_list = requisicoes_list.order_by('-data_requisicao', 'cd_item')
    
    # Estatísticas baseadas nos dados FILTRADOS (antes da paginação)
    total_count = requisicoes_list.count()
    itens_count = requisicoes_list.values('cd_item').distinct().count()
    centros_count = requisicoes_list.exclude(cd_centro_ativ__isnull=True).values('cd_centro_ativ').distinct().count()
    
    # Calcular valor total dos dados filtrados
    # vlr_movto_estoq já representa o valor total da linha (não é valor unitário)
    # IMPORTANTE: Apenas considerar cd_depo == 1 para custos (itens novos que geram gasto)
    # cd_depo == 3 são itens reutilizados que não geram custo
    valor_total = Decimal('0.00')
    for req in requisicoes_list:
        if req.vlr_movto_estoq and req.cd_depo == 1:
            valor_total += abs(req.vlr_movto_estoq)
    
    # Paginação (após calcular estatísticas)
    paginator = Paginator(requisicoes_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    requisicoes = paginator.get_page(page_number)
    
    # Meses para o template
    meses_choices = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]
    
    context = {
        'page_title': 'Consultar Requisições Almoxarifado',
        'active_page': 'consultar_requisicoes_almoxarifado',
        'requisicoes': requisicoes,
        'total_count': total_count,
        'itens_count': itens_count,
        'centros_count': centros_count,
        'valor_total': valor_total,
        # Filtros de data
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'meses_para_mostrar': meses_para_mostrar,
        'anos_disponiveis': anos_disponiveis,
        'meses_choices': meses_choices,
        # Preservar filtros no contexto
        'search_query': search_query,
        'filter_data': filter_data,
        'filter_item': filter_item,
        'filter_descricao': filter_descricao,
        'filter_quantidade': filter_quantidade,
        'filter_valor': filter_valor,
        'filter_centro': filter_centro,
        'filter_usuario_criou': filter_usuario_criou,
        'filter_usuario_atend': filter_usuario_atend,
        'filter_operacao': filter_operacao,
        'filter_local': filter_local,
        'sort_valor': sort_valor,
        'operacoes_disponiveis': operacoes_disponiveis,
        'has_operacao_vazio': has_operacao_vazio,
        'query_string_base': _build_query_string_base(request, exclude=['sort_valor', 'page']),
    }
    return render(request, 'consultar/consultar_requisicoes_almoxarifado.html', context)


def _build_query_string_base(request, exclude=None):
    """Build query string from request.GET excluding specified params (for pagination/sort links)."""
    exclude = exclude or []
    qd = request.GET.copy()
    for param in exclude:
        if param in qd:
            del qd[param]
    return qd.urlencode()


def visualizar_requisicao_almoxarifado(request, requisicao_id):
    """Visualizar detalhes de uma requisição almoxarifado específica"""
    from app.models import RequisicaoAlmoxarifado
    
    try:
        requisicao = RequisicaoAlmoxarifado.objects.get(id=requisicao_id)
    except RequisicaoAlmoxarifado.DoesNotExist:
        messages.error(request, 'Requisição não encontrada.')
        return redirect('consultar_requisicoes_almoxarifado')
    
    context = {
        'page_title': f'Visualizar Requisição {requisicao.id}',
        'active_page': 'consultar_requisicoes_almoxarifado',
        'requisicao': requisicao,
    }
    return render(request, 'almoxarifado/visualizar_requisicao_almoxarifado.html', context)


def consultar_paradas_maquina(request):
    """Consultar/listar paradas de máquinas com filtros de ano e mês"""
    from app.models import ParadaMaquina
    from datetime import datetime
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses (None)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Se não há meses selecionados, usar todos os meses
    meses_para_mostrar = meses_filtro_int if meses_filtro_int else list(range(1, 13))
    
    # Busca geral
    search_query = request.GET.get('search', '').strip()
    todas_paradas = ParadaMaquina.objects.all()
    
    # Obter anos disponíveis (baseado em data)
    anos_disponiveis = []
    for parada in todas_paradas:
        if parada.data:
            anos_disponiveis.append(parada.data.year)
    anos_disponiveis = sorted(list(set(anos_disponiveis)), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    if hoje.year not in anos_disponiveis:
        anos_disponiveis.insert(0, hoje.year)
    
    # Filtrar por ano e mês (data)
    paradas_filtradas_por_data = []
    for parada in todas_paradas:
        if parada.data and parada.data.year == ano_filtro:
            if not meses_filtro_int or parada.data.month in meses_filtro_int:
                paradas_filtradas_por_data.append(parada.id)
    
    if paradas_filtradas_por_data:
        paradas_list = todas_paradas.filter(id__in=paradas_filtradas_por_data)
    else:
        # Se não há paradas no ano/mês, retornar queryset vazio
        paradas_list = ParadaMaquina.objects.none()
    
    # Aplicar busca geral
    if search_query:
        paradas_list = paradas_list.filter(
            Q(cod_item__icontains=search_query) |
            Q(descr_item__icontains=search_query) |
            Q(cod_recurso__icontains=search_query) |
            Q(descr_recurso__icontains=search_query) |
            Q(cod_parada__icontains=search_query) |
            Q(descr_parada__icontains=search_query) |
            Q(descr_linha_producao__icontains=search_query) |
            Q(grupo_recurso__icontains=search_query) |
            Q(motivo__icontains=search_query) |
            Q(acao__icontains=search_query)
        )
    
    # Ordenar por data (mais recente primeiro) e horário inicial
    paradas_list = paradas_list.order_by('-data', '-horario_inicial', 'cod_recurso')
    
    # Estatísticas baseadas nos dados FILTRADOS (antes da paginação)
    total_count = paradas_list.count()
    recursos_count = paradas_list.exclude(cod_recurso__isnull=True).values('cod_recurso').distinct().count()
    linhas_count = paradas_list.exclude(linha_producao__isnull=True).values('linha_producao').distinct().count()
    
    # Calcular total de horas paradas (dif_hora)
    total_horas_paradas = 0.0
    for parada in paradas_list:
        if parada.dif_hora:
            total_horas_paradas += float(parada.dif_hora)
    
    # Paginação (após calcular estatísticas)
    paginator = Paginator(paradas_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    paradas = paginator.get_page(page_number)
    
    # Meses para o template
    meses_choices = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]
    
    context = {
        'page_title': 'Consultar Paradas de Máquina',
        'active_page': 'consultar_paradas_maquina',
        'paradas': paradas,
        'total_count': total_count,
        'recursos_count': recursos_count,
        'linhas_count': linhas_count,
        'total_horas_paradas': total_horas_paradas,
        # Filtros de data
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'meses_para_mostrar': meses_para_mostrar,
        'anos_disponiveis': anos_disponiveis,
        'meses_choices': meses_choices,
        # Preservar filtros no contexto
        'search_query': search_query,
    }
    
    return render(request, 'paradas_maquina/consultar_paradas_de_maquina.html', context)


def visualizar_parada_de_maquina(request, parada_id):
    """Visualiza todos os dados de um registro ParadaMaquina."""
    from app.models import ParadaMaquina
    from django.shortcuts import get_object_or_404

    parada = get_object_or_404(ParadaMaquina, pk=parada_id)
    context = {
        'page_title': f'Visualizar Parada {parada.id}',
        'active_page': 'consultar_paradas_maquina',
        'parada': parada,
    }
    return render(request, 'parada_de_maquina/visualizar_parada_de_maquina.html', context)


def configuracao_parada_maquina(request):
    """Página de configuração de paradas de máquina: carrega e salva ConfigParadaMaquina, ConfigRecursoParadaMaquina e ConfigLinhaProducaoCentroAtividade."""
    from app.models import ConfigParadaMaquina, ConfigRecursoParadaMaquina, ConfigLinhaProducaoCentroAtividade, ParadaMaquina, CentroAtividade
    from datetime import date
    from decimal import Decimal, InvalidOperation

    meses = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]
    ano_atual = date.today().year

    def _normalize_br_number(s):
        """Converte número no padrão BR (1.234,56) para formato Python (1234.56)."""
        s = str(s).strip()
        if not s:
            return ''
        s = s.replace('.', '').replace(',', '.')
        return s

    def _parse_int(val):
        if val is None or str(val).strip() == '':
            return None
        s = _normalize_br_number(val)
        if not s:
            return None
        try:
            return int(Decimal(s))
        except (ValueError, TypeError, InvalidOperation):
            return None

    def _parse_decimal(val):
        if val is None or str(val).strip() == '':
            return None
        s = _normalize_br_number(val)
        if not s:
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _parse_percentage(val):
        """Parse percentage value, stripping trailing '%' if present. Accepts BR format (85,50 ou 85,50%)."""
        if val is None or str(val).strip() == '':
            return None
        s = str(val).strip().rstrip('%').strip()
        return _parse_decimal(s)

    # POST: identificar qual formulário foi enviado
    if request.method == 'POST' and 'salvar_linha_ca' in request.POST:
        # Formulário "Linha de Produção × Centro de Atividade"
        grupos_recurso = list(
            ParadaMaquina.objects.exclude(descr_linha_producao__isnull=True)
            .exclude(descr_linha_producao='')
            .values_list('descr_linha_producao', flat=True)
            .distinct()
            .order_by('descr_linha_producao')
        )
        ConfigLinhaProducaoCentroAtividade.objects.all().delete()
        for i, linha in enumerate(grupos_recurso):
            ca_ids = request.POST.getlist(f'centro_linha_{i}')
            for cid in ca_ids:
                cid = (cid or '').strip()
                if cid.isdigit():
                    try:
                        ConfigLinhaProducaoCentroAtividade.objects.create(
                            descr_linha_producao=linha,
                            centro_atividade_id=int(cid)
                        )
                    except Exception:
                        pass
        messages.success(request, 'Associações Linha de Produção × Centro de Atividade salvas com sucesso.')
        return redirect('configuracao_parada_maquina')

    if request.method == 'POST' and 'salvar_recursos' in request.POST:
        # Formulário "Recursos incluídos na Análise"
        recurso_frigorifico = request.POST.getlist('recurso_frigorifico')
        recurso_industria = request.POST.getlist('recurso_industria')
        ConfigRecursoParadaMaquina.objects.filter(secao='frigorifico').delete()
        ConfigRecursoParadaMaquina.objects.filter(secao='industria').delete()
        for gr in recurso_frigorifico:
            gr = (gr or '').strip()
            if gr:
                ConfigRecursoParadaMaquina.objects.create(secao='frigorifico', descr_recurso=gr)
        for gr in recurso_industria:
            gr = (gr or '').strip()
            if gr:
                ConfigRecursoParadaMaquina.objects.create(secao='industria', descr_recurso=gr)
        messages.success(request, 'Recursos incluídos na análise salvos com sucesso.')
        return redirect('configuracao_parada_maquina')

    if request.method == 'POST':
        for mes_num, _ in meses:
            # Frigorífico
            ConfigParadaMaquina.objects.update_or_create(
                ano=ano_atual,
                mes=mes_num,
                secao='frigorifico',
                defaults={
                    'suinos_abatidos': _parse_int(request.POST.get(f'mes_{mes_num}_frigorifico_suinos_abatidos')),
                    'dias_uteis': _parse_int(request.POST.get(f'mes_{mes_num}_frigorifico_dias_uteis')),
                    'total_abate_planejado': _parse_int(request.POST.get(f'mes_{mes_num}_frigorifico_total_abate_planejado')),
                    'perda_maximo': _parse_percentage(request.POST.get(f'mes_{mes_num}_frigorifico_perda_maximo')),
                    'fator_eficiencia': _parse_percentage(request.POST.get(f'mes_{mes_num}_frigorifico_fator_eficiencia')),
                    'carcacas_por_minuto': _parse_decimal(request.POST.get(f'mes_{mes_num}_frigorifico_carcacas_por_minuto')),
                }
            )
            # Indústria
            ConfigParadaMaquina.objects.update_or_create(
                ano=ano_atual,
                mes=mes_num,
                secao='industria',
                defaults={
                    'dias_uteis_industria': _parse_int(request.POST.get(f'mes_{mes_num}_industria_dias_uteis')),
                    'total_producao_planejada_industria': _parse_int(request.POST.get(f'mes_{mes_num}_industria_total_producao')),
                    'perda_maximo_industria': _parse_percentage(request.POST.get(f'mes_{mes_num}_industria_perda_maximo')),
                    'fator_eficiencia_industria': _parse_percentage(request.POST.get(f'mes_{mes_num}_industria_fator_eficiencia')),
                }
            )
        messages.success(request, 'Configuração salva com sucesso.')
        return redirect('configuracao_parada_maquina')

    # GET: carregar configurações do ano atual
    configs = ConfigParadaMaquina.objects.filter(ano=ano_atual)
    config_frigorifico = {c.mes: c for c in configs if c.secao == 'frigorifico'}
    config_industria = {c.mes: c for c in configs if c.secao == 'industria'}
    config_por_mes = [
        (mes_num, mes_nome, config_frigorifico.get(mes_num), config_industria.get(mes_num))
        for mes_num, mes_nome in meses
    ]

    # Recursos incluídos na análise: distinct descr_linha_producao da tabela ParadaMaquina
    grupos_recurso = list(
        ParadaMaquina.objects.exclude(descr_linha_producao__isnull=True)
        .exclude(descr_linha_producao='')
        .values_list('descr_linha_producao', flat=True)
        .distinct()
        .order_by('descr_linha_producao')
    )
    recursos_frigorifico = set(
        ConfigRecursoParadaMaquina.objects.filter(secao='frigorifico').values_list('descr_recurso', flat=True)
    )
    recursos_industria = set(
        ConfigRecursoParadaMaquina.objects.filter(secao='industria').values_list('descr_recurso', flat=True)
    )

    # Linha de Produção × Centro de Atividade
    centros_atividade = list(CentroAtividade.objects.all().order_by('ca'))
    linha_ca_map = {}  # { descr_linha_producao: set(of centro_atividade ids) }
    for rel in ConfigLinhaProducaoCentroAtividade.objects.select_related('centro_atividade'):
        linha_ca_map.setdefault(rel.descr_linha_producao, set()).add(rel.centro_atividade_id)

    context = {
        'page_title': 'Configuração de Paradas de Máquina',
        'active_page': 'configuracao_parada_maquina',
        'meses': meses,
        'config_por_mes': config_por_mes,
        'ano_atual': ano_atual,
        'config_frigorifico': config_frigorifico,
        'config_industria': config_industria,
        'grupos_recurso': grupos_recurso,
        'recursos_frigorifico': recursos_frigorifico,
        'recursos_industria': recursos_industria,
        'centros_atividade': centros_atividade,
        'linha_ca_map': linha_ca_map,
    }
    return render(request, 'paradas_maquina/configuracao_parada_de_maquina.html', context)


def analise_paradas_maquina(request):
    """Análise de paradas de máquinas com estatísticas e gráficos. Frigorífico/Indústria conforme ConfigRecursoParadaMaquina (descr_recurso)."""
    from app.models import ParadaMaquina, ParadaMaquinaOS, OrdemServicoCorretiva, Maquina, ConfigRecursoParadaMaquina, ConfigParadaMaquina, ProducaoDiaria
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Q, Avg
    from collections import defaultdict
    import json
    from calendar import monthrange
    from decimal import Decimal
    
    # Obter anos e meses disponíveis no banco de dados
    anos_disponiveis = ParadaMaquina.objects.exclude(data__isnull=True).values_list('data__year', flat=True).distinct().order_by('-data__year')
    meses_disponiveis = {}
    for ano in anos_disponiveis:
        meses = ParadaMaquina.objects.filter(data__year=ano).exclude(data__isnull=True).values_list('data__month', flat=True).distinct().order_by('data__month')
        meses_disponiveis[ano] = list(meses)
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Construir queryset base com filtros
    queryset_base = ParadaMaquina.objects.exclude(data__isnull=True)
    
    # Aplicar filtro de ano
    if ano_filtro:
        queryset_base = queryset_base.filter(data__year=ano_filtro)
        
        # Aplicar filtro de meses (múltiplos)
        if meses_filtro_int:
            # Filtrar por múltiplos meses usando Q objects
            mes_conditions = Q()
            for mes in meses_filtro_int:
                mes_conditions |= Q(data__month=mes)
            queryset_base = queryset_base.filter(mes_conditions)

    # Frigorífico e Indústria conforme ConfigRecursoParadaMaquina (descr_recurso)
    recursos_frigorifico = set(
        ConfigRecursoParadaMaquina.objects.filter(secao='frigorifico').values_list('descr_recurso', flat=True)
    )
    recursos_industria = set(
        ConfigRecursoParadaMaquina.objects.filter(secao='industria').values_list('descr_recurso', flat=True)
    )
    qs_frigorifico = queryset_base.filter(descr_linha_producao__in=recursos_frigorifico) if recursos_frigorifico else ParadaMaquina.objects.none()
    qs_industria = queryset_base.filter(descr_linha_producao__in=recursos_industria) if recursos_industria else ParadaMaquina.objects.none()
    
    # Estatísticas gerais (usando queryset filtrado)
    hoje_date = hoje.date()
    total_paradas = queryset_base.count()
    
    # Últimos 30 dias (apenas se não houver filtros)
    if not ano_filtro:
        data_30_dias_atras = hoje_date - timedelta(days=30)
        paradas_recentes = queryset_base.filter(
            data__gte=data_30_dias_atras
        ).count()
    else:
        # Se há filtros, mostrar total filtrado
        paradas_recentes = total_paradas
    
    # Mês atual (apenas se não houver filtros)
    if not ano_filtro:
        primeiro_dia_mes = hoje_date.replace(day=1)
        paradas_mes_atual = queryset_base.filter(
            data__gte=primeiro_dia_mes
        ).count()
    else:
        # Se há filtros, mostrar total filtrado
        paradas_mes_atual = total_paradas
    
    # Recursos únicos
    recursos_unicos = queryset_base.exclude(cod_recurso__isnull=True).values('cod_recurso').distinct().count()
    
    # Linhas de produção únicas
    linhas_unicas = queryset_base.exclude(linha_producao__isnull=True).values('linha_producao').distinct().count()
    
    # Calcular total de horas paradas (dif_hora)
    total_horas_paradas = Decimal('0.00')
    for parada in queryset_base:
        if parada.dif_hora:
            total_horas_paradas += Decimal(str(parada.dif_hora))
    
    # Média de horas paradas por parada
    media_horas_paradas = total_horas_paradas / total_paradas if total_paradas > 0 else Decimal('0.00')

    # Frigorífico / Indústria: totais e horas
    total_paradas_frig = qs_frigorifico.count()
    total_horas_frig = Decimal('0.00')
    for parada in qs_frigorifico:
        if parada.dif_hora:
            total_horas_frig += Decimal(str(parada.dif_hora))
    total_paradas_ind = qs_industria.count()
    total_horas_ind = Decimal('0.00')
    for parada in qs_industria:
        if parada.dif_hora:
            total_horas_ind += Decimal(str(parada.dif_hora))
    
    # Soma capacidade (Frigorífico): soma da coluna capacidade em ParadaMaquina para dados do Frigorífico
    soma_capacidade_frig = qs_frigorifico.aggregate(s=Sum('capacidade'))['s']
    if soma_capacidade_frig is None:
        soma_capacidade_frig = Decimal('0.00')
    else:
        soma_capacidade_frig = Decimal(str(soma_capacidade_frig))
    
    # Soma capacidade (Indústria): soma da coluna capacidade para dados da Indústria (recursos incluídos na config)
    soma_capacidade_ind = qs_industria.aggregate(s=Sum('capacidade'))['s']
    if soma_capacidade_ind is None:
        soma_capacidade_ind = Decimal('0.00')
    else:
        soma_capacidade_ind = Decimal(str(soma_capacidade_ind))
    
    # Evolução temporal (últimos 12 meses ou período filtrado)
    meses_labels = []
    meses_data = []
    meses_horas = []
    meses_data_frig = []
    meses_horas_frig = []
    meses_data_ind = []
    meses_horas_ind = []
    
    # Determinar período para evolução temporal
    if ano_filtro:
        try:
            periodo_inicio = datetime(ano_filtro, 1, 1).date()
            if meses_filtro_int:
                # Se há meses selecionados, usar apenas os meses selecionados (não o intervalo entre primeiro e último)
                primeiro_mes = min(meses_filtro_int)
                ultimo_mes = max(meses_filtro_int)
                periodo_inicio = datetime(ano_filtro, primeiro_mes, 1).date()
                ultimo_dia = monthrange(ano_filtro, ultimo_mes)[1]
                periodo_fim = datetime(ano_filtro, ultimo_mes, ultimo_dia).date()
                if periodo_fim > hoje_date:
                    periodo_fim = hoje_date
                # Lista de meses a iterar: apenas os selecionados
                meses_para_iterar = [datetime(ano_filtro, m, 1).date() for m in meses_filtro_int]
            else:
                # Todos os meses do ano
                periodo_fim = datetime(ano_filtro, 12, 31).date()
                if periodo_fim > hoje_date:
                    periodo_fim = hoje_date
                meses_para_iterar = []
                d = periodo_inicio.replace(day=1)
                while d <= periodo_fim:
                    meses_para_iterar.append(d)
                    if d.month == 12:
                        d = d.replace(year=d.year + 1, month=1)
                    else:
                        d = d.replace(month=d.month + 1)
        except (ValueError, TypeError):
            periodo_inicio = (hoje_date - timedelta(days=365)).replace(day=1)
            periodo_fim = hoje_date
            meses_para_iterar = []
            d = periodo_inicio.replace(day=1)
            while d <= periodo_fim:
                meses_para_iterar.append(d)
                if d.month == 12:
                    d = d.replace(year=d.year + 1, month=1)
                else:
                    d = d.replace(month=d.month + 1)
    else:
        periodo_inicio = (hoje_date - timedelta(days=365)).replace(day=1)
        periodo_fim = hoje_date
        meses_para_iterar = []
        d = periodo_inicio.replace(day=1)
        while d <= periodo_fim:
            meses_para_iterar.append(d)
            if d.month == 12:
                d = d.replace(year=d.year + 1, month=1)
            else:
                d = d.replace(month=d.month + 1)
    
    # Produção esperada (Frigorífico): soma de total_abate_planejado da config para cada mês no período filtrado
    producao_esperada_frig = 0
    for _data_cfg in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_cfg.year, mes=_data_cfg.month, secao='frigorifico'
        ).first()
        if cfg and cfg.total_abate_planejado is not None:
            producao_esperada_frig += cfg.total_abate_planejado
    
    # Soma Suínos Abatidos (Frigorífico): soma de ProducaoDiaria.suinos_abatidos para cada mês no período filtrado
    soma_suinos_abatidos_frig = 0
    for _data_pd in meses_para_iterar:
        qs_pd = ProducaoDiaria.objects.filter(ano=_data_pd.year, mes=_data_pd.month)
        if _data_pd.year == periodo_fim.year and _data_pd.month == periodo_fim.month:
            qs_pd = qs_pd.filter(dia__lte=periodo_fim.day)
        result = qs_pd.aggregate(s=Sum('suinos_abatidos'))['s']
        if result is not None:
            soma_suinos_abatidos_frig += result
    
    # Soma Produção Indústria: soma de ProducaoDiaria.producao_industria para cada mês no período filtrado
    soma_producao_industria = Decimal('0.00')
    for _data_pi in meses_para_iterar:
        qs_pi = ProducaoDiaria.objects.filter(ano=_data_pi.year, mes=_data_pi.month)
        if _data_pi.year == periodo_fim.year and _data_pi.month == periodo_fim.month:
            qs_pi = qs_pi.filter(dia__lte=periodo_fim.day)
        result = qs_pi.aggregate(s=Sum('producao_industria'))['s']
        if result is not None:
            soma_producao_industria += Decimal(str(result))
    
    # Perda Máxima planejada (Frigorífico): soma por mês de total_abate_planejado × (perda_maximo / 100)
    perda_maxima_calculada_frig = Decimal('0.00')
    for _data_perda in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_perda.year, mes=_data_perda.month, secao='frigorifico'
        ).first()
        if cfg and cfg.perda_maximo is not None and cfg.total_abate_planejado is not None:
            perda_maxima_calculada_frig += (Decimal(str(cfg.perda_maximo)) / Decimal('100')) * Decimal(str(cfg.total_abate_planejado))
    
    # Perda Restante (Frigorífico): Perda Máx. (calculada) - Soma Capacidade (Perdas)
    perda_restante_frig = perda_maxima_calculada_frig - soma_capacidade_frig
    
    # Dias úteis (Frigorífico): soma de ConfigParadaMaquina.dias_uteis para cada mês no período
    dias_uteis_total_frig = 0
    for _data_du in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_du.year, mes=_data_du.month, secao='frigorifico'
        ).first()
        if cfg and cfg.dias_uteis is not None:
            dias_uteis_total_frig += cfg.dias_uteis
    
    # Dias úteis (Indústria): soma de ConfigParadaMaquina.dias_uteis_industria para cada mês no período
    dias_uteis_total_industria = 0
    for _data_du_ind in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_du_ind.year, mes=_data_du_ind.month, secao='industria'
        ).first()
        if cfg and cfg.dias_uteis_industria is not None:
            dias_uteis_total_industria += cfg.dias_uteis_industria
    
    # Total Produção Planejada (Indústria): soma de ConfigParadaMaquina.total_producao_planejada_industria para cada mês no período
    total_producao_planejada_industria = 0
    for _data_tpp in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_tpp.year, mes=_data_tpp.month, secao='industria'
        ).first()
        if cfg and cfg.total_producao_planejada_industria is not None:
            total_producao_planejada_industria += cfg.total_producao_planejada_industria
    
    # Produção diária planejada (Indústria): total planejada / dias úteis (média por dia útil)
    producao_diaria_planejada_industria = (
        total_producao_planejada_industria / dias_uteis_total_industria
        if dias_uteis_total_industria and dias_uteis_total_industria > 0
        else None
    )
    
    # Suínos Abatidos por dia (Frigorífico): média de ConfigParadaMaquina.suinos_abatidos no período
    suinos_abatidos_por_dia_soma = 0
    suinos_abatidos_por_dia_n = 0
    for _data_sad in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_sad.year, mes=_data_sad.month, secao='frigorifico'
        ).first()
        if cfg and cfg.suinos_abatidos is not None:
            suinos_abatidos_por_dia_soma += cfg.suinos_abatidos
            suinos_abatidos_por_dia_n += 1
    suinos_abatidos_por_dia_medio_frig = (suinos_abatidos_por_dia_soma / suinos_abatidos_por_dia_n) if suinos_abatidos_por_dia_n else None
    
    # % de Perda Máximo (Frigorífico): primeiro valor configurado no período (ConfigParadaMaquina.perda_maximo)
    perda_maximo_frig = None
    for _data_pm in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_pm.year, mes=_data_pm.month, secao='frigorifico'
        ).first()
        if cfg and cfg.perda_maximo is not None:
            perda_maximo_frig = cfg.perda_maximo
            break
    
    # % de Perda Máximo (Indústria): primeiro valor configurado no período (ConfigParadaMaquina.perda_maximo_industria)
    perda_maximo_industria = None
    for _data_pmi in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_pmi.year, mes=_data_pmi.month, secao='industria'
        ).first()
        if cfg and cfg.perda_maximo_industria is not None:
            perda_maximo_industria = cfg.perda_maximo_industria
            break
    
    # Perda Máxima planejada (Indústria): soma por mês de total_producao_planejada_industria × (perda_maximo_industria / 100)
    perda_maxima_planejada_industria = Decimal('0.00')
    for _data_pmpi in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_pmpi.year, mes=_data_pmpi.month, secao='industria'
        ).first()
        if cfg and cfg.perda_maximo_industria is not None and cfg.total_producao_planejada_industria is not None:
            perda_maxima_planejada_industria += (Decimal(str(cfg.perda_maximo_industria)) / Decimal('100')) * Decimal(str(cfg.total_producao_planejada_industria))
    
    # Perda Restante (Indústria): Perda Máxima planejada - Total de Perdas (Kg)
    perda_restante_industria = perda_maxima_planejada_industria - soma_capacidade_ind
    
    # Fator de Eficiência % (Indústria): primeiro valor configurado no período
    fator_eficiencia_industria = None
    for _data_fei in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_fei.year, mes=_data_fei.month, secao='industria'
        ).first()
        if cfg and cfg.fator_eficiencia_industria is not None:
            fator_eficiencia_industria = cfg.fator_eficiencia_industria
            break
    
    # Indicador sem Fator de Correção (Indústria): Total de Perdas (Kg) / Produção realizada (como %)
    if soma_producao_industria and soma_producao_industria > 0:
        indicador_sem_fator_industria = (soma_capacidade_ind / soma_producao_industria) * Decimal('100')
    else:
        indicador_sem_fator_industria = None
    
    # Indicador Atual (Indústria): (Total de Perdas (Kg) / Produção realizada) * Fator de Eficiência %
    if soma_producao_industria and soma_producao_industria > 0 and fator_eficiencia_industria is not None:
        indicador_atual_industria = (soma_capacidade_ind / soma_producao_industria) * Decimal(str(fator_eficiencia_industria))
    else:
        indicador_atual_industria = None
    
    # Fator de Eficiência % (Frigorífico): primeiro valor configurado no período (ConfigParadaMaquina.fator_eficiencia)
    fator_eficiencia_frig = None
    for _data_fe in meses_para_iterar:
        cfg = ConfigParadaMaquina.objects.filter(
            ano=_data_fe.year, mes=_data_fe.month, secao='frigorifico'
        ).first()
        if cfg and cfg.fator_eficiencia is not None:
            fator_eficiencia_frig = cfg.fator_eficiencia
            break
    
    # Indicador sem Fator de Correção (Frigorífico): Total carcaças Perdidas / Total suínos abatidos (como %)
    if soma_suinos_abatidos_frig and int(soma_suinos_abatidos_frig) > 0:
        indicador_sem_fator_frig = (Decimal(str(soma_capacidade_frig)) / Decimal(str(soma_suinos_abatidos_frig))) * Decimal('100')
    else:
        indicador_sem_fator_frig = None
    
    # Indicador Atual (Frigorífico): (Total carcaças Perdidas / Total suínos abatidos) * Fator de Eficiência %
    # Fator de Eficiência % vem de ConfigParadaMaquina (configuração)
    if soma_suinos_abatidos_frig and int(soma_suinos_abatidos_frig) > 0 and fator_eficiencia_frig is not None:
        indicador_atual_frig = (Decimal(str(soma_capacidade_frig)) / Decimal(str(soma_suinos_abatidos_frig))) * Decimal(str(fator_eficiencia_frig))
    else:
        indicador_atual_frig = None
    
    # Gerar meses do período — contagens diretas para garantir alinhamento com o filtro (apenas meses selecionados)
    for data_atual in meses_para_iterar:
        # Calcular último dia do mês
        ultimo_dia_mes = monthrange(data_atual.year, data_atual.month)[1]
        fim_mes_calc = datetime(data_atual.year, data_atual.month, ultimo_dia_mes).date()
        fim_mes = fim_mes_calc if fim_mes_calc <= periodo_fim else periodo_fim
        
        # Contagem direta por período (garante uso do filtro ano+mês)
        base_filter = ParadaMaquina.objects.exclude(data__isnull=True).filter(
            data__year=data_atual.year,
            data__month=data_atual.month,
            data__gte=data_atual,
            data__lte=fim_mes
        )
        count = base_filter.count()
        
        horas_mes = Decimal('0.00')
        for parada in base_filter:
            if parada.dif_hora:
                horas_mes += Decimal(str(parada.dif_hora))
        
        base_frig = base_filter.filter(descr_linha_producao__in=recursos_frigorifico) if recursos_frigorifico else ParadaMaquina.objects.none()
        count_frig = base_frig.count()
        horas_frig = Decimal('0.00')
        for parada in base_frig:
            if parada.dif_hora:
                horas_frig += Decimal(str(parada.dif_hora))
        base_ind = base_filter.filter(descr_linha_producao__in=recursos_industria) if recursos_industria else ParadaMaquina.objects.none()
        count_ind = base_ind.count()
        horas_ind = Decimal('0.00')
        for parada in base_ind:
            if parada.dif_hora:
                horas_ind += Decimal(str(parada.dif_hora))
        
        meses_labels.append(data_atual.strftime('%b/%Y'))
        meses_data.append(count)
        meses_horas.append(float(horas_mes))
        meses_data_frig.append(count_frig)
        meses_horas_frig.append(float(horas_frig))
        meses_data_ind.append(count_ind)
        meses_horas_ind.append(float(horas_ind))
    
    # Rótulo do período para o gráfico (ex: "Jan/2025 - Fev/2025" ou "2025")
    if meses_labels:
        periodo_chart_label = f"{meses_labels[0]} - {meses_labels[-1]}" if len(meses_labels) > 1 else meses_labels[0]
    else:
        periodo_chart_label = "Período filtrado"
    
    # Evolução Temporal Frigorífico: dados por DIA (Total Paradas + Total Perda/capacidade)
    temporal_frig_labels = []
    temporal_frig_paradas = []
    temporal_frig_perda = []
    _dia_atual = periodo_inicio
    while _dia_atual <= periodo_fim:
        temporal_frig_labels.append(_dia_atual.strftime('%d/%m'))
        qs_dia_frig = qs_frigorifico.filter(data=_dia_atual) if recursos_frigorifico else ParadaMaquina.objects.none()
        temporal_frig_paradas.append(qs_dia_frig.count())
        perda_dia = qs_dia_frig.aggregate(s=Sum('capacidade'))['s']
        temporal_frig_perda.append(float(perda_dia) if perda_dia is not None else 0.0)
        _dia_atual = _dia_atual + timedelta(days=1)
    
    # Top 10 recursos com mais paradas (por quantidade)
    top_recursos_qtd = queryset_base.exclude(
        cod_recurso__isnull=True
    ).values('cod_recurso', 'descr_recurso').annotate(
        total_count=Count('id')
    ).order_by('-total_count')[:10]
    
    top_recursos_labels = []
    top_recursos_data = []
    for recurso in top_recursos_qtd:
        descr = recurso['descr_recurso'] or f"Recurso {recurso['cod_recurso']}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_recursos_labels.append(f"{recurso['cod_recurso']} - {descr}")
        top_recursos_data.append(recurso['total_count'])
    
    # Top 10 recursos por horas paradas
    recursos_horas_dict = defaultdict(lambda: Decimal('0.00'))
    
    for parada in queryset_base.exclude(cod_recurso__isnull=True):
        if parada.dif_hora:
            recursos_horas_dict[parada.cod_recurso] += Decimal(str(parada.dif_hora))
    
    # Ordenar e pegar top 10
    sorted_recursos = sorted(recursos_horas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_recursos_horas_labels = []
    top_recursos_horas_data = []
    for cod_recurso, horas in sorted_recursos:
        parada = queryset_base.filter(cod_recurso=cod_recurso).first()
        descr = parada.descr_recurso if parada and parada.descr_recurso else f"Recurso {cod_recurso}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_recursos_horas_labels.append(f"{cod_recurso} - {descr}")
        top_recursos_horas_data.append(float(horas))
    
    # Top 10 paradas (cod_parada) por quantidade
    top_paradas_qtd = queryset_base.exclude(
        cod_parada__isnull=True
    ).exclude(cod_parada='').values('cod_parada', 'descr_parada').annotate(
        total_count=Count('id')
    ).order_by('-total_count')[:10]
    
    top_paradas_labels = []
    top_paradas_data = []
    for parada in top_paradas_qtd:
        descr = parada['descr_parada'] or f"Parada {parada['cod_parada']}"
        if len(descr) > 40:
            descr = descr[:37] + "..."
        top_paradas_labels.append(f"{parada['cod_parada']} - {descr}")
        top_paradas_data.append(parada['total_count'])
    
    # Distribuição por turno
    turnos_dict = defaultdict(lambda: {'count': 0, 'horas': Decimal('0.00')})
    
    for parada in queryset_base.exclude(turno__isnull=True).exclude(turno=''):
        turnos_dict[parada.turno]['count'] += 1
        if parada.dif_hora:
            turnos_dict[parada.turno]['horas'] += Decimal(str(parada.dif_hora))
    
    sorted_turnos = sorted(turnos_dict.items(), key=lambda x: x[1]['count'], reverse=True)
    
    turnos_labels = []
    turnos_data_count = []
    turnos_data_horas = []
    for turno, dados in sorted_turnos:
        turnos_labels.append(str(turno) if turno else 'Não informado')
        turnos_data_count.append(dados['count'])
        turnos_data_horas.append(float(dados['horas']))
    
    # Distribuição por linha de produção (top 10)
    linhas_dict = defaultdict(lambda: {'count': 0, 'horas': Decimal('0.00')})
    
    for parada in queryset_base.exclude(linha_producao__isnull=True):
        linhas_dict[parada.linha_producao]['count'] += 1
        if parada.dif_hora:
            linhas_dict[parada.linha_producao]['horas'] += Decimal(str(parada.dif_hora))
    
    sorted_linhas = sorted(linhas_dict.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
    
    linhas_labels = []
    linhas_data_count = []
    linhas_data_horas = []
    for linha_id, dados in sorted_linhas:
        parada = queryset_base.filter(linha_producao=linha_id).first()
        descr = parada.descr_linha_producao if parada and parada.descr_linha_producao else f"Linha {linha_id}"
        if len(descr) > 30:
            descr = descr[:27] + "..."
        linhas_labels.append(f"{linha_id} - {descr}")
        linhas_data_count.append(dados['count'])
        linhas_data_horas.append(float(dados['horas']))
    
    # Paradas recentes (últimas 20)
    paradas_recentes_list = queryset_base.order_by('-data', '-horario_inicial')[:20]

    def _build_section_charts(qs):
        """Construir top recursos, top paradas, turnos, linhas e paradas recentes para um queryset."""
        recursos_unicos_s = qs.exclude(cod_recurso__isnull=True).values('cod_recurso').distinct().count()
        linhas_unicas_s = qs.exclude(linha_producao__isnull=True).values('linha_producao').distinct().count()
        total_h_s = sum((Decimal(str(p.dif_hora)) for p in qs if p.dif_hora), Decimal('0.00'))
        count_s = qs.count()
        media_horas_s = total_h_s / count_s if count_s > 0 else Decimal('0.00')
        top_r_qtd = qs.exclude(cod_recurso__isnull=True).values('cod_recurso', 'descr_recurso').annotate(total_count=Count('id')).order_by('-total_count')[:10]
        top_r_labels = []
        top_r_data = []
        for r in top_r_qtd:
            descr = r['descr_recurso'] or f"Recurso {r['cod_recurso']}"
            if len(descr) > 40:
                descr = descr[:37] + "..."
            top_r_labels.append(f"{r['cod_recurso']} - {descr}")
            top_r_data.append(r['total_count'])
        recursos_horas_d = defaultdict(lambda: Decimal('0.00'))
        for p in qs.exclude(cod_recurso__isnull=True):
            if p.dif_hora:
                recursos_horas_d[p.cod_recurso] += Decimal(str(p.dif_hora))
        sorted_r = sorted(recursos_horas_d.items(), key=lambda x: x[1], reverse=True)[:10]
        top_r_h_labels = []
        top_r_h_data = []
        for cod, horas in sorted_r:
            parada = qs.filter(cod_recurso=cod).first()
            descr = parada.descr_recurso if parada and parada.descr_recurso else f"Recurso {cod}"
            if len(descr) > 40:
                descr = descr[:37] + "..."
            top_r_h_labels.append(f"{cod} - {descr}")
            top_r_h_data.append(float(horas))
        top_p_qtd = qs.exclude(cod_parada__isnull=True).exclude(cod_parada='').values('cod_parada', 'descr_parada').annotate(total_count=Count('id')).order_by('-total_count')[:10]
        top_p_labels = []
        top_p_data = []
        for p in top_p_qtd:
            descr = p['descr_parada'] or f"Parada {p['cod_parada']}"
            if len(descr) > 40:
                descr = descr[:37] + "..."
            top_p_labels.append(f"{p['cod_parada']} - {descr}")
            top_p_data.append(p['total_count'])
        turnos_d = defaultdict(lambda: {'count': 0, 'horas': Decimal('0.00')})
        for p in qs.exclude(turno__isnull=True).exclude(turno=''):
            turnos_d[p.turno]['count'] += 1
            if p.dif_hora:
                turnos_d[p.turno]['horas'] += Decimal(str(p.dif_hora))
        sorted_t = sorted(turnos_d.items(), key=lambda x: x[1]['count'], reverse=True)
        t_labels = [str(t) if t else 'Não informado' for t, _ in sorted_t]
        t_count = [d['count'] for _, d in sorted_t]
        t_horas = [float(d['horas']) for _, d in sorted_t]
        linhas_d = defaultdict(lambda: {'count': 0, 'horas': Decimal('0.00')})
        for p in qs.exclude(linha_producao__isnull=True):
            linhas_d[p.linha_producao]['count'] += 1
            if p.dif_hora:
                linhas_d[p.linha_producao]['horas'] += Decimal(str(p.dif_hora))
        sorted_l = sorted(linhas_d.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
        l_labels = []
        l_count = []
        l_horas = []
        for linha_id, dados in sorted_l:
            parada = qs.filter(linha_producao=linha_id).first()
            descr = parada.descr_linha_producao if parada and parada.descr_linha_producao else f"Linha {linha_id}"
            if len(descr) > 30:
                descr = descr[:27] + "..."
            l_labels.append(f"{linha_id} - {descr}")
            l_count.append(dados['count'])
            l_horas.append(float(dados['horas']))
        recentes = list(qs.order_by('-data', '-horario_inicial')[:20])

        # Paradas por Máquina (via ParadaMaquinaOS -> OrdemServicoCorretiva, ou fallback: extrair OS de motivo/acao)
        maquina_paradas_d = defaultdict(lambda: {'count': 0, 'descr': ''})
        parada_ids = list(qs.values_list('id', flat=True))
        
        def _add_maquina_from_os(os_num_str):
            try:
                if not os_num_str or not str(os_num_str).strip().isdigit():
                    return
                os_int = int(str(os_num_str).strip())
                os_obj = OrdemServicoCorretiva.objects.filter(cd_ordemserv=os_int).first()
                if os_obj and os_obj.cd_maquina is not None:
                    maquina_paradas_d[os_obj.cd_maquina]['count'] += 1
                    maquina_paradas_d[os_obj.cd_maquina]['descr'] = os_obj.descr_maquina or ''
            except (ValueError, TypeError):
                pass
        
        if parada_ids:
            # 1) Usar ParadaMaquinaOS (associações confirmadas)
            for pmo in ParadaMaquinaOS.objects.filter(parada_maquina_id__in=parada_ids).select_related('parada_maquina'):
                _add_maquina_from_os(pmo.os_numero)
        
        # 2) Fallback: se não há dados em ParadaMaquinaOS, extrair OS de motivo/acao (mesma lógica da página Analise Máquina por Parada)
        if not maquina_paradas_d:
            import re
            os_standalone = re.compile(r'\b\d{5,6}\b')
            os_com_prefixo = re.compile(
                r'(?:OS|ordem|n[°º]|#|ref|numero|número)(?:\s*:?\s*)?(\d{5,6})\b',
                re.IGNORECASE
            )
            def extract_os(text):
                if not text or not isinstance(text, str):
                    return []
                found = set()
                found.update(os_standalone.findall(text))
                found.update(os_com_prefixo.findall(text))
                found.update(re.findall(r'[\s\-_\.\/\(]\s*(\d{5,6})(?=[\s\-_\.\/\)]|$)', text))
                return list(found)
            
            for parada in qs:
                texto = ' '.join(filter(None, [
                    str(parada.motivo or ''), str(parada.acao or ''),
                    str(parada.nro or ''), str(parada.descr_parada or '')
                ]))
                for os_num in extract_os(texto):
                    if len(os_num) in (5, 6):
                        _add_maquina_from_os(os_num)
        
        sorted_maq = sorted(maquina_paradas_d.items(), key=lambda x: x[1]['count'], reverse=True)
        maq_labels = []
        maq_data = []
        maq_urls = []
        for cd_maq, d in sorted_maq:
            descr = (d['descr'] or '')[:40]
            if d['descr'] and len(d['descr']) > 40:
                descr += "..."
            maq_labels.append(f"{cd_maq} - {descr}" if descr else str(cd_maq))
            maq_data.append(d['count'])
            # Link para página da máquina (visualizar_maquina usa maquina_id)
            m_obj = Maquina.objects.filter(cd_maquina=cd_maq).first()
            maq_urls.append(f"/maquinas/visualizar/{m_obj.id}/" if m_obj else "")

        return {
            'recursos_unicos': recursos_unicos_s,
            'linhas_unicas': linhas_unicas_s,
            'media_horas': media_horas_s,
            'top_recursos_labels': top_r_labels,
            'top_recursos_data': top_r_data,
            'top_recursos_horas_labels': top_r_h_labels,
            'top_recursos_horas_data': top_r_h_data,
            'top_paradas_labels': top_p_labels,
            'top_paradas_data': top_p_data,
            'turnos_labels': t_labels,
            'turnos_data_count': t_count,
            'turnos_data_horas': t_horas,
            'linhas_labels': l_labels,
            'linhas_data_count': l_count,
            'linhas_data_horas': l_horas,
            'paradas_recentes_list': recentes,
            'maquinas_labels': maq_labels,
            'maquinas_data': maq_data,
            'maquinas_urls': maq_urls,
        }

    section_frig = _build_section_charts(qs_frigorifico) if recursos_frigorifico else _build_section_charts(ParadaMaquina.objects.none())
    section_ind = _build_section_charts(qs_industria) if recursos_industria else _build_section_charts(ParadaMaquina.objects.none())
    
    # Dados diários para o mês selecionado (para o gráfico de evolução diária)
    if ano_filtro and meses_filtro_int:
        try:
            # Usar o primeiro mês selecionado para o gráfico diário
            mes = meses_filtro_int[0]
            primeiro_dia_mes_atual = datetime(ano_filtro, mes, 1).date()
            ultimo_dia_mes_atual = datetime(ano_filtro, mes, monthrange(ano_filtro, mes)[1]).date()
            if ultimo_dia_mes_atual > hoje_date:
                ultimo_dia_mes_atual = hoje_date
        except (ValueError, TypeError):
            primeiro_dia_mes_atual = hoje_date.replace(day=1)
            if hoje_date.month == 12:
                ultimo_dia_mes_atual = hoje_date.replace(year=hoje_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                ultimo_dia_mes_atual = hoje_date.replace(month=hoje_date.month + 1, day=1) - timedelta(days=1)
            if ultimo_dia_mes_atual > hoje_date:
                ultimo_dia_mes_atual = hoje_date
    else:
        primeiro_dia_mes_atual = hoje_date.replace(day=1)
        if hoje_date.month == 12:
            ultimo_dia_mes_atual = hoje_date.replace(year=hoje_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            ultimo_dia_mes_atual = hoje_date.replace(month=hoje_date.month + 1, day=1) - timedelta(days=1)
        if ultimo_dia_mes_atual > hoje_date:
            ultimo_dia_mes_atual = hoje_date
    
    dias_labels = []
    dias_data = []
    dias_horas = []
    dias_data_frig = []
    dias_horas_frig = []
    dias_data_ind = []
    dias_horas_ind = []
    
    for dia in range(1, ultimo_dia_mes_atual.day + 1):
        data_dia = primeiro_dia_mes_atual.replace(day=dia)
        count = queryset_base.filter(data=data_dia).count()
        
        horas_dia = Decimal('0.00')
        for parada in queryset_base.filter(data=data_dia):
            if parada.dif_hora:
                horas_dia += Decimal(str(parada.dif_hora))
        
        count_frig = qs_frigorifico.filter(data=data_dia).count()
        horas_frig = Decimal('0.00')
        for parada in qs_frigorifico.filter(data=data_dia):
            if parada.dif_hora:
                horas_frig += Decimal(str(parada.dif_hora))
        count_ind = qs_industria.filter(data=data_dia).count()
        horas_ind = Decimal('0.00')
        for parada in qs_industria.filter(data=data_dia):
            if parada.dif_hora:
                horas_ind += Decimal(str(parada.dif_hora))
        
        dias_labels.append(data_dia.strftime('%d/%m'))
        dias_data.append(count)
        dias_horas.append(float(horas_dia))
        dias_data_frig.append(count_frig)
        dias_horas_frig.append(float(horas_frig))
        dias_data_ind.append(count_ind)
        dias_horas_ind.append(float(horas_ind))
    
    # Determinar mês selecionado para o gráfico diário
    if ano_filtro and meses_filtro_int:
        mes_selecionado_grafico = f"{ano_filtro}-{str(meses_filtro_int[0]).zfill(2)}"
    else:
        mes_selecionado_grafico = hoje.strftime('%Y-%m')
    
    # Nomes dos meses em português
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Obter lista de anos disponíveis (incluir ano atual e ano filtrado para permitir filtro em meses futuros)
    anos_disponiveis_list = list(anos_disponiveis)
    anos_set = set(anos_disponiveis_list) | {hoje.year}
    if ano_filtro:
        anos_set.add(ano_filtro)
    anos_disponiveis_list = sorted(anos_set, reverse=True)
    if not anos_disponiveis_list:
        anos_disponiveis_list = [hoje.year]
    
    context = {
        'page_title': 'Análise de Paradas de Máquina',
        'active_page': 'analise_paradas_maquina',
        'total_paradas': total_paradas,
        'paradas_recentes': paradas_recentes,
        'paradas_mes_atual': paradas_mes_atual,
        'recursos_unicos': recursos_unicos,
        'linhas_unicas': linhas_unicas,
        'total_horas_paradas': total_horas_paradas,
        'media_horas_paradas': media_horas_paradas,
        'total_paradas_frig': total_paradas_frig,
        'total_horas_frig': total_horas_frig,
        'producao_esperada_frig': producao_esperada_frig,
        'soma_capacidade_frig': soma_capacidade_frig,
        'soma_suinos_abatidos_frig': soma_suinos_abatidos_frig,
        'perda_maxima_calculada_frig': perda_maxima_calculada_frig,
        'perda_restante_frig': perda_restante_frig,
        'perda_maximo_frig': perda_maximo_frig,
        'perda_maximo_industria': perda_maximo_industria,
        'perda_maxima_planejada_industria': perda_maxima_planejada_industria,
        'perda_restante_industria': perda_restante_industria,
        'fator_eficiencia_industria': fator_eficiencia_industria,
        'indicador_sem_fator_industria': indicador_sem_fator_industria,
        'indicador_atual_industria': indicador_atual_industria,
        'fator_eficiencia_frig': fator_eficiencia_frig,
        'indicador_sem_fator_frig': indicador_sem_fator_frig,
        'indicador_atual_frig': indicador_atual_frig,
        'dias_uteis_total_frig': dias_uteis_total_frig,
        'dias_uteis_total_industria': dias_uteis_total_industria,
        'total_producao_planejada_industria': total_producao_planejada_industria,
        'producao_diaria_planejada_industria': producao_diaria_planejada_industria,
        'suinos_abatidos_por_dia_medio_frig': suinos_abatidos_por_dia_medio_frig,
        'total_paradas_ind': total_paradas_ind,
        'total_horas_ind': total_horas_ind,
        'soma_capacidade_ind': soma_capacidade_ind,
        'soma_producao_industria': soma_producao_industria,
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'meses_horas': json.dumps(meses_horas),
        'meses_data_frig': json.dumps(meses_data_frig),
        'meses_horas_frig': json.dumps(meses_horas_frig),
        'temporal_frig_labels': json.dumps(temporal_frig_labels),
        'temporal_frig_paradas': json.dumps(temporal_frig_paradas),
        'temporal_frig_perda': json.dumps(temporal_frig_perda),
        'meses_data_ind': json.dumps(meses_data_ind),
        'meses_horas_ind': json.dumps(meses_horas_ind),
        'dias_labels': json.dumps(dias_labels),
        'dias_data': json.dumps(dias_data),
        'dias_horas': json.dumps(dias_horas),
        'dias_data_frig': json.dumps(dias_data_frig),
        'dias_horas_frig': json.dumps(dias_horas_frig),
        'dias_data_ind': json.dumps(dias_data_ind),
        'dias_horas_ind': json.dumps(dias_horas_ind),
        'mes_selecionado': mes_selecionado_grafico,
        'top_recursos_labels': json.dumps(top_recursos_labels),
        'top_recursos_data': json.dumps(top_recursos_data),
        'top_recursos_horas_labels': json.dumps(top_recursos_horas_labels),
        'top_recursos_horas_data': json.dumps(top_recursos_horas_data),
        'top_paradas_labels': json.dumps(top_paradas_labels),
        'top_paradas_data': json.dumps(top_paradas_data),
        'turnos_labels': json.dumps(turnos_labels),
        'turnos_data_count': json.dumps(turnos_data_count),
        'turnos_data_horas': json.dumps(turnos_data_horas),
        'linhas_labels': json.dumps(linhas_labels),
        'linhas_data_count': json.dumps(linhas_data_count),
        'linhas_data_horas': json.dumps(linhas_data_horas),
        'paradas_recentes_list': paradas_recentes_list,
        'section_frig': section_frig,
        'section_ind': section_ind,
        'section_frig_top_recursos_labels': json.dumps(section_frig['top_recursos_labels']),
        'section_frig_top_recursos_data': json.dumps(section_frig['top_recursos_data']),
        'section_frig_top_recursos_horas_labels': json.dumps(section_frig['top_recursos_horas_labels']),
        'section_frig_top_recursos_horas_data': json.dumps(section_frig['top_recursos_horas_data']),
        'section_frig_top_paradas_labels': json.dumps(section_frig['top_paradas_labels']),
        'section_frig_top_paradas_data': json.dumps(section_frig['top_paradas_data']),
        'section_frig_turnos_labels': json.dumps(section_frig['turnos_labels']),
        'section_frig_turnos_data_count': json.dumps(section_frig['turnos_data_count']),
        'section_frig_turnos_data_horas': json.dumps(section_frig['turnos_data_horas']),
        'section_frig_linhas_labels': json.dumps(section_frig['linhas_labels']),
        'section_frig_linhas_data_count': json.dumps(section_frig['linhas_data_count']),
        'section_frig_linhas_data_horas': json.dumps(section_frig['linhas_data_horas']),
        'section_ind_top_recursos_labels': json.dumps(section_ind['top_recursos_labels']),
        'section_ind_top_recursos_data': json.dumps(section_ind['top_recursos_data']),
        'section_ind_top_recursos_horas_labels': json.dumps(section_ind['top_recursos_horas_labels']),
        'section_ind_top_recursos_horas_data': json.dumps(section_ind['top_recursos_horas_data']),
        'section_ind_top_paradas_labels': json.dumps(section_ind['top_paradas_labels']),
        'section_ind_top_paradas_data': json.dumps(section_ind['top_paradas_data']),
        'section_ind_turnos_labels': json.dumps(section_ind['turnos_labels']),
        'section_ind_turnos_data_count': json.dumps(section_ind['turnos_data_count']),
        'section_ind_turnos_data_horas': json.dumps(section_ind['turnos_data_horas']),
        'section_ind_linhas_labels': json.dumps(section_ind['linhas_labels']),
        'section_ind_linhas_data_count': json.dumps(section_ind['linhas_data_count']),
        'section_ind_linhas_data_horas': json.dumps(section_ind['linhas_data_horas']),
        'section_frig_maquinas_labels': json.dumps(section_frig['maquinas_labels']),
        'section_frig_maquinas_data': json.dumps(section_frig['maquinas_data']),
        'section_frig_maquinas_urls': json.dumps(section_frig['maquinas_urls']),
        'section_ind_maquinas_labels': json.dumps(section_ind['maquinas_labels']),
        'section_ind_maquinas_data': json.dumps(section_ind['maquinas_data']),
        'section_ind_maquinas_urls': json.dumps(section_ind['maquinas_urls']),
        # Filtros
        'anos_disponiveis': anos_disponiveis_list,
        'meses_nomes': meses_nomes,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'periodo_chart_label': periodo_chart_label,
    }
    return render(request, 'paradas_maquina/analise_paradas_de_maquina.html', context)


def analise_maquina_por_parada(request):
    """
    Análise de ParadaMaquina focada em extrair Ordens de Serviço (OS) dos campos motivo e acao.
    OS são identificadas por sequências de 5 ou 6 dígitos. Mesmos filtros (ano/mês) da página analise_paradas_maquina.
    Permite confirmar associações que são salvas em ParadaMaquinaOS para uso em outras análises.
    """
    from app.models import ParadaMaquina, ParadaMaquinaOS
    from datetime import datetime
    from django.db.models import Q
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.contrib import messages
    import re
    from collections import defaultdict

    # Processar confirmações (POST)
    if request.method == 'POST':
        ano_post = request.POST.get('ano')
        meses_post = request.POST.getlist('mes')
        # Pares (parada_id, os_numero) confirmados (checkboxes marcados)
        confirmados_set = set()
        for key in request.POST:
            if key.startswith('confirm_'):
                val = key.replace('confirm_', '')
                if ':' in val:
                    try:
                        pid, os_num = val.split(':', 1)
                        parada_id = int(pid)
                        if len(os_num) in (5, 6) and os_num.isdigit():
                            confirmados_set.add((parada_id, os_num))
                    except (ValueError, TypeError):
                        pass
        # Pares exibidos na tabela (hidden input displayed_pairs: "pid:os,pid:os,...")
        displayed_pairs = set()
        raw = request.POST.get('displayed_pairs', '')
        for part in raw.split(','):
            part = part.strip()
            if ':' in part:
                try:
                    pid, os_num = part.split(':', 1)
                    parada_id = int(pid)
                    if len(os_num) in (5, 6) and os_num.isdigit():
                        displayed_pairs.add((parada_id, os_num))
                except (ValueError, TypeError):
                    pass
        parada_ids_displayed = {p for p, _ in displayed_pairs}
        # Remover confirmações que estavam na tabela mas usuário desmarcou
        for parada_id in parada_ids_displayed:
            for rec in ParadaMaquinaOS.objects.filter(parada_maquina_id=parada_id):
                pair = (rec.parada_maquina_id, rec.os_numero)
                if pair in displayed_pairs and pair not in confirmados_set:
                    rec.delete()
        # Criar confirmações marcadas
        for parada_id, os_num in confirmados_set:
            ParadaMaquinaOS.objects.get_or_create(
                parada_maquina_id=parada_id,
                os_numero=os_num,
                defaults={}
            )
        # Processar OS manuais (Paradas sem OS)
        for key in request.POST:
            if key.startswith('manual_os_'):
                try:
                    parada_id = int(key.replace('manual_os_', ''))
                    valor = request.POST.get(key, '').strip()
                    if valor:
                        # Extrair apenas dígitos 5-6
                        import re as re_mod
                        nums = re_mod.findall(r'\d{5,6}', valor)
                        for os_num in nums:
                            if len(os_num) in (5, 6):
                                ParadaMaquinaOS.objects.get_or_create(
                                    parada_maquina_id=parada_id,
                                    os_numero=os_num,
                                    defaults={}
                                )
                except (ValueError, TypeError):
                    pass
        # Para paradas sem OS: se campo manual vazio, remover associação existente
        raw_manual_ids = request.POST.get('displayed_manual_ids', '')
        for pid_str in raw_manual_ids.split(','):
            try:
                parada_id = int(pid_str.strip())
                valor = request.POST.get(f'manual_os_{parada_id}', '').strip()
                if not valor:
                    ParadaMaquinaOS.objects.filter(parada_maquina_id=parada_id).delete()
            except (ValueError, TypeError):
                pass
        messages.success(request, 'Associações salvas com sucesso.')
        redirect_params = []
        if ano_post:
            redirect_params.append(f'ano={ano_post}')
        for m in (meses_post or []):
            redirect_params.append(f'mes={m}')
        url = reverse('analise_maquina_por_parada')
        if redirect_params:
            url += '?' + '&'.join(redirect_params)
        return redirect(url)

    # Anos e meses disponíveis (igual analise_paradas_maquina)
    anos_disponiveis = ParadaMaquina.objects.exclude(data__isnull=True).values_list('data__year', flat=True).distinct().order_by('-data__year')
    meses_disponiveis = {}
    for ano in anos_disponiveis:
        meses = ParadaMaquina.objects.filter(data__year=ano).exclude(data__isnull=True).values_list('data__month', flat=True).distinct().order_by('data__month')
        meses_disponiveis[ano] = list(meses)

    # Filtros de ano e meses
    hoje = datetime.now()
    has_explicit_filter = 'ano' in request.GET  # usuário submeteu o formulário
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year

    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    elif not has_explicit_filter:
        # Primeira carga: aplicar filtro do mês atual
        meses_filtro_int = [hoje.month]

    # Queryset filtrado
    queryset_base = ParadaMaquina.objects.exclude(data__isnull=True)
    if ano_filtro:
        queryset_base = queryset_base.filter(data__year=ano_filtro)
        if meses_filtro_int:
            mes_conditions = Q()
            for mes in meses_filtro_int:
                mes_conditions |= Q(data__month=mes)
            queryset_base = queryset_base.filter(mes_conditions)

    # Regex melhorada para Ordem de Serviço: standalone 5-6 dígitos OU após prefixos comuns
    os_standalone = re.compile(r'\b\d{5,6}\b')
    # Prefixos: OS, OS12345, OS 12345, Nº 12345, #12345, ordem 12345, ref 12345, etc.
    os_com_prefixo = re.compile(
        r'(?:OS|ordem|n[°º]|#|ref|numero|número)(?:\s*:?\s*)?(\d{5,6})\b',
        re.IGNORECASE
    )

    def extract_os(text):
        if not text or not isinstance(text, str):
            return []
        found = set()
        found.update(os_standalone.findall(text))
        found.update(os_com_prefixo.findall(text))
        # Sequência 5-6 dígitos após separadores (ex: "ref.12345", "-12345", "(12345)")
        found.update(re.findall(r'[\s\-_\.\/\(]\s*(\d{5,6})(?=[\s\-_\.\/\)]|$)', text))
        return list(found)

    # Analisar cada registro
    registros_com_os = []
    registros_sem_os = []
    os_to_registros = defaultdict(list)
    total_registros = 0

    # Carregar associações já confirmadas
    parada_ids_com_os = []
    for parada in queryset_base:
        total_registros += 1
        texto = ' '.join(filter(None, [str(parada.motivo or ''), str(parada.acao or '')]))
        os_encontradas = extract_os(texto)
        if os_encontradas:
            parada_ids_com_os.append(parada.id)
            registros_com_os.append({
                'parada': parada,
                'os_list': sorted(os_encontradas),
                'motivo_preview': (parada.motivo or '')[:200],
                'acao_preview': (parada.acao or '')[:200],
            })
            for os_num in os_encontradas:
                os_to_registros[os_num].append(parada.id)
        else:
            registros_sem_os.append(parada)

    os_confirmadas_set = set()
    if parada_ids_com_os:
        for rec in ParadaMaquinaOS.objects.filter(parada_maquina_id__in=parada_ids_com_os):
            os_confirmadas_set.add((rec.parada_maquina_id, rec.os_numero))
    # OS manuais para paradas sem detecção (podem ter mais de uma)
    manual_os_map = defaultdict(list)
    if registros_sem_os:
        parada_ids_sem = [p.id for p in registros_sem_os]
        for rec in ParadaMaquinaOS.objects.filter(parada_maquina_id__in=parada_ids_sem):
            manual_os_map[rec.parada_maquina_id].append(rec.os_numero)
    registros_sem_os_with_manual = [
        {'parada': p, 'manual_os_str': ', '.join(manual_os_map.get(p.id, []))}
        for p in registros_sem_os
    ]
    for item in registros_com_os:
        item['os_list_with_status'] = [
            (os_num, (item['parada'].id, os_num) in os_confirmadas_set)
            for os_num in item['os_list']
        ]

    # Resumo: OS únicas e contagem de paradas por OS
    os_resumo = []
    for os_num in sorted(os_to_registros.keys(), key=lambda x: (len(os_to_registros[x]), x), reverse=True):
        os_resumo.append({
            'os_num': os_num,
            'qtd_paradas': len(os_to_registros[os_num]),
        })

    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    anos_disponiveis_list = list(anos_disponiveis)
    anos_set = set(anos_disponiveis_list) | {hoje.year}
    if ano_filtro:
        anos_set.add(ano_filtro)
    anos_disponiveis_list = sorted(anos_set, reverse=True)
    if not anos_disponiveis_list:
        anos_disponiveis_list = [hoje.year]

    context = {
        'page_title': 'Análise Máquina por Parada (OS)',
        'active_page': 'analise_maquina_por_parada',
        'anos_disponiveis': anos_disponiveis_list,
        'meses_nomes': meses_nomes,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'total_registros': total_registros,
        'registros_com_os': registros_com_os,
        'registros_sem_os': registros_sem_os,
        'qtd_com_os': len(registros_com_os),
        'qtd_sem_os': len(registros_sem_os),
        'os_resumo': os_resumo,
        'os_unicas_count': len(os_to_registros),
        'qtd_confirmadas': len(os_confirmadas_set) + sum(len(v) for v in manual_os_map.values()),
        'displayed_pairs': ','.join(f"{item['parada'].id}:{os_num}" for item in registros_com_os for os_num in item['os_list']),
        'registros_sem_os_with_manual': registros_sem_os_with_manual,
        'displayed_manual_ids': ','.join(str(p.id) for p in registros_sem_os),
    }
    return render(request, 'parada_de_maquina/analise_maquina_por_parada.html', context)


def api_dados_diarios_paradas(request):
    """
    API endpoint para obter dados diários de paradas de máquina para um mês específico.
    Inclui série Frigorífico e Indústria conforme ConfigRecursoParadaMaquina (descr_recurso).
    """
    from app.models import ParadaMaquina, ConfigRecursoParadaMaquina
    from django.http import JsonResponse
    from datetime import datetime
    from calendar import monthrange
    from decimal import Decimal
    
    try:
        ano = int(request.GET.get('ano', datetime.now().year))
        mes = int(request.GET.get('mes', datetime.now().month))
        
        primeiro_dia = datetime(ano, mes, 1).date()
        ultimo_dia = datetime(ano, mes, monthrange(ano, mes)[1]).date()
        
        hoje = datetime.now().date()
        if ultimo_dia > hoje:
            ultimo_dia = hoje
        
        recursos_frigorifico = set(
            ConfigRecursoParadaMaquina.objects.filter(secao='frigorifico').values_list('descr_recurso', flat=True)
        )
        recursos_industria = set(
            ConfigRecursoParadaMaquina.objects.filter(secao='industria').values_list('descr_recurso', flat=True)
        )
        
        dias_labels = []
        dias_data = []
        dias_horas = []
        dias_data_frig = []
        dias_horas_frig = []
        dias_data_ind = []
        dias_horas_ind = []
        
        for dia in range(1, ultimo_dia.day + 1):
            data_dia = primeiro_dia.replace(day=dia)
            
            qs_dia = ParadaMaquina.objects.filter(data=data_dia)
            count = qs_dia.count()
            horas_dia = Decimal('0.00')
            for parada in qs_dia:
                if parada.dif_hora:
                    horas_dia += Decimal(str(parada.dif_hora))
            
            count_frig = qs_dia.filter(descr_linha_producao__in=recursos_frigorifico).count() if recursos_frigorifico else 0
            horas_frig = Decimal('0.00')
            for parada in qs_dia.filter(descr_linha_producao__in=recursos_frigorifico) if recursos_frigorifico else []:
                if parada.dif_hora:
                    horas_frig += Decimal(str(parada.dif_hora))
            count_ind = qs_dia.filter(descr_linha_producao__in=recursos_industria).count() if recursos_industria else 0
            horas_ind = Decimal('0.00')
            for parada in qs_dia.filter(descr_linha_producao__in=recursos_industria) if recursos_industria else []:
                if parada.dif_hora:
                    horas_ind += Decimal(str(parada.dif_hora))
            
            dias_labels.append(data_dia.strftime('%d/%m'))
            dias_data.append(count)
            dias_horas.append(float(horas_dia))
            dias_data_frig.append(count_frig)
            dias_horas_frig.append(float(horas_frig))
            dias_data_ind.append(count_ind)
            dias_horas_ind.append(float(horas_ind))
        
        secao = request.GET.get('secao', '').strip().lower()
        if secao == 'frigorifico':
            return JsonResponse({
                'labels': dias_labels,
                'data': dias_data_frig,
                'horas': dias_horas_frig
            })
        if secao == 'industria':
            return JsonResponse({
                'labels': dias_labels,
                'data': dias_data_ind,
                'horas': dias_horas_ind
            })
        return JsonResponse({
            'labels': dias_labels,
            'data': dias_data,
            'horas': dias_horas,
            'data_frig': dias_data_frig,
            'horas_frig': dias_horas_frig,
            'data_ind': dias_data_ind,
            'horas_ind': dias_horas_ind
        })
        
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': str(e)}, status=400)


def deletar_requisicoes_almoxarifado(request):
    """Deletar requisições de almoxarifado (individual ou em lote)"""
    from app.models import RequisicaoAlmoxarifado
    from django.urls import reverse
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('consultar_requisicoes_almoxarifado')
    
    # Verificar se é deleção individual ou em lote
    requisicao_id = request.POST.get('requisicao_id', None)
    requisicoes_ids = request.POST.getlist('requisicoes_ids', [])
    
    deleted_count = 0
    
    try:
        if requisicao_id:
            # Deleção individual
            try:
                requisicao = RequisicaoAlmoxarifado.objects.get(id=requisicao_id)
                requisicao.delete()
                deleted_count = 1
                messages.success(request, 'Requisição deletada com sucesso.')
            except RequisicaoAlmoxarifado.DoesNotExist:
                messages.error(request, 'Requisição não encontrada.')
        
        elif requisicoes_ids:
            # Deleção em lote
            try:
                ids = [int(id) for id in requisicoes_ids if id]
                if ids:
                    deleted_count = RequisicaoAlmoxarifado.objects.filter(id__in=ids).delete()[0]
                    if deleted_count > 0:
                        messages.success(request, f'{deleted_count} requisição(ões) deletada(s) com sucesso.')
                    else:
                        messages.warning(request, 'Nenhuma requisição foi deletada.')
                else:
                    messages.error(request, 'Nenhuma requisição selecionada.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Erro ao processar IDs: {str(e)}')
        else:
            messages.error(request, 'Nenhuma requisição especificada para deletar.')
    
    except Exception as e:
        messages.error(request, f'Erro ao deletar requisição(ões): {str(e)}')
        import traceback
        print(f"Erro ao deletar requisições: {traceback.format_exc()}")
    
    # Redirecionar mantendo os filtros
    redirect_url = reverse('consultar_requisicoes_almoxarifado')
    params = []
    
    # Preservar todos os parâmetros GET
    if request.GET.get('ano'):
        params.append(f"ano={request.GET.get('ano')}")
    for mes in request.GET.getlist('mes'):
        params.append(f"mes={mes}")
    if request.GET.get('search'):
        params.append(f"search={request.GET.get('search')}")
    if request.GET.get('filter_data'):
        params.append(f"filter_data={request.GET.get('filter_data')}")
    if request.GET.get('filter_item'):
        params.append(f"filter_item={request.GET.get('filter_item')}")
    if request.GET.get('filter_descricao'):
        params.append(f"filter_descricao={request.GET.get('filter_descricao')}")
    if request.GET.get('filter_quantidade'):
        params.append(f"filter_quantidade={request.GET.get('filter_quantidade')}")
    if request.GET.get('filter_valor'):
        params.append(f"filter_valor={request.GET.get('filter_valor')}")
    if request.GET.get('filter_centro'):
        params.append(f"filter_centro={request.GET.get('filter_centro')}")
    if request.GET.get('filter_usuario_criou'):
        params.append(f"filter_usuario_criou={request.GET.get('filter_usuario_criou')}")
    if request.GET.get('filter_usuario_atend'):
        params.append(f"filter_usuario_atend={request.GET.get('filter_usuario_atend')}")
    if request.GET.get('filter_operacao'):
        params.append(f"filter_operacao={request.GET.get('filter_operacao')}")
    if request.GET.get('filter_local'):
        params.append(f"filter_local={request.GET.get('filter_local')}")
    if request.GET.get('page'):
        params.append(f"page={request.GET.get('page')}")
    
    if params:
        redirect_url += '?' + '&'.join(params)
    
    return redirect(redirect_url)


def analise_requisicoes_data_importada(request):
    """Análise de datas importadas - Verifica quais dias têm dados de requisições"""
    from app.models import RequisicaoAlmoxarifado
    from datetime import datetime, date
    from calendar import monthrange
    from collections import defaultdict
    
    # Obter todas as datas únicas que têm dados
    datas_com_dados = RequisicaoAlmoxarifado.objects.exclude(
        data_requisicao__isnull=True
    ).values_list('data_requisicao', flat=True).distinct()
    
    # Converter para set para busca rápida
    datas_set = set(datas_com_dados)
    
    # Obter ano atual e ano anterior (para análise completa)
    hoje = datetime.now().date()
    ano_atual = hoje.year
    mes_atual = hoje.month
    
    # Organizar dados por mês
    meses_dados = defaultdict(dict)
    
    # Analisar cada mês do ano atual
    for mes in range(1, 13):
        # Obter número de dias no mês
        num_dias = monthrange(ano_atual, mes)[1]
        
        # Para cada dia do mês
        for dia in range(1, num_dias + 1):
            data_atual = date(ano_atual, mes, dia)
            
            # Verificar se já passou (não verificar dias futuros)
            if data_atual > hoje:
                break
            
            # Verificar se tem dados para este dia
            tem_dados = data_atual in datas_set
            
            # Traduzir dia da semana para português
            dias_semana_pt = {
                'Monday': 'Segunda',
                'Tuesday': 'Terça',
                'Wednesday': 'Quarta',
                'Thursday': 'Quinta',
                'Friday': 'Sexta',
                'Saturday': 'Sábado',
                'Sunday': 'Domingo'
            }
            dia_semana_en = data_atual.strftime('%A')
            dia_semana_pt = dias_semana_pt.get(dia_semana_en, dia_semana_en)
            
            # Armazenar informação do dia
            meses_dados[mes][dia] = {
                'data': data_atual,
                'tem_dados': tem_dados,
                'dia_semana': dia_semana_pt,
            }
    
    # Calcular estatísticas por mês
    meses_estatisticas = {}
    for mes, dias in meses_dados.items():
        total_dias = len(dias)
        dias_com_dados = sum(1 for dia_info in dias.values() if dia_info['tem_dados'])
        dias_sem_dados = total_dias - dias_com_dados
        percentual_completo = (dias_com_dados / total_dias * 100) if total_dias > 0 else 0
        
        meses_estatisticas[mes] = {
            'total_dias': total_dias,
            'dias_com_dados': dias_com_dados,
            'dias_sem_dados': dias_sem_dados,
            'percentual_completo': round(percentual_completo, 2),
        }
    
    # Nomes dos meses em português
    nomes_meses = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Estatísticas gerais
    total_dias_ano = sum(stats['total_dias'] for stats in meses_estatisticas.values())
    total_dias_com_dados = sum(stats['dias_com_dados'] for stats in meses_estatisticas.values())
    total_dias_sem_dados = sum(stats['dias_sem_dados'] for stats in meses_estatisticas.values())
    percentual_geral = (total_dias_com_dados / total_dias_ano * 100) if total_dias_ano > 0 else 0
    
    # Converter defaultdict para dict regular e ordenar meses
    meses_dados_dict = {}
    for mes in sorted(meses_dados.keys()):
        meses_dados_dict[mes] = dict(sorted(meses_dados[mes].items()))
    
    context = {
        'page_title': 'Análise de Datas Importadas',
        'active_page': 'analise_requisicoes_data_importada',
        'ano_atual': ano_atual,
        'mes_atual': mes_atual,
        'hoje': hoje,
        'meses_dados': meses_dados_dict,
        'meses_estatisticas': meses_estatisticas,
        'nomes_meses': nomes_meses,
        'total_dias_ano': total_dias_ano,
        'total_dias_com_dados': total_dias_com_dados,
        'total_dias_sem_dados': total_dias_sem_dados,
        'percentual_geral': round(percentual_geral, 2),
    }
    
    return render(request, 'almoxarifado/analise_requisicoes_data_importada.html', context)


def consultar_notas_fiscais(request):
    """Consultar/listar notas fiscais com filtros avançados"""
    from app.models import NotaFiscal
    from decimal import Decimal
    from django.db.models import Sum
    from datetime import datetime
    
    # Função auxiliar para parse de datas
    def parse_date(date_str):
        """Tenta fazer parse de data em vários formatos"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        if not date_str:
            return None
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses (None)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Se não há meses selecionados, usar todos os meses
    meses_para_mostrar = meses_filtro_int if meses_filtro_int else list(range(1, 13))
    
    # Obter anos disponíveis (baseado em data_emissao)
    anos_disponiveis = []
    todas_notas_anos = NotaFiscal.objects.exclude(data_emissao__isnull=True).exclude(data_emissao='')
    for nota in todas_notas_anos:
        data_emissao = parse_date(nota.data_emissao)
        if data_emissao:
            anos_disponiveis.append(data_emissao.year)
    anos_disponiveis = sorted(list(set(anos_disponiveis)), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    if hoje.year not in anos_disponiveis:
        anos_disponiveis.insert(0, hoje.year)
    
    # Busca geral
    search_query = request.GET.get('search', '').strip()
    notas_list = NotaFiscal.objects.all()
    
    # Filtrar por ano e mês usando somente data_emissao da NotaFiscal
    notas_filtradas_por_data = []
    for nota in notas_list:
        data_emissao = parse_date(nota.data_emissao)
        if data_emissao and data_emissao.year == ano_filtro:
            if not meses_filtro_int or data_emissao.month in meses_filtro_int:
                notas_filtradas_por_data.append(nota.id)
    
    if notas_filtradas_por_data:
        notas_list = notas_list.filter(id__in=notas_filtradas_por_data)
    else:
        # Se não há notas no ano/mês, retornar queryset vazio
        notas_list = NotaFiscal.objects.none()
    
    # Aplicar busca geral
    if search_query:
        # Tentar converter para número para busca exata em campos numéricos
        try:
            search_num = int(float(search_query))
            notas_list = notas_list.filter(
                Q(nota__icontains=search_query) |
                Q(emitente__icontains=search_query) |
                Q(nome_fantasia_emitente__icontains=search_query) |
                Q(total_nota=search_num)
            )
        except (ValueError, TypeError):
            notas_list = notas_list.filter(
                Q(nota__icontains=search_query) |
                Q(emitente__icontains=search_query) |
                Q(nome_fantasia_emitente__icontains=search_query) |
                Q(nome_unidade__icontains=search_query) |
                Q(situacao__icontains=search_query)
            )
    
    # Filtros específicos
    filtro_nota = request.GET.get('filtro_nota', '').strip()
    if filtro_nota:
        notas_list = notas_list.filter(nota__icontains=filtro_nota)
    
    filtro_emitente = request.GET.get('filtro_emitente', '').strip()
    if filtro_emitente:
        notas_list = notas_list.filter(
            Q(emitente__icontains=filtro_emitente) |
            Q(nome_fantasia_emitente__icontains=filtro_emitente)
        )
    
    filtro_unidade = request.GET.get('filtro_unidade', '').strip()
    if filtro_unidade:
        notas_list = notas_list.filter(
            Q(unidade__icontains=filtro_unidade) |
            Q(nome_unidade__icontains=filtro_unidade)
        )
    
    filtro_situacao = request.GET.get('filtro_situacao', '').strip()
    if filtro_situacao:
        notas_list = notas_list.filter(situacao__icontains=filtro_situacao)
    
    filtro_uso_contabil = request.GET.get('filtro_uso_contabil', '').strip()
    if filtro_uso_contabil:
        notas_list = notas_list.filter(uso_contabil__icontains=filtro_uso_contabil)
    
    filtro_data_emissao = request.GET.get('filtro_data_emissao', '').strip()
    if filtro_data_emissao:
        notas_list = notas_list.filter(data_emissao__icontains=filtro_data_emissao)
    
    filtro_data_autorizacao = request.GET.get('filtro_data_autorizacao', '').strip()
    if filtro_data_autorizacao:
        notas_list = notas_list.filter(data_autorizacao__icontains=filtro_data_autorizacao)
    
    ordenar_total = request.GET.get('ordenar_total', '').strip()
    
    # Obter valores únicos para os dropdowns de filtros (baseado nos dados já filtrados - ANTES da paginação)
    # Isso garante que os filtros mostrem apenas valores que existem na tabela filtrada
    situacoes_unicas_filtradas = notas_list.exclude(
        situacao__isnull=True
    ).exclude(
        situacao=''
    ).values_list('situacao', flat=True).distinct().order_by('situacao')
    
    uso_contabil_unicos_filtrados = notas_list.exclude(
        uso_contabil__isnull=True
    ).exclude(
        uso_contabil=''
    ).values_list('uso_contabil', flat=True).distinct().order_by('uso_contabil')
    
    # Obter datas únicas dos dados filtrados
    datas_emissao_unicas = notas_list.exclude(
        data_emissao__isnull=True
    ).exclude(
        data_emissao=''
    ).values_list('data_emissao', flat=True).distinct().order_by('data_emissao')
    
    datas_autorizacao_unicas = notas_list.exclude(
        data_autorizacao__isnull=True
    ).exclude(
        data_autorizacao=''
    ).values_list('data_autorizacao', flat=True).distinct().order_by('data_autorizacao')
    
    # Ordenar por data de emissão (mais recente primeiro) ou por total conforme seleção
    if ordenar_total == 'desc':
        notas_list = notas_list.order_by('-total_nota', '-data_emissao', '-created_at')
    elif ordenar_total == 'asc':
        notas_list = notas_list.order_by('total_nota', '-data_emissao', '-created_at')
    else:
        notas_list = notas_list.order_by('-data_emissao', '-created_at')
    
    # Estatísticas baseadas nos dados FILTRADOS (antes da paginação)
    total_count = notas_list.count()
    emitentes_count = notas_list.exclude(emitente__isnull=True).exclude(emitente='').values('emitente').distinct().count()
    unidades_count = notas_list.exclude(unidade__isnull=True).exclude(unidade='').values('unidade').distinct().count()
    _valor = notas_list.aggregate(s=Sum('total_nota'))['s']
    valor_total = Decimal(str(_valor)) if _valor is not None else Decimal('0.00')
    
    # Paginação
    paginator = Paginator(notas_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    notas = paginator.get_page(page_number)
    
    # Manter valores únicos gerais para dropdowns de filtros (todos os dados)
    situacoes_unicas = NotaFiscal.objects.exclude(
        situacao__isnull=True
    ).exclude(
        situacao=''
    ).values_list('situacao', flat=True).distinct().order_by('situacao')
    
    uso_contabil_unicos = NotaFiscal.objects.exclude(
        uso_contabil__isnull=True
    ).exclude(
        uso_contabil=''
    ).values_list('uso_contabil', flat=True).distinct().order_by('uso_contabil')
    
    unidades_unicas = NotaFiscal.objects.exclude(
        nome_unidade__isnull=True
    ).exclude(
        nome_unidade=''
    ).values_list('nome_unidade', flat=True).distinct().order_by('nome_unidade')
    
    context = {
        'page_title': 'Consultar Notas Fiscais',
        'active_page': 'consultar_notas_fiscais',
        'notas': notas,
        'total_count': total_count,
        'emitentes_count': emitentes_count,
        'unidades_count': unidades_count,
        'valor_total': valor_total,
        'situacoes_unicas': situacoes_unicas,
        'situacoes_unicas_filtradas': situacoes_unicas_filtradas,
        'uso_contabil_unicos': uso_contabil_unicos,
        'uso_contabil_unicos_filtrados': uso_contabil_unicos_filtrados,
        'datas_emissao_unicas': datas_emissao_unicas,
        'datas_autorizacao_unicas': datas_autorizacao_unicas,
        'unidades_unicas': unidades_unicas,
        # Filtros de ano e mês
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
        # Preservar filtros no contexto
        'search_query': search_query,
        'filtro_nota': filtro_nota,
        'filtro_emitente': filtro_emitente,
        'filtro_unidade': filtro_unidade,
        'filtro_situacao': filtro_situacao,
        'filtro_uso_contabil': filtro_uso_contabil,
        'filtro_data_emissao': filtro_data_emissao,
        'filtro_data_autorizacao': filtro_data_autorizacao,
        'ordenar_total': ordenar_total,
    }
    return render(request, 'consultar/consultar_notas_fiscais.html', context)


def visualizar_nota_fiscal(request, nota_id):
    """Visualizar detalhes de uma nota fiscal específica"""
    from app.models import NotaFiscal
    
    try:
        nota = NotaFiscal.objects.get(id=nota_id)
    except NotaFiscal.DoesNotExist:
        messages.error(request, 'Nota fiscal não encontrada.')
        return redirect('consultar_notas_fiscais')
    
    context = {
        'page_title': f'Visualizar Nota Fiscal {nota.nota or nota.id}',
        'active_page': 'consultar_notas_fiscais',
        'nota': nota,
    }
    return render(request, 'visualizar/visualizar_nota_fiscal.html', context)


def consultar_52_semanas(request):
    """Consultar/listar Semanas 52"""
    from app.models import Semana52
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Buscar todas as semanas
    semanas_list = Semana52.objects.all()
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        search_conditions = Q()
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(semana__icontains=search_query)
        )
        
        # Tentar buscar por data
        try:
            from datetime import datetime
            # Tentar diferentes formatos de data
            search_date = datetime.strptime(search_query, '%d/%m/%Y').date()
            search_conditions |= Q(inicio=search_date) | Q(fim=search_date)
        except (ValueError, TypeError):
            try:
                search_date = datetime.strptime(search_query, '%Y-%m-%d').date()
                search_conditions |= Q(inicio=search_date) | Q(fim=search_date)
            except (ValueError, TypeError):
                pass
        
        semanas_list = semanas_list.filter(search_conditions)
    
    # Filtros por coluna individual
    filter_semana = request.GET.get('filter_semana', '').strip()
    if filter_semana:
        semanas_list = semanas_list.filter(semana__icontains=filter_semana)
    
    filter_inicio = request.GET.get('filter_inicio', '').strip()
    if filter_inicio:
        # Tentar converter para data
        try:
            from datetime import datetime
            inicio_date = datetime.strptime(filter_inicio, '%d/%m/%Y').date()
            semanas_list = semanas_list.filter(inicio=inicio_date)
        except (ValueError, TypeError):
            # Se não for data válida, buscar como string na representação da data
            semanas_list = semanas_list.filter(inicio__icontains=filter_inicio)
    
    filter_fim = request.GET.get('filter_fim', '').strip()
    if filter_fim:
        # Tentar converter para data
        try:
            from datetime import datetime
            fim_date = datetime.strptime(filter_fim, '%d/%m/%Y').date()
            semanas_list = semanas_list.filter(fim=fim_date)
        except (ValueError, TypeError):
            # Se não for data válida, buscar como string na representação da data
            semanas_list = semanas_list.filter(fim__icontains=filter_fim)
    
    # Ordenar por data de início
    semanas_list = semanas_list.order_by('inicio', 'semana')
    
    # Adicionar campo calculado de duração
    from datetime import timedelta
    semanas_com_duracao = []
    for semana in semanas_list:
        duracao_dias = None
        if semana.inicio and semana.fim:
            duracao_dias = (semana.fim - semana.inicio).days + 1
        semanas_com_duracao.append({
            'semana': semana,
            'duracao_dias': duracao_dias
        })
    
    # Paginação
    paginator = Paginator(semanas_com_duracao, 25)  # 25 itens por página
    page_number = request.GET.get('page', 1)
    try:
        semanas = paginator.page(page_number)
    except:
        semanas = paginator.page(1)
    
    # Estatísticas
    total_count = Semana52.objects.count()
    
    context = {
        'page_title': 'Consultar 52 Semanas',
        'active_page': 'consultar_52_semanas',
        'semanas': semanas,
        'total_count': total_count,
        'search_query': search_query,
        'filter_semana': filter_semana,
        'filter_inicio': filter_inicio,
        'filter_fim': filter_fim,
    }
    return render(request, 'consultar/consultar_52_semanas.html', context)


def analise_geral_plano_preventiva_pcm(request):
    """Análise geral dos dados de Plano Preventiva PCM - Dashboard com estatísticas"""
    from app.models import (
        MeuPlanoPreventiva, PlanoPreventiva, RoteiroPreventiva,
        MaquinaPrimariaSecundaria, Maquina, MeuPlanoPreventivaDocumento, Semana52
    )
    from django.db.models import Count, Q
    from datetime import date, timedelta
    
    # ========== ESTATÍSTICAS MEU PLANO PREVENTIVA ==========
    total_meus_planos = MeuPlanoPreventiva.objects.count()
    meus_planos_com_roteiro = MeuPlanoPreventiva.objects.exclude(roteiro_preventiva__isnull=True).count()
    meus_planos_sem_roteiro = total_meus_planos - meus_planos_com_roteiro
    meus_planos_com_desc_detalhada = MeuPlanoPreventiva.objects.exclude(
        desc_detalhada_do_roteiro_preventiva__isnull=True
    ).exclude(desc_detalhada_do_roteiro_preventiva='').count()
    
    # Máquinas únicas em MeuPlanoPreventiva
    maquinas_unicas_meus_planos = MeuPlanoPreventiva.objects.exclude(
        cd_maquina__isnull=True
    ).values('cd_maquina').distinct().count()
    
    # Setores únicos
    setores_unicos_meus_planos = MeuPlanoPreventiva.objects.exclude(
        cd_setor__isnull=True
    ).exclude(cd_setor='').values('cd_setor').distinct().count()
    
    # Planos com documentos associados
    planos_com_documentos = MeuPlanoPreventiva.objects.filter(
        documentos_associados__isnull=False
    ).distinct().count()
    
    # ========== ESTATÍSTICAS ANÁLISE ROTEIRO/PLANO ==========
    total_planos = PlanoPreventiva.objects.count()
    total_roteiros = RoteiroPreventiva.objects.count()
    
    # Função para verificar correspondência exata
    def campos_correspondem(plano, roteiro):
        if not plano.cd_maquina or not roteiro.cd_maquina:
            return False
        if plano.cd_maquina != roteiro.cd_maquina:
            return False
        
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if descr_plano and descr_roteiro:
            if descr_plano != descr_roteiro:
                return False
        elif descr_plano or descr_roteiro:
            return False
        
        if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
            return False
        if plano.sequencia_tarefa != roteiro.cd_tarefamanu:
            return False
        
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if descr_tarefa_plano and descr_tarefa_roteiro:
            if descr_tarefa_plano != descr_tarefa_roteiro:
                return False
        elif descr_tarefa_plano or descr_tarefa_roteiro:
            return False
        
        if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
            return False
        if plano.sequencia_manutencao != roteiro.seq_seqplamanu:
            return False
        
        return True
    
    # Contar relacionamentos encontrados
    relacionamentos_encontrados = 0
    planos_processados = set()
    roteiros_processados = set()
    
    planos_list = PlanoPreventiva.objects.all()
    roteiros_list = RoteiroPreventiva.objects.all()
    
    for plano in planos_list:
        for roteiro in roteiros_list:
            if campos_correspondem(plano, roteiro):
                relacionamentos_encontrados += 1
                planos_processados.add(plano.id)
                roteiros_processados.add(roteiro.id)
                break
    
    planos_sem_relacao = total_planos - len(planos_processados)
    roteiros_sem_relacao = total_roteiros - len(roteiros_processados)
    
    # Relacionamentos já confirmados (salvos em MeuPlanoPreventiva)
    relacionamentos_confirmados = MeuPlanoPreventiva.objects.exclude(
        roteiro_preventiva__isnull=True
    ).count()
    
    relacionamentos_pendentes = max(0, relacionamentos_encontrados - relacionamentos_confirmados)
    
    # ========== ESTATÍSTICAS MÁQUINAS PRIMÁRIAS/SECUNDÁRIAS ==========
    maquinas_primarias_total = Maquina.objects.filter(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).count()
    
    maquinas_secundarias_total = Maquina.objects.exclude(
        descr_gerenc__iexact='MÁQUINAS PRINCIPAL'
    ).count()
    
    relacionamentos_maquinas = MaquinaPrimariaSecundaria.objects.count()
    
    # Máquinas primárias que têm relacionamentos
    primarias_com_relacionamentos = MaquinaPrimariaSecundaria.objects.values(
        'maquina_primaria_id'
    ).distinct().count()
    
    # Máquinas secundárias relacionadas
    secundarias_relacionadas = MaquinaPrimariaSecundaria.objects.values(
        'maquina_secundaria_id'
    ).distinct().count()
    
    # Máquinas primárias sem relacionamentos
    primarias_sem_relacionamentos = maquinas_primarias_total - primarias_com_relacionamentos
    
    # Máquinas secundárias disponíveis (não relacionadas)
    secundarias_disponiveis = maquinas_secundarias_total - secundarias_relacionadas
    
    # ========== PERCENTUAIS E TAXAS ==========
    taxa_cobertura_planos = (relacionamentos_encontrados / total_planos * 100) if total_planos > 0 else 0
    taxa_cobertura_roteiros = (relacionamentos_encontrados / total_roteiros * 100) if total_roteiros > 0 else 0
    taxa_confirmacao = (relacionamentos_confirmados / relacionamentos_encontrados * 100) if relacionamentos_encontrados > 0 else 0
    
    taxa_meus_planos_com_roteiro = (meus_planos_com_roteiro / total_meus_planos * 100) if total_meus_planos > 0 else 0
    taxa_meus_planos_com_desc = (meus_planos_com_desc_detalhada / total_meus_planos * 100) if total_meus_planos > 0 else 0
    
    taxa_primarias_relacionadas = (primarias_com_relacionamentos / maquinas_primarias_total * 100) if maquinas_primarias_total > 0 else 0
    
    # ========== ESTATÍSTICAS SEMANA52 ==========
    total_semanas = Semana52.objects.count()
    semanas_com_dados = Semana52.objects.exclude(inicio__isnull=True).exclude(fim__isnull=True).count()
    
    # Semanas ordenadas por data de início
    semanas_ordenadas = Semana52.objects.exclude(inicio__isnull=True).order_by('inicio')
    
    # Calcular estatísticas de datas
    semanas_list = list(semanas_ordenadas)
    if semanas_list:
        primeira_semana = semanas_list[0]
        ultima_semana = semanas_list[-1]
        primeira_data = primeira_semana.inicio if primeira_semana else None
        ultima_data = ultima_semana.fim if ultima_semana else None
        
        # Calcular duração total em dias
        if primeira_data and ultima_data:
            duracao_total_dias = (ultima_data - primeira_data).days + 1
        else:
            duracao_total_dias = 0
        
        # Calcular média de duração por semana e adicionar duração a cada semana
        duracoes = []
        semanas_com_duracao = []
        # Processar todas as semanas para calcular média, mas limitar preview a 10
        for semana in semanas_list:
            duracao_semana = None
            if semana.inicio and semana.fim:
                duracao_semana = (semana.fim - semana.inicio).days + 1
                duracoes.append(duracao_semana)
            # Adicionar apenas as primeiras 10 para preview
            if len(semanas_com_duracao) < 10:
                semanas_com_duracao.append({
                    'semana': semana,
                    'duracao_dias': duracao_semana
                })
        
        duracao_media = sum(duracoes) / len(duracoes) if duracoes else 0
    else:
        primeira_data = None
        ultima_data = None
        duracao_total_dias = 0
        duracao_media = 0
        semanas_com_duracao = []
    
    # Semanas do ano atual
    hoje = date.today()
    ano_atual = hoje.year
    semanas_ano_atual = Semana52.objects.filter(inicio__year=ano_atual).count()
    
    # Semanas futuras (ainda não iniciadas)
    semanas_futuras = Semana52.objects.filter(inicio__gt=hoje).count()
    
    # Semanas passadas (já finalizadas)
    semanas_passadas = Semana52.objects.filter(fim__lt=hoje).count()
    
    # Semana atual (hoje está entre inicio e fim)
    semana_atual = Semana52.objects.filter(inicio__lte=hoje, fim__gte=hoje).first()
    
    context = {
        'page_title': 'Análise Geral - Plano Preventiva PCM',
        'active_page': 'analise_geral_plano_preventiva_pcm',
        
        # Meus Planos Preventiva
        'total_meus_planos': total_meus_planos,
        'meus_planos_com_roteiro': meus_planos_com_roteiro,
        'meus_planos_sem_roteiro': meus_planos_sem_roteiro,
        'meus_planos_com_desc_detalhada': meus_planos_com_desc_detalhada,
        'maquinas_unicas_meus_planos': maquinas_unicas_meus_planos,
        'setores_unicos_meus_planos': setores_unicos_meus_planos,
        'planos_com_documentos': planos_com_documentos,
        'taxa_meus_planos_com_roteiro': taxa_meus_planos_com_roteiro,
        'taxa_meus_planos_com_desc': taxa_meus_planos_com_desc,
        
        # Análise Roteiro/Plano
        'total_planos': total_planos,
        'total_roteiros': total_roteiros,
        'relacionamentos_encontrados': relacionamentos_encontrados,
        'relacionamentos_confirmados': relacionamentos_confirmados,
        'relacionamentos_pendentes': relacionamentos_pendentes,
        'planos_sem_relacao': planos_sem_relacao,
        'roteiros_sem_relacao': roteiros_sem_relacao,
        'taxa_cobertura_planos': taxa_cobertura_planos,
        'taxa_cobertura_roteiros': taxa_cobertura_roteiros,
        'taxa_confirmacao': taxa_confirmacao,
        
        # Máquinas Primárias/Secundárias
        'maquinas_primarias_total': maquinas_primarias_total,
        'maquinas_secundarias_total': maquinas_secundarias_total,
        'relacionamentos_maquinas': relacionamentos_maquinas,
        'primarias_com_relacionamentos': primarias_com_relacionamentos,
        'secundarias_relacionadas': secundarias_relacionadas,
        'primarias_sem_relacionamentos': primarias_sem_relacionamentos,
        'secundarias_disponiveis': secundarias_disponiveis,
        'taxa_primarias_relacionadas': taxa_primarias_relacionadas,
        
        # Semana52
        'total_semanas': total_semanas,
        'semanas_com_dados': semanas_com_dados,
        'semanas_ordenadas': semanas_ordenadas[:10],  # Primeiras 10 para preview
        'semanas_com_duracao': semanas_com_duracao,  # Primeiras 10 com duração calculada
        'primeira_data': primeira_data,
        'ultima_data': ultima_data,
        'duracao_total_dias': duracao_total_dias,
        'duracao_media': duracao_media,
        'semanas_ano_atual': semanas_ano_atual,
        'semanas_futuras': semanas_futuras,
        'semanas_passadas': semanas_passadas,
        'semana_atual': semana_atual,
    }
    
    return render(request, 'planejamento/analise_geral_plano_preventiva_pcm.html', context)


def analise_ordens_de_servico(request):
    """Análise de Ordens de Serviço - Dashboard com estatísticas e filtros"""
    from app.models import OrdemServicoCorretiva, OrdemServicoPreventiva, OrdemServicoCorretivaFicha, CentroAtividade
    from django.db.models import Count, Q, Avg
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')
    
    # Valores padrão: ano atual e todos os meses
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Se nenhum mês foi selecionado, considerar todos os meses
    if not meses_filtro_int:
        meses_filtro_int = list(range(1, 13))
    
    # Função para fazer parse de dt_abertura_solicita
    def parse_dt_abertura_solicita(date_str):
        """Tenta fazer parse de dt_abertura_solicita em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para filtrar ordens por data
    def filtrar_ordens_por_data(queryset, ano, meses):
        """Filtra ordens baseado em dt_abertura_solicita"""
        ordens_filtradas = []
        for ordem in queryset:
            if ordem.dt_abertura_solicita:
                data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
                if data_parseada:
                    if data_parseada.year == ano:
                        if data_parseada.month in meses:
                            ordens_filtradas.append(ordem)
        return ordens_filtradas
    
    # Obter todas as ordens corretivas e filtrar
    todas_ordens_corretivas = OrdemServicoCorretiva.objects.all()
    ordens_corretivas_filtradas = filtrar_ordens_por_data(todas_ordens_corretivas, ano_filtro, meses_filtro_int)
    
    # Obter todas as ordens preventivas e filtrar
    todas_ordens_preventivas = OrdemServicoPreventiva.objects.all()
    ordens_preventivas_filtradas = filtrar_ordens_por_data(todas_ordens_preventivas, ano_filtro, meses_filtro_int)
    
    # Combinar ordens filtradas para estatísticas gerais
    ordens_filtradas = ordens_corretivas_filtradas + ordens_preventivas_filtradas
    
    # Estatísticas básicas (filtradas)
    total_corretivas = len(ordens_corretivas_filtradas)
    total_preventivas = len(ordens_preventivas_filtradas)
    total_ordens = total_corretivas + total_preventivas
    
    # ========== ESTATÍSTICAS ORDEMSERVICOCORRETIVA (FILTRADAS) ==========
    # Ordens por tipo de ordem (descr_tpordservtv) - MUITO IMPORTANTE
    ordens_por_tipo_os_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_tpordservtv:
            ordens_por_tipo_os_dict[ordem.descr_tpordservtv] += 1
    ordens_por_tipo_os = sorted(ordens_por_tipo_os_dict.items(), key=lambda x: x[1], reverse=True)
    tipos_os_labels = [item[0][:40] for item in ordens_por_tipo_os[:10]]
    tipos_os_data = [item[1] for item in ordens_por_tipo_os[:10]]
    
    # Ordens por setor (top 10)
    ordens_por_setor_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_setormanut:
            ordens_por_setor_dict[ordem.descr_setormanut] += 1
    ordens_por_setor_list = sorted(ordens_por_setor_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    ordens_por_setor = [{'descr_setormanut': item[0], 'total': item[1]} for item in ordens_por_setor_list]
    setores_labels = [item[0][:30] for item in ordens_por_setor_list]
    setores_data = [item[1] for item in ordens_por_setor_list]
    
    # Ordens por unidade (top 10)
    ordens_por_unidade_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nome_unid:
            ordens_por_unidade_dict[ordem.nome_unid] += 1
    ordens_por_unidade_list = sorted(ordens_por_unidade_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    ordens_por_unidade = [{'nome_unid': item[0], 'total': item[1]} for item in ordens_por_unidade_list]
    unidades_labels = [item[0][:30] for item in ordens_por_unidade_list]
    unidades_data = [item[1] for item in ordens_por_unidade_list]
    
    # Ordens por tipo de manutenção
    ordens_por_tipo_manut_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_tpmanuttv:
            ordens_por_tipo_manut_dict[ordem.descr_tpmanuttv] += 1
    ordens_por_tipo_manut_list = sorted(ordens_por_tipo_manut_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    ordens_por_tipo_manut = [{'descr_tpmanuttv': item[0], 'total': item[1]} for item in ordens_por_tipo_manut_list]
    
    # Ordens por situação
    ordens_por_situacao_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_sitordsetv:
            ordens_por_situacao_dict[ordem.descr_sitordsetv] += 1
    ordens_por_situacao = sorted(ordens_por_situacao_dict.items(), key=lambda x: x[1], reverse=True)

    # KPI Row 2: Ordens Abertas, Fechadas, Solicitações (baseado em descr_sitordsetv)
    ordens_abertas = sum(1 for ordem in ordens_filtradas if ordem.descr_sitordsetv and ordem.descr_sitordsetv.strip().upper() == 'ORDEM DE SERVIÇO EM EXECUÇÃO')
    ordens_fechadas = sum(1 for ordem in ordens_filtradas if ordem.descr_sitordsetv and ordem.descr_sitordsetv.strip().upper() == 'ORDEM DE SERVIÇO FECHADA')
    solicitacoes = sum(1 for ordem in ordens_filtradas if ordem.descr_sitordsetv and ordem.descr_sitordsetv.strip().upper() == 'SOLICITAÇÃO ABERTA')
    # Tempo Total Necessário: (Ordens Abertas + Solicitações) * 15 min → horas
    tempo_total_necessario_horas = (ordens_abertas + solicitacoes) * 15 / 60
    
    # Ordens com e sem máquina
    ordens_com_maquina = sum(1 for ordem in ordens_filtradas if ordem.cd_maquina)
    ordens_sem_maquina = total_ordens - ordens_com_maquina
    
    # Ordens com e sem funcionário executor
    ordens_com_executor = sum(1 for ordem in ordens_filtradas if ordem.nm_func_exec and ordem.nm_func_exec.strip())
    ordens_sem_executor = total_ordens - ordens_com_executor
    
    # Ordens com e sem funcionário solicitante
    ordens_com_solicitante = sum(1 for ordem in ordens_filtradas if ordem.nm_func_solic_os and ordem.nm_func_solic_os.strip())
    ordens_sem_solicitante = total_ordens - ordens_com_solicitante
    
    # Top 10 máquinas com mais ordens
    maquinas_dict = defaultdict(int)
    maquinas_desc = {}
    for ordem in ordens_filtradas:
        if ordem.cd_maquina:
            maquinas_dict[ordem.cd_maquina] += 1
            if ordem.descr_maquina:
                maquinas_desc[ordem.cd_maquina] = ordem.descr_maquina
    top_maquinas_list = sorted(maquinas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_maquinas = [{'cd_maquina': item[0], 'descr_maquina': maquinas_desc.get(item[0], ''), 'total': item[1]} for item in top_maquinas_list]
    maquinas_labels = [f"{item['cd_maquina']} - {item['descr_maquina'][:40]}" for item in top_maquinas]
    maquinas_data = [item['total'] for item in top_maquinas]
    
    # Top 10 funcionários executores
    executores_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nm_func_exec:
            executores_dict[ordem.nm_func_exec] += 1
    top_executores_list = sorted(executores_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_executores = [{'nm_func_exec': item[0], 'total': item[1]} for item in top_executores_list]
    executores_labels = [item['nm_func_exec'][:30] for item in top_executores]
    executores_data = [item['total'] for item in top_executores]
    
    # Ordens por dia do período filtrado (X=dias, Y=quantidade de ordens)
    from calendar import monthrange
    ordens_por_dia = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.dt_abertura_solicita:
            data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_parseada and data_parseada.year == ano_filtro and data_parseada.month in meses_filtro_int:
                chave_dia = (ano_filtro, data_parseada.month, data_parseada.day)
                ordens_por_dia[chave_dia] += 1
    
    # Listar todos os dias do período filtrado (ordenados)
    dias_ordenados = []
    meses_abrev = {
        1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
    }
    for mes in meses_filtro_int:
        _, ultimo_dia = monthrange(ano_filtro, mes)
        for dia in range(1, ultimo_dia + 1):
            chave = (ano_filtro, mes, dia)
            dias_ordenados.append(chave)
    
    meses_labels = [f"{dia:02d}/{meses_abrev.get(mes, mes)}" for ano, mes, dia in dias_ordenados]
    meses_data = [ordens_por_dia.get((ano, mes, dia), 0) for ano, mes, dia in dias_ordenados]
    
    # ========== ESTATÍSTICAS ORDEMSERVICOCORRETIVAFICHA (FILTRADAS) ==========
    # Obter IDs das ordens corretivas filtradas (fichas só existem para corretivas)
    ordens_filtradas_ids = [ordem.id for ordem in ordens_corretivas_filtradas]
    
    # Fichas relacionadas às ordens filtradas
    # SQLite tem limite de 999 variáveis por query, então dividimos em chunks
    fichas_filtradas_list = []
    chunk_size = 500  # Usar 500 para estar bem abaixo do limite de 999
    for i in range(0, len(ordens_filtradas_ids), chunk_size):
        chunk_ids = ordens_filtradas_ids[i:i + chunk_size]
        if chunk_ids:
            chunk_fichas = OrdemServicoCorretivaFicha.objects.filter(ordem_servico_id__in=chunk_ids)
            fichas_filtradas_list.extend(list(chunk_fichas))
    
    total_fichas = len(fichas_filtradas_list)
    ordens_com_fichas = len(set(ficha.ordem_servico_id for ficha in fichas_filtradas_list))
    ordens_sem_fichas = total_corretivas - ordens_com_fichas  # Fichas só existem para corretivas
    
    # Média de fichas por ordem
    if ordens_com_fichas > 0:
        media_fichas_por_ordem = total_fichas / ordens_com_fichas
    else:
        media_fichas_por_ordem = 0
    
    # Top 10 ordens com mais fichas
    ordens_com_fichas_dict = defaultdict(int)
    for ficha in fichas_filtradas_list:
        ordens_com_fichas_dict[ficha.ordem_servico_id] += 1
    top_ordens_fichas_list = sorted(ordens_com_fichas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_ordens_fichas = []
    for ordem_id, num_fichas in top_ordens_fichas_list:
        ordem = next((o for o in ordens_filtradas if o.id == ordem_id), None)
        if ordem:
            top_ordens_fichas.append({
                'cd_ordemserv': ordem.cd_ordemserv,
                'descr_maquina': ordem.descr_maquina or '-',
                'num_fichas': num_fichas
            })
    
    # Top 10 funcionários executores de fichas
    executores_fichas_dict = defaultdict(int)
    for ficha in fichas_filtradas_list:
        if ficha.nm_func_exec_os:
            executores_fichas_dict[ficha.nm_func_exec_os] += 1
    top_executores_fichas_list = sorted(executores_fichas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_executores_fichas = [{'nm_func_exec_os': item[0], 'total': item[1]} for item in top_executores_fichas_list]
    executores_fichas_labels = [item['nm_func_exec_os'][:30] for item in top_executores_fichas]
    executores_fichas_data = [item['total'] for item in top_executores_fichas]
    
    # Percentuais
    taxa_ordens_com_maquina = (ordens_com_maquina / total_ordens * 100) if total_ordens > 0 else 0
    taxa_ordens_com_executor = (ordens_com_executor / total_ordens * 100) if total_ordens > 0 else 0
    taxa_ordens_com_fichas = (ordens_com_fichas / total_corretivas * 100) if total_corretivas > 0 else 0  # Fichas só existem para corretivas
    
    # Contar setores e unidades únicos
    setores_unicos = set()
    unidades_unicas = set()
    for ordem in ordens_filtradas:
        if ordem.cd_setormanut:
            setores_unicos.add(ordem.cd_setormanut)
        if ordem.nome_unid:
            unidades_unicas.add(ordem.nome_unid)
    setores_count = len(setores_unicos)
    unidades_count = len(unidades_unicas)
    
    # Obter lista de anos disponíveis (de ambas as tabelas)
    anos_disponiveis = set()
    for ordem in todas_ordens_corretivas:
        if ordem.dt_abertura_solicita:
            data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_parseada:
                anos_disponiveis.add(data_parseada.year)
    for ordem in todas_ordens_preventivas:
        if ordem.dt_abertura_solicita:
            data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_parseada:
                anos_disponiveis.add(data_parseada.year)
    anos_disponiveis = sorted(anos_disponiveis, reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
        7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    context = {
        'page_title': 'Análise de Ordens de Serviço',
        'active_page': 'analise_ordens_de_servico',
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'meses_nomes': meses_nomes,
        'anos_disponiveis': anos_disponiveis,
        'total_ordens': total_ordens,
        'total_corretivas': total_corretivas,
        'total_preventivas': total_preventivas,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        
        # OrdemServicoCorretiva
        'ordens_por_setor': ordens_por_setor,
        'ordens_por_unidade': ordens_por_unidade,
        'ordens_por_tipo_manut': ordens_por_tipo_manut,
        'ordens_por_situacao': ordens_por_situacao,
        'ordens_com_maquina': ordens_com_maquina,
        'ordens_sem_maquina': ordens_sem_maquina,
        'ordens_com_executor': ordens_com_executor,
        'ordens_sem_executor': ordens_sem_executor,
        'ordens_com_solicitante': ordens_com_solicitante,
        'ordens_sem_solicitante': ordens_sem_solicitante,
        'ordens_abertas': ordens_abertas,
        'ordens_fechadas': ordens_fechadas,
        'solicitacoes': solicitacoes,
        'tempo_total_necessario_horas': tempo_total_necessario_horas,
        'top_maquinas': top_maquinas,
        'top_executores': top_executores,
        'taxa_ordens_com_maquina': taxa_ordens_com_maquina,
        'taxa_ordens_com_executor': taxa_ordens_com_executor,
        
        # OrdemServicoCorretivaFicha
        'total_fichas': total_fichas,
        'ordens_com_fichas': ordens_com_fichas,
        'ordens_sem_fichas': ordens_sem_fichas,
        'media_fichas_por_ordem': media_fichas_por_ordem,
        'top_ordens_fichas': top_ordens_fichas,
        'top_executores_fichas': top_executores_fichas,
        'taxa_ordens_com_fichas': taxa_ordens_com_fichas,
        
        # Dados para gráficos (JSON)
        'tipos_os_labels': json.dumps(tipos_os_labels),
        'tipos_os_data': json.dumps(tipos_os_data),
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'maquinas_labels': json.dumps(maquinas_labels),
        'maquinas_data': json.dumps(maquinas_data),
        'executores_labels': json.dumps(executores_labels),
        'executores_data': json.dumps(executores_data),
        'executores_fichas_labels': json.dumps(executores_fichas_labels),
        'executores_fichas_data': json.dumps(executores_fichas_data),
    }
    
    return render(request, 'ordens_de_servico/analise_geral_ordens_servico.html', context)


def analise_ordens_importadas_com_erro(request):
    """Análise de Ordens Importadas com Erro - Detecta padrões inconsistentes nos dados"""
    from app.models import OrdemServicoCorretiva
    from django.db.models import Count, Q
    from collections import defaultdict
    from datetime import datetime
    import json
    import re
    
    # Obter todas as ordens ordenadas por cd_ordemserv
    todas_ordens = list(OrdemServicoCorretiva.objects.all().order_by('cd_ordemserv'))
    total_ordens = int(len(todas_ordens))
    
    # Se não houver ordens, retornar página vazia
    if total_ordens == 0:
        context = {
            'page_title': 'Análise de Ordens Importadas com Erro',
            'active_page': 'analise_ordens_importadas_com_erro',
            'total_ordens': 0,
            'total_com_problemas': 0,
            'total_sem_problemas': 0,
            'percentual_problemas': 0,
            'intervalos_analise': [],
        }
        return render(request, 'ordens_de_servico/analise_ordens_importadas_com_erro.html', context)
    
    # Função para detectar padrão de data
    def detectar_formato_data(valor):
        """Detecta o formato de data mais comum"""
        if not valor or not str(valor).strip():
            return None
        
        valor_str = str(valor).strip()
        formatos_comuns = [
            r'\d{2}/\d{2}/\d{4}',  # DD/MM/YYYY
            r'\d{2}-\d{2}-\d{4}',  # DD-MM-YYYY
            r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
            r'\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}',  # DD/MM/YYYY HH:MM
        ]
        
        for formato in formatos_comuns:
            if re.match(formato, valor_str):
                return formato
        
        return None
    
    # Função para verificar se um valor segue o padrão da maioria
    def verificar_padrao_campo(campo_nome, valor, padroes_maioria):
        """Verifica se um valor segue o padrão da maioria"""
        problemas = []
        
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            # Campo vazio - verificar se a maioria tem valores
            if padroes_maioria.get(f'{campo_nome}_preenchido', 0) > 0.8:  # 80% preenchido
                problemas.append('Campo vazio quando deveria ter valor')
            return problemas
        
        valor_str = str(valor).strip()
        
        # Verificar formato de data para campos de data
        if 'dt_' in campo_nome.lower() or 'data' in campo_nome.lower():
            formato_esperado = padroes_maioria.get(f'{campo_nome}_formato_data')
            if formato_esperado:
                formato_atual = detectar_formato_data(valor_str)
                if formato_atual != formato_esperado:
                    problemas.append(f'Formato de data inconsistente')
        
        # Verificar comprimento para campos de texto
        if isinstance(valor, str):
            comprimento_medio = padroes_maioria.get(f'{campo_nome}_comprimento_medio', 0)
            comprimento_min = padroes_maioria.get(f'{campo_nome}_comprimento_min', 0)
            comprimento_max = padroes_maioria.get(f'{campo_nome}_comprimento_max', 0)
            
            # Garantir que são numéricos
            try:
                comprimento_medio = float(comprimento_medio) if comprimento_medio else 0
                comprimento_min = float(comprimento_min) if comprimento_min else 0
                comprimento_max = float(comprimento_max) if comprimento_max else 0
            except (ValueError, TypeError):
                comprimento_medio = 0
                comprimento_min = 0
                comprimento_max = 0
            
            if comprimento_medio > 0:
                comprimento_atual = len(valor_str)
                # Verificar se está muito fora do range normal
                if comprimento_max > comprimento_min:  # Se há variação
                    range_normal = comprimento_max - comprimento_min
                    if comprimento_atual < comprimento_min - range_normal * 0.5 or comprimento_atual > comprimento_max + range_normal * 0.5:
                        problemas.append(f'Comprimento anormal')
                elif abs(comprimento_atual - comprimento_medio) > comprimento_medio * 2:
                    problemas.append(f'Comprimento anormal')
        
        # Verificar caracteres especiais suspeitos (mas permitir acentos comuns)
        if valor_str and re.search(r'[^\w\s\-\/\.\:\(\)áàâãéêíóôõúçÁÀÂÃÉÊÍÓÔÕÚÇ]', valor_str, re.IGNORECASE):
            # Verificar se não são apenas caracteres de encoding comum
            if not re.match(r'^[\w\s\-\/\.\:\(\)áàâãéêíóôõúçÁÀÂÃÉÊÍÓÔÕÚÇ]+$', valor_str, re.IGNORECASE):
                problemas.append('Caracteres especiais suspeitos')
        
        # Verificar campos numéricos
        if 'cd_' in campo_nome.lower() and campo_nome.lower() not in ['cd_setormanut']:
            try:
                float(valor_str)
            except (ValueError, TypeError):
                if valor_str:  # Se não está vazio mas não é numérico
                    problemas.append('Valor não numérico em campo numérico')
        
        return problemas
    
    # Analisar padrões da maioria dos registros
    padroes_maioria = {}
    campos_analisar = [
        'dt_entrada', 'dt_abertura_solicita', 'dt_encordmanu', 'dt_aberordser',
        'dt_iniparmanu', 'dt_fimparmanu', 'dt_prev_exec',
        'cd_ordemserv', 'cd_maquina', 'cd_unid',
        'nome_unid', 'descr_setormanut', 'descr_maquina',
        'nm_func_exec', 'nm_func_solic_os', 'descr_tpordservtv',
        'descr_sitordsetv', 'descr_tpmanuttv', 'descr_clasorigos',
    ]
    
    # Calcular padrões para cada campo usando uma amostra maior
    # Garantir que total_ordens é um inteiro
    if not isinstance(total_ordens, int):
        total_ordens = int(total_ordens) if total_ordens else 0
    amostra_tamanho = 5000 if total_ordens > 5000 else total_ordens
    for campo_nome in campos_analisar:
        valores = []
        valores_preenchidos = 0
        formatos_data = defaultdict(int)
        valores_unicos = set()
        
        for ordem in todas_ordens[:amostra_tamanho]:
            try:
                valor = getattr(ordem, campo_nome, None)
                # Verificar se valor não é um HttpResponse ou outro tipo inválido
                from django.http import HttpResponse
                if isinstance(valor, HttpResponse):
                    continue
                
                if valor is not None:
                    if isinstance(valor, str):
                        if valor.strip():
                            valores_preenchidos += 1
                            valores.append(valor.strip())
                            valores_unicos.add(valor.strip())
                            # Detectar formato de data
                            if 'dt_' in campo_nome.lower():
                                formato = detectar_formato_data(valor.strip())
                                if formato:
                                    formatos_data[formato] += 1
                    else:
                        # Valor não-string (numérico, etc)
                        try:
                            valor_str = str(valor)
                            valores_preenchidos += 1
                            valores.append(valor_str)
                        except:
                            continue
            except Exception:
                continue
        
        total_amostra = amostra_tamanho
        padroes_maioria[f'{campo_nome}_preenchido'] = valores_preenchidos / total_amostra if total_amostra > 0 else 0
        
        if valores:
            # Garantir que todos os valores são strings válidas antes de calcular comprimentos
            valores_validos = []
            for v in valores:
                if isinstance(v, str):
                    valores_validos.append(v)
                else:
                    try:
                        valores_validos.append(str(v))
                    except:
                        continue
            
            if valores_validos:
                comprimentos_validos = []
                for v in valores_validos:
                    try:
                        # Verificar se v é uma string válida
                        if not isinstance(v, str):
                            continue
                        # Calcular comprimento
                        comprimento = len(v)
                        # Verificar se comprimento é um inteiro válido
                        if isinstance(comprimento, int) and not isinstance(comprimento, bool):
                            # Verificação adicional: garantir que não é HttpResponse
                            from django.http import HttpResponse
                            if not isinstance(comprimento, HttpResponse):
                                comprimentos_validos.append(comprimento)
                    except (TypeError, AttributeError, ValueError):
                        continue
                
                # Calcular estatísticas apenas se houver valores válidos
                if comprimentos_validos and len(comprimentos_validos) > 0:
                    try:
                        # Verificação final antes de usar min/max
                        # Filtrar qualquer coisa que não seja int
                        comprimentos_finais = [c for c in comprimentos_validos if isinstance(c, int) and not isinstance(c, bool) and type(c) is int]
                        
                        if comprimentos_finais and len(comprimentos_finais) > 0:
                            soma = sum(comprimentos_finais)
                            quantidade = len(comprimentos_finais)
                            min_val = comprimentos_finais[0]
                            max_val = comprimentos_finais[0]
                            
                            # Calcular min e max manualmente para evitar problemas
                            for c in comprimentos_finais:
                                if isinstance(c, int) and not isinstance(c, bool):
                                    if c < min_val:
                                        min_val = c
                                    if c > max_val:
                                        max_val = c
                            
                            padroes_maioria[f'{campo_nome}_comprimento_medio'] = float(soma / quantidade)
                            padroes_maioria[f'{campo_nome}_comprimento_min'] = float(min_val)
                            padroes_maioria[f'{campo_nome}_comprimento_max'] = float(max_val)
                        else:
                            padroes_maioria[f'{campo_nome}_comprimento_medio'] = 0.0
                            padroes_maioria[f'{campo_nome}_comprimento_min'] = 0.0
                            padroes_maioria[f'{campo_nome}_comprimento_max'] = 0.0
                    except Exception:
                        padroes_maioria[f'{campo_nome}_comprimento_medio'] = 0.0
                        padroes_maioria[f'{campo_nome}_comprimento_min'] = 0.0
                        padroes_maioria[f'{campo_nome}_comprimento_max'] = 0.0
                else:
                    padroes_maioria[f'{campo_nome}_comprimento_medio'] = 0.0
                    padroes_maioria[f'{campo_nome}_comprimento_min'] = 0.0
                    padroes_maioria[f'{campo_nome}_comprimento_max'] = 0.0
            else:
                padroes_maioria[f'{campo_nome}_comprimento_medio'] = 0.0
                padroes_maioria[f'{campo_nome}_comprimento_min'] = 0.0
                padroes_maioria[f'{campo_nome}_comprimento_max'] = 0.0
        
        if formatos_data:
            formato_mais_comum = max(formatos_data.items(), key=lambda x: x[1])[0]
            padroes_maioria[f'{campo_nome}_formato_data'] = formato_mais_comum
            padroes_maioria[f'{campo_nome}_formato_data_count'] = formatos_data[formato_mais_comum]
        
        # Detectar se campo tem valores muito variados (pode indicar problema)
        if len(valores_unicos) > 0:
            padroes_maioria[f'{campo_nome}_diversidade'] = len(valores_unicos) / len(valores) if valores else 0
    
    # Analisar cada ordem e detectar problemas
    ordens_com_problemas = []
    for ordem in todas_ordens:
        problemas_ordem = []
        
        for campo_nome in campos_analisar:
            valor = getattr(ordem, campo_nome, None)
            problemas_campo = verificar_padrao_campo(campo_nome, valor, padroes_maioria)
            if problemas_campo:
                problemas_ordem.extend([f'{campo_nome}: {p}' for p in problemas_campo])
        
        # Verificações adicionais específicas
        # Data de abertura antes da data de entrada (lógico)
        if ordem.dt_abertura_solicita and ordem.dt_entrada:
            try:
                dt_abertura_str = ordem.dt_abertura_solicita.split()[0] if ' ' in ordem.dt_abertura_solicita else ordem.dt_abertura_solicita
                dt_entrada_str = ordem.dt_entrada.split()[0] if ' ' in ordem.dt_entrada else ordem.dt_entrada
                
                # Tentar diferentes formatos
                formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']
                dt_abertura = None
                dt_entrada = None
                
                for fmt in formatos:
                    try:
                        dt_abertura = datetime.strptime(dt_abertura_str, fmt)
                        break
                    except:
                        continue
                
                for fmt in formatos:
                    try:
                        dt_entrada = datetime.strptime(dt_entrada_str, fmt)
                        break
                    except:
                        continue
                
                if dt_abertura and dt_entrada and dt_abertura < dt_entrada:
                    problemas_ordem.append('Data abertura anterior à data entrada')
            except:
                pass
        
        # Verificar se cd_ordemserv está presente e é válido
        if not ordem.cd_ordemserv:
            problemas_ordem.append('cd_ordemserv ausente')
        
        # Verificar campos críticos que geralmente devem estar preenchidos
        if not ordem.dt_abertura_solicita or not ordem.dt_abertura_solicita.strip():
            problemas_ordem.append('dt_abertura_solicita: Campo crítico vazio')
        
        # Verificar se descr_tpordservtv está presente (campo importante mencionado pelo usuário)
        if not ordem.descr_tpordservtv or not ordem.descr_tpordservtv.strip():
            problemas_ordem.append('descr_tpordservtv: Campo importante vazio')
        
        if problemas_ordem:
            ordens_com_problemas.append({
                'ordem': ordem,
                'problemas': problemas_ordem,
                'total_problemas': len(problemas_ordem)
            })
    
    # Agrupar por intervalos de 5000 baseado em cd_ordemserv
    intervalos_analise = []
    tamanho_intervalo = 5000
    
    # Encontrar o menor e maior cd_ordemserv
    min_cd_ordemserv = None
    max_cd_ordemserv = None
    for ordem in todas_ordens:
        if ordem.cd_ordemserv is not None:
            try:
                cd_num = int(ordem.cd_ordemserv)
                if min_cd_ordemserv is None or cd_num < min_cd_ordemserv:
                    min_cd_ordemserv = cd_num
                if max_cd_ordemserv is None or cd_num > max_cd_ordemserv:
                    max_cd_ordemserv = cd_num
            except (ValueError, TypeError):
                continue
    
    # Se não encontrou nenhum cd_ordemserv válido, retornar vazio
    if min_cd_ordemserv is None or max_cd_ordemserv is None:
        min_cd_ordemserv = 0
        max_cd_ordemserv = 0
    
    # Criar intervalos baseados em cd_ordemserv
    # Começar do múltiplo de 5000 mais próximo abaixo do min
    inicio_geral = (min_cd_ordemserv // tamanho_intervalo) * tamanho_intervalo
    fim_geral = ((max_cd_ordemserv // tamanho_intervalo) + 1) * tamanho_intervalo
    
    intervalo_numero = 1
    for intervalo_inicio in range(inicio_geral, fim_geral + 1, tamanho_intervalo):
        intervalo_fim = intervalo_inicio + tamanho_intervalo - 1
        
        # Filtrar ordens que estão neste intervalo de cd_ordemserv
        ordens_intervalo = []
        for ordem in todas_ordens:
            if ordem.cd_ordemserv is not None:
                try:
                    cd_num = int(ordem.cd_ordemserv)
                    if intervalo_inicio <= cd_num <= intervalo_fim:
                        ordens_intervalo.append(ordem)
                except (ValueError, TypeError):
                    continue
        
        if not ordens_intervalo:
            # Se não há ordens neste intervalo, pular
            continue
        
        ordens_intervalo_ids = {ordem.id for ordem in ordens_intervalo}
        
        # Encontrar problemas neste intervalo
        problemas_intervalo = []
        for item in ordens_com_problemas:
            if item['ordem'].id in ordens_intervalo_ids:
                problemas_intervalo.append(item)
        
        # Estatísticas do intervalo
        total_intervalo = len(ordens_intervalo)
        total_com_problemas_intervalo = len(problemas_intervalo)
        percentual_problemas = (total_com_problemas_intervalo / total_intervalo * 100) if total_intervalo > 0 else 0
        
        # Determinar status
        if percentual_problemas == 0:
            status = 'sem_problemas'
        elif percentual_problemas < 5:
            status = 'baixo'
        elif percentual_problemas < 15:
            status = 'medio'
        else:
            status = 'alto'
        
        # Obter primeira e última ordem do intervalo (por cd_ordemserv)
        ordens_ordenadas = sorted(ordens_intervalo, key=lambda x: int(x.cd_ordemserv) if x.cd_ordemserv else 0)
        primeira_ordem = ordens_ordenadas[0] if ordens_ordenadas else None
        ultima_ordem = ordens_ordenadas[-1] if ordens_ordenadas else None
        
        intervalos_analise.append({
            'inicio_idx': intervalo_numero,  # Número do intervalo
            'fim_idx': intervalo_numero,  # Mesmo número, pois é baseado em cd_ordemserv
            'inicio_numero': intervalo_inicio,  # Início do intervalo de cd_ordemserv
            'fim_numero': intervalo_fim,  # Fim do intervalo de cd_ordemserv
            'primeira_os': primeira_ordem.cd_ordemserv if primeira_ordem else None,
            'ultima_os': ultima_ordem.cd_ordemserv if ultima_ordem else None,
            'total': total_intervalo,
            'com_problemas': total_com_problemas_intervalo,
            'sem_problemas': total_intervalo - total_com_problemas_intervalo,
            'percentual_problemas': round(percentual_problemas, 2),
            'status': status,
            'ordens_problemas': problemas_intervalo[:100],  # Limitar a 100 por performance
        })
        
        intervalo_numero += 1
    
    # Estatísticas gerais
    total_com_problemas = len(ordens_com_problemas)
    percentual_geral = (total_com_problemas / total_ordens * 100) if total_ordens > 0 else 0
    
    context = {
        'page_title': 'Análise de Ordens Importadas com Erro',
        'active_page': 'analise_ordens_importadas_com_erro',
        'total_ordens': total_ordens,
        'total_com_problemas': total_com_problemas,
        'total_sem_problemas': total_ordens - total_com_problemas,
        'percentual_problemas': round(percentual_geral, 2),
        'intervalos_analise': intervalos_analise,
    }
    
    return render(request, 'ordens_de_servico/analise_ordens_importadas_com_erro.html', context)


def analise_faltantes_pelo_numero(request):
    """Análise de Faltantes pelo Número - Identifica números sequenciais faltantes em cd_ordemserv usando intervalos fixos de 5000"""
    from app.models import OrdemServicoCorretiva, OrdemServicoPreventiva
    from django.db.models import Min, Max, Count
    import json
    import builtins
    import math
    
    # Obter todos os valores de cd_ordemserv ordenados de ambas as tabelas
    ordens_corretivas = OrdemServicoCorretiva.objects.exclude(
        cd_ordemserv__isnull=True
    ).values_list('cd_ordemserv', flat=True)
    
    ordens_preventivas = OrdemServicoPreventiva.objects.exclude(
        cd_ordemserv__isnull=True
    ).values_list('cd_ordemserv', flat=True)
    
    # Combinar os valores de ambas as tabelas
    numeros_existentes = set(ordens_corretivas) | set(ordens_preventivas)
    
    # Estatísticas básicas
    total_ordens = len(numeros_existentes)
    total_corretivas = ordens_corretivas.count()
    total_preventivas = ordens_preventivas.count()
    
    if total_ordens == 0:
        context = {
            'page_title': 'Análise de Faltantes pelo Número',
            'active_page': 'analise_faltantes_pelo_numero',
            'total_ordens': 0,
            'total_corretivas': 0,
            'total_preventivas': 0,
            'min_numero': None,
            'max_numero': None,
            'intervalos_analise': [],
        }
        return render(request, 'ordens_de_servico/analise_faltantes_pelo_numero.html', context)
    
    # Obter min e max do conjunto combinado
    min_numero = min(numeros_existentes)
    max_numero = max(numeros_existentes)
    
    # Criar intervalos fixos de 5000
    intervalo_tamanho = 5000
    intervalos_analise = []
    
    # Começar do 0 ou do min_numero arredondado para baixo para múltiplo de 5000
    inicio_geral = (min_numero // intervalo_tamanho) * intervalo_tamanho
    fim_geral = ((max_numero // intervalo_tamanho) + 1) * intervalo_tamanho
    
    # Criar intervalos
    for intervalo_inicio in range(inicio_geral, fim_geral, intervalo_tamanho):
        intervalo_fim = intervalo_inicio + intervalo_tamanho - 1
        
        # Contar números existentes neste intervalo
        existentes_no_intervalo = sum(1 for num in numeros_existentes if intervalo_inicio <= num <= intervalo_fim)
        
        # Calcular faltantes
        total_esperado_intervalo = intervalo_tamanho
        faltantes_no_intervalo = total_esperado_intervalo - existentes_no_intervalo
        percentual_completo_intervalo = (existentes_no_intervalo / total_esperado_intervalo * 100) if total_esperado_intervalo > 0 else 0
        percentual_faltantes_intervalo = 100 - percentual_completo_intervalo
        
        # Determinar status do intervalo
        if faltantes_no_intervalo == 0:
            status = 'completo'
            status_class = 'success'
        elif percentual_faltantes_intervalo < 1:
            status = 'quase_completo'
            status_class = 'info'
        elif percentual_faltantes_intervalo < 5:
            status = 'bom'
            status_class = 'warning'
        else:
            status = 'critico'
            status_class = 'danger'
        
        intervalos_analise.append({
            'inicio': intervalo_inicio,
            'fim': intervalo_fim,
            'existentes': existentes_no_intervalo,
            'faltantes': faltantes_no_intervalo,
            'total_esperado': total_esperado_intervalo,
            'percentual_completo': round(percentual_completo_intervalo, 2),
            'percentual_faltantes': round(percentual_faltantes_intervalo, 2),
            'status': status,
            'status_class': status_class,
        })
    
    # Calcular totais gerais
    total_faltantes = sum(intervalo['faltantes'] for intervalo in intervalos_analise)
    total_esperado = sum(intervalo['total_esperado'] for intervalo in intervalos_analise)
    percentual_faltantes = (total_faltantes / total_esperado * 100) if total_esperado > 0 else 0
    percentual_completo = 100 - percentual_faltantes
    
    # Preparar dados para gráfico de distribuição por intervalo
    distribuicao_labels = [f"{intervalo['inicio']}-{intervalo['fim']}" for intervalo in intervalos_analise]
    distribuicao_existentes = [intervalo['existentes'] for intervalo in intervalos_analise]
    distribuicao_faltantes = [intervalo['faltantes'] for intervalo in intervalos_analise]
    
    context = {
        'page_title': 'Análise de Faltantes pelo Número',
        'active_page': 'analise_faltantes_pelo_numero',
        'total_ordens': total_ordens,
        'total_corretivas': total_corretivas,
        'total_preventivas': total_preventivas,
        'min_numero': min_numero,
        'max_numero': max_numero,
        'total_faltantes': total_faltantes,
        'total_esperado': total_esperado,
        'percentual_faltantes': round(percentual_faltantes, 2),
        'percentual_completo': round(percentual_completo, 2),
        'intervalos_analise': intervalos_analise,
        'distribuicao_labels': json.dumps(distribuicao_labels),
        'distribuicao_existentes': json.dumps(distribuicao_existentes),
        'distribuicao_faltantes': json.dumps(distribuicao_faltantes),
    }
    return render(request, 'ordens_de_servico/analise_faltantes_pelo_numero.html', context)


def config_analise_ordens(request):
    """Configuração de Análise de Ordens de Serviço"""
    from django.contrib import messages
    
    if request.method == 'POST':
        # Aqui você pode processar e salvar as configurações
        # Por enquanto, apenas mostra uma mensagem de sucesso
        messages.success(request, 'Configurações salvas com sucesso!')
        return redirect('config_analise_ordens')
    
    context = {
        'page_title': 'Configuração de Análise de Ordens de Serviço',
        'active_page': 'config_analise_ordens',
    }
    
    return render(request, 'ordens_de_servico/config_analise_ordens.html', context)


def agrupar_acoes_do_plano_por_data(request):
    """Agrupar ações do plano por data de execução"""
    from app.models import MeuPlanoPreventiva
    from django.db.models import Count, Q
    from collections import defaultdict
    from datetime import datetime
    
    # Buscar todos os planos
    planos = MeuPlanoPreventiva.objects.all().order_by('dt_execucao', 'cd_maquina', 'sequencia_manutencao')
    
    # Agrupar por data de execução
    planos_por_data = defaultdict(list)
    planos_sem_data = []
    
    for plano in planos:
        if plano.dt_execucao:
            # Tentar parsear a data (formato DD/MM/YYYY)
            try:
                # Remover espaços e tentar diferentes formatos
                data_str = plano.dt_execucao.strip()
                if '/' in data_str:
                    # Formato DD/MM/YYYY
                    data_obj = datetime.strptime(data_str, '%d/%m/%Y').date()
                elif '-' in data_str:
                    # Formato YYYY-MM-DD
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                else:
                    # Tentar outros formatos
                    data_obj = datetime.strptime(data_str, '%Y%m%d').date()
                
                planos_por_data[data_obj].append(plano)
            except (ValueError, AttributeError):
                # Se não conseguir parsear, adicionar aos sem data
                planos_sem_data.append(plano)
        else:
            planos_sem_data.append(plano)
    
    # Ordenar as datas
    datas_ordenadas = sorted(planos_por_data.keys())
    
    # Estatísticas
    total_planos = planos.count()
    total_com_data = sum(len(planos_por_data[data]) for data in datas_ordenadas)
    total_sem_data = len(planos_sem_data)
    total_datas_unicas = len(datas_ordenadas)
    
    # Agrupar por semana (opcional - para análise semanal)
    planos_por_semana = defaultdict(list)
    for data, planos_list in planos_por_data.items():
        # Calcular número da semana do ano
        semana_ano = data.isocalendar()[1]
        ano = data.year
        chave_semana = f"{ano}-W{semana_ano:02d}"
        planos_por_semana[chave_semana].extend(planos_list)
    
    semanas_ordenadas = sorted(planos_por_semana.keys())
    
    # Converter defaultdict para dict e criar lista de tuplas para facilitar acesso no template
    planos_por_data_list = [(data, planos_por_data[data]) for data in datas_ordenadas]
    planos_por_semana_list = [(semana, planos_por_semana[semana]) for semana in semanas_ordenadas]
    
    context = {
        'page_title': 'Agrupar Ações do Plano por Data',
        'active_page': 'agrupar_acoes_do_plano_por_data',
        'planos_por_data': dict(planos_por_data),
        'planos_por_data_list': planos_por_data_list,
        'datas_ordenadas': datas_ordenadas,
        'planos_sem_data': planos_sem_data,
        'planos_por_semana': dict(planos_por_semana),
        'planos_por_semana_list': planos_por_semana_list,
        'semanas_ordenadas': semanas_ordenadas,
        'total_planos': total_planos,
        'total_com_data': total_com_data,
        'total_sem_data': total_sem_data,
        'total_datas_unicas': total_datas_unicas,
    }
    
    return render(request, 'planejamento/agrupar_acoes_do_plano_por_data.html', context)


def criar_cronograma_planejado_preventiva(request):
    """Criar cronograma planejado de preventivas"""
    from app.models import MeuPlanoPreventiva, Semana52, Maquina
    from django.db.models import Q
    from datetime import datetime, date
    from collections import defaultdict
    
    # Parâmetros de seleção (para a nova função)
    selected_maquina_id = request.GET.get('maquina_id', None)
    selected_plano_id = request.GET.get('plano_id', None)
    selected_maquina = None
    selected_plano = None
    
    if selected_maquina_id:
        try:
            selected_maquina = Maquina.objects.get(id=selected_maquina_id)
        except Maquina.DoesNotExist:
            pass
    
    if selected_plano_id:
        try:
            selected_plano = MeuPlanoPreventiva.objects.get(id=selected_plano_id)
        except MeuPlanoPreventiva.DoesNotExist:
            pass
    
    # Buscar todas as máquinas e planos para popular os selects
    todas_maquinas = Maquina.objects.all().order_by('cd_maquina')
    todos_planos = MeuPlanoPreventiva.objects.all().order_by('cd_maquina', 'numero_plano', 'sequencia_manutencao')[:500]  # Limitar a 500 para performance
    
    # Buscar setores únicos para o filtro
    setores_unicos = Maquina.objects.exclude(
        cd_setormanut__isnull=True
    ).exclude(
        cd_setormanut=''
    ).values_list('cd_setormanut', flat=True).distinct().order_by('cd_setormanut')
    
    # Buscar todas as semanas do ano
    semanas = Semana52.objects.all().order_by('inicio')
    
    # Buscar todos os planos preventiva PCM
    planos = MeuPlanoPreventiva.objects.all().order_by('dt_execucao', 'cd_maquina', 'sequencia_manutencao')
    
    # Se uma máquina foi selecionada, filtrar planos por essa máquina
    if selected_maquina:
        planos = planos.filter(cd_maquina=selected_maquina.cd_maquina)
    
    # Se um plano foi selecionado, filtrar apenas esse plano
    if selected_plano:
        planos = planos.filter(id=selected_plano.id)
    
    # Buscar agendamentos de cronograma
    from app.models import AgendamentoCronograma
    agendamentos = AgendamentoCronograma.objects.all().select_related('maquina', 'plano_preventiva', 'semana').order_by('data_planejada')
    
    # Agrupar planos por semana
    planos_por_semana = defaultdict(list)
    planos_sem_data = []
    
    for plano in planos:
        if plano.dt_execucao:
            try:
                # Tentar parsear a data
                data_str = plano.dt_execucao.strip()
                if '/' in data_str:
                    data_obj = datetime.strptime(data_str, '%d/%m/%Y').date()
                elif '-' in data_str:
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                else:
                    data_obj = datetime.strptime(data_str, '%Y%m%d').date()
                
                # Encontrar a semana correspondente
                semana_encontrada = None
                for semana in semanas:
                    if semana.inicio and semana.fim:
                        if semana.inicio <= data_obj <= semana.fim:
                            semana_encontrada = semana
                            break
                
                if semana_encontrada:
                    planos_por_semana[semana_encontrada].append(plano)
                else:
                    planos_sem_data.append(plano)
            except (ValueError, AttributeError):
                planos_sem_data.append(plano)
        else:
            planos_sem_data.append(plano)
    
    # Agrupar agendamentos por semana
    agendamentos_por_semana = defaultdict(list)
    agendamentos_sem_semana = []
    
    for agendamento in agendamentos:
        if agendamento.semana:
            agendamentos_por_semana[agendamento.semana].append(agendamento)
        else:
            agendamentos_sem_semana.append(agendamento)
    
    # Estatísticas
    total_planos = planos.count()
    total_com_semana = sum(len(planos_por_semana[semana]) for semana in semanas)
    total_sem_semana = len(planos_sem_data)
    total_agendamentos = agendamentos.count()
    total_agendamentos_com_semana = sum(len(agendamentos_por_semana[semana]) for semana in semanas)
    
    # Criar lista de tuplas para facilitar acesso no template
    planos_por_semana_list = [(semana, planos_por_semana[semana]) for semana in semanas if semana in planos_por_semana]
    agendamentos_por_semana_list = [(semana, agendamentos_por_semana[semana]) for semana in semanas if semana in agendamentos_por_semana]
    
    # Criar lista combinada de semanas com agendamentos e planos
    semanas_com_dados = []
    for semana in semanas:
        agendamentos_semana = agendamentos_por_semana.get(semana, [])
        planos_semana = planos_por_semana.get(semana, [])
        if agendamentos_semana or planos_semana:
            semanas_com_dados.append((semana, agendamentos_semana, planos_semana))
    
    context = {
        'page_title': 'Criar Calendário Planejado de Preventivas',
        'active_page': 'criar_cronograma_planejado_preventiva',
        'semanas': semanas,
        'planos_por_semana': dict(planos_por_semana),
        'planos_por_semana_list': planos_por_semana_list,
        'planos_sem_data': planos_sem_data,
        'agendamentos_por_semana': dict(agendamentos_por_semana),
        'agendamentos_por_semana_list': agendamentos_por_semana_list,
        'agendamentos_sem_semana': agendamentos_sem_semana,
        'semanas_com_dados': semanas_com_dados,
        'total_planos': total_planos,
        'total_com_semana': total_com_semana,
        'total_sem_semana': total_sem_semana,
        'total_agendamentos': total_agendamentos,
        'total_agendamentos_com_semana': total_agendamentos_com_semana,
        'selected_maquina': selected_maquina,
        'selected_plano': selected_plano,
        'selected_maquina_id': selected_maquina_id,
        'selected_plano_id': selected_plano_id,
        'todas_maquinas': todas_maquinas,
        'todos_planos': todos_planos,
        'setores_unicos': setores_unicos,
    }
    
    return render(request, 'planejamento/criar_cronograma_planejado_preventiva.html', context)


def api_search_maquinas(request):
    """API endpoint para buscar máquinas"""
    from app.models import Maquina
    from django.http import JsonResponse
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    maquinas = Maquina.objects.all()
    
    # Buscar por código ou descrição
    try:
        query_num = int(float(query))
        maquinas = maquinas.filter(
            Q(cd_maquina=query_num) |
            Q(descr_maquina__icontains=query)
        )
    except (ValueError, TypeError):
        maquinas = maquinas.filter(
            Q(descr_maquina__icontains=query) |
            Q(cd_setormanut__icontains=query) |
            Q(nome_unid__icontains=query)
        )
    
    # Limitar a 20 resultados
    maquinas = maquinas[:20]
    
    results = []
    for maquina in maquinas:
        results.append({
            'id': maquina.id,
            'cd_maquina': maquina.cd_maquina,
            'descr_maquina': maquina.descr_maquina or '',
            'cd_setormanut': maquina.cd_setormanut or '',
            'nome_unid': maquina.nome_unid or '',
        })
    
    return JsonResponse({'results': results})


def api_search_planos_pcm(request):
    """API endpoint para buscar planos PCM"""
    from app.models import MeuPlanoPreventiva
    from django.http import JsonResponse
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    planos = MeuPlanoPreventiva.objects.all()
    
    # Buscar por código da máquina, número do plano ou descrição
    try:
        query_num = int(float(query))
        planos = planos.filter(
            Q(cd_maquina=query_num) |
            Q(numero_plano=query_num) |
            Q(descr_maquina__icontains=query) |
            Q(descr_tarefa__icontains=query)
        )
    except (ValueError, TypeError):
        planos = planos.filter(
            Q(descr_maquina__icontains=query) |
            Q(descr_tarefa__icontains=query) |
            Q(descr_plano__icontains=query)
        )
    
    # Limitar a 20 resultados
    planos = planos[:20]
    
    results = []
    for plano in planos:
        results.append({
            'id': plano.id,
            'cd_maquina': plano.cd_maquina,
            'descr_maquina': plano.descr_maquina or '',
            'numero_plano': plano.numero_plano,
            'sequencia_manutencao': plano.sequencia_manutencao,
            'sequencia_tarefa': plano.sequencia_tarefa,
            'descr_tarefa': plano.descr_tarefa or '',
        })
    
    return JsonResponse({'results': results})


def salvar_agendamentos_cronograma(request):
    """Salvar múltiplos agendamentos de cronograma com suporte a periodicidade"""
    from app.models import AgendamentoCronograma
    from django.http import JsonResponse
    from datetime import datetime, date, timedelta
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        agendamentos_data = data.get('agendamentos', [])
        
        if not agendamentos_data:
            return JsonResponse({'success': False, 'error': 'Nenhum agendamento fornecido'})
        
        saved_count = 0
        errors = []
        
        for agendamento_data in agendamentos_data:
            try:
                tipo = agendamento_data.get('tipo')
                item_id = agendamento_data.get('id')
                data_planejada_str = agendamento_data.get('data_planejada')
                nome_grupo = agendamento_data.get('nome_grupo', '').strip() or None
                periodicidade = agendamento_data.get('periodicidade')
                
                if not tipo or not item_id or not data_planejada_str:
                    errors.append(f'Agendamento inválido: campos obrigatórios faltando')
                    continue
                
                # Parse da data inicial
                try:
                    data_planejada = datetime.strptime(data_planejada_str, '%Y-%m-%d').date()
                except ValueError:
                    errors.append(f'Data inválida: {data_planejada_str}')
                    continue
                
                # Obter objeto máquina ou plano
                maquina_obj = None
                plano_obj = None
                
                if tipo == 'maquina':
                    from app.models import Maquina
                    try:
                        maquina_obj = Maquina.objects.get(id=item_id)
                    except Maquina.DoesNotExist:
                        errors.append(f'Máquina com ID {item_id} não encontrada')
                        continue
                elif tipo == 'plano':
                    from app.models import MeuPlanoPreventiva
                    try:
                        plano_obj = MeuPlanoPreventiva.objects.get(id=item_id)
                    except MeuPlanoPreventiva.DoesNotExist:
                        errors.append(f'Plano com ID {item_id} não encontrado')
                        continue
                else:
                    errors.append(f'Tipo de agendamento inválido: {tipo}')
                    continue
                
                # Calcular datas se houver periodicidade
                if periodicidade and periodicidade > 0:
                    # Calcular todas as datas até o final do ano
                    ano_atual = date.today().year
                    fim_do_ano = date(ano_atual, 12, 31)
                    
                    datas_agendamento = []
                    data_atual = data_planejada
                    
                    while data_atual <= fim_do_ano:
                        datas_agendamento.append(data_atual)
                        data_atual = data_atual + timedelta(days=periodicidade)
                else:
                    # Sem periodicidade, apenas uma data
                    datas_agendamento = [data_planejada]
                
                # Criar agendamentos para cada data
                for data_agendamento in datas_agendamento:
                    try:
                        agendamento = AgendamentoCronograma(
                            tipo_agendamento=tipo,
                            data_planejada=data_agendamento,
                            nome_grupo=nome_grupo,
                            periodicidade=periodicidade if periodicidade and periodicidade > 0 else None,
                            created_by=request.user.username if request.user.is_authenticated else 'Sistema'
                        )
                        
                        if tipo == 'maquina':
                            agendamento.maquina = maquina_obj
                        elif tipo == 'plano':
                            agendamento.plano_preventiva = plano_obj
                        
                        agendamento.full_clean()
                        agendamento.save()
                        saved_count += 1
                    except Exception as e:
                        errors.append(f'Erro ao salvar agendamento para data {data_agendamento}: {str(e)}')
                        continue
                
            except Exception as e:
                errors.append(f'Erro ao processar agendamento: {str(e)}')
                continue
        
        if saved_count > 0:
            return JsonResponse({
                'success': True,
                'saved_count': saved_count,
                'total': len(agendamentos_data),
                'errors': errors if errors else None
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Nenhum agendamento foi salvo',
                'errors': errors
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def consultar_roteiro_preventiva(request):
    """Consultar/listar roteiros de manutenção preventiva"""
    from app.models import RoteiroPreventiva
    
    # Buscar todos os roteiros preventiva
    roteiros_list = RoteiroPreventiva.objects.all().select_related('maquina')
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_maquina=search_num) | Q(cd_planmanut=search_num) | Q(seq_seqplamanu=search_num) | Q(cd_tarefamanu=search_num) | Q(cd_ordemserv=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(descr_tarefamanu__icontains=search_query) |
            Q(nome_funciomanu__icontains=search_query) |
            Q(cd_funciomanu__icontains=search_query) |
            Q(cd_setormanut__icontains=search_query) |
            Q(descr_setormanut__icontains=search_query) |
            Q(nome_unid__icontains=search_query) |
            Q(descr_planmanut__icontains=search_query) |
            Q(descr_item__icontains=search_query)
        )
        
        roteiros_list = roteiros_list.filter(search_conditions)
    
    # Filtros por coluna individual
    filter_maquina = request.GET.get('filter_maquina', '').strip()
    if filter_maquina:
        try:
            maquina_num = int(float(filter_maquina))
            roteiros_list = roteiros_list.filter(cd_maquina=maquina_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(
                Q(cd_maquina__icontains=filter_maquina) |
                Q(descr_maquina__icontains=filter_maquina)
            )
    
    filter_planmanut = request.GET.get('filter_planmanut', '').strip()
    if filter_planmanut:
        try:
            planmanut_num = int(float(filter_planmanut))
            roteiros_list = roteiros_list.filter(cd_planmanut=planmanut_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(
                Q(cd_planmanut__icontains=filter_planmanut) |
                Q(descr_planmanut__icontains=filter_planmanut)
            )
    
    filter_seq_plamanu = request.GET.get('filter_seq_plamanu', '').strip()
    if filter_seq_plamanu:
        try:
            seq_num = int(float(filter_seq_plamanu))
            roteiros_list = roteiros_list.filter(seq_seqplamanu=seq_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(seq_seqplamanu__icontains=filter_seq_plamanu)
    
    filter_tarefamanu = request.GET.get('filter_tarefamanu', '').strip()
    if filter_tarefamanu:
        try:
            tarefa_num = int(float(filter_tarefamanu))
            roteiros_list = roteiros_list.filter(cd_tarefamanu=tarefa_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(
                Q(cd_tarefamanu__icontains=filter_tarefamanu) |
                Q(descr_tarefamanu__icontains=filter_tarefamanu)
            )
    
    filter_ordemserv = request.GET.get('filter_ordemserv', '').strip()
    if filter_ordemserv:
        try:
            ordemserv_num = int(float(filter_ordemserv))
            roteiros_list = roteiros_list.filter(cd_ordemserv=ordemserv_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(cd_ordemserv__icontains=filter_ordemserv)
    
    filter_data_exec = request.GET.get('filter_data_exec', '').strip()
    if filter_data_exec:
        roteiros_list = roteiros_list.filter(dt_primexec__icontains=filter_data_exec)
    
    filter_periodo = request.GET.get('filter_periodo', '').strip()
    if filter_periodo:
        try:
            periodo_num = int(float(filter_periodo))
            roteiros_list = roteiros_list.filter(qtde_periodo=periodo_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(qtde_periodo__icontains=filter_periodo)
    
    filter_funcionario = request.GET.get('filter_funcionario', '').strip()
    if filter_funcionario:
        roteiros_list = roteiros_list.filter(
            Q(nome_funciomanu__icontains=filter_funcionario) |
            Q(cd_funciomanu__icontains=filter_funcionario)
        )
    
    filter_setor = request.GET.get('filter_setor', '').strip()
    if filter_setor:
        roteiros_list = roteiros_list.filter(
            Q(cd_setormanut__icontains=filter_setor) |
            Q(descr_setormanut__icontains=filter_setor)
        )
    
    filter_unidade = request.GET.get('filter_unidade', '').strip()
    if filter_unidade:
        try:
            unidade_num = int(float(filter_unidade))
            roteiros_list = roteiros_list.filter(cd_unid=unidade_num)
        except (ValueError, TypeError):
            roteiros_list = roteiros_list.filter(
                Q(nome_unid__icontains=filter_unidade)
            )
    
    # Ordenar por máquina, plano, sequência e tarefa
    roteiros_list = roteiros_list.order_by('cd_maquina', 'cd_planmanut', 'seq_seqplamanu', 'cd_tarefamanu')
    
    # Paginação
    paginator = Paginator(roteiros_list, 100)  # 100 itens por página
    page_number = request.GET.get('page', 1)
    roteiros = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = RoteiroPreventiva.objects.count()
    maquinas_count = RoteiroPreventiva.objects.exclude(cd_maquina__isnull=True).values('cd_maquina').distinct().count()
    setores_count = RoteiroPreventiva.objects.exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    planos_count = RoteiroPreventiva.objects.exclude(cd_planmanut__isnull=True).values('cd_planmanut').distinct().count()
    
    context = {
        'page_title': 'Consultar Roteiros Preventiva',
        'active_page': 'consultar_roteiro_preventiva',
        'roteiros': roteiros,
        'total_count': total_count,
        'maquinas_count': maquinas_count,
        'setores_count': setores_count,
        'planos_count': planos_count,
        # Preservar filtros no contexto
        'filter_maquina': filter_maquina,
        'filter_planmanut': filter_planmanut,
        'filter_seq_plamanu': filter_seq_plamanu,
        'filter_tarefamanu': filter_tarefamanu,
        'filter_ordemserv': filter_ordemserv,
        'filter_data_exec': filter_data_exec,
        'filter_periodo': filter_periodo,
        'filter_funcionario': filter_funcionario,
        'filter_setor': filter_setor,
        'filter_unidade': filter_unidade,
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_roteiro_preventiva.html', context)


def visualizar_roteiro_preventiva(request, roteiro_id):
    """Visualizar detalhes de um roteiro de manutenção preventiva específico"""
    from app.models import RoteiroPreventiva
    
    try:
        roteiro = RoteiroPreventiva.objects.select_related('maquina').get(id=roteiro_id)
    except RoteiroPreventiva.DoesNotExist:
        messages.error(request, 'Roteiro de manutenção preventiva não encontrado.')
        return redirect('consultar_roteiro_preventiva')
    
    context = {
        'page_title': f'Visualizar Roteiro Preventiva - Máquina {roteiro.cd_maquina}',
        'active_page': 'consultar_roteiro_preventiva',
        'roteiro': roteiro,
    }
    return render(request, 'visualizar/visualizar_roteiro_preventiva.html', context)


def visualizar_analise_plano_roteiro(request, plano_id, roteiro_id):
    """Visualizar análise detalhada da relação entre um PlanoPreventiva e um RoteiroPreventiva - Função limpa para recriar do zero"""
    from django.shortcuts import redirect
    from django.contrib import messages
    
    messages.info(request, 'Esta funcionalidade está sendo recriada.')
    return redirect('analise_roteiro_plano_preventiva')


def erro_analise_plano_roteiro(request, plano_id=None, roteiro_id=None):
    """Visualizar análise de erros - o que está faltando para encontrar match entre PlanoPreventiva e RoteiroPreventiva"""
    from app.models import PlanoPreventiva, RoteiroPreventiva
    from django.shortcuts import redirect
    from django.contrib import messages
    
    # Se não há IDs, mostrar visão geral de todos os registros sem match
    if plano_id is None or roteiro_id is None:
        return erro_analise_plano_roteiro_geral(request)
    
    try:
        plano = PlanoPreventiva.objects.select_related('maquina', 'roteiro_preventiva').get(id=plano_id)
    except PlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano de manutenção preventiva não encontrado.')
        return redirect('analise_roteiro_plano_preventiva')
    
    try:
        roteiro = RoteiroPreventiva.objects.select_related('maquina').get(id=roteiro_id)
    except RoteiroPreventiva.DoesNotExist:
        messages.error(request, 'Roteiro de manutenção preventiva não encontrado.')
        return redirect('analise_roteiro_plano_preventiva')
    
    # Analisar o que está faltando para ter match
    erros = []
    problemas = []
    campos_comparados = []  # Lista de todos os campos para exibição completa
    
    # Verificar cd_maquina
    if not plano.cd_maquina or not roteiro.cd_maquina:
        erros.append({
            'campo': 'cd_maquina',
            'label': 'Código da Máquina',
            'problema': 'Um ou ambos os campos estão vazios',
            'plano_valor': plano.cd_maquina,
            'roteiro_valor': roteiro.cd_maquina,
            'solucao': 'Ambos os registros precisam ter o código da máquina preenchido e devem ser iguais',
            'tipo': 'vazio'
        })
    elif plano.cd_maquina != roteiro.cd_maquina:
        erros.append({
            'campo': 'cd_maquina',
            'label': 'Código da Máquina',
            'problema': 'Os valores são diferentes',
            'plano_valor': plano.cd_maquina,
            'roteiro_valor': roteiro.cd_maquina,
            'solucao': f'O código da máquina no Plano ({plano.cd_maquina}) deve ser igual ao do Roteiro ({roteiro.cd_maquina})',
            'tipo': 'diferente'
        })
    
    # Verificar descr_maquina
    descr_plano = (plano.descr_maquina or '').strip().upper()
    descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
    campo_match = True
    if not descr_plano or not descr_roteiro:
        campo_match = False
        erros.append({
            'campo': 'descr_maquina',
            'label': 'Descrição da Máquina',
            'problema': 'Um ou ambos os campos estão vazios',
            'plano_valor': plano.descr_maquina,
            'roteiro_valor': roteiro.descr_maquina,
            'solucao': 'Ambos os registros precisam ter a descrição da máquina preenchida e devem ser iguais',
            'tipo': 'vazio'
        })
    elif descr_plano != descr_roteiro:
        campo_match = False
        erros.append({
            'campo': 'descr_maquina',
            'label': 'Descrição da Máquina',
            'problema': 'Os valores são diferentes',
            'plano_valor': plano.descr_maquina,
            'roteiro_valor': roteiro.descr_maquina,
            'solucao': 'As descrições da máquina devem ser idênticas (ignorando maiúsculas/minúsculas)',
            'tipo': 'diferente'
        })
    campos_comparados.append({
        'campo': 'descr_maquina',
        'label': 'Descrição da Máquina',
        'plano_valor': plano.descr_maquina,
        'roteiro_valor': roteiro.descr_maquina,
        'match': campo_match
    })
    
    # Verificar sequencia_tarefa vs cd_tarefamanu
    campo_match = True
    if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
        campo_match = False
        erros.append({
            'campo': 'sequencia_tarefa',
            'label': 'Sequência Tarefa (Plano) / Código Tarefa (Roteiro)',
            'problema': 'Um ou ambos os campos estão vazios',
            'plano_valor': plano.sequencia_tarefa,
            'roteiro_valor': roteiro.cd_tarefamanu,
            'solucao': 'O campo "Sequência Tarefa" do Plano deve ser igual ao campo "Código Tarefa" do Roteiro',
            'tipo': 'vazio'
        })
    elif plano.sequencia_tarefa != roteiro.cd_tarefamanu:
        campo_match = False
        erros.append({
            'campo': 'sequencia_tarefa',
            'label': 'Sequência Tarefa (Plano) / Código Tarefa (Roteiro)',
            'problema': 'Os valores são diferentes',
            'plano_valor': plano.sequencia_tarefa,
            'roteiro_valor': roteiro.cd_tarefamanu,
            'solucao': f'O campo "Sequência Tarefa" do Plano ({plano.sequencia_tarefa}) deve ser igual ao campo "Código Tarefa" do Roteiro ({roteiro.cd_tarefamanu})',
            'tipo': 'diferente'
        })
    campos_comparados.append({
        'campo': 'sequencia_tarefa',
        'label': 'Sequência Tarefa (Plano) / Código Tarefa (Roteiro)',
        'plano_valor': plano.sequencia_tarefa,
        'roteiro_valor': roteiro.cd_tarefamanu,
        'match': campo_match
    })
    
    # Verificar descr_tarefa vs descr_tarefamanu
    descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
    descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
    campo_match = True
    if not descr_tarefa_plano or not descr_tarefa_roteiro:
        campo_match = False
        erros.append({
            'campo': 'descr_tarefa',
            'label': 'Descrição Tarefa (Plano) / Descrição Tarefa (Roteiro)',
            'problema': 'Um ou ambos os campos estão vazios',
            'plano_valor': plano.descr_tarefa,
            'roteiro_valor': roteiro.descr_tarefamanu,
            'solucao': 'O campo "Descrição Tarefa" do Plano deve ser igual ao campo "Descrição Tarefa" do Roteiro',
            'tipo': 'vazio'
        })
    elif descr_tarefa_plano != descr_tarefa_roteiro:
        campo_match = False
        erros.append({
            'campo': 'descr_tarefa',
            'label': 'Descrição Tarefa (Plano) / Descrição Tarefa (Roteiro)',
            'problema': 'Os valores são diferentes',
            'plano_valor': plano.descr_tarefa,
            'roteiro_valor': roteiro.descr_tarefamanu,
            'solucao': 'As descrições da tarefa devem ser idênticas (ignorando maiúsculas/minúsculas)',
            'tipo': 'diferente'
        })
    campos_comparados.append({
        'campo': 'descr_tarefa',
        'label': 'Descrição Tarefa (Plano) / Descrição Tarefa (Roteiro)',
        'plano_valor': plano.descr_tarefa,
        'roteiro_valor': roteiro.descr_tarefamanu,
        'match': campo_match
    })
    
    # Verificar sequencia_manutencao vs seq_seqplamanu
    campo_match = True
    if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
        campo_match = False
        erros.append({
            'campo': 'sequencia_manutencao',
            'label': 'Sequência Manutenção (Plano) / Sequência Plano (Roteiro)',
            'problema': 'Um ou ambos os campos estão vazios',
            'plano_valor': plano.sequencia_manutencao,
            'roteiro_valor': roteiro.seq_seqplamanu,
            'solucao': 'O campo "Sequência Manutenção" do Plano deve ser igual ao campo "Sequência Plano" do Roteiro',
            'tipo': 'vazio'
        })
    elif plano.sequencia_manutencao != roteiro.seq_seqplamanu:
        campo_match = False
        erros.append({
            'campo': 'sequencia_manutencao',
            'label': 'Sequência Manutenção (Plano) / Sequência Plano (Roteiro)',
            'problema': 'Os valores são diferentes',
            'plano_valor': plano.sequencia_manutencao,
            'roteiro_valor': roteiro.seq_seqplamanu,
            'solucao': f'O campo "Sequência Manutenção" do Plano ({plano.sequencia_manutencao}) deve ser igual ao campo "Sequência Plano" do Roteiro ({roteiro.seq_seqplamanu})',
            'tipo': 'diferente'
        })
    campos_comparados.append({
        'campo': 'sequencia_manutencao',
        'label': 'Sequência Manutenção (Plano) / Sequência Plano (Roteiro)',
        'plano_valor': plano.sequencia_manutencao,
        'roteiro_valor': roteiro.seq_seqplamanu,
        'match': campo_match
    })
    
    # Resumo dos problemas
    total_erros = len(erros)
    campos_vazios = sum(1 for e in erros if e.get('tipo') == 'vazio')
    campos_diferentes = sum(1 for e in erros if e.get('tipo') == 'diferente')
    total_campos = 5  # Total de campos comparados
    campos_match = total_campos - total_erros
    percentual_match = (campos_match / total_campos * 100) if total_campos > 0 else 0
    
    context = {
        'page_title': f'Análise de Erros: Plano {plano.numero_plano} ↔ Roteiro {roteiro.cd_planmanut}',
        'active_page': 'analise_roteiro_plano_preventiva',
        'plano': plano,
        'roteiro': roteiro,
        'erros': erros,
        'total_erros': total_erros,
        'campos_vazios': campos_vazios,
        'campos_diferentes': campos_diferentes,
        'total_campos': total_campos,
        'campos_match': campos_match,
        'percentual_match': percentual_match,
        'campos_comparados': campos_comparados,
    }
    return render(request, 'planejamento/erro_analise_plano_roteiro.html', context)


def erro_analise_plano_roteiro_geral(request):
    """Visão geral de análise de erros - todos os registros sem match e o que está faltando"""
    from app.models import PlanoPreventiva, RoteiroPreventiva
    from django.core.paginator import Paginator
    
    # Buscar todos os registros
    planos = PlanoPreventiva.objects.all()
    roteiros = RoteiroPreventiva.objects.all()
    
    # Função para verificar se campos correspondem (mesma lógica da análise principal)
    def campos_correspondem(plano, roteiro):
        if not plano.cd_maquina or not roteiro.cd_maquina:
            return False
        if plano.cd_maquina != roteiro.cd_maquina:
            return False
        
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if descr_plano and descr_roteiro:
            if descr_plano != descr_roteiro:
                return False
        elif descr_plano or descr_roteiro:
            return False
        
        if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
            return False
        if plano.sequencia_tarefa != roteiro.cd_tarefamanu:
            return False
        
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if descr_tarefa_plano and descr_tarefa_roteiro:
            if descr_tarefa_plano != descr_tarefa_roteiro:
                return False
        elif descr_tarefa_plano or descr_tarefa_roteiro:
            return False
        
        if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
            return False
        if plano.sequencia_manutencao != roteiro.seq_seqplamanu:
            return False
        
        return True
    
    # Função para analisar erros de um par plano-roteiro
    def analisar_erros(plano, roteiro):
        erros = []
        
        # Verificar cd_maquina
        if not plano.cd_maquina or not roteiro.cd_maquina:
            erros.append('cd_maquina')
        elif plano.cd_maquina != roteiro.cd_maquina:
            erros.append('cd_maquina')
        
        # Verificar descr_maquina
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if not descr_plano or not descr_roteiro:
            erros.append('descr_maquina')
        elif descr_plano != descr_roteiro:
            erros.append('descr_maquina')
        
        # Verificar sequencia_tarefa vs cd_tarefamanu
        if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
            erros.append('sequencia_tarefa')
        elif plano.sequencia_tarefa != roteiro.cd_tarefamanu:
            erros.append('sequencia_tarefa')
        
        # Verificar descr_tarefa vs descr_tarefamanu
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if not descr_tarefa_plano or not descr_tarefa_roteiro:
            erros.append('descr_tarefa')
        elif descr_tarefa_plano != descr_tarefa_roteiro:
            erros.append('descr_tarefa')
        
        # Verificar sequencia_manutencao vs seq_seqplamanu
        if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
            erros.append('sequencia_manutencao')
        elif plano.sequencia_manutencao != roteiro.seq_seqplamanu:
            erros.append('sequencia_manutencao')
        
        return erros
    
    # Encontrar planos sem match
    planos_sem_match = []
    planos_processados = set()
    roteiros_processados = set()
    
    for plano in planos:
        tem_match = False
        melhor_match = None
        melhor_erros = []
        
        for roteiro in roteiros:
            if campos_correspondem(plano, roteiro):
                tem_match = True
                planos_processados.add(plano.id)
                roteiros_processados.add(roteiro.id)
                break
            else:
                # Analisar erros para encontrar o melhor match parcial
                erros = analisar_erros(plano, roteiro)
                if not melhor_match or len(erros) < len(melhor_erros):
                    melhor_match = roteiro
                    melhor_erros = erros
        
        if not tem_match:
            planos_sem_match.append({
                'plano': plano,
                'melhor_match': melhor_match,
                'erros': melhor_erros,
                'total_erros': len(melhor_erros) if melhor_erros else 5,
            })
    
    # Encontrar roteiros sem match
    roteiros_sem_match = []
    for roteiro in roteiros:
        if roteiro.id not in roteiros_processados:
            melhor_match = None
            melhor_erros = []
            
            for plano in planos:
                if plano.id not in planos_processados:
                    erros = analisar_erros(plano, roteiro)
                    if not melhor_match or len(erros) < len(melhor_erros):
                        melhor_match = plano
                        melhor_erros = erros
            
            roteiros_sem_match.append({
                'roteiro': roteiro,
                'melhor_match': melhor_match,
                'erros': melhor_erros,
                'total_erros': len(melhor_erros) if melhor_erros else 5,
            })
    
    # Estatísticas gerais
    total_planos_sem_match = len(planos_sem_match)
    total_roteiros_sem_match = len(roteiros_sem_match)
    
    # Contar tipos de erros mais comuns
    erros_comuns = {}
    for item in planos_sem_match + roteiros_sem_match:
        for erro in item['erros']:
            erros_comuns[erro] = erros_comuns.get(erro, 0) + 1
    
    context = {
        'page_title': 'Análise Geral de Erros - Correspondências não encontradas',
        'active_page': 'analise_roteiro_plano_preventiva',
        'planos_sem_match': planos_sem_match[:50],  # Limitar para performance
        'roteiros_sem_match': roteiros_sem_match[:50],
        'total_planos_sem_match': total_planos_sem_match,
        'total_roteiros_sem_match': total_roteiros_sem_match,
        'erros_comuns': sorted(erros_comuns.items(), key=lambda x: x[1], reverse=True),
        'is_geral': True,
    }
    return render(request, 'planejamento/erro_analise_plano_roteiro.html', context)


def relacionar_roteiro_plano(request):
    """Página para relacionar manualmente Roteiros e Planos que não têm match"""
    from app.models import PlanoPreventiva, RoteiroPreventiva, MeuPlanoPreventiva
    from django.contrib import messages
    from django.db import transaction
    from django.core.paginator import Paginator
    
    # Processar criação de relacionamento manual
    if request.method == 'POST' and 'criar_relacionamento' in request.POST:
        plano_id = request.POST.get('plano_id')
        roteiro_id = request.POST.get('roteiro_id')
        tipo = request.POST.get('tipo')  # 'roteiro_sem' ou 'plano_sem'
        
        if not plano_id or not roteiro_id:
            messages.error(request, 'Por favor, selecione tanto um Plano quanto um Roteiro.')
        else:
            try:
                plano = PlanoPreventiva.objects.get(id=plano_id)
                roteiro = RoteiroPreventiva.objects.get(id=roteiro_id)
                
                # Verificar se já existe um MeuPlanoPreventiva para este plano
                meu_plano, created = MeuPlanoPreventiva.objects.get_or_create(
                    cd_maquina=plano.cd_maquina,
                    sequencia_manutencao=plano.sequencia_manutencao,
                    sequencia_tarefa=plano.sequencia_tarefa,
                    defaults={
                        'cd_unid': plano.cd_unid,
                        'nome_unid': plano.nome_unid,
                        'cd_setor': plano.cd_setor,
                        'descr_setor': plano.descr_setor,
                        'cd_atividade': plano.cd_atividade,
                        'descr_maquina': plano.descr_maquina,
                        'nro_patrimonio': plano.nro_patrimonio,
                        'numero_plano': plano.numero_plano,
                        'descr_plano': plano.descr_plano,
                        'dt_execucao': plano.dt_execucao,
                        'quantidade_periodo': plano.quantidade_periodo,
                        'descr_tarefa': plano.descr_tarefa,
                        'cd_funcionario': plano.cd_funcionario,
                        'nome_funcionario': plano.nome_funcionario,
                        'descr_seqplamanu': plano.descr_seqplamanu,
                        'desc_detalhada_do_roteiro_preventiva': roteiro.descr_seqplamanu,
                        'roteiro_preventiva': roteiro,
                        'maquina': plano.maquina,
                    }
                )
                
                # Se já existia, atualizar
                if not created:
                    meu_plano.desc_detalhada_do_roteiro_preventiva = roteiro.descr_seqplamanu
                    meu_plano.roteiro_preventiva = roteiro
                    meu_plano.save()
                
                messages.success(request, f'Relacionamento criado com sucesso! Plano {plano.id} vinculado ao Roteiro {roteiro.id} em MeuPlanoPreventiva.')
            except PlanoPreventiva.DoesNotExist:
                messages.error(request, 'Plano não encontrado.')
            except RoteiroPreventiva.DoesNotExist:
                messages.error(request, 'Roteiro não encontrado.')
            except Exception as e:
                messages.error(request, f'Erro ao criar relacionamento: {str(e)}')
    
    # Buscar todos os registros
    planos = PlanoPreventiva.objects.all()
    roteiros = RoteiroPreventiva.objects.all()
    
    # Função para verificar se campos correspondem
    def campos_correspondem(plano, roteiro):
        if not plano.cd_maquina or not roteiro.cd_maquina:
            return False
        if plano.cd_maquina != roteiro.cd_maquina:
            return False
        
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        if descr_plano and descr_roteiro:
            if descr_plano != descr_roteiro:
                return False
        elif descr_plano or descr_roteiro:
            return False
        
        if not plano.sequencia_tarefa or not roteiro.cd_tarefamanu:
            return False
        if plano.sequencia_tarefa != roteiro.cd_tarefamanu:
            return False
        
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        if descr_tarefa_plano and descr_tarefa_roteiro:
            if descr_tarefa_plano != descr_tarefa_roteiro:
                return False
        elif descr_tarefa_plano or descr_tarefa_roteiro:
            return False
        
        if not plano.sequencia_manutencao or not roteiro.seq_seqplamanu:
            return False
        if plano.sequencia_manutencao != roteiro.seq_seqplamanu:
            return False
        
        return True
    
    # Encontrar relacionamentos existentes
    relacionamentos = []
    planos_processados = set()
    roteiros_processados = set()
    
    for plano in planos:
        for roteiro in roteiros:
            if campos_correspondem(plano, roteiro):
                relacionamentos.append((plano.id, roteiro.id))
                planos_processados.add(plano.id)
                roteiros_processados.add(roteiro.id)
                break
    
    # Encontrar planos sem match
    planos_sem_match = [p for p in planos if p.id not in planos_processados]
    
    # Encontrar roteiros sem match
    roteiros_sem_match = [r for r in roteiros if r.id not in roteiros_processados]
    
    # Paginação
    page_planos = request.GET.get('page_planos', 1)
    page_roteiros = request.GET.get('page_roteiros', 1)
    
    paginator_planos = Paginator(planos_sem_match, 20)
    paginator_roteiros = Paginator(roteiros_sem_match, 20)
    
    planos_paginated = paginator_planos.get_page(page_planos)
    roteiros_paginated = paginator_roteiros.get_page(page_roteiros)
    
    context = {
        'page_title': 'Relacionar Roteiro e Plano Manualmente',
        'active_page': 'relacionar_roteiro_plano',
        'planos_sem_match': planos_paginated,
        'roteiros_sem_match': roteiros_paginated,
        'total_planos_sem_match': len(planos_sem_match),
        'total_roteiros_sem_match': len(roteiros_sem_match),
        'todos_planos': list(planos.values('id', 'numero_plano', 'cd_maquina', 'descr_maquina', 'sequencia_manutencao', 'sequencia_tarefa')),
        'todos_roteiros': list(roteiros.values('id', 'cd_planmanut', 'cd_maquina', 'descr_maquina', 'seq_seqplamanu', 'cd_tarefamanu')),
    }
    return render(request, 'planejamento/relacionar_roteiro_plano.html', context)


def visualizar_comparacao_roteiro_plano(request, plano_id, roteiro_id):
    """Visualizar comparação detalhada entre um PlanoPreventiva e um RoteiroPreventiva"""
    from app.models import PlanoPreventiva, RoteiroPreventiva, MeuPlanoPreventiva
    from django.shortcuts import redirect
    from django.contrib import messages
    
    try:
        plano = PlanoPreventiva.objects.select_related('maquina', 'roteiro_preventiva').get(id=plano_id)
    except PlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano de manutenção preventiva não encontrado.')
        return redirect('analise_roteiro_plano_preventiva')
    
    try:
        roteiro = RoteiroPreventiva.objects.select_related('maquina').get(id=roteiro_id)
    except RoteiroPreventiva.DoesNotExist:
        messages.error(request, 'Roteiro de manutenção preventiva não encontrado.')
        return redirect('analise_roteiro_plano_preventiva')
    
    # Função para verificar se os campos correspondem
    def verificar_correspondencia(plano, roteiro):
        """Verifica se os campos principais correspondem exatamente"""
        comparacoes = {}
        
        # Comparar cd_maquina
        comparacoes['cd_maquina'] = {
            'plano': plano.cd_maquina,
            'roteiro': roteiro.cd_maquina,
            'match': plano.cd_maquina == roteiro.cd_maquina if plano.cd_maquina and roteiro.cd_maquina else False,
            'campo_plano': 'cd_maquina',
            'campo_roteiro': 'cd_maquina',
            'label': 'Código da Máquina'
        }
        
        # Comparar descr_maquina
        descr_plano = (plano.descr_maquina or '').strip().upper()
        descr_roteiro = (roteiro.descr_maquina or '').strip().upper()
        comparacoes['descr_maquina'] = {
            'plano': plano.descr_maquina,
            'roteiro': roteiro.descr_maquina,
            'match': descr_plano == descr_roteiro if descr_plano and descr_roteiro else False,
            'campo_plano': 'descr_maquina',
            'campo_roteiro': 'descr_maquina',
            'label': 'Descrição da Máquina'
        }
        
        # Comparar sequencia_tarefa (Plano) com cd_tarefamanu (Roteiro)
        comparacoes['sequencia_tarefa'] = {
            'plano': plano.sequencia_tarefa,
            'roteiro': roteiro.cd_tarefamanu,
            'match': plano.sequencia_tarefa == roteiro.cd_tarefamanu if plano.sequencia_tarefa and roteiro.cd_tarefamanu else False,
            'campo_plano': 'sequencia_tarefa',
            'campo_roteiro': 'cd_tarefamanu',
            'label': 'Sequência Tarefa / Código Tarefa'
        }
        
        # Comparar descr_tarefa (Plano) com descr_tarefamanu (Roteiro)
        descr_tarefa_plano = (plano.descr_tarefa or '').strip().upper()
        descr_tarefa_roteiro = (roteiro.descr_tarefamanu or '').strip().upper()
        comparacoes['descr_tarefa'] = {
            'plano': plano.descr_tarefa,
            'roteiro': roteiro.descr_tarefamanu,
            'match': descr_tarefa_plano == descr_tarefa_roteiro if descr_tarefa_plano and descr_tarefa_roteiro else False,
            'campo_plano': 'descr_tarefa',
            'campo_roteiro': 'descr_tarefamanu',
            'label': 'Descrição Tarefa'
        }
        
        # Comparar sequencia_manutencao (Plano) com seq_seqplamanu (Roteiro)
        comparacoes['sequencia_manutencao'] = {
            'plano': plano.sequencia_manutencao,
            'roteiro': roteiro.seq_seqplamanu,
            'match': plano.sequencia_manutencao == roteiro.seq_seqplamanu if plano.sequencia_manutencao and roteiro.seq_seqplamanu else False,
            'campo_plano': 'sequencia_manutencao',
            'campo_roteiro': 'seq_seqplamanu',
            'label': 'Sequência Manutenção'
        }
        
        return comparacoes
    
    # Verificar correspondências
    comparacoes = verificar_correspondencia(plano, roteiro)
    
    # Contar matches
    total_campos = len(comparacoes)
    campos_match = sum(1 for comp in comparacoes.values() if comp['match'])
    percentual_match = (campos_match / total_campos * 100) if total_campos > 0 else 0
    
    # Verificar se corresponde completamente (todos os campos)
    corresponde_completamente = all(comp['match'] for comp in comparacoes.values())
    
    # Verificar se já foi salvo em MeuPlanoPreventiva
    ja_salvo = MeuPlanoPreventiva.objects.filter(
        cd_maquina=plano.cd_maquina,
        sequencia_manutencao=plano.sequencia_manutencao,
        sequencia_tarefa=plano.sequencia_tarefa
    ).exists()
    
    context = {
        'page_title': f'Comparação: Plano {plano.numero_plano} ↔ Roteiro {roteiro.cd_planmanut}',
        'active_page': 'analise_roteiro_plano_preventiva',
        'plano': plano,
        'roteiro': roteiro,
        'comparacoes': comparacoes,
        'total_campos': total_campos,
        'campos_match': campos_match,
        'percentual_match': percentual_match,
        'corresponde_completamente': corresponde_completamente,
        'ja_salvo': ja_salvo,
        'descr_seqplamanu': roteiro.descr_seqplamanu,
    }
    return render(request, 'visualizar/visualizar_comparacao_roteiro_plano.html', context)


def visualizar_manutencao_preventiva(request, plano_id):
    """Visualizar detalhes de um plano de manutenção preventiva específico"""
    from app.models import PlanoPreventiva, PlanoPreventivaDocumento
    
    try:
        plano = PlanoPreventiva.objects.select_related('maquina', 'roteiro_preventiva').get(id=plano_id)
    except PlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano de manutenção preventiva não encontrado.')
        return redirect('consultar_manutencoes_preventivas')
    
    # Buscar documentos relacionados
    documentos = PlanoPreventivaDocumento.objects.filter(plano_preventiva=plano).order_by('-created_at')
    
    context = {
        'page_title': f'Visualizar Manutenção Preventiva - Plano {plano.numero_plano}',
        'active_page': 'consultar_manutencoes_preventivas',
        'plano': plano,
        'documentos': documentos,
    }
    return render(request, 'visualizar/visualizar_plano_preventiva.html', context)


def visualizar_plano_pcm(request, plano_id):
    """Visualizar detalhes de um MeuPlanoPreventiva específico"""
    from app.models import MeuPlanoPreventiva, MeuPlanoPreventivaDocumento
    from django.shortcuts import redirect
    from django.contrib import messages
    
    try:
        plano = MeuPlanoPreventiva.objects.select_related('maquina', 'roteiro_preventiva').get(id=plano_id)
    except MeuPlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano PCM não encontrado.')
        return redirect('consultar_meu_plano')
    
    # Buscar documentos associados
    documentos_associados = MeuPlanoPreventivaDocumento.objects.filter(
        meu_plano_preventiva=plano
    ).select_related('maquina_documento').order_by('-created_at')
    
    context = {
        'page_title': f'Visualizar Plano PCM - Plano {plano.numero_plano}',
        'active_page': 'consultar_meu_plano',
        'plano': plano,
        'documentos_associados': documentos_associados,
    }
    return render(request, 'visualizar/visualizar_plano_pcm.html', context)


def gerar_pdf_plano_pcm(request, plano_id):
    """Gerar PDF com informações do MeuPlanoPreventiva e documentos associados"""
    from app.models import MeuPlanoPreventiva, MeuPlanoPreventivaDocumento
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfgen import canvas
    from io import BytesIO
    import os
    from django.conf import settings
    
    try:
        plano = MeuPlanoPreventiva.objects.select_related('maquina', 'roteiro_preventiva').get(id=plano_id)
    except MeuPlanoPreventiva.DoesNotExist:
        from django.contrib import messages
        messages.error(request, 'Plano PCM não encontrado.')
        from django.shortcuts import redirect
        return redirect('consultar_meu_plano')
    
    # Buscar documentos associados
    documentos_associados = MeuPlanoPreventivaDocumento.objects.filter(
        meu_plano_preventiva=plano
    ).select_related('maquina_documento').order_by('-created_at')
    
    # Criar buffer para o PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    # Container para os elementos do PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF9800'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#424242'),
        spaceAfter=8,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14
    
    # Título
    elements.append(Paragraph("PLANO PCM - MANUTENÇÃO PREVENTIVA", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações do Plano
    elements.append(Paragraph("INFORMAÇÕES DO PLANO", heading_style))
    
    plano_data = [
        ['<b>Número do Plano:</b>', str(plano.numero_plano) if plano.numero_plano else 'Não informado'],
        ['<b>Descrição do Plano:</b>', plano.descr_plano or 'Não informado'],
        ['<b>Sequência Manutenção:</b>', str(plano.sequencia_manutencao) if plano.sequencia_manutencao else 'Não informado'],
        ['<b>Sequência Tarefa:</b>', str(plano.sequencia_tarefa) if plano.sequencia_tarefa else 'Não informado'],
        ['<b>Data Execução:</b>', plano.dt_execucao or 'Não informado'],
        ['<b>Período (dias):</b>', str(plano.quantidade_periodo) if plano.quantidade_periodo else 'Não informado'],
    ]
    
    plano_table = Table(plano_data, colWidths=[6*cm, 10*cm])
    plano_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(plano_table)
    elements.append(Spacer(1, 0.3*cm))
    
    # Descrição da Tarefa
    if plano.descr_tarefa:
        elements.append(Paragraph("<b>Descrição da Tarefa:</b>", subheading_style))
        elements.append(Paragraph(plano.descr_tarefa, normal_style))
        elements.append(Spacer(1, 0.3*cm))
    
    # DESCR_SEQPLAMANU
    if plano.descr_seqplamanu:
        elements.append(Paragraph("<b>Descrição Sequência Plano Manutenção (DESCR_SEQPLAMANU):</b>", subheading_style))
        elements.append(Paragraph(plano.descr_seqplamanu, normal_style))
        elements.append(Spacer(1, 0.3*cm))
    
    # Descrição Detalhada do Roteiro
    if plano.desc_detalhada_do_roteiro_preventiva:
        elements.append(Paragraph("<b>Descrição Detalhada do Roteiro Preventiva:</b>", subheading_style))
        elements.append(Paragraph(plano.desc_detalhada_do_roteiro_preventiva, normal_style))
        elements.append(Spacer(1, 0.3*cm))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações da Máquina
    elements.append(Paragraph("INFORMAÇÕES DA MÁQUINA", heading_style))
    
    maquina_data = [
        ['<b>Código da Máquina:</b>', str(plano.cd_maquina) if plano.cd_maquina else 'Não informado'],
        ['<b>Descrição da Máquina:</b>', plano.descr_maquina or 'Não informado'],
        ['<b>Nº Patrimônio:</b>', plano.nro_patrimonio or 'Não informado'],
    ]
    
    maquina_table = Table(maquina_data, colWidths=[6*cm, 10*cm])
    maquina_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(maquina_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações do Funcionário
    elements.append(Paragraph("FUNCIONÁRIO RESPONSÁVEL", heading_style))
    
    funcionario_data = [
        ['<b>Código Funcionário:</b>', plano.cd_funcionario or 'Não informado'],
        ['<b>Nome Funcionário:</b>', plano.nome_funcionario or 'Não informado'],
    ]
    
    funcionario_table = Table(funcionario_data, colWidths=[6*cm, 10*cm])
    funcionario_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(funcionario_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações de Unidade e Setor
    elements.append(Paragraph("UNIDADE E SETOR", heading_style))
    
    unidade_data = [
        ['<b>Código Unidade:</b>', str(plano.cd_unid) if plano.cd_unid else 'Não informado'],
        ['<b>Nome Unidade:</b>', plano.nome_unid or 'Não informado'],
        ['<b>Código Setor:</b>', plano.cd_setor or 'Não informado'],
        ['<b>Descrição Setor:</b>', plano.descr_setor or 'Não informado'],
        ['<b>Código Atividade:</b>', str(plano.cd_atividade) if plano.cd_atividade else 'Não informado'],
    ]
    
    unidade_table = Table(unidade_data, colWidths=[6*cm, 10*cm])
    unidade_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(unidade_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Documentos Associados
    elements.append(Paragraph("DOCUMENTOS ASSOCIADOS", heading_style))
    
    if documentos_associados:
        elements.append(Paragraph(f"Total de documentos associados: <b>{documentos_associados.count()}</b>", normal_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Cabeçalho da tabela de documentos
        doc_header = [['<b>#</b>', '<b>Nome do Arquivo</b>', '<b>Comentário Original</b>', '<b>Comentário Adicional</b>', '<b>Data Associação</b>']]
        
        doc_data = doc_header.copy()
        for idx, associacao in enumerate(documentos_associados, 1):
            nome_arquivo = os.path.basename(associacao.maquina_documento.arquivo.name) if associacao.maquina_documento.arquivo else 'N/A'
            comentario_original = associacao.maquina_documento.comentario or '-'
            comentario_adicional = associacao.comentario or '-'
            data_associacao = associacao.created_at.strftime('%d/%m/%Y %H:%M') if associacao.created_at else '-'
            
            doc_data.append([
                str(idx),
                nome_arquivo[:50] + '...' if len(nome_arquivo) > 50 else nome_arquivo,
                comentario_original[:40] + '...' if len(comentario_original) > 40 else comentario_original,
                comentario_adicional[:40] + '...' if len(comentario_adicional) > 40 else comentario_adicional,
                data_associacao
            ])
        
        doc_table = Table(doc_data, colWidths=[1*cm, 5*cm, 4*cm, 4*cm, 2*cm])
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(doc_table)
    else:
        elements.append(Paragraph("<i>Nenhum documento associado a este plano PCM.</i>", normal_style))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações do Sistema
    elements.append(Paragraph("INFORMAÇÕES DO SISTEMA", heading_style))
    
    sistema_data = [
        ['<b>ID do Registro:</b>', str(plano.id)],
        ['<b>Data de Criação:</b>', plano.created_at.strftime('%d/%m/%Y %H:%M:%S') if plano.created_at else 'N/A'],
        ['<b>Última Atualização:</b>', plano.updated_at.strftime('%d/%m/%Y %H:%M:%S') if plano.updated_at else 'N/A'],
    ]
    
    sistema_table = Table(sistema_data, colWidths=[6*cm, 10*cm])
    sistema_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(sistema_table)
    
    # Construir PDF principal
    doc.build(elements)
    
    # Obter o PDF principal
    buffer.seek(0)
    pdf_principal = buffer.getvalue()
    
    # Mesclar PDFs dos documentos associados
    try:
        from PyPDF2 import PdfReader, PdfWriter
        import tempfile
        
        # Criar um writer para o PDF final
        pdf_writer = PdfWriter()
        
        # Adicionar o PDF principal
        pdf_principal_reader = PdfReader(BytesIO(pdf_principal))
        for page in pdf_principal_reader.pages:
            pdf_writer.add_page(page)
        
        # Processar documentos associados
        pdfs_mesclados = 0
        pdfs_nao_mesclados = []
        
        for associacao in documentos_associados:
            if associacao.maquina_documento and associacao.maquina_documento.arquivo:
                arquivo_path = associacao.maquina_documento.arquivo.path
                nome_arquivo = os.path.basename(arquivo_path)
                extensao = os.path.splitext(nome_arquivo)[1].lower()
                
                # Verificar se é PDF
                if extensao == '.pdf' and os.path.exists(arquivo_path):
                    try:
                        # Ler o PDF do documento
                        with open(arquivo_path, 'rb') as pdf_file:
                            pdf_reader = PdfReader(pdf_file)
                            
                            # Adicionar diretamente todas as páginas do PDF do documento
                            for page in pdf_reader.pages:
                                pdf_writer.add_page(page)
                            
                            pdfs_mesclados += 1
                    except Exception as e:
                        # Se houver erro ao processar o PDF, apenas registrar e continuar
                        pdfs_nao_mesclados.append(nome_arquivo)
                        print(f"Erro ao mesclar PDF {nome_arquivo}: {str(e)}")
                else:
                    # Não é PDF ou arquivo não existe - criar página informativa
                    try:
                        info_buffer = BytesIO()
                        info_doc = SimpleDocTemplate(info_buffer, pagesize=A4)
                        info_elements = []
                        
                        info_elements.append(Spacer(1, 8*cm))
                        info_elements.append(Paragraph(f"<b>DOCUMENTO ANEXO:</b> {nome_arquivo}", heading_style))
                        info_elements.append(Spacer(1, 0.3*cm))
                        info_elements.append(Paragraph(f"<i>Este arquivo não é um PDF e não pode ser incluído diretamente no documento.</i>", normal_style))
                        info_elements.append(Spacer(1, 0.2*cm))
                        info_elements.append(Paragraph(f"<b>Tipo de arquivo:</b> {extensao or 'Desconhecido'}", normal_style))
                        if associacao.maquina_documento.comentario:
                            info_elements.append(Paragraph(f"<b>Comentário:</b> {associacao.maquina_documento.comentario}", normal_style))
                        if associacao.comentario:
                            info_elements.append(Paragraph(f"<b>Comentário Adicional:</b> {associacao.comentario}", normal_style))
                        
                        info_doc.build(info_elements)
                        info_buffer.seek(0)
                        info_reader = PdfReader(info_buffer)
                        if info_reader.pages:
                            pdf_writer.add_page(info_reader.pages[0])
                    except Exception as e:
                        print(f"Erro ao criar página informativa para {nome_arquivo}: {str(e)}")
                    pdfs_nao_mesclados.append(nome_arquivo)
        
        # Criar buffer final com o PDF mesclado
        buffer_final = BytesIO()
        pdf_writer.write(buffer_final)
        buffer_final.seek(0)
        pdf_final = buffer_final.getvalue()
        
    except ImportError:
        # Se PyPDF2 não estiver instalado, usar apenas o PDF principal
        pdf_final = pdf_principal
    except Exception as e:
        # Em caso de erro na mesclagem, usar apenas o PDF principal
        print(f"Erro ao mesclar PDFs: {str(e)}")
        pdf_final = pdf_principal
    
    # Criar resposta HTTP
    response = HttpResponse(pdf_final, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Plano_PCM_{plano.numero_plano}_{plano.cd_maquina}.pdf"'
    
    return response


def editar_plano_pcm(request, plano_id):
    """Editar um MeuPlanoPreventiva existente"""
    from app.forms import MeuPlanoPreventivaForm
    from app.models import MeuPlanoPreventiva, Maquina, RoteiroPreventiva, MaquinaDocumento, MeuPlanoPreventivaDocumento
    
    try:
        plano = MeuPlanoPreventiva.objects.get(id=plano_id)
    except MeuPlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano PCM não encontrado.')
        return redirect('consultar_meu_plano')
    
    if request.method == 'POST':
        form = MeuPlanoPreventivaForm(request.POST, instance=plano)
        
        if form.is_valid():
            try:
                plano = form.save()
                messages.success(request, f'Plano PCM {plano.numero_plano} atualizado com sucesso!')
                return redirect('visualizar_plano_pcm', plano_id=plano.id)
            except Exception as e:
                messages.error(request, f'Erro ao atualizar plano PCM: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = MeuPlanoPreventivaForm(instance=plano)
    
    # Buscar máquinas e roteiros para os selects
    maquinas = Maquina.objects.all().order_by('cd_maquina')[:100]  # Limitar para performance
    roteiros = RoteiroPreventiva.objects.all().order_by('cd_maquina', 'cd_planmanut')[:100]  # Limitar para performance
    
    # Buscar documentos da máquina associada (se houver)
    documentos_maquina = []
    documentos_associados = []
    associacoes = []
    documentos_associados_ids = []
    associacoes_dict = {}  # Dicionário para mapear documento_id -> associacao_id
    if plano.maquina:
        documentos_maquina = MaquinaDocumento.objects.filter(maquina=plano.maquina).order_by('-created_at')
        # Buscar associações existentes
        associacoes = MeuPlanoPreventivaDocumento.objects.filter(
            meu_plano_preventiva=plano
        ).select_related('maquina_documento').order_by('-created_at')
        documentos_associados_ids = list(associacoes.values_list('maquina_documento_id', flat=True))
        documentos_associados = MaquinaDocumento.objects.filter(id__in=documentos_associados_ids).order_by('-created_at')
        # Criar dicionário para facilitar busca no template
        for associacao in associacoes:
            associacoes_dict[associacao.maquina_documento.id] = associacao.id
    
    context = {
        'page_title': f'Editar Plano PCM - Plano {plano.numero_plano}',
        'active_page': 'consultar_meu_plano',
        'form': form,
        'plano': plano,
        'maquinas': maquinas,
        'roteiros': roteiros,
        'documentos_maquina': documentos_maquina,
        'documentos_associados': documentos_associados,
        'associacoes': associacoes,
        'documentos_associados_ids': documentos_associados_ids,
        'associacoes_dict': associacoes_dict,
    }
    return render(request, 'editar/editar_plano_pcm.html', context)


def associar_documento_plano_pcm(request, plano_id, documento_id):
    """Associar um documento de máquina a um MeuPlanoPreventiva"""
    from app.models import MeuPlanoPreventiva, MaquinaDocumento, MeuPlanoPreventivaDocumento
    
    # Aceitar tanto GET quanto POST
    if request.method not in ['GET', 'POST']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
        messages.error(request, 'Método não permitido.')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    
    try:
        plano = MeuPlanoPreventiva.objects.get(id=plano_id)
    except MeuPlanoPreventiva.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Plano PCM não encontrado'}, status=404)
        messages.error(request, 'Plano PCM não encontrado.')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    
    try:
        documento = MaquinaDocumento.objects.get(id=documento_id)
    except MaquinaDocumento.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Documento não encontrado'}, status=404)
        messages.error(request, 'Documento não encontrado.')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    
    # Verificar se já está associado
    if MeuPlanoPreventivaDocumento.objects.filter(meu_plano_preventiva=plano, maquina_documento=documento).exists():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Este documento já está associado a este plano'}, status=400)
        messages.warning(request, 'Este documento já está associado a este plano.')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    
    # Criar associação
    try:
        comentario = request.POST.get('comentario', '').strip() if request.method == 'POST' else ''
        
        # Debug: imprimir informações
        print(f"Associando documento {documento_id} ao plano {plano_id}")
        print(f"Plano: {plano}")
        print(f"Documento: {documento}")
        print(f"Comentário: {comentario}")
        
        associacao = MeuPlanoPreventivaDocumento.objects.create(
            meu_plano_preventiva=plano,
            maquina_documento=documento,
            comentario=comentario if comentario else None
        )
        
        print(f"Associação criada com sucesso! ID: {associacao.id}")
        
        # Se for requisição AJAX, retornar JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Documento associado com sucesso!',
                'associacao_id': associacao.id
            })
        
        messages.success(request, 'Documento associado com sucesso!')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    except Exception as e:
        import traceback
        print(f"ERRO ao associar documento {documento_id} ao plano {plano_id}:")
        traceback.print_exc()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': f'Erro ao associar documento: {str(e)}'}, status=500)
        messages.error(request, f'Erro ao associar documento: {str(e)}')
        return redirect('editar_plano_pcm', plano_id=plano_id)


def remover_documento_plano_pcm(request, plano_id, associacao_id):
    """Remover associação de documento de um MeuPlanoPreventiva"""
    from app.models import MeuPlanoPreventiva, MeuPlanoPreventivaDocumento
    
    try:
        plano = MeuPlanoPreventiva.objects.get(id=plano_id)
    except MeuPlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano PCM não encontrado.')
        return redirect('editar_plano_pcm', plano_id=plano_id)
    
    try:
        associacao = MeuPlanoPreventivaDocumento.objects.get(id=associacao_id, meu_plano_preventiva=plano)
        associacao.delete()
        messages.success(request, 'Associação de documento removida com sucesso!')
    except MeuPlanoPreventivaDocumento.DoesNotExist:
        messages.error(request, 'Associação não encontrada.')
    
    return redirect('editar_plano_pcm', plano_id=plano_id)


def adicionar_documento_plano_preventiva(request, plano_id):
    """Adicionar documento a um plano preventiva"""
    from app.models import PlanoPreventiva, PlanoPreventivaDocumento
    from app.forms import PlanoPreventivaDocumentoForm
    import os
    
    try:
        plano = PlanoPreventiva.objects.get(id=plano_id)
    except PlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano preventiva não encontrado.')
        return redirect('consultar_manutencoes_preventivas')
    
    if request.method == 'POST':
        print(f"DEBUG - Método POST recebido")
        print(f"DEBUG - request.FILES: {list(request.FILES.keys())}")
        print(f"DEBUG - request.POST: {dict(request.POST)}")
        
        # Verificar se arquivo foi enviado
        if 'arquivo' not in request.FILES:
            print("DEBUG - Arquivo não encontrado em request.FILES")
            messages.error(request, 'Por favor, selecione um arquivo para upload.')
            return redirect('visualizar_manutencao_preventiva', plano_id=plano_id)
        
        arquivo = request.FILES['arquivo']
        comentario = request.POST.get('comentario', '').strip()
        
        print(f"DEBUG - Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Content-Type: {arquivo.content_type}")
        print(f"DEBUG - Comentário: {comentario}")
        print(f"DEBUG - Plano ID: {plano.id}, Plano Preventiva: {plano}")
        
        # Criar documento diretamente
        try:
            documento = PlanoPreventivaDocumento(
                plano_preventiva=plano,
                arquivo=arquivo,
                comentario=comentario if comentario else None
            )
            documento.full_clean()  # Validar antes de salvar
            documento.save()
            print(f"DEBUG - Documento criado com sucesso! ID: {documento.id}, Arquivo: {documento.arquivo.name}")
            messages.success(request, 'Documento adicionado com sucesso!')
        except Exception as e:
            print(f"DEBUG - Erro ao criar documento: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Erro ao adicionar documento: {str(e)}')
    else:
        print(f"DEBUG - Método não é POST: {request.method}")
        messages.error(request, 'Método não permitido.')
    
    return redirect('visualizar_manutencao_preventiva', plano_id=plano_id)


def remover_documento_plano_preventiva(request, plano_id, documento_id):
    """Remover documento de um plano preventiva"""
    from app.models import PlanoPreventiva, PlanoPreventivaDocumento
    import os
    
    try:
        plano = PlanoPreventiva.objects.get(id=plano_id)
    except PlanoPreventiva.DoesNotExist:
        messages.error(request, 'Plano preventiva não encontrado.')
        return redirect('consultar_manutencoes_preventivas')
    
    try:
        documento = PlanoPreventivaDocumento.objects.get(id=documento_id, plano_preventiva=plano)
        # Deletar arquivo físico se existir
        if documento.arquivo:
            if os.path.isfile(documento.arquivo.path):
                os.remove(documento.arquivo.path)
        documento.delete()
        messages.success(request, 'Documento removido com sucesso!')
    except PlanoPreventivaDocumento.DoesNotExist:
        messages.error(request, 'Documento não encontrado.')
    
    return redirect('visualizar_manutencao_preventiva', plano_id=plano_id)


def adicionar_documento_maquina(request, maquina_id):
    """Adicionar documento a uma máquina"""
    from app.models import Maquina, MaquinaDocumento
    import os
    
    print(f"DEBUG - adicionar_documento_maquina chamado. Método: {request.method}, maquina_id: {maquina_id}")
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        print(f"DEBUG - Máquina {maquina_id} não encontrada")
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    if request.method == 'POST':
        print(f"DEBUG - Método POST recebido")
        print(f"DEBUG - request.FILES: {list(request.FILES.keys())}")
        print(f"DEBUG - request.POST: {dict(request.POST)}")
        
        # Verificar se arquivo foi enviado
        if 'arquivo' not in request.FILES:
            print("DEBUG - Arquivo não encontrado em request.FILES")
            messages.error(request, 'Por favor, selecione um arquivo para upload.')
            return redirect('editar_maquina', maquina_id=maquina_id)
        
        arquivo = request.FILES['arquivo']
        comentario = request.POST.get('comentario', '').strip()
        
        print(f"DEBUG - Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Content-Type: {arquivo.content_type}")
        print(f"DEBUG - Comentário: {comentario}")
        print(f"DEBUG - Máquina ID: {maquina.id}, Máquina: {maquina}")
        
        # Criar documento diretamente
        try:
            documento = MaquinaDocumento(
                maquina=maquina,
                arquivo=arquivo,
                comentario=comentario if comentario else None
            )
            documento.full_clean()  # Validar antes de salvar
            documento.save()
            print(f"DEBUG - Documento criado com sucesso! ID: {documento.id}, Arquivo: {documento.arquivo.name}")
            messages.success(request, 'Documento adicionado com sucesso!')
        except Exception as e:
            print(f"DEBUG - Erro ao criar documento: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Erro ao adicionar documento: {str(e)}')
    else:
        print(f"DEBUG - Método não é POST: {request.method}")
        messages.error(request, 'Método não permitido.')
    
    return redirect('editar_maquina', maquina_id=maquina_id)


def remover_documento_maquina(request, maquina_id, documento_id):
    """Remover documento de uma máquina"""
    from app.models import Maquina, MaquinaDocumento
    import os
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    try:
        documento = MaquinaDocumento.objects.get(id=documento_id, maquina=maquina)
        # Deletar arquivo físico se existir
        if documento.arquivo:
            if os.path.isfile(documento.arquivo.path):
                os.remove(documento.arquivo.path)
        documento.delete()
        messages.success(request, 'Documento removido com sucesso!')
    except MaquinaDocumento.DoesNotExist:
        messages.error(request, 'Documento não encontrado.')
    
    return redirect('editar_maquina', maquina_id=maquina_id)


def visualizar_maquina(request, maquina_id):
    """Visualizar detalhes de uma máquina específica"""
    from app.models import Maquina, ItemEstoque, MaquinaPeca, MaquinaPrimariaSecundaria, PlanoPreventiva, MaquinaDocumento, MeuPlanoPreventiva
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    # Buscar peças relacionadas a esta máquina
    pecas_relacionadas = MaquinaPeca.objects.filter(maquina=maquina).select_related('item_estoque').order_by('-created_at')
    
    # Buscar todos os itens de estoque para seleção (excluindo os já relacionados)
    itens_estoque_ids = pecas_relacionadas.values_list('item_estoque_id', flat=True)
    itens_disponiveis = ItemEstoque.objects.exclude(id__in=itens_estoque_ids).order_by('codigo_item')[:100]  # Limitar a 100 para performance
    
    # Buscar relacionamentos onde esta máquina é primária
    relacionamentos_como_primaria = MaquinaPrimariaSecundaria.objects.filter(
        maquina_primaria=maquina
    ).select_related('maquina_secundaria').order_by('maquina_secundaria__cd_maquina')
    
    # Buscar relacionamentos onde esta máquina é secundária
    relacionamentos_como_secundaria = MaquinaPrimariaSecundaria.objects.filter(
        maquina_secundaria=maquina
    ).select_related('maquina_primaria').order_by('maquina_primaria__cd_maquina')
    
    # Buscar planos preventiva relacionados a esta máquina
    # Primeiro pelo relacionamento direto, depois pelo código da máquina
    planos_preventiva = PlanoPreventiva.objects.filter(
        Q(maquina=maquina) | Q(cd_maquina=maquina.cd_maquina)
    ).order_by('numero_plano', 'sequencia_manutencao', 'sequencia_tarefa')
    
    # Buscar MeuPlanoPreventiva relacionados a esta máquina
    meus_planos_preventiva = MeuPlanoPreventiva.objects.filter(
        Q(maquina=maquina) | Q(cd_maquina=maquina.cd_maquina)
    ).order_by('dt_execucao', 'numero_plano', 'sequencia_manutencao')
    
    # Buscar documentos relacionados a esta máquina
    documentos_maquina = MaquinaDocumento.objects.filter(maquina=maquina).order_by('-created_at')
    
    # Verificar se é máquina principal
    is_maquina_principal = maquina.descr_gerenc and 'MÁQUINAS PRINCIPAL' in maquina.descr_gerenc.upper()
    
    # Se for máquina principal, buscar IDs das máquinas secundárias
    maquinas_secundarias_ids = []
    if is_maquina_principal and relacionamentos_como_primaria.exists():
        maquinas_secundarias_ids = relacionamentos_como_primaria.values_list('maquina_secundaria_id', flat=True)
    
    context = {
        'page_title': f'Visualizar Máquina {maquina.cd_maquina}',
        'active_page': 'consultar_maquinas',
        'maquina': maquina,
        'pecas_relacionadas': pecas_relacionadas,
        'itens_disponiveis': itens_disponiveis,
        'relacionamentos_como_primaria': relacionamentos_como_primaria,
        'relacionamentos_como_secundaria': relacionamentos_como_secundaria,
        'planos_preventiva': planos_preventiva,
        'meus_planos_preventiva': meus_planos_preventiva,
        'documentos_maquina': documentos_maquina,
        'is_maquina_principal': is_maquina_principal,
        'maquinas_secundarias_ids': list(maquinas_secundarias_ids),
    }
    return render(request, 'maquinas/visualizar_maquina.html', context)


def calendario_planos_maquina(request, maquina_id):
    """Endpoint JSON para fornecer eventos do calendário de MeuPlanoPreventiva para uma máquina"""
    from app.models import Maquina, MeuPlanoPreventiva
    from django.http import JsonResponse
    from datetime import datetime
    from django.db.models import Q
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        return JsonResponse({'error': 'Máquina não encontrada'}, status=404)
    
    # Buscar MeuPlanoPreventiva relacionados a esta máquina
    planos = MeuPlanoPreventiva.objects.filter(
        Q(maquina=maquina) | Q(cd_maquina=maquina.cd_maquina)
    ).exclude(
        dt_execucao__isnull=True
    ).exclude(
        dt_execucao=''
    )
    
    # Converter para formato de eventos do FullCalendar
    events = []
    for plano in planos:
        if plano.dt_execucao:
            try:
                # Tentar parsear a data (formato DD/MM/YYYY ou YYYY-MM-DD)
                data_str = plano.dt_execucao.strip()
                if '/' in data_str:
                    # Formato DD/MM/YYYY
                    data_obj = datetime.strptime(data_str, '%d/%m/%Y').date()
                elif '-' in data_str:
                    # Formato YYYY-MM-DD
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                else:
                    continue  # Pular se não conseguir parsear
                
                # Criar título do evento
                titulo_parts = []
                if plano.numero_plano:
                    titulo_parts.append(f"Plano {plano.numero_plano}")
                if plano.sequencia_manutencao:
                    titulo_parts.append(f"Seq: {plano.sequencia_manutencao}")
                if plano.descr_tarefa:
                    titulo_parts.append(plano.descr_tarefa[:50])
                
                titulo = " - ".join(titulo_parts) if titulo_parts else f"Manutenção Preventiva - {plano.cd_maquina}"
                
                # Criar descrição/tooltip
                descricao_parts = []
                if plano.descr_tarefa:
                    descricao_parts.append(f"Tarefa: {plano.descr_tarefa}")
                if plano.nome_funcionario:
                    descricao_parts.append(f"Funcionário: {plano.nome_funcionario}")
                if plano.descr_setor:
                    descricao_parts.append(f"Setor: {plano.descr_setor}")
                if plano.quantidade_periodo:
                    descricao_parts.append(f"Período: {plano.quantidade_periodo} dias")
                
                descricao = "\n".join(descricao_parts)
                
                # Determinar cor baseada em informações do plano
                cor = '#3788d8'  # Azul padrão
                if plano.quantidade_periodo and plano.quantidade_periodo > 30:
                    cor = '#dc3545'  # Vermelho para períodos longos
                elif plano.quantidade_periodo and plano.quantidade_periodo <= 7:
                    cor = '#28a745'  # Verde para períodos curtos
                
                events.append({
                    'id': plano.id,
                    'title': titulo,
                    'start': data_obj.isoformat(),
                    'allDay': True,
                    'backgroundColor': cor,
                    'borderColor': cor,
                    'textColor': '#ffffff',
                    'extendedProps': {
                        'plano_id': plano.id,
                        'numero_plano': plano.numero_plano,
                        'sequencia_manutencao': plano.sequencia_manutencao,
                        'descricao': descricao,
                        'url': f"/plano-pcm/visualizar/{plano.id}/" if plano.id else None,
                    }
                })
            except (ValueError, AttributeError):
                # Se não conseguir parsear a data, pular este plano
                continue
    
    return JsonResponse(events, safe=False)


def calendario_planos_secundarias(request, maquina_id):
    """Endpoint JSON para fornecer eventos do calendário de MeuPlanoPreventiva para máquinas secundárias de uma máquina principal"""
    from app.models import Maquina, MeuPlanoPreventiva, MaquinaPrimariaSecundaria
    from django.http import JsonResponse
    from datetime import datetime
    from django.db.models import Q
    
    try:
        maquina_principal = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        return JsonResponse({'error': 'Máquina não encontrada'}, status=404)
    
    # Verificar se é máquina principal
    is_maquina_principal = maquina_principal.descr_gerenc and 'MÁQUINAS PRINCIPAL' in maquina_principal.descr_gerenc.upper()
    
    if not is_maquina_principal:
        return JsonResponse({'error': 'Esta máquina não é uma máquina principal'}, status=400)
    
    # Buscar máquinas secundárias relacionadas
    relacionamentos = MaquinaPrimariaSecundaria.objects.filter(
        maquina_primaria=maquina_principal
    ).select_related('maquina_secundaria')
    
    if not relacionamentos.exists():
        return JsonResponse([], safe=False)  # Retornar lista vazia se não houver máquinas secundárias
    
    # Obter IDs e códigos das máquinas secundárias
    maquinas_secundarias_ids = relacionamentos.values_list('maquina_secundaria_id', flat=True)
    maquinas_secundarias_codigos = relacionamentos.values_list('maquina_secundaria__cd_maquina', flat=True)
    
    # Buscar MeuPlanoPreventiva relacionados às máquinas secundárias
    planos = MeuPlanoPreventiva.objects.filter(
        Q(maquina_id__in=maquinas_secundarias_ids) | Q(cd_maquina__in=maquinas_secundarias_codigos)
    ).exclude(
        dt_execucao__isnull=True
    ).exclude(
        dt_execucao=''
    )
    
    # Converter para formato de eventos do FullCalendar
    events = []
    for plano in planos:
        if plano.dt_execucao:
            try:
                # Tentar parsear a data (formato DD/MM/YYYY ou YYYY-MM-DD)
                data_str = plano.dt_execucao.strip()
                if '/' in data_str:
                    # Formato DD/MM/YYYY
                    data_obj = datetime.strptime(data_str, '%d/%m/%Y').date()
                elif '-' in data_str:
                    # Formato YYYY-MM-DD
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                else:
                    continue  # Pular se não conseguir parsear
                
                # Criar título do evento (incluir código da máquina secundária)
                titulo_parts = []
                titulo_parts.append(f"Máq: {plano.cd_maquina}")
                if plano.numero_plano:
                    titulo_parts.append(f"Plano {plano.numero_plano}")
                if plano.sequencia_manutencao:
                    titulo_parts.append(f"Seq: {plano.sequencia_manutencao}")
                if plano.descr_tarefa:
                    titulo_parts.append(plano.descr_tarefa[:40])
                
                titulo = " - ".join(titulo_parts) if titulo_parts else f"Manutenção Preventiva - {plano.cd_maquina}"
                
                # Criar descrição/tooltip
                descricao_parts = []
                descricao_parts.append(f"Máquina: {plano.cd_maquina} - {plano.descr_maquina or 'Sem descrição'}")
                if plano.descr_tarefa:
                    descricao_parts.append(f"Tarefa: {plano.descr_tarefa}")
                if plano.nome_funcionario:
                    descricao_parts.append(f"Funcionário: {plano.nome_funcionario}")
                if plano.descr_setor:
                    descricao_parts.append(f"Setor: {plano.descr_setor}")
                if plano.quantidade_periodo:
                    descricao_parts.append(f"Período: {plano.quantidade_periodo} dias")
                
                descricao = "\n".join(descricao_parts)
                
                # Determinar cor baseada em informações do plano (usar cor diferente para distinguir)
                cor = '#6c757d'  # Cinza para máquinas secundárias
                if plano.quantidade_periodo and plano.quantidade_periodo > 30:
                    cor = '#dc3545'  # Vermelho para períodos longos
                elif plano.quantidade_periodo and plano.quantidade_periodo <= 7:
                    cor = '#28a745'  # Verde para períodos curtos
                
                # Buscar ID da máquina relacionada para criar link
                maquina_relacionada_id = None
                if plano.maquina_id:
                    maquina_relacionada_id = plano.maquina_id
                else:
                    # Tentar encontrar pelo código
                    try:
                        maquina_obj = Maquina.objects.get(cd_maquina=plano.cd_maquina)
                        maquina_relacionada_id = maquina_obj.id
                    except Maquina.DoesNotExist:
                        pass
                
                events.append({
                    'id': f'sec_{plano.id}',
                    'title': titulo,
                    'start': data_obj.isoformat(),
                    'allDay': True,
                    'backgroundColor': cor,
                    'borderColor': cor,
                    'textColor': '#ffffff',
                    'extendedProps': {
                        'plano_id': plano.id,
                        'maquina_id': maquina_relacionada_id,
                        'maquina_codigo': plano.cd_maquina,
                        'numero_plano': plano.numero_plano,
                        'sequencia_manutencao': plano.sequencia_manutencao,
                        'descricao': descricao,
                        'url': f"/plano-pcm/visualizar/{plano.id}/" if plano.id else None,
                        'maquina_url': f"/maquinas/visualizar/{maquina_relacionada_id}/" if maquina_relacionada_id else None,
                    }
                })
            except (ValueError, AttributeError):
                # Se não conseguir parsear a data, pular este plano
                continue
    
    return JsonResponse(events, safe=False)


def editar_maquina(request, maquina_id):
    """Editar uma máquina existente"""
    from app.forms import MaquinaForm
    from app.models import Maquina
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')

    # Valores únicos de descr_gerenc no DB para o select
    gerenc_choices = list(
        Maquina.objects.exclude(descr_gerenc__isnull=True)
        .exclude(descr_gerenc='')
        .values_list('descr_gerenc', flat=True)
        .distinct()
        .order_by('descr_gerenc')
    )
    
    if request.method == 'POST':
        print(f"DEBUG - POST recebido para editar máquina {maquina_id}")
        print(f"DEBUG - request.FILES: {request.FILES}")
        print(f"DEBUG - 'foto' in request.FILES: {'foto' in request.FILES}")
        if 'foto' in request.FILES:
            print(f"DEBUG - Arquivo recebido: {request.FILES['foto'].name}, Tamanho: {request.FILES['foto'].size}")
        print(f"DEBUG - request.POST: {request.POST}")
        
        form = MaquinaForm(request.POST, request.FILES, instance=maquina, gerenc_choices=gerenc_choices)
        print(f"DEBUG - Form criado. is_valid(): {form.is_valid()}")
        
        if form.is_valid():
            try:
                print(f"DEBUG - Antes de salvar. Foto atual: {maquina.foto}")
                maquina = form.save()
                print(f"DEBUG - Máquina salva com sucesso. Foto: {maquina.foto}")
                print(f"DEBUG - Foto URL: {maquina.foto.url if maquina.foto else 'N/A'}")
                messages.success(request, f'Máquina {maquina.cd_maquina} atualizada com sucesso!')
                return redirect('visualizar_maquina', maquina_id=maquina.id)
            except Exception as e:
                print(f"DEBUG - Erro ao salvar: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Erro ao atualizar máquina: {str(e)}')
        else:
            print(f"DEBUG - Form inválido. Erros: {form.errors}")
            for field, errors in form.errors.items():
                print(f"DEBUG - Campo {field}: {errors}")
            handle_form_errors(form, request)
    else:
        form = MaquinaForm(instance=maquina, gerenc_choices=gerenc_choices)
    
    # Buscar documentos relacionados à máquina
    from app.models import MaquinaDocumento
    documentos = MaquinaDocumento.objects.filter(maquina=maquina).order_by('-created_at')
    
    context = {
        'page_title': f'Editar Máquina {maquina.cd_maquina}',
        'active_page': 'consultar_maquinas',
        'form': form,
        'maquina': maquina,
        'documentos': documentos,
    }
    return render(request, 'editar/editar_maquina.html', context)


def filtrar_locais_por_setormanut(request):
    """View AJAX para filtrar CentroAtividade baseado no cd_setormanut"""
    from app.models import CentroAtividade
    
    cd_setormanut = request.GET.get('cd_setormanut', '')
    
    if not cd_setormanut:
        return JsonResponse({'centros': []})
    
    try:
        # Tentar converter cd_setormanut para inteiro para comparar com ca
        if cd_setormanut.isdigit():
            ca_value = int(cd_setormanut)
            centros = CentroAtividade.objects.filter(ca=ca_value).order_by('ca')
            
            centros_data = [
                {
                    'id': centro.id,
                    'ca': centro.ca,
                    'sigla': centro.sigla or '',
                    'local': centro.local or '',
                    'descricao': centro.descricao or '',
                    'observacoes': centro.observacoes or ''
                }
                for centro in centros
            ]
            return JsonResponse({'centros': centros_data})
        else:
            return JsonResponse({'centros': []})
    except (ValueError, AttributeError) as e:
        return JsonResponse({'error': str(e), 'centros': []})


def cadastrar_corretiva_outros(request):
    """Cadastrar nova ordem corretiva/outros"""
    from app.forms import OrdemServicoCorretivaForm
    
    if request.method == 'POST':
        form = OrdemServicoCorretivaForm(request.POST)
        if form.is_valid():
            try:
                ordem = form.save()
                messages.success(request, f'Ordem de serviço {ordem.cd_ordemserv} cadastrada com sucesso!')
                return redirect('consultar_corretivas_outros')
            except Exception as e:
                messages.error(request, f'Erro ao cadastrar ordem: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = OrdemServicoCorretivaForm()
    
    context = {
        'page_title': 'Cadastrar Ordem Corretiva/Outros',
        'active_page': 'cadastrar_corretiva_outros',
        'form': form
    }
    return render(request, 'cadastrar/cadastrar_corretiva_outros.html', context)


def consultar_corretivas_outros(request):
    """Consultar/listar ordens corretivas cadastradas com filtros avançados"""
    from app.models import OrdemServicoCorretiva
    from datetime import datetime
    
    # Buscar todas as ordens
    ordens_list = OrdemServicoCorretiva.objects.all()
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_ordemserv=search_num)
            search_conditions |= Q(cd_maquina=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(cd_setormanut__icontains=search_query) |
            Q(descr_setormanut__icontains=search_query) |
            Q(nm_func_solic_os__icontains=search_query) |
            Q(nm_func_exec__icontains=search_query) |
            Q(descr_queixa__icontains=search_query) |
            Q(exec_tarefas__icontains=search_query)
        )
        
        ordens_list = ordens_list.filter(search_conditions)
    
    # Filtros específicos
    # Filtro por Setor de Manutenção
    filtro_setor = request.GET.get('filtro_setor', '')
    if filtro_setor:
        ordens_list = ordens_list.filter(descr_setormanut__icontains=filtro_setor)
    
    # Filtro por Unidade
    filtro_unidade = request.GET.get('filtro_unidade', '')
    if filtro_unidade:
        ordens_list = ordens_list.filter(nome_unid__icontains=filtro_unidade)
    
    # Filtro por Tipo de Ordem de Serviço
    filtro_tipo_os = request.GET.get('filtro_tipo_os', '')
    if filtro_tipo_os:
        ordens_list = ordens_list.filter(descr_tpordservtv__icontains=filtro_tipo_os)
    
    # Filtro por Situação da Ordem
    filtro_situacao = request.GET.get('filtro_situacao', '')
    if filtro_situacao:
        ordens_list = ordens_list.filter(descr_sitordsetv__icontains=filtro_situacao)
    
    # Filtro por Funcionário Solicitante
    filtro_solicitante = request.GET.get('filtro_solicitante', '')
    if filtro_solicitante:
        ordens_list = ordens_list.filter(nm_func_solic_os__icontains=filtro_solicitante)
    
    # Filtro por Funcionário Executor
    filtro_executor = request.GET.get('filtro_executor', '')
    if filtro_executor:
        ordens_list = ordens_list.filter(
            Q(nm_func_exec__icontains=filtro_executor)
        )
    
    # Filtro por Código da Máquina
    filtro_maquina = request.GET.get('filtro_maquina', '')
    if filtro_maquina:
        ordens_list = ordens_list.filter(cd_maquina__icontains=filtro_maquina)
    
    # Filtro por Data de Entrada (período)
    data_entrada_inicio = request.GET.get('data_entrada_inicio', '')
    data_entrada_fim = request.GET.get('data_entrada_fim', '')
    if data_entrada_inicio:
        try:
            data_inicio = datetime.strptime(data_entrada_inicio, '%Y-%m-%d')
            ordens_list = ordens_list.filter(created_at__gte=data_inicio)
        except ValueError:
            pass
    if data_entrada_fim:
        try:
            data_fim = datetime.strptime(data_entrada_fim, '%Y-%m-%d')
            # Adicionar 1 dia para incluir o dia final
            from datetime import timedelta
            data_fim = data_fim + timedelta(days=1)
            ordens_list = ordens_list.filter(created_at__lte=data_fim)
        except ValueError:
            pass
    
    # Filtro por Status da Ordem (Abertas/Fechadas)
    filtro_ordens_abertas = request.GET.get('filtro_ordens_abertas', '')
    filtro_ordens_fechadas = request.GET.get('filtro_ordens_fechadas', '')
    
    # Converter para boolean (se existe e não é vazio, é True)
    filtro_ordens_abertas = filtro_ordens_abertas == '1'
    filtro_ordens_fechadas = filtro_ordens_fechadas == '1'
    
    # Aplicar filtros baseado nos checkboxes marcados
    if filtro_ordens_abertas and filtro_ordens_fechadas:
        # Ambos marcados: mostrar todas (não aplicar filtro)
        pass
    elif filtro_ordens_abertas and not filtro_ordens_fechadas:
        # Apenas "Ordens Abertas" marcado: dt_encordmanu está vazio ou nulo
        ordens_list = ordens_list.filter(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    elif filtro_ordens_fechadas and not filtro_ordens_abertas:
        # Apenas "Ordens Fechadas" marcado: dt_encordmanu tem valor (não é nulo nem vazio)
        ordens_list = ordens_list.exclude(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    # Se nenhum está marcado, mostra todas (não aplicar filtro)
    
    # Ordenar por código da ordem de serviço (mais recente primeiro)
    ordens_list = ordens_list.order_by('-cd_ordemserv')
    
    # Paginação
    paginator = Paginator(ordens_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    ordens = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = OrdemServicoCorretiva.objects.count()
    setores_count = OrdemServicoCorretiva.objects.exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    unidades_count = OrdemServicoCorretiva.objects.exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros
    setores_unicos = OrdemServicoCorretiva.objects.exclude(
        descr_setormanut__isnull=True
    ).exclude(
        descr_setormanut=''
    ).values_list('descr_setormanut', flat=True).distinct().order_by('descr_setormanut')
    
    unidades_unicas = OrdemServicoCorretiva.objects.exclude(
        nome_unid__isnull=True
    ).exclude(
        nome_unid=''
    ).values_list('nome_unid', flat=True).distinct().order_by('nome_unid')
    
    tipos_os_unicos = OrdemServicoCorretiva.objects.exclude(
        descr_tpordservtv__isnull=True
    ).exclude(
        descr_tpordservtv=''
    ).values_list('descr_tpordservtv', flat=True).distinct().order_by('descr_tpordservtv')
    
    situacoes_unicas = OrdemServicoCorretiva.objects.exclude(
        descr_sitordsetv__isnull=True
    ).exclude(
        descr_sitordsetv=''
    ).values_list('descr_sitordsetv', flat=True).distinct().order_by('descr_sitordsetv')
    
    context = {
        'page_title': 'Consultar Ordens Corretivas/Outros',
        'active_page': 'consultar_corretivas_outros',
        'ordens': ordens,
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        # Valores para dropdowns
        'setores_unicos': setores_unicos,
        'unidades_unicas': unidades_unicas,
        'tipos_os_unicos': tipos_os_unicos,
        'situacoes_unicas': situacoes_unicas,
        # Valores dos filtros ativos
        'filtro_setor': filtro_setor,
        'filtro_unidade': filtro_unidade,
        'filtro_tipo_os': filtro_tipo_os,
        'filtro_situacao': filtro_situacao,
        'filtro_solicitante': filtro_solicitante,
        'filtro_executor': filtro_executor,
        'filtro_maquina': filtro_maquina,
        'data_entrada_inicio': data_entrada_inicio,
        'data_entrada_fim': data_entrada_fim,
        'filtro_ordens_abertas': '1' if filtro_ordens_abertas else '',
        'filtro_ordens_fechadas': '1' if filtro_ordens_fechadas else '',
    }
    return render(request, 'consultar/consultar_corretivas_outros.html', context)


def visualizar_corretiva_outros(request, ordem_id):
    """Visualizar detalhes de uma ordem corretiva específica"""
    from app.models import OrdemServicoCorretiva, Maquina
    
    try:
        ordem = OrdemServicoCorretiva.objects.get(id=ordem_id)
    except OrdemServicoCorretiva.DoesNotExist:
        messages.error(request, 'Ordem de serviço não encontrada.')
        return redirect('consultar_corretivas_outros')
    
    # Buscar a máquina correspondente se cd_maquina existir
    maquina = None
    if ordem.cd_maquina:
        try:
            maquina = Maquina.objects.get(cd_maquina=ordem.cd_maquina)
        except Maquina.DoesNotExist:
            maquina = None
    
    # Buscar fichas relacionadas (pode haver múltiplas fichas)
    fichas = ordem.fichas.all().order_by('-created_at')
    
    context = {
        'page_title': f'Visualizar OS {ordem.cd_ordemserv}',
        'active_page': 'consultar_corretivas_outros',
        'ordem': ordem,
        'maquina': maquina,
        'fichas': fichas,
    }
    return render(request, 'visualizar/visualizar_corretiva_outros.html', context)


def analise_corretiva_outros(request):
    """Página inicial da seção Manutenção Corretiva com análises e gráficos"""
    from app.models import OrdemServicoCorretiva, Maquina, CentroAtividade
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses (None)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Função para fazer parse de dt_abertura_solicita
    def parse_dt_abertura_solicita(date_str):
        """Tenta fazer parse de dt_abertura_solicita em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para filtrar ordens baseado em dt_abertura_solicita
    def filtrar_ordens_por_data(queryset, ano, meses=None):
        """Filtra ordens baseado em dt_abertura_solicita
        meses: lista de inteiros (1-12) ou None para todos os meses
        """
        ordens_filtradas = []
        for ordem in queryset:
            if ordem.dt_abertura_solicita:
                data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
                if data_parseada:
                    if data_parseada.year == ano:
                        if meses is None or len(meses) == 0 or data_parseada.month in meses:
                            ordens_filtradas.append(ordem)
        return ordens_filtradas
    
    # Obter todas as ordens e filtrar
    todas_ordens = OrdemServicoCorretiva.objects.all()
    ordens_filtradas = filtrar_ordens_por_data(todas_ordens, ano_filtro, meses_filtro_int if meses_filtro_int else None)
    
    # Estatísticas básicas (filtradas)
    total_count = len(ordens_filtradas)
    # Contar setores, unidades e máquinas únicos (filtrados)
    setores_unicos = set()
    unidades_unicas = set()
    for ordem in ordens_filtradas:
        if ordem.cd_setormanut:
            setores_unicos.add(ordem.cd_setormanut)
        if ordem.nome_unid:
            unidades_unicas.add(ordem.nome_unid)
    setores_count = len(setores_unicos)
    unidades_count = len(unidades_unicas)
    maquinas_count = Maquina.objects.count()
    
    # Ordens por setor (top 10) - filtradas
    ordens_por_setor_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_setormanut:
            ordens_por_setor_dict[ordem.descr_setormanut] += 1
    ordens_por_setor = sorted(ordens_por_setor_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    setores_labels = [item[0][:30] for item in ordens_por_setor]
    setores_data = [item[1] for item in ordens_por_setor]
    
    # Ordens por unidade (top 10) - filtradas
    ordens_por_unidade_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nome_unid:
            ordens_por_unidade_dict[ordem.nome_unid] += 1
    ordens_por_unidade = sorted(ordens_por_unidade_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    unidades_labels = [item[0][:30] for item in ordens_por_unidade]
    unidades_data = [item[1] for item in ordens_por_unidade]
    
    # Função para fazer parse de dt_abertura_solicita
    def parse_dt_abertura_solicita(date_str):
        """Tenta fazer parse de dt_abertura_solicita em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir (formato: "dd/mm/yyyy hh:mm" ou "dd/mm/yyyy hh:mm:ss")
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y',      # 26/09/2025
            '%d-%m-%Y',      # 26-09-2025
            '%d.%m.%Y',      # 26.09.2025
            '%Y-%m-%d',      # 2025-09-26
            '%Y/%m/%d',      # 2025/09/26
            '%d/%m/%y',      # 26/09/25
            '%d-%m-%y',      # 26-09-25
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Se nenhum formato funcionou, tentar parse manual para formato brasileiro comum
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    # Se ano tem 2 dígitos, assumir 2000+
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para fazer parse de dt_encordmanu (mesma lógica)
    def parse_dt_encordmanu(date_str):
        """Tenta fazer parse de dt_encordmanu em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Ordens por mês do ano filtrado baseado em dt_abertura_solicita
    ordens_por_mes = defaultdict(int)
    
    # Ordens abertas vs fechadas por mês (para gráfico comparativo)
    ordens_abertas_por_mes = defaultdict(int)
    ordens_fechadas_por_mes = defaultdict(int)
    
    # Se meses específicos foram selecionados, mostrar apenas esses meses
    if meses_filtro_int and len(meses_filtro_int) > 0:
        meses_para_mostrar = meses_filtro_int
    else:
        meses_para_mostrar = list(range(1, 13))
    
    # Contar ordens abertas e fechadas por mês
    for ordem in ordens_filtradas:
        # Contar abertas (dt_abertura_solicita)
        if ordem.dt_abertura_solicita:
            data_abertura = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_abertura and data_abertura.year == ano_filtro:
                mes_ano = data_abertura.strftime('%Y-%m')
                ordens_por_mes[mes_ano] += 1
                ordens_abertas_por_mes[mes_ano] += 1
        
        # Contar fechadas (dt_encordmanu)
        if ordem.dt_encordmanu:
            data_fechamento = parse_dt_encordmanu(ordem.dt_encordmanu)
            if data_fechamento and data_fechamento.year == ano_filtro:
                mes_ano = data_fechamento.strftime('%Y-%m')
                ordens_fechadas_por_mes[mes_ano] += 1
    
    # Preencher todos os meses do ano (ou apenas o mês selecionado)
    for mes in meses_para_mostrar:
        mes_ano = f"{ano_filtro}-{mes:02d}"
        if mes_ano not in ordens_por_mes:
            ordens_por_mes[mes_ano] = 0
        if mes_ano not in ordens_abertas_por_mes:
            ordens_abertas_por_mes[mes_ano] = 0
        if mes_ano not in ordens_fechadas_por_mes:
            ordens_fechadas_por_mes[mes_ano] = 0
    
    # Ordenar por data
    meses_ordenados = sorted(ordens_por_mes.keys())
    
    # Formatar labels para português brasileiro
    meses_abrev = {
        'Jan': 'Jan', 'Feb': 'Fev', 'Mar': 'Mar', 'Apr': 'Abr', 'May': 'Mai', 'Jun': 'Jun',
        'Jul': 'Jul', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Nov': 'Nov', 'Dec': 'Dez'
    }
    meses_labels = []
    for m in meses_ordenados:
        dt = datetime.strptime(m, '%Y-%m')
        mes_abrev = meses_abrev.get(dt.strftime('%b'), dt.strftime('%b'))
        meses_labels.append(f"{mes_abrev}/{dt.strftime('%Y')}")
    meses_data = [ordens_por_mes[m] for m in meses_ordenados]
    
    # Dados para gráfico comparativo (abertas vs fechadas)
    comparativo_labels = meses_labels.copy()
    comparativo_abertas_data = [ordens_abertas_por_mes[m] for m in meses_ordenados]
    comparativo_fechadas_data = [ordens_fechadas_por_mes[m] for m in meses_ordenados]
    
    # Top 10 máquinas com mais ordens - filtradas
    maquinas_dict = defaultdict(int)
    maquinas_desc = {}
    for ordem in ordens_filtradas:
        if ordem.cd_maquina and ordem.descr_maquina:
            maquinas_dict[ordem.cd_maquina] += 1
            maquinas_desc[ordem.cd_maquina] = ordem.descr_maquina
    top_maquinas_list = sorted(maquinas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    # Converter para formato compatível com template
    top_maquinas = [{'cd_maquina': item[0], 'descr_maquina': maquinas_desc.get(item[0], ''), 'total': item[1]} for item in top_maquinas_list]
    maquinas_labels = [f"{item['cd_maquina']} - {item['descr_maquina'][:40]}" for item in top_maquinas]
    maquinas_data = [item['total'] for item in top_maquinas]
    
    # Top 10 executores - filtradas
    executores_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nm_func_exec:
            executores_dict[ordem.nm_func_exec] += 1
    top_executores_list = sorted(executores_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    # Converter para formato compatível com template (usando nm_func_exec_os para compatibilidade)
    top_executores = [{'nm_func_exec_os': item[0], 'total': item[1]} for item in top_executores_list]
    executores_labels = [item['nm_func_exec_os'][:30] for item in top_executores]
    executores_data = [item['total'] for item in top_executores]
    
    # Distribuição por tipo de ordem de serviço - filtradas
    tipos_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_tpordservtv:
            tipos_dict[ordem.descr_tpordservtv] += 1
    ordens_por_tipo = sorted(tipos_dict.items(), key=lambda x: x[1], reverse=True)[:8]
    tipos_labels = [item[0][:30] for item in ordens_por_tipo]
    tipos_data = [item[1] for item in ordens_por_tipo]
    
    # Ordens por local (baseado em CentroAtividade)
    # Join OrdemServicoCorretiva com CentroAtividade onde cd_setormanut = sigla
    ordens_por_local = defaultdict(int)
    
    # Buscar todos os centros de atividade e criar um dicionário de sigla -> local
    centros_dict = {}
    for centro in CentroAtividade.objects.exclude(sigla__isnull=True).exclude(sigla=''):
        centros_dict[centro.sigla] = centro.local if centro.local else 'Indefinido'
    
    # Contar ordens por local usando o dicionário - filtradas
    for ordem in ordens_filtradas:
        if ordem.cd_setormanut:
            sigla = ordem.cd_setormanut
            local = centros_dict.get(sigla, 'Indefinido')
            ordens_por_local[local] += 1
    
    # Ordenar por quantidade e calcular percentuais
    ordens_por_local_sorted = sorted(ordens_por_local.items(), key=lambda x: x[1], reverse=True)
    local_labels = [item[0] for item in ordens_por_local_sorted]
    local_data = [item[1] for item in ordens_por_local_sorted]
    
    # Calcular percentuais
    total_ordens_local = sum(local_data)
    local_percentages = [(count / total_ordens_local * 100) if total_ordens_local > 0 else 0 for count in local_data]
    
    # Estatística 1: Ordens com dt_encordmanu vs sem dt_encordmanu
    ordens_com_fechamento = sum(1 for ordem in ordens_filtradas if ordem.dt_encordmanu and ordem.dt_encordmanu.strip())
    ordens_sem_fechamento = total_count - ordens_com_fechamento
    percentual_com_fechamento = (ordens_com_fechamento / total_count * 100) if total_count > 0 else 0
    percentual_sem_fechamento = (ordens_sem_fechamento / total_count * 100) if total_count > 0 else 0
    
    # Estatística 2: Ordens com nm_func_exec vs sem nm_func_exec
    ordens_com_executor = sum(1 for ordem in ordens_filtradas if ordem.nm_func_exec and ordem.nm_func_exec.strip())
    ordens_sem_executor = total_count - ordens_com_executor
    percentual_com_executor = (ordens_com_executor / total_count * 100) if total_count > 0 else 0
    percentual_sem_executor = (ordens_sem_executor / total_count * 100) if total_count > 0 else 0
    
    # Estatística 3: Ordens que foram abertas e fechadas no mesmo mês (regra do gráfico comparativo)
    ordens_mesmo_mes = 0
    for ordem in ordens_filtradas:
        if ordem.dt_abertura_solicita and ordem.dt_encordmanu:
            data_abertura = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            data_fechamento = parse_dt_encordmanu(ordem.dt_encordmanu)
            if data_abertura and data_fechamento:
                # Verificar se foram abertas e fechadas no mesmo mês e ano
                if data_abertura.year == data_fechamento.year and data_abertura.month == data_fechamento.month:
                    ordens_mesmo_mes += 1
    ordens_mes_diferente = total_count - ordens_mesmo_mes
    percentual_mesmo_mes = (ordens_mesmo_mes / total_count * 100) if total_count > 0 else 0
    percentual_mes_diferente = (ordens_mes_diferente / total_count * 100) if total_count > 0 else 0
    
    # Obter lista de anos disponíveis (baseado nas ordens)
    anos_disponiveis = set()
    for ordem in todas_ordens:
        if ordem.dt_abertura_solicita:
            data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_parseada:
                anos_disponiveis.add(data_parseada.year)
    anos_disponiveis = sorted(anos_disponiveis, reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    # Nomes dos meses para exibição
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
        7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    context = {
        'page_title': 'Manutenção Corretiva - Análise',
        'active_page': 'manutencao_corretiva',
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'meses_nomes': meses_nomes,
        'anos_disponiveis': anos_disponiveis,
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_count': maquinas_count,
        # Estatísticas para os cards KPI
        'ordens_com_fechamento': ordens_com_fechamento,
        'ordens_sem_fechamento': ordens_sem_fechamento,
        'percentual_com_fechamento': round(percentual_com_fechamento, 1),
        'percentual_sem_fechamento': round(percentual_sem_fechamento, 1),
        'ordens_com_executor': ordens_com_executor,
        'ordens_sem_executor': ordens_sem_executor,
        'percentual_com_executor': round(percentual_com_executor, 1),
        'percentual_sem_executor': round(percentual_sem_executor, 1),
        'ordens_mesmo_mes': ordens_mesmo_mes,
        'ordens_mes_diferente': ordens_mes_diferente,
        'percentual_mesmo_mes': round(percentual_mesmo_mes, 1),
        'percentual_mes_diferente': round(percentual_mes_diferente, 1),
        # Dados para gráficos (JSON)
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'maquinas_labels': json.dumps(maquinas_labels),
        'maquinas_data': json.dumps(maquinas_data),
        'executores_labels': json.dumps(executores_labels),
        'executores_data': json.dumps(executores_data),
        'tipos_labels': json.dumps(tipos_labels),
        'tipos_data': json.dumps(tipos_data),
        # Dados para gráfico de local
        'local_labels': json.dumps(local_labels),
        'local_data': json.dumps(local_data),
        'local_percentages': json.dumps(local_percentages),
        # Dados para gráfico comparativo (abertas vs fechadas)
        'comparativo_labels': json.dumps(comparativo_labels),
        'comparativo_abertas_data': json.dumps(comparativo_abertas_data),
        'comparativo_fechadas_data': json.dumps(comparativo_fechadas_data),
        # Dados para tabelas
        'top_maquinas': top_maquinas,
        'top_executores': top_executores,
    }
    return render(request, 'ordens_de_servico/analise_corretiva_outros.html', context)


def analise_ordens_preventivas(request):
    """Página de análise de ordens de serviço preventivas"""
    from app.models import OrdemServicoPreventiva, Maquina
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')
    
    # Valores padrão: ano atual e todos os meses
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Função para fazer parse de dt_abertura_solicita
    def parse_dt_abertura_solicita(date_str):
        """Tenta fazer parse de dt_abertura_solicita em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para fazer parse de dt_encordmanu
    def parse_dt_encordmanu(date_str):
        """Tenta fazer parse de dt_encordmanu em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para filtrar ordens baseado em dt_abertura_solicita
    def filtrar_ordens_por_data(queryset, ano, meses=None):
        """Filtra ordens baseado em dt_abertura_solicita"""
        ordens_filtradas = []
        for ordem in queryset:
            if ordem.dt_abertura_solicita:
                data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
                if data_parseada:
                    if data_parseada.year == ano:
                        if meses is None or len(meses) == 0 or data_parseada.month in meses:
                            ordens_filtradas.append(ordem)
        return ordens_filtradas
    
    # Obter todas as ordens preventivas e filtrar
    todas_ordens = OrdemServicoPreventiva.objects.all()
    ordens_filtradas = filtrar_ordens_por_data(todas_ordens, ano_filtro, meses_filtro_int if meses_filtro_int else None)
    
    # Estatísticas básicas (filtradas)
    total_count = len(ordens_filtradas)
    
    # Contar ordens abertas e fechadas
    ordens_abertas = sum(1 for o in ordens_filtradas if not o.dt_encordmanu or o.dt_encordmanu == '')
    ordens_fechadas = total_count - ordens_abertas
    taxa_fechamento = (ordens_fechadas / total_count * 100) if total_count > 0 else 0
    
    # Contar setores, unidades e máquinas únicos (filtrados)
    setores_unicos = set()
    unidades_unicas = set()
    maquinas_unicas = set()
    for ordem in ordens_filtradas:
        if ordem.cd_setormanut:
            setores_unicos.add(ordem.cd_setormanut)
        if ordem.nome_unid:
            unidades_unicas.add(ordem.nome_unid)
        if ordem.cd_maquina:
            maquinas_unicas.add(ordem.cd_maquina)
    setores_count = len(setores_unicos)
    unidades_count = len(unidades_unicas)
    maquinas_count = len(maquinas_unicas)
    
    # Ordens por setor (top 10) - filtradas
    ordens_por_setor_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_setormanut:
            ordens_por_setor_dict[ordem.descr_setormanut] += 1
    ordens_por_setor = sorted(ordens_por_setor_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    setores_labels = [item[0][:30] for item in ordens_por_setor]
    setores_data = [item[1] for item in ordens_por_setor]
    
    # Ordens por unidade (top 10) - filtradas
    ordens_por_unidade_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nome_unid:
            ordens_por_unidade_dict[ordem.nome_unid] += 1
    ordens_por_unidade = sorted(ordens_por_unidade_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    unidades_labels = [item[0][:30] for item in ordens_por_unidade]
    unidades_data = [item[1] for item in ordens_por_unidade]
    
    # Ordens por mês do ano filtrado baseado em dt_abertura_solicita
    ordens_por_mes = defaultdict(int)
    ordens_abertas_por_mes = defaultdict(int)
    ordens_fechadas_por_mes = defaultdict(int)
    
    # Se meses específicos foram selecionados, mostrar apenas esses meses
    if meses_filtro_int and len(meses_filtro_int) > 0:
        meses_para_mostrar = meses_filtro_int
    else:
        meses_para_mostrar = list(range(1, 13))
    
    # Contar ordens abertas e fechadas por mês
    for ordem in ordens_filtradas:
        # Contar abertas (dt_abertura_solicita)
        if ordem.dt_abertura_solicita:
            data_abertura = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_abertura and data_abertura.year == ano_filtro:
                mes_ano = data_abertura.strftime('%Y-%m')
                ordens_por_mes[mes_ano] += 1
                ordens_abertas_por_mes[mes_ano] += 1
        
        # Contar fechadas (dt_encordmanu)
        if ordem.dt_encordmanu:
            data_fechamento = parse_dt_encordmanu(ordem.dt_encordmanu)
            if data_fechamento and data_fechamento.year == ano_filtro:
                mes_ano = data_fechamento.strftime('%Y-%m')
                ordens_fechadas_por_mes[mes_ano] += 1
    
    # Preencher todos os meses do ano (ou apenas o mês selecionado)
    for mes in meses_para_mostrar:
        mes_ano = f"{ano_filtro}-{mes:02d}"
        if mes_ano not in ordens_por_mes:
            ordens_por_mes[mes_ano] = 0
        if mes_ano not in ordens_abertas_por_mes:
            ordens_abertas_por_mes[mes_ano] = 0
        if mes_ano not in ordens_fechadas_por_mes:
            ordens_fechadas_por_mes[mes_ano] = 0
    
    # Ordenar por data
    meses_ordenados = sorted(ordens_por_mes.keys())
    
    # Formatar labels para português brasileiro
    meses_abrev = {
        'Jan': 'Jan', 'Feb': 'Fev', 'Mar': 'Mar', 'Apr': 'Abr', 'May': 'Mai', 'Jun': 'Jun',
        'Jul': 'Jul', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Nov': 'Nov', 'Dec': 'Dez'
    }
    meses_labels = []
    for m in meses_ordenados:
        dt = datetime.strptime(m, '%Y-%m')
        mes_abrev = meses_abrev.get(dt.strftime('%b'), dt.strftime('%b'))
        meses_labels.append(f"{mes_abrev}/{dt.strftime('%Y')}")
    meses_data = [ordens_por_mes[m] for m in meses_ordenados]
    
    # Dados para gráfico comparativo (abertas vs fechadas)
    comparativo_labels = meses_labels.copy()
    comparativo_abertas_data = [ordens_abertas_por_mes[m] for m in meses_ordenados]
    comparativo_fechadas_data = [ordens_fechadas_por_mes[m] for m in meses_ordenados]
    
    # Top 10 máquinas com mais ordens preventivas - filtradas
    maquinas_dict = defaultdict(int)
    maquinas_desc = {}
    for ordem in ordens_filtradas:
        if ordem.cd_maquina and ordem.descr_maquina:
            maquinas_dict[ordem.cd_maquina] += 1
            maquinas_desc[ordem.cd_maquina] = ordem.descr_maquina
    top_maquinas_list = sorted(maquinas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_maquinas = [{'cd_maquina': item[0], 'descr_maquina': maquinas_desc.get(item[0], ''), 'total': item[1]} for item in top_maquinas_list]
    maquinas_labels = [f"{item['cd_maquina']}" for item in top_maquinas]
    maquinas_data = [item['total'] for item in top_maquinas]
    
    # Top 10 executores - filtradas
    executores_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        # Verificar executor na ordem principal
        if ordem.nm_func_exec:
            executores_dict[ordem.nm_func_exec] += 1
        # Verificar também nas fichas
        for ficha in ordem.fichas.all():
            if ficha.nm_func_exec_os:
                executores_dict[ficha.nm_func_exec_os] += 1
    top_executores_list = sorted(executores_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_executores = [{'nm_func_exec_os': item[0], 'total': item[1]} for item in top_executores_list]
    funcionarios_labels = [item['nm_func_exec_os'][:25] for item in top_executores]
    funcionarios_data = [item['total'] for item in top_executores]
    
    # Anos disponíveis para filtro
    anos_disponiveis = sorted(set(
        parse_dt_abertura_solicita(o.dt_abertura_solicita).year 
        for o in todas_ordens 
        if o.dt_abertura_solicita and parse_dt_abertura_solicita(o.dt_abertura_solicita)
    ), reverse=True)
    
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    context = {
        'page_title': 'Análise de Ordens Preventivas',
        'active_page': 'analise_ordens_preventivas',
        'total_count': total_count,
        'ordens_abertas': ordens_abertas,
        'ordens_fechadas': ordens_fechadas,
        'taxa_fechamento': round(taxa_fechamento, 1),
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_count': maquinas_count,
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'comparativo_labels': json.dumps(comparativo_labels),
        'comparativo_abertas_data': json.dumps(comparativo_abertas_data),
        'comparativo_fechadas_data': json.dumps(comparativo_fechadas_data),
        'maquinas_labels': json.dumps(maquinas_labels),
        'maquinas_data': json.dumps(maquinas_data),
        'top_maquinas': top_maquinas,
        'funcionarios_labels': json.dumps(funcionarios_labels),
        'funcionarios_data': json.dumps(funcionarios_data),
        'top_executores': top_executores,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
    }
    return render(request, 'ordens_de_servico/analise_ordens_preventivas.html', context)


def analise_lubrificacao(request):
    """Página de análise de lubrificações com gráficos e estatísticas"""
    from app.models import OrdemServicoCorretiva, Maquina
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e todos os meses (None)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Função para fazer parse de dt_abertura_solicita
    def parse_dt_abertura_solicita(date_str):
        """Tenta fazer parse de dt_abertura_solicita em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para fazer parse de dt_encordmanu
    def parse_dt_encordmanu(date_str):
        """Tenta fazer parse de dt_encordmanu em vários formatos diferentes."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        if not date_str:
            return None
        
        # Remover hora se existir
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        
        # Tentar diferentes formatos de data
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        
        # Parse manual para formato brasileiro
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        
        return None
    
    # Função para filtrar ordens baseado em dt_abertura_solicita
    def filtrar_ordens_por_data(queryset, ano, meses=None):
        """Filtra ordens baseado em dt_abertura_solicita"""
        ordens_filtradas = []
        for ordem in queryset:
            if ordem.dt_abertura_solicita:
                data_parseada = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
                if data_parseada:
                    if data_parseada.year == ano:
                        if meses is None or len(meses) == 0 or data_parseada.month in meses:
                            ordens_filtradas.append(ordem)
        return ordens_filtradas
    
    # Obter todas as ordens de lubrificação e filtrar
    todas_ordens = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    )
    ordens_filtradas = filtrar_ordens_por_data(todas_ordens, ano_filtro, meses_filtro_int if meses_filtro_int else None)
    
    # Estatísticas básicas (filtradas)
    total_count = len(ordens_filtradas)
    
    # Contar ordens abertas e fechadas
    ordens_abertas = sum(1 for o in ordens_filtradas if not o.dt_encordmanu or o.dt_encordmanu == '')
    ordens_fechadas = total_count - ordens_abertas
    taxa_fechamento = (ordens_fechadas / total_count * 100) if total_count > 0 else 0
    
    # Contar setores, unidades e máquinas únicos (filtrados)
    setores_unicos = set()
    unidades_unicas = set()
    maquinas_unicas = set()
    for ordem in ordens_filtradas:
        if ordem.cd_setormanut:
            setores_unicos.add(ordem.cd_setormanut)
        if ordem.nome_unid:
            unidades_unicas.add(ordem.nome_unid)
        if ordem.cd_maquina:
            maquinas_unicas.add(ordem.cd_maquina)
    setores_count = len(setores_unicos)
    unidades_count = len(unidades_unicas)
    maquinas_count = len(maquinas_unicas)
    
    # Ordens por setor (top 10) - filtradas
    ordens_por_setor_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.descr_setormanut:
            ordens_por_setor_dict[ordem.descr_setormanut] += 1
    ordens_por_setor = sorted(ordens_por_setor_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    setores_labels = [item[0][:30] for item in ordens_por_setor]
    setores_data = [item[1] for item in ordens_por_setor]
    
    # Ordens por unidade (top 10) - filtradas
    ordens_por_unidade_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        if ordem.nome_unid:
            ordens_por_unidade_dict[ordem.nome_unid] += 1
    ordens_por_unidade = sorted(ordens_por_unidade_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    unidades_labels = [item[0][:30] for item in ordens_por_unidade]
    unidades_data = [item[1] for item in ordens_por_unidade]
    
    # Ordens por mês do ano filtrado baseado em dt_abertura_solicita
    ordens_por_mes = defaultdict(int)
    ordens_abertas_por_mes = defaultdict(int)
    ordens_fechadas_por_mes = defaultdict(int)
    
    # Se meses específicos foram selecionados, mostrar apenas esses meses
    if meses_filtro_int and len(meses_filtro_int) > 0:
        meses_para_mostrar = meses_filtro_int
    else:
        meses_para_mostrar = list(range(1, 13))
    
    # Contar ordens abertas e fechadas por mês
    for ordem in ordens_filtradas:
        # Contar abertas (dt_abertura_solicita)
        if ordem.dt_abertura_solicita:
            data_abertura = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_abertura and data_abertura.year == ano_filtro:
                mes_ano = data_abertura.strftime('%Y-%m')
                ordens_por_mes[mes_ano] += 1
                ordens_abertas_por_mes[mes_ano] += 1
        
        # Contar fechadas (dt_encordmanu)
        if ordem.dt_encordmanu:
            data_fechamento = parse_dt_encordmanu(ordem.dt_encordmanu)
            if data_fechamento and data_fechamento.year == ano_filtro:
                mes_ano = data_fechamento.strftime('%Y-%m')
                ordens_fechadas_por_mes[mes_ano] += 1
    
    # Preencher todos os meses do ano (ou apenas o mês selecionado)
    for mes in meses_para_mostrar:
        mes_ano = f"{ano_filtro}-{mes:02d}"
        if mes_ano not in ordens_por_mes:
            ordens_por_mes[mes_ano] = 0
        if mes_ano not in ordens_abertas_por_mes:
            ordens_abertas_por_mes[mes_ano] = 0
        if mes_ano not in ordens_fechadas_por_mes:
            ordens_fechadas_por_mes[mes_ano] = 0
    
    # Ordenar por data
    meses_ordenados = sorted(ordens_por_mes.keys())
    
    # Formatar labels para português brasileiro
    meses_abrev = {
        'Jan': 'Jan', 'Feb': 'Fev', 'Mar': 'Mar', 'Apr': 'Abr', 'May': 'Mai', 'Jun': 'Jun',
        'Jul': 'Jul', 'Aug': 'Ago', 'Sep': 'Set', 'Oct': 'Out', 'Nov': 'Nov', 'Dec': 'Dez'
    }
    meses_labels = []
    for m in meses_ordenados:
        dt = datetime.strptime(m, '%Y-%m')
        mes_abrev = meses_abrev.get(dt.strftime('%b'), dt.strftime('%b'))
        meses_labels.append(f"{mes_abrev}/{dt.strftime('%Y')}")
    meses_data = [ordens_por_mes[m] for m in meses_ordenados]
    
    # Dados para gráfico comparativo (abertas vs fechadas)
    comparativo_labels = meses_labels.copy()
    comparativo_abertas_data = [ordens_abertas_por_mes[m] for m in meses_ordenados]
    comparativo_fechadas_data = [ordens_fechadas_por_mes[m] for m in meses_ordenados]
    
    # Top 10 máquinas com mais ordens de lubrificação - filtradas
    maquinas_dict = defaultdict(int)
    maquinas_desc = {}
    for ordem in ordens_filtradas:
        if ordem.cd_maquina and ordem.descr_maquina:
            maquinas_dict[ordem.cd_maquina] += 1
            maquinas_desc[ordem.cd_maquina] = ordem.descr_maquina
    top_maquinas_list = sorted(maquinas_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_maquinas = [{'cd_maquina': item[0], 'descr_maquina': maquinas_desc.get(item[0], ''), 'total': item[1]} for item in top_maquinas_list]
    maquinas_labels = [f"{item['cd_maquina']}" for item in top_maquinas]
    maquinas_data = [item['total'] for item in top_maquinas]
    
    # Top 10 executores - filtradas
    executores_dict = defaultdict(int)
    for ordem in ordens_filtradas:
        # Verificar executor na ordem principal
        if ordem.nm_func_exec:
            executores_dict[ordem.nm_func_exec] += 1
        # Verificar também nas fichas
        for ficha in ordem.fichas.all():
            if ficha.nm_func_exec_os:
                executores_dict[ficha.nm_func_exec_os] += 1
    top_executores_list = sorted(executores_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    top_executores = [{'nm_func_exec_os': item[0], 'total': item[1]} for item in top_executores_list]
    funcionarios_labels = [item['nm_func_exec_os'][:25] for item in top_executores]
    funcionarios_data = [item['total'] for item in top_executores]
    
    # Análise de frequência de lubrificação (dias entre lubrificações)
    frequencia_dict = defaultdict(list)
    for ordem in ordens_filtradas:
        if ordem.cd_maquina and ordem.dt_abertura_solicita:
            data_abertura = parse_dt_abertura_solicita(ordem.dt_abertura_solicita)
            if data_abertura:
                frequencia_dict[ordem.cd_maquina].append(data_abertura)
    
    # Calcular média de dias entre lubrificações por máquina
    frequencia_media = {}
    for maquina, datas in frequencia_dict.items():
        if len(datas) > 1:
            datas_ordenadas = sorted(datas)
            diferencas = []
            for i in range(1, len(datas_ordenadas)):
                diff = (datas_ordenadas[i] - datas_ordenadas[i-1]).days
                if diff > 0:
                    diferencas.append(diff)
            if diferencas:
                frequencia_media[maquina] = sum(diferencas) / len(diferencas)
    
    # Top máquinas por frequência (menor intervalo = mais frequente)
    top_frequencia = sorted(frequencia_media.items(), key=lambda x: x[1])[:5]
    
    # Anos disponíveis para filtro
    anos_disponiveis = sorted(set(
        parse_dt_abertura_solicita(o.dt_abertura_solicita).year 
        for o in todas_ordens 
        if o.dt_abertura_solicita and parse_dt_abertura_solicita(o.dt_abertura_solicita)
    ), reverse=True)
    
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    context = {
        'page_title': 'Análise de Lubrificação',
        'active_page': 'analise_lubrificacao',
        'total_count': total_count,
        'ordens_abertas': ordens_abertas,
        'ordens_fechadas': ordens_fechadas,
        'taxa_fechamento': round(taxa_fechamento, 1),
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_count': maquinas_count,
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'unidades_labels': json.dumps(unidades_labels),
        'unidades_data': json.dumps(unidades_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'comparativo_labels': json.dumps(comparativo_labels),
        'comparativo_abertas_data': json.dumps(comparativo_abertas_data),
        'comparativo_fechadas_data': json.dumps(comparativo_fechadas_data),
        'maquinas_labels': json.dumps(maquinas_labels),
        'maquinas_data': json.dumps(maquinas_data),
        'top_maquinas': top_maquinas,
        'funcionarios_labels': json.dumps(funcionarios_labels),
        'funcionarios_data': json.dumps(funcionarios_data),
        'top_executores': top_executores,
        'top_frequencia': top_frequencia,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
    }
    return render(request, 'lubrificacao/analise_lubrificacao.html', context)


def consultar_ordens_lubrificacao(request):
    """Consultar/listar ordens de lubrificação cadastradas com filtros avançados"""
    from app.models import OrdemServicoCorretiva
    from datetime import datetime
    
    # Buscar apenas ordens de lubrificação
    ordens_list = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    )
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos numéricos, tentar converter e fazer busca exata
        try:
            search_num = int(float(search_query))
            search_conditions |= Q(cd_ordemserv=search_num)
            search_conditions |= Q(cd_maquina=search_num)
        except (ValueError, TypeError):
            pass
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(descr_maquina__icontains=search_query) |
            Q(cd_setormanut__icontains=search_query) |
            Q(descr_setormanut__icontains=search_query) |
            Q(nm_func_solic_os__icontains=search_query) |
            Q(nm_func_exec__icontains=search_query) |
            Q(descr_queixa__icontains=search_query) |
            Q(exec_tarefas__icontains=search_query)
        )
        
        ordens_list = ordens_list.filter(search_conditions)
    
    # Filtros específicos
    # Filtro por Setor de Manutenção
    filtro_setor = request.GET.get('filtro_setor', '')
    if filtro_setor:
        ordens_list = ordens_list.filter(descr_setormanut__icontains=filtro_setor)
    
    # Filtro por Unidade
    filtro_unidade = request.GET.get('filtro_unidade', '')
    if filtro_unidade:
        ordens_list = ordens_list.filter(nome_unid__icontains=filtro_unidade)
    
    # Filtro por Tipo de Ordem de Serviço
    filtro_tipo_os = request.GET.get('filtro_tipo_os', '')
    if filtro_tipo_os:
        ordens_list = ordens_list.filter(descr_tpordservtv__icontains=filtro_tipo_os)
    
    # Filtro por Situação da Ordem
    filtro_situacao = request.GET.get('filtro_situacao', '')
    if filtro_situacao:
        ordens_list = ordens_list.filter(descr_sitordsetv__icontains=filtro_situacao)
    
    # Filtro por Funcionário Solicitante
    filtro_solicitante = request.GET.get('filtro_solicitante', '')
    if filtro_solicitante:
        ordens_list = ordens_list.filter(nm_func_solic_os__icontains=filtro_solicitante)
    
    # Filtro por Funcionário Executor
    filtro_executor = request.GET.get('filtro_executor', '')
    if filtro_executor:
        ordens_list = ordens_list.filter(
            Q(nm_func_exec__icontains=filtro_executor)
        )
    
    # Filtro por Código da Máquina
    filtro_maquina = request.GET.get('filtro_maquina', '')
    if filtro_maquina:
        ordens_list = ordens_list.filter(cd_maquina__icontains=filtro_maquina)
    
    # Filtro por Data de Entrada (período)
    data_entrada_inicio = request.GET.get('data_entrada_inicio', '')
    data_entrada_fim = request.GET.get('data_entrada_fim', '')
    if data_entrada_inicio:
        try:
            data_inicio = datetime.strptime(data_entrada_inicio, '%Y-%m-%d')
            ordens_list = ordens_list.filter(created_at__gte=data_inicio)
        except ValueError:
            pass
    if data_entrada_fim:
        try:
            data_fim = datetime.strptime(data_entrada_fim, '%Y-%m-%d')
            # Adicionar 1 dia para incluir o dia final
            from datetime import timedelta
            data_fim = data_fim + timedelta(days=1)
            ordens_list = ordens_list.filter(created_at__lte=data_fim)
        except ValueError:
            pass
    
    # Filtro por Status da Ordem (Abertas/Fechadas)
    filtro_ordens_abertas = request.GET.get('filtro_ordens_abertas', '')
    filtro_ordens_fechadas = request.GET.get('filtro_ordens_fechadas', '')
    
    # Converter para boolean (se existe e não é vazio, é True)
    filtro_ordens_abertas = filtro_ordens_abertas == '1'
    filtro_ordens_fechadas = filtro_ordens_fechadas == '1'
    
    # Aplicar filtros baseado nos checkboxes marcados
    if filtro_ordens_abertas and filtro_ordens_fechadas:
        # Ambos marcados: mostrar todas (não aplicar filtro)
        pass
    elif filtro_ordens_abertas and not filtro_ordens_fechadas:
        # Apenas "Ordens Abertas" marcado: dt_encordmanu está vazio ou nulo
        ordens_list = ordens_list.filter(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    elif filtro_ordens_fechadas and not filtro_ordens_abertas:
        # Apenas "Ordens Fechadas" marcado: dt_encordmanu tem valor (não é nulo nem vazio)
        ordens_list = ordens_list.exclude(
            Q(dt_encordmanu__isnull=True) | Q(dt_encordmanu='')
        )
    # Se nenhum está marcado, mostra todas (não aplicar filtro)
    
    # Ordenar por código da ordem de serviço (mais recente primeiro)
    ordens_list = ordens_list.order_by('-cd_ordemserv')
    
    # Paginação
    paginator = Paginator(ordens_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    ordens = paginator.get_page(page_number)
    
    # Estatísticas (apenas para ordens de lubrificação)
    total_count = OrdemServicoCorretiva.objects.filter(descr_tpordservtv__icontains='LUBRIFICAÇÃO').count()
    setores_count = OrdemServicoCorretiva.objects.filter(descr_tpordservtv__icontains='LUBRIFICAÇÃO').exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    unidades_count = OrdemServicoCorretiva.objects.filter(descr_tpordservtv__icontains='LUBRIFICAÇÃO').exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros (apenas para ordens de lubrificação)
    setores_unicos = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    ).exclude(
        descr_setormanut__isnull=True
    ).exclude(
        descr_setormanut=''
    ).values_list('descr_setormanut', flat=True).distinct().order_by('descr_setormanut')
    
    unidades_unicas = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    ).exclude(
        nome_unid__isnull=True
    ).exclude(
        nome_unid=''
    ).values_list('nome_unid', flat=True).distinct().order_by('nome_unid')
    
    tipos_os_unicos = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    ).exclude(
        descr_tpordservtv__isnull=True
    ).exclude(
        descr_tpordservtv=''
    ).values_list('descr_tpordservtv', flat=True).distinct().order_by('descr_tpordservtv')
    
    situacoes_unicas = OrdemServicoCorretiva.objects.filter(
        descr_tpordservtv__icontains='LUBRIFICAÇÃO'
    ).exclude(
        descr_sitordsetv__isnull=True
    ).exclude(
        descr_sitordsetv=''
    ).values_list('descr_sitordsetv', flat=True).distinct().order_by('descr_sitordsetv')
    
    context = {
        'page_title': 'Consultar Ordens de Lubrificação',
        'active_page': 'consultar_ordens_lubrificacao',
        'ordens': ordens,
        'total_count': total_count,
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        # Valores para dropdowns
        'setores_unicos': setores_unicos,
        'unidades_unicas': unidades_unicas,
        'tipos_os_unicos': tipos_os_unicos,
        'situacoes_unicas': situacoes_unicas,
        # Valores dos filtros ativos
        'filtro_setor': filtro_setor,
        'filtro_unidade': filtro_unidade,
        'filtro_tipo_os': filtro_tipo_os,
        'filtro_situacao': filtro_situacao,
        'filtro_solicitante': filtro_solicitante,
        'filtro_executor': filtro_executor,
        'filtro_maquina': filtro_maquina,
        'data_entrada_inicio': data_entrada_inicio,
        'data_entrada_fim': data_entrada_fim,
        'filtro_ordens_abertas': '1' if filtro_ordens_abertas else '',
        'filtro_ordens_fechadas': '1' if filtro_ordens_fechadas else '',
    }
    return render(request, 'consultar/consultar_ordens_lubrificacao.html', context)


def analise_corretiva_outros_com_parada(request):
    """Análise de Ordens Corretivas com informações de parada"""
    from app.models import OrdemServicoCorretiva, Maquina, CentroAtividade
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from collections import defaultdict
    import json
    import re
    
    # Filtrar apenas ordens com dt_iniparmanu preenchido (não nulo e não vazio)
    ordens_com_parada_qs = OrdemServicoCorretiva.objects.filter(
        Q(dt_iniparmanu__isnull=False) & ~Q(dt_iniparmanu='')
    )
    
    # Validar que dt_iniparmanu contém uma data válida
    # Tentar parsear as datas para garantir que são válidas
    def is_valid_date(date_str):
        """Verifica se a string contém uma data válida"""
        if not date_str or not isinstance(date_str, str):
            return False
        
        date_str = date_str.strip()
        if not date_str:
            return False
        
        # Tentar diferentes formatos de data comuns
        formatos_data = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S',
            '%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%d-%m-%y',
            '%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M'
        ]
        for formato in formatos_data:
            try:
                datetime.strptime(date_str, formato)
                return True
            except (ValueError, AttributeError):
                continue
        
        # Se não encontrou formato padrão, verificar se contém padrão de data
        if re.search(r'\d{1,2}[\s\-/]\d{1,2}[\s\-/]\d{2,4}', date_str):
            return True
        
        return False
    
    # Filtrar apenas ordens com datas válidas usando values_list para eficiência
    ordens_validas_ids = [
        ordem_id for ordem_id, dt_iniparmanu in 
        ordens_com_parada_qs.values_list('id', 'dt_iniparmanu')
        if is_valid_date(dt_iniparmanu)
    ]
    
    # Filtrar apenas ordens com datas válidas
    ordens_com_parada_qs = OrdemServicoCorretiva.objects.filter(id__in=ordens_validas_ids)
    
    total_com_parada = ordens_com_parada_qs.count()
    total_geral = OrdemServicoCorretiva.objects.count()
    
    # Ordens com início e fim de parada (ambos preenchidos e válidos)
    com_inicio_fim = ordens_com_parada_qs.filter(
        Q(dt_fimparmanu__isnull=False) & ~Q(dt_fimparmanu='')
    ).count()
    
    # Ordens apenas com início de parada (sem fim)
    apenas_inicio = ordens_com_parada_qs.filter(
        Q(dt_fimparmanu__isnull=True) | Q(dt_fimparmanu='')
    ).count()
    
    # Percentual de ordens com parada
    percentual_com_parada = (total_com_parada / total_geral * 100) if total_geral > 0 else 0
    
    # Estatísticas básicas
    setores_count = ordens_com_parada_qs.exclude(cd_setormanut__isnull=True).exclude(cd_setormanut='').values('cd_setormanut').distinct().count()
    unidades_count = ordens_com_parada_qs.exclude(nome_unid__isnull=True).exclude(nome_unid='').values('nome_unid').distinct().count()
    maquinas_count = ordens_com_parada_qs.exclude(cd_maquina__isnull=True).values('cd_maquina').distinct().count()
    
    # Ordens por setor (top 10)
    ordens_por_setor = ordens_com_parada_qs.exclude(
        descr_setormanut__isnull=True
    ).exclude(
        descr_setormanut=''
    ).values('descr_setormanut').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    setores_labels = [item['descr_setormanut'][:30] for item in ordens_por_setor]
    setores_data = [item['total'] for item in ordens_por_setor]
    
    # Ordens por mês (últimos 12 meses)
    ordens_por_mes = defaultdict(int)
    ordens = ordens_com_parada_qs.all().order_by('created_at')
    for ordem in ordens:
        if ordem.created_at:
            mes_ano = ordem.created_at.strftime('%Y-%m')
            ordens_por_mes[mes_ano] += 1
    
    # Ordenar por data e pegar últimos 12 meses
    meses_ordenados = sorted(ordens_por_mes.keys())[-12:]
    meses_labels = [datetime.strptime(m, '%Y-%m').strftime('%b/%Y') for m in meses_ordenados]
    meses_data = [ordens_por_mes[m] for m in meses_ordenados]
    
    # Top 10 máquinas com mais ordens com parada
    top_maquinas = ordens_com_parada_qs.exclude(
        descr_maquina__isnull=True
    ).exclude(
        descr_maquina=''
    ).values('cd_maquina', 'descr_maquina').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    maquinas_labels = [f"{item['cd_maquina']} - {item['descr_maquina'][:40]}" for item in top_maquinas]
    maquinas_data = [item['total'] for item in top_maquinas]
    
    # Top 10 executores
    top_executores = ordens_com_parada_qs.exclude(
        nm_func_exec__isnull=True
    ).exclude(
        nm_func_exec=''
    ).values('nm_func_exec').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    executores_labels = [item['nm_func_exec'][:30] for item in top_executores]
    executores_data = [item['total'] for item in top_executores]
    
    # Ordens por dia do mês atual usando dt_entrada
    hoje = datetime.now()
    primeiro_dia_mes = hoje.replace(day=1)
    ultimo_dia_mes = (primeiro_dia_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Inicializar contador para todos os dias do mês atual
    ordens_por_dia = defaultdict(int)
    for dia in range(1, ultimo_dia_mes.day + 1):
        ordens_por_dia[dia] = 0
    
    # Função para parsear dt_entrada e extrair a data
    def parse_dt_entrada(dt_entrada_str):
        """Tenta parsear dt_entrada e retornar o dia do mês se for do mês atual"""
        if not dt_entrada_str:
            return None
        
        dt_entrada_str = dt_entrada_str.strip()
        formatos_data = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S',
            '%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%d-%m-%y',
            '%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M'
        ]
        
        for formato in formatos_data:
            try:
                data_parseada = datetime.strptime(dt_entrada_str, formato)
                # Verificar se é do mês atual
                if data_parseada.year == hoje.year and data_parseada.month == hoje.month:
                    return data_parseada.day
            except (ValueError, AttributeError):
                continue
        
        # Tentar extrair dia usando regex se os formatos padrão não funcionarem
        match = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', dt_entrada_str)
        if match:
            dia_str, mes_str, ano_str = match.groups()
            try:
                dia = int(dia_str)
                mes = int(mes_str)
                ano = int(ano_str)
                if ano < 100:
                    ano += 2000 if ano < 50 else 1900
                
                if ano == hoje.year and mes == hoje.month:
                    return dia
            except (ValueError, TypeError):
                pass
        
        return None
    
    # Contar ordens por dia do mês atual
    for ordem in ordens_com_parada_qs:
        if ordem.dt_entrada:
            dia = parse_dt_entrada(ordem.dt_entrada)
            if dia and 1 <= dia <= ultimo_dia_mes.day:
                ordens_por_dia[dia] += 1
    
    # Preparar dados para o gráfico
    daily_labels = [f"{dia:02d}/{hoje.month:02d}" for dia in sorted(ordens_por_dia.keys())]
    daily_data = [ordens_por_dia[dia] for dia in sorted(ordens_por_dia.keys())]
    
    # Classificação por FRIGORÍFICO vs INDÚSTRIA baseado em cd_setormanut
    # A relação é: cd_setormanut em OrdemServicoCorretiva corresponde a sigla em CentroAtividade
    # Buscar Centros de Atividade classificados e obter suas siglas
    centros_frigorifico = CentroAtividade.objects.filter(
        local__icontains='FRIGORÍFICO'
    ).exclude(sigla__isnull=True).exclude(sigla='')
    
    centros_industria = CentroAtividade.objects.filter(
        local__icontains='INDÚSTRIA'
    ).exclude(sigla__isnull=True).exclude(sigla='')
    
    # Extrair siglas dos centros FRIGORÍFICO e INDÚSTRIA
    siglas_frigorifico = [centro.sigla.strip().upper() for centro in centros_frigorifico if centro.sigla]
    siglas_industria = [centro.sigla.strip().upper() for centro in centros_industria if centro.sigla]
    
    # Remover duplicatas
    siglas_frigorifico = list(set(siglas_frigorifico))
    siglas_industria = list(set(siglas_industria))
    
    # Classificar ordens por tipo baseado em cd_setormanut (que corresponde a sigla)
    # Usar Q objects para fazer match case-insensitive
    from django.db.models import Q
    
    q_frigorifico = Q()
    for sigla in siglas_frigorifico:
        q_frigorifico |= Q(cd_setormanut__iexact=sigla)
    
    q_industria = Q()
    for sigla in siglas_industria:
        q_industria |= Q(cd_setormanut__iexact=sigla)
    
    ordens_frigorifico = ordens_com_parada_qs.filter(q_frigorifico)
    ordens_industria = ordens_com_parada_qs.filter(q_industria)
    ordens_outros = ordens_com_parada_qs.exclude(q_frigorifico | q_industria).exclude(
        cd_setormanut__isnull=True
    ).exclude(cd_setormanut='')
    
    # Estatísticas por classificação
    total_frigorifico = ordens_frigorifico.count()
    total_industria = ordens_industria.count()
    total_outros = ordens_outros.count()
    
    # Percentuais
    percentual_frigorifico = (total_frigorifico / total_com_parada * 100) if total_com_parada > 0 else 0
    percentual_industria = (total_industria / total_com_parada * 100) if total_com_parada > 0 else 0
    percentual_outros = (total_outros / total_com_parada * 100) if total_com_parada > 0 else 0
    
    # Top máquinas por classificação
    top_maquinas_frigorifico = ordens_frigorifico.exclude(
        descr_maquina__isnull=True
    ).exclude(
        descr_maquina=''
    ).values('cd_maquina', 'descr_maquina').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    top_maquinas_industria = ordens_industria.exclude(
        descr_maquina__isnull=True
    ).exclude(
        descr_maquina=''
    ).values('cd_maquina', 'descr_maquina').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Ordens por dia separadas por classificação
    ordens_por_dia_frigorifico = defaultdict(int)
    ordens_por_dia_industria = defaultdict(int)
    for dia in range(1, ultimo_dia_mes.day + 1):
        ordens_por_dia_frigorifico[dia] = 0
        ordens_por_dia_industria[dia] = 0
    
    for ordem in ordens_frigorifico:
        if ordem.dt_entrada:
            dia = parse_dt_entrada(ordem.dt_entrada)
            if dia and 1 <= dia <= ultimo_dia_mes.day:
                ordens_por_dia_frigorifico[dia] += 1
    
    for ordem in ordens_industria:
        if ordem.dt_entrada:
            dia = parse_dt_entrada(ordem.dt_entrada)
            if dia and 1 <= dia <= ultimo_dia_mes.day:
                ordens_por_dia_industria[dia] += 1
    
    daily_labels_classificacao = [f"{dia:02d}/{hoje.month:02d}" for dia in sorted(ordens_por_dia.keys())]
    daily_data_frigorifico = [ordens_por_dia_frigorifico[dia] for dia in sorted(ordens_por_dia.keys())]
    daily_data_industria = [ordens_por_dia_industria[dia] for dia in sorted(ordens_por_dia.keys())]
    
    # Ordens por mês separadas por classificação (últimos 12 meses)
    ordens_por_mes_frigorifico = defaultdict(int)
    ordens_por_mes_industria = defaultdict(int)
    
    for ordem in ordens_frigorifico:
        if ordem.created_at:
            mes_ano = ordem.created_at.strftime('%Y-%m')
            ordens_por_mes_frigorifico[mes_ano] += 1
    
    for ordem in ordens_industria:
        if ordem.created_at:
            mes_ano = ordem.created_at.strftime('%Y-%m')
            ordens_por_mes_industria[mes_ano] += 1
    
    # Garantir que todos os meses tenham dados
    todos_meses = set(ordens_por_mes.keys()) | set(ordens_por_mes_frigorifico.keys()) | set(ordens_por_mes_industria.keys())
    meses_ordenados_classificacao = sorted(todos_meses)[-12:]
    meses_labels_classificacao = [datetime.strptime(m, '%Y-%m').strftime('%b/%Y') for m in meses_ordenados_classificacao]
    meses_data_frigorifico = [ordens_por_mes_frigorifico.get(m, 0) for m in meses_ordenados_classificacao]
    meses_data_industria = [ordens_por_mes_industria.get(m, 0) for m in meses_ordenados_classificacao]
    
    context = {
        'page_title': 'Análise Corretiva com Parada',
        'active_page': 'analise_corretiva_parada',
        'total_com_parada': total_com_parada,
        'com_inicio_fim': com_inicio_fim,
        'apenas_inicio': apenas_inicio,
        'percentual_com_parada': round(percentual_com_parada, 1),
        'setores_count': setores_count,
        'unidades_count': unidades_count,
        'maquinas_count': maquinas_count,
        # Dados para gráficos (JSON)
        'setores_labels': json.dumps(setores_labels),
        'setores_data': json.dumps(setores_data),
        'meses_labels': json.dumps(meses_labels),
        'meses_data': json.dumps(meses_data),
        'maquinas_labels': json.dumps(maquinas_labels),
        'maquinas_data': json.dumps(maquinas_data),
        'executores_labels': json.dumps(executores_labels),
        'executores_data': json.dumps(executores_data),
        'daily_labels': json.dumps(daily_labels),
        'daily_data': json.dumps(daily_data),
        # Dados para tabelas
        'top_maquinas': top_maquinas,
        'top_executores': top_executores,
        # Dados de classificação FRIGORÍFICO vs INDÚSTRIA
        'total_frigorifico': total_frigorifico,
        'total_industria': total_industria,
        'total_outros': total_outros,
        'percentual_frigorifico': round(percentual_frigorifico, 1),
        'percentual_industria': round(percentual_industria, 1),
        'percentual_outros': round(percentual_outros, 1),
        'top_maquinas_frigorifico': top_maquinas_frigorifico,
        'top_maquinas_industria': top_maquinas_industria,
        'daily_labels_classificacao': json.dumps(daily_labels_classificacao),
        'daily_data_frigorifico': json.dumps(daily_data_frigorifico),
        'daily_data_industria': json.dumps(daily_data_industria),
        'meses_labels_classificacao': json.dumps(meses_labels_classificacao),
        'meses_data_frigorifico': json.dumps(meses_data_frigorifico),
        'meses_data_industria': json.dumps(meses_data_industria),
    }
    return render(request, 'ordens_de_servico/analise_corretiva_outros_com_parada.html', context)


def consultar_manutencao_terceiros(request):
    """Consultar/listar manutenções de terceiros cadastradas com filtros avançados"""
    from app.models import ManutencaoTerceiro
    from datetime import datetime
    
    # Buscar todas as manutenções de terceiros
    manutencoes_list = ManutencaoTerceiro.objects.all()
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Criar lista de condições Q
        search_conditions = Q()
        
        # Para campos de texto, usar icontains
        search_conditions |= (
            Q(titulo__icontains=search_query) |
            Q(os__icontains=search_query) |
            Q(empresa__icontains=search_query) |
            Q(pedidodecompra__icontains=search_query) |
            Q(requisicaodecompra__icontains=search_query) |
            Q(descricao__icontains=search_query) |
            Q(maquina__descr_maquina__icontains=search_query) |
            Q(manutentor__Nome__icontains=search_query)
        )
        
        manutencoes_list = manutencoes_list.filter(search_conditions)
    
    # Filtros específicos
    # Filtro por Empresa
    filtro_empresa = request.GET.get('filtro_empresa', '')
    if filtro_empresa:
        manutencoes_list = manutencoes_list.filter(empresa__icontains=filtro_empresa)
    
    # Filtro por Tipo
    filtro_tipo = request.GET.get('filtro_tipo', '')
    if filtro_tipo:
        manutencoes_list = manutencoes_list.filter(tipo=filtro_tipo)
    
    # Filtro por Máquina
    filtro_maquina = request.GET.get('filtro_maquina', '')
    if filtro_maquina:
        try:
            maquina_id = int(filtro_maquina)
            manutencoes_list = manutencoes_list.filter(maquina_id=maquina_id)
        except ValueError:
            manutencoes_list = manutencoes_list.filter(maquina__descr_maquina__icontains=filtro_maquina)
    
    # Filtro por Manutentor
    filtro_manutentor = request.GET.get('filtro_manutentor', '')
    if filtro_manutentor:
        try:
            manutentor_id = filtro_manutentor
            manutencoes_list = manutencoes_list.filter(manutentor__Matricula=manutentor_id)
        except ValueError:
            manutencoes_list = manutencoes_list.filter(manutentor__Nome__icontains=filtro_manutentor)
    
    # Filtro por Data (período)
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    if data_inicio:
        try:
            data_ini = datetime.strptime(data_inicio, '%Y-%m-%d')
            manutencoes_list = manutencoes_list.filter(data__gte=data_ini)
        except ValueError:
            pass
    if data_fim:
        try:
            data_f = datetime.strptime(data_fim, '%Y-%m-%d')
            from datetime import timedelta
            data_f = data_f + timedelta(days=1)
            manutencoes_list = manutencoes_list.filter(data__lte=data_f)
        except ValueError:
            pass
    
    # Ordenar por data (mais recente primeiro)
    manutencoes_list = manutencoes_list.order_by('-data', '-created_at')
    
    # Paginação
    paginator = Paginator(manutencoes_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    manutencoes = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = ManutencaoTerceiro.objects.count()
    empresas_count = ManutencaoTerceiro.objects.exclude(empresa__isnull=True).exclude(empresa='').values('empresa').distinct().count()
    tipos_count = ManutencaoTerceiro.objects.exclude(tipo__isnull=True).exclude(tipo='').values('tipo').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros
    empresas_unicas = ManutencaoTerceiro.objects.exclude(
        empresa__isnull=True
    ).exclude(
        empresa=''
    ).values_list('empresa', flat=True).distinct().order_by('empresa')
    
    tipos_unicos = ManutencaoTerceiro.objects.exclude(
        tipo__isnull=True
    ).exclude(
        tipo=''
    ).values_list('tipo', flat=True).distinct().order_by('tipo')
    
    maquinas_unicas = ManutencaoTerceiro.objects.exclude(
        maquina__isnull=True
    ).select_related('maquina').values_list('maquina__id', 'maquina__descr_maquina').distinct().order_by('maquina__descr_maquina')
    
    manutentores_unicos = ManutencaoTerceiro.objects.exclude(
        manutentor__isnull=True
    ).select_related('manutentor').values_list('manutentor__Matricula', 'manutentor__Nome').distinct().order_by('manutentor__Nome')
    
    context = {
        'page_title': 'Consultar Manutenções Terceiros',
        'active_page': 'consultar_manutencao_terceiros',
        'manutencoes': manutencoes,
        'total_count': total_count,
        'empresas_count': empresas_count,
        'tipos_count': tipos_count,
        # Valores para dropdowns
        'empresas_unicas': empresas_unicas,
        'tipos_unicos': tipos_unicos,
        'maquinas_unicas': maquinas_unicas,
        'manutentores_unicos': manutentores_unicos,
        # Valores dos filtros ativos
        'filtro_empresa': filtro_empresa,
        'filtro_tipo': filtro_tipo,
        'filtro_maquina': filtro_maquina,
        'filtro_manutentor': filtro_manutentor,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    return render(request, 'consultar/consultar_manutencao_terceiros.html', context)


def cadastrar_manutencao_terceiro(request):
    """Cadastrar nova manutenção de terceiro"""
    print(f"\n{'='*80}")
    print(f"VIEW CADASTRAR_MANUTENCAO_TERCEIRO CALLED - Method: {request.method}")
    print(f"URL: {request.path}")
    print(f"{'='*80}\n")
    
    from app.forms import ManutencaoTerceiroForm
    
    if request.method == 'POST':
        print(f"\n{'='*80}")
        print("POST REQUEST RECEBIDO!")
        print(f"POST keys: {list(request.POST.keys())}")
        print(f"POST data: {dict(request.POST)}")
        print(f"{'='*80}\n")
        
        form = ManutencaoTerceiroForm(request.POST)
        
        print(f"\nForm is_valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"\n{'='*60}")
            print("FORMULÁRIO INVÁLIDO!")
            print(f"Erros: {form.errors}")
            print(f"Non-field errors: {form.non_field_errors()}")
            print(f"{'='*60}\n")
        
        if form.is_valid():
            try:
                print("Tentando salvar manutenção terceiro...")
                print(f"Cleaned data: {form.cleaned_data}")
                manutencao = form.save(commit=False)
                print(f"Manutenção objeto criado: {manutencao}")
                print(f"Título: {manutencao.titulo}, Empresa: {manutencao.empresa}")
                manutencao.save()
                print(f"Manutenção salva com sucesso! ID: {manutencao.id}")
                messages.success(request, f'Manutenção de terceiro "{manutencao.titulo}" cadastrada com sucesso!')
                return redirect('home')
            except Exception as e:
                import traceback
                print(f"DEBUG - Erro ao salvar manutenção terceiro: {str(e)}")
                print(f"DEBUG - Traceback: {traceback.format_exc()}")
                messages.error(request, f'Erro ao cadastrar manutenção de terceiro: {str(e)}')
        else:
            # Exibir erros de validação específicos
            print(f"\n{'='*60}")
            print("FORMULÁRIO INVÁLIDO - EXIBINDO ERROS")
            print(f"Total de campos com erro: {len(form.errors)}")
            print(f"Erros: {form.errors}")
            print(f"{'='*60}\n")
            
            missing_required = []
            for field, errors in form.errors.items():
                field_label = form.fields[field].label if field in form.fields else field
                print(f"  Campo '{field}' ({field_label}): {errors}")
                for error in errors:
                    error_str = str(error).lower()
                    if 'required' in error_str or 'obrigatório' in error_str:
                        if field_label not in missing_required:
                            missing_required.append(field_label)
                        messages.warning(request, f'<strong>{field_label}</strong>: Este campo é obrigatório e deve ser preenchido.')
                    else:
                        messages.error(request, f'<strong>{field_label}</strong>: {error}')
            
            # Mostrar erros não relacionados a campos específicos
            if form.non_field_errors():
                print(f"Non-field errors: {form.non_field_errors()}")
                for error in form.non_field_errors():
                    messages.error(request, f'Erro no formulário: {error}')
            
            if missing_required:
                messages.warning(request, f'<strong>Atenção:</strong> {len(missing_required)} campo(s) obrigatório(s) não preenchido(s): {", ".join(missing_required)}. Por favor, preencha todos os campos marcados com <span class="text-danger">*</span>.')
            elif form.errors:
                messages.error(request, 'Por favor, corrija os erros no formulário antes de continuar.')
            else:
                messages.error(request, 'Ocorreu um erro ao processar o formulário. Por favor, tente novamente.')
    else:
        print("GET request - mostrando formulário vazio")
        form = ManutencaoTerceiroForm()
    
    context = {
        'page_title': 'Cadastrar Manutenção Terceiro',
        'active_page': 'cadastrar_manutencao_terceiro',
        'form': form
    }
    return render(request, 'cadastrar/cadastrar_manutencao_terceiro.html', context)


def visualizar_manutencao_terceiro(request, manutencao_id):
    """Visualizar detalhes de uma manutenção de terceiro específica"""
    from app.models import ManutencaoTerceiro
    
    try:
        manutencao = ManutencaoTerceiro.objects.select_related('maquina', 'manutentor', 'os_importada').get(id=manutencao_id)
    except ManutencaoTerceiro.DoesNotExist:
        messages.error(request, 'Manutenção de terceiro não encontrada.')
        return redirect('consultar_manutencao_terceiros')
    
    context = {
        'page_title': f'Visualizar Manutenção Terceiro - {manutencao.titulo}',
        'active_page': 'consultar_manutencao_terceiros',
        'manutencao': manutencao,
    }
    return render(request, 'visualizar/visualizar_manutencao_terceiro.html', context)


def editar_manutencao_terceiro(request, manutencao_id):
    """Editar uma manutenção de terceiro existente"""
    from app.forms import ManutencaoTerceiroForm
    from app.models import ManutencaoTerceiro
    
    try:
        manutencao = ManutencaoTerceiro.objects.get(id=manutencao_id)
    except ManutencaoTerceiro.DoesNotExist:
        messages.error(request, 'Manutenção de terceiro não encontrada.')
        return redirect('consultar_manutencao_terceiros')
    
    if request.method == 'POST':
        form = ManutencaoTerceiroForm(request.POST, instance=manutencao)
        
        if form.is_valid():
            try:
                manutencao = form.save()
                messages.success(request, f'Manutenção de terceiro "{manutencao.titulo}" atualizada com sucesso!')
                return redirect('visualizar_manutencao_terceiro', manutencao_id=manutencao.id)
            except Exception as e:
                messages.error(request, f'Erro ao atualizar manutenção de terceiro: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = ManutencaoTerceiroForm(instance=manutencao)
    
    context = {
        'page_title': f'Editar Manutenção Terceiro - {manutencao.titulo}',
        'active_page': 'consultar_manutencao_terceiros',
        'form': form,
        'manutencao': manutencao,
    }
    return render(request, 'editar/editar_manutencao_terceiro.html', context)


def deletar_manutencao_terceiro(request, manutencao_id):
    """Deletar uma manutenção de terceiro"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('consultar_manutencao_terceiros')
    
    from app.models import ManutencaoTerceiro
    
    try:
        manutencao = ManutencaoTerceiro.objects.get(id=manutencao_id)
        titulo = manutencao.titulo
        manutencao.delete()
        messages.success(request, f'Manutenção de terceiro "{titulo}" deletada com sucesso!')
    except ManutencaoTerceiro.DoesNotExist:
        messages.error(request, 'Manutenção de terceiro não encontrada.')
    except Exception as e:
        messages.error(request, f'Erro ao deletar manutenção de terceiro: {str(e)}')
    
    return redirect('consultar_manutencao_terceiros')


def cadastrar_manutentor(request):
    """Cadastrar novo manutentor"""
    from app.forms import ManutentorForm
    from app.models import ManutentorMaquina, Maquina
    
    if request.method == 'POST':
        form = ManutentorForm(request.POST)
        if form.is_valid():
            try:
                manutentor = form.save()
                
                # Processar máquinas selecionadas
                maquinas_ids = request.POST.getlist('maquinas_selecionadas')
                maquinas_adicionadas = 0
                for maquina_id in maquinas_ids:
                    try:
                        maquina = Maquina.objects.get(id=maquina_id)
                        # Verificar se já existe relação
                        if not ManutentorMaquina.objects.filter(manutentor=manutentor, maquina=maquina).exists():
                            ManutentorMaquina.objects.create(
                                manutentor=manutentor,
                                maquina=maquina
                            )
                            maquinas_adicionadas += 1
                    except Maquina.DoesNotExist:
                        pass
                    except Exception as e:
                        print(f"Erro ao relacionar máquina {maquina_id}: {str(e)}")
                
                if maquinas_adicionadas > 0:
                    messages.success(request, f'Manutentor {manutentor.Matricula} cadastrado com sucesso! {maquinas_adicionadas} máquina(s) relacionada(s).')
                else:
                    messages.success(request, f'Manutentor {manutentor.Matricula} cadastrado com sucesso!')
                
                return redirect('consultar_manutentores')
            except Exception as e:
                import traceback
                print(f"DEBUG - Erro ao salvar manutentor: {str(e)}")
                print(f"DEBUG - Traceback: {traceback.format_exc()}")
                messages.error(request, f'Erro ao cadastrar manutentor: {str(e)}')
        else:
            # Exibir erros de validação específicos
            missing_required = []
            for field, errors in form.errors.items():
                field_label = form.fields[field].label
                for error in errors:
                    if 'required' in str(error).lower() or 'obrigatório' in str(error).lower():
                        missing_required.append(field_label)
                        messages.warning(request, f'<strong>{field_label}</strong>: Este campo é obrigatório e deve ser preenchido.')
                    else:
                        messages.error(request, f'<strong>{field_label}</strong>: {error}')
            
            if missing_required:
                messages.warning(request, f'<strong>Atenção:</strong> {len(missing_required)} campo(s) obrigatório(s) não preenchido(s). Por favor, preencha todos os campos marcados com <span class="text-danger">*</span>.')
            else:
                messages.error(request, 'Por favor, corrija os erros no formulário antes de continuar.')
    else:
        form = ManutentorForm()
    
    # Buscar todas as máquinas disponíveis
    maquinas_disponiveis = Maquina.objects.all().order_by('cd_maquina')
    
    context = {
        'page_title': 'Cadastrar Manutentor',
        'active_page': 'cadastrar_manutentor',
        'form': form,
        'maquinas_disponiveis': maquinas_disponiveis
    }
    return render(request, 'cadastrar/cadastrar_manutentor.html', context)


def analise_manutentores(request):
    """Análise de Manutentores - Dashboard com estatísticas"""
    from app.models import Manutentor, ManutentorMaquina, OrdemServicoCorretiva, OrdemServicoCorretivaFicha
    from django.db.models import Count, Q
    import json
    
    # Estatísticas básicas
    total_count = Manutentor.objects.count()
    
    # Contagem por turno (A, B, C no campo turno) e local (Externa, Civil no campo local_trab)
    turno_a_count = Manutentor.objects.filter(Q(turno__iexact='Turno A') | Q(turno__iexact='A')).count()
    turno_b_count = Manutentor.objects.filter(Q(turno__iexact='Turno B') | Q(turno__iexact='B')).count()
    turno_c_count = Manutentor.objects.filter(Q(turno__iexact='Turno C') | Q(turno__iexact='C')).count()
    externa_count = Manutentor.objects.filter(local_trab__iexact='Externa').count()
    civil_count = Manutentor.objects.filter(local_trab__iexact='Civil').count()
    frigorifico_count = Manutentor.objects.filter(local_trab__iexact='Frigorífico').count()
    industria_count = Manutentor.objects.filter(local_trab__iexact='Industria').count()
    
    # Manutentores com máquinas relacionadas (usado em outras partes)
    manutentores_com_maquinas = Manutentor.objects.filter(
        maquinas__isnull=False
    ).distinct().count()
    percentual_com_maquinas = (manutentores_com_maquinas / total_count * 100) if total_count > 0 else 0
    
    # Total de máquinas relacionadas
    total_maquinas_relacionadas = ManutentorMaquina.objects.count()
    
    # Manutentores com manutenções (via OrdemServicoCorretiva ou Fichas)
    # Buscar nomes de manutentores e contar ordens/fichas que correspondem
    manutentores_nomes = Manutentor.objects.values_list('Nome', flat=True).exclude(Nome__isnull=True).exclude(Nome='')
    manutentores_com_manutencoes = 0
    total_manutencoes_relacionadas = 0
    
    for nome in manutentores_nomes:
        ordens_count = OrdemServicoCorretiva.objects.filter(
            Q(nm_func_exec__icontains=nome)
        ).count()
        fichas_count = OrdemServicoCorretivaFicha.objects.filter(
            nm_func_exec_os__icontains=nome
        ).count()
        if ordens_count > 0 or fichas_count > 0:
            manutentores_com_manutencoes += 1
        total_manutencoes_relacionadas += ordens_count + fichas_count
    
    # Distribuição por Turno
    turnos_data_dict = Manutentor.objects.values('turno').annotate(total=Count('Matricula')).order_by('-total')
    turnos_labels = [item['turno'] or 'Não informado' for item in turnos_data_dict]
    turnos_data = [item['total'] for item in turnos_data_dict]
    
    # Distribuição por Local de Trabalho
    locais_data_dict = Manutentor.objects.values('local_trab').annotate(total=Count('Matricula')).order_by('-total')
    locais_labels = [item['local_trab'] or 'Não informado' for item in locais_data_dict]
    locais_data = [item['total'] for item in locais_data_dict]
    
    # Distribuição de Máquinas por Manutentor (mais granular: 0, 1, 2, 3, 4, 5, 6-10, 11-15, 16+)
    distribuicao_maquinas_labels = ['0 máquinas', '1 máquina', '2 máquinas', '3 máquinas', '4 máquinas', '5 máquinas', '6-10 máquinas', '11-15 máquinas', '16+ máquinas']
    distribuicao_maquinas_data = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    
    for manutentor in Manutentor.objects.all():
        qtd_maquinas = ManutentorMaquina.objects.filter(manutentor=manutentor).count()
        if qtd_maquinas == 0:
            distribuicao_maquinas_data[0] += 1
        elif qtd_maquinas == 1:
            distribuicao_maquinas_data[1] += 1
        elif qtd_maquinas == 2:
            distribuicao_maquinas_data[2] += 1
        elif qtd_maquinas == 3:
            distribuicao_maquinas_data[3] += 1
        elif qtd_maquinas == 4:
            distribuicao_maquinas_data[4] += 1
        elif qtd_maquinas == 5:
            distribuicao_maquinas_data[5] += 1
        elif qtd_maquinas <= 10:
            distribuicao_maquinas_data[6] += 1
        elif qtd_maquinas <= 15:
            distribuicao_maquinas_data[7] += 1
        else:
            distribuicao_maquinas_data[8] += 1
    
    # Top 10 Cargos
    cargos_data_dict = Manutentor.objects.exclude(Cargo__isnull=True).exclude(Cargo='').values('Cargo').annotate(
        total=Count('Matricula')
    ).order_by('-total')[:10]
    cargos_labels = [item['Cargo'][:30] for item in cargos_data_dict]
    cargos_data = [item['total'] for item in cargos_data_dict]
    
    # Top 10 Manutentores por Quantidade de Máquinas
    top_manutentores_maquinas = []
    for manutentor in Manutentor.objects.all():
        qtd_maquinas = ManutentorMaquina.objects.filter(manutentor=manutentor).count()
        if qtd_maquinas > 0:
            top_manutentores_maquinas.append({
                'manutentor': manutentor,
                'total': qtd_maquinas
            })
    top_manutentores_maquinas = sorted(top_manutentores_maquinas, key=lambda x: x['total'], reverse=True)[:10]
    
    # Top 10 Manutentores por Quantidade de Manutenções
    top_manutentores_manutencoes = []
    for manutentor in Manutentor.objects.all():
        if manutentor.Nome:
            ordens_count = OrdemServicoCorretiva.objects.filter(
                Q(nm_func_exec__icontains=manutentor.Nome)
            ).count()
            fichas_count = OrdemServicoCorretivaFicha.objects.filter(
                nm_func_exec_os__icontains=manutentor.Nome
            ).count()
            total_manut = ordens_count + fichas_count
            if total_manut > 0:
                top_manutentores_manutencoes.append({
                    'manutentor': manutentor,
                    'total': total_manut
                })
    top_manutentores_manutencoes = sorted(top_manutentores_manutencoes, key=lambda x: x['total'], reverse=True)[:10]
    
    context = {
        'page_title': 'Análise de Manutentores',
        'active_page': 'analise_manutentores',
        'total_count': total_count,
        'turno_a_count': turno_a_count,
        'turno_b_count': turno_b_count,
        'turno_c_count': turno_c_count,
        'externa_count': externa_count,
        'civil_count': civil_count,
        'frigorifico_count': frigorifico_count,
        'industria_count': industria_count,
        'manutentores_com_maquinas': manutentores_com_maquinas,
        'percentual_com_maquinas': round(percentual_com_maquinas, 1),
        'manutentores_com_manutencoes': manutentores_com_manutencoes,
        'total_maquinas_relacionadas': total_maquinas_relacionadas,
        'total_manutencoes_relacionadas': total_manutencoes_relacionadas,
        # Dados para gráficos (JSON)
        'turnos_labels': json.dumps(turnos_labels),
        'turnos_data': json.dumps(turnos_data),
        'locais_labels': json.dumps(locais_labels),
        'locais_data': json.dumps(locais_data),
        'distribuicao_maquinas_labels': json.dumps(distribuicao_maquinas_labels),
        'distribuicao_maquinas_data': json.dumps(distribuicao_maquinas_data),
        'distribuicao_maquinas_zip': list(zip(distribuicao_maquinas_labels, distribuicao_maquinas_data)),
        'cargos_labels': json.dumps(cargos_labels),
        'cargos_data': json.dumps(cargos_data),
        # Dados para tabelas
        'top_manutentores_maquinas': top_manutentores_maquinas,
        'top_manutentores_manutencoes': top_manutentores_manutencoes,
    }
    return render(request, 'analise/analise_manutentores.html', context)


def configuracao_manutentores(request):
    """Página de configuração de manutentores"""
    return render(request, 'manutentor/configuracao_manutentores.html')


def visualizar_manutentor(request, matricula):
    """Visualizar detalhes de um manutentor específico"""
    from app.models import Manutentor, ManutentorMaquina, Maquina
    
    try:
        manutentor = Manutentor.objects.get(Matricula=matricula)
    except Manutentor.DoesNotExist:
        messages.error(request, 'Manutentor não encontrado.')
        return redirect('consultar_manutentores')
    
    # Buscar máquinas relacionadas
    maquinas_relacionadas = ManutentorMaquina.objects.filter(manutentor=manutentor).select_related('maquina')
    
    # Buscar máquinas já relacionadas para excluir da lista de disponíveis
    maquinas_ids_relacionadas = maquinas_relacionadas.values_list('maquina_id', flat=True)
    maquinas_disponiveis = Maquina.objects.exclude(id__in=maquinas_ids_relacionadas).order_by('cd_maquina')
    
    context = {
        'page_title': f'Visualizar Manutentor {manutentor.Matricula}',
        'active_page': 'consultar_manutentores',
        'manutentor': manutentor,
        'maquinas_relacionadas': maquinas_relacionadas,
        'maquinas_disponiveis': maquinas_disponiveis,
    }
    return render(request, 'visualizar/visualizar_manutentor.html', context)


def editar_manutentor(request, matricula):
    """Editar um manutentor existente"""
    from app.forms import ManutentorForm
    from app.models import Manutentor
    
    try:
        manutentor = Manutentor.objects.get(Matricula=matricula)
    except Manutentor.DoesNotExist:
        messages.error(request, 'Manutentor não encontrado.')
        return redirect('consultar_manutentores')
    
    if request.method == 'POST':
        # Garantir que a Matricula não seja alterada (é a primary key)
        post_data = request.POST.copy()
        post_data['Matricula'] = manutentor.Matricula
        
        form = ManutentorForm(post_data, instance=manutentor)
        
        if form.is_valid():
            try:
                manutentor = form.save()
                messages.success(request, f'Manutentor {manutentor.Matricula} atualizado com sucesso!')
                return redirect('visualizar_manutentor', matricula=manutentor.Matricula)
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                print(f"Erro ao salvar manutentor: {error_detail}")
                messages.error(request, f'Erro ao atualizar manutentor: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = ManutentorForm(instance=manutentor)
        # Tornar o campo Matricula readonly na edição (é a primary key)
        form.fields['Matricula'].widget.attrs['readonly'] = True
        form.fields['Matricula'].widget.attrs['class'] = 'form-control bg-light'
    
    context = {
        'page_title': f'Editar Manutentor {manutentor.Matricula}',
        'active_page': 'consultar_manutentores',
        'form': form,
        'manutentor': manutentor,
    }
    return render(request, 'visualizar/editar_manutentor.html', context)


def consultar_manutentores(request):
    """Consultar/listar manutentores cadastrados com filtros avançados"""
    from app.models import Manutentor, TURNO, LOCAL_TRABALHO
    from datetime import datetime
    
    # Buscar todos os manutentores
    manutentores_list = Manutentor.objects.all()
    
    # Filtro de busca geral (texto)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        manutentores_list = manutentores_list.filter(
            Q(Matricula__icontains=search_query) |
            Q(Nome__icontains=search_query) |
            Q(Cargo__icontains=search_query) |
            Q(turno__icontains=search_query) |
            Q(local_trab__icontains=search_query)
        )
    
    # Filtros específicos
    # Filtro por Turno
    filtro_turno = request.GET.get('filtro_turno', '')
    if filtro_turno:
        manutentores_list = manutentores_list.filter(turno=filtro_turno)
    
    # Filtro por Local de Trabalho
    filtro_local_trab = request.GET.get('filtro_local_trab', '')
    if filtro_local_trab:
        manutentores_list = manutentores_list.filter(local_trab=filtro_local_trab)
    
    # Filtro por Cargo
    filtro_cargo = request.GET.get('filtro_cargo', '')
    if filtro_cargo:
        manutentores_list = manutentores_list.filter(Cargo__icontains=filtro_cargo)
    
    # Ordenar por nome e matricula
    manutentores_list = manutentores_list.order_by('Nome', 'Matricula')
    
    # Paginação
    paginator = Paginator(manutentores_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    manutentores = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = Manutentor.objects.count()
    turnos_count = Manutentor.objects.exclude(turno__isnull=True).exclude(turno='').values('turno').distinct().count()
    locais_count = Manutentor.objects.exclude(local_trab__isnull=True).exclude(local_trab='').values('local_trab').distinct().count()
    
    # Obter valores únicos para os dropdowns de filtros
    cargos_unicos = Manutentor.objects.exclude(
        Cargo__isnull=True
    ).exclude(
        Cargo=''
    ).values_list('Cargo', flat=True).distinct().order_by('Cargo')
    
    
    context = {
        'page_title': 'Consultar Manutentores',
        'active_page': 'consultar_manutentores',
        'manutentores': manutentores,
        'total_count': total_count,
        'turnos_count': turnos_count,
        'locais_count': locais_count,
        # Valores para dropdowns
        'turnos': TURNO,
        'locais_trabalho': LOCAL_TRABALHO,
        'cargos_unicos': cargos_unicos,
        # Valores dos filtros ativos
        'filtro_turno': filtro_turno,
        'filtro_local_trab': filtro_local_trab,
        'filtro_cargo': filtro_cargo,
    }
    return render(request, 'consultar/consultar_manutentores.html', context)


def consultar_agendamentos(request):
    """Consultar/listar agendamentos de cronograma cadastrados com visitas"""
    from app.models import AgendamentoCronograma, Visitas
    from django.db.models import Q
    from django.core.paginator import Paginator
    from datetime import datetime
    
    # Buscar todos os agendamentos com visitas relacionadas
    agendamentos_list = AgendamentoCronograma.objects.select_related('maquina', 'plano_preventiva', 'semana').prefetch_related('visitas').all()
    
    # Filtro de busca geral (incluindo campos de visitas)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        agendamentos_list = agendamentos_list.filter(
            Q(nome_grupo__icontains=search_query) |
            Q(maquina__cd_maquina__icontains=search_query) |
            Q(maquina__descr_maquina__icontains=search_query) |
            Q(plano_preventiva__numero_plano__icontains=search_query) |
            Q(plano_preventiva__descr_plano__icontains=search_query) |
            Q(observacoes__icontains=search_query) |
            Q(visitas__titulo__icontains=search_query) |
            Q(visitas__descricao__icontains=search_query) |
            Q(visitas__nome_contato__icontains=search_query)
        ).distinct()
    
    # Filtros específicos
    filtro_tipo = request.GET.get('filtro_tipo', '')
    if filtro_tipo:
        agendamentos_list = agendamentos_list.filter(tipo_agendamento=filtro_tipo)
    
    filtro_data_inicio = request.GET.get('filtro_data_inicio', '')
    if filtro_data_inicio:
        try:
            data_ini = datetime.strptime(filtro_data_inicio, '%Y-%m-%d').date()
            agendamentos_list = agendamentos_list.filter(data_planejada__gte=data_ini)
        except ValueError:
            pass
    
    filtro_data_fim = request.GET.get('filtro_data_fim', '')
    if filtro_data_fim:
        try:
            data_f = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
            agendamentos_list = agendamentos_list.filter(data_planejada__lte=data_f)
        except ValueError:
            pass
    
    # Ordenar por data planejada
    agendamentos_list = agendamentos_list.order_by('data_planejada', 'tipo_agendamento')
    
    # Paginação
    paginator = Paginator(agendamentos_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    agendamentos = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = AgendamentoCronograma.objects.count()
    tipo_maquina_count = AgendamentoCronograma.objects.filter(tipo_agendamento='maquina').count()
    tipo_plano_count = AgendamentoCronograma.objects.filter(tipo_agendamento='plano').count()
    visitas_count = Visitas.objects.count()
    
    context = {
        'page_title': 'Consultar Agendamentos',
        'active_page': 'consultar_agendamentos',
        'agendamentos': agendamentos,
        'total_count': total_count,
        'tipo_maquina_count': tipo_maquina_count,
        'tipo_plano_count': tipo_plano_count,
        'visitas_count': visitas_count,
        'filtro_tipo': filtro_tipo,
        'filtro_data_inicio': filtro_data_inicio,
        'filtro_data_fim': filtro_data_fim,
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_agendamentos.html', context)


def consultar_visitas(request):
    """Consultar/listar visitas cadastradas"""
    from app.models import Visitas
    from django.db.models import Q
    from django.core.paginator import Paginator
    from datetime import datetime
    
    # Buscar todas as visitas
    visitas_list = Visitas.objects.all()
    
    # Filtro de busca geral
    search_query = request.GET.get('search', '').strip()
    if search_query:
        visitas_list = visitas_list.filter(
            Q(titulo__icontains=search_query) |
            Q(descricao__icontains=search_query) |
            Q(nome_contato__icontains=search_query) |
            Q(numero_contato__icontains=search_query)
        )
    
    # Filtros específicos
    filtro_data_inicio = request.GET.get('filtro_data_inicio', '')
    if filtro_data_inicio:
        try:
            data_ini = datetime.strptime(filtro_data_inicio, '%Y-%m-%d').date()
            visitas_list = visitas_list.filter(data__date__gte=data_ini)
        except ValueError:
            pass
    
    filtro_data_fim = request.GET.get('filtro_data_fim', '')
    if filtro_data_fim:
        try:
            data_f = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
            visitas_list = visitas_list.filter(data__date__lte=data_f)
        except ValueError:
            pass
    
    # Ordenar por data (mais recente primeiro)
    visitas_list = visitas_list.order_by('-data', '-created_at')
    
    # Paginação
    paginator = Paginator(visitas_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    visitas = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = Visitas.objects.count()
    com_documentos_count = Visitas.objects.exclude(documento__isnull=True).exclude(documento='').count()
    com_contato_count = Visitas.objects.exclude(nome_contato__isnull=True).exclude(nome_contato='').count()
    
    context = {
        'page_title': 'Consultar Visitas',
        'active_page': 'consultar_visitas',
        'visitas': visitas,
        'total_count': total_count,
        'com_documentos_count': com_documentos_count,
        'com_contato_count': com_contato_count,
        'filtro_data_inicio': filtro_data_inicio,
        'filtro_data_fim': filtro_data_fim,
        'search_query': search_query,
    }
    return render(request, 'consultar/consultar_visitas.html', context)


def cadastrar_visita(request):
    """Cadastrar nova visita"""
    print(f"\n{'='*80}")
    print(f"VIEW CADASTRAR_VISITA CALLED - Method: {request.method}")
    print(f"URL: {request.path}")
    print(f"{'='*80}\n")
    
    from app.forms import VisitasForm
    from app.models import Visitas
    
    if request.method == 'POST':
        print(f"\n{'='*60}")
        print("DEBUG - POST request recebido")
        print(f"POST data: {dict(request.POST)}")
        print(f"FILES data: {dict(request.FILES)}")
        print(f"{'='*60}\n")
        
        form = VisitasForm(request.POST, request.FILES)
        
        print(f"Form is_valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"Form errors: {form.errors}")
            print(f"Form data: {form.data}")
            print(f"Form cleaned_data: {form.cleaned_data if hasattr(form, 'cleaned_data') else 'N/A'}")
        
        if form.is_valid():
            try:
                print("Tentando salvar visita...")
                print(f"Cleaned data: {form.cleaned_data}")
                visita = form.save(commit=False)
                print(f"Visita objeto criado: {visita}")
                print(f"Título: {visita.titulo}")
                visita.save()
                print(f"Visita salva com sucesso! ID: {visita.id}, Título: {visita.titulo}")
                messages.success(request, f'Visita "{visita.titulo}" cadastrada com sucesso!')
                return redirect('consultar_visitas')
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"\n{'='*60}")
                print(f"DEBUG - Erro ao salvar visita: {str(e)}")
                print(f"DEBUG - Traceback: {error_details}")
                print(f"{'='*60}\n")
                messages.error(request, f'Erro ao cadastrar visita: {str(e)}')
        else:
            # Usar helper function para exibir erros
            print(f"\n{'='*60}")
            print("Formulário inválido!")
            print(f"Erros: {form.errors}")
            print(f"{'='*60}\n")
            handle_form_errors(form, request)
            # Mostrar erros não relacionados a campos específicos
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, f'Erro no formulário: {error}')
    else:
        form = VisitasForm()
    
    context = {
        'page_title': 'Cadastrar Visita',
        'active_page': 'cadastrar_visita',
        'form': form
    }
    return render(request, 'cadastrar/cadastrar_visita.html', context)


def editar_visita(request, visita_id):
    """Editar uma visita existente"""
    from app.forms import VisitasForm
    from app.models import Visitas
    
    try:
        visita = Visitas.objects.get(id=visita_id)
    except Visitas.DoesNotExist:
        messages.error(request, 'Visita não encontrada.')
        return redirect('consultar_visitas')
    
    if request.method == 'POST':
        form = VisitasForm(request.POST, request.FILES, instance=visita)
        
        if form.is_valid():
            try:
                visita = form.save()
                messages.success(request, f'Visita "{visita.titulo}" atualizada com sucesso!')
                return redirect('consultar_visitas')
            except Exception as e:
                messages.error(request, f'Erro ao atualizar visita: {str(e)}')
        else:
            handle_form_errors(form, request)
    else:
        form = VisitasForm(instance=visita)
    
    context = {
        'page_title': f'Editar Visita - {visita.titulo}',
        'active_page': 'consultar_visitas',
        'form': form,
        'visita': visita,
    }
    return render(request, 'editar/editar_visita.html', context)


def deletar_visita(request, visita_id):
    """Deletar uma visita"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('consultar_visitas')
    
    from app.models import Visitas
    
    try:
        visita = Visitas.objects.get(id=visita_id)
        titulo = visita.titulo
        visita.delete()
        messages.success(request, f'Visita "{titulo}" deletada com sucesso!')
    except Visitas.DoesNotExist:
        messages.error(request, 'Visita não encontrada.')
    except Exception as e:
        messages.error(request, f'Erro ao deletar visita: {str(e)}')
    
    return redirect('consultar_visitas')


def visualizar_visita(request, visita_id):
    """Visualizar detalhes de uma visita específica"""
    from app.models import Visitas
    
    try:
        visita = Visitas.objects.get(id=visita_id)
    except Visitas.DoesNotExist:
        messages.error(request, 'Visita não encontrada.')
        return redirect('consultar_visitas')
    
    context = {
        'page_title': f'Visualizar Visita - {visita.titulo}',
        'active_page': 'consultar_visitas',
        'visita': visita,
    }
    return render(request, 'visualizar/visualizar_visita.html', context)


def agenda_geral(request):
    """Página de agenda geral com FullCalendar"""
    context = {
        'page_title': 'Agenda Geral',
        'active_page': 'agenda_geral',
    }
    return render(request, 'agenda_geral.html', context)


def api_eventos_calendario(request):
    """API endpoint para retornar eventos do calendário em formato JSON"""
    from app.models import Visitas, ManutencaoTerceiro, AgendamentoCronograma
    from django.http import JsonResponse
    from datetime import datetime
    
    # Obter filtros da query string
    tipos_eventos = request.GET.getlist('tipos[]', [])
    data_inicio = request.GET.get('start', None)
    data_fim = request.GET.get('end', None)
    
    eventos = []
    
    # Se nenhum tipo foi selecionado ou 'visitas' está selecionado
    if not tipos_eventos or 'visitas' in tipos_eventos:
        visitas = Visitas.objects.filter(data__isnull=False)
        if data_inicio and data_fim:
            try:
                start = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
                end = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
                visitas = visitas.filter(data__gte=start, data__lte=end)
            except:
                pass
        
        for visita in visitas:
            eventos.append({
                'id': f'visita_{visita.id}',
                'title': f'Visita: {visita.titulo}',
                'start': visita.data.isoformat() if visita.data else None,
                'color': '#0dcaf0',  # Cyan
                'textColor': '#000',
                'extendedProps': {
                    'tipo': 'visita',
                    'descricao': visita.descricao or '',
                    'url': f'/visitas/visualizar/{visita.id}/'
                }
            })
    
    # Se nenhum tipo foi selecionado ou 'manutencao_terceiro' está selecionado
    if not tipos_eventos or 'manutencao_terceiro' in tipos_eventos:
        manutencoes = ManutencaoTerceiro.objects.filter(data__isnull=False)
        if data_inicio and data_fim:
            try:
                start = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
                end = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
                manutencoes = manutencoes.filter(data__gte=start, data__lte=end)
            except:
                pass
        
        for manutencao in manutencoes:
            eventos.append({
                'id': f'manutencao_{manutencao.id}',
                'title': f'Manutenção: {manutencao.titulo}',
                'start': manutencao.data.isoformat() if manutencao.data else None,
                'color': '#ff9800',  # Orange
                'textColor': '#000',
                'extendedProps': {
                    'tipo': 'manutencao_terceiro',
                    'empresa': manutencao.empresa or '',
                    'maquina': manutencao.maquina.cd_maquina if manutencao.maquina else '',
                    'url': f'/manutencao-terceiro/visualizar/{manutencao.id}/'
                }
            })
    
    # Se nenhum tipo foi selecionado ou 'agendamento' está selecionado
    if not tipos_eventos or 'agendamento' in tipos_eventos:
        agendamentos = AgendamentoCronograma.objects.filter(data_planejada__isnull=False)
        if data_inicio and data_fim:
            try:
                start = datetime.fromisoformat(data_inicio.replace('Z', '+00:00')).date()
                end = datetime.fromisoformat(data_fim.replace('Z', '+00:00')).date()
                agendamentos = agendamentos.filter(data_planejada__gte=start, data_planejada__lte=end)
            except:
                pass
        
        for agendamento in agendamentos:
            if agendamento.tipo_agendamento == 'maquina' and agendamento.maquina:
                titulo = f'Agendamento: {agendamento.maquina.cd_maquina}'
            elif agendamento.tipo_agendamento == 'plano' and agendamento.plano_preventiva:
                titulo = f'Plano: {agendamento.plano_preventiva.numero_plano}'
            else:
                titulo = f'Agendamento: {agendamento.nome_grupo or "Sem nome"}'
            
            eventos.append({
                'id': f'agendamento_{agendamento.id}',
                'title': titulo,
                'start': agendamento.data_planejada.isoformat() if agendamento.data_planejada else None,
                'color': '#198754',  # Green
                'textColor': '#fff',
                'extendedProps': {
                    'tipo': 'agendamento',
                    'tipo_agendamento': agendamento.tipo_agendamento,
                    'observacoes': agendamento.observacoes or '',
                    'url': f'/agendamentos/consultar/'
                }
            })
    
    return JsonResponse(eventos, safe=False)


def gerenciar_projeto(request):
    """Página de gerenciamento administrativo do projeto"""
    from app.models import (
        Maquina, MaquinaDocumento, OrdemServicoCorretiva, OrdemServicoCorretivaFicha,
        CentroAtividade, Semana52, Manutentor, ManutentorMaquina,
        ItemEstoque, ManutencaoCsv, ManutencaoTerceiro, MaquinaPeca,
        MaquinaPrimariaSecundaria, PlanoPreventiva, PlanoPreventivaDocumento,
        MeuPlanoPreventiva, MeuPlanoPreventivaDocumento, AgendamentoCronograma,
        RoteiroPreventiva, RequisicaoAlmoxarifado, NotaFiscal, Visitas,
        ProjecaoGasto, RelacaoProjecaoNotaFiscal, DadosOrcamento, ControleRCeNF,
        ParadaMaquina
    )
    from datetime import date
    
    # Definir todos os modelos com suas informações
    modelos_info = [
        {
            'nome': 'Máquinas',
            'modelo': Maquina,
            'key': 'maquinas',
            'icone': 'fas fa-industry',
            'cor': 'primary',
            'descricao': 'Registros de máquinas cadastradas no sistema'
        },
        {
            'nome': 'Ordens Corretivas',
            'modelo': OrdemServicoCorretiva,
            'key': 'ordens',
            'icone': 'fas fa-wrench',
            'cor': 'info',
            'descricao': 'Ordens de serviço corretivas e outros fechadas'
        },
        {
            'nome': 'Fichas de Ordens Corretivas',
            'modelo': OrdemServicoCorretivaFicha,
            'key': 'ordens_ficha',
            'icone': 'fas fa-file-alt',
            'cor': 'info',
            'descricao': 'Fichas técnicas de ordens de serviço corretivas'
        },
        {
            'nome': 'Centros de Atividade',
            'modelo': CentroAtividade,
            'key': 'centros',
            'icone': 'fas fa-building',
            'cor': 'success',
            'descricao': 'Centros de atividade (CA) cadastrados com seus locais'
        },
        {
            'nome': 'Manutentores',
            'modelo': Manutentor,
            'key': 'manutentores',
            'icone': 'fas fa-user-tie',
            'cor': 'warning',
            'descricao': 'Manutentores cadastrados no sistema'
        },
        {
            'nome': 'Máquinas dos Manutentores',
            'modelo': ManutentorMaquina,
            'key': 'manutentor_maquina',
            'icone': 'fas fa-link',
            'cor': 'warning',
            'descricao': 'Relação entre manutentores e máquinas'
        },
        {
            'nome': 'Itens de Estoque',
            'modelo': ItemEstoque,
            'key': 'estoque',
            'icone': 'fas fa-boxes',
            'cor': 'secondary',
            'descricao': 'Itens de estoque cadastrados'
        },
        {
            'nome': 'Manutenções CSV',
            'modelo': ManutencaoCsv,
            'key': 'manutencao_csv',
            'icone': 'fas fa-file-csv',
            'cor': 'dark',
            'descricao': 'Manutenções importadas via CSV'
        },
        {
            'nome': 'Manutenções Terceiros',
            'modelo': ManutencaoTerceiro,
            'key': 'manutencao_terceiros',
            'icone': 'fas fa-tools',
            'cor': 'danger',
            'descricao': 'Manutenções de terceiros cadastradas'
        },
        {
            'nome': 'Peças das Máquinas',
            'modelo': MaquinaPeca,
            'key': 'maquina_peca',
            'icone': 'fas fa-cog',
            'cor': 'primary',
            'descricao': 'Relação entre máquinas e peças'
        },
        {
            'nome': 'Máquinas Primárias/Secundárias',
            'modelo': MaquinaPrimariaSecundaria,
            'key': 'maquina_primaria_secundaria',
            'icone': 'fas fa-sitemap',
            'cor': 'primary',
            'descricao': 'Relação entre máquinas primárias e secundárias'
        },
        {
            'nome': 'Planos Preventiva',
            'modelo': PlanoPreventiva,
            'key': 'plano_preventiva',
            'icone': 'fas fa-calendar-check',
            'cor': 'success',
            'descricao': 'Planos de manutenção preventiva'
        },
        {
            'nome': 'Documentos Planos Preventiva',
            'modelo': PlanoPreventivaDocumento,
            'key': 'plano_preventiva_documento',
            'icone': 'fas fa-file-upload',
            'cor': 'success',
            'descricao': 'Documentos relacionados aos planos preventiva'
        },
        {
            'nome': 'Meus Planos Preventiva',
            'modelo': MeuPlanoPreventiva,
            'key': 'meu_plano_preventiva',
            'icone': 'fas fa-calendar-alt',
            'cor': 'info',
            'descricao': 'Planos preventiva com descrição detalhada do roteiro'
        },
        {
            'nome': 'Roteiros Preventiva',
            'modelo': RoteiroPreventiva,
            'key': 'roteiro_preventiva',
            'icone': 'fas fa-route',
            'cor': 'primary',
            'descricao': 'Roteiros de manutenção preventiva'
        },
        {
            'nome': 'Documentos das Máquinas',
            'modelo': MaquinaDocumento,
            'key': 'maquina_documento',
            'icone': 'fas fa-file-pdf',
            'cor': 'primary',
            'descricao': 'Documentos relacionados às máquinas'
        },
        {
            'nome': 'Semanas 52',
            'modelo': Semana52,
            'key': 'semana52',
            'icone': 'fas fa-calendar-week',
            'cor': 'info',
            'descricao': 'Semanas do ano (52 semanas)'
        },
        {
            'nome': 'Documentos Meus Planos Preventiva',
            'modelo': MeuPlanoPreventivaDocumento,
            'key': 'meu_plano_preventiva_documento',
            'icone': 'fas fa-file-alt',
            'cor': 'info',
            'descricao': 'Documentos associados aos planos PCM'
        },
        {
            'nome': 'Agendamentos Cronograma',
            'modelo': AgendamentoCronograma,
            'key': 'agendamento_cronograma',
            'icone': 'fas fa-calendar-day',
            'cor': 'success',
            'descricao': 'Agendamentos de máquinas e planos no cronograma'
        },
        {
            'nome': 'Requisições Almoxarifado',
            'modelo': RequisicaoAlmoxarifado,
            'key': 'requisicao_almoxarifado',
            'icone': 'fas fa-shopping-cart',
            'cor': 'warning',
            'descricao': 'Requisições de itens retirados do almoxarifado (campo Data da Requisição define o mês)',
            'limpar_por_mes': True,
        },
        {
            'nome': 'Notas Fiscais',
            'modelo': NotaFiscal,
            'key': 'nota_fiscal',
            'icone': 'fas fa-file-invoice-dollar',
            'cor': 'primary',
            'descricao': 'Notas fiscais cadastradas no sistema (mês definido pelo campo Data Emissão)',
            'limpar_por_mes': True,
        },
        {
            'nome': 'Visitas',
            'modelo': Visitas,
            'key': 'visitas',
            'icone': 'fas fa-calendar-check',
            'cor': 'info',
            'descricao': 'Registros de visitas cadastradas'
        },
        {
            'nome': 'Projeções de Gastos',
            'modelo': ProjecaoGasto,
            'key': 'projecao_gasto',
            'icone': 'fas fa-chart-line',
            'cor': 'success',
            'descricao': 'Projeções de gastos e requisições de serviço'
        },
        {
            'nome': 'Relações Projeção-Nota Fiscal',
            'modelo': RelacaoProjecaoNotaFiscal,
            'key': 'relacao_projecao_nota_fiscal',
            'icone': 'fas fa-link',
            'cor': 'info',
            'descricao': 'Relações confirmadas entre projeções e notas fiscais'
        },
        {
            'nome': 'Dados de Orçamento',
            'modelo': DadosOrcamento,
            'key': 'dados_orcamento',
            'icone': 'fas fa-money-bill-wave',
            'cor': 'success',
            'descricao': 'Dados de orçamento por ano, mês e conta orçamentária'
        },
        {
            'nome': 'Controle RC e NF',
            'modelo': ControleRCeNF,
            'key': 'controle_rc_nf',
            'icone': 'fas fa-clipboard-list',
            'cor': 'warning',
            'descricao': 'Controle de RC (Requisição de Compra) e NF (Nota Fiscal)'
        },
        {
            'nome': 'Paradas de Máquina',
            'modelo': ParadaMaquina,
            'key': 'paradas_maquina',
            'icone': 'fas fa-stop-circle',
            'cor': 'warning',
            'descricao': 'Registros de paradas de máquinas (campo Data define o mês)',
            'limpar_por_mes': True,
        },
    ]
    
    # Contar registros em cada tabela
    tabelas_info = []
    total_geral = 0
    
    for info in modelos_info:
        count = info['modelo'].objects.count()
        total_geral += count
        tabelas_info.append({
            **info,
            'count': count
        })
    
    # Manter compatibilidade com o template atual (primeiros 4 para cards de estatísticas)
    # Buscar pelos keys específicos para garantir que pegamos os valores corretos
    maquinas_count = next((t['count'] for t in tabelas_info if t['key'] == 'maquinas'), 0)
    ordens_count = next((t['count'] for t in tabelas_info if t['key'] == 'ordens'), 0)
    centros_count = next((t['count'] for t in tabelas_info if t['key'] == 'centros'), 0)
    manutentores_count = next((t['count'] for t in tabelas_info if t['key'] == 'manutentores'), 0)
    
    # Anos e meses disponíveis para "Limpar por mês" (ParadaMaquina)
    anos_paradas = list(ParadaMaquina.objects.exclude(data__isnull=True).values_list('data__year', flat=True).distinct().order_by('-data__year'))
    ano_atual = date.today().year
    if ano_atual not in anos_paradas:
        anos_paradas = [ano_atual] + anos_paradas
    anos_disponiveis_paradas = sorted(set(anos_paradas), reverse=True) if anos_paradas else [ano_atual]
    # Anos disponíveis para "Limpar por mês" (RequisicaoAlmoxarifado)
    anos_requisicoes = list(RequisicaoAlmoxarifado.objects.exclude(data_requisicao__isnull=True).values_list('data_requisicao__year', flat=True).distinct().order_by('-data_requisicao__year'))
    if ano_atual not in anos_requisicoes:
        anos_requisicoes = [ano_atual] + anos_requisicoes
    anos_disponiveis_requisicoes = sorted(set(anos_requisicoes), reverse=True) if anos_requisicoes else [ano_atual]
    # Anos disponíveis para "Limpar por mês" (NotaFiscal — data_emissao é CharField, parse em Python)
    from datetime import datetime as dt
    def _parse_data_emissao(s):
        if not s or not str(s).strip():
            return None
        s = str(s).strip()
        if ' ' in s:
            s = s.split(' ')[0]
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
            try:
                return dt.strptime(s, fmt)
            except (ValueError, TypeError):
                continue
        if '/' in s:
            parts = s.split('/')
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return dt(y, m, d)
                except (ValueError, TypeError):
                    pass
        return None
    anos_notas_set = set()
    for nota in NotaFiscal.objects.exclude(data_emissao__isnull=True).exclude(data_emissao='').only('data_emissao'):
        d = _parse_data_emissao(nota.data_emissao)
        if d:
            anos_notas_set.add(d.year)
    if ano_atual not in anos_notas_set:
        anos_notas_set.add(ano_atual)
    anos_disponiveis_notas = sorted(anos_notas_set, reverse=True) if anos_notas_set else [ano_atual]
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Setores distintos para Projeções de Gastos (excluir por setor)
    from django.db.models import Count
    setores_projecao_gasto = list(
        ProjecaoGasto.objects.values('setor').annotate(count=Count('id')).order_by('setor')
    )
    for item in setores_projecao_gasto:
        if item['setor'] is None or (isinstance(item['setor'], str) and not item['setor'].strip()):
            item['setor_display'] = '(vazio / SEM SETOR)'
            item['setor_value'] = ''
        else:
            item['setor_display'] = item['setor']
            item['setor_value'] = item['setor']
    
    context = {
        'page_title': 'Gerenciar Projeto',
        'active_page': 'gerenciar_projeto',
        'maquinas_count': maquinas_count,
        'ordens_count': ordens_count,
        'centros_count': centros_count,
        'manutentores_count': manutentores_count,
        'tabelas_info': tabelas_info,
        'total_geral': total_geral,
        'anos_disponiveis_paradas': anos_disponiveis_paradas,
        'anos_disponiveis_requisicoes': anos_disponiveis_requisicoes,
        'anos_disponiveis_notas': anos_disponiveis_notas,
        'meses_nomes': meses_nomes,
        'ano_atual': ano_atual,
        'setores_projecao_gasto': setores_projecao_gasto,
    }
    return render(request, 'administrador/gerenciar_projeto.html', context)


def limpar_tabela(request):
    """Limpar registros de uma tabela específica ou todas as tabelas"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('gerenciar_projeto')
    
    from app.models import (
        Maquina, MaquinaDocumento, OrdemServicoCorretiva, OrdemServicoCorretivaFicha,
        CentroAtividade, Semana52, Manutentor, ManutentorMaquina,
        ItemEstoque, ManutencaoCsv, ManutencaoTerceiro, MaquinaPeca,
        MaquinaPrimariaSecundaria, PlanoPreventiva, PlanoPreventivaDocumento,
        MeuPlanoPreventiva, MeuPlanoPreventivaDocumento, AgendamentoCronograma,
        RoteiroPreventiva, RequisicaoAlmoxarifado, NotaFiscal, Visitas,
        ProjecaoGasto, RelacaoProjecaoNotaFiscal, DadosOrcamento, ControleRCeNF,
        ParadaMaquina
    )
    from django.db.models import Q
    
    # Mapeamento de tabelas para modelos
    tabelas_map = {
        'maquinas': {'modelo': Maquina, 'nome': 'Máquinas'},
        'ordens': {'modelo': OrdemServicoCorretiva, 'nome': 'Ordens Corretivas'},
        'ordens_ficha': {'modelo': OrdemServicoCorretivaFicha, 'nome': 'Fichas de Ordens Corretivas'},
        'centros': {'modelo': CentroAtividade, 'nome': 'Centros de Atividade'},
        'manutentores': {'modelo': Manutentor, 'nome': 'Manutentores'},
        'manutentor_maquina': {'modelo': ManutentorMaquina, 'nome': 'Máquinas dos Manutentores'},
        'estoque': {'modelo': ItemEstoque, 'nome': 'Itens de Estoque'},
        'manutencao_csv': {'modelo': ManutencaoCsv, 'nome': 'Manutenções CSV'},
        'manutencao_terceiros': {'modelo': ManutencaoTerceiro, 'nome': 'Manutenções Terceiros'},
        'maquina_peca': {'modelo': MaquinaPeca, 'nome': 'Peças das Máquinas'},
        'maquina_primaria_secundaria': {'modelo': MaquinaPrimariaSecundaria, 'nome': 'Máquinas Primárias/Secundárias'},
        'plano_preventiva': {'modelo': PlanoPreventiva, 'nome': 'Planos Preventiva'},
        'plano_preventiva_documento': {'modelo': PlanoPreventivaDocumento, 'nome': 'Documentos Planos Preventiva'},
        'meu_plano_preventiva': {'modelo': MeuPlanoPreventiva, 'nome': 'Meus Planos Preventiva'},
        'roteiro_preventiva': {'modelo': RoteiroPreventiva, 'nome': 'Roteiros Preventiva'},
        'maquina_documento': {'modelo': MaquinaDocumento, 'nome': 'Documentos das Máquinas'},
        'semana52': {'modelo': Semana52, 'nome': 'Semanas 52'},
        'meu_plano_preventiva_documento': {'modelo': MeuPlanoPreventivaDocumento, 'nome': 'Documentos Meus Planos Preventiva'},
        'agendamento_cronograma': {'modelo': AgendamentoCronograma, 'nome': 'Agendamentos Cronograma'},
        'requisicao_almoxarifado': {'modelo': RequisicaoAlmoxarifado, 'nome': 'Requisições Almoxarifado'},
        'nota_fiscal': {'modelo': NotaFiscal, 'nome': 'Notas Fiscais'},
        'visitas': {'modelo': Visitas, 'nome': 'Visitas'},
        'projecao_gasto': {'modelo': ProjecaoGasto, 'nome': 'Projeções de Gastos'},
        'relacao_projecao_nota_fiscal': {'modelo': RelacaoProjecaoNotaFiscal, 'nome': 'Relações Projeção-Nota Fiscal'},
        'dados_orcamento': {'modelo': DadosOrcamento, 'nome': 'Dados de Orçamento'},
        'controle_rc_nf': {'modelo': ControleRCeNF, 'nome': 'Controle RC e NF'},
        'paradas_maquina': {'modelo': ParadaMaquina, 'nome': 'Paradas de Máquina'},
    }
    
    tabela = request.POST.get('tabela', '')
    
    # Tratamento especial: ParadaMaquina — limpar por mês (campo data)
    if tabela == 'paradas_maquina':
        from django.db import transaction
        deletar_todos = request.POST.get('deletar_todos_paradas') == 'on'
        if deletar_todos:
            count = ParadaMaquina.objects.count()
            if count > 0:
                with transaction.atomic():
                    ParadaMaquina.objects.all().delete()
                messages.success(request, f'{count} registro(s) de Paradas de Máquina foram removidos permanentemente.')
            else:
                messages.info(request, 'Não há registros para limpar em Paradas de Máquina.')
        else:
            try:
                ano = int(request.POST.get('ano_paradas', 0))
                meses_raw = request.POST.getlist('mes_paradas')
                meses = [int(m) for m in meses_raw if m and 1 <= int(m) <= 12]
                meses = sorted(set(meses))
                if not ano or not meses:
                    messages.error(request, 'Selecione pelo menos um ano e um mês para excluir, ou marque "Deletar TODOS os registros".')
                    return redirect('gerenciar_projeto')
                mes_conditions = Q()
                for m in meses:
                    mes_conditions |= Q(data__month=m)
                qs = ParadaMaquina.objects.filter(data__year=ano).filter(mes_conditions)
                count = qs.count()
                if count > 0:
                    with transaction.atomic():
                        qs.delete()
                    nomes_meses = {1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}
                    nomes = [nomes_meses.get(m, str(m)) for m in meses]
                    messages.success(request, f'{count} registro(s) de Paradas de Máquina ({", ".join(nomes)}/{ano}) foram removidos permanentemente.')
                else:
                    messages.info(request, 'Não há registros para os meses selecionados em Paradas de Máquina.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Ano ou mês(es) inválidos: {str(e)}')
        return redirect('gerenciar_projeto')
    
    # Tratamento especial: RequisicaoAlmoxarifado — limpar por mês (campo data_requisicao)
    if tabela == 'requisicao_almoxarifado':
        from django.db import transaction
        deletar_todos = request.POST.get('deletar_todos_requisicoes') == 'on'
        if deletar_todos:
            count = RequisicaoAlmoxarifado.objects.count()
            if count > 0:
                with transaction.atomic():
                    RequisicaoAlmoxarifado.objects.all().delete()
                messages.success(request, f'{count} registro(s) de Requisições Almoxarifado foram removidos permanentemente.')
            else:
                messages.info(request, 'Não há registros para limpar em Requisições Almoxarifado.')
        else:
            try:
                ano = int(request.POST.get('ano_requisicoes', 0))
                meses_raw = request.POST.getlist('mes_requisicoes')
                meses = [int(m) for m in meses_raw if m and 1 <= int(m) <= 12]
                meses = sorted(set(meses))
                if not ano or not meses:
                    messages.error(request, 'Selecione pelo menos um ano e um mês para excluir, ou marque "Deletar TODOS os registros".')
                    return redirect('gerenciar_projeto')
                mes_conditions = Q()
                for m in meses:
                    mes_conditions |= Q(data_requisicao__month=m)
                qs = RequisicaoAlmoxarifado.objects.filter(data_requisicao__year=ano).filter(mes_conditions)
                count = qs.count()
                if count > 0:
                    with transaction.atomic():
                        qs.delete()
                    nomes_meses = {1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}
                    nomes = [nomes_meses.get(m, str(m)) for m in meses]
                    messages.success(request, f'{count} registro(s) de Requisições Almoxarifado ({", ".join(nomes)}/{ano}) foram removidos permanentemente.')
                else:
                    messages.info(request, 'Não há registros para os meses selecionados em Requisições Almoxarifado.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Ano ou mês(es) inválidos: {str(e)}')
        return redirect('gerenciar_projeto')
    
    # Tratamento especial: NotaFiscal — limpar por mês (campo data_emissao é CharField, parse em Python)
    if tabela == 'nota_fiscal':
        from django.db import transaction
        from datetime import datetime as _dt
        def _parse_nf_date(s):
            if not s or not str(s).strip():
                return None
            s = str(s).strip()
            if ' ' in s:
                s = s.split(' ')[0]
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
                try:
                    return _dt.strptime(s, fmt)
                except (ValueError, TypeError):
                    continue
            if '/' in s:
                parts = s.split('/')
                if len(parts) == 3:
                    try:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        if y < 100:
                            y += 2000
                        return _dt(y, m, d)
                    except (ValueError, TypeError):
                        pass
            return None
        deletar_todos = request.POST.get('deletar_todos_notas') == 'on'
        if deletar_todos:
            count = NotaFiscal.objects.count()
            if count > 0:
                with transaction.atomic():
                    NotaFiscal.objects.all().delete()
                messages.success(request, f'{count} registro(s) de Notas Fiscais foram removidos permanentemente.')
            else:
                messages.info(request, 'Não há registros para limpar em Notas Fiscais.')
        else:
            try:
                ano = int(request.POST.get('ano_notas', 0))
                meses_raw = request.POST.getlist('mes_notas')
                meses = [int(m) for m in meses_raw if m and 1 <= int(m) <= 12]
                meses = sorted(set(meses))
                if not ano or not meses:
                    messages.error(request, 'Selecione pelo menos um ano e um mês para excluir, ou marque "Deletar TODOS os registros".')
                    return redirect('gerenciar_projeto')
                ids_to_delete = []
                for nota in NotaFiscal.objects.only('id', 'data_emissao'):
                    d = _parse_nf_date(nota.data_emissao)
                    if d and d.year == ano and d.month in meses:
                        ids_to_delete.append(nota.id)
                if ids_to_delete:
                    with transaction.atomic():
                        NotaFiscal.objects.filter(id__in=ids_to_delete).delete()
                    nomes_meses = {1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}
                    nomes = [nomes_meses.get(m, str(m)) for m in meses]
                    messages.success(request, f'{len(ids_to_delete)} registro(s) de Notas Fiscais ({", ".join(nomes)}/{ano}) foram removidos permanentemente.')
                else:
                    messages.info(request, 'Não há registros para os meses selecionados em Notas Fiscais.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Ano ou mês(es) inválidos: {str(e)}')
        return redirect('gerenciar_projeto')
    
    try:
        if tabela == 'todos':
            # Limpar todas as tabelas
            total_removido = 0
            detalhes = []
            
            from django.db import transaction
            with transaction.atomic():
                for key, info in tabelas_map.items():
                    count = info['modelo'].objects.count()
                    if count > 0:
                        info['modelo'].objects.all().delete()
                        # Verificar se realmente foi deletado
                        remaining_count = info['modelo'].objects.count()
                        if remaining_count == 0:
                            total_removido += count
                            detalhes.append(f"{info['nome']} ({count})")
                        else:
                            detalhes.append(f"{info['nome']} ({count - remaining_count} removidos, {remaining_count} restantes)")
            
            if total_removido > 0:
                detalhes_str = '<br>'.join([f"- {d}" for d in detalhes])
                messages.success(request, f'Todas as tabelas foram limpas. Total de {total_removido} registro(s) removidos permanentemente.<br><br>{detalhes_str}')
            else:
                messages.info(request, 'Não há registros para limpar.')
        
        elif tabela in tabelas_map:
            # Limpar tabela específica
            info = tabelas_map[tabela]
            count = info['modelo'].objects.count()
            
            if count > 0:
                try:
                    # Forçar commit da transação para garantir que a deleção seja permanente
                    from django.db import transaction
                    with transaction.atomic():
                        deleted_count, deleted_details = info['modelo'].objects.all().delete()
                    # Verificar se realmente foi deletado
                    remaining_count = info['modelo'].objects.count()
                    if remaining_count == 0:
                        messages.success(request, f'{count} registro(s) de {info["nome"]} foram removidos permanentemente do banco de dados.')
                    else:
                        messages.warning(request, f'{count - remaining_count} registro(s) de {info["nome"]} foram removidos, mas {remaining_count} registro(s) ainda permanecem no banco de dados.')
                except Exception as delete_error:
                    import traceback
                    error_details = traceback.format_exc()
                    messages.error(request, f'Erro ao deletar registros de {info["nome"]}: {str(delete_error)}<br><br>Detalhes: <pre>{error_details}</pre>')
            else:
                messages.info(request, f'Não há registros para limpar em {info["nome"]}.')
        
        else:
            messages.error(request, f'Tabela "{tabela}" não reconhecida.')
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        messages.error(request, f'Erro ao limpar tabela: {str(e)}<br><br>Detalhes: <pre>{error_details}</pre>')
    
    return redirect('gerenciar_projeto')


def deletar_projecao_gasto_por_setor(request):
    """Excluir registros de ProjecaoGasto pelos setores selecionados."""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('gerenciar_projeto')
    
    from app.models import ProjecaoGasto
    from django.db.models import Q
    
    setores = request.POST.getlist('setores')
    if not setores:
        messages.error(request, 'Selecione pelo menos um setor para excluir.')
        return redirect('gerenciar_projeto')
    
    # __EMPTY__ = setor vazio (null ou string vazia)
    q = Q()
    if '__EMPTY__' in setores:
        q |= Q(setor__isnull=True) | Q(setor='')
    setores_val = [s for s in setores if s != '__EMPTY__']
    if setores_val:
        q |= Q(setor__in=setores_val)
    
    qs = ProjecaoGasto.objects.filter(q)
    count = qs.count()
    if count == 0:
        messages.info(request, 'Nenhum registro encontrado para os setores selecionados.')
        return redirect('gerenciar_projeto')
    
    try:
        from django.db import transaction
        with transaction.atomic():
            qs.delete()
        messages.success(request, f'{count} registro(s) de Projeções de Gastos do(s) setor(es) selecionado(s) foram excluídos permanentemente.')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {str(e)}')
    return redirect('gerenciar_projeto')


def adicionar_peca_maquina(request, maquina_id):
    """Adicionar uma peça de estoque a uma máquina"""
    print(f"=== ADICIONAR PECA MAQUINA === Method: {request.method}, Maquina ID: {maquina_id}")
    print(f"POST data: {request.POST}")
    print(f"GET data: {request.GET}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        # Redirecionar para a página de peças se vier de lá, senão para visualizar
        redirect_to = request.GET.get('redirect_to', 'visualizar_maquina')
        if redirect_to == 'maquinas_pecas':
            return redirect('maquinas_pecas', maquina_id=maquina_id)
        return redirect('visualizar_maquina', maquina_id=maquina_id)
    
    from app.models import Maquina, ItemEstoque, MaquinaPeca
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    item_estoque_id = request.POST.get('item_estoque_id')
    quantidade = request.POST.get('quantidade', '1')
    observacoes = request.POST.get('observacoes', '')
    
    if not item_estoque_id:
        messages.error(request, 'Por favor, selecione um item de estoque.')
        # Redirecionar para a página de peças se vier de lá, senão para visualizar
        redirect_to = request.GET.get('redirect_to', 'visualizar_maquina')
        if redirect_to == 'maquinas_pecas':
            return redirect('maquinas_pecas', maquina_id=maquina_id)
        return redirect('visualizar_maquina', maquina_id=maquina_id)
    
    try:
        item_estoque = ItemEstoque.objects.get(id=item_estoque_id)
        
        # Verificar se já existe relação
        if MaquinaPeca.objects.filter(maquina=maquina, item_estoque=item_estoque).exists():
            messages.warning(request, f'Esta peça já está relacionada à máquina {maquina.cd_maquina}.')
        else:
            # Converter quantidade para Decimal
            from decimal import Decimal
            try:
                quantidade_decimal = Decimal(str(quantidade))
            except (ValueError, TypeError):
                quantidade_decimal = Decimal('1.0')
            
            # Criar relação
            try:
                MaquinaPeca.objects.create(
                    maquina=maquina,
                    item_estoque=item_estoque,
                    quantidade=quantidade_decimal,
                    observacoes=observacoes if observacoes else None
                )
                messages.success(request, f'Peça "{item_estoque.descricao_item or item_estoque.codigo_item}" adicionada com sucesso à máquina {maquina.cd_maquina}.')
            except Exception as create_error:
                from django.db import IntegrityError
                if isinstance(create_error, IntegrityError):
                    messages.warning(request, f'Esta peça já está relacionada à máquina {maquina.cd_maquina}.')
                else:
                    raise create_error
    
    except ItemEstoque.DoesNotExist:
        messages.error(request, 'Item de estoque não encontrado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        messages.error(request, f'Erro ao adicionar peça: {str(e)}')
        print(f"Erro ao adicionar peça: {error_detail}")  # Debug
    
    # Redirecionar para a página de peças se vier de lá, senão para visualizar
    redirect_to = request.GET.get('redirect_to', 'visualizar_maquina')
    if redirect_to == 'maquinas_pecas':
        return redirect('maquinas_pecas', maquina_id=maquina_id)
    return redirect('visualizar_maquina', maquina_id=maquina_id)


def adicionar_maquina_manutentor(request, matricula):
    """Adicionar uma máquina a um manutentor"""
    from app.models import Manutentor, ManutentorMaquina, Maquina
    from django.db import IntegrityError
    
    print(f"=== ADICIONAR MAQUINA MANUTENTOR === Method: {request.method}, Cadastro: {cadastro}")
    print(f"POST data: {request.POST}")
    print(f"GET data: {request.GET}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('visualizar_manutentor', cadastro=cadastro)
    
    try:
        manutentor = Manutentor.objects.get(Cadastro=cadastro)
    except Manutentor.DoesNotExist:
        messages.error(request, 'Manutentor não encontrado.')
        return redirect('consultar_manutentores')
    
    maquina_id = request.POST.get('maquina_id')
    observacoes = request.POST.get('observacoes', '')
    
    if not maquina_id:
        messages.error(request, 'Por favor, selecione uma máquina.')
        return redirect('visualizar_manutentor', cadastro=cadastro)
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
        
        # Verificar se já existe relação
        if ManutentorMaquina.objects.filter(manutentor=manutentor, maquina=maquina).exists():
            messages.warning(request, f'Esta máquina já está relacionada ao manutentor {manutentor.Cadastro}.')
        else:
            # Criar relação
            try:
                ManutentorMaquina.objects.create(
                    manutentor=manutentor,
                    maquina=maquina,
                    observacoes=observacoes if observacoes else None
                )
                messages.success(request, f'Máquina "{maquina.cd_maquina}" adicionada com sucesso ao manutentor {manutentor.Matricula}.')
            except Exception as create_error:
                if isinstance(create_error, IntegrityError):
                    messages.warning(request, f'Esta máquina já está relacionada ao manutentor {manutentor.Matricula}.')
                else:
                    raise create_error
    
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        messages.error(request, f'Erro ao adicionar máquina: {str(e)}')
        print(f"Erro ao adicionar máquina: {error_detail}")  # Debug
    
    return redirect('visualizar_manutentor', matricula=matricula)


def remover_maquina_manutentor(request, matricula, manutentor_maquina_id):
    """Remover uma máquina de um manutentor"""
    from app.models import Manutentor, ManutentorMaquina
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('visualizar_manutentor', matricula=matricula)
    
    try:
        manutentor = Manutentor.objects.get(Matricula=matricula)
    except Manutentor.DoesNotExist:
        messages.error(request, 'Manutentor não encontrado.')
        return redirect('consultar_manutentores')
    
    try:
        manutentor_maquina = ManutentorMaquina.objects.get(id=manutentor_maquina_id, manutentor=manutentor)
        maquina_codigo = manutentor_maquina.maquina.cd_maquina
        manutentor_maquina.delete()
        messages.success(request, f'Máquina "{maquina_codigo}" removida com sucesso do manutentor {manutentor.Matricula}.')
    except ManutentorMaquina.DoesNotExist:
        messages.error(request, 'Relação não encontrada.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        messages.error(request, f'Erro ao remover máquina: {str(e)}')
        print(f"Erro ao remover máquina: {error_detail}")  # Debug
    
    return redirect('visualizar_manutentor', cadastro=cadastro)


def remover_peca_maquina(request, maquina_id, peca_id):
    """Remover uma peça de estoque de uma máquina"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        # Redirecionar para a página de peças se vier de lá, senão para visualizar
        redirect_to = request.GET.get('redirect_to', 'visualizar_maquina')
        if redirect_to == 'maquinas_pecas':
            return redirect('maquinas_pecas', maquina_id=maquina_id)
        return redirect('visualizar_maquina', maquina_id=maquina_id)
    
    from app.models import MaquinaPeca
    
    try:
        peca = MaquinaPeca.objects.get(id=peca_id, maquina_id=maquina_id)
        item_descricao = peca.item_estoque.descricao_item or peca.item_estoque.codigo_item
        peca.delete()
        messages.success(request, f'Peça "{item_descricao}" removida com sucesso da máquina.')
    except MaquinaPeca.DoesNotExist:
        messages.error(request, 'Relação de peça não encontrada.')
    except Exception as e:
        messages.error(request, f'Erro ao remover peça: {str(e)}')
    
    # Redirecionar para a página de peças se vier de lá, senão para visualizar
    redirect_to = request.GET.get('redirect_to', 'visualizar_maquina')
    if redirect_to == 'maquinas_pecas':
        return redirect('maquinas_pecas', maquina_id=maquina_id)
    return redirect('visualizar_maquina', maquina_id=maquina_id)


def deletar_maquina(request, maquina_id):
    """Deletar uma máquina e seus relacionamentos"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('consultar_maquinas')
    
    from app.models import Maquina, ManutentorMaquina, MaquinaPeca, MaquinaPrimariaSecundaria
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
        cd_maquina = maquina.cd_maquina
        descr_maquina = maquina.descr_maquina or 'Sem descrição'
        
        # Contar relacionamentos que serão deletados
        relacionamentos_manutentor = ManutentorMaquina.objects.filter(maquina=maquina).count()
        relacionamentos_pecas = MaquinaPeca.objects.filter(maquina=maquina).count()
        relacionamentos_primaria = MaquinaPrimariaSecundaria.objects.filter(maquina_primaria=maquina).count()
        relacionamentos_secundaria = MaquinaPrimariaSecundaria.objects.filter(maquina_secundaria=maquina).count()
        
        # Deletar a máquina (os relacionamentos serão deletados automaticamente devido ao CASCADE)
        maquina.delete()
        
        # Mensagem de sucesso com detalhes
        detalhes = []
        if relacionamentos_manutentor > 0:
            detalhes.append(f'{relacionamentos_manutentor} relacionamento(s) com manutentor(es)')
        if relacionamentos_pecas > 0:
            detalhes.append(f'{relacionamentos_pecas} relacionamento(s) com peça(s)')
        if relacionamentos_primaria > 0:
            detalhes.append(f'{relacionamentos_primaria} relacionamento(s) como máquina primária')
        if relacionamentos_secundaria > 0:
            detalhes.append(f'{relacionamentos_secundaria} relacionamento(s) como máquina secundária')
        
        mensagem = f'Máquina "{cd_maquina} - {descr_maquina}" deletada com sucesso.'
        if detalhes:
            mensagem += f' Também foram removidos: {", ".join(detalhes)}.'
        
        messages.success(request, mensagem)
        
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        messages.error(request, f'Erro ao deletar máquina: {str(e)}')
        print(f"Erro ao deletar máquina: {error_detail}")
    
    # Redirecionar de volta para a página de consulta, preservando filtros
    redirect_url = 'consultar_maquinas'
    if request.GET.get('search'):
        redirect_url += f"?search={request.GET.get('search')}"
    if request.GET.get('page'):
        redirect_url += f"{'&' if '?' in redirect_url else '?'}page={request.GET.get('page')}"
    
    return redirect(redirect_url)


def atualizar_codigo_aurora(request, maquina_id):
    """Atualizar foto do código Aurora de uma máquina"""
    from app.models import Maquina
    
    print(f"=== ATUALIZAR CODIGO AURORA === Method: {request.method}, Maquina ID: {maquina_id}")
    print(f"POST data: {request.POST}")
    print(f"FILES data: {request.FILES}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('maquinas_pecas', maquina_id=maquina_id)
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    try:
        if 'codigo_aurora' in request.FILES:
            arquivo = request.FILES['codigo_aurora']
            print(f"Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Tipo: {arquivo.content_type}")
            maquina.codigo_aurora = arquivo
            maquina.save()
            print(f"Foto salva com sucesso: {maquina.codigo_aurora.url if maquina.codigo_aurora else 'N/A'}")
            messages.success(request, 'Foto do código Aurora atualizada com sucesso!')
        else:
            print("Nenhum arquivo encontrado em request.FILES")
            messages.error(request, 'Nenhum arquivo foi enviado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Erro ao atualizar código Aurora: {error_detail}")
        messages.error(request, f'Erro ao atualizar foto: {str(e)}')
    
    return redirect('maquinas_pecas', maquina_id=maquina_id)


def atualizar_codigo_fabricante(request, maquina_id):
    """Atualizar foto do código do fabricante de uma máquina"""
    from app.models import Maquina
    
    print(f"=== ATUALIZAR CODIGO FABRICANTE === Method: {request.method}, Maquina ID: {maquina_id}")
    print(f"POST data: {request.POST}")
    print(f"FILES data: {request.FILES}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('maquinas_pecas', maquina_id=maquina_id)
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    try:
        if 'codigo_fabricante' in request.FILES:
            arquivo = request.FILES['codigo_fabricante']
            print(f"Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Tipo: {arquivo.content_type}")
            maquina.codigo_fabricante = arquivo
            maquina.save()
            print(f"Foto salva com sucesso: {maquina.codigo_fabricante.url if maquina.codigo_fabricante else 'N/A'}")
            messages.success(request, 'Foto do código do fabricante atualizada com sucesso!')
        else:
            print("Nenhum arquivo encontrado em request.FILES")
            messages.error(request, 'Nenhum arquivo foi enviado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Erro ao atualizar código Fabricante: {error_detail}")
        messages.error(request, f'Erro ao atualizar foto: {str(e)}')
    
    return redirect('maquinas_pecas', maquina_id=maquina_id)


def maquinas_pecas(request, maquina_id):
    """Página para gerenciar peças de reposição de uma máquina"""
    from app.models import Maquina, ItemEstoque, MaquinaPeca
    
    try:
        maquina = Maquina.objects.get(id=maquina_id)
    except Maquina.DoesNotExist:
        messages.error(request, 'Máquina não encontrada.')
        return redirect('consultar_maquinas')
    
    # Buscar peças relacionadas a esta máquina
    pecas_relacionadas = MaquinaPeca.objects.filter(maquina=maquina).select_related('item_estoque').order_by('-created_at')
    
    # Buscar todos os itens de estoque para seleção (excluindo os já relacionados)
    itens_estoque_ids = pecas_relacionadas.values_list('item_estoque_id', flat=True)
    itens_disponiveis = ItemEstoque.objects.exclude(id__in=itens_estoque_ids).order_by('codigo_item')[:100]  # Limitar a 100 para performance
    
    context = {
        'page_title': f'Peças de Reposição - Máquina {maquina.cd_maquina}',
        'active_page': 'consultar_maquinas',
        'maquina': maquina,
        'pecas_relacionadas': pecas_relacionadas,
        'itens_disponiveis': itens_disponiveis,
    }
    return render(request, 'visualizar/visualizar_maquina_pecas.html', context)


def visualizar_item_estoque(request, item_id):
    """Visualizar detalhes de um item de estoque específico"""
    from app.models import ItemEstoque, MaquinaPeca
    
    try:
        item = ItemEstoque.objects.get(id=item_id)
    except ItemEstoque.DoesNotExist:
        messages.error(request, 'Item de estoque não encontrado.')
        return redirect('consultar_estoque')
    
    # Buscar máquinas relacionadas a este item
    maquinas_relacionadas = MaquinaPeca.objects.filter(item_estoque=item).select_related('maquina').order_by('-created_at')
    
    context = {
        'page_title': f'Visualizar Item de Estoque {item.codigo_item}',
        'active_page': 'consultar_estoque',
        'item': item,
        'maquinas_relacionadas': maquinas_relacionadas,
    }
    return render(request, 'visualizar/visualizar_item_peca.html', context)


def atualizar_foto_item(request, item_id):
    """Atualizar foto do item de estoque"""
    from app.models import ItemEstoque
    
    print(f"=== ATUALIZAR FOTO ITEM === Method: {request.method}, Item ID: {item_id}")
    print(f"POST data: {request.POST}")
    print(f"FILES data: {request.FILES}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('visualizar_item_estoque', item_id=item_id)
    
    try:
        item = ItemEstoque.objects.get(id=item_id)
    except ItemEstoque.DoesNotExist:
        messages.error(request, 'Item de estoque não encontrado.')
        return redirect('consultar_estoque')
    
    try:
        if 'foto_item' in request.FILES:
            arquivo = request.FILES['foto_item']
            print(f"Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Tipo: {arquivo.content_type}")
            item.foto_item = arquivo
            item.save()
            print(f"Foto salva com sucesso: {item.foto_item.url if item.foto_item else 'N/A'}")
            messages.success(request, 'Foto do item atualizada com sucesso!')
        else:
            print("Nenhum arquivo encontrado em request.FILES")
            messages.error(request, 'Nenhum arquivo foi enviado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Erro ao atualizar foto do item: {error_detail}")
        messages.error(request, f'Erro ao atualizar foto: {str(e)}')
    
    return redirect('visualizar_item_estoque', item_id=item_id)


def atualizar_documentacao_tecnica(request, item_id):
    """Atualizar documentação técnica do item de estoque"""
    from app.models import ItemEstoque
    
    print(f"=== ATUALIZAR DOCUMENTACAO TECNICA === Method: {request.method}, Item ID: {item_id}")
    print(f"POST data: {request.POST}")
    print(f"FILES data: {request.FILES}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('visualizar_item_estoque', item_id=item_id)
    
    try:
        item = ItemEstoque.objects.get(id=item_id)
    except ItemEstoque.DoesNotExist:
        messages.error(request, 'Item de estoque não encontrado.')
        return redirect('consultar_estoque')
    
    try:
        if 'documentacao_tecnica' in request.FILES:
            arquivo = request.FILES['documentacao_tecnica']
            print(f"Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Tipo: {arquivo.content_type}")
            item.documentacao_tecnica = arquivo
            item.save()
            print(f"Documentação salva com sucesso: {item.documentacao_tecnica.url if item.documentacao_tecnica else 'N/A'}")
            messages.success(request, 'Documentação técnica atualizada com sucesso!')
        else:
            print("Nenhum arquivo encontrado em request.FILES")
            messages.error(request, 'Nenhum arquivo foi enviado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Erro ao atualizar documentação técnica: {error_detail}")
        messages.error(request, f'Erro ao atualizar documentação: {str(e)}')
    
    return redirect('visualizar_item_estoque', item_id=item_id)


def atualizar_foto_detalhada(request, item_id):
    """Atualizar foto detalhada do item de estoque"""
    from app.models import ItemEstoque
    
    print(f"=== ATUALIZAR FOTO DETALHADA === Method: {request.method}, Item ID: {item_id}")
    print(f"POST data: {request.POST}")
    print(f"FILES data: {request.FILES}")
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('visualizar_item_estoque', item_id=item_id)
    
    try:
        item = ItemEstoque.objects.get(id=item_id)
    except ItemEstoque.DoesNotExist:
        messages.error(request, 'Item de estoque não encontrado.')
        return redirect('consultar_estoque')
    
    try:
        if 'foto_detalhada' in request.FILES:
            arquivo = request.FILES['foto_detalhada']
            print(f"Arquivo recebido: {arquivo.name}, Tamanho: {arquivo.size}, Tipo: {arquivo.content_type}")
            item.foto_detalhada = arquivo
            item.save()
            print(f"Foto detalhada salva com sucesso: {item.foto_detalhada.url if item.foto_detalhada else 'N/A'}")
            messages.success(request, 'Foto detalhada atualizada com sucesso!')
        else:
            print("Nenhum arquivo encontrado em request.FILES")
            messages.error(request, 'Nenhum arquivo foi enviado.')
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Erro ao atualizar foto detalhada: {error_detail}")
        messages.error(request, f'Erro ao atualizar foto: {str(e)}')
    
    return redirect('visualizar_item_estoque', item_id=item_id)


def dados_orcamento(request):
    """Página para gerenciar dados de orçamento por ano, mês e conta orçamentária"""
    from app.models import DadosOrcamento, ProjecaoGasto, NotaFiscal, Semana52, SaldoOrcamentarioSemanal
    from django.db.models import Count, Sum
    from collections import defaultdict
    from decimal import Decimal
    from datetime import datetime, date
    from calendar import monthrange
    
    # Processar POST para criar/atualizar registro mensal
    if request.method == 'POST' and 'action' not in request.POST:
        try:
            ano = int(request.POST.get('ano'))
            mes = int(request.POST.get('mes'))
            conta_orcamentaria = request.POST.get('conta_orcamentaria', '').strip()
            valor_orcamento_str = request.POST.get('valor_orcamento', '0').replace(',', '.')
            valor_final_desejado_str = request.POST.get('valor_final_desejado', '0').replace(',', '.')
            
            if not conta_orcamentaria:
                messages.error(request, 'Conta orçamentária é obrigatória.')
            else:
                try:
                    valor_orcamento = Decimal(valor_orcamento_str)
                    valor_final_desejado = Decimal(valor_final_desejado_str)
                    
                    # Criar ou atualizar registro
                    dados_orcamento, created = DadosOrcamento.objects.update_or_create(
                        ano=ano,
                        mes=mes,
                        conta_orcamentaria=conta_orcamentaria,
                        defaults={
                            'valor_orcamento': valor_orcamento,
                            'valor_final_desejado': valor_final_desejado,
                        }
                    )
                    
                    if created:
                        messages.success(request, f'Registro criado com sucesso para {conta_orcamentaria}!')
                    else:
                        messages.info(request, f'Registro atualizado com sucesso para {conta_orcamentaria}!')
                except ValueError:
                    messages.error(request, 'Valores inválidos. Use números válidos.')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro ao processar dados: {str(e)}')
        
        return redirect('dados_orcamento')
    
    # Processar POST para salvar saldo semanal
    if request.method == 'POST' and request.POST.get('action') == 'save_weekly_balance':
        try:
            ano = int(request.POST.get('ano'))
            mes = int(request.POST.get('mes'))
            conta_orcamentaria = request.POST.get('conta_orcamentaria', '').strip()
            semana_id = int(request.POST.get('semana_id'))
            saldo_str = request.POST.get('saldo_orcamentario_desejado', '0').replace(',', '.')
            
            if not conta_orcamentaria:
                messages.error(request, 'Conta orçamentária é obrigatória.')
            else:
                try:
                    saldo = Decimal(saldo_str)
                    semana = Semana52.objects.get(id=semana_id)
                    
                    # Criar ou atualizar saldo semanal
                    saldo_semanal, created = SaldoOrcamentarioSemanal.objects.update_or_create(
                        ano=ano,
                        mes=mes,
                        conta_orcamentaria=conta_orcamentaria,
                        semana=semana,
                        defaults={
                            'saldo_orcamentario_desejado': saldo,
                        }
                    )
                    
                    if created:
                        messages.success(request, f'Saldo semanal criado com sucesso!')
                    else:
                        messages.info(request, f'Saldo semanal atualizado com sucesso!')
                except (ValueError, Semana52.DoesNotExist) as e:
                    messages.error(request, f'Erro ao processar saldo semanal: {str(e)}')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro ao processar dados: {str(e)}')
        
        return redirect('dados_orcamento')
    
    # Buscar todos os dados de orçamento
    todos_dados = DadosOrcamento.objects.all().order_by('ano', 'mes', 'conta_orcamentaria')
    
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Anos que sempre devem aparecer (2025 e 2026)
    anos_obrigatorios = [2025, 2026]
    
    # Obter anos únicos dos dados existentes
    anos_com_dados = sorted(set(todos_dados.values_list('ano', flat=True)), reverse=True)
    
    # Combinar anos obrigatórios com anos que têm dados, removendo duplicatas
    todos_anos = sorted(list(set(anos_obrigatorios + anos_com_dados)), reverse=True)
    
    # Organizar por ano e mês, incluindo semanas
    anos_dados = []
    
    for ano in todos_anos:
        dados_ano = todos_dados.filter(ano=ano)
        
        # Buscar todas as semanas do ano
        semanas_ano = Semana52.objects.filter(
            inicio__year=ano
        ).order_by('inicio')
        
        # Sempre criar estrutura para todos os 12 meses
        meses_data = []
        for mes in range(1, 13):
            dados_mes = dados_ano.filter(mes=mes)
            
            # Buscar semanas que pertencem a este mês
            # Uma semana pertence ao mês se a data de início ou fim está no mês
            semanas_mes = []
            for semana in semanas_ano:
                if semana.inicio:
                    # Se a semana começa ou termina neste mês, ela pertence ao mês
                    if semana.inicio.month == mes or (semana.fim and semana.fim.month == mes):
                        # Buscar saldos semanais para cada registro deste mês
                        semanas_mes.append(semana)
            
            # Para cada registro de dados, buscar saldos semanais
            dados_com_saldos = []
            for dado in dados_mes:
                saldos_semanais = SaldoOrcamentarioSemanal.objects.filter(
                    ano=ano,
                    mes=mes,
                    conta_orcamentaria=dado.conta_orcamentaria
                ).select_related('semana').order_by('semana__inicio')
                
                # Criar dicionário de saldos por semana_id para fácil acesso
                saldos_dict = {}
                for saldo in saldos_semanais:
                    saldos_dict[saldo.semana_id] = {
                        'saldo': saldo,
                        'valor_formatado': f"{saldo.saldo_orcamentario_desejado:.2f}".replace('.', ',')
                    }
                
                dados_com_saldos.append({
                    'dado': dado,
                    'saldos_semanais': saldos_dict
                })
            
            meses_data.append({
                'mes': mes,
                'mes_nome': meses_nomes.get(mes, f'Mês {mes}'),
                'dados': dados_com_saldos,
                'semanas': semanas_mes
            })
        
        anos_dados.append({
            'ano': ano,
            'meses': meses_data
        })
    
    # Calcular estatísticas
    total_projecoes = ProjecaoGasto.objects.count()
    total_notas_fiscais = NotaFiscal.objects.count()
    total_requisicoes = DadosOrcamento.objects.count()
    total_relacoes_confirmadas = 0  # Placeholder - ajustar conforme necessário
    
    context = {
        'page_title': 'Dados de Orçamento',
        'active_page': 'dados_orcamento',
        'anos_dados': anos_dados,
        'total_projecoes': total_projecoes,
        'total_notas_fiscais': total_notas_fiscais,
        'total_requisicoes': total_requisicoes,
        'total_relacoes_confirmadas': total_relacoes_confirmadas,
    }
    
    return render(request, 'orcamento/dados_orcamento.html', context)


def editar_dados_orcamento(request, registro_id):
    """Editar um registro de dados de orçamento existente"""
    from app.models import DadosOrcamento
    from decimal import Decimal
    
    try:
        registro = DadosOrcamento.objects.get(id=registro_id)
    except DadosOrcamento.DoesNotExist:
        messages.error(request, 'Registro não encontrado.')
        return redirect('dados_orcamento')
    
    if request.method == 'POST':
        try:
            ano = int(request.POST.get('ano'))
            mes = int(request.POST.get('mes'))
            conta_orcamentaria = request.POST.get('conta_orcamentaria', '').strip()
            valor_orcamento_str = request.POST.get('valor_orcamento', '0').replace(',', '.')
            valor_final_desejado_str = request.POST.get('valor_final_desejado', '0').replace(',', '.')
            
            if not conta_orcamentaria:
                messages.error(request, 'Conta orçamentária é obrigatória.')
            else:
                try:
                    valor_orcamento = Decimal(valor_orcamento_str)
                    valor_final_desejado = Decimal(valor_final_desejado_str)
                    
                    # Verificar se já existe outro registro com a mesma combinação (ano, mes, conta_orcamentaria)
                    # mas com ID diferente (para evitar conflito de unique_together)
                    registro_existente = DadosOrcamento.objects.filter(
                        ano=ano,
                        mes=mes,
                        conta_orcamentaria=conta_orcamentaria
                    ).exclude(id=registro_id).first()
                    
                    if registro_existente:
                        messages.error(request, f'Já existe um registro para {conta_orcamentaria} em {ano}/{mes:02d}.')
                    else:
                        # Atualizar registro
                        registro.ano = ano
                        registro.mes = mes
                        registro.conta_orcamentaria = conta_orcamentaria
                        registro.valor_orcamento = valor_orcamento
                        registro.valor_final_desejado = valor_final_desejado
                        registro.save()
                        
                        messages.success(request, f'Registro atualizado com sucesso para {conta_orcamentaria}!')
                        return redirect('dados_orcamento')
                except ValueError:
                    messages.error(request, 'Valores inválidos. Use números válidos.')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro ao processar dados: {str(e)}')
    
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    # Formatar valores monetários para o template (formato brasileiro: 1234,56)
    # Converter Decimal para string com 2 casas decimais e substituir ponto por vírgula
    valor_orcamento_formatado = f"{registro.valor_orcamento:.2f}".replace('.', ',')
    valor_final_desejado_formatado = f"{registro.valor_final_desejado:.2f}".replace('.', ',')
    
    context = {
        'page_title': 'Editar Dados de Orçamento',
        'active_page': 'dados_orcamento',
        'registro': registro,
        'meses_nomes': meses_nomes,
        'valor_orcamento_formatado': valor_orcamento_formatado,
        'valor_final_desejado_formatado': valor_final_desejado_formatado,
    }
    
    return render(request, 'orcamento/editar_dados_orcamento.html', context)


def analise_geral_orcamento(request):
    """Página de análise geral de orçamento com filtros de ano e mês"""
    from app.models import (
        DadosOrcamento, ProjecaoGasto, NotaFiscal, 
        RequisicaoAlmoxarifado, RelacaoProjecaoNotaFiscal,
        SaldoOrcamentarioSemanal, Semana52
    )
    from django.db.models import Sum, Count, Q, Avg
    from datetime import datetime, timedelta
    from collections import defaultdict
    from decimal import Decimal
    import json
    
    # Obter filtros de ano e meses (múltiplos)
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: ano atual e mês atual (quando não há filtros na URL)
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    # Converter para inteiro
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    # Converter meses para inteiros e validar
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        # Remover duplicatas e ordenar
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    # Se não há meses selecionados, usar o mês atual
    if not meses_filtro_int:
        meses_filtro_int = [hoje.month]
    meses_para_mostrar = meses_filtro_int
    
    # Função auxiliar para parse de datas
    def parse_date(date_str):
        """Tenta fazer parse de data em vários formatos"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        if not date_str:
            return None
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        return None
    
    # ========== DADOS DE ORÇAMENTO ==========
    dados_orcamento_filtrados = DadosOrcamento.objects.filter(
        ano=ano_filtro,
        mes__in=meses_para_mostrar
    )
    total_orcamento_disponivel = dados_orcamento_filtrados.aggregate(
        total=Sum('valor_orcamento')
    )['total'] or Decimal('0')
    
    # ========== PROJEÇÕES DE GASTOS ==========
    # Ano: ano_referencia OU created_at quando ano_referencia vazio (igual analise_projecao_gastos)
    q_ano_proj = (
        Q(ano_referencia=ano_filtro) |
        (Q(ano_referencia__isnull=True) & Q(created_at__year=ano_filtro))
    )
    projecoes_filtradas = ProjecaoGasto.objects.filter(q_ano_proj)
    if meses_filtro_int:
        # mes_referencia no DB pode ser '01','02',...'12' (import) ou nome; também previsao_execucao e created_at
        meses_str = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
                     'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        meses_selecionados_str = [meses_str[m-1] for m in meses_filtro_int]
        q_mes_proj = Q()
        for mes_num in meses_filtro_int:
            q_mes_proj |= Q(mes_referencia__in=[str(mes_num), f'{mes_num:02d}'])
        for mes_str in meses_selecionados_str:
            q_mes_proj |= Q(mes_referencia__iexact=mes_str)
        meses_nomes_variantes = {
            1: ['JANEIRO', 'JAN'], 2: ['FEVEREIRO', 'FEV'], 3: ['MARÇO', 'MARCO', 'MAR'],
            4: ['ABRIL', 'ABR'], 5: ['MAIO', 'MAI'], 6: ['JUNHO', 'JUN'],
            7: ['JULHO', 'JUL'], 8: ['AGOSTO', 'AGO'], 9: ['SETEMBRO', 'SET'],
            10: ['OUTUBRO', 'OUT'], 11: ['NOVEMBRO', 'NOV'], 12: ['DEZEMBRO', 'DEZ']
        }
        for mes_num in meses_filtro_int:
            for nome in meses_nomes_variantes.get(mes_num, []):
                q_mes_proj |= Q(previsao_execucao__icontains=nome)
        for mes_num in meses_filtro_int:
            q_mes_proj |= Q(
                (Q(mes_referencia__isnull=True) | Q(mes_referencia='')) &
                (Q(previsao_execucao__isnull=True) | Q(previsao_execucao='')) &
                Q(created_at__month=mes_num)
            )
        projecoes_filtradas = projecoes_filtradas.filter(q_mes_proj)
    
    total_projecoes = projecoes_filtradas.count()
    valor_total_projecoes = projecoes_filtradas.aggregate(
        total=Sum('valor_total')
    )['total'] or Decimal('0')
    valor_total_projetado = valor_total_projecoes

    # KPIs Previsões de Gastos (para accordion na análise geral, mesmo da página analise-projecao-gastos)
    qs_servico_sim = projecoes_filtradas.filter(servico_concluido__iexact='SIM')
    projecoes_servico_concluido_sim = qs_servico_sim.count()
    valor_servico_concluido_sim = qs_servico_sim.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    percentual_servico_concluido_sim = (projecoes_servico_concluido_sim / total_projecoes * 100) if total_projecoes > 0 else 0

    qs_servico_nao = projecoes_filtradas.filter(servico_concluido__iexact='NÃO')
    projecoes_servico_concluido_nao = qs_servico_nao.count()
    valor_servico_concluido_nao = qs_servico_nao.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    percentual_servico_concluido_nao = (projecoes_servico_concluido_nao / total_projecoes * 100) if total_projecoes > 0 else 0

    qs_servico_indefinido = projecoes_filtradas.exclude(servico_concluido__iexact='SIM').exclude(servico_concluido__iexact='NÃO')
    projecoes_servico_concluido_indefinido = qs_servico_indefinido.count()
    valor_servico_concluido_indefinido = qs_servico_indefinido.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    percentual_servico_concluido_indefinido = (projecoes_servico_concluido_indefinido / total_projecoes * 100) if total_projecoes > 0 else 0

    qs_com_numero_nf = projecoes_filtradas.exclude(numero_nf__isnull=True).exclude(numero_nf='')
    projecoes_com_numero_nf = qs_com_numero_nf.count()
    valor_com_numero_nf = qs_com_numero_nf.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    percentual_com_numero_nf = (projecoes_com_numero_nf / total_projecoes * 100) if total_projecoes > 0 else 0

    valor_orcamento_periodo = total_orcamento_disponivel  # já calculado abaixo em DadosOrcamento
    percentual_impacto_orcamento = None
    if total_orcamento_disponivel and total_orcamento_disponivel > 0:
        percentual_impacto_orcamento = float((valor_total_projetado / total_orcamento_disponivel) * 100)

    # ========== CORRELAÇÃO PROJEÇÕES E NOTAS FISCAIS ==========
    from app.utils import find_matching_notas_fiscais
    
    projecoes_com_nota_match = 0
    projecoes_sem_nota_match = 0
    valor_projecoes_com_match = Decimal('0')
    valor_projecoes_sem_match = Decimal('0')
    
    # Analisar cada projeção para encontrar matches
    for projecao in projecoes_filtradas:
        matches = find_matching_notas_fiscais(projecao)
        if matches:
            projecoes_com_nota_match += 1
            valor_projecoes_com_match += (projecao.valor_total or Decimal('0'))
        else:
            projecoes_sem_nota_match += 1
            valor_projecoes_sem_match += (projecao.valor_total or Decimal('0'))
    
    percentual_projecoes_com_match = (projecoes_com_nota_match / total_projecoes * 100) if total_projecoes > 0 else 0
    
    # Projeções com notas fiscais relacionadas
    projecoes_com_nf = projecoes_filtradas.filter(
        relacoes_notas_fiscais__isnull=False
    ).distinct().count()
    valor_projecoes_com_nf = RelacaoProjecaoNotaFiscal.objects.filter(
        projecao__in=projecoes_filtradas
    ).aggregate(
        total=Sum('projecao__valor_total')
    )['total'] or Decimal('0')
    
    # ========== NOTAS FISCAIS ==========
    # IMPORTANTE: Usar APENAS data_emissao para determinar o mês de pagamento previsto
    # A data_emissao indica em qual mês a nota deve ser prevista para pagamento
    notas_filtradas = []
    todas_notas = NotaFiscal.objects.all()
    for nota in todas_notas:
        data_emissao = parse_date(nota.data_emissao)
        # Usar APENAS data_emissao para determinar o mês
        if data_emissao and data_emissao.year == ano_filtro:
            if not meses_filtro_int or data_emissao.month in meses_filtro_int:
                notas_filtradas.append(nota)
    
    total_notas = len(notas_filtradas)
    valor_total_notas = sum(
        (nota.total_nota or Decimal('0')) for nota in notas_filtradas
    )
    
    # Notas fiscais com situacao = "LANÇADA" e uso_contabil = "242" 
    # IMPORTANTE: Usar APENAS data_emissao para determinar o mês de pagamento previsto
    # A data_emissao indica em qual mês a nota deve ser prevista para pagamento
    notas_autorizadas_242_filtradas = []
    todas_notas_autorizadas = NotaFiscal.objects.filter(
        Q(situacao__icontains='LANÇADA') | Q(situacao__icontains='LANCADA')
    ).filter(
        uso_contabil='242'
    )
    
    for nota in todas_notas_autorizadas:
        # Usar APENAS data_emissao para determinar o mês
        data_emissao = parse_date(nota.data_emissao)
        
        # Se não houver filtro de mês, incluir todas as notas do ano (ou sem data)
        if not meses_filtro_int:
            if not data_emissao:
                # Se não há data_emissao, incluir (assumindo que é do ano filtrado)
                notas_autorizadas_242_filtradas.append(nota)
            elif data_emissao.year == ano_filtro:
                notas_autorizadas_242_filtradas.append(nota)
        else:
            # Filtrar por ano e mês específicos baseado em data_emissao
            if data_emissao and data_emissao.year == ano_filtro:
                if data_emissao.month in meses_filtro_int:
                    notas_autorizadas_242_filtradas.append(nota)
    
    valor_total_notas_lancadas = sum(
        (nota.total_nota or Decimal('0')) for nota in notas_autorizadas_242_filtradas
    )
    
    # Notas relacionadas a projeções
    notas_relacionadas = RelacaoProjecaoNotaFiscal.objects.filter(
        projecao__in=projecoes_filtradas
    ).values_list('nota_fiscal_id', flat=True).distinct()
    total_notas_relacionadas = len(notas_relacionadas)
    valor_notas_relacionadas = sum(
        (nota.total_nota or Decimal('0')) 
        for nota in NotaFiscal.objects.filter(id__in=notas_relacionadas)
    )
    
    # ========== REQUISIÇÕES ALMOXARIFADO ==========
    requisicoes_filtradas = RequisicaoAlmoxarifado.objects.filter(
        data_requisicao__year=ano_filtro,
        data_requisicao__month__in=meses_para_mostrar
    )
    total_requisicoes = requisicoes_filtradas.count()
    # IMPORTANTE: Apenas considerar cd_depo == 1 para custos (itens novos que geram gasto)
    # cd_depo == 3 são itens reutilizados que não geram custo
    valor_total_requisicoes = Decimal('0.00')
    for req in requisicoes_filtradas.filter(cd_depo=1):
        if req.vlr_movto_estoq:
            # vlr_movto_estoq já representa o valor total da transação (pode ser negativo para saídas)
            valor_total_requisicoes += abs(req.vlr_movto_estoq)
    
    # Valor total de requisições para o card (já filtrado por ano/mês)
    valor_total_requisicoes_filtrado = valor_total_requisicoes

    # Requisições associadas a Notas Fiscais (mesmo critério da página analise-requisicoes)
    mensagem_nf_requisicao = "REQUISIÇÃO GERADA PELO PROCESSO DE ENTRADA DE NOTA FISCAL"
    requisicoes_com_nf = requisicoes_filtradas.filter(obs_rm__icontains=mensagem_nf_requisicao)
    total_requisicoes_com_nf = requisicoes_com_nf.count()
    valor_total_requisicoes_com_nf = Decimal('0.00')
    for req in requisicoes_com_nf.filter(cd_depo=1):
        if req.vlr_movto_estoq:
            valor_total_requisicoes_com_nf += abs(req.vlr_movto_estoq)

    # Requisições sem associação a Nota Fiscal (excluir as que têm a mensagem NF)
    requisicoes_sem_nf = requisicoes_filtradas.exclude(obs_rm__icontains=mensagem_nf_requisicao)
    total_requisicoes_sem_nf = requisicoes_sem_nf.count()
    valor_total_requisicoes_sem_nf = Decimal('0.00')
    for req in requisicoes_sem_nf.filter(cd_depo=1):
        if req.vlr_movto_estoq:
            valor_total_requisicoes_sem_nf += abs(req.vlr_movto_estoq)

    # Percentual que Requisições sem NF representam sobre o orçamento total (para o último box)
    percentual_requisicoes_sem_nf_sobre_orcamento = 0
    if total_orcamento_disponivel and total_orcamento_disponivel > 0:
        percentual_requisicoes_sem_nf_sobre_orcamento = float(
            (valor_total_requisicoes_sem_nf / total_orcamento_disponivel) * 100
        )

    # Percentual que Valor NF Lançadas representa sobre o orçamento total (último box da linha Notas Fiscais)
    percentual_notas_lancadas_sobre_orcamento = 0
    if total_orcamento_disponivel and total_orcamento_disponivel > 0:
        percentual_notas_lancadas_sobre_orcamento = float(
            (valor_total_notas_lancadas / total_orcamento_disponivel) * 100
        )

    # ========== KPIs ==========
    # Total de gastos (projeções + notas + requisições)
    total_gastos = valor_total_projecoes + valor_total_notas + valor_total_requisicoes
    
    # Percentual do orçamento utilizado
    percentual_utilizado = 0
    if total_orcamento_disponivel > 0:
        percentual_utilizado = float((total_gastos / total_orcamento_disponivel) * 100)
    
    # Saldo disponível
    saldo_disponivel = total_orcamento_disponivel - total_gastos
    
    # Saldo Parcial = Orçamento Disponível − Requisições sem NF − Gastos Previstos (Serv. Concluído NÃO) − Valor NF Lançadas
    saldo_parcial = total_orcamento_disponivel - valor_total_requisicoes_sem_nf - valor_servico_concluido_nao - valor_total_notas_lancadas
    
    # Percentual de projeções com NF relacionada
    percentual_projecoes_com_nf = 0
    if total_projecoes > 0:
        percentual_projecoes_com_nf = (projecoes_com_nf / total_projecoes) * 100
    
    # ========== EVOLUÇÃO TEMPORAL (DIÁRIA CUMULATIVA) ==========
    from calendar import monthrange
    
    dias_labels = []
    dias_orcamento = []
    dias_requisicoes = []
    dias_notas_fiscais = []
    dias_saldo = []
    
    # Criar dicionário de orçamento por mês (para distribuir nos dias)
    orcamento_por_mes = {}
    for mes in meses_para_mostrar:
        orc_mes = DadosOrcamento.objects.filter(
            ano=ano_filtro, mes=mes
        ).aggregate(total=Sum('valor_orcamento'))['total'] or Decimal('0')
        orcamento_por_mes[mes] = orc_mes
    
    # Variáveis acumulativas
    req_acumulado = Decimal('0')
    notas_acumulado = Decimal('0')
    
    # Gerar todos os dias dos meses filtrados
    for mes in meses_para_mostrar:
        # Obter número de dias no mês
        num_dias = monthrange(ano_filtro, mes)[1]
        
        for dia in range(1, num_dias + 1):
            data_atual = datetime(ano_filtro, mes, dia)
            dias_labels.append(data_atual.strftime('%d/%m'))
            
            # Orçamento do dia (usar o valor mensal para todos os dias do mês - fixo)
            orc_dia = orcamento_por_mes.get(mes, Decimal('0'))
            dias_orcamento.append(float(orc_dia))
            
            # Requisições do dia (RequisicaoAlmoxarifado) - valor do dia
            # IMPORTANTE: Apenas considerar cd_depo == 1 para custos (itens novos que geram gasto)
            req_dia = Decimal('0.00')
            for req in RequisicaoAlmoxarifado.objects.filter(
                data_requisicao__year=ano_filtro,
                data_requisicao__month=mes,
                data_requisicao__day=dia,
                cd_depo=1  # Apenas itens que geram custo
            ):
                if req.vlr_movto_estoq:
                    req_dia += abs(req.vlr_movto_estoq)
            # Acumular
            req_acumulado += req_dia
            dias_requisicoes.append(float(req_acumulado))
            
            # Notas Fiscais do dia (situacao="LANÇADA" e uso_contabil="242") - valor do dia
            # IMPORTANTE: Usar APENAS data_emissao para determinar o dia/mês de pagamento previsto
            notas_dia = Decimal('0')
            todas_notas_dia = NotaFiscal.objects.filter(
                Q(situacao__icontains='LANÇADA') | Q(situacao__icontains='LANCADA')
            ).filter(
                uso_contabil='242'
            )
            
            for nota in todas_notas_dia:
                # Usar data_emissao para determinar o dia/mês
                data_emissao = parse_date(nota.data_emissao)
                if data_emissao and data_emissao.year == ano_filtro and data_emissao.month == mes and data_emissao.day == dia:
                    notas_dia += (nota.total_nota or Decimal('0'))
            # Acumular
            notas_acumulado += notas_dia
            dias_notas_fiscais.append(float(notas_acumulado))
            
            # Saldo do dia (Orçamento - Requisições Acumuladas - Notas Fiscais Acumuladas)
            saldo_dia = orc_dia - req_acumulado - notas_acumulado
            dias_saldo.append(float(saldo_dia))
    
    # ========== SALDO ORÇAMENTÁRIO SEMANAL ==========
    # Buscar todos os SaldoOrcamentarioSemanal para o ano e meses filtrados
    saldos_semanais_filtrados = SaldoOrcamentarioSemanal.objects.filter(
        ano=ano_filtro,
        mes__in=meses_para_mostrar
    ).select_related('semana')
    
    # Agrupar por semana (usando a data de início da semana) e somar os valores
    # Criar um dicionário: (data_inicio_semana, data_fim_semana) -> soma dos saldos
    saldos_por_semana = {}
    
    for saldo_semanal in saldos_semanais_filtrados:
        if saldo_semanal.semana and saldo_semanal.semana.inicio and saldo_semanal.semana.fim:
            # Usar tupla (inicio, fim) como chave
            chave_semana = (saldo_semanal.semana.inicio, saldo_semanal.semana.fim)
            # Verificar se a semana pertence ao ano filtrado
            if saldo_semanal.semana.inicio.year == ano_filtro:
                if chave_semana not in saldos_por_semana:
                    saldos_por_semana[chave_semana] = Decimal('0')
                saldos_por_semana[chave_semana] += (saldo_semanal.saldo_orcamentario_desejado or Decimal('0'))
    
    # Criar lista de saldos semanais para cada dia
    # Para cada dia, verificar se ele pertence a alguma semana e usar o valor correspondente
    dias_saldo_semanal = []
    
    for mes in meses_para_mostrar:
        num_dias = monthrange(ano_filtro, mes)[1]
        for dia in range(1, num_dias + 1):
            data_atual = datetime(ano_filtro, mes, dia).date()
            
            # Encontrar a semana que contém este dia
            saldo_semana_atual = Decimal('0')
            for (data_inicio_semana, data_fim_semana), valor_semana in saldos_por_semana.items():
                if data_inicio_semana <= data_atual <= data_fim_semana:
                    saldo_semana_atual = valor_semana
                    break
            
            dias_saldo_semanal.append(float(saldo_semana_atual))
    
    # Manter variáveis antigas para compatibilidade
    meses_labels = dias_labels
    meses_orcamento = dias_orcamento
    meses_requisicoes = dias_requisicoes
    meses_notas_fiscais = dias_notas_fiscais
    meses_saldo = dias_saldo
    
    # ========== PREPARAR NOTAS PARA GRÁFICOS ==========
    # Combinar todas as notas filtradas por ano/mês
    # Incluir notas filtradas por data_emissao/vencimento E notas filtradas por data_autorizacao
    notas_para_graficos_ids = set()
    notas_para_graficos = []
    
    # Adicionar notas filtradas por data_autorizacao (LANÇADAS com uso 242)
    for nota in notas_autorizadas_242_filtradas:
        if nota.id not in notas_para_graficos_ids:
            notas_para_graficos.append(nota)
            notas_para_graficos_ids.add(nota.id)
    
    # Adicionar outras notas filtradas por data_emissao (evitando duplicatas)
    for nota in notas_filtradas:
        if nota.id not in notas_para_graficos_ids:
            notas_para_graficos.append(nota)
            notas_para_graficos_ids.add(nota.id)
    
    # ========== DISTRIBUIÇÃO POR SETOR ==========
    # Combinar dados de projeções, notas fiscais e requisições filtradas por ano/mês
    setores_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0')})
    
    # Projeções (já filtradas por ano/mês)
    for proj in projecoes_filtradas:
        if proj.setor and proj.setor.strip():
            setor = proj.setor.strip()[:30]
            setores_dict[setor]['count'] += 1
            setores_dict[setor]['valor'] += (proj.valor_total or Decimal('0'))
    
    # Notas fiscais - usar centro_atividade como setor (se disponível)
    for nota in notas_para_graficos:
        setor = None
        if nota.centro_atividade and nota.centro_atividade.strip():
            setor = nota.centro_atividade.strip()[:30]
        elif nota.nome_centro_atividade and nota.nome_centro_atividade.strip():
            setor = nota.nome_centro_atividade.strip()[:30]
        
        if setor:
            setores_dict[setor]['count'] += 1
            setores_dict[setor]['valor'] += (nota.total_nota or Decimal('0'))
    
    # Requisições - não têm setor diretamente, mas podem ter informações relacionadas
    # (por enquanto, não incluímos requisições no gráfico de setores)
    
    # Ordenar por valor e pegar top 10
    setores_sorted = sorted(
        setores_dict.items(),
        key=lambda x: x[1]['valor'],
        reverse=True
    )[:10]
    
    # Garantir que sempre temos listas (mesmo que vazias)
    setores_labels = [item[0] for item in setores_sorted] if setores_sorted else []
    setores_data = [float(item[1]['valor']) for item in setores_sorted] if setores_sorted else []
    
    # ========== DISTRIBUIÇÃO POR USO CONTÁBIL ==========
    # Usar apenas dados filtrados por ano/mês
    uso_contabil_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0')})
    
    # Projeções (já filtradas por ano/mês)
    for proj in projecoes_filtradas:
        if proj.uso_contabil:
            uso_contabil_dict[proj.uso_contabil]['count'] += 1
            uso_contabil_dict[proj.uso_contabil]['valor'] += (proj.valor_total or Decimal('0'))
    
    # Notas fiscais - usar as notas já preparadas acima
    for nota in notas_para_graficos:
        if nota.uso_contabil:
            uso_contabil_dict[nota.uso_contabil]['count'] += 1
            uso_contabil_dict[nota.uso_contabil]['valor'] += (nota.total_nota or Decimal('0'))
    
    for nota in notas_para_graficos:
        if nota.uso_contabil:
            uso_contabil_dict[nota.uso_contabil]['count'] += 1
            uso_contabil_dict[nota.uso_contabil]['valor'] += (nota.total_nota or Decimal('0'))
    
    # Requisições (já filtradas por ano/mês)
    for req in requisicoes_filtradas:
        if req.descr_uso_ctb:
            uso_contabil_dict[req.descr_uso_ctb]['count'] += 1
            valor_req = req.vlr_movto_estoq or Decimal('0')
            if valor_req < 0:
                valor_req = abs(valor_req)
            uso_contabil_dict[req.descr_uso_ctb]['valor'] += valor_req
    
    uso_contabil_sorted = sorted(
        uso_contabil_dict.items(),
        key=lambda x: x[1]['valor'],
        reverse=True
    )[:10]
    
    uso_contabil_labels = [item[0][:30] for item in uso_contabil_sorted]
    uso_contabil_data = [float(item[1]['valor']) for item in uso_contabil_sorted]
    
    # ========== TOP FORNECEDORES ==========
    # Usar apenas dados filtrados por ano/mês
    fornecedores_dict = defaultdict(lambda: {'count': 0, 'valor': Decimal('0')})
    
    # Projeções (já filtradas por ano/mês)
    for proj in projecoes_filtradas:
        if proj.fornecedor_nome_fantasia:
            fornecedores_dict[proj.fornecedor_nome_fantasia]['count'] += 1
            fornecedores_dict[proj.fornecedor_nome_fantasia]['valor'] += (proj.valor_total or Decimal('0'))
    
    # Notas fiscais - usar as mesmas notas filtradas usadas no gráfico de uso contábil
    for nota in notas_para_graficos:
        if nota.nome_fantasia_emitente:
            fornecedores_dict[nota.nome_fantasia_emitente]['count'] += 1
            fornecedores_dict[nota.nome_fantasia_emitente]['valor'] += (nota.total_nota or Decimal('0'))
    
    fornecedores_sorted = sorted(
        fornecedores_dict.items(),
        key=lambda x: x[1]['valor'],
        reverse=True
    )[:10]
    
    fornecedores_labels = [item[0][:50] for item in fornecedores_sorted]
    fornecedores_data = [float(item[1]['valor']) for item in fornecedores_sorted]
    
    # ========== ANOS DISPONÍVEIS ==========
    anos_orcamento = sorted(
        DadosOrcamento.objects.values_list('ano', flat=True).distinct(),
        reverse=True
    )
    anos_projecoes = sorted(
        ProjecaoGasto.objects.exclude(ano_referencia__isnull=True).values_list('ano_referencia', flat=True).distinct(),
        reverse=True
    )
    anos_requisicoes = sorted(
        RequisicaoAlmoxarifado.objects.exclude(data_requisicao__isnull=True).values_list('data_requisicao__year', flat=True).distinct(),
        reverse=True
    )
    anos_disponiveis = sorted(
        set(anos_orcamento + anos_projecoes + anos_requisicoes),
        reverse=True
    )
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    context = {
        'page_title': 'Análise Geral de Orçamento',
        'active_page': 'analise_geral_orcamento',
        'total_requisicoes_com_nf': total_requisicoes_com_nf,
        'valor_total_requisicoes_com_nf': valor_total_requisicoes_com_nf,
        'total_requisicoes_sem_nf': total_requisicoes_sem_nf,
        'valor_total_requisicoes_sem_nf': valor_total_requisicoes_sem_nf,
        'percentual_requisicoes_sem_nf_sobre_orcamento': percentual_requisicoes_sem_nf_sobre_orcamento,
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
        # KPIs
        'total_orcamento_disponivel': total_orcamento_disponivel,
        'total_gastos': total_gastos,
        'saldo_disponivel': saldo_disponivel,
        'percentual_utilizado': percentual_utilizado,
        'total_projecoes': total_projecoes,
        'valor_total_projecoes': valor_total_projecoes,
        'valor_total_projetado': valor_total_projetado,
        'valor_servico_concluido_sim': valor_servico_concluido_sim,
        'projecoes_servico_concluido_sim': projecoes_servico_concluido_sim,
        'percentual_servico_concluido_sim': percentual_servico_concluido_sim,
        'valor_servico_concluido_nao': valor_servico_concluido_nao,
        'projecoes_servico_concluido_nao': projecoes_servico_concluido_nao,
        'percentual_servico_concluido_nao': percentual_servico_concluido_nao,
        'valor_servico_concluido_indefinido': valor_servico_concluido_indefinido,
        'projecoes_servico_concluido_indefinido': projecoes_servico_concluido_indefinido,
        'percentual_servico_concluido_indefinido': percentual_servico_concluido_indefinido,
        'valor_com_numero_nf': valor_com_numero_nf,
        'projecoes_com_numero_nf': projecoes_com_numero_nf,
        'percentual_com_numero_nf': percentual_com_numero_nf,
        'percentual_impacto_orcamento': percentual_impacto_orcamento,
        'valor_orcamento_periodo': valor_orcamento_periodo,
        'projecoes_com_nf': projecoes_com_nf,
        'percentual_projecoes_com_nf': percentual_projecoes_com_nf,
        'total_notas': total_notas,
        'valor_total_notas': valor_total_notas,
        'total_notas_relacionadas': total_notas_relacionadas,
        'valor_notas_relacionadas': valor_notas_relacionadas,
        'total_requisicoes': total_requisicoes,
        'valor_total_requisicoes': valor_total_requisicoes,
        'valor_total_requisicoes_filtrado': valor_total_requisicoes_filtrado,
        'valor_total_notas_lancadas': valor_total_notas_lancadas,
        'percentual_notas_lancadas_sobre_orcamento': percentual_notas_lancadas_sobre_orcamento,
        'saldo_parcial': saldo_parcial,
        # Gráficos
        'meses_labels': json.dumps(meses_labels, ensure_ascii=False),
        'meses_orcamento': json.dumps(meses_orcamento, ensure_ascii=False),
        'meses_requisicoes': json.dumps(meses_requisicoes, ensure_ascii=False),
        'meses_notas_fiscais': json.dumps(meses_notas_fiscais, ensure_ascii=False),
        'meses_saldo': json.dumps(meses_saldo, ensure_ascii=False),
        'meses_saldo_semanal': json.dumps(dias_saldo_semanal, ensure_ascii=False),
        'setores_labels': json.dumps(setores_labels, ensure_ascii=False),
        'setores_data': json.dumps(setores_data, ensure_ascii=False),
        'uso_contabil_labels': json.dumps(uso_contabil_labels, ensure_ascii=False),
        'uso_contabil_data': json.dumps(uso_contabil_data, ensure_ascii=False),
        'fornecedores_labels': json.dumps(fornecedores_labels, ensure_ascii=False),
        'fornecedores_data': json.dumps(fornecedores_data, ensure_ascii=False),
    }
    
    return render(request, 'orcamento/analise_geral_orcamento.html', context)


def consultar_projecao_gastos(request):
    """Consultar/listar projeções de gastos com filtros avançados - apenas tabela"""
    from app.models import ProjecaoGasto
    from decimal import Decimal
    from django.db.models import Q
    from django.core.paginator import Paginator
    from datetime import datetime
    
    # ========== FILTRO DE ANO E MÊS (igual analise_geral_orcamento) ==========
    # Verificar se o usuário quer mostrar todos os registros sem filtros
    mostrar_todos = request.GET.get('mostrar_todos', 'false').lower() == 'true'
    
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    
    # Valores padrão: usar o ano mais recente dos dados ou ano atual
    hoje = datetime.now()
    
    # Anos disponíveis para o filtro (de ano_referencia E de previsao_execucao)
    import re
    anos_disponiveis_set = set()
    anos_ref = ProjecaoGasto.objects.exclude(ano_referencia__isnull=True).values_list('ano_referencia', flat=True).distinct()
    for ano in anos_ref:
        if ano:
            anos_disponiveis_set.add(int(ano))
    # Extrair anos de previsao_execucao (ex: "DEZEMBRO / 2025", "OUTUBRO 2024")
    prev_exec_values = ProjecaoGasto.objects.exclude(previsao_execucao__isnull=True).exclude(
        previsao_execucao=''
    ).values_list('previsao_execucao', flat=True).distinct()
    for texto in prev_exec_values:
        if texto:
            ano_match = re.search(r'\b(20\d{2}|19\d{2})\b', str(texto))
            if ano_match:
                anos_disponiveis_set.add(int(ano_match.group(1)))
    anos_disponiveis_set.add(hoje.year)
    anos_disponiveis = sorted(list(anos_disponiveis_set), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    # Se mostrar_todos está ativo, não aplicar filtros de ano/mês
    # Aplicar ano/mês só quando o usuário enviou 'ano' na requisição (ex.: pelo accordion); assim, ao usar só o filtro de Setor na tabela, não aplicamos ano/mês
    ano_mes_explicito = 'ano' in request.GET
    if mostrar_todos:
        ano_filtro = None
        meses_filtro_int = []
        meses_para_mostrar = []
    else:
        # Se não há filtro de ano, usar o ano mais recente dos dados (ou ano atual se não houver dados) apenas para exibição
        if not ano_filtro:
            if anos_disponiveis:
                ano_filtro = anos_disponiveis[0]  # Usar o ano mais recente
            else:
                ano_filtro = hoje.year
        
        # Converter para inteiro
        try:
            ano_filtro = int(ano_filtro)
        except (ValueError, TypeError):
            if anos_disponiveis:
                ano_filtro = anos_disponiveis[0]
            else:
                ano_filtro = hoje.year
        
        # Converter meses para inteiros e validar
        meses_filtro_int = []
        if meses_filtro:
            for mes in meses_filtro:
                try:
                    mes_int = int(mes)
                    if 1 <= mes_int <= 12:
                        meses_filtro_int.append(mes_int)
                except (ValueError, TypeError):
                    continue
            # Remover duplicatas e ordenar
            meses_filtro_int = sorted(list(set(meses_filtro_int)))
        
        # Se não há meses selecionados, usar todos os meses (só quando ano foi enviado)
        meses_para_mostrar = meses_filtro_int if meses_filtro_int else list(range(1, 13))
    
    # ========== FILTROS DE PROJEÇÕES ==========
    try:
        # Busca geral
        search_query = request.GET.get('search', '').strip()
        # Filtrar apenas registros com dados significativos (não apenas ID)
        # Incluir apenas registros que têm pelo menos um campo principal preenchido
        projecoes_list = ProjecaoGasto.objects.filter(
            Q(setor__isnull=False) & ~Q(setor='') |
            Q(descricao__isnull=False) & ~Q(descricao='') |
            Q(valor_total__isnull=False) & ~Q(valor_total=0) |
            Q(fornecedor_nome_fantasia__isnull=False) & ~Q(fornecedor_nome_fantasia='')
        )
    except Exception as e:
        import traceback
        print(f"Erro ao acessar ProjecaoGasto: {e}")
        print(traceback.format_exc())
        projecoes_list = ProjecaoGasto.objects.none()
        search_query = ''
    
    # Aplicar filtro de ano/mês apenas quando o usuário enviou 'ano' na URL (ex.: filtro do accordion); não aplicar quando só usa filtros da tabela (ex.: Setor)
    if not mostrar_todos and ano_mes_explicito:
        q_ano = Q(ano_referencia=ano_filtro) | Q(previsao_execucao__icontains=str(ano_filtro))
        projecoes_list = projecoes_list.filter(q_ano)
        
        # Aplicar filtro de meses baseado em previsao_execucao
        if meses_filtro_int:
            # Mapear números de mês para nomes em português (usado em previsao_execucao)
            meses_nomes = {
            1: ['JANEIRO', 'JAN'],
            2: ['FEVEREIRO', 'FEV'],
            3: ['MARÇO', 'MARCO', 'MAR'],
            4: ['ABRIL', 'ABR'],
            5: ['MAIO', 'MAI'],
            6: ['JUNHO', 'JUN'],
            7: ['JULHO', 'JUL'],
            8: ['AGOSTO', 'AGO'],
            9: ['SETEMBRO', 'SET'],
            10: ['OUTUBRO', 'OUT'],
            11: ['NOVEMBRO', 'NOV'],
                12: ['DEZEMBRO', 'DEZ']
            }
            
            # Criar Q object para filtrar por mes_referencia (extraído durante import)
            meses_selecionados_str = []
            for m in meses_filtro_int:
                meses_selecionados_str.append(str(m))  # Formato sem zero: '1', '2', ..., '12'
                meses_selecionados_str.append(f"{m:02d}")  # Formato com zero: '01', '02', ..., '12'
            meses_selecionados_str = list(set(meses_selecionados_str))
            
            # Criar Q object para filtrar por previsao_execucao diretamente (fallback)
            q_previsao = Q()
            for m in meses_filtro_int:
                nomes_mes = meses_nomes.get(m, [])
                for nome in nomes_mes:
                    # Buscar por nome do mês em previsao_execucao (case-insensitive)
                    q_previsao |= Q(previsao_execucao__icontains=nome)
            
            # Filtrar: mes_referencia OU previsao_execucao contém o mês
            q_mes = Q(mes_referencia__in=meses_selecionados_str) | q_previsao
            projecoes_list = projecoes_list.filter(q_mes)
    
    # Aplicar busca geral (inclui previsao_execucao para buscar por "DEZEMBRO 2025", etc.)
    if search_query:
        try:
            search_num = Decimal(search_query.replace(',', '.'))
            projecoes_list = projecoes_list.filter(
                Q(setor__icontains=search_query) |
                Q(descricao__icontains=search_query) |
                Q(fornecedor_nome_fantasia__icontains=search_query) |
                Q(numero_requisicao_compra__icontains=search_query) |
                Q(numero_ordem_servico__icontains=search_query) |
                Q(previsao_execucao__icontains=search_query) |
                Q(valor_total=search_num)
            )
        except (ValueError, TypeError):
            try:
                search_id = int(search_query)
                projecoes_list = projecoes_list.filter(
                    Q(id_excel=search_id) |
                    Q(setor__icontains=search_query) |
                    Q(descricao__icontains=search_query) |
                    Q(fornecedor_nome_fantasia__icontains=search_query) |
                    Q(numero_requisicao_compra__icontains=search_query) |
                    Q(numero_ordem_servico__icontains=search_query) |
                    Q(previsao_execucao__icontains=search_query) |
                    Q(solicitante__icontains=search_query)
                )
            except (ValueError, TypeError):
                projecoes_list = projecoes_list.filter(
                    Q(setor__icontains=search_query) |
                    Q(descricao__icontains=search_query) |
                    Q(fornecedor_nome_fantasia__icontains=search_query) |
                    Q(numero_requisicao_compra__icontains=search_query) |
                    Q(numero_ordem_servico__icontains=search_query) |
                    Q(previsao_execucao__icontains=search_query) |
                    Q(solicitante__icontains=search_query)
                )
    
    # Filtros específicos (mantidos para compatibilidade com filtros da tabela)
    filtro_setor = request.GET.get('filtro_setor', '').strip()
    if filtro_setor:
        projecoes_list = projecoes_list.filter(setor__icontains=filtro_setor)
    
    filtro_fornecedor = request.GET.get('filtro_fornecedor', '').strip()
    if filtro_fornecedor:
        projecoes_list = projecoes_list.filter(
            Q(fornecedor_nome_fantasia__icontains=filtro_fornecedor) |
            Q(fornecedor_cnpj__icontains=filtro_fornecedor)
        )
    
    filtro_descricao = request.GET.get('filtro_descricao', '').strip()
    if filtro_descricao:
        projecoes_list = projecoes_list.filter(descricao__icontains=filtro_descricao)

    filtro_uso_contabil = request.GET.get('filtro_uso_contabil', '').strip()
    if filtro_uso_contabil:
        projecoes_list = projecoes_list.filter(uso_contabil__icontains=filtro_uso_contabil)

    filtro_servico_concluido = request.GET.get('filtro_servico_concluido', '').strip()
    if filtro_servico_concluido:
        projecoes_list = projecoes_list.filter(servico_concluido__iexact=filtro_servico_concluido)

    # Helpers para normalizar Mês/Ano: 01/2026 = JANEIRO / 2026 (todos os meses)
    MES_NUM_TO_NOME = {
        1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO', 4: 'ABRIL', 5: 'MAIO', 6: 'JUNHO',
        7: 'JULHO', 8: 'AGOSTO', 9: 'SETEMBRO', 10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO',
    }
    MES_NOME_TO_NUM = {}
    for _n, _nome in MES_NUM_TO_NOME.items():
        MES_NOME_TO_NUM[_nome.upper()] = _n
    MES_NOME_TO_NUM['MARCO'] = 3
    for _n, _a in [(1,'JAN'),(2,'FEV'),(3,'MAR'),(4,'ABR'),(5,'MAI'),(6,'JUN'),(7,'JUL'),(8,'AGO'),(9,'SET'),(10,'OUT'),(11,'NOV'),(12,'DEZ')]:
        MES_NOME_TO_NUM[_a] = _n

    def _parse_mes_ano(val):
        """Converte '01/2026' ou 'JANEIRO / 2026' -> (1, 2026). Retorna None se inválido."""
        if not val or not str(val).strip():
            return None
        val = str(val).strip().upper()
        ano_match = re.search(r'\b(20\d{2}|19\d{2})\b', val)
        ano = int(ano_match.group(1)) if ano_match else None
        if ano is None:
            return None
        # Formato "01/2026" ou "1/2026"
        if '/' in val:
            parts = val.split('/', 1)
            mes_part = parts[0].strip()
            if mes_part.isdigit():
                mes_int = int(mes_part)
                if 1 <= mes_int <= 12:
                    return (mes_int, ano)
        # Formato "JANEIRO / 2026" ou "JAN 2026" (verificar nomes mais longos primeiro)
        for nome in sorted(MES_NOME_TO_NUM.keys(), key=lambda x: -len(x)):
            if len(nome) >= 2 and nome in val:
                return (MES_NOME_TO_NUM[nome], ano)
        return None

    def _canonical_mes_ano(mes_int, ano_int):
        """(1, 2026) -> 'JANEIRO / 2026'."""
        return f"{MES_NUM_TO_NOME.get(mes_int, str(mes_int))} / {ano_int}"

    filtro_mes_ano = request.GET.get('filtro_mes_ano', '').strip()
    if filtro_mes_ano:
        parsed = _parse_mes_ano(filtro_mes_ano)
        if parsed:
            mes_int, ano_int = parsed
            mes_strs = [str(mes_int), f"{mes_int:02d}"]
            nome_mes = MES_NUM_TO_NOME.get(mes_int, '')
            abrev_mes = dict((_n, _a) for _n, _a in [(1,'JAN'),(2,'FEV'),(3,'MAR'),(4,'ABR'),(5,'MAI'),(6,'JUN'),(7,'JUL'),(8,'AGO'),(9,'SET'),(10,'OUT'),(11,'NOV'),(12,'DEZ')]).get(mes_int, '')
            q_numeric = Q(ano_referencia=ano_int) & (Q(mes_referencia__in=mes_strs) if nome_mes else Q(mes_referencia__icontains=filtro_mes_ano))
            q_previsao = (Q(previsao_execucao__icontains=nome_mes) | Q(previsao_execucao__icontains=abrev_mes)) & Q(previsao_execucao__icontains=str(ano_int)) if nome_mes else Q(previsao_execucao__icontains=filtro_mes_ano)
            projecoes_list = projecoes_list.filter(q_numeric | q_previsao)
        else:
            projecoes_list = projecoes_list.filter(
                Q(mes_referencia__icontains=filtro_mes_ano) |
                Q(previsao_execucao__icontains=filtro_mes_ano)
            )

    filtro_id = request.GET.get('filtro_id', '').strip()
    if filtro_id:
        try:
            id_val = int(filtro_id)
            projecoes_list = projecoes_list.filter(id_excel=id_val)
        except (ValueError, TypeError):
            pass
    
    filtro_numero_nf = request.GET.get('filtro_numero_nf', '').strip()
    if filtro_numero_nf:
        projecoes_list = projecoes_list.filter(numero_nf__icontains=filtro_numero_nf)
    
    filtro_numero_requisicao = request.GET.get('filtro_numero_requisicao', '').strip()
    if filtro_numero_requisicao:
        projecoes_list = projecoes_list.filter(numero_requisicao_compra__icontains=filtro_numero_requisicao)
    
    filtro_valor_min = request.GET.get('filtro_valor_min', '').strip()
    if filtro_valor_min:
        try:
            valor_min = Decimal(filtro_valor_min.replace(',', '.'))
            projecoes_list = projecoes_list.filter(valor_total__gte=valor_min)
        except (ValueError, TypeError):
            pass
    
    filtro_valor_max = request.GET.get('filtro_valor_max', '').strip()
    if filtro_valor_max:
        try:
            valor_max = Decimal(filtro_valor_max.replace(',', '.'))
            projecoes_list = projecoes_list.filter(valor_total__lte=valor_max)
        except (ValueError, TypeError):
            pass
    
    # Ordenação: permitir ordenar por ID (crescente ou decrescente)
    order_id = (request.GET.get('order_id') or '').strip().lower()
    default_order = ['-ano_referencia', '-mes_referencia', '-data_abertura_requisicao', '-created_at']
    try:
        if order_id == 'asc':
            projecoes_list = projecoes_list.order_by('id_excel', *default_order)
        elif order_id == 'desc':
            projecoes_list = projecoes_list.order_by('-id_excel', *default_order)
        else:
            projecoes_list = projecoes_list.order_by(*default_order)
    except Exception:
        try:
            projecoes_list = projecoes_list.order_by('-created_at', '-id')
        except Exception:
            pass
    
    # Paginação
    try:
        paginator = Paginator(projecoes_list, 50)
        page_number = request.GET.get('page', 1)
        try:
            page_number = int(page_number)
        except (ValueError, TypeError):
            page_number = 1
        projecoes = paginator.get_page(page_number)
    except Exception as e:
        empty_list = ProjecaoGasto.objects.none()
        paginator = Paginator(empty_list, 50)
        projecoes = paginator.page(1)
    
    # Obter valores únicos para os dropdowns de filtros da tabela
    setores_unicos = ProjecaoGasto.objects.exclude(
        setor__isnull=True
    ).exclude(
        setor=''
    ).values_list('setor', flat=True).distinct().order_by('setor')
    
    # Opções para o filtro Mês/Ano Ref.: normalizar para uma opção por (mês, ano), ex: "JANEIRO / 2026"
    mes_ano_opcoes_set = set()
    for pe in ProjecaoGasto.objects.exclude(previsao_execucao__isnull=True).exclude(previsao_execucao='').values_list('previsao_execucao', flat=True).distinct():
        if pe:
            parsed = _parse_mes_ano(pe)
            if parsed:
                mes_ano_opcoes_set.add(_canonical_mes_ano(parsed[0], parsed[1]))
    for mes_ref, ano_ref in ProjecaoGasto.objects.exclude(mes_referencia__isnull=True).exclude(ano_referencia__isnull=True).exclude(mes_referencia='').values_list('mes_referencia', 'ano_referencia').distinct():
        if mes_ref is not None and ano_ref is not None:
            try:
                mes_int = int(mes_ref) if str(mes_ref).isdigit() else None
                if mes_int and 1 <= mes_int <= 12:
                    mes_ano_opcoes_set.add(_canonical_mes_ano(mes_int, int(ano_ref)))
            except (ValueError, TypeError):
                pass
    def _sort_key_mes_ano(s):
        p = _parse_mes_ano(s)
        return (p[1], p[0], s) if p else (0, 0, s)
    mes_ano_opcoes = sorted(mes_ano_opcoes_set, key=_sort_key_mes_ano, reverse=True)
    
    # URLs para ordenar coluna ID (preservando todos os filtros)
    from urllib.parse import urlencode
    get_copy_asc = request.GET.copy()
    get_copy_asc['order_id'] = 'asc'
    get_copy_desc = request.GET.copy()
    get_copy_desc['order_id'] = 'desc'
    url_ordenar_id_asc = request.path + '?' + urlencode(get_copy_asc)
    url_ordenar_id_desc = request.path + '?' + urlencode(get_copy_desc)
    
    context = {
        'page_title': 'Consultar Projeção de Gastos',
        'active_page': 'consultar_projecao_gastos',
        'projecoes': projecoes,
        'search_query': search_query or '',
        'setores_unicos': setores_unicos or [],
        'mes_ano_opcoes': mes_ano_opcoes,
        'filtro_setor': filtro_setor or '',
        'filtro_fornecedor': filtro_fornecedor or '',
        'filtro_descricao': filtro_descricao or '',
        'filtro_uso_contabil': filtro_uso_contabil or '',
        'filtro_mes_ano': filtro_mes_ano or '',
        'filtro_mes_ano_canonical': _canonical_mes_ano(parsed[0], parsed[1]) if (filtro_mes_ano and (parsed := _parse_mes_ano(filtro_mes_ano))) else (filtro_mes_ano or ''),
        'filtro_id': filtro_id or '',
        'filtro_numero_nf': filtro_numero_nf or '',
        'filtro_numero_requisicao': filtro_numero_requisicao or '',
        'filtro_valor_min': filtro_valor_min or '',
        'filtro_valor_max': filtro_valor_max or '',
        'filtro_servico_concluido': filtro_servico_concluido or '',
        # Filtros de ano/mês (igual analise_geral_orcamento)
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
        'mostrar_todos': mostrar_todos,
        # Ordenação ID
        'order_id': order_id,
        'url_ordenar_id_asc': url_ordenar_id_asc,
        'url_ordenar_id_desc': url_ordenar_id_desc,
    }
    
    return render(request, 'orcamento/consultar_projecao_gastos.html', context)


def visualizar_projecao_gasto(request, projecao_id):
    """Visualizar detalhes de uma projeção de gasto específica"""
    from app.models import ProjecaoGasto, RelacaoProjecaoNotaFiscal

    try:
        projecao = ProjecaoGasto.objects.get(id=projecao_id)
    except ProjecaoGasto.DoesNotExist:
        messages.error(request, 'Projeção de gasto não encontrada.')
        return redirect('consultar_projecao_gastos')

    # Notas fiscais relacionadas (se houver)
    relacoes_nf = RelacaoProjecaoNotaFiscal.objects.filter(projecao=projecao).select_related('nota_fiscal')

    context = {
        'page_title': f'Visualizar Projeção de Gasto #{projecao.id}',
        'active_page': 'consultar_projecao_gastos',
        'projecao': projecao,
        'relacoes_nf': relacoes_nf,
    }
    return render(request, 'orcamento/visualizar_projecao_gasto.html', context)


def visualizar_projecoes_gastos_por_setor(request):
    """Lista e analisa projeções de gastos de um único setor. setor via GET (permite nomes com /)."""
    from app.models import ProjecaoGasto
    from django.core.paginator import Paginator
    from django.db.models import Sum, Count, Q
    from decimal import Decimal
    from django.shortcuts import redirect
    import json

    setor_nome = request.GET.get('setor', '').strip()
    if not setor_nome:
        return redirect('analise_projecao_gastos')
    projecoes_list = ProjecaoGasto.objects.filter(setor=setor_nome).order_by(
        '-ano_referencia', '-mes_referencia', '-data_abertura_requisicao', '-created_at'
    )
    total_registros = projecoes_list.count()
    valor_total_setor = projecoes_list.aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')

    # Previsões para o mês atual (mes_referencia + ano_referencia = mês/ano corrente)
    from datetime import datetime
    hoje = datetime.now()
    q_mes_atual = Q(ano_referencia=hoje.year) & (Q(mes_referencia=str(hoje.month)) | Q(mes_referencia=f'{hoje.month:02d}'))
    projecoes_mes_atual = projecoes_list.filter(q_mes_atual)
    valor_mes_atual = projecoes_mes_atual.aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')
    count_mes_atual = projecoes_mes_atual.count()

    # Serviço concluído: SIM / NÃO / INDEFINIDO (resto)
    qs_sim = projecoes_list.filter(servico_concluido__iexact='SIM')
    qs_nao = projecoes_list.filter(servico_concluido__iexact='NÃO')
    qs_indef = projecoes_list.exclude(servico_concluido__iexact='SIM').exclude(servico_concluido__iexact='NÃO')
    count_sim = qs_sim.count()
    count_nao = qs_nao.count()
    count_indef = qs_indef.count()
    valor_sim = qs_sim.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0.00')
    valor_nao = qs_nao.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0.00')
    valor_indef = qs_indef.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0.00')
    pct_sim = (count_sim / total_registros * 100) if total_registros > 0 else 0
    pct_nao = (count_nao / total_registros * 100) if total_registros > 0 else 0
    pct_indef = (count_indef / total_registros * 100) if total_registros > 0 else 0

    # Com número de NF preenchido
    qs_com_nf = projecoes_list.exclude(numero_nf__isnull=True).exclude(numero_nf='')
    count_com_nf = qs_com_nf.count()
    valor_com_nf = qs_com_nf.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0.00')
    pct_com_nf = (count_com_nf / total_registros * 100) if total_registros > 0 else 0

    # Top fornecedores do setor (valor total)
    top_fornecedores = list(
        projecoes_list.exclude(fornecedor_nome_fantasia__isnull=True).exclude(fornecedor_nome_fantasia='')
        .values('fornecedor_nome_fantasia')
        .annotate(valor_total=Sum('valor_total'), qtd=Count('id'))
        .order_by('-valor_total')[:10]
    )
    # Valor por uso contábil
    uso_contabil_agregado = list(
        projecoes_list.exclude(uso_contabil__isnull=True).exclude(uso_contabil='')
        .values('uso_contabil')
        .annotate(valor_total=Sum('valor_total'), qtd=Count('id'))
        .order_by('-valor_total')
    )
    # Evolução por mês/ano (mes_referencia + ano_referencia) — por ano para filtro no gráfico
    _meses_nome = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    anos_grafico = list(
        projecoes_list.exclude(ano_referencia__isnull=True)
        .values_list('ano_referencia', flat=True).distinct().order_by('-ano_referencia')
    )
    if not anos_grafico:
        from datetime import datetime
        anos_grafico = [datetime.now().year]
    evolucao_por_ano = {}
    for ano in anos_grafico:
        labels_ano = []
        valores_ano = []
        valores_sim = []
        valores_nao = []
        for mes in range(1, 13):
            q_mes = Q(ano_referencia=ano) & (Q(mes_referencia=str(mes)) | Q(mes_referencia=f'{mes:02d}'))
            base_qs = projecoes_list.filter(q_mes)
            agg = base_qs.aggregate(v=Sum('valor_total'), c=Count('id'))
            v = agg['v'] or Decimal('0')
            agg_sim = base_qs.filter(servico_concluido__iexact='SIM').aggregate(v=Sum('valor_total'))
            agg_nao = base_qs.filter(servico_concluido__iexact='NÃO').aggregate(v=Sum('valor_total'))
            v_sim = agg_sim['v'] or Decimal('0')
            v_nao = agg_nao['v'] or Decimal('0')
            labels_ano.append(f"{_meses_nome[mes - 1]}/{str(ano)[2:]}")
            valores_ano.append(float(v))
            valores_sim.append(float(v_sim))
            valores_nao.append(float(v_nao))
        evolucao_por_ano[str(ano)] = {
            'labels': labels_ano,
            'valores': valores_ano,
            'valores_sim': valores_sim,
            'valores_nao': valores_nao,
        }
    # Manter compatibilidade: primeiro ano = dados default do gráfico
    ano_default = str(anos_grafico[0])
    evolucao_por_mes = [{'label': evolucao_por_ano[ano_default]['labels'][i], 'valor': evolucao_por_ano[ano_default]['valores'][i]} for i in range(12)]
    evolucao_labels = json.dumps(evolucao_por_ano[ano_default]['labels'], ensure_ascii=False)
    evolucao_valores = json.dumps(evolucao_por_ano[ano_default]['valores'])
    evolucao_por_ano_json = json.dumps(evolucao_por_ano, ensure_ascii=False)
    anos_grafico_json = json.dumps([str(a) for a in anos_grafico], ensure_ascii=False)

    # Análise de dados faltantes (por campo) para orientar o preenchimento
    def _missing_char(qs, field):
        return qs.filter(Q(**{f'{field}__isnull': True}) | Q(**{field: ''})).count()
    def _missing_date(qs, field):
        return qs.filter(**{f'{field}__isnull': True}).count()
    def _missing_valor(qs):
        return qs.filter(Q(valor_total__isnull=True) | Q(valor_total=0)).count()
    def _missing_ano(qs):
        return qs.filter(ano_referencia__isnull=True).count()

    campos_analise = [
        ('descricao', 'Descrição do Serviço'),
        ('valor_total', 'Valor Total'),
        ('fornecedor_nome_fantasia', 'Fornecedor (Nome Fantasia)'),
        ('fornecedor_cnpj', 'Fornecedor (CNPJ)'),
        ('data_abertura_requisicao', 'Data de Abertura da Requisição'),
        ('previsao_execucao', 'Previsão para Execução'),
        ('mes_referencia', 'Mês Referência'),
        ('ano_referencia', 'Ano Referência'),
        ('uso_contabil', 'Uso Contábil'),
        ('numero_nf', 'Número da NF'),
        ('numero_ordem_servico', 'Número Ordem de Serviço'),
        ('numero_requisicao_compra', 'Número da Requisição de Compra'),
        ('numero_pedido_compra', 'Número do Pedido de Compra'),
        ('servico_concluido', 'Serviço Concluído'),
        ('solicitante', 'Solicitante'),
        ('tipo_solicitacao', 'Tipo de Solicitação'),
        ('nf_servico_recebida', 'NF de Serviço Recebida'),
        ('nf_enviada_lancamento', 'NF Enviada para Lançamento'),
        ('observacoes', 'Observações'),
    ]
    dados_faltantes = []
    for campo, label in campos_analise:
        if campo == 'valor_total':
            n = _missing_valor(projecoes_list)
        elif campo == 'ano_referencia':
            n = _missing_ano(projecoes_list)
        elif campo == 'data_abertura_requisicao':
            n = _missing_date(projecoes_list, campo)
        else:
            n = _missing_char(projecoes_list, campo)
        pct = (n / total_registros * 100) if total_registros > 0 else 0
        dados_faltantes.append({
            'campo': campo,
            'label': label,
            'count_missing': n,
            'pct_missing': round(pct, 1),
            'count_ok': total_registros - n,
        })
    dados_faltantes.sort(key=lambda x: (-x['count_missing'], x['label']))

    paginator = Paginator(projecoes_list, 50)
    page_number = request.GET.get('page', 1)
    try:
        page_number = int(page_number)
    except (ValueError, TypeError):
        page_number = 1
    projecoes = paginator.get_page(page_number)

    context = {
        'page_title': f'Projeções de Gastos - Setor: {setor_nome}',
        'active_page': 'analise_projecao_gastos',
        'setor_nome': setor_nome,
        'projecoes': projecoes,
        'valor_total_setor': valor_total_setor,
        'valor_mes_atual': valor_mes_atual,
        'count_mes_atual': count_mes_atual,
        'total_registros': total_registros,
        'count_sim': count_sim,
        'count_nao': count_nao,
        'count_indef': count_indef,
        'valor_sim': valor_sim,
        'valor_nao': valor_nao,
        'valor_indef': valor_indef,
        'pct_sim': pct_sim,
        'pct_nao': pct_nao,
        'pct_indef': pct_indef,
        'count_com_nf': count_com_nf,
        'valor_com_nf': valor_com_nf,
        'pct_com_nf': pct_com_nf,
        'top_fornecedores': top_fornecedores,
        'uso_contabil_agregado': uso_contabil_agregado,
        'evolucao_labels': evolucao_labels,
        'evolucao_valores': evolucao_valores,
        'evolucao_por_mes': evolucao_por_mes,
        'evolucao_por_ano_json': evolucao_por_ano_json,
        'anos_grafico': anos_grafico,
        'anos_grafico_json': anos_grafico_json,
        'dados_faltantes': dados_faltantes,
    }
    return render(request, 'orcamento/visualizar_projecoes_gastos_por_setor.html', context)


def analise_projecao_gastos(request):
    """Página de análise detalhada de projeções de gastos com filtros e gráficos"""
    from app.models import ProjecaoGasto, RelacaoProjecaoNotaFiscal, DadosOrcamento
    from django.db.models import Sum, Count, Q
    from datetime import datetime, timedelta
    from decimal import Decimal
    import json
    
    # --- Lógica de Filtro ---
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')
    setores_filtro = request.GET.getlist('setor')
    setores_filtro = [s.strip() for s in setores_filtro if s and s.strip()]
    uso_contabil_filtro = request.GET.get('uso_contabil', '').strip()

    hoje = datetime.now()
    
    # Anos disponíveis para o filtro (baseado em ano_referencia OU created_at)
    anos_disponiveis_set = set()
    # Buscar anos de ano_referencia
    anos_ref = ProjecaoGasto.objects.exclude(ano_referencia__isnull=True).values_list('ano_referencia', flat=True).distinct()
    for ano in anos_ref:
        if ano:
            anos_disponiveis_set.add(int(ano))
    # Buscar anos de created_at
    projecoes_com_data = ProjecaoGasto.objects.exclude(created_at__isnull=True)
    for proj in projecoes_com_data:
        if proj.created_at:
            anos_disponiveis_set.add(proj.created_at.year)
    
    # Sempre incluir o ano atual na lista de anos disponíveis
    anos_disponiveis_set.add(hoje.year)
    
    anos_disponiveis = sorted(list(anos_disponiveis_set), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    # Se não há filtro de ano, usar o ano atual (para incluir dados novos)
    if not ano_filtro:
        ano_filtro = hoje.year
    else:
        try:
            ano_filtro = int(ano_filtro)
        except (ValueError, TypeError):
            ano_filtro = hoje.year
    
    if ano_filtro not in anos_disponiveis:
        anos_disponiveis.insert(0, ano_filtro)
        anos_disponiveis = sorted(list(set(anos_disponiveis)), reverse=True)

    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))

    # Se nenhum mês for selecionado, usar o mês atual
    if not meses_filtro_int:
        meses_filtro_int = [hoje.month]
    meses_para_mostrar = meses_filtro_int

    meses_choices = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]

    # Mapear meses numéricos para strings
    meses_str_map = {
        1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO', 4: 'ABRIL',
        5: 'MAIO', 6: 'JUNHO', 7: 'JULHO', 8: 'AGOSTO',
        9: 'SETEMBRO', 10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO'
    }
    meses_selecionados_str = [meses_str_map[m] for m in meses_para_mostrar]

    # Parser de previsao_execucao (reutilizado em correlação mensal e gráfico Gastos por Previsão)
    import re
    _meses_nome_to_num = {
        'JANEIRO': 1, 'FEVEREIRO': 2, 'MARCO': 3, 'MARÇO': 3, 'ABRIL': 4, 'MAIO': 5, 'JUNHO': 6,
        'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9, 'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12
    }
    def _parse_previsao_execucao(texto):
        if not texto or not str(texto).strip():
            return None, None
        texto = str(texto).strip().upper()
        ano_match = re.search(r'\b(20\d{2}|19\d{2})\b', texto)
        ano = int(ano_match.group(1)) if ano_match else None
        mes_num = None
        for nome, num in _meses_nome_to_num.items():
            if nome in texto:
                mes_num = num
                break
        return ano, mes_num

    # Obter valores únicos de setor e uso_contabil para os filtros (de todos os dados, não apenas filtrados)
    setores_disponiveis = ProjecaoGasto.objects.exclude(
        setor__isnull=True
    ).exclude(
        setor=''
    ).values_list('setor', flat=True).distinct().order_by('setor')
    
    uso_contabil_disponiveis = ProjecaoGasto.objects.exclude(
        uso_contabil__isnull=True
    ).exclude(
        uso_contabil=''
    ).values_list('uso_contabil', flat=True).distinct().order_by('uso_contabil')
    
    # --- Queryset Filtrado ---
    # Começar com todos os dados, mas excluir registros que só têm ID (sem dados significativos)
    # Incluir apenas registros que têm pelo menos um campo principal preenchido
    projecoes_qs = ProjecaoGasto.objects.filter(
        Q(setor__isnull=False) & ~Q(setor='') |
        Q(descricao__isnull=False) & ~Q(descricao='') |
        Q(valor_total__isnull=False) & ~Q(valor_total=0) |
        Q(fornecedor_nome_fantasia__isnull=False) & ~Q(fornecedor_nome_fantasia='')
    )
    # Queryset sem filtro de ano (para o gráfico Gastos por Previsão, que tem seletor de ano próprio)
    projecoes_grafico_previsao = projecoes_qs
    
    # Aplicar filtro de ano (ano_referencia OU created_at quando ano_referencia vazio)
    q_ano = (
        Q(ano_referencia=ano_filtro) |  # Ano de referência coincide
        (Q(ano_referencia__isnull=True) & Q(created_at__year=ano_filtro))  # Fallback: sem ano_ref, criado no ano
    )
    projecoes_qs = projecoes_qs.filter(q_ano)
    
    # Se há meses selecionados, aplicar filtro de mês
    # mes_referencia no DB é armazenado como '01','02',...'12' (extraído do import), não 'JANEIRO','FEVEREIRO'
    # Também suportar previsao_execucao como fallback (ex: "DEZEMBRO / 2025")
    if meses_filtro_int:
        q_mes = Q()
        # Formato numérico: mes_referencia in ['01','02',...]
        for mes_num in meses_filtro_int:
            q_mes |= Q(mes_referencia__in=[str(mes_num), f'{mes_num:02d}'])
        # Formato texto (fallback para imports antigos ou diferentes)
        for mes_str in meses_selecionados_str:
            q_mes |= Q(mes_referencia__iexact=mes_str)
        # previsao_execucao contém nome do mês (ex: "JANEIRO / 2025")
        meses_nomes_variantes = {
            1: ['JANEIRO', 'JAN'], 2: ['FEVEREIRO', 'FEV'], 3: ['MARÇO', 'MARCO', 'MAR'],
            4: ['ABRIL', 'ABR'], 5: ['MAIO', 'MAI'], 6: ['JUNHO', 'JUN'],
            7: ['JULHO', 'JUL'], 8: ['AGOSTO', 'AGO'], 9: ['SETEMBRO', 'SET'],
            10: ['OUTUBRO', 'OUT'], 11: ['NOVEMBRO', 'NOV'], 12: ['DEZEMBRO', 'DEZ']
        }
        for mes_num in meses_filtro_int:
            for nome in meses_nomes_variantes.get(mes_num, []):
                q_mes |= Q(previsao_execucao__icontains=nome)
        # Sem mes_referencia nem previsao: usar created_at
        for mes_num in meses_filtro_int:
            q_mes |= Q(
                (Q(mes_referencia__isnull=True) | Q(mes_referencia='')) &
                (Q(previsao_execucao__isnull=True) | Q(previsao_execucao='')) &
                Q(created_at__month=mes_num)
            )
        projecoes_qs = projecoes_qs.filter(q_mes)
        projecoes_grafico_previsao = projecoes_grafico_previsao.filter(q_mes)
    
    # Aplicar filtro de setor (múltiplos)
    if setores_filtro:
        projecoes_qs = projecoes_qs.filter(setor__in=setores_filtro)
        projecoes_grafico_previsao = projecoes_grafico_previsao.filter(setor__in=setores_filtro)
    
    # Aplicar filtro de uso contábil
    if uso_contabil_filtro:
        projecoes_qs = projecoes_qs.filter(uso_contabil=uso_contabil_filtro)
        projecoes_grafico_previsao = projecoes_grafico_previsao.filter(uso_contabil=uso_contabil_filtro)

    # --- KPIs ---
    total_projecoes = projecoes_qs.count()
    
    valor_total_projetado = projecoes_qs.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)

    # Projeções com servico_concluido = "SIM"
    qs_servico_sim = projecoes_qs.filter(servico_concluido__iexact='SIM')
    projecoes_servico_concluido_sim = qs_servico_sim.count()
    valor_servico_concluido_sim = qs_servico_sim.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)
    percentual_servico_concluido_sim = (projecoes_servico_concluido_sim / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções com servico_concluido = "NÃO"
    qs_servico_nao = projecoes_qs.filter(servico_concluido__iexact='NÃO')
    projecoes_servico_concluido_nao = qs_servico_nao.count()
    valor_servico_concluido_nao = qs_servico_nao.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)
    percentual_servico_concluido_nao = (projecoes_servico_concluido_nao / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções com servico_concluido = "INDEFINIDO" (nulo, vazio ou valor "INDEFINIDO")
    qs_servico_indefinido = projecoes_qs.exclude(servico_concluido__iexact='SIM').exclude(servico_concluido__iexact='NÃO')
    projecoes_servico_concluido_indefinido = qs_servico_indefinido.count()
    valor_servico_concluido_indefinido = qs_servico_indefinido.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)
    percentual_servico_concluido_indefinido = (projecoes_servico_concluido_indefinido / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções com Nota Fiscal relacionada
    projecoes_com_nf_ids = RelacaoProjecaoNotaFiscal.objects.filter(
        projecao__in=projecoes_qs
    ).values_list('projecao_id', flat=True).distinct()
    projecoes_com_nf = len(projecoes_com_nf_ids)
    percentual_com_nf = (projecoes_com_nf / total_projecoes * 100) if total_projecoes > 0 else 0
    
    # ========== CORRELAÇÃO PROJEÇÕES E NOTAS FISCAIS (Análise Automática) ==========
    from app.utils import find_matching_notas_fiscais
    
    projecoes_com_nota_match = 0
    projecoes_sem_nota_match = 0
    valor_projecoes_com_match = Decimal('0')
    valor_projecoes_sem_match = Decimal('0')
    
    # Lista detalhada de correlações para a tabela
    correlacoes_detalhadas = []
    
    # Dados mensais para o gráfico de correlação
    correlacao_mensal_labels = []
    correlacao_mensal_valor_total = []
    correlacao_mensal_valor_com_match = []
    correlacao_mensal_valor_sem_match = []
    
    # Analisar cada projeção para encontrar matches
    for projecao in projecoes_qs:
        matches = find_matching_notas_fiscais(projecao)
        if matches:
            projecoes_com_nota_match += 1
            valor_projecoes_com_match += (projecao.valor_total or Decimal('0'))
            # Adicionar à lista detalhada (pegar o melhor match - primeiro da lista ordenada)
            melhor_match = matches[0] if matches else None
            if melhor_match:
                nota_fiscal, match_details = melhor_match
                correlacoes_detalhadas.append({
                    'projecao': projecao,
                    'nota_fiscal': nota_fiscal,
                    'match_details': match_details,
                    'tem_match': True
                })
        else:
            projecoes_sem_nota_match += 1
            valor_projecoes_sem_match += (projecao.valor_total or Decimal('0'))
            # Adicionar projeções sem match também
            correlacoes_detalhadas.append({
                'projecao': projecao,
                'nota_fiscal': None,
                'match_details': None,
                'tem_match': False
            })
    
    percentual_projecoes_com_match = (projecoes_com_nota_match / total_projecoes * 100) if total_projecoes > 0 else 0
    
    # Ordenar correlações: primeiro as com match, depois as sem match
    correlacoes_detalhadas.sort(key=lambda x: (not x['tem_match'], x['projecao'].id_excel or 0))
    
    # Calcular valores mensais de correlação: alocar gasto ao mês/ano conforme previsao_execucao (igual ao gráfico Gastos por Previsão)
    meses_nomes_curtos = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    por_mes_corr = {(ano_filtro, m): {'total': Decimal('0'), 'com_match': Decimal('0'), 'sem_match': Decimal('0')} for m in meses_para_mostrar}
    for item in correlacoes_detalhadas:
        proj = item['projecao']
        ano, mes = _parse_previsao_execucao(proj.previsao_execucao)
        if ano is not None and mes is not None and ano == ano_filtro and mes in meses_para_mostrar:
            v = proj.valor_total or Decimal('0')
            por_mes_corr[(ano_filtro, mes)]['total'] += v
            if item['tem_match']:
                por_mes_corr[(ano_filtro, mes)]['com_match'] += v
            else:
                por_mes_corr[(ano_filtro, mes)]['sem_match'] += v
    for mes_num in meses_para_mostrar:
        label = f"{meses_nomes_curtos[mes_num-1]}/{str(ano_filtro)[2:]}"
        correlacao_mensal_labels.append(label)
        d = por_mes_corr[(ano_filtro, mes_num)]
        correlacao_mensal_valor_total.append(float(d['total']))
        correlacao_mensal_valor_com_match.append(float(d['com_match']))
        correlacao_mensal_valor_sem_match.append(float(d['sem_match']))

    # Projeções com Ordem de Serviço (para gráfico)
    projecoes_com_os = projecoes_qs.exclude(numero_ordem_servico__isnull=True).exclude(numero_ordem_servico='').count()
    percentual_com_os = (projecoes_com_os / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções com numero_nf preenchido (box "Com Ordem de Serviço" repurposed)
    qs_com_numero_nf = projecoes_qs.exclude(numero_nf__isnull=True).exclude(numero_nf='')
    projecoes_com_numero_nf = qs_com_numero_nf.count()
    valor_com_numero_nf = qs_com_numero_nf.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)
    percentual_com_numero_nf = (projecoes_com_numero_nf / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções com serviço concluído
    projecoes_concluidas = projecoes_qs.exclude(servico_concluido__isnull=True).count()
    percentual_concluidas = (projecoes_concluidas / total_projecoes * 100) if total_projecoes > 0 else 0

    # Projeções pendentes (sem serviço concluído) - mantido para gráfico Status
    projecoes_pendentes = total_projecoes - projecoes_concluidas
    percentual_pendentes = (projecoes_pendentes / total_projecoes * 100) if total_projecoes > 0 else 0

    # Impacto no Orçamento: % do Valor Total Projetado sobre orçamento total do período (DadosOrcamento)
    # Ex.: filtro FEVEREIRO -> orçamento de fevereiro; múltiplos meses -> soma dos orçamentos dos meses
    valor_orcamento_periodo = DadosOrcamento.objects.filter(
        ano=ano_filtro, mes__in=meses_para_mostrar
    ).aggregate(Sum('valor_orcamento'))['valor_orcamento__sum'] or Decimal(0)
    if valor_orcamento_periodo and valor_orcamento_periodo > 0:
        percentual_impacto_orcamento = float((valor_total_projetado / valor_orcamento_periodo) * 100)
    else:
        percentual_impacto_orcamento = None  # sem orçamento cadastrado para o período

    # Setores únicos
    setores_unicos = projecoes_qs.exclude(setor__isnull=True).exclude(setor='').values('setor').distinct().count()
    
    # Obter valores únicos de setor e uso_contabil para os filtros (de todos os dados, não apenas filtrados)
    setores_disponiveis = ProjecaoGasto.objects.exclude(
        setor__isnull=True
    ).exclude(
        setor=''
    ).values_list('setor', flat=True).distinct().order_by('setor')
    
    uso_contabil_disponiveis = ProjecaoGasto.objects.exclude(
        uso_contabil__isnull=True
    ).exclude(
        uso_contabil=''
    ).values_list('uso_contabil', flat=True).distinct().order_by('uso_contabil')

    # --- Dados para Gráficos ---
    # Gráfico 1: Distribuição por Setor (Top 10) - usando aggregate para eficiência
    setores_agregados = projecoes_qs.exclude(setor__isnull=True).exclude(setor='').values('setor').annotate(
        valor_total=Sum('valor_total')
    ).order_by('-valor_total')[:10]
    
    setores_labels = [s['setor'] for s in setores_agregados]
    setores_data = [float(s['valor_total'] or 0) for s in setores_agregados]

    # Gráfico 2: Top 10 Fornecedores por Valor - usando aggregate para eficiência
    fornecedores_agregados = projecoes_qs.exclude(fornecedor_nome_fantasia__isnull=True).exclude(fornecedor_nome_fantasia='').values('fornecedor_nome_fantasia').annotate(
        valor_total=Sum('valor_total')
    ).order_by('-valor_total')[:10]
    
    fornecedores_labels = [f['fornecedor_nome_fantasia'][:30] + '...' if len(f['fornecedor_nome_fantasia']) > 30 else f['fornecedor_nome_fantasia'] for f in fornecedores_agregados]
    fornecedores_data = [float(f['valor_total'] or 0) for f in fornecedores_agregados]

    # Gráfico 3: Evolução Temporal (Mês a Mês)
    evolucao_labels = []
    evolucao_quantidade_data = []
    evolucao_valor_data = []
    
    meses_nomes_curtos = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    for mes_num, mes_nome in meses_choices:
        if mes_num in meses_para_mostrar:
            label = f"{meses_nomes_curtos[mes_num-1]}/{str(ano_filtro)[2:]}"
            evolucao_labels.append(label)
            
            # Alocar projeção ao mês: mes_referencia ('01','02'...) OU nome em previsao_execucao OU created_at
            q_mes_temporal = (
                Q(mes_referencia__in=[str(mes_num), f'{mes_num:02d}']) |
                Q(mes_referencia__iexact=meses_str_map[mes_num]) |
                Q(previsao_execucao__icontains=meses_str_map[mes_num]) |
                Q(created_at__month=mes_num)
            )
            qtd_mes = projecoes_qs.filter(q_mes_temporal).count()
            evolucao_quantidade_data.append(qtd_mes)
            
            # Valor total do mês
            valor_mes = projecoes_qs.filter(q_mes_temporal).aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)
            evolucao_valor_data.append(float(valor_mes))

    # Gráfico 4: Status das Projeções
    status_labels = []
    status_data = []
    
    # Com NF
    status_labels.append('Com Nota Fiscal')
    status_data.append(projecoes_com_nf)
    
    # Sem NF
    status_labels.append('Sem Nota Fiscal')
    status_data.append(total_projecoes - projecoes_com_nf)
    
    # Com OS
    status_labels.append('Com Ordem de Serviço')
    status_data.append(projecoes_com_os)
    
    # Concluídas
    status_labels.append('Serviços Concluídos')
    status_data.append(projecoes_concluidas)
    
    # Pendentes
    status_labels.append('Pendentes')
    status_data.append(projecoes_pendentes)

    # Gráfico 5: Distribuição por Uso Contábil - usando aggregate para eficiência
    uso_contabil_agregados = projecoes_qs.exclude(uso_contabil__isnull=True).exclude(uso_contabil='').values('uso_contabil').annotate(
        valor_total=Sum('valor_total')
    ).order_by('-valor_total')[:15]
    
    uso_contabil_labels = [u['uso_contabil'][:40] + '...' if len(u['uso_contabil']) > 40 else u['uso_contabil'] for u in uso_contabil_agregados]
    uso_contabil_data = [float(u['valor_total'] or 0) for u in uso_contabil_agregados]

    # Gráfico 6: Gastos por Previsão de Execução — mesmo processo do gráfico Evolução (setor): por ano, 3 barras (Total, SIM, NÃO)
    from collections import defaultdict
    meses_nomes_curtos_previsao = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    total_previsao = defaultdict(lambda: Decimal('0'))
    sim_previsao = defaultdict(lambda: Decimal('0'))
    nao_previsao = defaultdict(lambda: Decimal('0'))
    for proj in projecoes_grafico_previsao.exclude(previsao_execucao__isnull=True).exclude(previsao_execucao=''):
        ano, mes = _parse_previsao_execucao(proj.previsao_execucao)
        if ano is not None and mes is not None:
            v = proj.valor_total or Decimal('0')
            key = (ano, mes)
            total_previsao[key] += v
            if (proj.servico_concluido or '').strip().upper() == 'SIM':
                sim_previsao[key] += v
            elif (proj.servico_concluido or '').strip().upper() == 'NÃO':
                nao_previsao[key] += v
    anos_previsao_set = set(k[0] for k in total_previsao.keys())
    anos_grafico_previsao = sorted(anos_previsao_set, reverse=True) if anos_previsao_set else [datetime.now().year]
    gastos_previsao_por_ano = {}
    for ano in anos_grafico_previsao:
        labels_ano = [f"{meses_nomes_curtos_previsao[mes - 1]}/{str(ano)[2:]}" for mes in range(1, 13)]
        valores_ano = [float(total_previsao.get((ano, mes), Decimal('0'))) for mes in range(1, 13)]
        valores_sim = [float(sim_previsao.get((ano, mes), Decimal('0'))) for mes in range(1, 13)]
        valores_nao = [float(nao_previsao.get((ano, mes), Decimal('0'))) for mes in range(1, 13)]
        gastos_previsao_por_ano[str(ano)] = {
            'labels': labels_ano,
            'valores': valores_ano,
            'valores_sim': valores_sim,
            'valores_nao': valores_nao,
        }
    gastos_previsao_por_ano_json = json.dumps(gastos_previsao_por_ano, ensure_ascii=False)
    anos_grafico_previsao_json = json.dumps([str(a) for a in anos_grafico_previsao], ensure_ascii=False)

    # Gráfico Valor por Setor: X = data (mês/ano), Y = valor (R$); uma série por setor
    meses_nomes_curtos = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    # Pares (ano, mes) presentes nos dados (mes_referencia + ano_referencia)
    valor_por_setor_mes = defaultdict(lambda: defaultdict(lambda: Decimal('0')))  # setor -> (ano, mes) -> valor
    keys_meses_setor = set()  # (ano, mes)
    for proj in projecoes_qs.exclude(ano_referencia__isnull=True):
        ano = proj.ano_referencia
        mes_str = (proj.mes_referencia or '').strip()
        mes = None
        if mes_str.isdigit():
            mes = int(mes_str) if len(mes_str) <= 2 else int(mes_str[:2])
        if mes is None or not (1 <= mes <= 12):
            continue
        setor_nome = (proj.setor or '').strip() or 'Não informado'
        key = (ano, mes)
        keys_meses_setor.add(key)
        valor_por_setor_mes[setor_nome][key] += (proj.valor_total or Decimal('0'))
    keys_meses_ordenados = sorted(keys_meses_setor, key=lambda x: (x[0], x[1]))
    valor_por_setor_labels = [f"{meses_nomes_curtos[k[1] - 1]}/{str(k[0])[2:]}" for k in keys_meses_ordenados]
    setores_ordenados = sorted(set(valor_por_setor_mes.keys()), key=lambda s: (s == 'Não informado', s))
    cores_setores = [
        '#0d6efd', '#198754', '#ffc107', '#dc3545', '#0dcaf0', '#6f42c1',
        '#fd7e14', '#20c997', '#e83e8c', '#6c757d', '#17a2b8', '#28a745',
        '#ff6b6b', '#4ecdc4', '#95e1d3', '#a0522d', '#9370db', '#2f4f4f'
    ]
    valor_por_setor_datasets = []
    for i, setor in enumerate(setores_ordenados):
        valores = [float(valor_por_setor_mes[setor].get(k, Decimal('0'))) for k in keys_meses_ordenados]
        valor_por_setor_datasets.append({
            'label': setor,
            'data': valores,
            'backgroundColor': cores_setores[i % len(cores_setores)],
            'borderColor': cores_setores[i % len(cores_setores)],
            'borderWidth': 1
        })
    valor_por_setor_labels_json = json.dumps(valor_por_setor_labels, ensure_ascii=False)
    valor_por_setor_datasets_json = json.dumps(valor_por_setor_datasets, ensure_ascii=False)

    # --- Tabela por Setor e servico_concluido (abaixo do gráfico Gastos por Previsão) ---
    setores_raw = list(projecoes_qs.values_list('setor', flat=True).distinct())
    setores_para_tabela = []  # list of (display_name, filter Q object)
    seen = set()
    for s in setores_raw:
        if s is None or (isinstance(s, str) and not str(s).strip()):
            display = 'Não informado'
            q_filter = Q(setor__isnull=True) | Q(setor='')
            key = 'Não informado'  # consolidar null e vazio em uma linha
        else:
            display = str(s).strip() or 'Não informado'
            q_filter = Q(setor=s)  # use raw value for exact DB match
            key = s
        if key not in seen:
            seen.add(key)
            setores_para_tabela.append((display, q_filter))
    setores_para_tabela.sort(key=lambda x: (x[0] == 'Não informado', x[0]))

    # Fallback: se não encontrou setores mas há dados, mostrar linha "Total"
    if not setores_para_tabela and projecoes_qs.exists():
        setores_para_tabela = [('Total', Q())]

    tabela_setor_servico = []
    for setor_display, q_filter in setores_para_tabela:
        qs_setor = projecoes_qs.filter(q_filter)
        sim = qs_setor.filter(servico_concluido__iexact='SIM').aggregate(s=Sum('valor_total'))['s'] or Decimal(0)
        nao = qs_setor.filter(servico_concluido__iexact='NÃO').aggregate(s=Sum('valor_total'))['s'] or Decimal(0)
        nao_informado = qs_setor.filter(Q(servico_concluido__isnull=True) | Q(servico_concluido='')).aggregate(s=Sum('valor_total'))['s'] or Decimal(0)
        indefinido = qs_setor.exclude(servico_concluido__iexact='SIM').exclude(servico_concluido__iexact='NÃO').exclude(servico_concluido__isnull=True).exclude(servico_concluido='').aggregate(s=Sum('valor_total'))['s'] or Decimal(0)
        total = qs_setor.aggregate(s=Sum('valor_total'))['s'] or Decimal(0)
        tabela_setor_servico.append({
            'setor': setor_display,
            'sim': sim,
            'nao': nao,
            'indefinido': indefinido,
            'nao_informado': nao_informado,
            'total': total,
        })

    # --- Tabelas Detalhadas ---
    # Top 10 Setores por Valor
    top_setores = projecoes_qs.exclude(setor__isnull=True).exclude(setor='').values('setor').annotate(
        quantidade=Count('id'),
        valor_total=Sum('valor_total')
    ).order_by('-valor_total')[:10]

    # Top 10 Fornecedores por Valor
    top_fornecedores = projecoes_qs.exclude(fornecedor_nome_fantasia__isnull=True).exclude(fornecedor_nome_fantasia='').values('fornecedor_nome_fantasia').annotate(
        quantidade=Count('id'),
        valor_total=Sum('valor_total')
    ).order_by('-valor_total')[:10]

    setores_para_links = list(setores_disponiveis)[:9]

    context = {
        'page_title': 'Análise de Projeções de Gastos',
        'active_page': 'analise_projecao_gastos',
        'setores_para_links': setores_para_links,
        'anos_disponiveis': anos_disponiveis,
        'ano_selecionado': ano_filtro,
        'meses_choices': meses_choices,
        'meses_selecionados': meses_filtro_int,
        'setores_disponiveis': setores_disponiveis,
        'uso_contabil_disponiveis': uso_contabil_disponiveis,
        'setores_selecionados': setores_filtro,
        'uso_contabil_filtro': uso_contabil_filtro,

        # KPIs
        'total_projecoes': total_projecoes,
        'valor_total_projetado': valor_total_projetado,
        'valor_medio': valor_total_projetado / total_projecoes if total_projecoes > 0 else Decimal(0),  # mantido para compatibilidade
        'valor_servico_concluido_sim': valor_servico_concluido_sim,
        'projecoes_servico_concluido_sim': projecoes_servico_concluido_sim,
        'percentual_servico_concluido_sim': percentual_servico_concluido_sim,
        'valor_servico_concluido_nao': valor_servico_concluido_nao,
        'projecoes_servico_concluido_nao': projecoes_servico_concluido_nao,
        'percentual_servico_concluido_nao': percentual_servico_concluido_nao,
        'valor_servico_concluido_indefinido': valor_servico_concluido_indefinido,
        'projecoes_servico_concluido_indefinido': projecoes_servico_concluido_indefinido,
        'percentual_servico_concluido_indefinido': percentual_servico_concluido_indefinido,
        'projecoes_com_nf': projecoes_com_nf,
        'percentual_com_nf': percentual_com_nf,
        'projecoes_com_os': projecoes_com_os,
        'percentual_com_os': percentual_com_os,
        'valor_com_numero_nf': valor_com_numero_nf,
        'projecoes_com_numero_nf': projecoes_com_numero_nf,
        'percentual_com_numero_nf': percentual_com_numero_nf,
        'projecoes_concluidas': projecoes_concluidas,
        'percentual_concluidas': percentual_concluidas,
        'projecoes_pendentes': projecoes_pendentes,
        'percentual_pendentes': percentual_pendentes,
        'percentual_impacto_orcamento': percentual_impacto_orcamento,
        'valor_orcamento_periodo': valor_orcamento_periodo,
        'setores_unicos': setores_unicos,
        # Correlação Projeções e Notas Fiscais
        'projecoes_com_nota_match': projecoes_com_nota_match,
        'projecoes_sem_nota_match': projecoes_sem_nota_match,
        'percentual_projecoes_com_match': percentual_projecoes_com_match,
        'valor_projecoes_com_match': valor_projecoes_com_match,
        'valor_projecoes_sem_match': valor_projecoes_sem_match,
        'correlacoes_detalhadas': correlacoes_detalhadas,
        
        # Dados mensais para gráfico de correlação
        'correlacao_mensal_labels': json.dumps(correlacao_mensal_labels, ensure_ascii=False),
        'correlacao_mensal_valor_total': json.dumps(correlacao_mensal_valor_total),
        'correlacao_mensal_valor_com_match': json.dumps(correlacao_mensal_valor_com_match),
        'correlacao_mensal_valor_sem_match': json.dumps(correlacao_mensal_valor_sem_match),

        # Gráficos
        'setores_labels': json.dumps(setores_labels, ensure_ascii=False),
        'setores_data': json.dumps(setores_data),
        
        'fornecedores_labels': json.dumps(fornecedores_labels, ensure_ascii=False),
        'fornecedores_data': json.dumps(fornecedores_data),
        
        'evolucao_labels': json.dumps(evolucao_labels, ensure_ascii=False),
        'evolucao_quantidade_data': json.dumps(evolucao_quantidade_data),
        'evolucao_valor_data': json.dumps(evolucao_valor_data),
        
        'status_labels': json.dumps(status_labels, ensure_ascii=False),
        'status_data': json.dumps(status_data),
        
        'uso_contabil_labels': json.dumps(uso_contabil_labels, ensure_ascii=False),
        'uso_contabil_data': json.dumps(uso_contabil_data),

        # Gastos por Previsão de Execução (por ano, 3 barras: Total, SIM, NÃO — mesmo processo do gráfico Evolução setor)
        'gastos_previsao_por_ano_json': gastos_previsao_por_ano_json,
        'anos_grafico_previsao': anos_grafico_previsao,
        'anos_grafico_previsao_json': anos_grafico_previsao_json,

        # Gráfico Valor por Setor
        'valor_por_setor_labels_json': valor_por_setor_labels_json,
        'valor_por_setor_datasets_json': valor_por_setor_datasets_json,

        # Tabelas
        'top_setores': top_setores,
        'top_fornecedores': top_fornecedores,
        'tabela_setor_servico': tabela_setor_servico,
    }

    return render(request, 'orcamento/analise_projecao_gastos.html', context)


def analise_notas_fiscais(request):
    """Página de análise detalhada de notas fiscais com filtros e gráficos"""
    from app.models import NotaFiscal, RelacaoProjecaoNotaFiscal
    from django.db.models import Sum, Count, Q
    from datetime import datetime
    from collections import defaultdict
    from decimal import Decimal
    import json
    
    # --- Lógica de Filtro ---
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')
    uso_contabil_filtro = request.GET.get('uso_contabil', '').strip()
    situacao_filtro = request.GET.get('situacao', '').strip()
    emitente_filtro = request.GET.get('emitente', '').strip()
    centro_filtro = request.GET.get('centro', '').strip()

    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)

    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year

    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))

    if not meses_filtro_int:
        meses_para_mostrar = list(range(1, 13))
    else:
        meses_para_mostrar = meses_filtro_int

    # Função auxiliar para parse de datas
    def parse_date(date_str):
        """Tenta fazer parse de data em vários formatos"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        if not date_str:
            return None
        if ' ' in date_str:
            date_part = date_str.split(' ')[0]
        else:
            date_part = date_str
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(date_part, fmt)
            except (ValueError, TypeError):
                continue
        if '/' in date_part:
            parts = date_part.split('/')
            if len(parts) == 3:
                try:
                    day, month, year = parts
                    if len(year) == 2:
                        year = '20' + year
                    return datetime(int(year), int(month), int(day))
                except (ValueError, TypeError):
                    pass
        return None

    # Anos disponíveis para o filtro (baseado em data_emissao)
    todas_notas_list = NotaFiscal.objects.all()
    anos_set = set()
    for nota in todas_notas_list:
        data_emissao = parse_date(nota.data_emissao)
        if data_emissao:
            anos_set.add(data_emissao.year)
    anos_disponiveis = sorted(list(anos_set), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    if ano_filtro not in anos_disponiveis:
        anos_disponiveis.insert(0, ano_filtro)
        anos_disponiveis = sorted(list(set(anos_disponiveis)), reverse=True)

    meses_choices = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
    ]
    
    # Obter valores únicos para uso_contabil e situacao (via ORM, não da lista filtrada)
    uso_contabil_unicos = list(NotaFiscal.objects.exclude(
        uso_contabil__isnull=True
    ).exclude(uso_contabil='').values_list('uso_contabil', flat=True).distinct().order_by('uso_contabil'))
    uso_contabil_unicos = sorted(set(u.strip() for u in uso_contabil_unicos if u and u.strip()))
    
    situacao_unicas = list(NotaFiscal.objects.exclude(
        situacao__isnull=True
    ).exclude(situacao='').values_list('situacao', flat=True).distinct().order_by('situacao'))
    situacao_unicas = sorted(set(s.strip() for s in situacao_unicas if s and s.strip()))

    # Emitentes e centros para filtro (nome fantasia quando disponível)
    emitentes_set = set()
    for nome, emit in NotaFiscal.objects.exclude(
        Q(nome_fantasia_emitente__isnull=True) & Q(emitente__isnull=True)
    ).values_list('nome_fantasia_emitente', 'emitente'):
        v = (nome or emit or '').strip()
        if v:
            emitentes_set.add(v)
    emitentes_unicos_opcoes = sorted(emitentes_set)
    
    centros_set = set()
    for nome, centro in NotaFiscal.objects.exclude(
        Q(nome_centro_atividade__isnull=True) & Q(centro_atividade__isnull=True)
    ).values_list('nome_centro_atividade', 'centro_atividade'):
        v = (nome or centro or '').strip()
        if v:
            centros_set.add(v)
    centros_unicos_opcoes = sorted(centros_set)

    # Normalização para comparação (evitar problemas com Ç, acentos etc.)
    from unicodedata import normalize as _n
    def _norm(s):
        if not s or not str(s).strip():
            return ''
        return _n('NFKD', str(s).strip()).encode('ASCII', 'ignore').decode('ASCII').lower()

    # --- Queryset Filtrado (baseado APENAS em data_emissao) ---
    # IMPORTANTE: Usar APENAS data_emissao para determinar o mês de pagamento previsto
    notas_filtradas = []
    _uso_norm = _norm(uso_contabil_filtro)
    _situ_norm = _norm(situacao_filtro)
    for nota in todas_notas_list:
        data_emissao = parse_date(nota.data_emissao)
        if not data_emissao or data_emissao.year != ano_filtro:
            continue
        if meses_filtro_int and data_emissao.month not in meses_filtro_int:
            continue
        if _uso_norm and _norm(nota.uso_contabil) != _uso_norm:
            continue
        if _situ_norm and _norm(nota.situacao) != _situ_norm:
            continue
        if emitente_filtro:
            ef = emitente_filtro.lower()
            emit_ok = (nota.emitente and ef in nota.emitente.lower()) or (nota.nome_fantasia_emitente and ef in nota.nome_fantasia_emitente.lower())
            if not emit_ok:
                continue
        if centro_filtro:
            cf = centro_filtro.lower()
            centro_ok = (nota.centro_atividade and cf in nota.centro_atividade.lower()) or (nota.nome_centro_atividade and cf in nota.nome_centro_atividade.lower())
            if not centro_ok:
                continue
        notas_filtradas.append(nota)

    # --- KPIs ---
    total_notas = len(notas_filtradas)
    
    valor_total_notas = sum(
        (nota.total_nota or Decimal(0)) for nota in notas_filtradas
    )
    
    valor_medio = Decimal(0)
    if total_notas > 0:
        valor_medio = valor_total_notas / total_notas

    # Notas com projeção relacionada
    notas_com_projecao_ids = RelacaoProjecaoNotaFiscal.objects.filter(
        nota_fiscal_id__in=[n.id for n in notas_filtradas]
    ).values_list('nota_fiscal_id', flat=True).distinct()
    notas_com_projecao = len(notas_com_projecao_ids)
    percentual_com_projecao = (notas_com_projecao / total_notas * 100) if total_notas > 0 else 0

    # Emitentes únicos
    emitentes_unicos = len(set(
        nota.emitente for nota in notas_filtradas if nota.emitente
    ))

    # Notas vencidas (data_vencimento < hoje)
    hoje_date = hoje.date()
    notas_vencidas = sum(
        1 for nota in notas_filtradas
        if parse_date(nota.data_vencimento) and parse_date(nota.data_vencimento).date() < hoje_date
    )
    percentual_vencidas = (notas_vencidas / total_notas * 100) if total_notas > 0 else 0

    # Notas pagas (situação contém "paga" ou similar)
    notas_pagas = sum(
        1 for nota in notas_filtradas
        if nota.situacao and ('paga' in nota.situacao.lower() or 'liquidada' in nota.situacao.lower() or 'quitada' in nota.situacao.lower())
    )
    percentual_pagas = (notas_pagas / total_notas * 100) if total_notas > 0 else 0

    # Centros de atividade únicos
    centros_atividade_unicos = len(set(
        nota.centro_atividade for nota in notas_filtradas if nota.centro_atividade
    ))

    # Contadores por Situação
    notas_autorizadas = sum(
        1 for nota in notas_filtradas
        if nota.situacao and ('AUTORIZADA' in nota.situacao.upper() and 'AGUARDANDO' not in nota.situacao.upper())
    )
    valor_autorizadas = sum(
        (nota.total_nota or Decimal(0)) for nota in notas_filtradas
        if nota.situacao and ('AUTORIZADA' in nota.situacao.upper() and 'AGUARDANDO' not in nota.situacao.upper())
    )
    percentual_autorizadas = (notas_autorizadas / total_notas * 100) if total_notas > 0 else 0

    notas_lancadas = sum(
        1 for nota in notas_filtradas
        if nota.situacao and ('LANÇADA' in nota.situacao.upper() or 'LANCADA' in nota.situacao.upper())
    )
    valor_lancadas = sum(
        (nota.total_nota or Decimal(0)) for nota in notas_filtradas
        if nota.situacao and ('LANÇADA' in nota.situacao.upper() or 'LANCADA' in nota.situacao.upper())
    )
    percentual_lancadas = (notas_lancadas / total_notas * 100) if total_notas > 0 else 0

    notas_aguardando = sum(
        1 for nota in notas_filtradas
        if nota.situacao and ('AGUARDANDO AUTORIZAÇÃO' in nota.situacao.upper() or 'AGUARDANDO AUTORIZACAO' in nota.situacao.upper())
    )
    valor_aguardando = sum(
        (nota.total_nota or Decimal(0)) for nota in notas_filtradas
        if nota.situacao and ('AGUARDANDO AUTORIZAÇÃO' in nota.situacao.upper() or 'AGUARDANDO AUTORIZACAO' in nota.situacao.upper())
    )
    percentual_aguardando = (notas_aguardando / total_notas * 100) if total_notas > 0 else 0

    notas_pendentes = sum(
        1 for nota in notas_filtradas
        if nota.situacao and 'PENDENTE' in nota.situacao.upper()
    )
    valor_pendentes = sum(
        (nota.total_nota or Decimal(0)) for nota in notas_filtradas
        if nota.situacao and 'PENDENTE' in nota.situacao.upper()
    )
    percentual_pendentes = (notas_pendentes / total_notas * 100) if total_notas > 0 else 0

    # --- Dados para Gráficos ---
    # Gráfico 1: Distribuição por Emitente (Top 10)
    emitentes_data_dict = defaultdict(Decimal)
    for nota in notas_filtradas:
        emitente_key = nota.nome_fantasia_emitente or nota.emitente or 'Não informado'
        if emitente_key:
            emitentes_data_dict[emitente_key] += (nota.total_nota or Decimal(0))
    
    emitentes_sorted = sorted(emitentes_data_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    emitentes_labels = [e[0][:30] + '...' if len(e[0]) > 30 else e[0] for e in emitentes_sorted]
    emitentes_data = [float(e[1]) for e in emitentes_sorted]

    # Gráfico 2: Top 10 Emitentes por Valor (mesmo que acima, mas para bar chart)
    top_emitentes_labels = emitentes_labels
    top_emitentes_data = emitentes_data

    # Gráfico 3: Evolução Temporal (Mês a Mês)
    evolucao_labels = []
    evolucao_quantidade_data = []
    evolucao_valor_data = []
    
    meses_nomes_curtos = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    for mes_num in meses_para_mostrar:
        label = f"{meses_nomes_curtos[mes_num-1]}/{str(ano_filtro)[2:]}"
        evolucao_labels.append(label)
        
        # Quantidade de notas no mês (usar APENAS data_emissao)
        qtd_mes = sum(
            1 for nota in notas_filtradas
            if parse_date(nota.data_emissao) and parse_date(nota.data_emissao).month == mes_num
        )
        evolucao_quantidade_data.append(qtd_mes)
        
        # Valor total do mês (usar APENAS data_emissao)
        valor_mes = sum(
            (nota.total_nota or Decimal(0)) for nota in notas_filtradas
            if parse_date(nota.data_emissao) and parse_date(nota.data_emissao).month == mes_num
        )
        evolucao_valor_data.append(float(valor_mes))

    # Stats por Situação (6 boxes fixos): ordem de checagem para categorias mutuamente exclusivas
    # LANÇADA, PENDENTE, AUTORIZADA, AGUARDANDO AUTORIZAÇÃO, CANCELADA, Sem situação (vazio)
    def classify_situacao(s):
        if not s or not s.strip():
            return 'Sem situação'
        su = s.upper()
        if 'AGUARDANDO AUTORIZAÇÃO' in su or 'AGUARDANDO AUTORIZACAO' in su:
            return 'AGUARDANDO AUTORIZAÇÃO'
        if 'LANÇADA' in su or 'LANCADA' in su:
            return 'LANÇADA'
        if 'PENDENTE' in su:
            return 'PENDENTE'
        if 'AUTORIZADA' in su:
            return 'AUTORIZADA'
        if 'CANCELADA' in su:
            return 'CANCELADA'
        return 'Sem situação'  # Outros valores → Sem situação
    
    situacao_order = ['LANÇADA', 'PENDENTE', 'AUTORIZADA', 'AGUARDANDO AUTORIZAÇÃO', 'CANCELADA', 'Sem situação']
    situacao_counts = defaultdict(lambda: {'quantidade': 0, 'valor_total': Decimal(0)})
    for nota in notas_filtradas:
        cat = classify_situacao(nota.situacao)
        situacao_counts[cat]['quantidade'] += 1
        situacao_counts[cat]['valor_total'] += (nota.total_nota or Decimal(0))
    
    situacao_stats = []
    for label in situacao_order:
        d = situacao_counts[label]
        situacao_stats.append({
            'situacao': label,
            'quantidade': d['quantidade'],
            'valor_total': d['valor_total'],
            'percentual': (d['quantidade'] / total_notas * 100) if total_notas > 0 else 0,
        })
    
    # Gráfico 4: Distribuição por Situação
    situacao_data_dict = defaultdict(int)
    for nota in notas_filtradas:
        situacao_key = nota.situacao or 'Não informado'
        situacao_data_dict[situacao_key] += 1
    
    situacao_sorted = sorted(situacao_data_dict.items(), key=lambda x: x[1], reverse=True)[:6]
    situacao_labels = [s[0][:40] + '...' if len(s[0]) > 40 else s[0] for s in situacao_sorted]
    situacao_data = [s[1] for s in situacao_sorted]

    # Gráfico 5: Distribuição por Uso Contábil
    uso_contabil_data_dict = defaultdict(Decimal)
    for nota in notas_filtradas:
        if nota.uso_contabil:
            uso_contabil_data_dict[nota.uso_contabil] += (nota.total_nota or Decimal(0))
    
    uso_contabil_sorted = sorted(uso_contabil_data_dict.items(), key=lambda x: x[1], reverse=True)[:15]
    uso_contabil_labels = [u[0][:40] + '...' if len(u[0]) > 40 else u[0] for u in uso_contabil_sorted]
    uso_contabil_data = [float(u[1]) for u in uso_contabil_sorted]

    # Gráfico 6: Distribuição por Centro de Atividade
    centro_atividade_data_dict = defaultdict(Decimal)
    for nota in notas_filtradas:
        centro_key = nota.nome_centro_atividade or nota.centro_atividade or 'Não informado'
        if centro_key:
            centro_atividade_data_dict[centro_key] += (nota.total_nota or Decimal(0))
    
    centro_atividade_sorted = sorted(centro_atividade_data_dict.items(), key=lambda x: x[1], reverse=True)[:10]
    centro_atividade_labels = [c[0][:30] + '...' if len(c[0]) > 30 else c[0] for c in centro_atividade_sorted]
    centro_atividade_data = [float(c[1]) for c in centro_atividade_sorted]

    # --- Tabelas Detalhadas ---
    # Top 10 Emitentes por Valor
    top_emitentes = []
    emitentes_dict = defaultdict(lambda: {'quantidade': 0, 'valor_total': Decimal(0), 'nome': ''})
    for nota in notas_filtradas:
        emitente_key = nota.emitente or 'Não informado'
        nome_key = nota.nome_fantasia_emitente or emitente_key
        emitentes_dict[emitente_key]['quantidade'] += 1
        emitentes_dict[emitente_key]['valor_total'] += (nota.total_nota or Decimal(0))
        emitentes_dict[emitente_key]['nome'] = nome_key
    
    top_emitentes = sorted(
        [
            {
                'emitente': k,
                'nome_fantasia_emitente': v['nome'],
                'quantidade': v['quantidade'],
                'valor_total': v['valor_total']
            }
            for k, v in emitentes_dict.items()
        ],
        key=lambda x: x['valor_total'],
        reverse=True
    )[:10]

    # Top 10 Centros de Atividade por Valor
    top_centros_atividade = []
    centros_dict = defaultdict(lambda: {'quantidade': 0, 'valor_total': Decimal(0), 'nome': ''})
    for nota in notas_filtradas:
        centro_key = nota.centro_atividade or 'Não informado'
        nome_key = nota.nome_centro_atividade or centro_key
        centros_dict[centro_key]['quantidade'] += 1
        centros_dict[centro_key]['valor_total'] += (nota.total_nota or Decimal(0))
        centros_dict[centro_key]['nome'] = nome_key
    
    top_centros_atividade = sorted(
        [
            {
                'centro_atividade': k,
                'nome_centro_atividade': v['nome'],
                'quantidade': v['quantidade'],
                'valor_total': v['valor_total']
            }
            for k, v in centros_dict.items()
        ],
        key=lambda x: x['valor_total'],
        reverse=True
    )[:10]

    context = {
        'page_title': 'Análise de Notas Fiscais',
        'active_page': 'analise_notas_fiscais',
        'anos_disponiveis': anos_disponiveis,
        'ano_selecionado': ano_filtro,
        'meses_choices': meses_choices,
        'meses_selecionados': meses_filtro_int,
        'uso_contabil_unicos': uso_contabil_unicos,
        'situacao_unicas': situacao_unicas,
        'emitentes_unicos_opcoes': emitentes_unicos_opcoes,
        'centros_unicos_opcoes': centros_unicos_opcoes,
        'uso_contabil_filtro': uso_contabil_filtro,
        'situacao_filtro': situacao_filtro,
        'emitente_filtro': emitente_filtro,
        'centro_filtro': centro_filtro,

        # KPIs
        'total_notas': total_notas,
        'valor_total_notas': valor_total_notas,
        'valor_medio': valor_medio,
        'notas_com_projecao': notas_com_projecao,
        'percentual_com_projecao': percentual_com_projecao,
        'emitentes_unicos': emitentes_unicos,
        'notas_vencidas': notas_vencidas,
        'percentual_vencidas': percentual_vencidas,
        'notas_pagas': notas_pagas,
        'percentual_pagas': percentual_pagas,
        'centros_atividade_unicos': centros_atividade_unicos,
        
        # Stats por Situação (dinâmico)
        'situacao_stats': situacao_stats,
        
        # Situações
        'notas_autorizadas': notas_autorizadas,
        'valor_autorizadas': valor_autorizadas,
        'percentual_autorizadas': percentual_autorizadas,
        'notas_lancadas': notas_lancadas,
        'valor_lancadas': valor_lancadas,
        'percentual_lancadas': percentual_lancadas,
        'notas_aguardando': notas_aguardando,
        'valor_aguardando': valor_aguardando,
        'percentual_aguardando': percentual_aguardando,
        'notas_pendentes': notas_pendentes,
        'valor_pendentes': valor_pendentes,
        'percentual_pendentes': percentual_pendentes,

        # Gráficos
        'emitentes_labels': json.dumps(emitentes_labels, ensure_ascii=False),
        'emitentes_data': json.dumps(emitentes_data),
        
        'top_emitentes_labels': json.dumps(top_emitentes_labels, ensure_ascii=False),
        'top_emitentes_data': json.dumps(top_emitentes_data),
        
        'evolucao_labels': json.dumps(evolucao_labels, ensure_ascii=False),
        'evolucao_quantidade_data': json.dumps(evolucao_quantidade_data),
        'evolucao_valor_data': json.dumps(evolucao_valor_data),
        
        'situacao_labels': json.dumps(situacao_labels, ensure_ascii=False),
        'situacao_data': json.dumps(situacao_data),
        
        'uso_contabil_labels': json.dumps(uso_contabil_labels, ensure_ascii=False),
        'uso_contabil_data': json.dumps(uso_contabil_data),

        'centro_atividade_labels': json.dumps(centro_atividade_labels, ensure_ascii=False),
        'centro_atividade_data': json.dumps(centro_atividade_data),

        # Tabelas
        'top_emitentes': top_emitentes,
        'top_centros_atividade': top_centros_atividade,
    }

    return render(request, 'orcamento/analise_notas_fiscais.html', context)


def relacionar_projecao_nota_fiscal(request):
    """Relacionar Projeções de Gastos com Notas Fiscais e Controle RC e NF - encontrar matches e confirmar relações"""
    from app.models import ProjecaoGasto, NotaFiscal, ControleRCeNF, RelacaoProjecaoNotaFiscal
    from django.db.models import Q
    from decimal import Decimal
    from django.contrib import messages
    
    # Processar ações POST
    if request.method == 'POST':
        acao = request.POST.get('acao')
        projecao_id = request.POST.get('projecao_id')
        nota_id = request.POST.get('nota_id')
        
        if acao == 'confirmar':
            try:
                projecao = ProjecaoGasto.objects.get(id=projecao_id)
                nota = NotaFiscal.objects.get(id=nota_id)
                controle_id = request.POST.get('controle_id', '').strip()
                score_match = request.POST.get('score_match', '0')
                observacoes = request.POST.get('observacoes', '').strip()
                
                controle_rc_nf = None
                if controle_id:
                    try:
                        controle_rc_nf = ControleRCeNF.objects.get(id_excel=controle_id)
                    except ControleRCeNF.DoesNotExist:
                        pass
                
                relacao, created = RelacaoProjecaoNotaFiscal.objects.update_or_create(
                    projecao=projecao,
                    nota_fiscal=nota,
                    controle_rc_nf=controle_rc_nf,
                    defaults={
                        'status': 'confirmado',
                        'score_match': Decimal(score_match) if score_match else None,
                        'confirmado_por': request.user.username if request.user.is_authenticated else 'Anônimo',
                        'observacoes': observacoes if observacoes else None,
                    }
                )
                
                if created:
                    messages.success(request, f'Relação confirmada com sucesso entre Projeção e Nota Fiscal {nota.nota}!')
                else:
                    messages.success(request, f'Relação atualizada com sucesso!')
                    
            except ProjecaoGasto.DoesNotExist:
                messages.error(request, 'Projeção de gasto não encontrada.')
            except NotaFiscal.DoesNotExist:
                messages.error(request, 'Nota fiscal não encontrada.')
            except Exception as e:
                messages.error(request, f'Erro ao confirmar relação: {str(e)}')
        
        elif acao == 'rejeitar':
            try:
                projecao = ProjecaoGasto.objects.get(id=projecao_id)
                nota = NotaFiscal.objects.get(id=nota_id)
                controle_id = request.POST.get('controle_id', '').strip()
                observacoes = request.POST.get('observacoes', '').strip()
                
                controle_rc_nf = None
                if controle_id:
                    try:
                        controle_rc_nf = ControleRCeNF.objects.get(id_excel=controle_id)
                    except ControleRCeNF.DoesNotExist:
                        pass
                
                relacao, created = RelacaoProjecaoNotaFiscal.objects.update_or_create(
                    projecao=projecao,
                    nota_fiscal=nota,
                    controle_rc_nf=controle_rc_nf,
                    defaults={
                        'status': 'rejeitado',
                        'confirmado_por': request.user.username if request.user.is_authenticated else 'Anônimo',
                        'observacoes': observacoes if observacoes else None,
                    }
                )
                
                messages.success(request, f'Relação rejeitada com sucesso.')
                    
            except ProjecaoGasto.DoesNotExist:
                messages.error(request, 'Projeção de gasto não encontrada.')
            except NotaFiscal.DoesNotExist:
                messages.error(request, 'Nota fiscal não encontrada.')
            except Exception as e:
                messages.error(request, f'Erro ao rejeitar relação: {str(e)}')
        
        elif acao == 'remover':
            try:
                projecao = ProjecaoGasto.objects.get(id=projecao_id)
                nota = NotaFiscal.objects.get(id=nota_id)
                controle_id = request.POST.get('controle_id', '').strip()
                
                filtro = Q(projecao=projecao, nota_fiscal=nota)
                if controle_id:
                    try:
                        controle_rc_nf = ControleRCeNF.objects.get(id_excel=controle_id)
                        filtro &= Q(controle_rc_nf=controle_rc_nf)
                    except ControleRCeNF.DoesNotExist:
                        pass
                
                RelacaoProjecaoNotaFiscal.objects.filter(filtro).delete()
                
                messages.success(request, f'Relação removida com sucesso.')
                    
            except Exception as e:
                messages.error(request, f'Erro ao remover relação: {str(e)}')
        
        return redirect('relacionar_projecao_nota_fiscal')
    
    # Função para calcular score de match (duas tabelas - original)
    def calcular_match_score(projecao, nota):
        """Calcula score de correspondência entre projeção e nota fiscal (0-100)"""
        score = 0
        total_peso = 0
        detalhes = []
        
        # Valor (peso 30)
        if projecao.valor_total and nota.total_nota:
            total_peso += 30
            diff_percent = abs(float(projecao.valor_total - nota.total_nota)) / float(projecao.valor_total) * 100
            if diff_percent <= 5:  # Diferença de até 5%
                score += 30
                detalhes.append(f"Valor: Match perfeito (diff: {diff_percent:.1f}%)")
            elif diff_percent <= 10:
                score += 20
                detalhes.append(f"Valor: Match bom (diff: {diff_percent:.1f}%)")
            elif diff_percent <= 20:
                score += 10
                detalhes.append(f"Valor: Match parcial (diff: {diff_percent:.1f}%)")
            else:
                detalhes.append(f"Valor: Diferença significativa (diff: {diff_percent:.1f}%)")
        
        # Fornecedor/Emitente (peso 25)
        fornecedor_proj = (projecao.fornecedor_nome_fantasia or '').strip().upper()
        emitente_nota = (nota.nome_fantasia_emitente or nota.emitente or '').strip().upper()
        if fornecedor_proj and emitente_nota:
            total_peso += 25
            if fornecedor_proj == emitente_nota:
                score += 25
                detalhes.append("Fornecedor/Emitente: Match perfeito")
            elif fornecedor_proj in emitente_nota or emitente_nota in fornecedor_proj:
                score += 15
                detalhes.append("Fornecedor/Emitente: Match parcial")
            else:
                detalhes.append("Fornecedor/Emitente: Não corresponde")
        
        # CNPJ (peso 20)
        cnpj_proj = (projecao.fornecedor_cnpj or '').strip().replace('.', '').replace('/', '').replace('-', '')
        cnpj_nota = (nota.emitente or '').strip().replace('.', '').replace('/', '').replace('-', '')
        if cnpj_proj and cnpj_nota:
            total_peso += 20
            if cnpj_proj == cnpj_nota:
                score += 20
                detalhes.append("CNPJ: Match perfeito")
            else:
                detalhes.append("CNPJ: Não corresponde")
        
        # Data (peso 15)
        if projecao.data_abertura_requisicao and nota.data_emissao:
            total_peso += 15
            # Tentar parsear data_emissao que é string
            try:
                from datetime import datetime
                # Tentar diferentes formatos de data
                data_nota_str = nota.data_emissao.strip()
                data_nota = None
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        data_nota = datetime.strptime(data_nota_str, fmt).date()
                        break
                    except ValueError:
                        continue
                
                if data_nota:
                    diff_dias = abs((projecao.data_abertura_requisicao - data_nota).days)
                    if diff_dias <= 7:
                        score += 15
                        detalhes.append(f"Data: Match próximo ({diff_dias} dias)")
                    elif diff_dias <= 30:
                        score += 10
                        detalhes.append(f"Data: Match razoável ({diff_dias} dias)")
                    elif diff_dias <= 90:
                        score += 5
                        detalhes.append(f"Data: Match distante ({diff_dias} dias)")
                    else:
                        detalhes.append(f"Data: Diferença grande ({diff_dias} dias)")
                else:
                    detalhes.append("Data: Formato não reconhecido")
            except Exception:
                detalhes.append("Data: Erro ao comparar")
        
        # Número NF (peso 10)
        nf_proj = (projecao.numero_nf or '').strip()
        nf_nota = (nota.nota or '').strip()
        if nf_proj and nf_nota:
            total_peso += 10
            if nf_proj == nf_nota:
                score += 10
                detalhes.append("Número NF: Match perfeito")
            else:
                detalhes.append("Número NF: Não corresponde")
        
        if total_peso == 0:
            return 0, []
        
        percentual = (score / total_peso * 100)
        return round(percentual, 2), detalhes
    
    # Função para calcular score de match (três tabelas)
    def calcular_match_score_tres_tabelas(projecao, nota, controle=None):
        """Calcula score de correspondência entre projeção, nota fiscal e controle RC e NF (0-100)"""
        if not controle:
            # Se não há controle, usar função de duas tabelas
            return calcular_match_score(projecao, nota)
        
        score = 0
        total_peso = 0
        detalhes = []
        
        # Valor - comparar entre as três tabelas (peso 30)
        valores = []
        if projecao.valor_total:
            valores.append(('Projeção', projecao.valor_total))
        if nota.total_nota:
            valores.append(('Nota Fiscal', nota.total_nota))
        if controle:
            if controle.valor_nf:
                valores.append(('Controle NF', controle.valor_nf))
            elif controle.valor_total_pedido:
                valores.append(('Controle Pedido', controle.valor_total_pedido))
        
        if len(valores) >= 2:
            total_peso += 30
            # Comparar todos os valores
            valores_decimais = [float(v[1]) for v in valores]
            max_valor = max(valores_decimais)
            min_valor = min(valores_decimais)
            if max_valor > 0:
                diff_percent = abs(max_valor - min_valor) / max_valor * 100
                if diff_percent <= 5:
                    score += 30
                    detalhes.append(f"Valor (3 tabelas): Match perfeito (diff: {diff_percent:.1f}%)")
                elif diff_percent <= 10:
                    score += 20
                    detalhes.append(f"Valor (3 tabelas): Match bom (diff: {diff_percent:.1f}%)")
                elif diff_percent <= 20:
                    score += 10
                    detalhes.append(f"Valor (3 tabelas): Match parcial (diff: {diff_percent:.1f}%)")
                else:
                    detalhes.append(f"Valor (3 tabelas): Diferença significativa (diff: {diff_percent:.1f}%)")
        
        # Fornecedor/Emitente/Empresa - comparar entre as três tabelas (peso 25)
        fornecedores = []
        fornecedor_proj = (projecao.fornecedor_nome_fantasia or '').strip().upper()
        if fornecedor_proj:
            fornecedores.append(('Projeção', fornecedor_proj))
        emitente_nota = (nota.nome_fantasia_emitente or nota.emitente or '').strip().upper()
        if emitente_nota:
            fornecedores.append(('Nota Fiscal', emitente_nota))
        if controle and controle.empresa:
            empresa_controle = (controle.empresa or '').strip().upper()
            if empresa_controle:
                fornecedores.append(('Controle RC/NF', empresa_controle))
        
        if len(fornecedores) >= 2:
            total_peso += 25
            # Verificar se todos são iguais ou similares
            nomes = [f[1] for f in fornecedores]
            if len(set(nomes)) == 1:
                score += 25
                detalhes.append("Fornecedor/Emitente/Empresa (3 tabelas): Match perfeito")
            elif any(nomes[0] in n or n in nomes[0] for n in nomes[1:]):
                score += 15
                detalhes.append("Fornecedor/Emitente/Empresa (3 tabelas): Match parcial")
            else:
                detalhes.append("Fornecedor/Emitente/Empresa (3 tabelas): Não corresponde")
        
        # Número NF - comparar entre as três tabelas (peso 20)
        nfs = []
        nf_proj = (projecao.numero_nf or '').strip()
        if nf_proj:
            nfs.append(('Projeção', nf_proj))
        nf_nota = (nota.nota or '').strip()
        if nf_nota:
            nfs.append(('Nota Fiscal', nf_nota))
        if controle:
            if controle.nf_saida:
                nfs.append(('Controle NF Saída', controle.nf_saida.strip()))
            if controle.nf_servico:
                nfs.append(('Controle NF Serviço', controle.nf_servico.strip()))
        
        if len(nfs) >= 2:
            total_peso += 20
            # Normalizar números de NF para comparação
            from app.utils import normalize_nota_numero
            nfs_normalizados = [(f[0], normalize_nota_numero(f[1])) for f in nfs if f[1]]
            if len(set([n[1] for n in nfs_normalizados])) == 1:
                score += 20
                detalhes.append("Número NF (3 tabelas): Match perfeito")
            else:
                detalhes.append("Número NF (3 tabelas): Não corresponde")
        
        # RC/Pedido - comparar Projeção com Controle (peso 15)
        if controle:
            total_peso += 15
            match_rc_pedido = False
            # Comparar número de requisição
            if projecao.numero_requisicao_compra and controle.rc:
                rc_proj = (projecao.numero_requisicao_compra or '').strip()
                rc_controle = (controle.rc or '').strip()
                if rc_proj == rc_controle:
                    score += 10
                    detalhes.append("RC: Match perfeito")
                    match_rc_pedido = True
            # Comparar número de pedido
            if projecao.numero_pedido_compra and controle.pedido:
                pedido_proj = (projecao.numero_pedido_compra or '').strip()
                pedido_controle = (controle.pedido or '').strip()
                if pedido_proj == pedido_controle:
                    score += 5
                    detalhes.append("Pedido: Match perfeito")
                    match_rc_pedido = True
            if not match_rc_pedido:
                detalhes.append("RC/Pedido: Não corresponde")
        
        # Data - comparar entre as três tabelas (peso 10)
        datas = []
        if projecao.data_abertura_requisicao:
            datas.append(('Projeção', projecao.data_abertura_requisicao))
        if nota.data_emissao:
            try:
                from datetime import datetime
                data_nota_str = nota.data_emissao.strip()
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        data_nota = datetime.strptime(data_nota_str, fmt).date()
                        datas.append(('Nota Fiscal', data_nota))
                        break
                    except ValueError:
                        continue
            except Exception:
                pass
        if controle and controle.data_rc:
            datas.append(('Controle RC', controle.data_rc.date() if hasattr(controle.data_rc, 'date') else controle.data_rc))
        if controle and controle.emissao:
            datas.append(('Controle Emissão', controle.emissao.date() if hasattr(controle.emissao, 'date') else controle.emissao))
        
        if len(datas) >= 2:
            total_peso += 10
            # Calcular diferença máxima entre datas
            from datetime import date
            datas_only = [d[1] for d in datas if isinstance(d[1], date)]
            if len(datas_only) >= 2:
                max_data = max(datas_only)
                min_data = min(datas_only)
                diff_dias = abs((max_data - min_data).days)
                if diff_dias <= 7:
                    score += 10
                    detalhes.append(f"Data (3 tabelas): Match próximo ({diff_dias} dias)")
                elif diff_dias <= 30:
                    score += 5
                    detalhes.append(f"Data (3 tabelas): Match razoável ({diff_dias} dias)")
                else:
                    detalhes.append(f"Data (3 tabelas): Diferença grande ({diff_dias} dias)")
        
        if total_peso == 0:
            return 0, []
        
        percentual = (score / total_peso * 100)
        return round(percentual, 2), detalhes
    
    # Função para calcular score de match (duas tabelas - mantida para compatibilidade)
    def calcular_match_score(projecao, nota):
        """Calcula score de correspondência entre projeção e nota fiscal (0-100)"""
        score = 0
        total_peso = 0
        detalhes = []
        
        # Valor (peso 30)
        if projecao.valor_total and nota.total_nota:
            total_peso += 30
            diff_percent = abs(float(projecao.valor_total - nota.total_nota)) / float(projecao.valor_total) * 100
            if diff_percent <= 5:
                score += 30
                detalhes.append(f"Valor: Match perfeito (diff: {diff_percent:.1f}%)")
            elif diff_percent <= 10:
                score += 20
                detalhes.append(f"Valor: Match bom (diff: {diff_percent:.1f}%)")
            elif diff_percent <= 20:
                score += 10
                detalhes.append(f"Valor: Match parcial (diff: {diff_percent:.1f}%)")
            else:
                detalhes.append(f"Valor: Diferença significativa (diff: {diff_percent:.1f}%)")
        
        # Fornecedor/Emitente (peso 25)
        fornecedor_proj = (projecao.fornecedor_nome_fantasia or '').strip().upper()
        emitente_nota = (nota.nome_fantasia_emitente or nota.emitente or '').strip().upper()
        if fornecedor_proj and emitente_nota:
            total_peso += 25
            if fornecedor_proj == emitente_nota:
                score += 25
                detalhes.append("Fornecedor/Emitente: Match perfeito")
            elif fornecedor_proj in emitente_nota or emitente_nota in fornecedor_proj:
                score += 15
                detalhes.append("Fornecedor/Emitente: Match parcial")
            else:
                detalhes.append("Fornecedor/Emitente: Não corresponde")
        
        # CNPJ (peso 20)
        cnpj_proj = (projecao.fornecedor_cnpj or '').strip().replace('.', '').replace('/', '').replace('-', '')
        cnpj_nota = (nota.emitente or '').strip().replace('.', '').replace('/', '').replace('-', '')
        if cnpj_proj and cnpj_nota:
            total_peso += 20
            if cnpj_proj == cnpj_nota:
                score += 20
                detalhes.append("CNPJ: Match perfeito")
            else:
                detalhes.append("CNPJ: Não corresponde")
        
        # Data (peso 15)
        if projecao.data_abertura_requisicao and nota.data_emissao:
            total_peso += 15
            try:
                from datetime import datetime
                data_nota_str = nota.data_emissao.strip()
                data_nota = None
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        data_nota = datetime.strptime(data_nota_str, fmt).date()
                        break
                    except ValueError:
                        continue
                
                if data_nota:
                    diff_dias = abs((projecao.data_abertura_requisicao - data_nota).days)
                    if diff_dias <= 7:
                        score += 15
                        detalhes.append(f"Data: Match próximo ({diff_dias} dias)")
                    elif diff_dias <= 30:
                        score += 10
                        detalhes.append(f"Data: Match razoável ({diff_dias} dias)")
                    elif diff_dias <= 90:
                        score += 5
                        detalhes.append(f"Data: Match distante ({diff_dias} dias)")
                    else:
                        detalhes.append(f"Data: Diferença grande ({diff_dias} dias)")
                else:
                    detalhes.append("Data: Formato não reconhecido")
            except Exception:
                detalhes.append("Data: Erro ao comparar")
        
        # Número NF (peso 10)
        nf_proj = (projecao.numero_nf or '').strip()
        nf_nota = (nota.nota or '').strip()
        if nf_proj and nf_nota:
            total_peso += 10
            if nf_proj == nf_nota:
                score += 10
                detalhes.append("Número NF: Match perfeito")
            else:
                detalhes.append("Número NF: Não corresponde")
        
        if total_peso == 0:
            return 0, []
        
        percentual = (score / total_peso * 100)
        return round(percentual, 2), detalhes
    
    # Obter filtros
    filtro_centro_atividade = request.GET.get('filtro_centro_atividade', '').strip()
    filtro_ano = request.GET.get('filtro_ano', '').strip()
    filtro_tipo = request.GET.get('filtro_tipo', '').strip()
    filtro_valor_min = request.GET.get('filtro_valor_min', '').strip()
    filtro_valor_max = request.GET.get('filtro_valor_max', '').strip()
    mostrar_apenas_proximos = request.GET.get('mostrar_apenas_proximos', '') == 'on'
    mostrar_apenas_nao_confirmados = request.GET.get('mostrar_apenas_nao_confirmados', '') == 'on'
    
    # Análise invertida: Nota Fiscal como base → encontrar Projeção de Gasto que corresponde
    # Score mínimo para mostrar match (evita pares irrelevantes 0%)
    SCORE_MINIMO = 20
    MAX_MATCHES_POR_NOTA = 5  # Máximo de projeções candidatas por nota
    
    # Buscar projeções com filtros (todas que passam nos filtros - base menor)
    projecoes_qs = ProjecaoGasto.objects.all()
    if filtro_centro_atividade:
        projecoes_qs = projecoes_qs.filter(setor__icontains=filtro_centro_atividade)
    if filtro_ano:
        try:
            ano = int(filtro_ano)
            projecoes_qs = projecoes_qs.filter(ano_referencia=ano)
        except ValueError:
            pass
    if filtro_tipo:
        projecoes_qs = projecoes_qs.filter(tipo_solicitacao__icontains=filtro_tipo)
    if filtro_valor_min:
        try:
            projecoes_qs = projecoes_qs.filter(valor_total__gte=Decimal(filtro_valor_min))
        except (ValueError, TypeError):
            pass
    if filtro_valor_max:
        try:
            projecoes_qs = projecoes_qs.filter(valor_total__lte=Decimal(filtro_valor_max))
        except (ValueError, TypeError):
            pass
    
    projecoes = list(projecoes_qs)
    projecoes_total_count = len(projecoes)
    resultados_limitados = False  # Sem limite artificial com análise por NF
    
    # Buscar notas fiscais separadas por uso_contabil (NF como base da análise)
    notas_242 = list(NotaFiscal.objects.filter(uso_contabil='242'))
    notas_5 = list(NotaFiscal.objects.filter(uso_contabil='5'))
    notas_170 = list(NotaFiscal.objects.filter(uso_contabil='170'))
    
    # Buscar todos os ControleRCeNF para matching (em memória para evitar N+1 no loop)
    controles_rc_nf_list = list(ControleRCeNF.objects.all())
    
    # Índices para lookup O(1) em vez de scan O(n) por par projecao-nota
    from app.utils import normalize_nota_numero
    _normalize = normalize_nota_numero
    controles_by_nf = {}
    controles_by_rc = {}
    controles_by_pedido = {}
    for c in controles_rc_nf_list:
        for nf_val in (_normalize(c.nf_saida), _normalize(c.nf_servico)):
            if nf_val:
                controles_by_nf.setdefault(nf_val, []).append(c)
        rc_val = (c.rc or '').strip()
        if rc_val:
            controles_by_rc.setdefault(rc_val, []).append(c)
        ped_val = (c.pedido or '').strip()
        if ped_val:
            controles_by_pedido.setdefault(ped_val, []).append(c)
    
    # Buscar relações confirmadas (incluindo controle_rc_nf)
    relacoes_confirmadas = RelacaoProjecaoNotaFiscal.objects.filter(
        status='confirmado'
    ).select_related('projecao', 'nota_fiscal', 'controle_rc_nf')
    
    # Criar dicionário de relações confirmadas para lookup rápido
    # Chave: (projecao_id, nota_id, controle_id_excel ou None)
    relacoes_dict = {}
    for relacao in relacoes_confirmadas:
        controle_key = relacao.controle_rc_nf.id_excel if relacao.controle_rc_nf else None
        relacoes_dict[(relacao.projecao.id, relacao.nota_fiscal.id, controle_key)] = relacao
    
    # Função auxiliar para encontrar ControleRCeNF (usa índices para O(1) lookup)
    def encontrar_controle_rc_nf(projecao, nota, idx_nf, idx_rc, idx_pedido, controles_list):
        """Encontra ControleRCeNF que corresponde à Projeção e Nota Fiscal"""
        controles_candidatos = []
        seen_ids = set()
        
        def add_candidatos(lst):
            for c in lst:
                if id(c) not in seen_ids:
                    seen_ids.add(id(c))
                    controles_candidatos.append(c)
                    if len(controles_candidatos) >= 3:
                        return
        
        # Buscar por número de NF (lookup O(1))
        nf_busca = _normalize(projecao.numero_nf) or _normalize(nota.nota)
        if nf_busca and nf_busca in idx_nf:
            add_candidatos(idx_nf[nf_busca])
            if len(controles_candidatos) >= 3:
                return controles_candidatos[:3]
        
        # Buscar por RC (lookup O(1))
        rc_busca = (projecao.numero_requisicao_compra or '').strip()
        if rc_busca and rc_busca in idx_rc:
            add_candidatos(idx_rc[rc_busca])
            if len(controles_candidatos) >= 3:
                return controles_candidatos[:3]
        
        # Buscar por Pedido (lookup O(1))
        pedido_busca = (projecao.numero_pedido_compra or '').strip()
        if pedido_busca and pedido_busca in idx_pedido:
            add_candidatos(idx_pedido[pedido_busca])
        
        # Se não encontrou, fallback: buscar por empresa/fornecedor
        if not controles_candidatos and (projecao.fornecedor_nome_fantasia or nota.nome_fantasia_emitente):
            fornecedor_busca = (projecao.fornecedor_nome_fantasia or nota.nome_fantasia_emitente or '').strip().upper()
            if fornecedor_busca:
                for controle in controles_list:
                    if (controle.empresa or '').strip().upper().find(fornecedor_busca) >= 0:
                        controles_candidatos.append(controle)
                        if len(controles_candidatos) >= 3:
                            break
        
        return controles_candidatos[:3]
    
    # Análise por Nota Fiscal: para cada NF, encontrar Projeções que correspondem
    # Score mínimo: SCORE_MINIMO (ou 50 se "apenas próximos" ativo)
    def calcular_matches_para_notas(notas_list):
        score_min = 50 if mostrar_apenas_proximos else SCORE_MINIMO
        matches_list = []
        for nota in notas_list:
            candidatos_nota = []
            for projecao in projecoes:
                controles_candidatos = encontrar_controle_rc_nf(
                    projecao, nota,
                    controles_by_nf, controles_by_rc, controles_by_pedido,
                    controles_rc_nf_list
                )
                controle = controles_candidatos[0] if controles_candidatos else None
                controle_key = controle.id_excel if controle else None
                relacao_confirmada = relacoes_dict.get((projecao.id, nota.id, controle_key))
                if mostrar_apenas_nao_confirmados and relacao_confirmada:
                    continue
                score, detalhes = calcular_match_score_tres_tabelas(projecao, nota, controle)
                if score < score_min:
                    continue
                candidatos_nota.append({
                    'projecao': projecao,
                    'nota': nota,
                    'controle': controle,
                    'match_info': {'percentual': score, 'detalhes': detalhes},
                    'relacao_confirmada': relacao_confirmada,
                })
            # Ordenar por score e pegar até MAX_MATCHES_POR_NOTA por nota
            candidatos_nota.sort(key=lambda x: x['match_info']['percentual'], reverse=True)
            matches_list.extend(candidatos_nota[:MAX_MATCHES_POR_NOTA])
        matches_list.sort(key=lambda x: x['match_info']['percentual'], reverse=True)
        return matches_list
    
    # Calcular matches para cada uso_contabil
    matches_242 = calcular_matches_para_notas(notas_242)
    matches_5 = calcular_matches_para_notas(notas_5)
    matches_170 = calcular_matches_para_notas(notas_170)
    
    # Paginação por seção (25 matches por página)
    from django.core.paginator import Paginator
    PER_PAGE = 25
    page_242 = request.GET.get('page_242', 1)
    page_5 = request.GET.get('page_5', 1)
    page_170 = request.GET.get('page_170', 1)
    try:
        page_242 = max(1, int(page_242))
        page_5 = max(1, int(page_5))
        page_170 = max(1, int(page_170))
    except (ValueError, TypeError):
        page_242 = page_5 = page_170 = 1
    
    paginator_242 = Paginator(matches_242, PER_PAGE)
    paginator_5 = Paginator(matches_5, PER_PAGE)
    paginator_170 = Paginator(matches_170, PER_PAGE)
    matches_242_page = paginator_242.get_page(page_242)
    matches_5_page = paginator_5.get_page(page_5)
    matches_170_page = paginator_170.get_page(page_170)
    
    # Estatísticas (notas são listas após o limit)
    total_projecoes = projecoes_total_count
    total_notas_242 = len(notas_242)
    total_notas_5 = len(notas_5)
    total_notas_170 = len(notas_170)
    total_notas = total_notas_242 + total_notas_5 + total_notas_170
    total_controles_rc_nf = len(controles_rc_nf_list)
    total_relacoes_confirmadas = RelacaoProjecaoNotaFiscal.objects.filter(status='confirmado').count()
    
    context = {
        'page_title': 'Relacionar Projeção vs Nota Fiscal vs Controle RC/NF',
        'active_page': 'relacionar_projecao_nota_fiscal',
        'matches_242': matches_242_page,
        'matches_5': matches_5_page,
        'matches_170': matches_170_page,
        'matches_242_count': len(matches_242),
        'matches_5_count': len(matches_5),
        'matches_170_count': len(matches_170),
        'total_projecoes': total_projecoes,
        'total_notas': total_notas,
        'total_notas_242': total_notas_242,
        'total_notas_5': total_notas_5,
        'total_notas_170': total_notas_170,
        'total_controles_rc_nf': total_controles_rc_nf,
        'total_relacoes_confirmadas': total_relacoes_confirmadas,
        'filtro_centro_atividade': filtro_centro_atividade,
        'filtro_ano': filtro_ano,
        'filtro_tipo': filtro_tipo,
        'filtro_valor_min': filtro_valor_min,
        'filtro_valor_max': filtro_valor_max,
        'mostrar_apenas_proximos': mostrar_apenas_proximos,
        'mostrar_apenas_nao_confirmados': mostrar_apenas_nao_confirmados,
        'resultados_limitados': resultados_limitados,
        'score_minimo': SCORE_MINIMO,
        'max_matches_por_nota': MAX_MATCHES_POR_NOTA,
        'active_uc_tab': '170' if page_170 > 1 else ('5' if page_5 > 1 else '242'),
    }
    
    return render(request, 'orcamento/relacionar_projecao_nota_fiscal.html', context)


def visualizar_relacao_projecao_nota(request):
    """Visualizar detalhes da relação entre uma Projeção de Gasto e uma Nota Fiscal"""
    from app.models import ProjecaoGasto, NotaFiscal, RelacaoProjecaoNotaFiscal
    from django.contrib import messages
    from datetime import datetime
    from decimal import Decimal
    
    # Obter IDs dos parâmetros
    projecao_id = request.GET.get('projecao_id')
    nota_id = request.GET.get('nota_id')
    
    if not projecao_id or not nota_id:
        messages.error(request, 'Parâmetros projecao_id e nota_id são obrigatórios.')
        return redirect('relacionar_projecao_nota_fiscal')
    
    try:
        projecao = ProjecaoGasto.objects.get(id=projecao_id)
        nota = NotaFiscal.objects.get(id=nota_id)
    except ProjecaoGasto.DoesNotExist:
        messages.error(request, 'Projeção de gasto não encontrada.')
        return redirect('relacionar_projecao_nota_fiscal')
    except NotaFiscal.DoesNotExist:
        messages.error(request, 'Nota fiscal não encontrada.')
        return redirect('relacionar_projecao_nota_fiscal')
    
    # Buscar relação se existir
    relacao = None
    try:
        relacao = RelacaoProjecaoNotaFiscal.objects.get(projecao=projecao, nota_fiscal=nota)
    except RelacaoProjecaoNotaFiscal.DoesNotExist:
        pass
    
    # Calcular match score e criar estrutura detalhada
    score = 0
    total_peso = 0
    comparacoes = {}
    
    # Centro de Atividade (setor)
    centro_proj = (projecao.setor or '').strip()
    centro_nota = (nota.centro_atividade or '').strip()
    if centro_proj and centro_nota:
        if centro_proj == centro_nota:
            comparacoes['centro_atividade'] = {
                'projecao': centro_proj,
                'nota': centro_nota,
                'status': 'exato'
            }
        elif centro_proj in centro_nota or centro_nota in centro_proj:
            comparacoes['centro_atividade'] = {
                'projecao': centro_proj,
                'nota': centro_nota,
                'status': 'parcial'
            }
        else:
            comparacoes['centro_atividade'] = {
                'projecao': centro_proj,
                'nota': centro_nota,
                'status': 'diferente'
            }
    elif centro_proj or centro_nota:
        comparacoes['centro_atividade'] = {
            'projecao': centro_proj or '-',
            'nota': centro_nota or '-',
            'status': 'diferente'
        }
    
    # Valor
    valor_proj = projecao.valor_total
    valor_nota = nota.total_nota
    if valor_proj and valor_nota:
        total_peso += 30
        diff = abs(float(valor_proj - valor_nota))
        diff_percent = (diff / float(valor_proj)) * 100
        
        if diff_percent <= 5:
            score += 30
            status_valor = 'match'
        elif diff_percent <= 10:
            score += 20
            status_valor = 'parcial'
        elif diff_percent <= 20:
            score += 10
            status_valor = 'parcial'
        else:
            status_valor = 'diferente'
        
        comparacoes['valor'] = {
            'tipo': 'Total',
            'projecao': float(valor_proj),
            'nota': float(valor_nota),
            'diferenca': diff,
            'diferenca_percent': diff_percent,
            'status': status_valor
        }
    elif valor_proj or valor_nota:
        comparacoes['valor'] = {
            'tipo': 'Total',
            'projecao': float(valor_proj) if valor_proj else None,
            'nota': float(valor_nota) if valor_nota else None,
            'status': 'diferente'
        }
    
    # Data
    data_proj = projecao.data_abertura_requisicao
    data_nota = None
    if nota.data_emissao:
        data_nota_str = nota.data_emissao.strip()
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                data_nota = datetime.strptime(data_nota_str, fmt).date()
                break
            except ValueError:
                continue
    
    if data_proj and data_nota:
        total_peso += 15
        diff_dias = abs((data_proj - data_nota).days)
        if diff_dias <= 7:
            score += 15
            status_data = 'exato'
        elif diff_dias <= 30:
            score += 10
            status_data = 'proximo'
        elif diff_dias <= 90:
            score += 5
            status_data = 'parcial'
        else:
            status_data = 'diferente'
        
        comparacoes['data'] = {
            'projecao': data_proj,
            'nota': data_nota,
            'diff_dias': diff_dias,
            'status': status_data
        }
    elif data_proj or data_nota:
        comparacoes['data'] = {
            'projecao': data_proj,
            'nota': data_nota,
            'status': 'diferente'
        }
    
    # Fornecedor/Emitente
    fornecedor_proj = (projecao.fornecedor_nome_fantasia or '').strip()
    emitente_nota = (nota.nome_fantasia_emitente or nota.emitente or '').strip()
    if fornecedor_proj and emitente_nota:
        total_peso += 25
        fornecedor_proj_upper = fornecedor_proj.upper()
        emitente_nota_upper = emitente_nota.upper()
        if fornecedor_proj_upper == emitente_nota_upper:
            score += 25
            status_fornecedor = 'exato'
        elif fornecedor_proj_upper in emitente_nota_upper or emitente_nota_upper in fornecedor_proj_upper:
            score += 15
            status_fornecedor = 'parcial'
        else:
            status_fornecedor = 'diferente'
        
        comparacoes['fornecedor'] = {
            'projecao': fornecedor_proj,
            'nota': emitente_nota,
            'status': status_fornecedor
        }
    elif fornecedor_proj or emitente_nota:
        comparacoes['fornecedor'] = {
            'projecao': fornecedor_proj or '-',
            'nota': emitente_nota or '-',
            'status': 'diferente'
        }
    
    # Número (Requisição vs Nota Fiscal)
    numero_proj = (projecao.numero_requisicao_compra or projecao.numero_nf or '').strip()
    numero_nota = (nota.nota or '').strip()
    if numero_proj and numero_nota:
        total_peso += 10
        if numero_proj == numero_nota:
            score += 10
            status_numero = 'exato'
        else:
            status_numero = 'diferente'
        
        comparacoes['numero'] = {
            'projecao': numero_proj,
            'nota': numero_nota,
            'status': status_numero
        }
    elif numero_proj or numero_nota:
        comparacoes['numero'] = {
            'projecao': numero_proj or '-',
            'nota': numero_nota or '-',
            'status': 'diferente'
        }
    
    # Calcular percentual
    if total_peso == 0:
        percentual = 0
    else:
        percentual = (score / total_peso * 100)
    
    match_info = {
        'percentual': round(percentual, 2),
        'score': score,
        'max_score': total_peso,
        'comparacoes': comparacoes
    }
    
    context = {
        'page_title': 'Visualizar Relação Projeção vs Nota Fiscal',
        'active_page': 'relacionar_projecao_nota_fiscal',
        'projecao': projecao,
        'nota': nota,
        'match_info': match_info,
        'relacao': relacao,
    }
    
    return render(request, 'visualizar/visualizar_relacao_projecao_nota.html', context)


def consultar_planilha_rc(request):
    """Consultar/listar dados do Controle RC e NF com filtros"""
    from app.models import ControleRCeNF
    from django.db.models import Q, Sum, Count
    from decimal import Decimal
    from django.core.paginator import Paginator
    
    # Busca geral
    search_query = request.GET.get('search', '').strip()
    controles_list = ControleRCeNF.objects.all()
    
    # Aplicar busca geral
    if search_query:
        try:
            search_num = Decimal(search_query.replace(',', '.'))
            controles_list = controles_list.filter(
                Q(id_excel__icontains=search_query) |
                Q(solicitante__icontains=search_query) |
                Q(empresa__icontains=search_query) |
                Q(nf_saida__icontains=search_query) |
                Q(descricao_servico__icontains=search_query) |
                Q(rc__icontains=search_query) |
                Q(pedido__icontains=search_query) |
                Q(os__icontains=search_query) |
                Q(valor_total_pedido=search_num) |
                Q(valor_nf=search_num)
            )
        except (ValueError, TypeError):
            controles_list = controles_list.filter(
                Q(id_excel__icontains=search_query) |
                Q(solicitante__icontains=search_query) |
                Q(empresa__icontains=search_query) |
                Q(nf_saida__icontains=search_query) |
                Q(descricao_servico__icontains=search_query) |
                Q(rc__icontains=search_query) |
                Q(pedido__icontains=search_query) |
                Q(os__icontains=search_query) |
                Q(status__icontains=search_query)
            )
    
    # Filtros específicos
    filtro_id = request.GET.get('filtro_id', '').strip()
    if filtro_id:
        controles_list = controles_list.filter(id_excel__icontains=filtro_id)
    
    filtro_empresa = request.GET.get('filtro_empresa', '').strip()
    if filtro_empresa:
        controles_list = controles_list.filter(empresa__icontains=filtro_empresa)
    
    filtro_solicitante = request.GET.get('filtro_solicitante', '').strip()
    if filtro_solicitante:
        controles_list = controles_list.filter(solicitante__icontains=filtro_solicitante)
    
    filtro_rc = request.GET.get('filtro_rc', '').strip()
    if filtro_rc:
        controles_list = controles_list.filter(rc__icontains=filtro_rc)
    
    filtro_os = request.GET.get('filtro_os', '').strip()
    if filtro_os:
        controles_list = controles_list.filter(os__icontains=filtro_os)
    
    filtro_status = request.GET.get('filtro_status', '').strip()
    if filtro_status:
        controles_list = controles_list.filter(status__icontains=filtro_status)
    
    filtro_uso = request.GET.get('filtro_uso', '').strip()
    if filtro_uso:
        controles_list = controles_list.filter(uso__icontains=filtro_uso)
    
    filtro_valor_min = request.GET.get('filtro_valor_min', '').strip()
    if filtro_valor_min:
        try:
            controles_list = controles_list.filter(
                Q(valor_total_pedido__gte=Decimal(filtro_valor_min.replace(',', '.'))) |
                Q(valor_nf__gte=Decimal(filtro_valor_min.replace(',', '.')))
            )
        except (ValueError, TypeError):
            pass
    
    filtro_valor_max = request.GET.get('filtro_valor_max', '').strip()
    if filtro_valor_max:
        try:
            controles_list = controles_list.filter(
                Q(valor_total_pedido__lte=Decimal(filtro_valor_max.replace(',', '.'))) |
                Q(valor_nf__lte=Decimal(filtro_valor_max.replace(',', '.')))
            )
        except (ValueError, TypeError):
            pass
    
    # Ordenação
    order_by = request.GET.get('order_by', '-data_rc')
    if order_by:
        controles_list = controles_list.order_by(order_by)
    
    # Paginação
    paginator = Paginator(controles_list, 50)  # 50 itens por página
    page_number = request.GET.get('page', 1)
    controles = paginator.get_page(page_number)
    
    # Estatísticas
    total_count = ControleRCeNF.objects.count()
    empresas_count = ControleRCeNF.objects.exclude(empresa__isnull=True).exclude(empresa='').values('empresa').distinct().count()
    
    # Calcular valor total
    valor_total_result = ControleRCeNF.objects.aggregate(
        total_pedido=Sum('valor_total_pedido'),
        total_nf=Sum('valor_nf')
    )
    valor_total_pedido = valor_total_result['total_pedido'] or Decimal('0.00')
    valor_total_nf = valor_total_result['total_nf'] or Decimal('0.00')
    
    # Obter valores únicos para os dropdowns de filtros
    empresas_unicas = ControleRCeNF.objects.exclude(
        empresa__isnull=True
    ).exclude(
        empresa=''
    ).values_list('empresa', flat=True).distinct().order_by('empresa')
    
    solicitantes_unicos = ControleRCeNF.objects.exclude(
        solicitante__isnull=True
    ).exclude(
        solicitante=''
    ).values_list('solicitante', flat=True).distinct().order_by('solicitante')
    
    status_unicos = ControleRCeNF.objects.exclude(
        status__isnull=True
    ).exclude(
        status=''
    ).values_list('status', flat=True).distinct().order_by('status')
    
    usos_unicos = ControleRCeNF.objects.exclude(
        uso__isnull=True
    ).exclude(
        uso=''
    ).values_list('uso', flat=True).distinct().order_by('uso')
    
    # Detectar duplicatas
    # Duplicatas são detectadas baseadas em combinações de campos chave
    duplicatas_analise = []
    
    # 1. Duplicatas por RC + NF Saída (mesma RC e mesma NF)
    duplicatas_rc_nf = ControleRCeNF.objects.values('rc', 'nf_saida').annotate(
        count=Count('id_excel')
    ).filter(count__gt=1).exclude(rc__isnull=True).exclude(rc='').exclude(nf_saida__isnull=True).exclude(nf_saida='')
    
    for dup in duplicatas_rc_nf:
        registros = ControleRCeNF.objects.filter(rc=dup['rc'], nf_saida=dup['nf_saida']).order_by('id_excel')
        ids = list(registros.values_list('id_excel', flat=True))
        duplicatas_analise.append({
            'tipo': 'RC + NF Saída',
            'campos': {'RC': dup['rc'], 'NF Saída': dup['nf_saida']},
            'quantidade': dup['count'],
            'ids': ids,
            'registros': registros,
            'motivo': f"Registros com RC '{dup['rc']}' e NF Saída '{dup['nf_saida']}' aparecem {dup['count']} vezes no banco de dados."
        })
    
    # 2. Duplicatas por RC + Pedido (mesma RC e mesmo Pedido)
    duplicatas_rc_pedido = ControleRCeNF.objects.values('rc', 'pedido').annotate(
        count=Count('id_excel')
    ).filter(count__gt=1).exclude(rc__isnull=True).exclude(rc='').exclude(pedido__isnull=True).exclude(pedido='')
    
    for dup in duplicatas_rc_pedido:
        # Verificar se já não foi incluído na análise anterior
        if not any(d['tipo'] == 'RC + Pedido' and d['campos']['RC'] == dup['rc'] and d['campos']['Pedido'] == dup['pedido'] for d in duplicatas_analise):
            registros = ControleRCeNF.objects.filter(rc=dup['rc'], pedido=dup['pedido']).order_by('id_excel')
            ids = list(registros.values_list('id_excel', flat=True))
            duplicatas_analise.append({
                'tipo': 'RC + Pedido',
                'campos': {'RC': dup['rc'], 'Pedido': dup['pedido']},
                'quantidade': dup['count'],
                'ids': ids,
                'registros': registros,
                'motivo': f"Registros com RC '{dup['rc']}' e Pedido '{dup['pedido']}' aparecem {dup['count']} vezes no banco de dados."
            })
    
    # 3. Duplicatas por NF Saída + Pedido (mesma NF e mesmo Pedido)
    duplicatas_nf_pedido = ControleRCeNF.objects.values('nf_saida', 'pedido').annotate(
        count=Count('id_excel')
    ).filter(count__gt=1).exclude(nf_saida__isnull=True).exclude(nf_saida='').exclude(pedido__isnull=True).exclude(pedido='')
    
    for dup in duplicatas_nf_pedido:
        # Verificar se já não foi incluído
        if not any(d['tipo'] == 'NF Saída + Pedido' and d['campos']['NF Saída'] == dup['nf_saida'] and d['campos']['Pedido'] == dup['pedido'] for d in duplicatas_analise):
            registros = ControleRCeNF.objects.filter(nf_saida=dup['nf_saida'], pedido=dup['pedido']).order_by('id_excel')
            ids = list(registros.values_list('id_excel', flat=True))
            duplicatas_analise.append({
                'tipo': 'NF Saída + Pedido',
                'campos': {'NF Saída': dup['nf_saida'], 'Pedido': dup['pedido']},
                'quantidade': dup['count'],
                'ids': ids,
                'registros': registros,
                'motivo': f"Registros com NF Saída '{dup['nf_saida']}' e Pedido '{dup['pedido']}' aparecem {dup['count']} vezes no banco de dados."
            })
    
    # 4. Duplicatas por RC apenas (mesma RC, mas pode ter NFs diferentes)
    duplicatas_rc_apenas = ControleRCeNF.objects.values('rc').annotate(
        count=Count('id_excel')
    ).filter(count__gt=1).exclude(rc__isnull=True).exclude(rc='')
    
    for dup in duplicatas_rc_apenas:
        # Verificar se já não foi incluído em análises anteriores
        if not any(d['campos'].get('RC') == dup['rc'] for d in duplicatas_analise):
            registros = ControleRCeNF.objects.filter(rc=dup['rc']).order_by('id_excel')
            ids = list(registros.values_list('id_excel', flat=True))
            nfs_unicas = registros.exclude(nf_saida__isnull=True).exclude(nf_saida='').values_list('nf_saida', flat=True).distinct()
            if len(nfs_unicas) == 1:
                # Se todas têm a mesma NF, já foi detectado acima
                continue
            duplicatas_analise.append({
                'tipo': 'RC (múltiplas ocorrências)',
                'campos': {'RC': dup['rc']},
                'quantidade': dup['count'],
                'ids': ids,
                'registros': registros,
                'motivo': f"RC '{dup['rc']}' aparece {dup['count']} vezes no banco de dados com diferentes NFs ou Pedidos."
            })
    
    context = {
        'page_title': 'Consultar Planilha RC',
        'active_page': 'consultar_planilha_rc',
        'controles': controles,
        'search_query': search_query,
        'total_count': total_count,
        'empresas_count': empresas_count,
        'valor_total_pedido': valor_total_pedido,
        'valor_total_nf': valor_total_nf,
        'empresas_unicas': empresas_unicas,
        'solicitantes_unicos': solicitantes_unicos,
        'status_unicos': status_unicos,
        'usos_unicos': usos_unicos,
        'filtro_empresa': filtro_empresa,
        'filtro_solicitante': filtro_solicitante,
        'filtro_rc': filtro_rc,
        'filtro_os': filtro_os,
        'filtro_status': filtro_status,
        'filtro_uso': filtro_uso,
        'filtro_valor_min': filtro_valor_min,
        'filtro_valor_max': filtro_valor_max,
        'duplicatas_analise': duplicatas_analise,
    }
    
    return render(request, 'orcamento/consultar_planilha_rc.html', context)


def visualizar_planilha_rc(request, controle_id):
    """Visualizar detalhes completos de um registro do Controle RC e NF"""
    from app.models import ControleRCeNF
    from django.contrib import messages
    
    try:
        controle = ControleRCeNF.objects.get(id=controle_id)
    except ControleRCeNF.DoesNotExist:
        messages.error(request, 'Registro não encontrado.')
        return redirect('consultar_planilha_rc')
    
    context = {
        'page_title': f'Visualizar Planilha RC - {controle.rc or controle.nf_saida or controle.id}',
        'active_page': 'consultar_planilha_rc',
        'controle': controle,
    }
    return render(request, 'visualizar/visualizar_planilha_rc.html', context)


def analise_planilha_rc(request):
    """Análise da Planilha RC e NF com foco em status"""
    from app.models import ControleRCeNF
    from django.db.models import Count, Sum, Q, Avg
    from decimal import Decimal
    from collections import defaultdict
    import json
    from datetime import datetime, timedelta
    
    # ========== FILTROS ==========
    ano_filtro = request.GET.get('ano', None)
    meses_filtro = request.GET.getlist('mes')  # getlist para múltiplos valores
    uso_filtro = request.GET.get('uso', '').strip() or None
    classificacao_filtro = request.GET.get('classificacao', '').strip() or None
    status_filtro = request.GET.get('status', '').strip() or None
    previsao_ano_filtro = request.GET.get('previsao_ano', '').strip() or None
    previsao_meses_filtro = request.GET.getlist('previsao_mes')
    
    hoje = datetime.now()
    if not ano_filtro:
        ano_filtro = str(hoje.year)
    
    try:
        ano_filtro = int(ano_filtro)
    except (ValueError, TypeError):
        ano_filtro = hoje.year
    
    meses_filtro_int = []
    if meses_filtro:
        for mes in meses_filtro:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    meses_filtro_int.append(mes_int)
            except (ValueError, TypeError):
                continue
        meses_filtro_int = sorted(list(set(meses_filtro_int)))
    
    meses_para_mostrar = meses_filtro_int if meses_filtro_int else list(range(1, 13))
    
    # Previsão para uso: ano e meses
    previsao_ano_int = None
    if previsao_ano_filtro:
        try:
            previsao_ano_int = int(previsao_ano_filtro)
        except (ValueError, TypeError):
            pass
    previsao_meses_int = []
    for m in previsao_meses_filtro:
        try:
            mi = int(m)
            if 1 <= mi <= 12:
                previsao_meses_int.append(mi)
        except (ValueError, TypeError):
            continue
    previsao_meses_int = sorted(list(set(previsao_meses_int))) if previsao_meses_int else []
    
    # Valores distintos para dropdowns (do banco completo)
    usos_disponiveis = list(
        ControleRCeNF.objects.exclude(Q(uso__isnull=True) | Q(uso=''))
        .values_list('uso', flat=True).distinct().order_by('uso')
    )
    classificacoes_disponiveis = list(
        ControleRCeNF.objects.exclude(Q(classificacao__isnull=True) | Q(classificacao=''))
        .values_list('classificacao', flat=True).distinct().order_by('classificacao')
    )
    status_disponiveis = list(
        ControleRCeNF.objects.exclude(Q(status__isnull=True) | Q(status=''))
        .values_list('status', flat=True).distinct().order_by('status')
    )
    
    # Anos disponíveis (data_rc + previsao_para_uso + ano atual)
    anos_from_db = list(
        ControleRCeNF.objects.exclude(data_rc__isnull=True)
        .dates('data_rc', 'year')
    )
    anos_previsao_from_db = list(
        ControleRCeNF.objects.exclude(previsao_para_uso__isnull=True)
        .dates('previsao_para_uso', 'year')
    )
    anos_disponiveis = sorted(
        set(d.year for d in anos_from_db) | set(d.year for d in anos_previsao_from_db) | {hoje.year},
        reverse=True
    )
    if not anos_disponiveis:
        anos_disponiveis = [hoje.year]
    
    # Base queryset com filtros
    queryset_base = ControleRCeNF.objects.filter(data_rc__year=ano_filtro)
    queryset_base = queryset_base.filter(data_rc__month__in=meses_para_mostrar)
    
    if uso_filtro:
        queryset_base = queryset_base.filter(uso=uso_filtro)
    if classificacao_filtro:
        queryset_base = queryset_base.filter(classificacao=classificacao_filtro)
    if status_filtro:
        queryset_base = queryset_base.filter(status=status_filtro)
    if previsao_ano_int:
        queryset_base = queryset_base.filter(previsao_para_uso__year=previsao_ano_int)
        if previsao_meses_int:
            queryset_base = queryset_base.filter(previsao_para_uso__month__in=previsao_meses_int)
    
    # Estatísticas básicas (usando queryset filtrado)
    total_registros = queryset_base.count()
    
    # Análise por STATUS (mais importante)
    status_distribution = queryset_base.exclude(
        status__isnull=True
    ).exclude(
        status=''
    ).values('status').annotate(
        total=Count('id_excel'),
        valor_total_pedido=Sum('valor_total_pedido'),
        valor_total_nf=Sum('valor_nf')
    ).order_by('-total')
    
    status_labels = [item['status'][:50] for item in status_distribution]
    status_data = [item['total'] for item in status_distribution]
    status_valores_pedido = [float(item['valor_total_pedido'] or 0) for item in status_distribution]
    status_valores_nf = [float(item['valor_total_nf'] or 0) for item in status_distribution]
    
    # Calcular percentuais para status
    status_percentages = []
    if total_registros > 0:
        status_percentages = [(item['total'] / total_registros * 100) for item in status_distribution]
    else:
        status_percentages = [0] * len(status_distribution)
    
    # Registros sem status (usado na tabela Detalhamento)
    registros_sem_status = queryset_base.filter(
        Q(status__isnull=True) | Q(status='')
    ).count()

    # Total com Previsão de uso Atual: registros onde previsao_para_uso está dentro do período filtrado
    # Usar período da Previsão para Uso quando aplicado; senão usar período da Data RC
    _ano_pp = previsao_ano_int if previsao_ano_int else ano_filtro
    _meses_pp = previsao_meses_int if previsao_meses_int else meses_para_mostrar
    _qs_previsao_atual = queryset_base.exclude(
        previsao_para_uso__isnull=True
    ).filter(
        previsao_para_uso__year=_ano_pp,
        previsao_para_uso__month__in=_meses_pp
    )
    total_com_previsao_uso_atual = _qs_previsao_atual.count()
    valor_previsao_uso_atual = _qs_previsao_atual.aggregate(
        total=Sum('valor_total_pedido'),
        nf=Sum('valor_nf')
    )
    valor_previsao_uso_atual = (valor_previsao_uso_atual.get('total') or Decimal(0)) + (valor_previsao_uso_atual.get('nf') or Decimal(0))
    
    # Análise por Empresa
    empresa_distribution = queryset_base.exclude(
        empresa__isnull=True
    ).exclude(
        empresa=''
    ).values('empresa').annotate(
        total=Count('id_excel'),
        valor_total_pedido=Sum('valor_total_pedido'),
        valor_nf=Sum('valor_nf')
    ).order_by('-total')[:10]
    
    empresa_labels = [item['empresa'][:30] for item in empresa_distribution]
    empresa_data = [item['total'] for item in empresa_distribution]
    empresa_valores_pedido = [float(item['valor_total_pedido'] or 0) for item in empresa_distribution]
    empresa_valores_nf = [float(item['valor_nf'] or 0) for item in empresa_distribution]
    
    # Análise por Solicitante
    solicitante_distribution = queryset_base.exclude(
        solicitante__isnull=True
    ).exclude(
        solicitante=''
    ).values('solicitante').annotate(
        total=Count('id_excel'),
        valor_total_pedido=Sum('valor_total_pedido'),
        valor_nf=Sum('valor_nf')
    ).order_by('-total')[:10]
    
    solicitante_labels = [item['solicitante'][:30] for item in solicitante_distribution]
    solicitante_data = [item['total'] for item in solicitante_distribution]
    solicitante_valores_pedido = [float(item['valor_total_pedido'] or 0) for item in solicitante_distribution]
    solicitante_valores_nf = [float(item['valor_nf'] or 0) for item in solicitante_distribution]
    
    # Análise por Uso
    uso_distribution = queryset_base.exclude(
        uso__isnull=True
    ).exclude(
        uso=''
    ).values('uso').annotate(
        total=Count('id_excel')
    ).order_by('-total')
    
    uso_labels = [item['uso'] for item in uso_distribution]
    uso_data = [item['total'] for item in uso_distribution]
    
    # Análise temporal por Data RC (período filtrado)
    registros_por_mes = defaultdict(int)
    registros_com_data = queryset_base.exclude(data_rc__isnull=True)
    
    for registro in registros_com_data:
        if registro.data_rc:
            mes_ano = registro.data_rc.strftime('%Y-%m')
            registros_por_mes[mes_ano] += 1
    
    # Meses no período filtrado (ano + meses selecionados)
    for m in meses_para_mostrar:
        key = f'{ano_filtro}-{m:02d}'
        if key not in registros_por_mes:
            registros_por_mes[key] = 0
    meses_ordenados = sorted(registros_por_mes.keys())
    meses_labels = [datetime.strptime(m, '%Y-%m').strftime('%b/%Y') for m in meses_ordenados]
    meses_data = [registros_por_mes[m] for m in meses_ordenados]
    
    # Valores totais
    valor_total_pedido_geral = queryset_base.aggregate(
        total=Sum('valor_total_pedido')
    )['total'] or Decimal('0.00')
    
    valor_total_nf_geral = queryset_base.aggregate(
        total=Sum('valor_nf')
    )['total'] or Decimal('0.00')
    
    # Valores médios
    valor_medio_pedido = queryset_base.exclude(
        valor_total_pedido__isnull=True
    ).aggregate(
        media=Avg('valor_total_pedido')
    )['media'] or Decimal('0.00')
    
    valor_medio_nf = queryset_base.exclude(
        valor_nf__isnull=True
    ).aggregate(
        media=Avg('valor_nf')
    )['media'] or Decimal('0.00')
    
    # Top 10 status por valor total
    top_status_por_valor = queryset_base.exclude(
        status__isnull=True
    ).exclude(
        status=''
    ).values('status').annotate(
        total_registros=Count('id_excel'),
        valor_total_pedido=Sum('valor_total_pedido'),
        valor_total_nf=Sum('valor_nf')
    ).order_by('-valor_total_nf')[:10]
    
    top_status_valor_labels = [item['status'][:40] for item in top_status_por_valor]
    top_status_valor_pedido = [float(item['valor_total_pedido'] or 0) for item in top_status_por_valor]
    top_status_valor_nf = [float(item['valor_total_nf'] or 0) for item in top_status_por_valor]
    
    # Estatísticas adicionais
    registros_com_rc = queryset_base.exclude(
        Q(rc__isnull=True) | Q(rc='')
    ).count()
    
    registros_com_pedido = queryset_base.exclude(
        Q(pedido__isnull=True) | Q(pedido='')
    ).count()
    
    registros_com_nf_saida = queryset_base.exclude(
        Q(nf_saida__isnull=True) | Q(nf_saida='')
    ).count()
    
    registros_com_nf_servico = queryset_base.exclude(
        Q(nf_servico__isnull=True) | Q(nf_servico='')
    ).count()
    
    registros_com_saldo_residual_pedido = queryset_base.exclude(
        saldo_residual_pedido__isnull=True
    ).count()
    
    # Resumo dos filtros ativos para exibição
    MESES_NOMES = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    filtros_ativos_resumo = []
    # Data RC
    if meses_para_mostrar and len(meses_para_mostrar) == 12:
        filtros_ativos_resumo.append(f"Data RC: {ano_filtro} (todos os 12 meses)")
    elif meses_para_mostrar:
        meses_str = ', '.join(MESES_NOMES[i - 1] for i in meses_para_mostrar)
        filtros_ativos_resumo.append(f"Data RC: {ano_filtro} ({meses_str})")
    else:
        filtros_ativos_resumo.append(f"Data RC: {ano_filtro} (todos)")
    # Previsão para Uso
    if previsao_ano_int:
        if previsao_meses_int:
            if len(previsao_meses_int) == 12:
                filtros_ativos_resumo.append(f"Previsão para Uso: {previsao_ano_int} (todos os 12 meses)")
            else:
                meses_str = ', '.join(MESES_NOMES[i - 1] for i in previsao_meses_int)
                filtros_ativos_resumo.append(f"Previsão para Uso: {previsao_ano_int} ({meses_str})")
        else:
            filtros_ativos_resumo.append(f"Previsão para Uso: {previsao_ano_int} (todos os meses)")
    if uso_filtro:
        filtros_ativos_resumo.append(f"Uso: {uso_filtro}")
    if classificacao_filtro:
        filtros_ativos_resumo.append(f"Classificação: {classificacao_filtro}")
    if status_filtro:
        filtros_ativos_resumo.append(f"Status: {status_filtro}")
    
    # Tabela Status x Mês (ano atual): duas linhas por status - Data RC e Previsão p/ Uso
    ano_atual = hoje.year
    # Base: data_rc
    qs_rc = ControleRCeNF.objects.filter(data_rc__year=ano_atual).exclude(
        Q(status__isnull=True) | Q(status='')
    )
    agg_rc = qs_rc.values('status', 'data_rc__month').annotate(
        vp=Sum('valor_total_pedido'), vn=Sum('valor_nf')
    ).order_by('status', 'data_rc__month')
    tabela_valores_rc = {}
    status_set = set()
    for row in agg_rc:
        st, mes = row['status'], row['data_rc__month']
        if st and mes:
            status_set.add(st)
            val = (row['vp'] or Decimal(0)) + (row['vn'] or Decimal(0))
            tabela_valores_rc[(st, mes)] = val
    # Base: previsao_para_uso
    qs_previsao = ControleRCeNF.objects.filter(
        previsao_para_uso__year=ano_atual
    ).exclude(previsao_para_uso__isnull=True).exclude(
        Q(status__isnull=True) | Q(status='')
    )
    agg_previsao = qs_previsao.values('status', 'previsao_para_uso__month').annotate(
        vp=Sum('valor_total_pedido'), vn=Sum('valor_nf')
    ).order_by('status', 'previsao_para_uso__month')
    tabela_valores_previsao = {}
    for row in agg_previsao:
        st, mes = row['status'], row['previsao_para_uso__month']
        if st and mes:
            status_set.add(st)
            val = (row['vp'] or Decimal(0)) + (row['vn'] or Decimal(0))
            tabela_valores_previsao[(st, mes)] = val
    # Build grouped rows: [(status, row_rc, row_previsao), ...] for visual grouping
    tabela_status_list = sorted(status_set)
    tabela_status_mes_rows = []  # kept for backward compat / empty check
    tabela_status_grouped = []
    tabela_totais_rc = {f'mes_{m}': Decimal(0) for m in range(1, 13)}
    tabela_totais_rc['total'] = Decimal(0)
    tabela_totais_previsao = {f'mes_{m}': Decimal(0) for m in range(1, 13)}
    tabela_totais_previsao['total'] = Decimal(0)
    for st in tabela_status_list:
        row_rc = {'base': 'data_rc', 'total': Decimal(0)}
        row_previsao = {'base': 'previsao', 'total': Decimal(0)}
        for m in range(1, 13):
            v_rc = tabela_valores_rc.get((st, m), Decimal(0))
            v_previsao = tabela_valores_previsao.get((st, m), Decimal(0))
            row_rc[f'mes_{m}'] = v_rc
            row_previsao[f'mes_{m}'] = v_previsao
            row_rc['total'] += v_rc
            row_previsao['total'] += v_previsao
            tabela_totais_rc[f'mes_{m}'] += v_rc
            tabela_totais_previsao[f'mes_{m}'] += v_previsao
        tabela_totais_rc['total'] += row_rc['total']
        tabela_totais_previsao['total'] += row_previsao['total']
        tabela_status_grouped.append({'status': st, 'row_rc': row_rc, 'row_previsao': row_previsao})
        tabela_status_mes_rows.append(row_rc)
        tabela_status_mes_rows.append(row_previsao)
    
    context = {
        'page_title': 'Análise Planilha RC',
        'active_page': 'analise_planilha_rc',
        'ano_filtro': ano_filtro,
        'meses_filtro': meses_filtro_int,
        'anos_disponiveis': anos_disponiveis,
        'uso_filtro': uso_filtro,
        'classificacao_filtro': classificacao_filtro,
        'status_filtro': status_filtro,
        'previsao_ano_filtro': previsao_ano_int,
        'previsao_meses_filtro': previsao_meses_int,
        'usos_disponiveis': usos_disponiveis,
        'classificacoes_disponiveis': classificacoes_disponiveis,
        'status_disponiveis': status_disponiveis,
        'total_registros': total_registros,
        'registros_sem_status': registros_sem_status,
        'total_com_previsao_uso_atual': total_com_previsao_uso_atual,
        'valor_previsao_uso_atual': valor_previsao_uso_atual,
        'registros_com_rc': registros_com_rc,
        'registros_com_pedido': registros_com_pedido,
        'registros_com_nf_saida': registros_com_nf_saida,
        'registros_com_nf_servico': registros_com_nf_servico,
        'registros_com_saldo_residual_pedido': registros_com_saldo_residual_pedido,
        'valor_total_pedido_geral': valor_total_pedido_geral,
        'valor_total_nf_geral': valor_total_nf_geral,
        'valor_possiveis_gastos': (valor_previsao_uso_atual or Decimal(0)) - (valor_total_nf_geral or Decimal(0)),
        'filtros_ativos_resumo': filtros_ativos_resumo,
        'valor_medio_pedido': valor_medio_pedido,
        'valor_medio_nf': valor_medio_nf,
        # Status (principal)
        'status_distribution': status_distribution,
        'status_labels': json.dumps(status_labels, ensure_ascii=False),
        'status_data': json.dumps(status_data),
        'status_percentages': json.dumps([round(p, 1) for p in status_percentages]),
        'status_valores_pedido': json.dumps(status_valores_pedido),
        'status_valores_nf': json.dumps(status_valores_nf),
        # Empresa
        'empresa_labels': json.dumps(empresa_labels, ensure_ascii=False),
        'empresa_data': json.dumps(empresa_data),
        'empresa_valores_pedido': json.dumps(empresa_valores_pedido),
        'empresa_valores_nf': json.dumps(empresa_valores_nf),
        # Solicitante
        'solicitante_labels': json.dumps(solicitante_labels, ensure_ascii=False),
        'solicitante_data': json.dumps(solicitante_data),
        'solicitante_valores_pedido': json.dumps(solicitante_valores_pedido),
        'solicitante_valores_nf': json.dumps(solicitante_valores_nf),
        # Uso
        'uso_labels': json.dumps(uso_labels, ensure_ascii=False),
        'uso_data': json.dumps(uso_data),
        # Temporal
        'meses_labels': json.dumps(meses_labels, ensure_ascii=False),
        'meses_data': json.dumps(meses_data),
        # Top status por valor
        'top_status_valor_labels': json.dumps(top_status_valor_labels, ensure_ascii=False),
        'top_status_valor_pedido': json.dumps(top_status_valor_pedido),
        'top_status_valor_nf': json.dumps(top_status_valor_nf),
        # Tabela Status x Mês
        'tabela_status_mes_ano': ano_atual,
        'tabela_status_mes_rows': tabela_status_mes_rows,
        'tabela_status_grouped': tabela_status_grouped,
        'tabela_totais_rc': tabela_totais_rc,
        'tabela_totais_previsao': tabela_totais_previsao,
    }
    
    return render(request, 'orcamento/analise_planilha_rc.html', context)
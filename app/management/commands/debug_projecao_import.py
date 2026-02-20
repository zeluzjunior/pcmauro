"""
Debug do import de Projeção de Gastos - mostra estrutura exata do Excel.
Uso: python manage.py debug_projecao_import "caminho/para/arquivo.xlsx"
"""
import os
import re
import unicodedata

from django.core.management.base import BaseCommand


def normalize_key(s):
    if not s:
        return ''
    s = str(s).replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'\s+', ' ', s).strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s


class Command(BaseCommand):
    help = 'Debug: mostra colunas e dados do Excel para identificar SERVIÇO CONCLUÍDO'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Caminho do arquivo Excel')
        parser.add_argument('--rows', type=int, default=5, help='Linhas de dados a mostrar (default: 5)')

    def handle(self, *args, **options):
        file_path = options['file_path']
        n_rows = options['rows']
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'Arquivo não encontrado: {file_path}'))
            return

        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('openpyxl não instalado'))
            return

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = [s for s in wb.sheetnames if 'GASTO' in s.upper() or s == wb.sheetnames[0]]
        if not sheet_names:
            sheet_names = wb.sheetnames[:1]
        self.stdout.write(f'Planilhas: {wb.sheetnames}\n')

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            self.stdout.write(self.style.SUCCESS(f'\n=== Planilha: {sheet_name} ===\n'))

            for header_row in [1, 2]:
                headers = []
                for cell in sheet[header_row]:
                    v = cell.value
                    headers.append(v if v is not None and str(v).strip() else f'col_{len(headers)}')
                self.stdout.write(f'\n--- Cabeçalho linha {header_row} ({len(headers)} colunas) ---')
                for i, h in enumerate(headers):
                    self.stdout.write(f'  {i:2}: {repr(h)[:70]}')
                data_start = header_row + 1
                self.stdout.write(f'\n--- Dados linhas {data_start}-{data_start + n_rows - 1} ---')
                for r in range(data_start, min(data_start + n_rows, sheet.max_row + 1)):
                    row = [c.value for c in sheet[r]]
                    self.stdout.write(f'\nLinha {r}:')
                    for i in range(min(len(headers), len(row))):
                        v = row[i] if i < len(row) else None
                        if v is not None and str(v).strip():
                            self.stdout.write(f'  {headers[i][:40]}: {repr(v)[:50]}')
                self.stdout.write('')
            break
        col_p_idx = 15
        self.stdout.write(self.style.SUCCESS(f'\n--- Coluna P (índice {col_p_idx}) ---'))
        sheet = wb[sheet_names[0]]
        headers = [c.value for c in sheet[1]]
        if len(headers) > col_p_idx:
            self.stdout.write(f'Nome: {repr(headers[col_p_idx])}')
            for r in range(2, min(7, sheet.max_row + 1)):
                row = [c.value for c in sheet[r]]
                if len(row) > col_p_idx:
                    self.stdout.write(f'  Linha {r}: {repr(row[col_p_idx])}')
        self.stdout.write(self.style.SUCCESS('\nConcluído.'))

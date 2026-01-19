"""
Script to check if an Excel file matches the ProjecaoGasto model structure
"""
import openpyxl
import sys

def check_excel_file(file_path):
    """Check if Excel file has the correct structure for ProjecaoGasto"""
    
    print(f"Checking file: {file_path}\n")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Check if GASTOS sheet exists
        if 'GASTOS' not in wb.sheetnames:
            print(f"[X] ERROR: Sheet 'GASTOS' not found!")
            print(f"Available sheets: {wb.sheetnames}")
            return False
        
        sheet = wb['GASTOS']
        print(f"[OK] Found sheet 'GASTOS'")
        print(f"  Total rows: {sheet.max_row}")
        print(f"  Total columns: {sheet.max_column}\n")
        
        # Get headers
        headers = [cell.value for cell in sheet[1]]
        print("Headers found in file:")
        for i, header in enumerate(headers, 1):
            if header:
                print(f"  {i}. {repr(header)}")
        
        print("\n" + "=" * 80)
        print("Expected columns for ProjecaoGasto model:")
        print("=" * 80)
        
        # Expected columns (with variations)
        expected_columns = {
            'SETOR': ['SETOR ', 'SETOR', 'setor'],
            'SOLICITANTE': ['SOLICITANTE', 'solicitante'],
            'FORNECEDOR NOME FANTASIA': ['FORNECEDOR\nNOME FANTASIA', 'FORNECEDOR NOME FANTASIA', 'FORNECEDOR'],
            'FORNECEDOR CNPJ': ['FORNECEDOR\nCNPJ', 'FORNECEDOR CNPJ'],
            'DESCRIÇÃO DO SERVIÇO': ['DESCRIÇÃO DO SERVIÇO', 'DESCRIO DO SERVIO', 'DESCRICAO DO SERVICO'],
            'VALOR TOTAL': ['VALOR TOTAL', 'valor_total'],
            'PREVISÃO P/ EXECUÇÃO': ['PREVISÃO \nP/ EXECUÇÃO', 'PREVISÃO P/ EXECUÇÃO', 'PREVISO \nP/ EXECUO'],
            'USO CONTÁBIL': ['USO \nCONTÁBIL', 'USO CONTÁBIL', 'USO \nCONTBIL'],
            'NÚMERO DA NOTA FISCAL': ['NÚMERO DA \nNOTA FISCAL', 'NÚMERO DA NOTA FISCAL', 'NMERO DA \nNOTA FISCAL'],
            'ORDEM DE SERVIÇO': ['ORDEM \nDE SERVIÇO', 'ORDEM DE SERVIÇO', 'ORDEM \nDE SERVIO'],
            'DATA DE ABERTURA DA REQUISIÇÃO': ['DATA DE ABERTURA \nDA REQUISIÇÃO', 'DATA DE ABERTURA DA REQUISIÇÃO', 'DATA DE ABERTURA \nDA REQUISIO'],
            'NÚMERO DA REQUISIÇÃO DE COMPRA': ['NÚMERO DA REQUISIÇÃO \nDE COMPRA', 'NÚMERO DA REQUISIÇÃO DE COMPRA', 'NEMRO DA REQUISIO \nDE COMPRA'],
            'NÚMERO DO PEDIDO DE COMPRA': ['NÚMERO DO \nPEDIDO DE COMPRA', 'NÚMERO DO PEDIDO DE COMPRA', 'NMERO DO \nPEDIDO DE COMPRA'],
            'SERVIÇO CONCLUÍDO': ['SERVIÇO CONCLUÍDO', 'SERVIO CONCLUDO'],
            'NF DE SERVIÇO RECEBIDA': ['NF DE SERVIÇO\n RECEBIDA', 'NF DE SERVIÇO RECEBIDA', 'NF DE SERVIO\n RECEBIDA'],
            'NF ENVIADA PARA LANÇAMENTO': ['NF ENVIADA\n PARA LANÇAMENTO ', 'NF ENVIADA PARA LANÇAMENTO', 'NF ENVIADA\n PARA LANAMENTO '],
            'OBSERVAÇÕES': ['OBSERVAÇÕES', 'OBSERVAES', 'OBSERVACOES', 'OBSERVAÇÕES ']
        }
        
        # Normalize headers for comparison
        def normalize_header(header):
            if not header:
                return None
            # Remove extra spaces and normalize line breaks
            import re
            normalized = re.sub(r'\s+', ' ', str(header).replace('\n', ' ').strip())
            return normalized.upper()
        
        normalized_headers = {normalize_header(h): h for h in headers if h}
        
        print("\nColumn matching results:")
        print("-" * 80)
        
        all_found = True
        for expected_name, variations in expected_columns.items():
            found = False
            matched_variation = None
            
            for variation in variations:
                normalized_var = normalize_header(variation)
                if normalized_var in normalized_headers:
                    found = True
                    matched_variation = normalized_headers[normalized_var]
                    break
            
            if found:
                print(f"[OK] {expected_name:40s} -> Found: {repr(matched_variation)}")
            else:
                print(f"[X] {expected_name:40s} -> NOT FOUND")
                all_found = False
        
        print("\n" + "=" * 80)
        
        # Check for extra columns
        print("\nExtra columns found (not in model):")
        found_expected = set()
        for expected_name, variations in expected_columns.items():
            for variation in variations:
                normalized_var = normalize_header(variation)
                if normalized_var in normalized_headers:
                    found_expected.add(normalized_var)
        
        extra_cols = [h for h in normalized_headers.keys() if h not in found_expected]
        if extra_cols:
            for col in extra_cols:
                print(f"  - {normalized_headers[col]}")
        else:
            print("  (None)")
        
        print("\n" + "=" * 80)
        
        # Check sample data
        print("\nSample data (first 2 rows):")
        print("-" * 80)
        for row_num in range(2, min(4, sheet.max_row + 1)):
            row_data = [cell.value for cell in sheet[row_num]]
            if any(row_data):  # Skip empty rows
                print(f"\nRow {row_num}:")
                for i, (header, value) in enumerate(zip(headers, row_data)):
                    if header and value:
                        print(f"  {header}: {value}")
        
        print("\n" + "=" * 80)
        
        if all_found:
            print("\n[SUCCESS] This file appears to be compatible with ProjecaoGasto model!")
            print("   All required columns are present (with encoding variations handled).")
        else:
            print("\n[WARNING] Some expected columns are missing.")
            print("   The file may still work if missing columns are optional.")
        
        return all_found
        
    except Exception as e:
        print(f"[ERROR] Error reading file: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "_files/GASTOS/UTILIDADES_teste_projecao_gastos.xlsx"
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    
    check_excel_file(file_path)

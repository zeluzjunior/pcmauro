import openpyxl

wb = openpyxl.load_workbook('52_semanas_2026.xlsx', data_only=True)
ws = wb.active

print('Headers:')
headers = [cell.value for cell in ws[1]]
print(headers)

print('\nFirst 5 rows:')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
    print(f'Row {i}: {row}')

